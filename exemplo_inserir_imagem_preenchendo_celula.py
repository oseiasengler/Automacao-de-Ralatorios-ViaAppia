"""
Exemplo: inserir imagem preenchendo uma célula (ou merged range) no Excel.
Evita miniatura definindo a âncora com extent em EMU.

Uso: ajuste caminho_foto, cell_addr e opcionalmente (largura_px, altura_px).
"""
from io import BytesIO

from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import get_column_letter, column_index_from_string
from PIL import Image as PILImage

# 1 pixel ≈ 9525 EMU (96 DPI) — o Excel usa EMU no extent da âncora
EMU = 9525


def redimensionar_imagem_bytes(caminho_foto: str, largura: int, altura: int) -> bytes:
    """Redimensiona a imagem e retorna bytes JPEG."""
    buf = BytesIO()
    with PILImage.open(caminho_foto) as img:
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")
        img = img.resize((largura, altura), PILImage.Resampling.LANCZOS)
        img.save(buf, format="JPEG", quality=90)
    return buf.getvalue()


class ImageFromBytes(XLImage):
    """Imagem a partir de bytes (openpyxl lê no save(); arquivo temp sumiria)."""
    def __init__(self, data: bytes):
        super().__init__(BytesIO(data))
        self._bytes_data = data

    def _data(self):
        return self._bytes_data


def _get_merge_topleft(ws, col: int, row: int):
    """Retorna (min_col, min_row) do merge que contém (col, row). Se não mesclado, (col, row)."""
    for mr in ws.merged_cells.ranges:
        if mr.min_col <= col <= mr.max_col and mr.min_row <= row <= mr.max_row:
            return mr.min_col, mr.min_row
    return col, row


def inserir_imagem_preenchendo(ws, cell_addr: str, caminho_foto: str, largura_px: int = 300, altura_px: int = 225):
    """
    Insere a imagem na célula (ou no merged range que a contém) no tamanho desejado.
    Sem isso, o openpyxl pode desenhar a imagem em tamanho miniatura.
    """
    col_letter = "".join(c for c in cell_addr if c.isalpha())
    row_num = int("".join(c for c in cell_addr if c.isdigit()))
    col_num = column_index_from_string(col_letter)
    col_ancora, row_ancora = _get_merge_topleft(ws, col_num, row_num)

    # Tamanho em EMU para o extent da âncora (isso define o “quadro” que a imagem preenche)
    w_emu = largura_px * EMU
    h_emu = altura_px * EMU

    # 1. Redimensionar imagem com Pillow e obter bytes
    data = redimensionar_imagem_bytes(caminho_foto, largura_px, altura_px)

    # 2. Criar objeto de imagem para openpyxl (BytesIO ou path — NUNCA passar objeto PIL)
    xl_img = ImageFromBytes(data)
    xl_img.width = largura_px
    xl_img.height = altura_px

    # 3. O PULO DO GATO: âncora com extent em EMU (col/row são 0-based no AnchorMarker)
    anchor = OneCellAnchor()
    anchor._from = AnchorMarker(col=col_ancora - 1, colOff=0, row=row_ancora - 1, rowOff=0)
    anchor.ext = XDRPositiveSize2D(cx=w_emu, cy=h_emu)
    xl_img.anchor = anchor

    # 4. Adicionar à planilha (não use patch_add_image para posicionar — ele não faz isso)
    ws.add_image(xl_img)


# ——— Exemplo de uso ———
if __name__ == "__main__":
    from openpyxl import load_workbook

    caminho_foto = "sua_foto.jpg"  # ajuste
    cell_addr = "B2"
    largura_px, altura_px = 300, 225

    wb = load_workbook("seu_arquivo.xlsx")  # ajuste
    ws = wb.active

    inserir_imagem_preenchendo(ws, cell_addr, caminho_foto, largura_px, altura_px)

    wb.save("saida_com_foto.xlsx")
