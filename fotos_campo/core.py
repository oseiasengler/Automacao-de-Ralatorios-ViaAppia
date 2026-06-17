"""
fotos_campo/core.py
────────────────────────────────────────────────────────────
Lógica de processamento de fotos de campo.
Replica as Macros Panelas (Coord Renomear) de forma mais ágil e consistente.
Sem Tkinter — pode ser usado na GUI desktop e em endpoints web.

Módulos expostos pelo fotos_router (web):
  1. listar       – listar_de_zip / listar_arquivos_subpastas
  2. coordenadas  – coordenadas_km_bytes / processar_coordenadas_km
  3. renomear     – copiar_renomear_xlsx (uso desktop/local)

As funções dos módulos 4 (relatório Foto 2 Lados) e 5 (exportar KCor-Kria)
estão implementadas neste arquivo mas pertencem ao pipeline NC ARTESP.

Dependências:
  pip install openpyxl pillow piexif

Desenvolvedor: Ozeias Engler
"""

from __future__ import annotations

import io
import logging
import math
import os
import shutil
import sys
import tempfile
import zipfile
from collections import Counter
from copy import copy as copy_obj
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

_REPO_ROOT_FC = Path(__file__).resolve().parent.parent


def _extrair_zip_para_pasta(zf: zipfile.ZipFile, destino: Path) -> None:
    """No Windows usa ``\\\\?\\`` via nc_artesp.utils.helpers (caminhos longos)."""
    if os.name != "nt":
        destino.mkdir(parents=True, exist_ok=True)
        zf.extractall(str(destino))
        return
    if str(_REPO_ROOT_FC) not in sys.path:
        sys.path.insert(0, str(_REPO_ROOT_FC))
    try:
        from nc_artesp.utils.helpers import extrair_zipfile_para_pasta

        extrair_zipfile_para_pasta(zf, destino)
    except Exception:
        destino.mkdir(parents=True, exist_ok=True)
        zf.extractall(str(destino))


try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImg
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.formula.translate import Translator
    import openpyxl.utils
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    from PIL import Image as PILImage
    import piexif
    PIL_OK = True
except ImportError:
    PIL_OK = False

_LogCb = Optional[Callable[[str], None]]
_log_draft = logging.getLogger("fotos_campo.draft")

def _log_draft_ram(ident: str, size_before: Tuple[int, int], size_after: Tuple[int, int], channels: int = 3) -> None:
    """Log estimativa de RAM economizada por draft() (só em nível DEBUG)."""
    if not _log_draft.isEnabledFor(logging.DEBUG):
        return
    w0, h0 = size_before
    w1, h1 = size_after
    full_mb = (w0 * h0 * channels) / (1024 * 1024)
    after_mb = (w1 * h1 * channels) / (1024 * 1024)
    saved = max(0.0, full_mb - after_mb)
    _log_draft.debug("[draft] %s: %dx%d → %dx%d | ~%.2f MB RAM economizados", ident, w0, h0, w1, h1, saved)

# UTILITÁRIOS INTERNOS

def _log(cb: _LogCb, msg: str) -> None:
    if cb:
        cb(msg)


def _sanitizar_nome(s: str) -> str:
    """Remove caracteres inválidos para nome de arquivo Windows."""
    if not s or not isinstance(s, str):
        return ""
    for c in ('\\', '/', ':', '*', '?', '"', '<', '>', '|'):
        s = s.replace(c, "_")
    return s.strip()[:80]


# MÓDULO 1 – LISTAR ARQUIVOS + GPS EXIF

EXTS_FOTO = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp"}

COLUNAS_M1 = [
    "Caminho Foto", "Nome Arquivo", "Tamanho", "Tipo",
    "Latitude", "Longitude", "Rodovia", "km", "Sentido", "Distancia",
    "Novo Caminho", "Servico_Descricao", "Pasta",
    "Nova Pasta a Ser Salva", "Nome Simplificado", "Titulo",
    "Caminho Relatório",
]
NOME_ABA_FOTOS = "Fotos"
COL_CAMINHO_RELATORIO = 17  # coluna da cópia reduzida/renomeada para relatório


def exif_gps(path: str) -> Tuple[str, str]:
    """Extrai lat/lon do EXIF da foto. Retorna ('', '') se não encontrar."""
    if not PIL_OK:
        return "", ""
    try:
        exif = piexif.load(path)
        gps = exif.get("GPS", {})
        if not gps:
            return "", ""

        def to_deg(val):
            d, m, s = val
            return d[0] / d[1] + m[0] / m[1] / 60 + s[0] / s[1] / 3600

        lat_val = gps.get(piexif.GPSIFD.GPSLatitude)
        lat_ref = gps.get(piexif.GPSIFD.GPSLatitudeRef, b"S")
        lon_val = gps.get(piexif.GPSIFD.GPSLongitude)
        lon_ref = gps.get(piexif.GPSIFD.GPSLongitudeRef, b"W")
        if not lat_val or not lon_val:
            return "", ""
        lat = to_deg(lat_val)
        lon = to_deg(lon_val)
        if lat_ref in (b"S", "S"):
            lat = -lat
        if lon_ref in (b"W", "W"):
            lon = -lon
        return round(lat, 7), round(lon, 7)
    except Exception:
        return "", ""


def exif_data_foto(path: str) -> str:
    """Retorna data da foto do EXIF (DateTimeOriginal), formato DD/MM/YYYY HH:MM."""
    if not PIL_OK:
        return ""
    try:
        exif = piexif.load(path)
        dt = exif.get("Exif", {}).get(36867) or exif.get("Exif", {}).get(36868)
        if not dt:
            return ""
        if isinstance(dt, bytes):
            dt = dt.decode("utf-8", errors="ignore").strip()
        else:
            dt = str(dt)
        if " " in dt:
            data_part, hora_part = dt.split(" ", 1)
            partes = data_part.replace(":", "-").split("-")
            if len(partes) == 3:
                d, m, a = partes[2], partes[1], partes[0]
                h = hora_part.split(":")[:2]
                h_str = ":".join(h) if len(h) >= 2 else hora_part[:5]
                return f"{d}/{m}/{a} {h_str}"
        return dt[:16] if len(dt) >= 16 else dt
    except Exception:
        return ""


def listar_arquivos_subpastas(pasta_raiz: str, log_cb: _LogCb = None) -> List[dict]:
    """
    Percorre subpastas recursivamente, extrai metadados e GPS EXIF.
    Retorna lista de registros compatível com COLUNAS_M1.
    """
    registros = []
    raiz = Path(pasta_raiz)
    for dirpath, _, files in os.walk(str(raiz)):
        for nome_arq in sorted(files):
            caminho = Path(dirpath) / nome_arq
            ext = caminho.suffix.lower()
            lat, lon = ("", "")
            if ext in EXTS_FOTO:
                lat, lon = exif_gps(str(caminho))
            try:
                tamanho = caminho.stat().st_size
            except Exception:
                tamanho = ""
            registros.append({
                "Caminho Foto":           str(caminho),
                "Nome Arquivo":           nome_arq,
                "Tamanho":                tamanho,
                "Tipo":                   ext.lstrip(".").upper(),
                "Latitude":               lat,
                "Longitude":              lon,
                "Rodovia":                "",
                "km":                     "",
                "Sentido":                "",
                "Distancia":              "",
                "Novo Caminho":           "",
                "Servico_Descricao":      "",
                "Pasta":                  str(Path(dirpath)) + os.sep,
                "Nova Pasta a Ser Salva": "",
                "Nome Simplificado":      nome_arq,
                "Titulo":                 "",
                "Caminho Relatório":      "",
            })
            _log(log_cb, f"  {caminho.name}  lat={lat}  lon={lon}")
    return registros


def _escrever_aba_fotos(ws, registros: List[dict]) -> None:
    """Preenche a aba com cabeçalho COLUNAS_M1 e registros."""
    header_fill = PatternFill("solid", fgColor="4A4E69")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, col_nome in enumerate(COLUNAS_M1, 1):
        c = ws.cell(row=1, column=col_idx, value=col_nome)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")
    for row_idx, reg in enumerate(registros, 2):
        for col_idx, col_nome in enumerate(COLUNAS_M1, 1):
            ws.cell(row=row_idx, column=col_idx, value=reg.get(col_nome, ""))
    larguras = {
        "Caminho Foto": 60, "Pasta": 50, "Novo Caminho": 60,
        "Nova Pasta a Ser Salva": 40, "Nome Arquivo": 30,
        "Nome Simplificado": 30, "Titulo": 25, "Caminho Relatório": 60,
    }
    for col_idx, col_nome in enumerate(COLUNAS_M1, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = larguras.get(col_nome, 14)


def salvar_xlsx_modulo1(registros: List[dict], arquivo_saida: str) -> None:
    """Cria planilha nova com aba 'Fotos' (nova listagem do zero)."""
    if not OPENPYXL_OK:
        raise ImportError("openpyxl nao instalado. pip install openpyxl")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = NOME_ABA_FOTOS
    _escrever_aba_fotos(ws, registros)
    Path(arquivo_saida).parent.mkdir(parents=True, exist_ok=True)
    wb.save(arquivo_saida)


def criar_planilha_primeira_vez(arquivo_saida: str) -> None:
    """Cria planilha vazia (só cabeçalho, aba Fotos). Passo 1 do fluxo."""
    salvar_xlsx_modulo1([], arquivo_saida)


def adicionar_aba_xlsx_modulo1(novos_registros: List[dict],
                                arquivo_saida: str) -> Tuple[int, int, int, int]:
    """
    Adiciona nova aba sequencial (Fotos, Fotos_2, Fotos_3…) ao arquivo existente.
    Cria o arquivo se não existir.
    Retorna (abas_anteriores, qtd_novos, total_abas, total_registros).
    """
    if not OPENPYXL_OK:
        raise ImportError("openpyxl nao instalado. pip install openpyxl")
    if not os.path.exists(arquivo_saida):
        salvar_xlsx_modulo1(novos_registros, arquivo_saida)
        return 0, len(novos_registros), 1, len(novos_registros)
    wb = openpyxl.load_workbook(arquivo_saida)
    titulo = NOME_ABA_FOTOS
    n = 1
    while titulo in wb.sheetnames:
        n += 1
        titulo = f"{NOME_ABA_FOTOS}_{n}"
    ws_nova = wb.create_sheet(title=titulo)
    _escrever_aba_fotos(ws_nova, novos_registros)
    total_abas = len(wb.sheetnames)
    total_registros = sum(
        max(0, wb[sn].max_row - 1)
        for sn in wb.sheetnames
        if wb[sn].cell(1, 1).value and "Caminho" in str(wb[sn].cell(1, 1).value)
    )
    Path(arquivo_saida).parent.mkdir(parents=True, exist_ok=True)
    wb.save(arquivo_saida)
    return total_abas - 1, len(novos_registros), total_abas, total_registros


def listar_de_zip(zip_bytes: bytes, log_cb: _LogCb = None) -> Tuple[bytes, int]:
    """
    Versão web do Módulo 1: recebe ZIP com fotos, extrai em temp, lista, retorna XLSX em bytes.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        zip_path = Path(tmpdir) / "fotos.zip"
        zip_path.write_bytes(zip_bytes)
        pasta_fotos = Path(tmpdir) / "fotos"
        pasta_fotos.mkdir()
        with zipfile.ZipFile(str(zip_path)) as zf:
            _extrair_zip_para_pasta(zf, pasta_fotos)
        registros = listar_arquivos_subpastas(str(pasta_fotos), log_cb=log_cb)
        # Ajusta caminhos para relativos (útil na web)
        for reg in registros:
            orig = reg.get("Caminho Foto", "")
            rel = os.path.relpath(orig, str(pasta_fotos)) if orig else orig
            reg["Caminho Foto"] = rel
            reg["Pasta"] = str(Path(rel).parent) + os.sep if rel else ""
        xlsx_path = Path(tmpdir) / "fotos_listagem.xlsx"
        salvar_xlsx_modulo1(registros, str(xlsx_path))
        return xlsx_path.read_bytes(), len(registros)


# MÓDULO 2 – COORDENADAS → RODOVIA / KM / SENTIDO

def haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Distância Haversine entre dois pontos GPS (em km)."""
    R = 6371.0
    d_lat = math.radians(lat2 - lat1)
    d_lon = math.radians(lon2 - lon1)
    a = (math.sin(d_lat / 2) ** 2 +
         math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) *
         math.sin(d_lon / 2) ** 2)
    return R * 2 * math.asin(math.sqrt(max(0, min(1, a))))


def processar_coordenadas_km(planilha_dados: str, relacao_total: str,
                              planilha_saida: Optional[str] = None,
                              log_cb: _LogCb = None) -> int:
    """
    Preenche Rodovia/km/Sentido/Distância em todas as abas da planilha de dados
    usando a Relação Total (A=Rodovia, B=km, C=Sentido, D=Lat, E=Lon).
    Retorna número de registros com rodovia/km gravados.
    """
    if not OPENPYXL_OK:
        raise ImportError("pip install openpyxl")

    wb_base = openpyxl.load_workbook(relacao_total, data_only=True)
    ws_base = wb_base.active
    pontos_base = []
    for row in ws_base.iter_rows(min_row=2, values_only=True):
        try:
            rodov = row[0];  km_val = row[1];  sent = row[2]
            lat_b = float(str(row[3]).replace(",", "."))
            lon_b = float(str(row[4]).replace(",", "."))
            pontos_base.append((rodov, km_val, sent, lat_b, lon_b))
        except Exception:
            continue

    if not pontos_base:
        raise ValueError(
            "Relação Total sem dados válidos nas colunas A-E.\n"
            "Estrutura esperada: A=Rodovia  B=km  C=Sentido  D=Latitude  E=Longitude"
        )
    _log(log_cb, f"[OK] Relação Total: {len(pontos_base)} pontos")

    wb = openpyxl.load_workbook(planilha_dados)

    def _vazio(v):
        if v is None: return True
        if isinstance(v, str) and v.strip() in ("", "None", "N/A", "nan"): return True
        return False

    processados = sem_coords = total_linhas = abas_processadas = 0

    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        a1 = ws.cell(row=1, column=1).value
        if not a1 or "Caminho" not in str(a1):
            continue
        abas_processadas += 1
        linhas_aba = max(0, ws.max_row - 1)
        total_linhas += linhas_aba
        _log(log_cb, f"[OK] Aba '{nome_aba}': {linhas_aba} linhas")

        ws["G1"] = "Rodovia";  ws["H1"] = "km";  ws["I1"] = "Sentido";  ws["J1"] = "Distancia"

        for row_idx in range(2, ws.max_row + 1):
            lat_cell = ws.cell(row=row_idx, column=5).value
            lon_cell = ws.cell(row=row_idx, column=6).value
            if _vazio(lat_cell) or _vazio(lon_cell):
                for col in [7, 8, 9, 10]:
                    ws.cell(row=row_idx, column=col).value = ""
                sem_coords += 1
                continue
            try:
                lat = float(str(lat_cell).replace(",", "."))
                lon = float(str(lon_cell).replace(",", "."))
            except ValueError:
                for col in [7, 8, 9, 10]:
                    ws.cell(row=row_idx, column=col).value = ""
                sem_coords += 1
                continue

            melhor = min(pontos_base, key=lambda p: haversine_km(lat, lon, p[3], p[4]))
            dist = haversine_km(lat, lon, melhor[3], melhor[4])
            try:
                km_fmt = f"{float(str(melhor[1]).replace(',', '.')):.3f}"
            except Exception:
                km_fmt = str(melhor[1]) if melhor[1] else ""

            ws.cell(row=row_idx, column=7).value  = str(melhor[0]) if melhor[0] else ""
            ws.cell(row=row_idx, column=8).value  = km_fmt
            ws.cell(row=row_idx, column=9).value  = str(melhor[2]) if melhor[2] else ""
            ws.cell(row=row_idx, column=10).value = round(dist, 4)
            processados += 1
            if log_cb and processados % 50 == 0:
                log_cb(f"  {processados} processados...")

    destino = (planilha_saida.strip() if planilha_saida and str(planilha_saida).strip()
               else None) or planilha_dados
    wb.save(destino)
    _log(log_cb, f"\n[INFO] Abas: {abas_processadas}  Linhas: {total_linhas}")
    _log(log_cb, f"[INFO] Com GPS: {processados}  Sem GPS: {sem_coords}")
    return processados


def coordenadas_km_bytes(xlsx_dados_bytes: bytes, xlsx_relacao_bytes: bytes,
                          log_cb: _LogCb = None) -> bytes:
    """Versão web do Módulo 2: opera com bytes em memória, retorna XLSX atualizado."""
    with tempfile.TemporaryDirectory() as tmpdir:
        p_dados = Path(tmpdir) / "dados.xlsx"
        p_relac = Path(tmpdir) / "relacao.xlsx"
        p_dados.write_bytes(xlsx_dados_bytes)
        p_relac.write_bytes(xlsx_relacao_bytes)
        processar_coordenadas_km(str(p_dados), str(p_relac), log_cb=log_cb)
        return p_dados.read_bytes()


def listar_rodovia_por_caminho(xlsx_listagem_km_bytes: bytes) -> List[Tuple[str, str]]:
    """
    Lê o XLSX da listagem com Rodovia/km (saída do Módulo 2) e retorna lista de
    (caminho_normalizado, rodovia). Caminho usa / e é relativo (ex: 'foto.jpg' ou 'pasta/foto.jpg').
    Usado para montar ZIP por rodovia.
    """
    if not OPENPYXL_OK:
        return []
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_listagem_km_bytes), data_only=True)
    out = []
    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        if not str(ws.cell(1, 1).value or "").startswith("Caminho"):
            continue
        for row in range(2, (ws.max_row or 0) + 1):
            caminho = ws.cell(row=row, column=1).value
            rodovia = ws.cell(row=row, column=7).value
            if not caminho:
                continue
            caminho_n = str(caminho).strip().replace("\\", "/").lstrip("/")
            rodovia_s = _sanitizar_nome(str(rodovia or "").strip()) or "Sem_rodovia"
            out.append((caminho_n, rodovia_s))
    return out


# MÓDULO 3 – COPIAR E RENOMEAR ARQUIVOS

MODOS_DESCRICAO: Dict[int, str] = {
    1: "Pasta Original  →  Rodovia - Sentido - km.jpg",
    2: "Pasta Original  →  NomeOriginal - Rodovia - Sentido - km.jpg",
    3: "Pasta Original  →  Rodovia - Sentido - km - NomeOriginal.jpg",
    4: "Pasta Nova (col N)  →  Rodovia - Sentido - km.jpg",
    5: "Pasta Nova (col N)  →  NomeOriginal - Rodovia - Sentido - km.jpg",
    6: "Pasta Nova (col N)  →  Rodovia - Sentido - km - NomeOriginal.jpg",
}


def montar_nome_destino(modo: int, pasta_orig: str, pasta_nova: str,
                         rodovia: str, sentido: str, km_val, arquivo_1: str) -> str:
    try:
        km_fmt = f"{float(str(km_val).replace(',', '.')):.3f}"
    except Exception:
        km_fmt = str(km_val)
    base_orig = pasta_orig.rstrip("/\\") + os.sep
    base_nova = pasta_nova.rstrip("/\\") + os.sep if pasta_nova else base_orig
    mapa = {
        1: base_orig + f"{rodovia} - {sentido} - km {km_fmt}.jpg",
        2: base_orig + f"{arquivo_1} - {rodovia} - {sentido} - km {km_fmt}.jpg",
        3: base_orig + f"{rodovia} - {sentido} - km {km_fmt} - {arquivo_1}.jpg",
        4: base_nova + f"{rodovia} - {sentido} - km {km_fmt}.jpg",
        5: base_nova + f"{arquivo_1} - {rodovia} - {sentido} - km {km_fmt}.jpg",
        6: base_nova + f"{rodovia} - {sentido} - km {km_fmt} - {arquivo_1}.jpg",
    }
    return mapa[modo]


def copiar_renomear_xlsx(planilha: str, modo: int,
                          planilha_saida: Optional[str] = None,
                          log_cb: _LogCb = None) -> dict:
    """Copia e renomeia arquivos conforme modo; atualiza coluna 'Novo Caminho' no XLSX."""
    if not OPENPYXL_OK:
        raise ImportError("pip install openpyxl")
    wb = openpyxl.load_workbook(planilha)
    stats = {"copiados": 0, "erros": 0, "ignorados": 0}

    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        a1 = ws.cell(row=1, column=1).value
        if not a1 or "Caminho" not in str(a1):
            continue
        _log(log_cb, f"[OK] Aba '{nome_aba}' ({ws.max_row - 1} linhas)")

        for row in range(2, ws.max_row + 1):
            caminho_orig = ws.cell(row=row, column=1).value
            if not caminho_orig:
                stats["ignorados"] += 1
                continue
            caminho_orig = str(caminho_orig)
            rodovia   = str(ws.cell(row=row, column=7).value  or "")
            km_val    = ws.cell(row=row, column=8).value
            sentido   = str(ws.cell(row=row, column=9).value  or "")
            pasta_n   = str(ws.cell(row=row, column=14).value or "")
            arquivo_1 = str(ws.cell(row=row, column=2).value  or "")
            if not arquivo_1:
                arquivo_1 = Path(caminho_orig).name

            ultima_barra = max(caminho_orig.rfind("\\"), caminho_orig.rfind("/"))
            pasta_orig = caminho_orig[:ultima_barra + 1] if ultima_barra != -1 else ""
            ws.cell(row=row, column=13).value = pasta_orig

            if not rodovia or not km_val:
                _log(log_cb, f"[AVISO] Linha {row}: sem rodovia/km — execute Módulo 2 primeiro")
                stats["ignorados"] += 1
                continue
            try:
                destino = montar_nome_destino(modo, pasta_orig, pasta_n,
                                              rodovia, sentido, km_val, arquivo_1)
            except Exception as e:
                _log(log_cb, f"[ERRO] Linha {row}: {e}")
                stats["erros"] += 1
                continue

            ws.cell(row=row, column=11).value = destino
            Path(destino).parent.mkdir(parents=True, exist_ok=True)

            try:
                orig_norm = os.path.normpath(os.path.abspath(caminho_orig))
                dest_norm = os.path.normpath(os.path.abspath(destino))
                if orig_norm == dest_norm:
                    stats["copiados"] += 1
                    continue
            except Exception:
                pass
            try:
                if os.path.exists(caminho_orig):
                    shutil.copy2(caminho_orig, destino)
                    stats["copiados"] += 1
                    _log(log_cb, f"[OK] {Path(caminho_orig).name} → {Path(destino).name}")
                else:
                    _log(log_cb, f"[AVISO] Não encontrado: {caminho_orig}")
                    stats["ignorados"] += 1
            except Exception as e:
                _log(log_cb, f"[ERRO] {Path(caminho_orig).name}: {e}")
                stats["erros"] += 1

    destino_xlsx = (planilha_saida.strip()
                    if planilha_saida and str(planilha_saida).strip()
                    else None) or planilha
    wb.save(destino_xlsx)
    return stats


# MÓDULO 4 – RELATÓRIO FOTOGRÁFICO 2 LADOS
# Template: nc_artesp/assets/templates/Planilha Modelo Conservação - Foto 2 Lados.xlsx
# Estrutura validada no template (aba Relat_Foto_Cons):
# - Fotos: C8:H18 (esq.) e J8:O18 (dir.), âncora C8/J8
# - Dados: rodovia/km/sentido em j (linha 19), serviço em j+1 (linha 20)
# - Data da Constatação (EXIF): j+3 (linha 22), data em G/N e hora em H/O
#
LINHA_INICIO_BLOCO  = 19
TAMANHO_BLOCO       = 16   # padrão macro/template: 19 a 34
LINHA_DADOS_ABAIXO_FOTO = 0    # j+0 = linha 19 (rodovia, km, sentido)
LINHA_DATA_NO_BLOCO = 3        # j+3 = linha 22 (Data da Constatação)
LINHA_IMAGEM_NO_BLOCO = -11    # foto em j-11 (linha 8) — âncora C8/J8

# Colunas bloco esquerdo (C–H): C=3, D=4, E=5, F=6, G=7, H=8
COL_ESQ_RODOVIA = 5   # E
COL_ESQ_KM      = 7   # G
COL_ESQ_SENTIDO = 8   # H
COL_ESQ_SERVICO = 3   # C (serviço linha 19)
COL_ESQ_FOTO    = 3   # C (âncora C8:H18)
COL_ESQ_DATA    = 7   # G (data na linha 22)
COL_ESQ_HORA    = 8   # H (hora na linha 22)
# Colunas bloco direito (J–O)
COL_DIR_RODOVIA = 12  # L
COL_DIR_KM      = 14  # N
COL_DIR_SENTIDO = 15  # O
COL_DIR_SERVICO = 10  # J (serviço linha 19)
COL_DIR_FOTO    = 10  # J (âncora J8:O18)
COL_DIR_DATA    = 14  # N (data na linha 22)
COL_DIR_HORA    = 15  # O (hora na linha 22)

# ─── Inserção de imagem preenchendo célula/merged range (igual NC ARTESP) ─────
EMU = 9525  # 1 pixel ≈ 9525 EMU (96 DPI)
# Tamanho do QUADRO no template (cm): a foto deve caber exatamente aqui.
# (Macro Art_022 usa 275×210 pt ≈ 9,7×7,4 cm, maior que este quadro.)
FOTO2LADOS_QUADRO_W_CM = 8.34
FOTO2LADOS_QUADRO_H_CM = 6.35
# OOXML: 1 inch = 914400 EMU → 1 cm = 914400/2.54 EMU
EMU_PER_CM = 914400 / 2.54
# Pixels a 96 DPI: 1 cm = 96/2.54 px
PX_PER_CM = 96 / 2.54
# Macro Art_022 (referência): AddPicture Width:=275, Height:=210 pt → 367×280 px
FOTO2LADOS_MACRO_W_PT = 275
FOTO2LADOS_MACRO_H_PT = 210
FOTO2LADOS_MIN_W_PX = int(FOTO2LADOS_MACRO_W_PT * 96 / 72)
FOTO2LADOS_MIN_H_PX = int(FOTO2LADOS_MACRO_H_PT * 96 / 72)


def _col_px(ws, letter: str) -> int:
    """Largura aproximada da coluna em pixels (Excel width em unidades de caractere → ~px a 96 DPI)."""
    w = ws.column_dimensions.get(letter)
    width = w.width if (w and w.width) else 10
    # Fator mais generoso para preencher o quadro (1 unidade Excel ≈ 7–8 px no viewer)
    return int(width * 12 + 4)


def _col_px_display(ws, letter: str) -> int:
    """Largura da coluna como o Excel desenha (~7 px por unidade). Usado no extent da âncora para a foto caber no merge."""
    w = ws.column_dimensions.get(letter)
    width = w.width if (w and w.width) else 10
    return max(1, int((width + 0.5) * 7))


def _row_px(ws, row_num: int) -> int:
    """Altura aproximada da linha em pixels (Excel height em pontos → px a 96 DPI)."""
    r = ws.row_dimensions.get(row_num)
    height = r.height if (r and r.height) else 20
    # 1 ponto = 96/72 px; mínimo 20 pt para áreas de foto
    return int(height * 96 / 72)


def _find_merged_range(ws, col: int, row: int) -> Tuple[int, int]:
    """Retorna (col_fim, row_fim) do merged range que contém (col, row). Se não mesclado, (col, row)."""
    for mr in ws.merged_cells.ranges:
        if mr.min_col <= col <= mr.max_col and mr.min_row <= row <= mr.max_row:
            return mr.max_col, mr.max_row
    return col, row


def _get_merge_topleft(ws, col: int, row: int) -> Tuple[int, int]:
    """Retorna (min_col, min_row) do merge que contém (col, row). Se não mesclado, (col, row). Usado para âncora da imagem no topo-esquerdo do merge."""
    for mr in ws.merged_cells.ranges:
        if mr.min_col <= col <= mr.max_col and mr.min_row <= row <= mr.max_row:
            return mr.min_col, mr.min_row
    return col, row


def _merged_range_px_raw(ws, cell_addr: str) -> Tuple[int, int]:
    """Retorna (largura_px, altura_px) do merged range com _col_px (usado em _merged_range_px para resize)."""
    col_letter = "".join(c for c in cell_addr if c.isalpha())
    row_num = int("".join(c for c in cell_addr if c.isdigit()))
    col_num = column_index_from_string(col_letter)
    col_fim, row_fim = _find_merged_range(ws, col_num, row_num)
    w = sum(_col_px(ws, get_column_letter(c)) for c in range(col_num, col_fim + 1))
    h = sum(_row_px(ws, r) for r in range(row_num, row_fim + 1))
    return max(w, 1), max(h, 1)


def _merged_range_px_extent(ws, cell_addr: str) -> Tuple[int, int]:
    """Tamanho do merge como o Excel desenha (_col_px_display), para extent da âncora — foto não ultrapassa o quadro."""
    col_letter = "".join(c for c in cell_addr if c.isalpha())
    row_num = int("".join(c for c in cell_addr if c.isdigit()))
    col_num = column_index_from_string(col_letter)
    col_fim, row_fim = _find_merged_range(ws, col_num, row_num)
    w = sum(_col_px_display(ws, get_column_letter(c)) for c in range(col_num, col_fim + 1))
    h = sum(_row_px(ws, r) for r in range(row_num, row_fim + 1))
    return max(w, 1), max(h, 1)


def _merged_range_px(ws, cell_addr: str) -> Tuple[int, int]:
    """Retorna (largura_px, altura_px) para redimensionar a imagem. Aplica mínimo para não virar micro foto."""
    w_raw, h_raw = _merged_range_px_raw(ws, cell_addr)
    w = max(w_raw, FOTO2LADOS_MIN_W_PX)
    h = max(h_raw, FOTO2LADOS_MIN_H_PX)
    return w, h


def _redimensionar_imagem_bytes(path_foto: str, largura: int, altura: int) -> bytes:
    """Redimensiona para (largura, altura) e retorna JPEG. Usa draft() em JPEG para menos RAM; resize() garante tamanho exato."""
    buf = io.BytesIO()
    with PILImage.open(path_foto) as img:
        if getattr(img, "format", None) == "JPEG" and (img.width > largura or img.height > altura):
            before = (img.width, img.height)
            try:
                img.draft("RGB", (largura, altura))
                _log_draft_ram(Path(path_foto).name, before, (img.width, img.height))
            except (AttributeError, TypeError, ValueError):
                pass
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")
        img_resized = img.resize((largura, altura), PILImage.Resampling.LANCZOS)
        img_resized.save(buf, format="JPEG", quality=90)
    return buf.getvalue()


class _ImageFromBytes(XLImg):
    """Imagem a partir de bytes (openpyxl lê no save(); arquivo temp sumiria)."""

    def __init__(self, data: bytes):
        super().__init__(io.BytesIO(data))
        self._bytes_data = data

    def _data(self):
        return self._bytes_data


def _inserir_imagem_preenchendo_celula(ws, cell_addr: str, path_foto: str) -> None:
    """
    Insere a imagem preenchendo o merged range (igual NC ARTESP).
    Âncora from = canto superior-esquerdo do merge; to = canto inferior-direito (ou mesmo célula + tamanho em EMU).
    """
    col_letter = "".join(c for c in cell_addr if c.isalpha())
    row_num = int("".join(c for c in cell_addr if c.isdigit()))
    col_num = column_index_from_string(col_letter)
    col_fim, row_fim = _find_merged_range(ws, col_num, row_num)
    col_ancora, row_ancora = _get_merge_topleft(ws, col_num, row_num)
    cell_ancora = f"{get_column_letter(col_ancora)}{row_ancora}"
    # Extent = tamanho do quadro em cm (6,35 × 8,31) para a foto caber exatamente
    w_emu = int(FOTO2LADOS_QUADRO_W_CM * EMU_PER_CM)
    h_emu = int(FOTO2LADOS_QUADRO_H_CM * EMU_PER_CM)
    w_px = max(int(FOTO2LADOS_QUADRO_W_CM * PX_PER_CM), 1)
    h_px = max(int(FOTO2LADOS_QUADRO_H_CM * PX_PER_CM), 1)
    data = _redimensionar_imagem_bytes(path_foto, w_px, h_px)
    xl_img = _ImageFromBytes(data)
    xl_img.width = w_px
    xl_img.height = h_px
    # OneCellAnchor: extent = quadro 6,35 × 8,31 cm
    anchor = OneCellAnchor()
    anchor._from = AnchorMarker(col=col_ancora - 1, colOff=0, row=row_ancora - 1, rowOff=0)
    anchor.ext = XDRPositiveSize2D(cx=w_emu, cy=h_emu)
    xl_img.anchor = anchor
    ws.add_image(xl_img)


def _celula_gravavel(ws, row, column):
    """Retorna célula gravável: topo-esquerdo do merge se (row,col) estiver em um merge."""
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= column <= mr.max_col:
            return ws.cell(row=mr.min_row, column=mr.min_col)
    return ws.cell(row=row, column=column)


def _escrever_em_celula_bloco(ws, row: int, col: int, value) -> None:
    """Escreve valor preservando o estilo (principalmente bordas) do template.

    - Se (row,col) estiver em merge, grava no topo-esquerdo do merge.
    - Faz snapshot do estilo completo e reaplica após escrever, evitando sumiço de bordas.
    """

    def _forcar_fonte_template(cell):
        try:
            f = getattr(cell, "font", None)
            if f is None:
                cell.font = Font(name="Arial", size=10)
                return
            nome = (getattr(f, "name", None) or "").strip().lower()
            tamanho = getattr(f, "size", None)
            if nome in ("", "calibri") or tamanho in (None, 11):
                cell.font = Font(
                    name="Arial",
                    size=10,
                    bold=getattr(f, "bold", False),
                    italic=getattr(f, "italic", False),
                    underline=getattr(f, "underline", None),
                    color=getattr(f, "color", None),
                )
        except Exception:
            pass

    c = _celula_gravavel(ws, row, col)
    val = value if value is not None else ""

    snap = {}
    for attr in ("font", "fill", "border", "alignment", "number_format", "protection"):
        try:
            snap[attr] = copy_obj(getattr(c, attr))
        except Exception:
            snap[attr] = None
    try:
        snap["_style"] = copy_obj(getattr(c, "_style", None))
    except Exception:
        snap["_style"] = None

    try:
        c.value = val
        for attr, v in snap.items():
            if v is None:
                continue
            try:
                if attr == "_style":
                    c._style = v
                else:
                    setattr(c, attr, v)
            except Exception:
                pass
        _forcar_fonte_template(c)

    except AttributeError:
        min_col, min_row = _get_merge_topleft(ws, col, row)
        from openpyxl.cell import Cell
        old = ws._cells.get((min_row, min_col))
        new_cell = Cell(ws, row=min_row, column=min_col, value=val)
        style_src = old or c
        for attr in ("font", "fill", "border", "alignment", "number_format", "protection"):
            try:
                setattr(new_cell, attr, copy_obj(getattr(style_src, attr)))
            except Exception:
                pass
        try:
            if getattr(style_src, "_style", None) is not None:
                new_cell._style = copy_obj(style_src._style)
        except Exception:
            pass
        ws._cells[(min_row, min_col)] = new_cell
        _forcar_fonte_template(new_cell)


def _fill_branco():
    # Branco sólido (sem canal alfa). '00FFFFFF' pode virar cinza em alguns Excel.
    return PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
def _forcar_fonte_arial10(ws, row: int, col: int) -> None:
    """Força fonte Arial 10 na célula alvo (resolvendo merge)."""
    try:
        c = _celula_gravavel(ws, row, col)
        c.font = Font(
            name="Arial",
            size=10,
            bold=getattr(c.font, "bold", False),
            italic=getattr(c.font, "italic", False),
            underline=getattr(c.font, "underline", None),
            color=getattr(c.font, "color", None),
        )
    except Exception:
        pass


def _normalizar_merge_rotulo_data(ws, row_data: int) -> None:
    """
    Ajusta apenas a linha do rótulo 'Data da Constatação' no bloco-base:
    - C:G -> C:F
    - J:N -> J:M
    e mantém rótulos alinhados à esquerda.
    """
    # Remove qualquer merge horizontal que intercepte a área do rótulo da data
    # (esquerda C..H, direita J..O), para evitar a data cair dentro do merge do rótulo.
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row == row_data and mr.max_row == row_data:
            try:
                overlap_left = (mr.min_col <= 8 and mr.max_col >= 3)
                overlap_right = (mr.min_col <= 15 and mr.max_col >= 10)
                if overlap_left or overlap_right:
                    ws.unmerge_cells(str(mr))
            except Exception:
                pass

    # Recria padrão fixo do rótulo em todos os blocos:
    # esquerda C:F (data em G, hora em H), direita J:M (data em N, hora em O)
    try:
        ws.merge_cells(start_row=row_data, start_column=3, end_row=row_data, end_column=6)
    except Exception:
        pass
    try:
        ws.merge_cells(start_row=row_data, start_column=10, end_row=row_data, end_column=13)
    except Exception:
        pass
    try:
        ws.cell(row=row_data, column=3).alignment = Alignment(horizontal="left", vertical="center")
    except Exception:
        pass
    try:
        ws.cell(row=row_data, column=10).alignment = Alignment(horizontal="left", vertical="center")
    except Exception:
        pass


def _copiar_estilo(src, dst):
    """Copia o estilo COMPLETO da célula src para dst.

    Em templates com bordas/mesclas, copiar apenas alguns atributos pode apagar bordas.
    Preserva fonte, preenchimento, bordas, alinhamento, formato numérico e proteção.
    Também tenta copiar o _style interno quando disponível.
    """
    for attr in ("font", "fill", "border", "alignment", "number_format", "protection"):
        v = getattr(src, attr, None)
        if v is not None:
            try:
                setattr(dst, attr, copy_obj(v))
            except Exception:
                pass
    try:
        if getattr(src, "_style", None) is not None:
            dst._style = copy_obj(src._style)
    except Exception:
        pass


def _replicar_bloco(ws, linha_origem: int, linha_destino: int,
                    n_linhas: int = TAMANHO_BLOCO) -> None:
    """Replica um bloco completo do template, incluindo a área da FOTO (linhas acima do rótulo).

    No modelo Foto 2 Lados, a foto fica em j-11 (linha 8 do 1º bloco) e o rótulo inicia em j (linha 19).
    Se copiar apenas 19–34, o novo bloco não herda as mesclas C8:H18/J8:O18 e a imagem vira miniatura.
    """

    linha_origem_top = linha_origem + LINHA_IMAGEM_NO_BLOCO
    linha_dest_top   = linha_destino + LINHA_IMAGEM_NO_BLOCO
    n_linhas_total   = n_linhas - LINHA_IMAGEM_NO_BLOCO if LINHA_IMAGEM_NO_BLOCO < 0 else n_linhas

    max_col = ws.max_column or 20

    # Copiar alturas
    for offset in range(n_linhas_total):
        r_orig = linha_origem_top + offset
        r_dest = linha_dest_top + offset
        if r_orig in ws.row_dimensions and ws.row_dimensions[r_orig].height is not None:
            ws.row_dimensions[r_dest].height = ws.row_dimensions[r_orig].height

    # Copiar células + valores fixos/fórmulas
    for offset in range(n_linhas_total):
        r_orig = linha_origem_top + offset
        r_dest = linha_dest_top + offset
        for col in range(1, max_col + 1):
            src = ws.cell(row=r_orig, column=col)
            dst = _celula_gravavel(ws, r_dest, col)
            _copiar_estilo(src, dst)
            if getattr(src, "value", None) not in (None, ""):
                try:
                    if isinstance(src.value, str) and src.value.startswith("="):
                        try:
                            dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
                        except Exception:
                            dst.value = src.value
                    else:
                        dst.value = src.value
                except Exception:
                    pass

    # Replicar mesclagens
    for mr in list(ws.merged_cells.ranges):
        if linha_origem_top <= mr.min_row and mr.max_row < linha_origem_top + n_linhas_total:
            off_min = mr.min_row - linha_origem_top
            off_max = mr.max_row - linha_origem_top
            try:
                ws.merge_cells(
                    start_row=linha_dest_top + off_min, start_column=mr.min_col,
                    end_row=linha_dest_top + off_max,   end_column=mr.max_col,
                )
            except Exception:
                pass


def preparar_fotos_para_relatorio(planilha_dados: str, pasta_destino: str,
                                   max_lado: int = 1200, qualidade: int = 85,
                                   log_cb: _LogCb = None) -> int:
    """
    Copia fotos para subpasta 'Fotos_YYYYMMDD_HHMM', redimensiona e renomeia.
    Atualiza coluna 'Caminho Relatório' (col 17) na planilha de dados.
    Retorna número de fotos únicas gravadas.
    """
    if not OPENPYXL_OK: raise ImportError("pip install openpyxl")
    if not PIL_OK:       raise ImportError("pip install pillow")

    Path(pasta_destino).mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    pasta_fotos = Path(pasta_destino) / f"Fotos_{ts}"
    pasta_fotos.mkdir(parents=True, exist_ok=True)
    _log(log_cb, f"[OK] Pasta de fotos: {pasta_fotos}")

    wb = openpyxl.load_workbook(planilha_dados)
    caminho_por_chave: dict = {}
    total = erros = reutilizadas = 0

    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        if not str(ws.cell(1, 1).value or "").startswith("Caminho"):
            continue
        _log(log_cb, f"[OK] Aba '{nome_aba}'")
        for row in range(2, ws.max_row + 1):
            caminho_orig = ws.cell(row, 1).value
            rodovia = str(ws.cell(row, 7).value or "").strip()
            km_v    = ws.cell(row, 8).value
            sentido = str(ws.cell(row, 9).value or "").strip()
            if not caminho_orig or not rodovia or km_v is None:
                continue
            try:
                km_fmt = f"{float(str(km_v).replace(',', '.')):.3f}"
            except Exception:
                km_fmt = str(km_v)
            chave = (rodovia, sentido, km_fmt)
            if chave in caminho_por_chave:
                ws.cell(row, COL_CAMINHO_RELATORIO).value = caminho_por_chave[chave]
                reutilizadas += 1
                continue
            orig = str(caminho_orig).strip()
            if not os.path.exists(orig):
                _log(log_cb, f"  [AVISO] Não encontrado: {Path(orig).name}")
                erros += 1
                continue
            rod_s  = _sanitizar_nome(rodovia) or "Rodovia"
            sent_s = _sanitizar_nome(sentido) or "Sentido"
            nome   = f"{rod_s} - {sent_s} - km {km_fmt}.jpg"
            dest   = str(pasta_fotos / nome)
            try:
                img = PILImage.open(orig)
                w, h = img.size
                if max(w, h) > max_lado:
                    if w >= h:
                        nw, nh = max_lado, int(h * max_lado / w)
                    else:
                        nw, nh = int(w * max_lado / h), max_lado
                    if getattr(img, "format", None) == "JPEG":
                        before = (img.width, img.height)
                        try:
                            img.draft("RGB", (nw, nh))
                            _log_draft_ram(Path(orig).name, before, (img.width, img.height))
                        except (AttributeError, TypeError, ValueError):
                            pass
                if img.mode in ("RGBA", "P"):
                    img = img.convert("RGB")
                if max(w, h) > max_lado:
                    img = img.resize((nw, nh), PILImage.Resampling.LANCZOS)
                img.save(dest, "JPEG", quality=min(95, max(50, qualidade)), optimize=True)
                caminho_por_chave[chave] = dest
                ws.cell(row, COL_CAMINHO_RELATORIO).value = dest
                total += 1
                if log_cb and total <= 30:
                    log_cb(f"  {Path(orig).name} → {nome}")
            except Exception as e:
                _log(log_cb, f"  [ERRO] {Path(orig).name}: {e}")
                erros += 1

    wb.save(planilha_dados)
    _log(log_cb, f"\n[OK] {total} fotos gravadas em {pasta_fotos}")
    if erros:
        _log(log_cb, f"[AVISO] {erros} erros.")
    return total


def gerar_relatorio_foto2lados(planilha_dados: str, modelo_xlsx: str,
                                pasta_saida: str, assunto: str,
                                log_cb: _LogCb = None) -> str:
    """
    Gera relatório fotográfico 2 lados. Retorna caminho do XLSX gerado.
    Também gera planilha de excluídas se houver fotos sem km ou com erro.
    """
    if not OPENPYXL_OK: raise ImportError("pip install openpyxl")

    wb_d = openpyxl.load_workbook(planilha_dados, data_only=True)
    registros = []
    for nome_aba in wb_d.sheetnames:
        ws_d = wb_d[nome_aba]
        if not str(ws_d.cell(1, 1).value or "").startswith("Caminho"):
            continue
        for row in range(2, ws_d.max_row + 1):
            foto         = ws_d.cell(row, 1).value
            rodov        = ws_d.cell(row, 7).value
            km_v         = ws_d.cell(row, 8).value
            # Sentido: col 9 (I) = preenchido por processar_coordenadas_km a partir da coluna C da Relação Total
            sent         = ws_d.cell(row, 9).value
            serv         = ws_d.cell(row, 12).value
            caminho_rel  = ws_d.cell(row, COL_CAMINHO_RELATORIO).value
            if not foto:
                continue
            caminho_usar = str(caminho_rel or "").strip()
            foto_usar = (caminho_usar if caminho_usar and os.path.exists(caminho_usar)
                         else str(foto).strip())
            registros.append({
                "rodovia": str(rodov or ""),
                "km":      str(km_v  or ""),
                "sentido": str(sent  or ""),
                "servico": str(serv  or ""),
                "foto":    foto_usar,
            })

    if not registros:
        raise ValueError("Nenhum registro encontrado na planilha.")

    registros_com_km = []
    excluidas = []
    for reg in registros:
        if reg["rodovia"].strip() and str(reg["km"]).strip():
            registros_com_km.append(reg)
        else:
            excluidas.append((reg["foto"], "Sem localização km"))

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    shutil.copy2(modelo_xlsx, tmp.name)
    wb_m = openpyxl.load_workbook(tmp.name)
    ws_m = wb_m.active
    # Normaliza o rótulo de data no bloco-base; os próximos blocos herdam pela replicação.
    _normalizar_merge_rotulo_data(ws_m, LINHA_INICIO_BLOCO + LINHA_DATA_NO_BLOCO)

    j  = LINHA_INICIO_BLOCO
    m  = 0
    fotos_erro = []

    # Preencher fotos em pares (esquerda/direita). Replicar bloco do template sempre que
    # formos usar um novo bloco (template tem ~20 blocos; se houver mais fotos, replicar).
    while m < len(registros_com_km):
        reg = registros_com_km[m]
        impar = (m % 2 == 0)

        foto_num = m + 1
        rodovia_txt = reg["rodovia"].strip()
        km_txt      = str(reg["km"]).strip()
        sentido_txt = reg["sentido"].strip()

        # Replicar bloco modelo (19–34) para a posição j quando necessário
        if impar and j > LINHA_INICIO_BLOCO:
            # O template já vem com vários blocos prontos (ex.: C35='Foto', C51='Foto'...).
            # Só replica quando o bloco NÃO existir ainda, evitando erro de mesclagem duplicada.
            if str(ws_m.cell(row=j, column=3).value or '').strip().lower() != 'foto':
                _replicar_bloco(ws_m, LINHA_INICIO_BLOCO, j, TAMANHO_BLOCO)

        def _gravar_foto(col_rodovia, col_km, col_sentido, col_servico, col_data, col_hora, col_img_letra):
            # Padrão macro/template: dados em j/j+1, data em j+3 e foto em j-11.
            _escrever_em_celula_bloco(ws_m, j + LINHA_DADOS_ABAIXO_FOTO, col_rodovia, rodovia_txt)
            _escrever_em_celula_bloco(ws_m, j + LINHA_DADOS_ABAIXO_FOTO, col_km,      km_txt)
            _escrever_em_celula_bloco(ws_m, j + LINHA_DADOS_ABAIXO_FOTO, col_sentido, sentido_txt)
            # Numeração sequencial (não depende das fórmulas do template)
            # Esquerda: coluna D (4); Direita: coluna K (11).
            if col_servico == COL_ESQ_SERVICO:
                _escrever_em_celula_bloco(ws_m, j + LINHA_DADOS_ABAIXO_FOTO, 4, foto_num)
            else:
                _escrever_em_celula_bloco(ws_m, j + LINHA_DADOS_ABAIXO_FOTO, 11, foto_num)
            _escrever_em_celula_bloco(ws_m, j + 1, col_servico, str(reg.get("servico") or ""))
            _forcar_fonte_arial10(ws_m, j + LINHA_DADOS_ABAIXO_FOTO, col_rodovia)
            _forcar_fonte_arial10(ws_m, j + LINHA_DADOS_ABAIXO_FOTO, col_km)
            _forcar_fonte_arial10(ws_m, j + LINHA_DADOS_ABAIXO_FOTO, col_sentido)
            _forcar_fonte_arial10(ws_m, j + 1, col_servico)
            data_f = exif_data_foto(reg["foto"]) if reg.get("foto") else ""
            data_val = ""
            hora_val = ""
            if data_f:
                parts = str(data_f).split(" ", 1)
                data_val = parts[0]
                hora_val = parts[1] if len(parts) > 1 else ""
            row_data = j + LINHA_DATA_NO_BLOCO
            # Normaliza a mesclagem da linha da data em C:F e J:M para TODOS os blocos
            _normalizar_merge_rotulo_data(ws_m, row_data)
            # Garante rótulo em todos os blocos (caso não venha na replicação do template)
            lbl = _celula_gravavel(ws_m, row_data, col_servico)
            if not str(lbl.value or "").strip():
                _escrever_em_celula_bloco(ws_m, row_data, col_servico, "Data da Constatação")
            _escrever_em_celula_bloco(ws_m, row_data, col_data, data_val)
            _escrever_em_celula_bloco(ws_m, row_data, col_hora, hora_val)
            _forcar_fonte_arial10(ws_m, row_data, col_data)
            _forcar_fonte_arial10(ws_m, row_data, col_hora)
            # Força alinhamento à esquerda na célula da data/hora
            c_data = _celula_gravavel(ws_m, row_data, col_data)
            c_data.alignment = Alignment(horizontal="left", vertical="center")
            c_hora = _celula_gravavel(ws_m, row_data, col_hora)
            c_hora.alignment = Alignment(horizontal="left", vertical="center")
            if os.path.exists(reg["foto"]):
                try:
                    cell_anchor = f"{col_img_letra}{j + LINHA_IMAGEM_NO_BLOCO}"
                    _inserir_imagem_preenchendo_celula(ws_m, cell_anchor, reg["foto"])
                    _log(log_cb, f"  Foto: {Path(reg['foto']).name}")
                except Exception as e:
                    fotos_erro.append((reg["foto"], str(e)))
            else:
                fotos_erro.append((reg["foto"], "Arquivo não encontrado"))

        if impar:
            _gravar_foto(COL_ESQ_RODOVIA, COL_ESQ_KM, COL_ESQ_SENTIDO, COL_ESQ_SERVICO, COL_ESQ_DATA, COL_ESQ_HORA, get_column_letter(COL_ESQ_FOTO))
        else:
            _gravar_foto(COL_DIR_RODOVIA, COL_DIR_KM, COL_DIR_SENTIDO, COL_DIR_SERVICO, COL_DIR_DATA, COL_DIR_HORA, get_column_letter(COL_DIR_FOTO))
            j += TAMANHO_BLOCO
        m += 1

    # Fundo branco na área de conteúdo dos blocos
    fill_b = _fill_branco()
    linha_fim = (j - 1) if j > LINHA_INICIO_BLOCO else (LINHA_INICIO_BLOCO + TAMANHO_BLOCO - 1)
    linha_fim = max(linha_fim, ws_m.max_row or 0)
    for row in range(LINHA_INICIO_BLOCO + LINHA_IMAGEM_NO_BLOCO, linha_fim + 1):
        for col in range(1, 51):
            try:
                _celula_gravavel(ws_m, row, col).fill = fill_b
            except Exception:
                pass

    Path(pasta_saida).mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d-%H%M")
    caminho_saida = str(Path(pasta_saida) / f"{ts} - {assunto}.xlsx")

    aba_principal = ws_m.title
    for nome_aba in list(wb_m.sheetnames):
        if nome_aba != aba_principal:
            del wb_m[nome_aba]
    wb_m.save(caminho_saida)

    try:
        os.unlink(tmp.name)
    except Exception:
        pass

    for c, motivo in fotos_erro:
        excluidas.append((c, motivo))

    if excluidas:
        wb_ex = openpyxl.Workbook()
        ws_ex = wb_ex.active
        ws_ex.title = "Excluidas"
        hf = PatternFill("solid", fgColor="4A4E69")
        hfont = Font(bold=True, color="FFFFFF")
        for col, titulo in enumerate(["Caminho Foto", "Nome Arquivo", "Motivo"], 1):
            c = ws_ex.cell(row=1, column=col, value=titulo)
            c.fill = hf;  c.font = hfont
        for idx, (caminho, motivo) in enumerate(excluidas, 2):
            ws_ex.cell(idx, 1, caminho)
            ws_ex.cell(idx, 2, Path(caminho).name)
            ws_ex.cell(idx, 3, motivo)
        ws_ex.column_dimensions["A"].width = 70
        ws_ex.column_dimensions["B"].width = 45
        ws_ex.column_dimensions["C"].width = 28
        wb_ex.save(str(Path(pasta_saida) / f"{ts} - Excluidas.xlsx"))

    _log(log_cb, f"\n[OK] Relatório: {caminho_saida}")
    _log(log_cb, f"[OK] Incluídas: {len(registros_com_km)}  Excluídas: {len(excluidas)}")
    return caminho_saida


def relatorio_foto2lados_bytes(xlsx_dados_bytes: bytes, xlsx_modelo_bytes: bytes,
                                zip_fotos_bytes: Optional[bytes], assunto: str,
                                log_cb: _LogCb = None) -> bytes:
    """
    Versão web do Módulo 4.
    Se zip_fotos_bytes for fornecido, extrai as fotos para temp e ajusta os caminhos no XLSX.
    Retorna ZIP com: relatório XLSX + excluídas XLSX (se houver).
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        p_dados  = Path(tmpdir) / "dados.xlsx"
        p_modelo = Path(tmpdir) / "modelo.xlsx"
        p_saida  = Path(tmpdir) / "saida"
        p_saida.mkdir()
        p_dados.write_bytes(xlsx_dados_bytes)
        p_modelo.write_bytes(xlsx_modelo_bytes)

        if zip_fotos_bytes:
            pasta_fotos = Path(tmpdir) / "fotos"
            pasta_fotos.mkdir()
            with zipfile.ZipFile(io.BytesIO(zip_fotos_bytes)) as zf:
                _extrair_zip_para_pasta(zf, pasta_fotos)
            # Lista de imagens em ordem determinística (para fallback por posição — primeiro bloco)
            _extensoes_foto = (".jpg", ".jpeg", ".png", ".JPG", ".JPEG", ".PNG")
            lista_arquivos_ordem = sorted(
                p for p in pasta_fotos.rglob("*")
                if p.is_file() and p.suffix in _extensoes_foto
            )
            # Ajusta caminhos no XLSX para apontar para o temp dir (prioridade: primeiro bloco)
            wb_adj = openpyxl.load_workbook(str(p_dados))
            for sn in wb_adj.sheetnames:
                ws_adj = wb_adj[sn]
                if not str(ws_adj.cell(1, 1).value or "").startswith("Caminho"):
                    continue
                max_row = ws_adj.max_row or 0
                # Garantir primeiro bloco: linhas 2 e 3 recebem sempre os 2 primeiros arquivos do ZIP
                if len(lista_arquivos_ordem) >= 2 and max_row >= 3:
                    ws_adj.cell(2, 1).value = str(lista_arquivos_ordem[0])
                    ws_adj.cell(2, COL_CAMINHO_RELATORIO).value = str(lista_arquivos_ordem[0])
                    ws_adj.cell(3, 1).value = str(lista_arquivos_ordem[1])
                    ws_adj.cell(3, COL_CAMINHO_RELATORIO).value = str(lista_arquivos_ordem[1])
                for r in range(2, max_row + 1):
                    # Linhas 2 e 3 já garantidas para o primeiro bloco; não sobrescrever
                    if r <= 3:
                        continue
                    nome_arq = ws_adj.cell(r, 2).value
                    caminho_cel = ws_adj.cell(r, 1).value
                    encontrado = None
                    if nome_arq:
                        nome_busca = str(nome_arq).strip()
                        for p in pasta_fotos.rglob("*"):
                            if p.is_file() and p.name == nome_busca:
                                encontrado = p
                                break
                        if not encontrado:
                            for p in pasta_fotos.rglob("*"):
                                if p.is_file() and p.name.lower() == nome_busca.lower():
                                    encontrado = p
                                    break
                    if not encontrado and caminho_cel:
                        nome_de_caminho = Path(str(caminho_cel).strip()).name
                        if nome_de_caminho:
                            for p in pasta_fotos.rglob("*"):
                                if p.is_file() and (p.name == nome_de_caminho or p.name.lower() == nome_de_caminho.lower()):
                                    encontrado = p
                                    break
                    if not encontrado and caminho_cel:
                        # Caminho relativo ao ZIP extraído (ex.: fotos/IMG_001.jpg)
                        try:
                            p_rel = (pasta_fotos / str(caminho_cel).strip().lstrip("/\\")).resolve()
                            if p_rel.is_file() and p_rel.resolve().is_relative_to(pasta_fotos.resolve()):
                                encontrado = p_rel
                        except Exception:
                            pass
                    if encontrado:
                        ws_adj.cell(r, 1).value = str(encontrado)
                        ws_adj.cell(r, COL_CAMINHO_RELATORIO).value = str(encontrado)
                # Fallback por posição: linhas sem arquivo válido recebem os primeiros arquivos da lista (evita foto 1 e 2 faltando no primeiro bloco)
                idx_arquivo = 0
                for r in range(2, max_row + 1):
                    if idx_arquivo >= len(lista_arquivos_ordem):
                        break
                    path_atual = ws_adj.cell(r, 1).value
                    try:
                        existe = path_atual and Path(str(path_atual).strip()).exists()
                    except Exception:
                        existe = False
                    if not existe:
                        p = lista_arquivos_ordem[idx_arquivo]
                        ws_adj.cell(r, 1).value = str(p)
                        ws_adj.cell(r, COL_CAMINHO_RELATORIO).value = str(p)
                        idx_arquivo += 1
            wb_adj.save(str(p_dados))

        gerar_relatorio_foto2lados(str(p_dados), str(p_modelo), str(p_saida),
                                   assunto, log_cb=log_cb)

        # Empacotar saída em ZIP
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf_out:
            for f in p_saida.iterdir():
                zf_out.write(f, f.name)
        return buf.getvalue()


# MÓDULO 5 – EXPORTAR PARA KCOR-KRIA
# Colunas espelho macro M03 (Art_031) / nc_artesp.modulos.inserir_nc_kria — conservação.

MAPA_RODOVIAS: Dict[str, str] = {
    "SP 075": "SP075",  "SP 127": "SP127",  "SP 280": "SP280",
    "SP 300": "SP300",  "SP 102": "SPI102/300",
    "CP 147": "FORA",   "CP 308": "FORA",
}

_K_A, _K_B, _K_C, _K_D, _K_E = 1, 2, 3, 4, 5
_K_F, _K_G, _K_H, _K_I, _K_J = 6, 7, 8, 9, 10
_K_K, _K_L, _K_M, _K_N, _K_O = 11, 12, 13, 14, 15
_K_P, _K_Q, _K_R, _K_S, _K_T = 16, 17, 18, 19, 20
_K_U, _K_V, _K_W, _K_X, _K_Y = 21, 22, 23, 24, 25

_COLS_KCOR_M03 = tuple(range(1, 26))


def _exif_dd_mm_yyyy(exif_full: str) -> str:
    if not exif_full:
        return ""
    return str(exif_full).strip().split()[0]


def _servico_para_kcor_m03(desc_serv: str) -> Tuple[str, str, str]:
    """(tipo col E, classificacao D, executor L) como SERVICO_NC / default M03."""
    d = (desc_serv or "").strip()
    try:
        from nc_artesp.config import SERVICO_NC
        if d in SERVICO_NC:
            sn, cl, ex = SERVICO_NC[d]
            return (sn or "")[:120], cl, ex
        dl = d.lower()
        for k, (sn, cl, ex) in SERVICO_NC.items():
            if k.lower() == dl or k.lower() in dl or dl in k.lower():
                return (sn or "")[:120], cl, ex
    except Exception:
        pass
    if d:
        return d[:120], "Conservação Rotina", "Soluciona - Conserva"
    return "", "Conservação Rotina", "Soluciona - Conserva"


def _sentido_kcor_m03(sent: str) -> str:
    try:
        from nc_artesp.modulos.analisar_pdf_ma import _sentido_para_texto
        return (_sentido_para_texto(sent or "") or "").strip()[:120]
    except Exception:
        return (sent or "").strip()[:120]


def _kcor_unmerge_linha(ws, row: int, col_ini: int, col_fim: int) -> None:
    try:
        from openpyxl.utils.cell import range_boundaries
    except ImportError:
        return
    for mc in list(ws.merged_cells.ranges):
        try:
            min_c, min_r, max_c, max_r = range_boundaries(str(mc))
        except Exception:
            continue
        if min_r <= row <= max_r and not (col_fim < min_c or col_ini > max_c):
            try:
                ws.unmerge_cells(str(mc))
            except Exception:
                pass


def exportar_kcor(planilha_dados: str, modelo_kcor: str,
                  pasta_saida: str, assunto: str,
                  mapa_rodovias: Optional[dict] = None,
                  log_cb: _LogCb = None) -> str:
    """
    Exporta planilha Módulos 1–3 (Caminho col A…) para KCor-Kria.
    Preenche A–Y como macro M03 conservação (inserir_nc_kria).
    """
    if not OPENPYXL_OK:
        raise ImportError("pip install openpyxl")
    mapa = mapa_rodovias or MAPA_RODOVIAS

    wb_d = openpyxl.load_workbook(planilha_dados, data_only=True)
    todos_registros: List[dict] = []

    for nome_aba in wb_d.sheetnames:
        ws = wb_d[nome_aba]
        a1 = ws.cell(1, 1).value
        if not a1 or "Caminho" not in str(a1):
            continue
        _log(log_cb, f"[OK] Aba '{nome_aba}': {max(0, ws.max_row - 1)} linhas")
        for row in range(2, ws.max_row + 1):
            caminho = ws.cell(row, 1).value
            if not caminho:
                continue
            caminho_s = str(caminho).strip()
            rodov = str(ws.cell(row, 7).value or "").strip()
            km_i = str(ws.cell(row, 8).value or "").strip()
            sent = str(ws.cell(row, 9).value or "").strip()
            serv = str(ws.cell(row, 12).value or "").strip()
            titulo = str(ws.cell(row, 16).value or "").strip()
            exif_f = exif_data_foto(caminho_s)
            dt_str = _exif_dd_mm_yyyy(exif_f)
            pasta13 = str(ws.cell(row, 13).value or "").strip()
            nome15 = str(ws.cell(row, 15).value or "").strip()
            col11 = ws.cell(row, 11).value
            dest_renomeado = str(col11).strip() if col11 else ""

            rodov_kcor = mapa.get(rodov[:6], rodov)
            if dest_renomeado and Path(dest_renomeado).is_file():
                dir_v = str(Path(dest_renomeado).resolve().parent)
                arq_w = Path(dest_renomeado).name
            elif dest_renomeado and Path(dest_renomeado).is_dir():
                dir_v = dest_renomeado.rstrip("\\/")
                arq_w = nome15 or Path(caminho_s).name
            else:
                dir_v = pasta13 or (str(Path(caminho_s).parent) if caminho_s else "")
                arq_w = nome15 or Path(caminho_s).name

            serv_nc, classifica, executor = _servico_para_kcor_m03(serv)

            todos_registros.append({
                "Rodovia": rodov_kcor,
                "KMi": km_i,
                "KMf": km_i,
                "Sentido": _sentido_kcor_m03(sent),
                "serv_nc": serv_nc,
                "classifica": classifica,
                "executor": executor,
                "dt_str": dt_str,
                "titulo": titulo,
                "dir_v": dir_v,
                "arq_w": arq_w,
                "serv_raw": serv,
            })

    if not todos_registros:
        raise ValueError("Nenhum registro encontrado na planilha.")

    Path(pasta_saida).mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d-%H%M")
    caminho_saida = str(Path(pasta_saida) / f"{ts} - {assunto}.xlsx")

    wb_k = openpyxl.load_workbook(modelo_kcor)
    ws_k = None
    for sheet in wb_k.worksheets:
        a1 = sheet.cell(1, 1).value
        if a1 is not None and "numitem" in str(a1).strip().lower():
            ws_k = sheet
            break
    if ws_k is None:
        ws_k = wb_k.active

    def _tem_borda(cell) -> bool:
        b = getattr(cell, "border", None)
        if b is None:
            return False
        try:
            return any(getattr(side, "style", None) is not None
                       for side in (b.left, b.right, b.top, b.bottom))
        except Exception:
            return False

    linha_ref = 2 if ws_k.max_row >= 2 else 1
    for rr in range(2, (ws_k.max_row or 2) + 1):
        if any(_tem_borda(ws_k.cell(rr, cc)) for cc in _COLS_KCOR_M03):
            linha_ref = rr
            break

    borda_ref = {}
    for cc in _COLS_KCOR_M03:
        ref_cell = ws_k.cell(linha_ref, cc)
        if getattr(ref_cell, "border", None):
            try:
                borda_ref[cc] = copy_obj(ref_cell.border)
            except Exception:
                pass

    for seq, reg in enumerate(todos_registros, 1):
        r = seq + 1
        dt = reg["dt_str"]
        obs_u = " ".join(x for x in (reg.get("serv_raw"), reg.get("titulo")) if x).strip()

        _kcor_unmerge_linha(ws_k, r, _K_Q, _K_T)

        ws_k.cell(r, _K_A).value = seq
        ws_k.cell(r, _K_B).value = "Artesp"
        ws_k.cell(r, _K_C).value = "2"
        ws_k.cell(r, _K_D).value = reg["classifica"]
        ws_k.cell(r, _K_E).value = reg["serv_nc"]
        ws_k.cell(r, _K_F).value = reg["Rodovia"]
        ws_k.cell(r, _K_G).value = reg["KMi"]
        ws_k.cell(r, _K_H).value = reg["KMf"]
        ws_k.cell(r, _K_I).value = reg["Sentido"]
        ws_k.cell(r, _K_J).value = ""
        ws_k.cell(r, _K_K).value = "Conservação"
        ws_k.cell(r, _K_L).value = reg["executor"]
        ws_k.cell(r, _K_M).value = dt
        ws_k.cell(r, _K_N).value = ""
        ws_k.cell(r, _K_O).value = dt
        ws_k.cell(r, _K_P).value = dt
        ws_k.cell(r, _K_Q).value = ""
        ws_k.cell(r, _K_R).value = ""
        ws_k.cell(r, _K_S).value = ""
        ws_k.cell(r, _K_T).value = reg["titulo"]
        ws_k.cell(r, _K_U).value = obs_u
        ws_k.cell(r, _K_V).value = reg["dir_v"]
        ws_k.cell(r, _K_W).value = reg["arq_w"]
        ws_k.cell(r, _K_X).value = ""
        ws_k.cell(r, _K_Y).value = ""

        for col in _COLS_KCOR_M03:
            dst = ws_k.cell(r, col)
            if col in borda_ref:
                try:
                    dst.border = copy_obj(borda_ref[col])
                except Exception:
                    pass

    wb_k.active = ws_k
    wb_k.save(caminho_saida)
    _log(log_cb, f"\n[OK] KCor-Kria (layout M03): {caminho_saida}  ({len(todos_registros)} registros)")
    return caminho_saida


def exportar_kcor_bytes(xlsx_dados_bytes: bytes, xlsx_modelo_bytes: bytes,
                         assunto: str, log_cb: _LogCb = None) -> bytes:
    """Versão web do Módulo 5: retorna XLSX em bytes."""
    with tempfile.TemporaryDirectory() as tmpdir:
        p_dados  = Path(tmpdir) / "dados.xlsx"
        p_modelo = Path(tmpdir) / "modelo.xlsx"
        p_saida  = Path(tmpdir) / "saida"
        p_saida.mkdir()
        p_dados.write_bytes(xlsx_dados_bytes)
        p_modelo.write_bytes(xlsx_modelo_bytes)
        resultado = exportar_kcor(str(p_dados), str(p_modelo), str(p_saida),
                                  assunto, log_cb=log_cb)
        return Path(resultado).read_bytes()
