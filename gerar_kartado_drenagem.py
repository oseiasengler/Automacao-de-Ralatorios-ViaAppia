"""
Gera arquivos Excel no formato Kartado a partir do banco SQLite do APP_Inventarios.
Inclui as fotos associadas e empacota tudo em um ZIP para upload no Kartado.

Uso:
    python gerar_kartado_drenagem.py
    python gerar_kartado_drenagem.py --rodovia SP-280
    python gerar_kartado_drenagem.py --db /caminho/coleta.db --output /caminho/saida
    python gerar_kartado_drenagem.py --fotos-dir /caminho/fotos
"""
import argparse
import json
import os
import shutil
import sqlite3
import zipfile
from datetime import date, datetime
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Configuração: mapeamento tipo PWA → bucket Kartado
# ---------------------------------------------------------------------------

TIPO_PARA_BUCKET = {
    "canaleta":          "superficial",
    "sarjeta":           "superficial",
    "valeta_protecao":   "superficial",
    "meio_fio":          "superficial",
    "descida_dagua":     "escadas",
    "dreno_longitudinal":"drenos",
    "espinha_peixe":     "drenos",
    "dhp":               "drenos",
    "colchao_drenante":  "drenos",
    "caixa_coletora":    "caixas_alas",
    "saida_dagua":       "caixas_alas",
    "dissipador_energia":"caixas_alas",
    "ala":               "caixas_alas",
    "bueiro_tubular":    "tubulacao",
    "bueiro_celular":    "tubulacao",
}

NOME_ARQUIVO_SAIDA = {
    "superficial":  "Drenagem_Superficial_Kartado.xlsx",
    "drenos":       "Drenos_Kartado.xlsx",
    "escadas":      "Escadas_Hidraulica_Kartado.xlsx",
    "caixas_alas":  "Caixas_Alas_Kartado.xlsx",
    "tubulacao":    "Tubulacao_Aduelas_Kartado.xlsx",
}

NOME_TEMPLATE = {
    "superficial":  "Inv. - Drenagem Superficial.xlsx",
    "drenos":       "Inv. - Drenagem - Drenos.xlsx",
    "escadas":      "Inv. - Drenagem - Escadas Hidraulica Rápidos.xlsx",
    "caixas_alas":  "Inv. - Drenagem Profunda - Caixas_Alas.xlsx",
    "tubulacao":    "Inv. - Drenagem Profunda - Tubulação_Aduelas.xlsx",
}

# ---------------------------------------------------------------------------
# Leitura do banco SQLite
# ---------------------------------------------------------------------------

def ler_dispositivos(db_path: str, rodovia: str | None = None) -> list[dict]:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    sql = "SELECT * FROM dispositivo WHERE deleted = 0"
    params = []
    if rodovia:
        sql += " AND rodovia = ?"
        params.append(rodovia)
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    registros = []
    for row in rows:
        r = dict(row)
        for campo_json in ("atributos", "fotos", "ncs"):
            if isinstance(r.get(campo_json), str):
                try:
                    r[campo_json] = json.loads(r[campo_json])
                except (json.JSONDecodeError, TypeError):
                    r[campo_json] = {} if campo_json == "atributos" else []
        registros.append(r)
    return registros


# ---------------------------------------------------------------------------
# Helpers de mapeamento
# ---------------------------------------------------------------------------

def _comum(r: dict) -> dict:
    """Campos presentes em todos os templates Kartado."""
    d = {}
    d["Código do Inventário para vinculo com apontamento"] = r.get("id", "")
    d["Latitude"]  = r.get("lat_ini", "")
    d["Longitude"] = r.get("lon_ini", "")
    d["km"]        = r.get("km_ini", "")
    d["km final"]  = r.get("km_fim", "")
    d["Sentido"]   = r.get("sentido", "")
    d["Rodovia"]   = r.get("rodovia", "")
    d["Encontrado em"] = r.get("data_inspecao", "")
    d["Observações"]   = r.get("observacoes", "")
    fotos = r.get("fotos") or []
    for i, f in enumerate(fotos[:10], start=1):
        d[f"Foto_{i}"] = f
    return d


def map_superficial(r: dict) -> dict:
    a = r.get("atributos") or {}
    d = _comum(r)
    d["Tipo de Drenagem"]          = r.get("tipo", "")
    d["Material Superfície Drenagem"] = a.get("material", "")
    d["Extensão Drenagem"]         = r.get("extensao_m", "")
    d["Largura Lado 1"]            = a.get("largura_cm", "")
    d["Largura Lado 2 ou Espelho"] = a.get("largura2_cm", "")
    d["Largura Fundo Base"]        = a.get("largura_fundo_cm", "")
    d["Altura Fundo"]              = a.get("altura_cm", "")
    d["Espessura Concreto"]        = a.get("espessura_cm", "")
    d["Km Montante"]               = r.get("km_ini", "")
    d["Latitude Montante"]         = r.get("lat_ini", "")
    d["Longitude Montante"]        = r.get("lon_ini", "")
    d["Lado da Via Montante"]      = r.get("lado", "")
    d["Km Jusante"]                = r.get("km_fim", "")
    d["Latitude Jusante"]          = r.get("lat_fim", "")
    d["Longitude Jusante"]         = r.get("lon_fim", "")
    d["Lado da Via Jusante"]       = r.get("lado", "")
    d["Tipo De Dispositivo Entrada Montante"] = a.get("tipo_entrada_montante", "")
    d["Tipo de Dispositivo Saída Jusante"]    = a.get("tipo_saida_jusante", "")
    d["Tipo de Dissipador Jusante"]           = a.get("tipo_dissipador", "")
    return d


def map_drenos(r: dict) -> dict:
    a = r.get("atributos") or {}
    d = _comum(r)
    d["Lado da Via"]               = r.get("lado", "")
    d["Tipo de Ativos Dreno"]      = r.get("tipo", "")
    d["Tipo de Dispositivo Saída Jusante"] = a.get("tipo_saida_jusante", "")
    d["Tipo de Dissipador Jusante"]        = a.get("tipo_dissipador", "")
    d["Km Montante"]               = r.get("km_ini", "")
    d["Latitude Montante"]         = r.get("lat_ini", "")
    d["Longitude Montante"]        = r.get("lon_ini", "")
    d["Lado da Via Montante"]      = r.get("lado", "")
    d["Km Jusante"]                = r.get("km_fim", "")
    d["Latitude Jusante"]          = r.get("lat_fim", "")
    d["Longitude Jusante"]         = r.get("lon_fim", "")
    d["Lado da Via Jusante"]       = r.get("lado", "")
    d["Largura Caixa Dreno"]       = a.get("largura_cm", "")
    d["Altura Caixa Dreno"]        = a.get("profundidade_cm", "")
    d["Diametro Tubo"]             = a.get("diametro_tubo_mm", "")
    return d


def map_escadas(r: dict) -> dict:
    a = r.get("atributos") or {}
    d = _comum(r)
    d["Lado da Via"]                    = r.get("lado", "")
    d["Tipo Escada Rápida"]             = a.get("tipo_descida", r.get("tipo", ""))
    d["Tipo Material Escadas Hidraulica"] = a.get("material", "")
    d["Extensão"]                       = r.get("extensao_m", "")
    d["Quantidade De Degraus"]          = a.get("n_degraus", "")
    d["Altura Média Degraus"]           = a.get("altura_espelho_cm", "")
    d["Largura Base Degraus"]           = a.get("largura_cm", "")
    d["Profundidade Base Degraus"]      = a.get("profundidade_cm", "")
    d["Altura Parede Lateral"]          = a.get("altura_parede_cm", "")
    d["Espessura Parede Lateral"]       = a.get("espessura_parede_cm", "")
    d["Largura Abertura Superior"]      = a.get("largura_abertura_sup_cm", "")
    d["Largura Base"]                   = a.get("largura_base_cm", "")
    d["Altura"]                         = a.get("altura_cm", "")
    d["Espessura"]                      = a.get("espessura_cm", "")
    d["Tipo de Dispositivo Saída Jusante"] = a.get("tipo_saida_jusante", "")
    d["Tipo de Dissipador Jusante"]        = a.get("tipo_dissipador", "")
    return d


def map_caixas_alas(r: dict) -> dict:
    a = r.get("atributos") or {}
    d = _comum(r)
    d["Lado da Via"]            = r.get("lado", "")
    d["Tipo de Caixa Ala"]      = r.get("tipo", "")
    d["Tipo Material Caixa Ala"] = a.get("material", "")
    d["Largura"]                = a.get("largura_cm", "")
    d["Comprimento"]            = a.get("comprimento_cm", "")
    d["Altura"]                 = a.get("profundidade_cm", a.get("altura_cm", ""))
    d["Espessura"]              = a.get("espessura_cm", "")
    d["Descrição Caixa Alas"]   = a.get("descricao", "")
    return d


def map_tubulacao(r: dict) -> dict:
    a = r.get("atributos") or {}
    d = _comum(r)
    tipo_bueiro = {
        "bueiro_tubular": "Tubular",
        "bueiro_celular": "Celular",
    }.get(r.get("tipo", ""), r.get("tipo", ""))
    d["Tipo de Bueiro"]                   = tipo_bueiro
    d["Tipo Tubulação Material"]          = a.get("material", "")
    d["Extensão da Linha de Tubo"]        = r.get("extensao_m", "")
    d["Lado da Via"]                      = r.get("lado", "")
    d["Latitude Montante"]                = r.get("lat_ini", "")
    d["Longitude Montante"]               = r.get("lon_ini", "")
    d["Lado da Via Montante"]             = r.get("lado", "")  # mesma via, ambas extremidades
    d["Largura Externo Seção Montante"]   = a.get("largura_cm", "")
    d["Altura Externo Seção Montante"]    = a.get("altura_cm", "")
    d["Diametro Externo Seção Montante"]  = a.get("diametro_mm", "")
    d["Espessura Parede Montante"]        = a.get("espessura_parede_mm", "")
    d["Tipo De Dispositivo Entrada Montante"] = a.get("tipo_entrada", "")
    d["Km Jusante"]                       = r.get("km_fim", "")
    d["Latitude Jusante"]                 = r.get("lat_fim", "")
    d["Longitude Jusante"]                = r.get("lon_fim", "")
    d["Lado da Via Jusante"]              = r.get("lado", "")
    d["Largura Externo Seção Jusante"]    = a.get("largura_cm", "")
    d["Altura Externo Seção Jusante"]     = a.get("altura_cm", "")
    d["Diametro Externo Seção Jusante"]   = a.get("diametro_mm", "")
    d["Espessura Parede Jusante"]         = a.get("espessura_parede_mm", "")
    d["Tipo de Dispositivo Saída Jusante"] = a.get("tipo_saida_jusante", "")
    d["Tipo de Dissipador Jusante"]       = a.get("tipo_dissipador", "")
    n_linhas = a.get("n_linhas", "")
    n_celulas = a.get("n_celulas", "")
    descr_parts = []
    if n_linhas:
        descr_parts.append(f"{n_linhas} linhas")
    if n_celulas:
        descr_parts.append(f"{n_celulas} células")
    if a.get("obstrucao_pct"):
        descr_parts.append(f"obstrução {a['obstrucao_pct']}%")
    d["Descrição Drenagem Profunda Tubulação Aduelas"] = "; ".join(descr_parts)
    return d


MAPEADORES = {
    "superficial": map_superficial,
    "drenos":      map_drenos,
    "escadas":     map_escadas,
    "caixas_alas": map_caixas_alas,
    "tubulacao":   map_tubulacao,
}


# ---------------------------------------------------------------------------
# Geração do Excel
# ---------------------------------------------------------------------------

def ler_cabecalho_template(template_path: str) -> list[str]:
    wb = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    ws = wb.active
    header = [str(c) if c is not None else "" for c in next(ws.iter_rows(max_row=1, values_only=True))]
    wb.close()
    return header


def gerar_excel(template_path: str, linhas: list[dict], output_path: str) -> None:
    header = ler_cabecalho_template(template_path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for linha in linhas:
        row = [linha.get(col, "") for col in header]
        ws.append(row)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


# ---------------------------------------------------------------------------
# Coleta de fotos
# ---------------------------------------------------------------------------

def coletar_fotos(registros: list[dict], fotos_dir: str | None) -> dict[str, str]:
    """Retorna {nome_arquivo: caminho_absoluto} para fotos encontradas no sistema."""
    encontradas = {}
    if not fotos_dir or not os.path.isdir(fotos_dir):
        return encontradas
    for r in registros:
        for foto in (r.get("fotos") or []):
            nome = os.path.basename(foto)
            candidato = os.path.join(fotos_dir, nome)
            if os.path.isfile(candidato):
                encontradas[nome] = candidato
            elif os.path.isfile(foto):
                encontradas[os.path.basename(foto)] = foto
    return encontradas


# ---------------------------------------------------------------------------
# ZIP final
# ---------------------------------------------------------------------------

def empacotar_zip(excels: dict[str, str], fotos: dict[str, str], zip_path: str) -> None:
    """Cria o ZIP com os Excel e a pasta fotos/."""
    os.makedirs(os.path.dirname(zip_path), exist_ok=True)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for nome_destino, caminho_excel in excels.items():
            zf.write(caminho_excel, nome_destino)
        for nome_foto, caminho_foto in fotos.items():
            zf.write(caminho_foto, f"fotos/{nome_foto}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Gera Excel Kartado a partir do banco do APP_Inventarios")
    parser.add_argument("--db",        default="coleta.db",       help="Caminho para o SQLite (padrão: coleta.db)")
    parser.add_argument("--output",    default="Output_Kartado",  help="Diretório de saída (padrão: Output_Kartado)")
    parser.add_argument("--rodovia",   default=None,              help="Filtrar por rodovia (ex.: SP-280)")
    parser.add_argument("--fotos-dir", default=None,              help="Diretório com as fotos capturadas")
    parser.add_argument("--templates", default="Templates/Inventarios", help="Diretório dos templates Excel")
    args = parser.parse_args()

    db_path = args.db
    if not os.path.isfile(db_path):
        print(f"[ERRO] Banco não encontrado: {db_path}")
        return

    print(f"[INFO] Lendo registros de '{db_path}'" + (f" (rodovia={args.rodovia})" if args.rodovia else ""))
    registros = ler_dispositivos(db_path, args.rodovia)
    print(f"[INFO] {len(registros)} registros carregados")

    # Agrupa por bucket
    grupos: dict[str, list[dict]] = {b: [] for b in MAPEADORES}
    ignorados = []
    for r in registros:
        bucket = TIPO_PARA_BUCKET.get(r.get("tipo", ""))
        if bucket:
            dados = MAPEADORES[bucket](r)
            grupos[bucket].append(dados)
        else:
            ignorados.append(r.get("tipo", "?"))

    if ignorados:
        print(f"[AVISO] Tipos sem mapeamento (ignorados): {set(ignorados)}")

    # Diretório de saída com data
    hoje = date.today().isoformat()
    sufixo = f"_{args.rodovia}" if args.rodovia else ""
    pasta_saida = os.path.join(args.output, f"{hoje}{sufixo}")

    # Gera os Excel
    excels_gerados: dict[str, str] = {}
    for bucket, linhas in grupos.items():
        if not linhas:
            continue
        nome_template = NOME_TEMPLATE[bucket]
        template_path = os.path.join(args.templates, nome_template)
        if not os.path.isfile(template_path):
            print(f"[AVISO] Template não encontrado: {template_path}")
            continue
        nome_saida = NOME_ARQUIVO_SAIDA[bucket]
        output_path = os.path.join(pasta_saida, nome_saida)
        gerar_excel(template_path, linhas, output_path)
        excels_gerados[nome_saida] = output_path
        print(f"[OK]   {nome_saida}  ->  {len(linhas)} registros")

    if not excels_gerados:
        print("[INFO] Nenhum Excel gerado (sem dados de drenagem no banco).")
        return

    # Coleta fotos
    fotos = coletar_fotos(registros, args.fotos_dir)
    if fotos:
        print(f"[INFO] {len(fotos)} foto(s) encontrada(s) para incluir no ZIP")
    else:
        print("[INFO] Nenhuma foto encontrada (ZIP conterá apenas os Excel)")

    # Empacota ZIP
    zip_name = f"kartado_drenagem_{hoje}{sufixo}.zip"
    zip_path = os.path.join(pasta_saida, zip_name)
    empacotar_zip(excels_gerados, fotos, zip_path)
    print(f"\n[ZIP]  {zip_path}")
    print(f"       Conteúdo: {len(excels_gerados)} Excel + {len(fotos)} foto(s)")
    print("\nProntos para upload no Kartado.")


if __name__ == "__main__":
    main()
