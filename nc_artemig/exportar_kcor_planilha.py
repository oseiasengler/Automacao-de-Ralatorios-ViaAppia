r"""
Gera «Exportar Kcor.xlsx» (lote 50 Artemig), alinhado às macros Nas01 (col. T/U) e Nas02 (V/W).

Inserções: uma linha por NC a partir da **linha 2** (aba Dados), como Nas01; insere linhas se o modelo tiver menos linhas que o lote.

T (Obs. Gestor): Trecho + Notificação + Nº Consol **numa única linha** (sem ``\\n`` — o Excel parte o texto ao editar se houver quebras).
U (Observações): valores na ordem (código · SH · indicador · patologia · …) **sem** rótulos «Notificação:» etc.; consol no fim; **uma linha**; limite ~32k (Excel).
G/H (KMi/KMf): se valor > 500, divide por 1000 (metros → km), como Nas01.
V (Diretório): barras só ``\\``, por segmento (letra ``C:\\`` + pastas); W (Arquivos): nomes sem ``\\``/``/``; PDF primeiro, depois nc (.jpg), «;» (Nas02).

Excel: T/U conteúdo **sem quebras** na célula (barra de fórmulas / edição num só bloco); V/W como antes. ``_escapar_inicio_formula_excel``; sem wrap. Altura não forçada; ``None`` se vazio. Desfaz mesclagens A–Y na linha de dados.
Valores que começam por = + - @ são gravados como texto (prefixo ') para a barra de fórmulas mostrar o conteúdo inteiro; colunas de texto usam formato @ e fonte Calibri preta (evita herdar cor branca do modelo).
"""
from __future__ import annotations

import io
import logging
import os
import re
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# Excel interpreta início com = + - @ como fórmula: barra de fórmulas truncada e parte do texto «invisível».
_EXCEL_INICIO_FORMULA = frozenset("=+-@")


def _strip_controles_invisiveis_excel(s: str, *, multiline: bool) -> str:
    """Remove C0/C1 invisíveis do PDF; preserva quebras só se multiline (col. T/U)."""
    if not s:
        return ""
    t = s.replace("\r\n", "\n").replace("\r", "\n")
    t = t.replace("\u2028", "\n").replace("\u2029", "\n")
    out: list[str] = []
    for ch in t:
        o = ord(ch)
        if multiline and ch == "\n":
            out.append(ch)
        elif ch == "\t":
            out.append(" ")
        elif o < 32:
            continue
        elif o == 0x7F:
            continue
        else:
            out.append(ch)
    u = "".join(out)
    if not multiline:
        return re.sub(r"\s+", " ", u).strip()
    u = re.sub(r"\n{4,}", "\n\n\n", u).strip()
    linhas = [re.sub(r"\s+", " ", ln).strip() for ln in u.split("\n")]
    return "\n".join(x for x in linhas if x)


def _compactar_quebras_multilinha_excel(s: str) -> str:
    """
    Normaliza ``\\r\\n`` / ``\\r`` → ``\\n`` e remove linhas em branco repetidas.
    Necessário porque ``\\r\\n\\r\\n`` não é captado por ``\\n{2,}`` e o Excel mostra
    linhas «fantasma» entre Trecho / Notificação / Nº Consol.
    """
    if not s:
        return ""
    t = str(s).replace("\r\n", "\n").replace("\r", "\n")
    t = re.sub(r"\n{2,}", "\n", t)
    return t.strip()


def _escapar_inicio_formula_excel(s: str) -> str:
    """Força texto literal se o primeiro carácter visível for = + - @ (regra do Excel)."""
    if not s:
        return ""
    stripped = s.lstrip(" \t")
    if not stripped:
        return s
    if stripped[0] not in _EXCEL_INICIO_FORMULA:
        return s
    i = len(s) - len(stripped)
    return s[:i] + "'" + stripped


def _valor_linha_unica_excel_final(s: str) -> str:
    t = _linha_unica_espacos(s)
    t = _strip_controles_invisiveis_excel(t, multiline=False)
    if not t:
        return ""
    return _escapar_inicio_formula_excel(t)


def _aplicar_bordas_linha_kcor(ws, row: int, col_fim: int = 25) -> None:
    from openpyxl.styles import Border, Side

    b = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    for col in range(1, col_fim + 1):
        ws.cell(row=row, column=col).border = b


def _desfazer_merge_colunas_linha_kcor(ws, row: int, col_ini: int, col_fim: int) -> None:
    """Remove mesclagens que cruzem a linha de dados (evita texto «partido» entre T/U/V)."""
    from openpyxl.utils.cell import range_boundaries

    to_unmerge = []
    for mc in list(ws.merged_cells.ranges):
        try:
            min_col, min_row, max_col, max_row = range_boundaries(str(mc))
        except Exception:
            continue
        if row >= min_row and row <= max_row and not (col_fim < min_col or col_ini > max_col):
            to_unmerge.append(mc)
    for mc in to_unmerge:
        try:
            ws.unmerge_cells(str(mc))
        except Exception:
            pass


def _copiar_estilo_linha_kcor(ws, row_origem: int, row_destino: int, col_fim: int = 25) -> None:
    for col in range(1, col_fim + 1):
        src = ws.cell(row=row_origem, column=col)
        dst = ws.cell(row=row_destino, column=col)
        if src.has_style:
            dst.font = src.font.copy()
            dst.border = src.border.copy()
            dst.fill = src.fill.copy()
            dst.number_format = src.number_format
            dst.alignment = src.alignment.copy()
    _aplicar_bordas_linha_kcor(ws, row_destino, col_fim)


def _linha_unica_espacos(s: str) -> str:
    """Uma linha: NBSP + ``\\s+`` → um espaço (`limpeza_profunda`)."""
    from nc_artemig.texto_pdf import limpeza_profunda

    return limpeza_profunda(s)


def _patologia_fonte_nc_kcor(nc: Any) -> str:
    """Patologia para Nas01/Kcor: campo PDF; se vazio, col. Tipo da planilha de análise; por último Grupo."""
    for attr in ("patologia_artemig", "tipo_atividade", "grupo_atividade"):
        v = (getattr(nc, attr, None) or "").strip()
        if v:
            return v
    return ""


_RE_ATIVIDADE_COMO_PAT_U = re.compile(
    r"(?i)e\s*/\s*ou|pista\s+de\s+rolamento|panelas?\s+na\s+pista|buracos?\s+e/ou",
)
_RE_PAT_U_INCOMPLETA_E = re.compile(r"(?i)(?:panelas?|buracos?|pavimento)\s+e\s*$")
_RE_MARCADOR_ESCAPE_EXCEL = re.compile(r"(?i)_x000[0-9a-f]{1}_")


def _patologia_fonte_observacao_col_u(nc: Any) -> str:
    """Col. U: patologia truncada no PDF — prolongar com tipo/grupo do mesmo registo ou com atividade típica."""
    p = (getattr(nc, "patologia_artemig", None) or "").strip()
    t = (getattr(nc, "tipo_atividade", None) or "").strip()
    g = (getattr(nc, "grupo_atividade", None) or "").strip()
    base = (_patologia_fonte_nc_kcor(nc) or "").strip()
    if p and _RE_PAT_U_INCOMPLETA_E.search(p):
        if t and p.lower() in t.lower() and len(t) >= len(p) + 3:
            base = t
        elif g and p.lower() in g.lower() and len(g) >= len(p) + 3:
            base = g
    atv = _linha_unica_espacos(getattr(nc, "atividade", None) or "").strip()
    if atv:
        try:
            from nc_artesp.modulos.analisar_pdf_nc import _texto_e_bloco_legenda_atividade_artemig

            if _texto_e_bloco_legenda_atividade_artemig(atv):
                atv = ""
        except Exception:
            pass
    cands = [base] if base else []
    if atv and _RE_ATIVIDADE_COMO_PAT_U.search(atv):
        cands.append(atv)
    if not cands:
        return ""
    return max(cands, key=len)


def _remover_marcadores_escape_excel(s: str) -> str:
    t = (s or "").strip()
    if not t:
        return ""
    return _RE_MARCADOR_ESCAPE_EXCEL.sub("", t).strip()


def _indicador_fonte_nc_kcor(nc: Any, patologia_bruta: str) -> str:
    """Indicador dedicado; Grupo só se distinto da patologia já resolvida (Grupo = indicador no Excel Artemig)."""
    i = (getattr(nc, "indicador_artemig", None) or "").strip()
    if i:
        return i
    g = (getattr(nc, "grupo_atividade", None) or "").strip()
    pb = (patologia_bruta or "").strip()
    if g and pb and g != pb:
        return g
    return ""


def _normalizar_texto_celula_kcor(s: str) -> str:
    """Multilinha (ex. observação PDF): NFKC + colapso por linha."""
    from nc_artemig.texto_pdf import colapsar_espacos_pdf

    return colapsar_espacos_pdf(s, multiline=True)


def _primeira_linha_dados_planilha_kcor(ws, c: dict[str, int]) -> int:
    """
    Linha da primeira NC (abaixo dos cabeçalhos). Modelos Nas01 variam: cabeçalhos na 1 ou 2.
    """
    c_a = c.get("NumItem", 1)
    c_b = c.get("Origem", 2)
    lim = min(ws.max_row + 1, 12)
    for r in range(1, lim):
        a = str(ws.cell(r, c_a).value or "").strip().lower()
        a_alnum = re.sub(r"[^a-z0-9]", "", a)
        b = str(ws.cell(r, c_b).value or "").strip().lower()
        if "numitem" in a_alnum or ("item" in a_alnum and "num" in a_alnum):
            return r + 1
        if b == "origem":
            return r + 1
    return 2


def _larguras_min_colunas_texto_kcor(ws, c: dict[str, int]) -> None:
    from openpyxl.utils import get_column_letter

    alvo = {
        c["Origem"]: 16.0,
        c["Obs_Gestor"]: 56.0,
        c["Observacoes"]: 62.0,
        c["Diretorio"]: 54.0,
        c["Arquivos"]: 58.0,
    }
    for col_idx, wmin in alvo.items():
        le = get_column_letter(col_idx)
        dim = ws.column_dimensions[le]
        cur = dim.width
        try:
            cval = float(cur) if cur is not None else 0.0
        except (TypeError, ValueError):
            cval = 0.0
        dim.width = max(wmin, cval)


def _resolver_caminho_modelo_kcor() -> Path:
    """Ficheiro em disco ou `ARTEMIG_MODELO_KCOR_KRIA` (env); senão default em config."""
    from nc_artemig.config import MODELO_KCOR_KRIA

    env = (os.environ.get("ARTEMIG_MODELO_KCOR_KRIA") or "").strip()
    return Path(env) if env else Path(MODELO_KCOR_KRIA)


def _workbook_modelo_kcor_minimo() -> Any:
    """Modelo mínimo quando o .xlsx da rede/repo não existe — evita ZIP sem Exportar Kcor."""
    import openpyxl
    from openpyxl.styles import Border, Font, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dados"
    thin = Side(style="thin", color="000000")
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = [
        "NumItem",
        "Origem",
        "Motivo",
        "Classificacao",
        "Tipo",
        "Rodovia",
        "KMi",
        "KMf",
        "Sentido",
        "Local",
        "Gestor",
        "Executores",
        "Data Solicitação",
        "Data de Suspensão",
        "Dt.Inicio Prog.",
        "Dt.Fim Prog.",
        "Dt.Inicio Exec.",
        "Dt.Fim Exec.",
        "Prazo (em Dias)",
        "Observação Gestor",
        "Observações",
        "Diretorio",
        "Arquivos (Separados por ;)",
        "Indicador",
        "Unidade",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True)
        c.border = bord
    for col in range(1, 26):
        ws.cell(2, col).border = bord
    return wb


_CLASS = "Eng. QID"


def _norm_macro_patologia(s: str) -> str:
    import unicodedata

    t = unicodedata.normalize("NFKC", (s or "").strip())
    return " ".join(t.lower().split())


def _patologia_para_kcor(
    pat: str, _indicador: str, _atividade: str, _nc: Any | None = None
) -> tuple[str, str]:
    """
    Col. E (Tipo Kcor) — conversão Nas01 / macro VBA, **só** sobre o texto de patologia.

    Vários ``If`` seguidos (sem ``ElseIf``): a última condição verdadeira prevalece.
    ``_indicador``, ``_atividade`` e ``_nc`` não entram no mapa (assinatura mantida para chamadores).
    Exceção: buracos/panelas diferencia emergencial de reparo técnico quando ``_nc.tipo_panela=True``.
    Se nada bater, ``""`` (célula vazia).
    """
    pl = _norm_macro_patologia(_linha_unica_espacos((pat or "").strip()))
    kcor: str | None = None

    def _set(v: str) -> None:
        nonlocal kcor
        kcor = v.rstrip() + (" " if v.endswith(" ") else "")

    if len(pl) >= 6 and pl[:6] == "trilha":
        _set("Afundamento nas trilhas de rodas")
    if len(pl) >= 9 and pl[:9] == "alambrado":
        _set("Alambrado")
    if pl == _norm_macro_patologia("Dispositivo de segurança (alambrado)"):
        _set("Alambrado")
    if pl.startswith("guarda corpo"):
        _set("Barreira rígida ")
    if pl == _norm_macro_patologia("Inexistência de elementos refletivos"):
        _set("Barreira rígida ")
    if len(pl) >= 7 and pl[:7] == "buracos":
        _eh_reparo_tec = (
            _nc is not None
            and getattr(_nc, "tipo_panela", False)
            and not getattr(_nc, "emergencial", True)
        )
        if _eh_reparo_tec:
            _set("Buracos e panelas - Reparo técnico")
        else:
            _set("Buracos e panelas - Emergencial ")
    if len(pl) >= 7 and pl[-7:] == "caiação":
        _set("Caiação")
    if len(pl) >= 5 and pl[:5] == "cerca":
        _set("Cerca")
    if len(pl) >= 6 and pl[:6] == "erosão":
        _set("Conservação de terraplenos e contenções")
    if len(pl) >= 7 and pl[:7] == "defensa":
        _set("Defensa metálica")
    if len(pl) >= 11 and pl[:11] == "deformações":
        _set("Deformação permanente ")
    if len(pl) >= 6 and pl[:6] == "degrau":
        if "acostamento" in pl:
            _set("Degraus em acostamentos no maximo 5cm")
        else:
            _set("Degraus")
    if len(pl) >= 20 and pl[:20] == "sinalização vertical":
        _set("Demais placas")
    if len(pl) >= 13 and pl[:13] == "demais placas":
        _set("Demais placas")
    if pl == _norm_macro_patologia("Vandalismo demais placas"):
        _set("Demais placas")
    if len(pl) >= 6 and pl[-6:] == "perigo":
        _set("Demais placas")
    if len(pl) >= 8 and pl[-8:] == "vertical":
        _set("Demais placas")
    if len(pl) >= 10 and pl[:10] == "iluminação":
        _set("Dispositivos de Iluminação")
    if len(pl) >= 20 and pl[:20] == "drenagem subterrânea":
        _set("Drenagem Subterrânea")
    if len(pl) >= 20 and pl[:20] == "drenagem superficial":
        _set("Drenagem Superficial")
    if len(pl) >= 6 and pl[:6] == "grelha":
        _set("Drenagem Superficial")
    if len(pl) >= 7 and pl[:7] == "entulho":
        _set("Entulho")
    if len(pl) >= 32 and pl[:32] == "vandalismo placas de advertência":
        _set("Placas - Regulam. / Advertência")
    if len(pl) >= 21 and pl[:21] == "placas de advertência":
        _set("Placas - Regulam. / Advertência")
    if len(pl) >= 10 and pl[-10:] == "horizontal":
        _set("Sinalização horizontal")
    if len(pl) >= 22 and pl[:22] == "sinalização horizontal":
        _set("Sinalização horizontal")
    if len(pl) >= 7 and pl[-7:] == "tachões":
        _set("Tachas e tachões")
    if len(pl) >= 9 and pl[:9] == "vegetação":
        _set("Vegetação")
    # ── Novas regras — parametrização ArteM IG confirmada ────────────────────
    if pl.startswith("afundamento"):
        _set("Afundamento nas trilhas de rodas")
    if pl.startswith("detrito"):
        _set("Entulho")
    if pl.startswith("inexistência de cerca"):
        _set("Cerca")
    if pl.startswith("inexistência de defensa"):
        _set("Defensa metálica")
    if pl.startswith("inexistência de ilumina"):
        _set("Dispositivos de Iluminação")
    if pl.startswith("inexistência de sinalização h"):
        _set("Sinalização horizontal")
    if pl.startswith("repintura"):
        _set("Sinalização horizontal")
    if pl.startswith("tacha"):
        _set("Tachas e tachões")
    if pl.startswith("talude"):
        _set("Conservação de terraplenos e contenções")
    if pl.startswith("altura"):
        _set("Vegetação")
    if pl.startswith("inexistência de marco"):
        _set("Demais placas")
    if pl.startswith("dispositivos aux"):
        _set("Demais placas")
    if pl.startswith("abrigo"):
        _set("Abrigo de passageiros")
    # ─────────────────────────────────────────────────────────────────────────

    if kcor is None:
        return "", _CLASS
    return kcor, _CLASS


def _parse_dt(s: str) -> datetime | None:
    t = (s or "").strip()
    if " " in t:
        t = t.split()[0].strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(t, fmt)
        except ValueError:
            continue
    return None


def _prazo_dias_efetivo(nc: Any) -> int:
    d = getattr(nc, "prazo_dias", None)
    if d is None:
        return 0
    try:
        n = int(d)
    except (TypeError, ValueError):
        return 0
    if n == 24 and getattr(nc, "emergencial", False):
        return 1
    return max(0, n)


def _codigo_fiscalizacao_arquivos(nc: Any) -> str:
    """Código da fiscalização (ex.: 202506784) — sem whitespace (PDF/NBSP)."""
    from nc_artemig.texto_pdf import identificador_pdf_sem_whitespace

    c = identificador_pdf_sem_whitespace(getattr(nc, "codigo", None) or "")
    if not c:
        return ""
    if re.fullmatch(r"\d{6,14}", c):
        return c
    m = re.search(r"\b(\d{8,10})\b", c)
    return identificador_pdf_sem_whitespace(m.group(1)) if m else c


def _rodovia_coluna_f(rod: str) -> str:
    r = re.sub(r"\s+", " ", (rod or "").strip().upper()).replace("-", " ")
    m = re.match(r"^(MG|BR)\s+(\d+)$", r)
    if m:
        pref, num = m.group(1), int(m.group(2))
        return f"{pref}-{num:03d}"
    if " " in r:
        return r.replace(" ", "-", 1)
    return (rod or "").strip()


def _local_coluna_j(nc: Any) -> str:
    blob = f"{nc.atividade or ''} {getattr(nc, 'tipo_atividade', None) or ''} {getattr(nc, 'grupo_atividade', None) or ''}".upper()
    if (
        ("DOM" in blob and "NIO" in blob)
        or "FAIXA DE DOM" in blob
        or "FAIXA DE DOMINIO" in blob
        or "FX." in blob
    ):
        return "Faixa de Domínio"
    return "Faixa de Rolamento"


def _data_kcor_so_data(nc: Any) -> tuple[str, datetime.date | None]:
    """Data Kcor como dd/mm/aaaa (hora fica na coluna Hora)."""
    dt = _parse_dt(nc.data_con or "")
    if not dt:
        return "", None
    d = dt.date()
    return d.strftime("%d/%m/%Y"), d


def _inferir_tag_subpasta_artemig_fallback(nc: Any) -> str:
    from nc_artemig.texto_pdf import limpeza_profunda

    blob = limpeza_profunda(
        " ".join(
            str(x)
            for x in (
                getattr(nc, "grupo_atividade", None) or "",
                getattr(nc, "tipo_atividade", None) or "",
                getattr(nc, "indicador_artemig", None) or "",
                getattr(nc, "patologia_artemig", None) or "",
                getattr(nc, "atividade", None) or "",
            )
            if x
        )
    ).upper()
    b = blob.replace("Â", "A").replace("Ã", "A")
    if "SINALIZ" in b:
        return "SINALIZACAO"
    if "DRENAGEM" in b:
        return "DRENAGEM"
    if "PARAMETROS" in b and "GERAIS" in b:
        return "PARAMETROS_GERAIS"
    if "DEFENS" in b:
        return "DEFENSA"
    if any(k in b for k in ("PAVIMENTO", "BURACO", "PANELA", "RECAP")):
        return "PAVIMENTO"
    return "OUTROS"


def _stem_subpasta_fotos(nc: Any) -> str:
    """Pasta col. V: stem do PDF (NOT-…_DRENAGEM_CE…); só sintetiza NOT-…_TAG_CE… se stem ausente."""
    from nc_artemig.texto_pdf import identificador_pdf_sem_whitespace, limpeza_profunda

    stem = limpeza_profunda((getattr(nc, "artemig_pdf_stem", None) or ""))
    if stem:
        return stem

    cod = _codigo_fiscalizacao_arquivos(nc)
    cons = identificador_pdf_sem_whitespace(getattr(nc, "num_consol", None) or "")
    if len(cod) >= 9 and cons.isdigit():
        yy = cod[2:4]
        seq = cod[4:9]
        tag = _inferir_tag_subpasta_artemig_fallback(nc)
        return f"NOT-{yy}-{seq}_{tag}_CE{cons}"
    return stem


def _parte_texto_caminho_v(p: str) -> str:
    """
    Caminho Windows: ``/`` → ``\\``; espaços parasitas por *segmento*; barras duplicadas
    via ``PureWindowsPath``. Letra de unidade: ``C:`` + ``pasta`` → ``C:\\pasta`` (não ``C:pasta``).
    UNC ``\\\\servidor\\...`` e ``\\\\?\\`` / ``\\\\.\\`` tratados pelo pathlib.
    """
    import re
    import unicodedata
    from pathlib import PureWindowsPath

    from nc_artemig.texto_pdf import colapsar_espacos_pdf

    if not p:
        return ""
    raw = unicodedata.normalize("NFKC", str(p)).strip().strip('"')
    raw = raw.replace("\r", "").replace("\n", " ").replace("/", "\\")
    # Caminhos longos / dispositivo — não partir em segmentos manuais
    if re.match(r"^(\\\\\?\\|\\\\\.\\)", raw):
        try:
            return str(PureWindowsPath(raw))
        except Exception:
            return raw.rstrip("\\").strip()
    unc = raw.startswith("\\\\")
    body = raw[2:] if unc else raw
    segs = [colapsar_espacos_pdf(x, multiline=False).strip() for x in body.split("\\")]
    segs = [x for x in segs if x]
    if not segs:
        return ""
    if unc:
        return str(PureWindowsPath("\\\\", *segs))
    first = segs[0]
    if len(first) == 2 and first[1] == ":" and first[0].isalpha():
        first = first + "\\"
    path = PureWindowsPath(first)
    for x in segs[1:]:
        path = path / x
    return str(path)


def _caminho_coluna_v_windows(base: str, stem: str) -> str:
    """Coluna V: só ``\\``, sem barras duplicadas nem espaços parasitas entre pastas."""
    from pathlib import PureWindowsPath

    b = _parte_texto_caminho_v(base)
    s = _parte_texto_caminho_v(stem)
    if not b:
        return ""
    if not s:
        return str(PureWindowsPath(b))
    return str(PureWindowsPath(b) / s)


def _lista_arquivos_coluna_w_sanear(s: str) -> str:
    """Col. W: ``;`` entre nomes; só ficheiros (sem ``\\`` ou ``/`` no nome); NBSP + espaços colapsados."""
    from nc_artemig.texto_pdf import limpeza_profunda

    partes: list[str] = []
    for seg in (s or "").split(";"):
        t = limpeza_profunda(seg)
        t = t.replace("\\", "").replace("/", "").replace("_", " ")
        t = re.sub(r" +", " ", t).strip()
        if t:
            partes.append(t)
    return ";".join(partes)


def _excel_valor_texto_ou_none(s: str | None) -> str | None:
    """Evita célula «vazia» com só espaços/NBSP: grava ``None`` (T/U/V/W)."""
    if s is None:
        return None
    t = str(s).strip()
    return None if not t else t


def _km_normalizado_nas01(val: float | None) -> float | None:
    """Nas01: só divide por 1000 quando o valor parece metros inteiros (ex.: 653400), nunca km decimal (653,4)."""
    if val is None:
        return None
    try:
        x = float(val)
    except (TypeError, ValueError):
        return None
    if abs(x - round(x)) > 1e-6:
        return x
    xi = int(round(x))
    if xi >= 10_000:
        return xi / 1000.0
    return x


def _linha_rotulo_identificador_kcor(rotulo_com_dois_pontos: str, valor_bruto: str) -> str | None:
    """Rótulo + valor com espaço após «:» (ex.: «Trecho Homogênio: SH05»); valor só IDs sem lixo PDF."""
    from nc_artemig.texto_pdf import identificador_pdf_sem_whitespace

    rot = (rotulo_com_dois_pontos or "").replace("\n", "").replace("\r", "").strip()
    v = identificador_pdf_sem_whitespace(_remover_marcadores_escape_excel(valor_bruto))
    if not v:
        return None
    linha = f"{rot} {v}"
    return re.sub(r"[\n\r\t]+", "", linha).strip()


def _bloco_obs_gestor_nas01_linhas_raw(nc: Any) -> str:
    """Trecho / Notificação / Nº Consol colados sem ``\\n`` (uma linha; evita «3 partes» ao selecionar no Excel)."""
    from nc_artemig.texto_pdf import limpeza_linha_excel_pdf

    linhas: list[str] = []
    for rotulo, attr in (
        ("Trecho Homogênio:", getattr(nc, "sh_artemig", None) or ""),
        ("Notificação:", nc.codigo or ""),
        ("Nº Consol:", getattr(nc, "num_consol", None) or ""),
    ):
        ln = _linha_rotulo_identificador_kcor(rotulo, attr)
        if ln:
            linhas.append(limpeza_linha_excel_pdf(ln))
    return "".join(linhas)


def _bloco_obs_gestor_nas01(nc: Any) -> str:
    """Nas01 col. T: uma linha + ``_escapar_inicio_formula_excel`` (como V/W)."""
    raw = _bloco_obs_gestor_nas01_linhas_raw(nc)
    if not raw:
        return ""
    return _escapar_inicio_formula_excel(raw)


def _observacao_para_col_u(nc: Any, desc: str) -> str:
    """Trecho extra do PDF: não repetir só código/notificação, Nº Consol nem texto já igual à atividade."""
    from nc_artemig.texto_pdf import identificador_pdf_sem_whitespace

    obs = _normalizar_texto_celula_kcor(getattr(nc, "observacao", None) or "")
    try:
        from nc_artesp.modulos.analisar_pdf_nc import (
            _limpar_legendas_campo_artemig,
            _strip_descricao_rotulos_repetidos_artemig,
            _texto_e_bloco_legenda_atividade_artemig,
        )

        obs = _limpar_legendas_campo_artemig(obs)
        obs = _strip_descricao_rotulos_repetidos_artemig(obs)
        if _texto_e_bloco_legenda_atividade_artemig(obs):
            return ""
    except Exception:
        pass
    if not obs:
        return ""
    if obs == desc or (desc and obs in desc):
        return ""
    cons = identificador_pdf_sem_whitespace(getattr(nc, "num_consol", None) or "")
    cod = identificador_pdf_sem_whitespace(nc.codigo or "")
    obs_nw = identificador_pdf_sem_whitespace(obs)
    if cons and obs_nw == cons:
        return ""
    if cod and obs_nw == cod:
        return ""
    if obs_nw.isdigit() and cons and obs_nw == cons:
        return ""
    # Multilinha: tirar linhas que são só o Nº CONSOL (já no rodapé da U).
    if cons and re.fullmatch(r"\d{6,10}", cons):
        linhas = [
            ln
            for ln in obs.split("\n")
            if identificador_pdf_sem_whitespace(ln) != cons
        ]
        obs = "\n".join(linhas).strip()
    if not obs:
        return ""
    return obs.strip()[:32000]


def _sanear_rotulo_pdf_col_u(s: str) -> str:
    """Remove «/» inicial típico de cópia de tabela; trim."""
    t = (s or "").strip()
    return re.sub(r"^[/\\|]+\s*", "", t).strip()


def _patologia_texto_completo_col_u(nc: Any, pat_san: str) -> str:
    tipo = _sanear_rotulo_pdf_col_u(_linha_unica_espacos((getattr(nc, "tipo_atividade", None) or "")))
    grp = _sanear_rotulo_pdf_col_u(_linha_unica_espacos((getattr(nc, "grupo_atividade", None) or "")))
    atv = _sanear_rotulo_pdf_col_u(_linha_unica_espacos((getattr(nc, "atividade", None) or "")))
    p, t, g, a = (pat_san or "").strip(), (tipo or "").strip(), (grp or "").strip(), (atv or "").strip()
    if not p:
        return t or g or a
    lowp = p.lower()
    for cand in (t, g):
        if cand and lowp in cand.lower() and len(cand) >= len(p) + 3:
            return cand
    atv_ok = a
    if atv_ok:
        try:
            from nc_artesp.modulos.analisar_pdf_nc import _texto_e_bloco_legenda_atividade_artemig

            if _texto_e_bloco_legenda_atividade_artemig(atv_ok):
                atv_ok = ""
        except Exception:
            pass
    if (
        atv_ok
        and _RE_ATIVIDADE_COMO_PAT_U.search(atv_ok)
        and len(atv_ok) > len(p)
        and _RE_PAT_U_INCOMPLETA_E.search(p)
    ):
        return atv_ok
    return p or t or g


def _indicador_texto_completo_col_u(nc: Any, ind_san: str) -> str:
    grp = _sanear_rotulo_pdf_col_u(_linha_unica_espacos((getattr(nc, "grupo_atividade", None) or "")))
    i, g = (ind_san or "").strip(), (grp or "").strip()
    if g and i and "/" not in i and "/" in g and (
        g.lower().startswith(i.lower() + " ")
        or g.lower().startswith(i.lower() + "/")
    ):
        return g
    return i if i else g


def _indicador_patologia_deduplicados_col_u(ind: str, pat: str) -> tuple[str, str]:
    """Indicador e patologia iguais (ex.: Excel/planilha repetem «Gerais (Parâmetros)») → só patologia na U."""
    i, p = (ind or "").strip(), (pat or "").strip()
    if i and p and _norm_macro_patologia(i) == _norm_macro_patologia(p):
        return "", p
    return i, p


def _indicador_prefixo_patologia_col_u(ind: str, pat: str) -> str:
    """«Buracos» só como token e patologia já traz «Buracos e/ou …» → não repetir na U."""
    i, p = (ind or "").strip(), (pat or "").strip()
    if not i or not p or " " in i or "/" in i:
        return i
    low = p.lower()
    if len(p) > len(i) + 12 and low.startswith(i.lower()) and (
        "e/ou" in low or "rolamento" in low
    ):
        return ""
    return i


def _bloco_observacao_col_u_valores(nc: Any, pat: str, ind: str) -> str:
    """Col. U: só conteúdos (sem títulos), ordem estável — código · SH · indicador · patologia."""
    from nc_artemig.texto_pdf import identificador_pdf_sem_whitespace

    sep = " · "
    vals: list[str] = []
    cod = identificador_pdf_sem_whitespace(nc.codigo or "")
    if cod:
        vals.append(cod)
    sh = identificador_pdf_sem_whitespace(getattr(nc, "sh_artemig", None) or "")
    if sh:
        vals.append(sh)
    if (ind or "").strip():
        vals.append((ind or "").strip())
    if (pat or "").strip():
        vals.append((pat or "").strip())
    return sep.join(vals).strip()


def _num_consol_valor_col_u(nc: Any) -> str:
    from nc_artemig.texto_pdf import identificador_pdf_sem_whitespace

    return identificador_pdf_sem_whitespace(getattr(nc, "num_consol", None) or "")


def _nc_deve_texto_padrao_vegetacao_parametros_gerais_col_u(nc: Any) -> bool:
    from nc_artemig.texto_pdf import limpeza_profunda

    blob = limpeza_profunda(
        " ".join(
            str(getattr(nc, a, None) or "")
            for a in (
                "patologia_artemig",
                "tipo_atividade",
                "grupo_atividade",
                "indicador_artemig",
                "atividade",
            )
        )
    ).lower()
    veg = "vegetação" in blob or "vegetacao" in blob
    par = "parâmetros gerais" in blob or "parametros gerais" in blob or "gerais (parâmetro" in blob
    return bool(veg and par)


def _texto_col_u_padrao_notificacao_vegetacao(nc: Any) -> str:
    from nc_artemig.texto_pdf import identificador_pdf_sem_whitespace

    sh = identificador_pdf_sem_whitespace(
        _remover_marcadores_escape_excel(getattr(nc, "sh_artemig", None) or "")
    )
    cod = identificador_pdf_sem_whitespace(
        _remover_marcadores_escape_excel(getattr(nc, "codigo", None) or "")
    )
    cons = identificador_pdf_sem_whitespace(
        _remover_marcadores_escape_excel(getattr(nc, "num_consol", None) or "")
    )
    descricao = (
        "Foi verificado, em rotina diária de monitoramento, vegetação fora do padrão permitido "
        "(Altura > 0,30 m), causando aspecto visual desagradável, facilitando a propagação de incêndios "
        "e impactando nas condições de segurança da rodovia."
    )
    partes = [
        "Vegetação fora padrão (Parâmetros Gerais)",
        "Prazo para Atendimento à Notificação:",
        "Em até 5 (cinco) dias, a partir da data do recebimento desta notificação.",
        "Reincidência No:",
        descricao,
    ]
    if sh:
        partes.append(sh)
    if cod:
        partes.append(f"Notificação: {cod}")
    if cons:
        partes.append(f"Nº Consol: {cons}")
    return " ".join(partes).strip()


def _limpar_texto_final_obs_u(nc: Any, texto: str) -> str:
    """Aspas de cópia PDF e linhas duplicadas do Nº CONSOL."""
    from nc_artemig.texto_pdf import identificador_pdf_sem_whitespace

    t = (texto or "").strip()
    if len(t) >= 2 and t[0] in "\"\u201c\u00ab" and t[-1] in "\"\u201d\u00bb":
        t = t[1:-1].strip()
    cons = identificador_pdf_sem_whitespace(getattr(nc, "num_consol", None) or "")
    if cons and re.fullmatch(r"\d{6,10}", cons):
        linhas = [ln for ln in t.split("\n") if identificador_pdf_sem_whitespace(ln) != cons]
        t = "\n".join(linhas).strip()
    t = t.replace("\r\n", "\n").replace("\r", "\n")
    t = re.sub(r"\n{2,}", "\n", t)
    return t.strip()


def _colapsar_linhas_duplas_obs_u(texto: str) -> str:
    """Uma quebra lógica por linha (inclui normalização CRLF)."""
    return _compactar_quebras_multilinha_excel(texto or "")


def _texto_observacoes_nas01(nc: Any) -> str:
    """Col. U — Nas01: valores (código, SH, indicador, patologia, …) sem rótulos; consol só o número no fim."""
    if _nc_deve_texto_padrao_vegetacao_parametros_gerais_col_u(nc):
        tp = _texto_col_u_padrao_notificacao_vegetacao(nc)
        if tp:
            t0 = re.sub(r"[\r\n]+", " ", tp).strip()
            t0 = re.sub(r" +", " ", t0).strip()
            return _escapar_inicio_formula_excel(t0[:32700])
    pat_bruto_kcor = _patologia_fonte_nc_kcor(nc)
    pat_bruto = _patologia_fonte_observacao_col_u(nc)
    pat = _sanear_rotulo_pdf_col_u(_linha_unica_espacos(pat_bruto))
    ind = _sanear_rotulo_pdf_col_u(
        _linha_unica_espacos(_indicador_fonte_nc_kcor(nc, pat_bruto_kcor))
    )
    pat = _patologia_texto_completo_col_u(nc, pat)
    ind = _indicador_texto_completo_col_u(nc, ind)
    ind, pat = _indicador_patologia_deduplicados_col_u(ind, pat)
    ind = _indicador_prefixo_patologia_col_u(ind, pat)

    cabeca = _bloco_observacao_col_u_valores(nc, pat, ind)

    desc_raw = _sanear_rotulo_pdf_col_u(_linha_unica_espacos(nc.atividade or ""))
    try:
        from nc_artesp.modulos.analisar_pdf_nc import _strip_descricao_rotulos_repetidos_artemig

        desc_raw = _strip_descricao_rotulos_repetidos_artemig(desc_raw)
    except Exception:
        pass
    desc = ""
    if desc_raw:
        if _norm_macro_patologia(desc_raw) == _norm_macro_patologia(pat):
            desc = ""
        else:
            try:
                from nc_artesp.modulos.analisar_pdf_nc import _texto_e_bloco_legenda_atividade_artemig

                if not _texto_e_bloco_legenda_atividade_artemig(desc_raw):
                    desc = desc_raw
            except Exception:
                desc = desc_raw
    obs_extra = _observacao_para_col_u(nc, desc)
    cons_val = _num_consol_valor_col_u(nc)

    partes: list[str] = []
    if cabeca:
        partes.append(cabeca)
    if desc:
        partes.append(desc)
    if obs_extra:
        partes.append(obs_extra)

    sep = " · "
    corpo = sep.join(p for p in partes if p)
    if cons_val:
        texto = f"{corpo}{sep}{cons_val}".strip() if corpo else cons_val
    else:
        texto = corpo
    texto = _limpar_texto_final_obs_u(nc, texto)
    texto = _colapsar_linhas_duplas_obs_u(texto)
    from nc_artemig.texto_pdf import limpeza_multilinha_excel_pdf

    texto = limpeza_multilinha_excel_pdf(texto)
    t = texto.strip()[:32700]
    if not t:
        return ""
    # U numa linha: quebras do PDF não podem ficar na célula (Excel segmenta ao editar).
    t = re.sub(r"[\r\n]+", " ", t)
    t = re.sub(r" +", " ", t).strip()
    return _escapar_inicio_formula_excel(t)


def _montar_v_w_kcor(nc: Any) -> tuple[str, str]:
    """V: pasta com \\. W: ordem Nas02 — .pdf primeiro, depois nc (...).jpg (lista real da extração quando existir)."""
    from nc_artemig.config import DIR_BASE_FOTOS_KCOR
    from nc_artesp.pdf_extractor import nome_pdf_original_seguro_zip

    base = re.sub(
        r"\s+",
        " ",
        (DIR_BASE_FOTOS_KCOR or os.environ.get("ARTEMIG_KCOR_DIR_FOTOS") or "").strip(),
    ).strip()
    stem = _stem_subpasta_fotos(nc)
    cod = _codigo_fiscalizacao_arquivos(nc)
    pags = list(getattr(nc, "artemig_kcor_paginas_jpg", None) or [])
    nomes_extraidos = [
        str(x).strip()
        for x in (getattr(nc, "artemig_kcor_nomes_arquivos", None) or [])
        if str(x).strip()
    ]

    v = _caminho_coluna_v_windows(base, stem) if base else ""

    if not cod:
        return v, ""

    stem_pdf = (getattr(nc, "artemig_pdf_stem", None) or "").strip()
    pdf_nome = nome_pdf_original_seguro_zip(f"{stem_pdf}.pdf" if stem_pdf else None)
    stem_arquivos = Path(pdf_nome).stem if pdf_nome.lower().endswith(".pdf") else stem_pdf
    if nomes_extraidos:
        fotos = nomes_extraidos
    else:
        n_nc = max(1, len(pags)) if pags else 1
        fotos = [f"nc ({cod}).jpg"]
        for i in range(1, n_nc):
            fotos.append(f"nc ({cod}) ({i}).jpg")
    if stem_arquivos:
        prefixo = f"{stem_arquivos} "
        fotos = [f if f.startswith(prefixo) else f"{prefixo}{f}" for f in fotos]
    w_s = _lista_arquivos_coluna_w_sanear(";".join([pdf_nome] + fotos))
    return v, w_s


def _ordenar_ncs_por_codigo_kcor(ncs: list[Any]) -> list[Any]:
    """Uma linha por fiscalização; ordem estável por número do código (ligação Excel ↔ ficheiros)."""

    def chave(nc: Any) -> tuple:
        c = _codigo_fiscalizacao_arquivos(nc) or ""
        try:
            n = int(c) if c.isdigit() else 0
        except ValueError:
            n = 0
        return (n, nc.km_ini or 0.0, (nc.rodovia or ""), c)

    return sorted(ncs, key=chave)


def gerar_exportar_kcor_xlsx_bytes(
    ncs: list[Any],
) -> tuple[bytes | None, dict[str, Any]]:
    """Gera o XLSX; devolve também metadados (modelo em disco vs. mínimo gerado)."""
    from nc_artemig.config import COL_KCOR_KRIA

    meta: dict[str, Any] = {
        "ok": False,
        "motivo": "",
        "modelo_caminho": "",
        "modelo_arquivo_existe": False,
        "modelo_minimo_gerado": False,
    }
    try:
        from nc_artemig.sanear_pipeline import norm_lote_numero, sanear_ncs_lote50_consol
    except ImportError as ex:
        logger.error("exportar_kcor: sanear_pipeline: %s", ex)
        meta["motivo"] = "sanear_pipeline_ausente"
        return None, meta
    try:
        sanear_ncs_lote50_consol(ncs, forcar_todas=True)
    except Exception as ex:
        logger.warning("exportar_kcor: saneamento lote 50: %s", ex)

    ncs50 = [n for n in ncs if norm_lote_numero(getattr(n, "lote", None) or "") == "50"]
    if not ncs50:
        meta["motivo"] = "sem_ncs_lote_50"
        return None, meta
    ncs50 = _ordenar_ncs_por_codigo_kcor(ncs50)
    modelo = _resolver_caminho_modelo_kcor()
    meta["modelo_caminho"] = str(modelo)
    meta["modelo_arquivo_existe"] = modelo.is_file()
    try:
        import openpyxl
    except ImportError:
        logger.error("exportar_kcor: openpyxl necessário")
        meta["motivo"] = "openpyxl_ausente"
        return None, meta

    c = COL_KCOR_KRIA
    try:
        if modelo.is_file():
            wb = openpyxl.load_workbook(modelo)
        else:
            meta["modelo_minimo_gerado"] = True
            logger.warning(
                "exportar_kcor: modelo XLSX não encontrado em %s — a gerar modelo mínimo (coloque o ficheiro real ou defina ARTEMIG_MODELO_KCOR_KRIA).",
                modelo,
            )
            wb = _workbook_modelo_kcor_minimo()
        ws = wb["Dados"] if "Dados" in wb.sheetnames else wb.active

        from openpyxl.styles import Alignment, Border, Font, Side

        _side_k = Side(style="thin", color="000000")
        _border_linha_k = Border(
            left=_side_k, right=_side_k, top=_side_k, bottom=_side_k
        )
        _alinh_texto = Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=False,
            shrink_to_fit=False,
        )
        _alinh_texto_wrap = Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=True,
            shrink_to_fit=False,
        )
        _alinh_centro = Alignment(
            horizontal="center",
            vertical="top",
            wrap_text=False,
            shrink_to_fit=False,
        )
        _alinh_s_prazo = Alignment(
            horizontal="center",
            vertical="top",
            wrap_text=False,
            shrink_to_fit=False,
        )
        _font_txt_visivel = Font(name="Calibri", size=11, bold=False, color="FF000000")
        _cols_kcor_texto_formato_arroba = (
            c["Origem"],
            c["Motivo"],
            c["Classificacao"],
            c["Tipo"],
            c["Rodovia"],
            c["Sentido"],
            c["Local"],
            c["Obs_Gestor"],
            c["Observacoes"],
            c["Diretorio"],
            c["Arquivos"],
            c["Indicador"],
        )

        # Nas01: primeira linha de dados = após cabeçalhos; limpar até ao fim da folha para não deixar «Orgão Fiscalizador» do modelo.
        lin_mod = _primeira_linha_dados_planilha_kcor(ws, c)
        n_lin = len(ncs50)
        for r in range(lin_mod + 1, lin_mod + n_lin):
            if r > ws.max_row:
                ws.insert_rows(r, 1)
                _copiar_estilo_linha_kcor(ws, lin_mod, r, 25)
        r_limpar_fim = max(ws.max_row, lin_mod + max(n_lin, 1) - 1)
        for r in range(lin_mod, r_limpar_fim + 1):
            _desfazer_merge_colunas_linha_kcor(ws, r, 1, 25)
            for col in range(1, 26):
                ws.cell(r, col).value = None

        for idx, nc in enumerate(ncs50, start=1):
            r = lin_mod + idx - 1
            cod_linha = _codigo_fiscalizacao_arquivos(nc)
            if not cod_linha:
                logger.warning(
                    "Exportar Kcor linha %s: sem código fiscalização; col. W vazia para esta NC",
                    idx,
                )
            pat_bruto = _patologia_fonte_nc_kcor(nc)
            pat = _linha_unica_espacos(pat_bruto)
            ind = _linha_unica_espacos(_indicador_fonte_nc_kcor(nc, pat_bruto))
            kcor, classe = _patologia_para_kcor(
                pat, ind, _linha_unica_espacos(nc.atividade or ""), nc
            )

            ws.cell(r, c["NumItem"], idx)
            # Export Kcor só existe para CONSOL lote 50 Artemig: Origem Nas01 é sempre 0-QID (não depender de nc.lote por PDF/Excel).
            ws.cell(r, c["Origem"], "0-QID")
            ws.cell(r, c["Motivo"], "Conservação de Rotina")
            ws.cell(
                r,
                c["Classificacao"],
                _excel_valor_texto_ou_none(_valor_linha_unica_excel_final(classe)),
            )
            ws.cell(
                r,
                c["Tipo"],
                _excel_valor_texto_ou_none(_valor_linha_unica_excel_final(kcor)),
            )
            ws.cell(
                r,
                c["Rodovia"],
                _excel_valor_texto_ou_none(
                    _valor_linha_unica_excel_final(_rodovia_coluna_f(nc.rodovia or ""))
                ),
            )
            g = nc.km_ini if nc.km_ini is not None else _km_f(nc.km_ini_str)
            h = nc.km_fim if nc.km_fim is not None else g
            g = _km_normalizado_nas01(g)
            h = _km_normalizado_nas01(h)
            ws.cell(r, c["KMi"], g if g is not None else None)
            ws.cell(r, c["KMf"], h if h is not None else None)
            ws.cell(
                r,
                c["Sentido"],
                _excel_valor_texto_ou_none(_valor_linha_unica_excel_final(nc.sentido or "")),
            )
            ws.cell(
                r,
                c["Local"],
                _excel_valor_texto_ou_none(
                    _valor_linha_unica_excel_final(_local_coluna_j(nc))
                ),
            )
            ws.cell(r, c["Gestor"], None)
            ws.cell(r, c["Executores"], None)

            ds, d0 = _data_kcor_so_data(nc)
            pd = _prazo_dias_efetivo(nc)
            if ds and d0:
                ws.cell(r, c["Data_Solicitacao"], ds)
                ws.cell(r, c["Dt_Inicio_Prog"], ds)
                ws.cell(r, c["Dt_Inicio_Exec"], ds)
                fim = d0 + timedelta(days=pd) if pd else d0
                ws.cell(r, c["Dt_Fim_Prog"], fim.strftime("%d/%m/%Y"))
            if pd:
                ws.cell(r, c["Prazo"], pd)
            elif getattr(nc, "prazo_dias", None) is not None:
                try:
                    ws.cell(r, c["Prazo"], int(nc.prazo_dias))
                except (TypeError, ValueError):
                    pass

            t_obs = _bloco_obs_gestor_nas01(nc)
            u_obs = _texto_observacoes_nas01(nc)
            v_dir, w_arq = _montar_v_w_kcor(nc)
            # T, U, V, W: mesma regra — ``_escapar_inicio_formula_excel`` já está em T/U nos geradores; aqui só ``None`` se vazio.
            ws.cell(
                r,
                c["Obs_Gestor"],
                _excel_valor_texto_ou_none(t_obs),
            )
            ws.cell(
                r,
                c["Observacoes"],
                _excel_valor_texto_ou_none(u_obs),
            )
            v_xlsx = _excel_valor_texto_ou_none(
                _escapar_inicio_formula_excel(v_dir) if v_dir else ""
            )
            w_xlsx = _excel_valor_texto_ou_none(
                _escapar_inicio_formula_excel(w_arq) if w_arq else ""
            )
            ws.cell(r, c["Diretorio"], v_xlsx)
            ws.cell(r, c["Arquivos"], w_xlsx)
            ws.cell(r, c["Indicador"], None)
            ws.cell(r, c["Unidade"], None)

            for col_k in (
                c["Data_Solicitacao"],
                c["Data_Suspensao"],
                c["Dt_Inicio_Prog"],
                c["Dt_Fim_Prog"],
                c["Dt_Inicio_Exec"],
                c["Dt_Fim_Exec"],
            ):
                cl = ws.cell(r, col_k)
                if cl.value is not None and str(cl.value).strip():
                    sv = str(cl.value).strip()
                    if " " in sv and re.match(r"^\d{1,2}/\d{1,2}/\d{4}", sv):
                        cl.value = sv.split()[0][:10]
                cl.number_format = "@"
            for col_k in range(1, 26):
                ws.cell(r, col_k).border = _border_linha_k
            for col_tf in _cols_kcor_texto_formato_arroba:
                clf = ws.cell(r, col_tf)
                clf.font = _font_txt_visivel
                v = clf.value
                if v is not None and str(v).strip() != "":
                    clf.number_format = "@"
            for _ca in (
                c["Origem"],
                c["Motivo"],
                c["Classificacao"],
                c["Tipo"],
                c["Rodovia"],
                c["Sentido"],
                c["Local"],
                c["Indicador"],
            ):
                ws.cell(r, _ca).alignment = _alinh_texto
            for _ca in (
                c["Obs_Gestor"],
                c["Observacoes"],
                c["Diretorio"],
                c["Arquivos"],
            ):
                ws.cell(r, _ca).alignment = _alinh_texto_wrap
            ws.cell(r, c["NumItem"]).alignment = _alinh_centro
            for _cn in (c["KMi"], c["KMf"]):
                cln = ws.cell(r, _cn)
                if cln.value is not None:
                    cln.alignment = _alinh_centro
            for col_k in (
                c["Data_Solicitacao"],
                c["Data_Suspensao"],
                c["Dt_Inicio_Prog"],
                c["Dt_Fim_Prog"],
                c["Dt_Inicio_Exec"],
                c["Dt_Fim_Exec"],
            ):
                cl = ws.cell(r, col_k)
                if cl.value is not None and str(cl.value).strip():
                    cl.alignment = _alinh_centro
            ws.cell(r, c["Prazo"]).alignment = _alinh_s_prazo
            for _ce in (c["Gestor"], c["Executores"], c["Unidade"]):
                ws.cell(r, _ce).alignment = _alinh_texto
            # Altura padrão (não forçar pt por nº de quebras): com wrap_text=False, uma altura
            # calculada por \n deixava a linha «alta» como várias linhas mescladas e o modo
            # edição empilhava o texto. Limpar altura herdada do modelo.
            ws.row_dimensions[r].height = None

        _larguras_min_colunas_texto_kcor(ws, c)

        buf = io.BytesIO()
        wb.save(buf)
        meta["ok"] = True
        meta["motivo"] = ""
        return buf.getvalue(), meta
    except Exception as e:
        logger.exception("exportar_kcor: %s", e)
        meta["motivo"] = f"erro: {e!s}"
        return None, meta


def _km_f(s: str) -> float | None:
    if not s:
        return None
    m = re.match(r"(\d+)\s*\+\s*(\d+)", str(s).strip())
    if m:
        return int(m.group(1)) + int(m.group(2)) / 1000.0
    try:
        return float(str(s).replace(",", "."))
    except ValueError:
        return None
