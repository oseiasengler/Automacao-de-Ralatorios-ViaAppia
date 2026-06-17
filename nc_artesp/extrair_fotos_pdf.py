"""
Extrai imagens dos PDFs das NCs, redimensiona (nc 800×500 px, 222×319 DPI; PDF 960×401) e salva
nc (N).jpg / PDF (N).jpg. Entrada: PDF pareado com Excel ou pasta de PDFs. Execução: antes do M02.
"""

import logging
import unicodedata
from pathlib import Path

from config import (
    M01_LINHA_INICIO,
    M02_EXTRACAO_RENDER_DPI,
    M02_FOTOS_NC,
    M02_FOTOS_PDF,
    M02_FOTO_W,
    M02_FOTO_H,
    M02_FOTO_DPI_X,
    M02_FOTO_DPI_Y,
    M02_FOTO_PDF_W,
    M02_FOTO_PDF_H,
    SERVICO_ABREV,
)
from utils.helpers import garantir_pasta, str_caminho_io_windows

logger = logging.getLogger(__name__)

_extrair_imagens_pdf_cached = None


def _extrair_imagens_pdf_fn():
    global _extrair_imagens_pdf_cached
    if _extrair_imagens_pdf_cached is not None:
        return _extrair_imagens_pdf_cached
    import sys

    nc_root = Path(__file__).resolve().parent.parent
    repo = nc_root.parent
    for p in (str(repo), str(nc_root)):
        if p not in sys.path:
            sys.path.insert(0, p)
    from nc_artesp.pdf_extractor import extrair_imagens_pdf as fn

    _extrair_imagens_pdf_cached = fn
    return fn

def _log_draft_ram(ident: str, size_before: tuple, size_after: tuple, channels: int = 3) -> None:
    """Log em DEBUG: RAM economizada por draft()."""
    if not logger.isEnabledFor(logging.DEBUG):
        return
    w0, h0 = size_before
    w1, h1 = size_after
    full_mb = (w0 * h0 * channels) / (1024 * 1024)
    after_mb = (w1 * h1 * channels) / (1024 * 1024)
    saved = max(0.0, full_mb - after_mb)
    logger.debug("[draft] %s: %dx%d → %dx%d | ~%.2f MB RAM economizados", ident, w0, h0, w1, h1, saved)

COL_CODIGO_NC = 3
COL_TIPO_NC = 17  # tipo/serviço NC (EAF)

try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    openpyxl = None
    load_workbook = None

try:
    import xlrd
except ImportError:
    xlrd = None


def _normalizar_texto(s: str) -> str:
    """Remove acentos e deixa maiúsculo para comparação."""
    if not s:
        return ""
    nfd = unicodedata.normalize("NFD", s)
    sem_acento = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
    return sem_acento.upper().strip()


def _redimensionar_e_salvar(
    img_bytes: bytes, dest: Path, largura: int, altura: int, dpi: tuple[int, int] | None = None
) -> bool:
    """Redimensiona para (largura, altura), salva JPG. Opcional dpi=(x,y) no JPEG."""
    try:
        from PIL import Image
        import io
        img = Image.open(io.BytesIO(img_bytes))
        if getattr(img, "format", None) == "JPEG" and (img.width > largura or img.height > altura):
            before = (img.width, img.height)
            try:
                img.draft("RGB", (largura, altura))
                _log_draft_ram(dest.name, before, (img.width, img.height))
            except (AttributeError, TypeError, ValueError):
                pass
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")
        img_resized = img.resize((largura, altura), Image.LANCZOS)
        garantir_pasta(dest.parent)
        save_kw = {"quality": 90}
        if dpi:
            save_kw["dpi"] = dpi
        img_resized.save(str_caminho_io_windows(dest), "JPEG", **save_kw)
        return True
    except Exception as e:
        logger.warning(f"  Erro ao processar imagem: {e}")
        return False


def _extrair_primeira_imagem_pdf(pdf_path: Path) -> bytes | None:
    """Extrai a primeira imagem do PDF. Retorna bytes ou None."""
    imagens = _extrair_todas_imagens_pdf(pdf_path)
    return imagens[0] if imagens else None


def _normalizar_codigo_nc(valor: object) -> tuple[str, str]:
    """
    Retorna (codigo_completo, codigo_base) para matching.
    Exemplos:
      - "31-6033" -> ("31-6033", "6033")
      - "6033" -> ("6033", "6033")
    """
    s = str(valor or "").strip()
    if not s:
        return "", ""
    m = __import__("re").fullmatch(r"\s*(\d{1,4})\s*-\s*(\d{3,})\s*", s)
    if m:
        return f"{int(m.group(1))}-{int(m.group(2))}", str(int(m.group(2)))
    try:
        n = int(float(s.replace(",", ".")))
        if n > 0:
            ns = str(n)
            return ns, ns
    except Exception:
        pass
    return s, s


def _tokens_codigo_no_texto(texto: str) -> set[str]:
    """Extrai candidatos de código de NC presentes no texto da página."""
    import re

    out: set[str] = set()
    t = str(texto or "")
    for m in re.finditer(r"\b(\d{1,4})\s*-\s*(\d{3,})\b", t):
        out.add(f"{int(m.group(1))}-{int(m.group(2))}")
        out.add(str(int(m.group(2))))
    for m in re.finditer(r"\b\d{3,}\b", t):
        try:
            out.add(str(int(m.group(0))))
        except Exception:
            continue
    return out


def _ler_ncs_do_excel(arquivo_excel: Path) -> list[tuple[int, str, str, str, str]]:
    """
    Lê a planilha EAF e retorna a lista de NCs em ordem:
    (índice_1based, tipo_nc, abrev, codigo_completo, codigo_base).
    """
    path = arquivo_excel
    if not path.is_file():
        return []
    suff = path.suffix.lower()
    ncs: list[tuple[int, str, str, str, str]] = []

    if suff == ".xls" and xlrd:
        try:
            book = xlrd.open_workbook(str(path))
            sheet = book.sheet_by_index(0)
            for r in range(M01_LINHA_INICIO - 1, sheet.nrows):
                tipo = (sheet.cell_value(r, COL_TIPO_NC - 1) or "").strip()
                if not tipo:
                    continue
                codigo_val = sheet.cell_value(r, COL_CODIGO_NC - 1)
                cod_full, cod_base = _normalizar_codigo_nc(codigo_val)
                abrev = SERVICO_ABREV.get(tipo, tipo[:30].replace("/", " ").replace("  ", " "))
                ncs.append((len(ncs) + 1, tipo, abrev, cod_full, cod_base))
        except Exception as e:
            logger.warning(f"Erro ao ler .xls para NCs: {e}")
            return []
    elif suff in (".xlsx", ".xlsm") and load_workbook:
        try:
            wb = load_workbook(str(path), data_only=True)
            ws = wb.active
            for r in range(M01_LINHA_INICIO, ws.max_row + 1):
                val = ws.cell(row=r, column=COL_TIPO_NC).value
                tipo = (val or "").strip() if val is not None else ""
                if not tipo:
                    continue
                cod_full, cod_base = _normalizar_codigo_nc(ws.cell(row=r, column=COL_CODIGO_NC).value)
                abrev = SERVICO_ABREV.get(tipo, tipo[:30].replace("/", " ").replace("  ", " "))
                ncs.append((len(ncs) + 1, tipo, abrev, cod_full, cod_base))
        except Exception as e:
            logger.warning(f"Erro ao ler .xlsx para NCs: {e}")
            return []
    return ncs


def _extrair_um_pdf_via_pdf_extractor(
    arquivo_pdf: Path,
    pasta_saida_pdf: Path,
    pasta_saida_nc: Path,
) -> int:
    """
    Mesma extração que /nc/extrair-pdf (recorte por NC). O legado deste módulo usava só a 1.ª
    imagem embutida ou a página inteira — ignorava ``pdf_extractor`` por completo.
    Pastas: neste módulo ``nc (N).jpg`` vai para ``pasta_saida_pdf`` e ``PDF (N).jpg`` para
    ``pasta_saida_nc`` (nomes herdados); no ``pdf_extractor`` é o contrário, por isso o swap.
    """
    extrair_imagens_pdf = _extrair_imagens_pdf_fn()
    garantir_pasta(pasta_saida_pdf)
    garantir_pasta(pasta_saida_nc)
    salvos = extrair_imagens_pdf(
        str(arquivo_pdf.resolve()),
        pasta_saida_nc=str(pasta_saida_pdf.resolve()),
        pasta_saida_pdf=str(pasta_saida_nc.resolve()),
        dpi=int(M02_EXTRACAO_RENDER_DPI) if M02_EXTRACAO_RENDER_DPI else None,
    )
    return len([s for s in (salvos or []) if s and str(s).lower().endswith(".jpg")])


def _extrair_paginas_com_imagem_e_texto(pdf_path: Path) -> list[tuple[str, bytes]]:
    """Para cada página com imagem, retorna (texto_da_página, bytes_imagem)."""
    if not fitz:
        return []
    doc = fitz.open(str(pdf_path))
    resultado: list[tuple[str, bytes]] = []
    try:
        for page_num in range(len(doc)):
            page = doc[page_num]
            text = (page.get_text() or "").strip()
            img_bytes = None

            # Método 1: extrair imagens embutidas
            for img in doc.get_page_images(page_num):
                xref = img[0]
                try:
                    base_img = doc.extract_image(xref)
                    b = base_img["image"]
                    if len(b) > 1000:
                        img_bytes = b
                        break
                except Exception:
                    continue

            # Método 2 (fallback): renderizar página como imagem (PDFs digitalizados)
            if img_bytes is None:
                try:
                    pix = page.get_pixmap(dpi=M02_EXTRACAO_RENDER_DPI)
                    img_bytes = pix.tobytes("png")
                    if img_bytes and len(img_bytes) > 1000:
                        logger.debug(f"  Página {page_num + 1}: renderizada (sem imagens embutidas)")
                except Exception as e:
                    logger.debug(f"  Página {page_num + 1}: falha ao renderizar: {e}")

            if img_bytes:
                resultado.append((text, img_bytes))
    finally:
        doc.close()
    return resultado


def _associar_imagens_a_ncs(
    paginas: list[tuple[str, bytes]],
    ncs: list[tuple[int, str, str, str, str]],
) -> list[tuple[int, bytes]]:
    """
    Associa cada (texto_página, imagem) ao índice da NC no Excel que melhor combina
    com o texto (por abrev ou palavras do tipo). Retorna [(índice_nc_1based, imagem), ...].
    Se várias NCs têm o mesmo tipo, atribui na ordem do Excel (primeira página que der match
    vai para a primeira NC daquele tipo, etc.).
    """
    if not ncs:
        return [(i + 1, img) for i, (_, img) in enumerate(paginas)]
    ncs_norm = [
        (
            idx,
            tipo,
            _normalizar_texto(abrev),
            _normalizar_texto(tipo),
            str(cod_full or "").strip(),
            str(cod_base or "").strip(),
        )
        for idx, tipo, abrev, cod_full, cod_base in ncs
    ]
    usados: set[int] = set()
    resultado: list[tuple[int, bytes]] = []

    for page_text, img_bytes in paginas:
        page_norm = _normalizar_texto(page_text)
        page_codes = _tokens_codigo_no_texto(page_text)
        candidatos: list[tuple[int, int]] = []  # (idx, pontos)
        for idx, tipo, abrev_norm, tipo_norm, cod_full, cod_base in ncs_norm:
            pontos = 0
            if cod_full and cod_full in page_codes:
                pontos += 10000
            if cod_base and cod_base in page_codes:
                pontos += 5000
            if abrev_norm and abrev_norm in page_norm:
                pontos += 100 + len(abrev_norm)
            palavras = [p for p in tipo_norm.split() if len(p) > 3]
            for p in palavras:
                if p in page_norm:
                    pontos += 10
            if pontos > 0:
                candidatos.append((idx, pontos))
        # Escolher o melhor candidato que ainda não foi usado (mesmo tipo: pegar o de menor índice)
        candidatos.sort(key=lambda x: (-x[1], x[0]))
        escolhido = None
        for idx, _ in candidatos:
            if idx not in usados:
                escolhido = idx
                break
        if escolhido is not None:
            usados.add(escolhido)
            resultado.append((escolhido, img_bytes))
            logger.debug(f"  Página → NC {escolhido}")
        else:
            # Nenhum match ou todos já usados: fallback = próxima NC livre por ordem
            for idx, _, _, _, _, _ in ncs_norm:
                if idx not in usados:
                    usados.add(idx)
                    resultado.append((idx, img_bytes))
                    logger.info(f"  Página sem match → NC {idx} (fallback)")
                    break

    return sorted(resultado, key=lambda x: x[0])


def _extrair_todas_imagens_pdf(pdf_path: Path) -> list[bytes]:
    """Extrai todas as imagens do PDF em ordem (uma por página, ignorando ícones)."""
    if not fitz:
        logger.error("PyMuPDF não instalado. Execute: pip install pymupdf")
        return []
    doc = fitz.open(str(pdf_path))
    imagens: list[bytes] = []
    try:
        for page_num in range(len(doc)):
            page = doc[page_num]
            img_bytes = None
            for img in doc.get_page_images(page_num):
                xref = img[0]
                try:
                    base_img = doc.extract_image(xref)
                    b = base_img["image"]
                    if len(b) > 1000:
                        img_bytes = b
                        break
                except Exception:
                    continue
            if img_bytes is None:
                try:
                    pix = page.get_pixmap(dpi=M02_EXTRACAO_RENDER_DPI)
                    img_bytes = pix.tobytes("png")
                except Exception:
                    pass
            if img_bytes and len(img_bytes) > 1000:
                imagens.append(img_bytes)
    finally:
        doc.close()
    return imagens


def executar(
    pasta_pdfs: Path | None = None,
    arquivo_pdf: Path | None = None,
    arquivo_excel: Path | None = None,
    pasta_saida_pdf: Path | None = None,
    pasta_saida_nc: Path | None = None,
    callback_progresso=None,
) -> int:
    """
    Extrai imagens e salva como pdf (N).jpg e nc (N).jpg.

    Modo 1 (PDF + Excel pareados): usa o texto de cada página do PDF para associar
    a imagem à NC correta do Excel (nc (1) = 1ª NC do Excel, etc.).
    Modo 2 (só PDF): ordem das páginas = ordem nc (1), nc (2)...
    Modo 3 (pasta): pasta_pdfs = pasta com vários PDFs (um por NC).

    Retorna o número de fotos extraídas.
    """
    pasta_saida_pdf = pasta_saida_pdf or M02_FOTOS_PDF
    pasta_saida_nc = pasta_saida_nc or M02_FOTOS_NC

    if not fitz:
        logger.error("Instale PyMuPDF: pip install pymupdf")
        return 0

    extraidas = 0

    # Modo 1: PDF pareado + Excel → alinhar imagem ao tipo de NC pelo texto da página
    if arquivo_pdf and arquivo_pdf.exists():
        logger.info(f"Extraindo fotos do PDF pareado: {arquivo_pdf.name}")
        ncs_excel = _ler_ncs_do_excel(arquivo_excel) if arquivo_excel and arquivo_excel.exists() else []
        if not ncs_excel:
            try:
                n_ext = _extrair_um_pdf_via_pdf_extractor(
                    arquivo_pdf, pasta_saida_pdf, pasta_saida_nc
                )
                if n_ext > 0:
                    logger.info(
                        "  %s: %s JPG(s) via nc_artesp.pdf_extractor (recorte constatação).",
                        arquivo_pdf.name,
                        n_ext,
                    )
                    return n_ext
            except Exception as e:
                logger.warning(
                    "  pdf_extractor indisponível ou falhou (%s); a usar extração legado (1.ª imagem/página).",
                    e,
                )
        paginas = _extrair_paginas_com_imagem_e_texto(arquivo_pdf)
        if not paginas:
            logger.warning(
                "Nenhuma imagem encontrada no PDF. "
                "Verifique: 1) PyMuPDF instalado (pip install pymupdf) 2) PDF não corrompido 3) Páginas com conteúdo visual."
            )
            return 0
        if ncs_excel:
            logger.info(f"Alinhando {len(paginas)} imagem(ns) às {len(ncs_excel)} NC(s) do Excel (por texto da página).")
            associadas = _associar_imagens_a_ncs(paginas, ncs_excel)
        else:
            associadas = [(i + 1, img) for i, (_, img) in enumerate(paginas)]
        total = len(associadas)
        for pos, (n, img_bytes) in enumerate(associadas, start=1):
            if callback_progresso:
                callback_progresso(pos, total, f"Extraindo NC {n}/{total}")
            dest_pdf = pasta_saida_pdf / f"nc ({n}).jpg"
            dest_nc = pasta_saida_nc / f"PDF ({n}).jpg"
            ok_pdf = _redimensionar_e_salvar(img_bytes, dest_pdf, M02_FOTO_PDF_W, M02_FOTO_PDF_H)
            ok_nc = _redimensionar_e_salvar(
                img_bytes, dest_nc, M02_FOTO_W, M02_FOTO_H, dpi=(M02_FOTO_DPI_X, M02_FOTO_DPI_Y)
            )
            if ok_pdf or ok_nc:
                extraidas += 1
                logger.info(f"  ✓ NC {n}: PDF ({n}).jpg (pasta NC), nc ({n}).jpg (pasta PDF)")
        logger.info(f"Extração concluída: {extraidas} foto(s).")
        return extraidas

    # Modo 2: pasta com vários PDFs (um por NC)
    if pasta_pdfs and pasta_pdfs.exists():
        pdfs = sorted(p for p in pasta_pdfs.glob("*.pdf") if not p.name.startswith("~"))
        if not pdfs:
            logger.warning(f"Nenhum PDF encontrado em: {pasta_pdfs}")
            return 0
        logger.info(f"Extraindo fotos de {len(pdfs)} PDF(s)...")
        for idx, pdf_path in enumerate(pdfs, start=1):
            if callback_progresso:
                callback_progresso(idx, len(pdfs), f"Extraindo: {pdf_path.name[:50]}")
            img_bytes = _extrair_primeira_imagem_pdf(pdf_path)
            if img_bytes:
                n = idx
                dest_pdf = pasta_saida_pdf / f"nc ({n}).jpg"
                dest_nc = pasta_saida_nc / f"PDF ({n}).jpg"
                ok_pdf = _redimensionar_e_salvar(img_bytes, dest_pdf, M02_FOTO_PDF_W, M02_FOTO_PDF_H)
                ok_nc = _redimensionar_e_salvar(
                    img_bytes, dest_nc, M02_FOTO_W, M02_FOTO_H, dpi=(M02_FOTO_DPI_X, M02_FOTO_DPI_Y)
                )
                if ok_pdf or ok_nc:
                    extraidas += 1
                    logger.info(f"  ✓ {pdf_path.name} → PDF ({n}).jpg (pasta NC), nc ({n}).jpg (pasta PDF)")
            else:
                logger.warning(f"  Nenhuma imagem encontrada em: {pdf_path.name}")
        logger.info(f"Extração concluída: {extraidas} foto(s).")
        return extraidas

    return 0
