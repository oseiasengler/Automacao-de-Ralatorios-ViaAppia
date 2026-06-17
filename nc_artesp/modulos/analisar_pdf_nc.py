"""
nc_artesp/modulos/analisar_pdf_nc.py
────────────────────────────────────────────────────────────────────────────
Analisa o PDF de Constatação de Rotina Artesp:
  1. Extrai os apontamentos (NCs) do texto do PDF sem alterar nada.
  2. Ordena por rodovia → sentido → km e detecta saltos de KM.
  3. Alerta sobre trechos sem apontamento onde panelas (buraco/recalque)
     podem estar não registradas — prazo legal: 24 h.
  4. Gera relatório PDF com:
       • Capa com resumo e estatísticas
       • Alertas de salto de KM
       • NCs emergenciais (prazo ≤ 24 h)
       • NCs agrupadas por tipo, em sequência de KM, dados originais intactos

Dependências: PyMuPDF (fitz), reportlab — já presentes no requirements.txt.
Desenvolvedor: Ozeias Engler
"""

from __future__ import annotations

import io
import logging
import os
import re
from copy import deepcopy
import unicodedata
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)

try:
    import fitz
    FITZ_OK = True
except ImportError:
    FITZ_OK = False

if FITZ_OK:
    try:
        fitz.TOOLS.mupdf_display_errors(False)
    except Exception:
        pass

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, PageBreak, KeepTogether,
    )
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    openpyxl = None
    OPENPYXL_OK = False

# CONFIGURAÇÕES

LIMIAR_GAP_KM   = 2.0    # gap entre NCs consecutivas (km) para gerar alerta
PRAZO_EMERG_MAX = 1      # prazo ≤ 1 dia = emergencial (24 h)


def _norm_lote_numero(lote: Any) -> str:
    """Primeiro grupo de dígitos normalizado (ex.: '050' → '50'). Vazio se não houver."""
    s = str(lote if lote is not None else "").strip()
    m = re.search(r"\d+", s)
    if not m:
        return ""
    try:
        return str(int(m.group(0)))
    except ValueError:
        return m.group(0)


# Carrega o mapa EAF do config (pode ser sobreposto via parâmetro)
try:
    from nc_artesp.config import MAPA_EAF as _MAPA_EAF_PADRAO
except ImportError:
    try:
        from config import MAPA_EAF as _MAPA_EAF_PADRAO
    except ImportError:
        _MAPA_EAF_PADRAO = []

# Palavras-chave que identificam possível panela/buraco no tipo de NC
PALAVRAS_PANELA = [
    "buraco", "panela", "remendo", "afundamento", "recalque",
    "ruptura", "deformação", "deformacao", "trinca", "fissura",
    "escorregamento", "deslizamento", "depressão", "depressao",
    "irregular", "desgaste", "trilha",
]

# Atividades cujo prazo legal ARTESP é 24 h (buraco/panela)
ATIVIDADES_24H = [
    "buraco", "panela", "remendo profundo", "recapeamento de buraco",
]


# MODELO DE DADOS

@dataclass
class NcItem:
    codigo: str           = ""
    data_con: str         = ""   # DD/MM/AAAA
    horario_fiscalizacao: str = ""  # HH:MM:SS (col 5 do relatório)
    km_ini_str: str       = ""   # "50 + 950"
    km_fim_str: str       = ""
    km_ini: float         = 0.0  # valor decimal (50.950)
    km_fim: float         = 0.0
    sentido: str          = ""
    atividade: str        = ""   # descrição completa (col 17; template Artemig: «Observações da fiscalização»)
    tipo_atividade: str   = ""   # ex.: Faixa de Domínio (col 15)
    grupo_atividade: str  = ""   # ex.: Limpeza (col 16)
    observacao: str       = ""
    rodovia: str          = ""   # "SP 075"
    rodovia_nome: str     = ""
    lote: str             = ""
    concessionaria: str   = ""
    prazo_str: str        = ""   # DD/MM/AAAA
    prazo_dias: Optional[int] = None
    emergencial: bool     = False
    tipo_panela: bool     = False  # atividade sugere panela/buraco
    grupo: int            = 0    # grupo EAF (col V template)
    empresa: str          = ""   # nome da empresa fiscalizadora (col 20 EAF)
    nome_fiscal: str      = ""   # responsável técnico (col 21)
    origem_ma: bool       = False
    tipo_artemig: str     = ""   # ex.: QID (col A template Artemig)
    sh_artemig: str       = ""   # ex.: SH02 (col B)
    num_consol: str       = ""   # só para dedupe; exibição em observacao (Artemig)
    patologia_artemig: str = ""  # mapeamento Nas01 → col. Tipo (Kcor)
    indicador_artemig: str = ""
    artemig_pdf_stem: str = ""  # nome do PDF sem .p/Exportar Kcor col. V/W
    artemig_kcor_paginas_jpg: list = field(default_factory=list)  # páginas c/ foto (heurística legada)
    artemig_kcor_nomes_arquivos: list = field(default_factory=list)  # nomes reais nc (cod).jpg / _N.jpg → col. W



@dataclass
class GapAlerta:
    """Salto de KM dentro de um mesmo grupo/rodovia/sentido."""
    grupo: int
    empresa: str
    rodovia: str
    sentido: str
    km_antes: float
    km_depois: float
    gap_km: float
    nc_antes: str    # código fiscalização
    nc_depois: str


@dataclass
class CodigoGapAlerta:
    """Salto na numeração sequencial do Código Fiscalização dentro de um grupo.
    Indica apontamentos que foram gerados mas não foram entregues à concessionária
    — potencialmente buracos/panelas na pista (prazo 24 h)."""
    grupo: int
    empresa: str
    codigo_antes: str        # último código entregue
    codigo_depois: str       # próximo código entregue
    n_faltantes: int         # quantos estão faltando entre os dois
    codigos_faltantes: list  # lista dos códigos ausentes (até 10 por alerta)


def _atividade_exibicao_relatorio(nc: NcItem) -> str:
    try:
        from modulos.nome_resposta_saida import _texto_atividade_exibicao_pendentes
    except ImportError:
        from .nome_resposta_saida import _texto_atividade_exibicao_pendentes

    base = (nc.atividade or "").strip()
    if base:
        return _texto_atividade_exibicao_pendentes(
            {"atividade": base, "mae_atividade_q": "", "tipo_nc": ""}
        )
    blob = " ".join(
        filter(
            None,
            [(nc.tipo_atividade or "").strip(), (nc.grupo_atividade or "").strip()],
        )
    ).strip()
    if blob:
        return _texto_atividade_exibicao_pendentes(
            {"atividade": blob, "mae_atividade_q": "", "tipo_nc": ""}
        )
    return ""


# PARSER DO PDF

def _km_para_float(s: str) -> float:
    """'50 + 950' → 50.950 ; '67 + 000' → 67.0"""
    m = re.match(r'(\d+)\s*\+\s*(\d+)', s.strip())
    if m:
        return int(m.group(1)) + int(m.group(2)) / 1000.0
    try:
        return float(s.replace(",", "."))
    except Exception:
        return 0.0


def _parse_data(s: str) -> Optional[datetime]:
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s.strip(), fmt)
        except ValueError:
            pass
    return None


def _prazo_dias(data_con: str, prazo: str) -> Optional[int]:
    d1 = _parse_data(data_con)
    d2 = _parse_data(prazo)
    if d1 and d2:
        return max(0, (d2 - d1).days)
    return None


def _is_panela(atividade: str) -> bool:
    a = (atividade or "").lower()
    return any(p in a for p in PALAVRAS_PANELA)


def _is_panela_artemig_nc(nc: NcItem) -> bool:
    """Artemig: atividade vazia; usar indicador/patologia/grupo/tipo."""
    if getattr(nc, "tipo_panela", False):
        return True
    if (getattr(nc, "lote", None) or "").strip() != "50":
        return False
    blob = " ".join(
        [
            nc.tipo_atividade or "",
            nc.grupo_atividade or "",
            nc.atividade or "",
            getattr(nc, "observacao", "") or "",
        ]
    )
    return _is_panela(blob)


def _tipo_evento_panela_nascentes(base: str, modo: str) -> str:
    t = (base or "").strip() or "Panelas"
    t = re.sub(r"(?i)\s*-\s*(emergencial|reparo\s+t[eé]cnico)\s*$", "", t).strip()
    if not t:
        t = "Panelas"
    if modo == "emergencial":
        return f"{t} - Emergencial"
    return f"{t} - Reparo Técnico"


def _expandir_eventos_panela_nascentes(ncs: list[NcItem], lote_num: str) -> list[NcItem]:
    """Nascentes: para panela/buraco, gerar evento emergencial (1 dia) e reparo técnico (15 dias)."""
    if (lote_num or "").strip() != "50":
        return ncs
    out: list[NcItem] = []
    for nc in ncs:
        if not _is_panela_artemig_nc(nc):
            out.append(nc)
            continue
        d0 = _parse_data(nc.data_con or "")
        if not d0:
            out.append(nc)
            continue
        base_tipo = (nc.tipo_atividade or nc.atividade or "Panelas").strip()
        desc_em = _tipo_evento_panela_nascentes(base_tipo, "emergencial")
        desc_tec = _tipo_evento_panela_nascentes(base_tipo, "tecnico")
        nc_emerg = deepcopy(nc)
        nc_emerg.prazo_dias = 1
        nc_emerg.prazo_str = (d0.date() + timedelta(days=1)).strftime("%d/%m/%Y")
        nc_emerg.emergencial = True
        nc_emerg.tipo_panela = True
        nc_emerg.tipo_atividade = desc_em
        base_act = (nc.atividade or "").strip()
        nc_emerg.atividade = (
            f"{base_act} — {desc_em}" if base_act else desc_em
        )
        nc_tec = deepcopy(nc)
        nc_tec.prazo_dias = 15
        nc_tec.prazo_str = (d0.date() + timedelta(days=15)).strftime("%d/%m/%Y")
        nc_tec.emergencial = False
        nc_tec.tipo_panela = True
        nc_tec.tipo_atividade = desc_tec
        nc_tec.atividade = (
            f"{base_act} — {desc_tec}" if base_act else desc_tec
        )
        out.extend([nc_emerg, nc_tec])
    return out


# Grupo = resumo da atividade; Tipo = descrição do tipo. (palavras na atividade, grupo, tipo)
_MAPA_ATIVIDADE_GRUPO_TIPO: list[tuple[list[str], str, str]] = [
    (["erosao", "erosão", "bueiro", "ruptura de bueiro"], "Erosão", "Recuperação de erosão"),
    (["defensa", "defensas", "acidentada", "reposição de defensa"], "Defensas", "Reposição de defensa acidentada"),
    (["pavimento", "buraco", "panela", "recapeamento", "depressao", "depressão"], "Pavimento", "Pavimento"),
    (["cerca", "cercas", "alambrado", "telamento", "vedos"], "Segurança", "Segurança"),
    (["limpeza", "lixo", "faixa de dominio", "faixa de domínio"], "Limpeza", "Faixa de Domínio"),
    (["drenagem", "dreno", "sarjeta", "galeria"], "Drenagem", "Drenagem"),
    (["sinalizacao", "sinalização", "conifica", "conificação", "fita zebrada"], "Sinalização", "Sinalização"),
]


def _inferir_grupo_tipo_da_atividade(atividade: str) -> tuple[str, str]:
    """Infere grupo (resumo) e tipo de atividade a partir do texto da atividade quando não preenchidos."""
    if not (atividade or "").strip():
        return ("", "")
    a = (atividade or "").strip().lower()
    for keywords, grupo, tipo in _MAPA_ATIVIDADE_GRUPO_TIPO:
        if any(kw in a for kw in keywords):
            return (grupo, tipo)
    return ("", "")


# Mínimo de caracteres por página para considerar que há texto nativo (senão tenta OCR)
_EXTRAIR_TEXTO_MIN_LEN = 40


def _extrair_texto_pdf(pdf_bytes: bytes) -> str:
    """Extrai todo o texto do PDF em uma string única. Inclui versão por blocos para tabelas.
    Para PDFs escaneados (sem camada de texto), usa OCR (pytesseract) se disponível."""
    if not FITZ_OK:
        raise ImportError("PyMuPDF não instalado: pip install pymupdf")
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    partes = []
    partes_blocks = []
    for pag in doc:
        txt = pag.get_text("text") or ""
        if len(txt.strip()) < _EXTRAIR_TEXTO_MIN_LEN:
            try:
                from nc_artesp.pdf_ocr import texto_de_pagina_ocr
                ocr = texto_de_pagina_ocr(pag, dpi=200)
                if ocr:
                    txt = ocr
            except Exception as e:
                logger.debug("OCR página %s: %s", pag.number + 1, e)
        partes.append(txt)
        try:
            blocs = pag.get_text("blocks")
            if blocs:
                # Blocos PyMuPDF não vêm em ordem de leitura
                blocs.sort(key=lambda b: (round(b[1], 0), round(b[0], 0)))
                blocos_str = "\n".join((b[4] or "").strip() for b in blocs if (b[4] or "").strip())
                if len(blocos_str.strip()) < _EXTRAIR_TEXTO_MIN_LEN and txt.strip():
                    blocos_str = txt
                partes_blocks.append(blocos_str)
            elif txt.strip():
                partes_blocks.append(txt)
        except Exception:
            if txt.strip():
                partes_blocks.append(txt)
    doc.close()
    texto = "\n".join(partes)
    if partes_blocks:
        texto_blocks = "\n".join(partes_blocks)
        if texto_blocks and texto_blocks != texto:
            texto = texto + "\n" + texto_blocks
    return texto


def _sentido_para_texto(s: str) -> str:
    """Converte letra (L/O/N/S/I/E) para nome completo no relatório de análise."""
    s = (s or "").strip().upper()
    if not s:
        return ""
    letra = s[0] if s else ""
    if letra == "0":
        letra = "O"
    mapa = {"L": "Leste", "O": "Oeste", "N": "Norte", "S": "Sul", "I": "Interna", "E": "Externa"}
    return mapa.get(letra, s)


def _texto_observacao_pdf_valido(s: str) -> bool:
    t = (s or "").strip()
    if len(t) < 2:
        return False
    if re.fullmatch(r"[\s:;|,.:-]+", t):
        return False
    if re.fullmatch(r"\d{5,}", t):
        return False
    return True


def _linha_boilerplate_apos_observacao_artesp(ln: str) -> bool:
    t = (ln or "").strip()
    if not t:
        return True
    if re.match(r"^Observa[cç][aã]o\s*:?\s*$", t, re.IGNORECASE):
        return True
    if re.match(r"^Obs\.?\s*:?\s*$", t, re.IGNORECASE):
        return True
    if re.match(r"^Rodovia\s*\(SP\)\s*:", t, re.IGNORECASE):
        return True
    if re.match(r"^\d{2}/\d{2}/\d{4}$", t):
        return True
    if re.match(r"^Rodovia\s*:", t, re.IGNORECASE):
        return True
    if re.match(r"^Constata[cç][aã]o\s*-\s*", t, re.IGNORECASE):
        return True
    if re.match(r"^C[oó]digo\s+Fiscaliza[cç][aã]o\s*:", t, re.IGNORECASE):
        return True
    if re.match(r"^Lote\s*:", t, re.IGNORECASE):
        return True
    if re.match(r"^Concession[aá]ria\s*:", t, re.IGNORECASE):
        return True
    if re.match(r"^Km\+m\s*-\s*", t, re.IGNORECASE):
        return True
    if re.match(r"^Km\s*[-\s]*Inicial", t, re.IGNORECASE):
        return True
    if re.match(r"^Inicial\s*:?\s*$", t, re.IGNORECASE) or re.match(r"^Final\s*:?\s*$", t, re.IGNORECASE):
        return True
    if re.match(r"^\d+\s*\+\s*\d+", t) and len(t) < 48:
        return True
    if re.match(r"^Sentido\s*:", t, re.IGNORECASE):
        return True
    if re.match(r"^Data\s+Limite\b", t, re.IGNORECASE):
        return True
    if re.match(r"^Atividade\s*:", t, re.IGNORECASE):
        return True
    if re.fullmatch(r"Leste|Oeste|Norte|Sul|Interna|Externa", t, re.IGNORECASE):
        return True
    if re.fullmatch(r"Marginal\s+(Oeste|Leste)", t, re.IGNORECASE):
        return True
    return False


def _outro_codigo_fiscal_linha_artesp(ln: str, codigo_nc: str) -> bool:
    if not (codigo_nc or "").strip():
        return False
    m = re.match(r"^C[oó]digo\s+Fiscaliza[cç][aã]o\s*:\s*(\d+)", (ln or "").strip(), re.IGNORECASE)
    if not m:
        return False
    return m.group(1).strip() != str(codigo_nc).strip()


def _coletar_observacao_artesp_linhas(lines: list[str], codigo_nc: str, *, ambito_estrito: bool) -> str:
    codigo_nc = (codigo_nc or "").strip()
    candidatos: list[str] = []
    n = len(lines)
    for i, ln in enumerate(lines):
        if not (
            re.search(r"Observa[cç][aã]o\s*:", ln, re.IGNORECASE)
            or re.search(r"^Obs\.?\s*:", ln, re.IGNORECASE)
        ):
            continue
        for pat in (
            r"Observa[cç][aã]o\s*:+\s*(.+)",
            r"Obs\.?\s*:+\s*(.+)",
        ):
            m_inline = re.search(pat, ln, re.IGNORECASE)
            if not m_inline:
                continue
            frag = (m_inline.group(1) or "").strip()
            if _texto_observacao_pdf_valido(frag) and not _linha_boilerplate_apos_observacao_artesp(frag):
                candidatos.append(frag)
        if i > 0:
            prev = lines[i - 1].strip()
            if (
                _texto_observacao_pdf_valido(prev)
                and not _linha_boilerplate_apos_observacao_artesp(prev)
                and not re.match(r"^Atividade\s*:", prev, re.IGNORECASE)
            ):
                candidatos.append(prev)
        j = i + 1
        while j < n and j < i + 55:
            seg = lines[j]
            if ambito_estrito and codigo_nc and _outro_codigo_fiscal_linha_artesp(seg, codigo_nc):
                break
            if (
                re.match(r"^Observa[cç][aã]o\s*:?\s*$", seg, re.IGNORECASE)
                or re.match(r"^Obs\.?\s*:?\s*$", seg, re.IGNORECASE)
            ) and j > i:
                break
            if _linha_boilerplate_apos_observacao_artesp(seg):
                j += 1
                continue
            cand = seg.strip()
            if _texto_observacao_pdf_valido(cand):
                candidatos.append(cand)
            j += 1
    if not candidatos:
        return ""
    return max(candidatos, key=len)[:12000]


def _parse_nc_block(block: str) -> Optional[NcItem]:
    """
    Analisa um bloco de texto correspondente a uma NC.

    Estrutura observada no PDF ARTESP (com variações de layout/OCR):
      {data_con}   Constatação -
      Código Fiscalização: Lote: {codigo} Concessionária: {conc}
      {km_ini}   Km+m - Inicial: {km_fim}   Km+m - Final: {sentido}   Sentido:
      Data Limite para Reparo -
      Atividade: {atividade}
      [{observacao}]   Observação: / Obs.:
      Rodovia (SP): {rodovia}
    """
    nc = NcItem()
    lines = [ln.strip() for ln in block.splitlines() if ln.strip()]
    # Uma única linha com espaços (para regex que cruzam quebra de linha)
    texto_flat = " ".join(lines)

    # ── data constatação (primeira linha com padrão de data)
    for ln in lines[:5]:
        m = re.match(r'^(\d{2}/\d{2}/\d{4})', ln)
        if m:
            nc.data_con = m.group(1)
            break
    if not nc.data_con:
        m = re.search(r'Data\s+(?:Constata[cç][aã]o\s*[:\-]?\s*)?(\d{2}/\d{2}/\d{4})', texto_flat, re.IGNORECASE)
        if m:
            nc.data_con = m.group(1)
    if not nc.data_con:
        for i, ln in enumerate(lines[:8]):
            if re.match(r'^Data\s*$', ln, re.IGNORECASE) or re.match(r'^Data\s+Constata[cç][aã]o\s*$', ln, re.IGNORECASE):
                if i + 1 < len(lines) and re.match(r'^\d{2}/\d{2}/\d{4}', lines[i + 1]):
                    nc.data_con = lines[i + 1].strip()[:10]
                    break

    # ── código fiscalização + concessionária (vários formatos)
    for ln in lines:
        m = re.search(
            r'C[oó]digo\s+Fiscaliza[cç][aã]o:\s*Lote:\s*(\S+)\s+Concession[aá]ria:\s*(.+)',
            ln, re.IGNORECASE
        )
        if m:
            nc.codigo = m.group(1).strip()
            nc.concessionaria = m.group(2).strip()
            break
    if not nc.codigo:
        for ln in lines:
            m = re.search(r'Cod\.?\s*Fiscal\.?\s*:?\s*(\d+)', ln, re.IGNORECASE)
            if m:
                nc.codigo = m.group(1).strip()
                break
    if not nc.codigo:
        m = re.search(r'C[oó]digo\s+Fiscal[.:]?\s*(\d+)', texto_flat, re.IGNORECASE)
        if m:
            nc.codigo = m.group(1).strip()
    if not nc.codigo:
        # Layout de sinalização costuma trazer "Código 6031" (sem "Fiscalização").
        m = re.search(r'\bC[oó]digo\s*:?\s*(\d{4,})\b', texto_flat, re.IGNORECASE)
        if m:
            nc.codigo = m.group(1).strip()
    if not nc.codigo:
        m = re.search(r'C[oó]digo(?!\s+da)(?!\s+Fiscaliza)\s*:?\s*(\d{4,})', texto_flat, re.IGNORECASE)
        if m:
            nc.codigo = m.group(1).strip()
    if not nc.codigo:
        m = re.search(r'(\d{4,})\s+C[oó]digo(?!\s+da)(?!\s+Fiscaliza)', texto_flat, re.IGNORECASE)
        if m:
            nc.codigo = m.group(1).strip()
    if not nc.codigo:
        for i, ln in enumerate(lines):
            if re.match(r'^C[oó]digo\s*:?\s*$', ln, re.IGNORECASE):
                if i + 1 < len(lines) and re.match(r'^\d{4,}\s*$', lines[i + 1].strip()):
                    nc.codigo = lines[i + 1].strip()
                    break
    if not nc.codigo:
        for i, ln in enumerate(lines):
            if re.match(r'^Codigo\s*:?\s*$', ln, re.IGNORECASE):
                if i + 1 < len(lines) and re.match(r'^\d{4,}\s*$', lines[i + 1].strip()):
                    nc.codigo = lines[i + 1].strip()
                    break
    if not nc.codigo:
        # Último recurso: lote só é aceito como código para valores > 3 dígitos.
        m = re.search(r'Lote:\s*(\d+)', texto_flat, re.IGNORECASE)
        if m:
            cand = m.group(1).strip()
            if len(cand) > 3:
                nc.codigo = cand
    if not nc.codigo and re.search(r'\bC[oó]digo\b', texto_flat, re.IGNORECASE):
        for i, ln in enumerate(lines[:12]):
            if not re.match(r'^\d{4,}\s*$', ln.strip()) or ln.strip().upper().startswith("LOTE"):
                continue
            adj = (lines[i - 1] if i > 0 else "") + " " + (lines[i + 1] if i + 1 < len(lines) else "")
            if re.search(r'C[oó]digo', adj, re.IGNORECASE):
                nc.codigo = ln.strip()
                break
    # Layout emergencial: "Código" numa linha, valor (5+ dígitos) mais abaixo, antes de "Lote:" ou "Data da Constatação"
    if not nc.codigo and re.search(r'\bC[oó]digo\b', texto_flat, re.IGNORECASE):
        idx_codigo = next((i for i, ln in enumerate(lines) if re.match(r'^C[oó]digo\s*$', ln, re.IGNORECASE)), -1)
        if idx_codigo >= 0:
            for ln in lines[idx_codigo + 1: min(idx_codigo + 15, len(lines))]:
                if re.match(r'^Lote\s*:?\s*', ln, re.IGNORECASE) or re.match(r'^Data\s+da\s+Constata', ln, re.IGNORECASE):
                    break
                if re.match(r'^\d{5,}\s*$', ln.strip()):
                    nc.codigo = ln.strip()
                    break

    # ── km inicial / km final / sentido (uma linha ou texto contínuo)
    for ln in lines:
        m = re.search(
            r'(\d+\s*\+\s*\d+)\s+Km\+m\s*-\s*Inicial:\s*(\d+\s*\+\s*\d+)\s+Km\+m\s*-\s*Final:\s*(.+?)\s+Sentido:',
            ln, re.IGNORECASE
        )
        if m:
            nc.km_ini_str = m.group(1).strip()
            nc.km_fim_str = m.group(2).strip()
            nc.sentido = _sentido_para_texto(m.group(3).strip())
            nc.km_ini = _km_para_float(nc.km_ini_str)
            nc.km_fim = _km_para_float(nc.km_fim_str)
            break
    if not nc.km_ini_str:
        m = re.search(
            r'(\d+\s*\+\s*\d+)\s+Km\+m\s*-\s*Inicial:\s*(\d+\s*\+\s*\d+)\s+Km\+m\s*-\s*Final:\s*(.+?)\s+Sentido:',
            texto_flat, re.IGNORECASE
        )
        if m:
            nc.km_ini_str = m.group(1).strip()
            nc.km_fim_str = m.group(2).strip()
            nc.sentido = _sentido_para_texto(m.group(3).strip())
            nc.km_ini = _km_para_float(nc.km_ini_str)
            nc.km_fim = _km_para_float(nc.km_fim_str)
    if not nc.km_ini_str:
        # Alternativa: Km (sem +m), Inicial/Final em qualquer ordem
        m = re.search(
            r'(\d+\s*\+\s*\d+)\s+Km\s*[-\s]*Inicial\s*:?\s*(\d+\s*\+\s*\d+)\s+Km\s*[-\s]*Final\s*:?\s*([LONSIE0])\s*Sentido',
            texto_flat, re.IGNORECASE
        )
        if m:
            nc.km_ini_str = m.group(1).strip()
            nc.km_fim_str = m.group(2).strip()
            nc.sentido = _sentido_para_texto(m.group(3).strip())
            nc.km_ini = _km_para_float(nc.km_ini_str)
            nc.km_fim = _km_para_float(nc.km_fim_str)
    if not nc.km_ini_str:
        # Padrão mais solto: dois "km+m" e depois sentido
        m = re.search(
            r'(\d+\s*\+\s*\d+)\s+.*?Inicial\s*:?\s*(\d+\s*\+\s*\d+)\s+.*?Final\s*:?\s*([LONSIE0]?)\s*Sentido',
            texto_flat, re.IGNORECASE | re.DOTALL
        )
        if m:
            nc.km_ini_str = m.group(1).strip()
            nc.km_fim_str = m.group(2).strip()
            nc.sentido = _sentido_para_texto((m.group(3) or "").strip())
            nc.km_ini = _km_para_float(nc.km_ini_str)
            nc.km_fim = _km_para_float(nc.km_fim_str)
    if not nc.km_ini_str and re.search(r'Km|Inicial|Final|Sentido', texto_flat, re.I):
        # Último recurso: dois "número + número" (km ini e km fim) no bloco
        m = re.findall(r'(\d+\s*\+\s*\d+)', texto_flat)
        if len(m) >= 2:
            nc.km_ini_str = m[0].strip()
            nc.km_fim_str = m[1].strip()
            nc.km_ini = _km_para_float(nc.km_ini_str)
            nc.km_fim = _km_para_float(nc.km_fim_str)
        elif len(m) == 1:
            nc.km_ini_str = nc.km_fim_str = m[0].strip()
            nc.km_ini = nc.km_fim = _km_para_float(nc.km_ini_str)
    if not nc.km_ini_str and re.search(r'Km\+m', texto_flat, re.I):
        for i, ln in enumerate(lines):
            if re.search(r'Km\+m', ln, re.IGNORECASE):
                for j in range(i + 1, min(i + 5, len(lines))):
                    km_m = re.search(r'(\d+\s*\+\s*\d+)', lines[j])
                    if km_m:
                        nc.km_ini_str = nc.km_fim_str = km_m.group(1).strip()
                        nc.km_ini = nc.km_fim = _km_para_float(nc.km_ini_str)
                        break
                break
    if not nc.km_ini_str or not nc.km_fim_str:
        for i, ln in enumerate(lines):
            if re.match(r'^Inicial\s*:?\s*$', ln, re.IGNORECASE) and i + 1 < len(lines):
                km = re.search(r'(\d+\s*\+\s*\d+)', lines[i + 1])
                if km and not nc.km_ini_str:
                    nc.km_ini_str = km.group(1).strip()
                    nc.km_ini = _km_para_float(nc.km_ini_str)
            if re.match(r'^Final\s*:?\s*$', ln, re.IGNORECASE) and i + 1 < len(lines):
                km = re.search(r'(\d+\s*\+\s*\d+)', lines[i + 1])
                if km and not nc.km_fim_str:
                    nc.km_fim_str = km.group(1).strip()
                    nc.km_fim = _km_para_float(nc.km_fim_str)
    if not nc.sentido and nc.km_ini_str:
        m = re.search(r'Sentido\s*:?\s*([LONSIE0])\b', texto_flat, re.IGNORECASE)
        if m:
            nc.sentido = _sentido_para_texto(m.group(1).strip())
        if not nc.sentido:
            m = re.search(r'\b([LONSIE0])\s+Sentido\b', texto_flat, re.IGNORECASE)
            if m:
                nc.sentido = _sentido_para_texto(m.group(1).strip())
    if not nc.sentido:
        for i, ln in enumerate(lines):
            if re.match(r'^Sentido\s*:?\s*$', ln, re.IGNORECASE) and i + 1 < len(lines):
                next_ln = lines[i + 1].strip()
                s = re.search(r'\b([LONSIE0])\b', next_ln, re.IGNORECASE)
                if s:
                    nc.sentido = _sentido_para_texto(s.group(1).strip())
                    break
                if next_ln.lower() in ("sul", "norte", "leste", "oeste", "interna", "externa"):
                    nc.sentido = next_ln
                    break

    # ── lote (formato emergencial "Lote: 13" ou após "Data Limite para Reparo"; não confundir com "Lote: 896643" do código)
    m = re.search(r'Lote\s*:?\s*(\d+)', texto_flat, re.IGNORECASE)
    if m:
        val = m.group(1).strip()
        if val != (nc.codigo or "").strip() or len(val) <= 3:
            nc.lote = val
    if not nc.lote:
        for i, ln in enumerate(lines):
            if re.match(r'^Lote\s*:?\s*$', ln, re.IGNORECASE) and i + 1 < len(lines):
                next_val = lines[i + 1].strip()
                if re.match(r'^\d+$', next_val) and (next_val != (nc.codigo or "").strip() or len(next_val) <= 3):
                    nc.lote = next_val
                    break
    if not nc.lote:
        limite_idx = next(
            (i for i, ln in enumerate(lines) if "Limite para Reparo" in ln), -1
        )
        if limite_idx >= 0:
            for ln in lines[limite_idx + 1: limite_idx + 4]:
                if re.match(r'^\d+$', ln.strip()):
                    nc.lote = ln.strip()
                    break

    # ── atividade (padrões normais primeiro; "Evento:" só como fallback para PDF emergencial)
    for ln in lines:
        m = re.match(r'Atividade\s*:?\s*(.+?)(?=\s*Grupo\s|\s*Tipo\s|$)', ln, re.IGNORECASE | re.DOTALL)
        if m and m.group(1).strip():
            nc.atividade = m.group(1).strip()
            break
    if not nc.atividade:
        for ln in lines:
            m = re.match(r'Atividade:\s*(.+)', ln, re.IGNORECASE)
            if m:
                nc.atividade = m.group(1).strip()
                break
    if not nc.atividade:
        for i, ln in enumerate(lines):
            if re.match(r'^Atividade\s*:?\s*$', ln, re.IGNORECASE):
                if i + 1 < len(lines) and lines[i + 1].strip():
                    nc.atividade = lines[i + 1].strip()[:200]
                    break
    if not nc.atividade:
        for ln in lines:
            m = re.search(r'Evento\s*:?\s*(.+)', ln, re.IGNORECASE)
            if m and m.group(1).strip():
                nc.atividade = m.group(1).strip()[:200]
                break
    if not nc.atividade:
        for i, ln in enumerate(lines):
            if re.match(r'^Evento\s*:?\s*$', ln, re.IGNORECASE):
                if i + 1 < len(lines) and lines[i + 1].strip():
                    nc.atividade = lines[i + 1].strip()[:200]
                    break

    # ── horário da fiscalização (col E): consta nos PDFs como "Horário da(s) constatação(ões)" ou "Horário da Fiscalização"
    _padrao_hora = r'(\d{1,2}:\d{2}(?::\d{2})?)'
    _padrao_hora_h = r'(\d{1,2})h\s*(\d{2})'
    m = re.search(r'Hor[áa]rio\s*(?:da\s*)?Fiscaliza[cç][aã]o\s*:?\s*' + _padrao_hora, texto_flat, re.IGNORECASE)
    if m:
        nc.horario_fiscalizacao = m.group(1).strip()
    if not nc.horario_fiscalizacao:
        m = re.search(r'Hor[áa]rio\s*(?:das?\s*)?Constata[cç][oõ]es?\s*:?\s*' + _padrao_hora, texto_flat, re.IGNORECASE)
        if m:
            nc.horario_fiscalizacao = m.group(1).strip()
    if not nc.horario_fiscalizacao:
        m = re.search(r'Hora\s*:?\s*' + _padrao_hora, texto_flat, re.IGNORECASE)
        if m:
            nc.horario_fiscalizacao = m.group(1).strip()
    if not nc.horario_fiscalizacao:
        m = re.search(r'Constata[cç][aã]o\s*[:\-]?\s*' + _padrao_hora, texto_flat, re.IGNORECASE)
        if m:
            nc.horario_fiscalizacao = m.group(1).strip()
        if not nc.horario_fiscalizacao:
            m = re.search(_padrao_hora + r'\s*[-\s]*Constata[cç][aã]o', texto_flat, re.IGNORECASE)
            if m:
                nc.horario_fiscalizacao = m.group(1).strip()
    if not nc.horario_fiscalizacao:
        m = re.search(_padrao_hora, texto_flat)
        if m:
            nc.horario_fiscalizacao = m.group(1)
    if not nc.horario_fiscalizacao:
        m = re.search(r'(\d{1,2})h\s*(\d{2})', texto_flat)
        if m:
            nc.horario_fiscalizacao = "{}:{}".format(m.group(1), m.group(2))
    for i, ln in enumerate(lines):
        if re.search(r'Hor[áa]rio\s*(?:da\s*)?(?:Fiscaliza[cç][aã]o|Constata[cç][oõ]es?)|Hora\b|Constata[cç][aã]o\s*-\s*$', ln, re.IGNORECASE):
            m = re.search(_padrao_hora, ln)
            if m:
                nc.horario_fiscalizacao = m.group(1).strip()
                break
            if i + 1 < len(lines):
                m = re.search(r'(\d{1,2}:\d{2}(?::\d{2})?|\d{1,2}h\s*\d{2})', lines[i + 1])
                if m:
                    h = m.group(1).replace("h", ":").replace(" ", "")
                    nc.horario_fiscalizacao = h
                    break
            break
    for i, ln in enumerate(lines):
        if not nc.horario_fiscalizacao and re.search(r'\bConstata[cç][aã]o\b', ln, re.IGNORECASE):
            m = re.search(_padrao_hora, ln)
            if m:
                nc.horario_fiscalizacao = m.group(1).strip()
                break
            if i + 1 < len(lines):
                m = re.search(_padrao_hora, lines[i + 1])
                if m:
                    nc.horario_fiscalizacao = m.group(1).strip()
                    break

    # ── tipo de atividade (col O) / grupo de atividade (col P)
    # PDF pode vir: "Atividade ... Grupo Pavimento Tipo Faixa de domínio" ou "Grupo: Limpeza Tipo: Segurança"
    # Valores comuns: Tipo = Faixa de Domínio, Pavimento, Segurança; Grupo = Limpeza, Pavimento, Vedos Cercas Alambrados e Telamento
    def _limpar_valor(s: str, max_len: int) -> str:
        s = (s or "").strip()
        for prefix in ("Atividade", "Grupo", "Tipo", "Observação", "Obs"):
            m = re.match(re.escape(prefix) + r"\s*[,:]?\s*", s, re.IGNORECASE)
            if m:
                s = s[m.end() :].strip()
                break
        return s[:max_len] if s else ""

    # texto_flat: padrões "Tipo de Atividade: X" / "Grupo de Atividade: X" ou "Tipo X" / "Grupo X"
    # Parar antes de Grupo, Tipo, Atividade, Data, Observação (para não capturar o próximo campo)
    m = re.search(
        r'Tipo\s*(?:de\s*)?[Aa]tividade\s*[,:]?\s*([^\n|;]+?)(?=\s*Grupo\s|\s*Atividade\s*[,:]|\s*Data\s|Observa[cç][aã]o|\s*$)',
        texto_flat, re.IGNORECASE
    )
    if m:
        nc.tipo_atividade = _limpar_valor(m.group(1), 200)
    m = re.search(
        r'Grupo\s*(?:de\s*)?[Aa]tividade\s*[,:]?\s*([^\n|;]+?)(?=\s*Tipo\s|\s*Atividade\s*[,:]|\s*Data\s|Observa[cç][aã]o|\s*$)',
        texto_flat, re.IGNORECASE
    )
    if m:
        nc.grupo_atividade = _limpar_valor(m.group(1), 100)
    # "Tipo" / "Grupo" sozinhos: "Tipo Faixa de dominio", "Grupo Limpeza", "Tipo, Segurança", "Grupo, Vedos, Cercas..."
    if not nc.tipo_atividade:
        m = re.search(r'\bTipo\s*(?:de\s*atividade)?\s*[,:]?\s*([^\n;]+?)(?=\s*Grupo\s|\s*Atividade\s*[,:]|\s*Data\s|Observa[cç][aã]o|\s*$)', texto_flat, re.IGNORECASE)
        if m:
            nc.tipo_atividade = _limpar_valor(m.group(1), 200)
    if not nc.grupo_atividade:
        m = re.search(r'\bGrupo\s*(?:de\s*atividade)?\s*[,:]?\s*([^\n;]+?)(?=\s*Tipo\s|\s*Atividade\s*[,:]|\s*Data\s|Observa[cç][aã]o|\s*$)', texto_flat, re.IGNORECASE)
        if m:
            nc.grupo_atividade = _limpar_valor(m.group(1), 100)

    # Por linha: "Tipo de Atividade: Faixa de Domínio" ou "Tipo Faixa de dominio" ou "Tipo, Segurança"
    for i, ln in enumerate(lines):
        m = re.match(r'Tipo\s*(?:de\s*)?[Aa]tividade\s*[,:]?\s*(.+)', ln, re.IGNORECASE)
        if m and m.group(1).strip():
            nc.tipo_atividade = _limpar_valor(m.group(1), 200)
            break
        m = re.match(r'Tipo\s*[,:]?\s*(.+)', ln, re.IGNORECASE)
        if m and m.group(1).strip() and not re.match(r'^(de\s+)?[Aa]tividade\s*[,:]?\s*$', m.group(1).strip(), re.IGNORECASE):
            nc.tipo_atividade = _limpar_valor(m.group(1), 200)
            break
        if re.match(r'Tipo\s*(?:de\s*)?[Aa]tividade\s*[,:]?\s*$', ln, re.IGNORECASE) and i + 1 < len(lines):
            nc.tipo_atividade = _limpar_valor(lines[i + 1], 200)
            break
        if re.match(r'Tipo\s*[,:]?\s*$', ln, re.IGNORECASE) and i + 1 < len(lines):
            nc.tipo_atividade = _limpar_valor(lines[i + 1], 200)
            break
    for i, ln in enumerate(lines):
        m = re.match(r'Grupo\s*(?:de\s*)?[Aa]tividade\s*[,:]?\s*(.+)', ln, re.IGNORECASE)
        if m and m.group(1).strip():
            nc.grupo_atividade = _limpar_valor(m.group(1), 100)
            break
        m = re.match(r'Grupo\s*[,:]?\s*(.+)', ln, re.IGNORECASE)
        if m and m.group(1).strip() and not re.match(r'^(de\s+)?[Aa]tividade\s*[,:]?\s*$', m.group(1).strip(), re.IGNORECASE):
            nc.grupo_atividade = _limpar_valor(m.group(1), 100)
            break
        if re.match(r'Grupo\s*(?:de\s*)?[Aa]tividade\s*[,:]?\s*$', ln, re.IGNORECASE) and i + 1 < len(lines):
            nc.grupo_atividade = _limpar_valor(lines[i + 1], 100)
            break
        if re.match(r'Grupo\s*[,:]?\s*$', ln, re.IGNORECASE) and i + 1 < len(lines):
            nc.grupo_atividade = _limpar_valor(lines[i + 1], 100)
            break

    # Fallback: "grupo" / "tipo" em qualquer posição da linha (ex.: "depressão atividade grupo pavimento tipo Pavimento")
    for ln in lines:
        if not nc.grupo_atividade:
            m = re.search(r'\bgrupo\b\s*(?:de\s*atividade)?\s*[,:]?\s*(.+?)(?=\s*tipo\b|\s*atividade\s*[,:]|\s*$)', ln, re.IGNORECASE)
            if m and m.group(1).strip():
                nc.grupo_atividade = _limpar_valor(m.group(1), 100)
        if not nc.tipo_atividade:
            m = re.search(r'\btipo\b\s*(?:de\s*atividade)?\s*[,:]?\s*(.+?)(?=\s*grupo\b|\s*atividade\s*[,:]|\s*$)', ln, re.IGNORECASE)
            if m and m.group(1).strip():
                nc.tipo_atividade = _limpar_valor(m.group(1), 200)
    for ln in lines:
        if not nc.grupo_atividade:
            m = re.search(r'\bgrupo\b\s*[,:]?\s*(.+)', ln, re.IGNORECASE)
            if m and m.group(1).strip():
                v = _limpar_valor(m.group(1), 200)
                if v and v.lower() not in ("de", "atividade"):
                    if " tipo " in (" " + v.lower() + " "):
                        v = v.split(" tipo ")[0].strip()
                    nc.grupo_atividade = v[:100]
                    break
        if not nc.tipo_atividade:
            m = re.search(r'\btipo\b\s*[,:]?\s*(.+)', ln, re.IGNORECASE)
            if m and m.group(1).strip():
                v = _limpar_valor(m.group(1), 200)
                if v and v.lower() not in ("de", "atividade"):
                    if " grupo " in (" " + v.lower() + " "):
                        v = v.split(" grupo ")[0].strip()
                    nc.tipo_atividade = v[:200]
                    break

    # ── observação (vários formatos; PDFs com «Observação::» não podem gravar só «:»)
    codigo_obs = (nc.codigo or "").strip()
    for ln in lines:
        m = re.match(r"^(.+?)\s+Observa[cç][aã]o\s*:?\s*$", ln)
        if m and _texto_observacao_pdf_valido(m.group(1)):
            nc.observacao = m.group(1).strip()
            break
    if not _texto_observacao_pdf_valido(nc.observacao):
        nc.observacao = ""
    if not nc.observacao:
        cand = _coletar_observacao_artesp_linhas(lines, codigo_obs, ambito_estrito=True)
        if cand:
            nc.observacao = cand
    if not nc.observacao:
        cand = _coletar_observacao_artesp_linhas(lines, codigo_obs, ambito_estrito=False)
        if cand:
            nc.observacao = cand
    if not nc.observacao:
        for ln in lines:
            m = re.search(r"Observa[cç][aã]o\s*:+\s*(.+)", ln, re.IGNORECASE)
            if m and _texto_observacao_pdf_valido(m.group(1)):
                frag = m.group(1).strip()
                if not _linha_boilerplate_apos_observacao_artesp(frag):
                    nc.observacao = frag
                    break
    if not nc.observacao:
        for ln in lines:
            m = re.search(r"Obs\.?\s*:+\s*(.+)", ln, re.IGNORECASE)
            if m and _texto_observacao_pdf_valido(m.group(1)):
                frag = m.group(1).strip()
                if not _linha_boilerplate_apos_observacao_artesp(frag):
                    nc.observacao = frag
                    break
    try:
        from utils.kartado_observacao_pdf import normalizar_observacao_extraida_pdf
    except ImportError:
        try:
            from nc_artesp.utils.kartado_observacao_pdf import normalizar_observacao_extraida_pdf
        except ImportError:
            normalizar_observacao_extraida_pdf = None
    if normalizar_observacao_extraida_pdf and nc.observacao:
        nc.observacao = normalizar_observacao_extraida_pdf(nc.observacao)[:12000]

    # ── rodovia SP (ex: "SP 075"); formato emergencial "RodoviaSP 075" (sem espaço)
    for ln in lines:
        m = re.search(r'Rodovia\s*\(SP\)\s*:?\s*(.+)', ln, re.IGNORECASE)
        if m and m.group(1).strip():
            nc.rodovia = m.group(1).strip()
            break
    if not nc.rodovia:
        m = re.search(r'Rodovia\s*SP\s*(\d+)', texto_flat, re.IGNORECASE)
        if m:
            nc.rodovia = "SP " + m.group(1).strip()
    if not nc.rodovia:
        m = re.search(r'RodoviaSP\s*(\d+)', texto_flat, re.IGNORECASE)
        if m:
            nc.rodovia = "SP " + m.group(1).strip()
    if not nc.rodovia:
        m = re.search(r'Rodovia\s*\(SP\)\s*:?\s*(\S+(?:\s+\S+)?)', texto_flat, re.IGNORECASE)
        if m:
            nc.rodovia = m.group(1).strip()
    if not nc.rodovia:
        for i, ln in enumerate(lines):
            if re.match(r'^Rodovia\s*\(SP\)\s*:?\s*$', ln, re.IGNORECASE):
                if i + 1 < len(lines) and lines[i + 1].strip():
                    nc.rodovia = lines[i + 1].strip().split()[0] if lines[i + 1].strip() else ""
                    if len(lines[i + 1].strip().split()) > 1:
                        nc.rodovia = lines[i + 1].strip()[:30]
                    break
    if not nc.rodovia:
        for i, ln in enumerate(lines):
            if re.match(r'^Rodovia\s*:?\s*$', ln, re.IGNORECASE):
                if i + 1 < len(lines) and re.search(r'SP\s*\d+', lines[i + 1], re.IGNORECASE):
                    nc.rodovia = lines[i + 1].strip()[:30]
                    break

    # ── prazo: data isolada APÓS a rodovia SP ou "Prazo" / "Data Limite" / "Data Programada Término" (emergencial)
    m = re.search(r'Data\s+Programada\s+T[eé]rmino\s*:?\s*(\d{2}/\d{2}/\d{4})', texto_flat, re.IGNORECASE)
    if m:
        nc.prazo_str = m.group(1)
    if not nc.prazo_str:
        m = re.search(r'Data\s+de\s+Execu[cç][aã]o\s+T[eé]rmino\s*:?\s*(\d{2}/\d{2}/\d{4})', texto_flat, re.IGNORECASE)
        if m:
            nc.prazo_str = m.group(1)
    if not nc.prazo_str:
        m = re.search(r'Prazo\s*:?\s*(\d{2}/\d{2}/\d{4})', texto_flat, re.IGNORECASE)
        if m:
            nc.prazo_str = m.group(1)
    if not nc.prazo_str:
        m = re.search(r'Data\s+Limite\s+(?:para\s+Reparo\s*)?:?\s*(\d{2}/\d{2}/\d{4})', texto_flat, re.IGNORECASE)
        if m:
            nc.prazo_str = m.group(1)
    if not nc.prazo_str:
        for i, ln in enumerate(lines):
            if re.search(r'Data\s+Programada\s+T[eé]rmino\s*:?\s*$', ln, re.IGNORECASE) or re.search(r'Data\s+de\s+Execu[cç][aã]o\s+T[eé]rmino\s*:?\s*$', ln, re.IGNORECASE):
                if i + 1 < len(lines) and re.match(r'^\d{2}/\d{2}/\d{4}', lines[i + 1]):
                    nc.prazo_str = lines[i + 1].strip()[:10]
                    break
    if not nc.prazo_str:
        for i, ln in enumerate(lines):
            if re.match(r'^Prazo\s*:?\s*$', ln, re.IGNORECASE) or re.search(r'Data\s+Limite\s+para\s+Reparo\s*$', ln, re.IGNORECASE):
                if i + 1 < len(lines) and re.match(r'^\d{2}/\d{2}/\d{4}', lines[i + 1]):
                    nc.prazo_str = lines[i + 1].strip()[:10]
                    break
    if not nc.prazo_str:
        rodsp_idx = next(
            (i for i, ln in enumerate(lines) if "Rodovia (SP):" in ln or re.search(r'Rodovia\s*\(SP\)', ln)), -1
        )
        if rodsp_idx >= 0:
            for ln in lines[rodsp_idx + 1: rodsp_idx + 5]:
                if re.match(r'^\d{2}/\d{2}/\d{4}$', ln.strip()):
                    nc.prazo_str = ln.strip()
                    break

    for ln in lines:
        m = re.match(r'Rodovia:\s*(.+)', ln, re.IGNORECASE)
        if m:
            nc.rodovia_nome = m.group(1).strip()
            break

    if nc.data_con and nc.prazo_str:
        nc.prazo_dias = _prazo_dias(nc.data_con, nc.prazo_str)
        if nc.prazo_dias is not None:
            nc.emergencial = nc.prazo_dias <= PRAZO_EMERG_MAX

    nc.tipo_panela = _is_panela(nc.atividade)

    # Se KM final não foi extraído, usar o mesmo que KM inicial (evento localizado; extensão vem na NC quando houver)
    if (nc.km_ini_str or nc.km_ini) and (not nc.km_fim_str or nc.km_fim == 0.0):
        nc.km_fim_str = nc.km_ini_str or ""
        nc.km_fim = nc.km_ini

    # Fallback final: preencher E, O, P a partir de padrões soltos no bloco
    if not nc.horario_fiscalizacao:
        m = re.search(r'\b(\d{1,2}:\d{2}(?::\d{2})?)\b', block)
        if m:
            nc.horario_fiscalizacao = m.group(1)
    if not nc.tipo_atividade or not nc.grupo_atividade:
        VALORES_TIPO = ("Faixa de Domínio", "Faixa de dominio", "Pavimento", "Segurança", "Conservação")
        VALORES_GRUPO = ("Limpeza", "Pavimento", "Segurança", "Vedos", "Cercas Alambrados e Telamento", "Conservação")
        for i, ln in enumerate(lines):
            ln_clean = ln.strip()
            if not nc.tipo_atividade:
                for v in VALORES_TIPO:
                    if v.lower() in ln_clean.lower() and len(ln_clean) < 80:
                        if i > 0 and re.search(r'\btipo\b', lines[i - 1], re.IGNORECASE):
                            nc.tipo_atividade = ln_clean[:200]
                            break
                        if re.match(r'^' + re.escape(v) + r'\s*$', ln_clean, re.IGNORECASE):
                            nc.tipo_atividade = v[:200]
                            break
                if not nc.tipo_atividade and re.search(r'\b(Faixa\s+de\s+[Dd]om[ií]nio|Pavimento|Seguran[cç]a)\b', ln_clean):
                    nc.tipo_atividade = re.search(r'(Faixa\s+de\s+[Dd]om[ií]nio|Pavimento|Seguran[cç]a)', ln_clean, re.IGNORECASE).group(1).strip()[:200]
            if not nc.grupo_atividade:
                for v in VALORES_GRUPO:
                    if v.lower() in ln_clean.lower() and len(ln_clean) < 120:
                        if i > 0 and re.search(r'\bgrupo\b', lines[i - 1], re.IGNORECASE):
                            nc.grupo_atividade = ln_clean[:100]
                            break
                        if re.match(r'^' + re.escape(v) + r'\s*$', ln_clean, re.IGNORECASE):
                            nc.grupo_atividade = v[:100]
                            break
                if not nc.grupo_atividade and re.search(r'\b(Limpeza|Pavimento|Vedos|Cercas)', ln_clean, re.IGNORECASE):
                    m = re.search(r'((?:Vedos[,\s]*)?Cercas\s+Alambrados\s+e\s+Telamento|Limpeza|Pavimento|Seguran[cç]a)', ln_clean, re.IGNORECASE)
                    if m:
                        nc.grupo_atividade = m.group(1).strip()[:100]

    # Regra de negócio: pavimento → Pavimento (tipo e grupo); cerca/cercas → Segurança; Limpeza permanece no grupo
    texto_nc = " ".join(filter(None, [nc.atividade, nc.tipo_atividade, nc.grupo_atividade, block])).lower()
    if re.search(r'\b(pavimento|depress[aã]o|buraco|recapeamento|panela|reparo\s+de\s+pavimento)\b', texto_nc):
        nc.tipo_atividade = "Pavimento"
        nc.grupo_atividade = "Pavimento"
    elif re.search(r'\b(cerca|cercas|alambrado|telamento|vedos|reparo\s+de\s+cerca)\b', texto_nc):
        nc.tipo_atividade = "Segurança"
        nc.grupo_atividade = "Segurança"
    # Limpeza: grupo = Limpeza, tipo = Faixa de Domínio (não sobrescrever se já for Limpeza)
    if re.search(r'\b(limpeza|remo[cç][aã]o\s+de\s+lixo|faixa\s+de\s+dom[ií]nio)\b', texto_nc) and nc.grupo_atividade:
        if "limpeza" in (nc.grupo_atividade or "").lower():
            if not nc.tipo_atividade:
                nc.tipo_atividade = "Faixa de Domínio"

    # Grupo = resumo da atividade; Tipo = descrição do tipo. Inferir da atividade quando vazios.
    if (not nc.grupo_atividade or not nc.tipo_atividade) and nc.atividade:
        grupo_inf, tipo_inf = _inferir_grupo_tipo_da_atividade(nc.atividade)
        if grupo_inf and not nc.grupo_atividade:
            nc.grupo_atividade = grupo_inf
        if tipo_inf and not nc.tipo_atividade:
            nc.tipo_atividade = tipo_inf

    # NC válida precisa ao menos de código ou atividade
    if not (nc.codigo or nc.atividade):
        return None
    if logger.isEnabledFor(logging.DEBUG) and (not nc.tipo_atividade or not nc.grupo_atividade or not nc.horario_fiscalizacao):
        logger.debug(
            "NC block (tipo=%r grupo=%r horario=%r) text sample:\n%.800s",
            nc.tipo_atividade, nc.grupo_atividade, nc.horario_fiscalizacao, block
        )
    return nc


def _atribuir_grupo(nc: NcItem, mapa_eaf: list[dict]) -> None:
    """
    Atribui grupo EAF e empresa por trecho (MAPA_EAF). Só sobrescreve quando o mapa
    devolve valor; mapa vazio (ex. lote 50) não apaga empresa já vinda do PDF (rodapé EAF).
    """
    from nc_artesp.utils.helpers import obter_grupo_empresa_por_trecho
    grupo, empresa = obter_grupo_empresa_por_trecho(nc.rodovia, nc.km_ini, mapa_eaf)
    if grupo:
        nc.grupo = grupo
    if empresa:
        nc.empresa = empresa


def parse_pdf_nc(pdf_bytes: bytes) -> list[NcItem]:
    """Extrai NCs do PDF (Constatação/Código). Atributos NcItem = CAMPOS_TEMPLATE_LIST. Ordenado por rodovia, sentido, km_ini."""
    texto = _extrair_texto_pdf(pdf_bytes)

    # Layout tabular de Sinalização (Quadro Resumo): prioriza caminho rápido/estável
    # para evitar colapso em um único código no parser legado.
    texto_norm = (
        (texto or "")
        .replace("‑", "-")
        .replace("–", "-")
        .replace("—", "-")
    )
    padrao_tabela_header = r"C.{0,2}digo\s*A.*C.{0,2}digo\s*B.*C.{0,2}digo\s*C"
    padrao_tabela_linha = (
        r"(?<!\d)\d{1,4}\s+\d+\+\d{3}\s*-\s*[A-Za-zÀ-ÿ]+\s+"
        r"\d{1,2}\s+\d{1,2}\s+\d{1,2}\s+.+?\s+\d{2}/\d{2}/\d{4}\s+\d{2}/\d{2}/\d{4}\s+"
        r"(?:\d+\s*semana(?:s)?|\d+\s*dias?|(?:\d+|48)\s*horas?)\s+"
        r"(?:A\s+Verificar|Pendente)(?:\s+0)?\b"
    )
    parece_tabela_sinalizacao = bool(
        re.search(padrao_tabela_header, texto_norm, re.IGNORECASE | re.DOTALL)
        or re.search(padrao_tabela_linha, texto_norm, re.IGNORECASE | re.DOTALL)
    )
    if parece_tabela_sinalizacao:
        ncs_tabela = _parse_pdf_tabela_sinalizacao(texto_norm)
        if len(ncs_tabela) >= 2:
            for nc in ncs_tabela:
                _atribuir_grupo(nc, _MAPA_EAF_PADRAO)
            ncs_tabela.sort(key=lambda n: (n.grupo or 999, n.rodovia, n.sentido, n.km_ini))
            return ncs_tabela

    # Remove cabeçalhos de página e rodapés
    texto = re.sub(r'Relat[oó]rio de Conserva[cç][aã]o de Rotina\s*', '', texto)
    texto = re.sub(r'--\s*\d+\s*of\s*\d+\s*--', '', texto)
    texto = re.sub(r'^\s*\d+\s*\d{2}/\d{2}/\d{4}\s*$', '', texto, flags=re.MULTILINE)

    # Divide nos blocos de NC (começa em data + "Constatação -")
    partes = re.split(r'(?=\d{2}/\d{2}/\d{4}\s+Constatação)', texto)

    ncs: list[NcItem] = []
    for bloco in partes:
        if "Constatação" not in bloco:
            continue
        nc = _parse_nc_block(bloco)
        if nc:
            ncs.append(nc)

    # PDF com legenda só "Código" (sem "Constatação" ao lado da data): divide por "Código" + número
    if len(ncs) == 0 and re.search(r'C[oó]digo\s*:?\s*\d{4,}', texto, re.IGNORECASE):
        partes_alt = re.split(r'(?=C[oó]digo\s*:?\s*\d{4,})', texto, flags=re.IGNORECASE)
        for bloco in partes_alt:
            bloco = bloco.strip()
            if not bloco or not re.search(r'\d{4,}', bloco):
                continue
            nc = _parse_nc_block(bloco)
            if nc:
                ncs.append(nc)
    # Layout emergencial: cada NC começa com linha só "Código" (valor em linha abaixo ou mais adiante)
    if len(ncs) == 0 and re.search(r'(?m)^C[oó]digo\s*$', texto, re.IGNORECASE):
        partes_alt = re.split(r'(?m)(?=^C[oó]digo\s*$)', texto, flags=re.IGNORECASE)
        for bloco in partes_alt:
            bloco = bloco.strip()
            if not bloco or not re.search(r'C[oó]digo', bloco, re.IGNORECASE):
                continue
            nc = _parse_nc_block(bloco)
            if nc:
                ncs.append(nc)

    # Layout de tabela (Sinalização): prioriza parser tabular quando o layout é detectado.
    padrao_tabela_sinalizacao = r"C.{0,2}digo\s*A.*C.{0,2}digo\s*B.*C.{0,2}digo\s*C"
    if re.search(padrao_tabela_sinalizacao, texto, re.IGNORECASE):
        ncs_tabela = _parse_pdf_tabela_sinalizacao(texto)
        if ncs_tabela and (len(ncs) == 0 or _ncs_codigos_pouco_confiaveis(ncs) or len(ncs_tabela) >= len(ncs)):
            ncs = ncs_tabela

    # Atribui EAF/Grupo e ordena: Grupo → Rodovia → Sentido → KM
    for nc in ncs:
        _atribuir_grupo(nc, _MAPA_EAF_PADRAO)
    ncs.sort(key=lambda n: (n.grupo or 999, n.rodovia, n.sentido, n.km_ini))
    return ncs


def _parse_pdf_tabela_sinalizacao(texto: str) -> list[NcItem]:
    """Extrai NCs de PDF em formato tabular (Sinalização) sem Excel de apoio."""
    if not texto:
        return []

    m_rod = re.search(r"\b(SP|MG|BR)\s*[- ]?\s*(\d{2,3})\b", texto, re.IGNORECASE)
    rodovia = ""
    if m_rod:
        rodovia = f"{m_rod.group(1).upper()} {int(m_rod.group(2)):03d}"

    m_dc = re.search(r"\b(\d{2}/\d{2}/\d{4})\b", texto)
    data_con = m_dc.group(1) if m_dc else ""

    out: list[NcItem] = []
    texto_lin = re.sub(r"[ \t]+", " ", (texto or "").replace("\xa0", " "))
    texto_lin = (
        texto_lin.replace("‑", "-")
        .replace("–", "-")
        .replace("—", "-")
    )

    # Captura robusta no texto inteiro (mesmo com quebras de linha no meio da descrição).
    padrao_global = re.compile(
        r"(?<!\d)(\d{1,4})\s+(\d+\+\d{3})\s*[-]\s*([A-Za-zÀ-ÿ]+)\s+"
        r"(\d{1,2})\s+(\d{1,2})\s+(\d{1,2})\s+"
        r"(.+?)\s+"
        r"(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+"
        r"((?:\d+\s*semana(?:s)?|\d+\s*dias?|(?:\d+|48)\s*horas?))\s+"
        r"(A\s+Verificar|Pendente)(?:\s+0)?\b",
        re.IGNORECASE | re.DOTALL,
    )

    encontrados = list(padrao_global.finditer(texto_lin))
    for m in encontrados:
        item_id = m.group(1).strip()
        km_txt = m.group(2).strip().replace(" ", "")
        sentido = m.group(3).strip().title()
        cod_a = m.group(4).strip()
        cod_b = m.group(5).strip().zfill(2)
        cod_c = m.group(6).strip()
        atividade = re.sub(r"\s+", " ", (m.group(7) or "").strip())[:200]
        data_const = (m.group(8) or "").strip() or data_con
        prazo_txt = (m.group(10) or "").strip().lower()

        prazo_dias: Optional[int] = None
        if "hora" in prazo_txt:
            prazo_dias = 1
        else:
            m_sem = re.search(r"(\d+)\s*semana", prazo_txt)
            m_dia = re.search(r"(\d+)\s*dias?", prazo_txt)
            if m_sem:
                prazo_dias = int(m_sem.group(1)) * 7
            elif m_dia:
                prazo_dias = int(m_dia.group(1))

        prazo_str = ""
        if prazo_dias is not None and data_const:
            d0 = _parse_data(data_const)
            if d0:
                prazo_str = (d0.date() + timedelta(days=prazo_dias)).strftime("%d/%m/%Y")
        if not prazo_str:
            prazo_str = data_const

        grupo_inf, tipo_inf = _inferir_grupo_tipo_da_atividade(atividade)
        km_val = _km_para_float(km_txt.replace("+", " + "))
        codigo_legado = f"{cod_a}{cod_b}{cod_c}"
        codigo_saida = f"{item_id}-{codigo_legado}"

        out.append(
            NcItem(
                codigo=codigo_saida,
                data_con=data_const,
                km_ini_str=km_txt.replace("+", " + "),
                km_fim_str=km_txt.replace("+", " + "),
                km_ini=km_val,
                km_fim=km_val,
                sentido=sentido,
                atividade=atividade,
                tipo_atividade=tipo_inf,
                grupo_atividade=grupo_inf,
                rodovia=rodovia,
                prazo_str=prazo_str,
                prazo_dias=prazo_dias,
                emergencial=(prazo_dias is not None and prazo_dias <= PRAZO_EMERG_MAX),
                tipo_panela=_is_panela(atividade),
            )
        )

    # Fallback legado: linha única completa.
    if not out:
        padrao_linha = re.compile(
            r"^\s*(\d{1,4})\s+(\d+\+\d{3})\s*[-]\s*([A-Za-zÀ-ÿ]+)\s+"
            r"(\d{1,2})\s+(\d{1,2})\s+(\d{1,2})\s+(.+?)\s+"
            r"(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+(.+?)\s+(?:A\s+Verificar|Pendente)\b",
            re.IGNORECASE,
        )
        for ln in texto.splitlines():
            linha = re.sub(r"\s+", " ", (ln or "").strip())
            if not linha:
                continue
            m = padrao_linha.match(linha)
            if not m:
                continue
            item_id = m.group(1).strip()
            km_txt = m.group(2).strip().replace(" ", "")
            sentido = m.group(3).strip().title()
            cod_a = m.group(4).strip()
            cod_b = m.group(5).strip().zfill(2)
            cod_c = m.group(6).strip()
            atividade = m.group(7).strip()[:200]
            data_const = (m.group(8) or "").strip() or data_con
            prazo_txt = (m.group(10) or "").strip().lower()
            prazo_dias: Optional[int] = None
            if "hora" in prazo_txt:
                prazo_dias = 1
            else:
                m_sem = re.search(r"(\d+)\s*semana", prazo_txt)
                m_dia = re.search(r"(\d+)\s*dias?", prazo_txt)
                if m_sem:
                    prazo_dias = int(m_sem.group(1)) * 7
                elif m_dia:
                    prazo_dias = int(m_dia.group(1))
            prazo_str = ""
            if prazo_dias is not None and data_const:
                d0 = _parse_data(data_const)
                if d0:
                    prazo_str = (d0.date() + timedelta(days=prazo_dias)).strftime("%d/%m/%Y")
            if not prazo_str:
                prazo_str = data_const
            grupo_inf, tipo_inf = _inferir_grupo_tipo_da_atividade(atividade)
            km_val = _km_para_float(km_txt.replace("+", " + "))
            codigo_legado = f"{cod_a}{cod_b}{cod_c}"
            codigo_saida = f"{item_id}-{codigo_legado}"
            out.append(
                NcItem(
                    codigo=codigo_saida,
                    data_con=data_const,
                    km_ini_str=km_txt.replace("+", " + "),
                    km_fim_str=km_txt.replace("+", " + "),
                    km_ini=km_val,
                    km_fim=km_val,
                    sentido=sentido,
                    atividade=atividade,
                    tipo_atividade=tipo_inf,
                    grupo_atividade=grupo_inf,
                    rodovia=rodovia,
                    prazo_str=prazo_str,
                    prazo_dias=prazo_dias,
                    emergencial=(prazo_dias is not None and prazo_dias <= PRAZO_EMERG_MAX),
                    tipo_panela=_is_panela(atividade),
                )
            )

    return out


def _ncs_codigos_pouco_confiaveis(ncs: list[NcItem]) -> bool:
    cods = [(n.codigo or "").strip() for n in ncs if (n.codigo or "").strip()]
    if not cods:
        return True
    unq = set(cods)
    if len(unq) == 1 and len(cods) > 1:
        return True
    if all(re.fullmatch(r"\d{1,3}", c or "") for c in cods):
        return True
    return False


def _texto_pdf_indica_layout_artemig(texto: str) -> bool:
    """Layout típico de notificação Artemig (MG). Constatação ARTESP (SP) não dispara."""
    if not texto or len(texto.strip()) < 45:
        return False
    t = texto[:20000]
    if not re.search(r"(MG[- ]?050|BR[- ]?265|BR[- ]?491)", t, re.I):
        return False
    return bool(
        re.search(
            r"(NOTIFICA[OÇ]|N[ºo°]?\s*da\s*CONSOL|LOCALIZA[ÇC][AÃ]?O|LOCALIZA[OÇ][AÃ]O|REGISTRO\s+FOTOGR)",
            t,
            re.I,
        )
    )


def _nc_parece_artemig(nc: NcItem) -> bool:
    """NC vinda de PDF Artemig (campos exclusivos ou rodovia da malha MG)."""
    if (getattr(nc, "tipo_artemig", None) or "").strip():
        return True
    if (getattr(nc, "num_consol", None) or "").strip():
        return True
    r = re.sub(r"[\s._-]", "", (nc.rodovia or "").upper())
    if "MG050" in r:
        return True
    if "BR265" in r or "BR491" in r:
        return True
    return False


_LOTES_ANALISE = frozenset({"13", "21", "26", "50"})


def _lotes_indicados_no_texto(texto: str) -> set[str]:
    """Trechos típicos de constatação ARTESP / notificação Artemig → números de lote."""
    s: set[str] = set()
    if not texto or len(texto.strip()) < 20:
        return s
    tu = texto[:35000].upper()
    if _texto_pdf_indica_layout_artemig(texto):
        s.add("50")
    for n in _LOTES_ANALISE:
        if re.search(rf"\bLOTE\s*:?\s*{n}\b", tu):
            s.add(n)
    if "RODOVIAS DAS COLINAS" in tu:
        s.add("13")
    if re.search(r"RODOVIAS DO TIET", tu):
        s.add("21")
    if re.search(r"\bLOTE\s*:?\s*26\b", tu):
        s.add("26")
    if "NASCENTES DAS GERAIS" in tu:
        s.add("50")
    return {x for x in s if x in _LOTES_ANALISE}


def _indicios_lote_um_arquivo(texto: str, parcial: list[NcItem]) -> set[str]:
    out = _lotes_indicados_no_texto(texto)
    for nc in parcial:
        n = _lote_num_do_pdf(nc)
        if n and n in _LOTES_ANALISE:
            out.add(n)
    for nc in parcial:
        if _nc_parece_artemig(nc):
            out.add("50")
    return {x for x in out if x in _LOTES_ANALISE}


def _validar_lotes_pdf_vs_selecionado(
    bloques: list[tuple[str, str, list[NcItem]]],
    lote_num: str,
) -> None:
    """
    Um arquivo não pode misturar lotes; o lote inferido de cada PDF deve coincidir
    com o selecionado e ser único entre os arquivos.
    """
    try:
        from nc_artesp.config import _LOTE_CONCESSIONARIA as _LC
    except Exception:
        _LC = {}

    def _nome_l(n: str) -> str:
        return _LC.get(n, f"Lote {n}")

    msg_artemig = (
        "PDF(s) no formato Artemig (MG). Selecione o lote 50 (CONSOL) "
        "ou use constatações ARTESP (lotes 13, 21 ou 26)."
    )
    msg_nao_artemig = (
        "Este(s) PDF(s) são de constatação ARTESP (lotes 13, 21 ou 26). "
        "Para notificação Artemig, selecione o lote 50."
    )

    por_arquivo: list[tuple[str, str | None]] = []
    for src, texto, parcial in bloques:
        if not parcial:
            por_arquivo.append((src, None))
            continue
        ind = _indicios_lote_um_arquivo(texto, parcial)
        if len(ind) > 1:
            raise ValueError(
                f'O arquivo «{src}» indica mais de um lote ({", ".join(sorted(ind))}). '
                "Use um PDF por lote, sem misturar concessionárias."
            )
        por_arquivo.append((src, next(iter(ind)) if ind else None))

    detectados = {v for _, v in por_arquivo if v is not None}
    if len(detectados) > 1:
        raise ValueError(
            "Os PDFs são de lotes diferentes ({}). Envie apenas arquivos do **mesmo** lote.".format(
                ", ".join(_nome_l(x) for x in sorted(detectados))
            )
        )

    if len(detectados) == 1:
        unico = next(iter(detectados))
        if unico != lote_num:
            raise ValueError(
                "Os PDFs são do {} (lote {}), mas você selecionou {} (lote {}). "
                "Ajuste o lote no menu ou troque os arquivos.".format(
                    _nome_l(unico), unico, _nome_l(lote_num), lote_num
                )
            )
        return

    # Nenhum indício explícito de número de lote no conjunto
    if lote_num == "50":
        for src, texto, parcial in bloques:
            if not parcial:
                continue
            if not (
                _texto_pdf_indica_layout_artemig(texto)
                or any(_nc_parece_artemig(nc) for nc in parcial)
            ):
                raise ValueError(
                    f'Lote 50: o arquivo «{src}» não parece ser notificação Artemig (MG). {msg_artemig}'
                )
    else:
        for src, texto, parcial in bloques:
            if not parcial:
                continue
            if _texto_pdf_indica_layout_artemig(texto) or any(_nc_parece_artemig(nc) for nc in parcial):
                raise ValueError(
                    f'O arquivo «{src}» é Artemig (lote 50). {msg_nao_artemig}'
                )


def _data_artemig_dd_mm_yyyy(s: str) -> str:
    """DD/MM/AA ou DD/MM/AAAA → DD/MM/AAAA (ano com 2 dígitos: 00–69 → 20xx)."""
    try:
        from nc_artemig.texto_pdf import limpeza_profunda, normalizar_texto_extraido_pdf

        s = limpeza_profunda(normalizar_texto_extraido_pdf(s or ""))
    except ImportError:
        s = re.sub(r"\s+", " ", (s or "").replace("\xa0", " ").strip())
    m = re.search(r"(\d{2})/(\d{2})/(\d{2}|\d{4})\b", s)
    if not m:
        return s
    d, mo, y = m.group(1), m.group(2), m.group(3)
    if len(y) == 2:
        yi = int(y)
        y = str(2000 + yi if yi < 70 else 1900 + yi)
    return f"{d}/{mo}/{y}"


def _prazo_artemig(texto: str, data_con: str) -> tuple[str, Optional[int], bool]:
    """
    Artemig: data de reprovação (se houver), senão prazo em dias, senão data em linha
    útil após «Prazo para Atendimento» (nunca legenda «à Notificação:»).
    Com «Remendo Emergencial» antes de «Remendo Técnico», usa só o primeiro bloco (ex.: 24 h, não os 15 dias).
    """
    prazo_str = ""
    prazo_dias: Optional[int] = None
    emerg = False
    d0 = _parse_data(data_con) if data_con else None

    repro_m = re.search(
        r"(?:[Dd]ata\s+de\s+)?[Rr]epro(?:va[çc][aã]o|va[cç][aã]o)\s*:?\s*(\d{2}/\d{2}/\d{2,4})",
        texto,
        re.I,
    ) or re.search(
        r"[Dd]ata\s+repro\S*\s*:?\s*(\d{2}/\d{2}/\d{2,4})",
        texto,
        re.I,
    ) or re.search(
        r"[Rr]epro\.?\s*(?:[Dd]ata|em)\s*:?\s*(\d{2}/\d{2}/\d{2,4})",
        texto,
        re.I,
    )
    if repro_m:
        cand = _data_artemig_dd_mm_yyyy(repro_m.group(1).strip())
        dlim = _parse_data(cand) if cand else None
        if dlim and (not d0 or dlim.date() >= d0.date()):
            prazo_str = cand
            if data_con:
                prazo_dias = _prazo_dias(data_con, prazo_str)

    pm = re.search(r"Prazo\s+para\s+Atendimento", texto, re.I)
    janela = texto[pm.end() : pm.end() + 800] if pm else texto
    if pm:
        m_rt = re.search(r"(?i)remendo\s+t[eé]cnico", janela)
        m_re = re.search(r"(?i)remendo\s+emergencial", janela)
        if m_rt is not None and m_re is not None and m_re.start() < m_rt.start():
            janela = janela[: m_rt.start()]
    _em_ate_ndias = (
        r"em\s+at[eé]?\s+(\d{1,3})\s*(?:\([^)]{0,200}\)\s*)?dias?\s*(?:úteis|uteis|corridos)?"
    )
    em_m = re.search(_em_ate_ndias, janela, re.I) or re.search(_em_ate_ndias, texto, re.I)
    if not prazo_str and em_m and data_con:
        n = int(em_m.group(1))
        if d0:
            prazo_dias = n
            prazo_str = (d0.date() + timedelta(days=n)).strftime("%d/%m/%Y")
    elif not prazo_str and em_m:
        prazo_str = f"em até {em_m.group(1)} dias"
        prazo_dias = int(em_m.group(1))
    if prazo_dias is None and not prazo_str:
        hor_scope = janela if pm else texto
        hor_m = re.search(
            r"em\s+at[eé]?\s+(\d+)\s*(?:\([^)]{0,200}\)\s*)?horas?\b",
            hor_scope,
            re.I,
        )
        if not hor_m:
            hor_m = re.search(
                r"em\s+at[eé]?\s+(\d+)\s*(?:\([^)]{0,200}\)\s*)?horas?\b",
                texto,
                re.I,
            )
        if hor_m:
            nh = int(hor_m.group(1))
            nh = max(1, min(nh, 24 * 14))
            dias = max(1, (nh + 23) // 24)
            prazo_dias = dias
            if d0:
                prazo_str = (d0.date() + timedelta(days=dias)).strftime("%d/%m/%Y")
            else:
                prazo_str = f"em até {nh} horas"
    if prazo_dias is None and not prazo_str and pm:
        m_prazo_nd = re.search(
            r"(?i)(?:prazo\s+m[aá]ximo\s+de|no\s+prazo\s+(?:m[aá]ximo\s+)?de)\s+"
            r"(\d{1,3})\s*(?:\([^)]{0,200}\))?\s*dias?",
            janela,
        )
        if m_prazo_nd:
            n = int(m_prazo_nd.group(1))
            if 1 <= n <= 366:
                prazo_dias = n
                if d0:
                    prazo_str = (d0.date() + timedelta(days=n)).strftime("%d/%m/%Y")
                else:
                    prazo_str = f"em até {n} dias"
    if not prazo_str:
        for raw in (janela.split("\n") if pm else []):
            ln = (raw or "").strip()
            if len(ln) < 8:
                continue
            if re.search(
                r"(?i)^à\s*Notifica|^\s*Notifica[çc][aã]o\s*:?\s*$|Atendimento\s+a\s+Notifica",
                ln,
            ):
                continue
            if not re.search(r"\d{2}/\d{2}/\d{2,4}", ln):
                continue
            dm = re.search(r"(\d{2}/\d{2}/\d{2,4})", ln)
            if not dm:
                continue
            cand = _data_artemig_dd_mm_yyyy(dm.group(1).strip())
            dlim = _parse_data(cand) if cand else None
            if dlim and (not d0 or dlim.date() >= d0.date()):
                prazo_str = cand
                if data_con:
                    prazo_dias = _prazo_dias(data_con, prazo_str)
                break
    if prazo_dias is not None:
        emerg = prazo_dias <= PRAZO_EMERG_MAX
    elif prazo_str and data_con:
        pd = _prazo_dias(data_con, prazo_str)
        if pd is not None:
            prazo_dias = pd
            emerg = pd <= PRAZO_EMERG_MAX
    return prazo_str, prazo_dias, emerg


def _limpar_legenda_consol_artemig(s: str) -> str:
    """Remove rótulos Nº CONSOL / CONSOL da notificação que vazam para atividade/indicador."""
    if not (s or "").strip():
        return ""
    t = re.sub(r"(?i)N[ºoO°]?\s*(?:da\s*)?CONSOl?\s*:?\s*", "", s)
    t = re.sub(r"(?i)\bCONSOl?\s*,?\s*da\s+notifica[çc][aã]o\s*:?\s*", "", t)
    return re.sub(r"\s{2,}", " ", t).strip(" ,.;—-|")


def _limpar_legendas_campo_artemig(s: str) -> str:
    """Remove legendas típicas do PDF Artemig; mantém só o conteúdo útil do campo."""
    if not (s or "").strip():
        return ""
    t = s.strip()
    for pat in (
        r"(?i)\bNotifica[çc][aã]o\s*\|\s*Data\s*\|\s*Hora\s*\|\s*Indicador\s*\|\s*Patologia\b[^\n]*",
        r"(?i)\bNotifica[çc][aã]o\s+Data\s+Hora\s+Indicador\s+Patologia\b[^\n]*",
        r"(?i)Prazo\s+para\s+Atendimento(\s+a\s+Notifica[çc][aã]o)?[^\n]*",
        r"(?i)LOCALIZA[ÇC][AÃ]?O\s*:?",
        r"(?i)à\s+Notifica[çc][aã]o\s*:?",
        r"(?i)\bRegistro\s+Fotogr[aá]fico\b\s*",
        r"(?i)\bHor[aá]rio\s+da\s+Fiscaliza[çc][aã]o\s*:?",
        r"(?i)\bData\s+Fiscaliza[çc][aã]o\s*:?",
        r"(?i)\bC[oó]d\.?\s*Fiscaliza[çc][aã]o\s*:?",
        r"(?i)\bConcession[aá]ria\s+Lote\s*:?",
        r"(?i)\bTrecho\s*:?",
        r"(?i)\bKM\s*[Ii]nicial\b\s*:?",
        r"(?i)\bKM\s*[Ff]inal\b\s*:?",
        r"(?i)\bSentido\s*:?\s*(?=(CRESCENTE|DECRESCENTE|AMBOS)\b)",
    ):
        t = re.sub(pat, " ", t)
    t = re.sub(
        r"(?i)^\s*(Indicador|Patologia|Tipo|Descri[çc][aã]o|Observa[çc][aã]o)\s*:?\s*",
        "",
        t,
    )
    t = _limpar_legenda_consol_artemig(t)
    _toks = r"Notifica[çc][aã]o|Data|Hora|Indicador|Patologia"
    for _ in range(8):
        n = re.sub(rf"(?i)^(?:{_toks})(\s+(?:{_toks}))*\s+", "", t).strip()
        if n == t:
            break
        t = n
    t = re.sub(r"\s{2,}", " ", t).strip(" ,.;—-|")
    if re.fullmatch(
        r"(?i)(Notifica[çc][aã]o|Data|Hora|Indicador|Patologia|Tipo|SH)\s*:?",
        t,
    ):
        return ""
    try:
        from nc_artemig.texto_pdf import colapsar_espacos_pdf

        t = colapsar_espacos_pdf(t, multiline=False)
    except Exception:
        t = re.sub(r"\s+", " ", (t or "").strip())
    return t[:500] if len(t) > 500 else t


def _sentido_artemig_normalizado(rodovia: str, sentido_bruto: str) -> str:
    """CRESCENTE/DECRESCENTE/AMBOS → texto Kcor (mesma regra do XLSX / Nas01)."""
    s = re.sub(r"(?i)^\s*sentido\s*:?\s*", "", (sentido_bruto or "").strip())
    try:
        from nc_artemig.sentido_kcor import sentido_artemig_para_kcor

        return sentido_artemig_para_kcor(rodovia or "", s)
    except Exception:
        return s


def _prazo_str_valido_artemig(s: str) -> str:
    """Evita gravar legendas do PDF na coluna de data."""
    x = (s or "").strip()
    if not x:
        return ""
    if re.search(r"(?i)notifica[çc][aã]o|atendimento\s+a\s+notifica", x) and not re.search(
        r"\d{2}/\d{2}/\d{2,4}", x
    ):
        return ""
    return x


def _texto_e_bloco_legenda_atividade_artemig(s: str) -> bool:
    """PDF linearizado: CONSOL + registro + tabela localização + prazo colados numa linha."""
    if not s or len(s) < 80:
        return False
    u = s.upper()
    hits = sum(
        [
            "CONSOL" in u and ("Nº" in s or "N " in s or "DA CONSOL" in u),
            "REGISTRO" in u and "FOTO" in u,
            "KM INICIAL" in u or "KM FINAL" in u,
            "RODOVIA" in u and "SH" in u,
            "DECRESCENTE" in u or "CRESCENTE" in u,
            "EM ATÉ" in u or "(CINCO) DIAS" in u or "NOTIFICA" in u,
            "DESCRI" in u and "ÃO" in s,
        ]
    )
    return hits >= 4


_DESC_ROTULO_PREFIX = re.compile(r"^Descri[çc][aã]o\s*:?\s*", re.IGNORECASE)


def _strip_descricao_rotulos_repetidos_artemig(s: str) -> str:
    """PDF com «Descrição:» duplicado na mesma linha — fica só o corpo."""
    t = (s or "").strip()
    for _ in range(6):
        t2 = _DESC_ROTULO_PREFIX.sub("", t, count=1).strip()
        if t2 == t:
            break
        t = t2
    return t


def _extrair_descricao_atividade_artemig(texto: str) -> str:
    """
    Texto após «Descrição:» para a col. atividade (Artemig).
    QID costuma repetir o rótulo (primeiro vazio + tabela; segundo com o parágrafo) antes de LOCALIZAÇÃO.
    Fallback: alguns PDFs extraem o texto fora de ordem — «Em até N dias» fica antes da descrição real,
    ou o segundo «Descrição:» aparece após LOCALIZAÇÃO. Ambos os casos são tratados.
    """
    if not texto:
        return ""

    prose = re.compile(
        r"(?is)\b(foi\s+verificado|verificado|existência|existencia|constatad|"
        r"monitoramento|vandalismo|aspecto\s+visual|desconforto|impacto|faixa\s+de\s+dom[ií]nio)\b"
    )

    def _limpa_frag(frag: str) -> str:
        frag = _strip_descricao_rotulos_repetidos_artemig(frag)
        frag = re.split(r"(?i)\s+Prazo\s+para\s+Atendimento\b", frag)[0]
        # Alguns PDFs (ex.: FISC-NOT-26-00043) extraem o texto reordenado: «Em até N dias»
        # aparece NO MEIO do fragmento, e a descrição real vem DEPOIS dele.
        # Se a prosa descritiva estiver após «Em até», usamos esse trecho; caso contrário
        # mantemos o comportamento original (tudo antes de «Em até»).
        partes_em = re.split(r"(?i)\s+Em\s+até\s+\d+[^\n]*\n?", frag)
        if len(partes_em) > 1:
            resto = " ".join(partes_em[1:]).strip()
            frag = resto if prose.search(resto) else partes_em[0]
        frag = re.sub(r"\s+", " ", frag).strip()
        frag = _limpar_legendas_campo_artemig(frag)
        frag = _strip_descricao_rotulos_repetidos_artemig(frag)
        return frag

    def _melhor_candidato(candidatos: list[str]) -> str:
        limpos = [_limpa_frag(c) for c in candidatos]
        ok = [f for f in limpos if len(f) > 25]
        if not ok:
            curto = max(limpos, key=len) if limpos else ""
            return curto[:500] if len(curto) > 10 else ""
        for f in reversed(ok):
            if prose.search(f):
                return f[:500]
        return ok[-1][:500]

    loc_m = re.search(r"(?is)\bLOCALIZA[ÇC][AÃ]O\b", texto)
    bloco = texto[: loc_m.start()] if loc_m else texto
    partes = re.split(r"(?is)Descri[çc][aã]o\s*:?\s*", bloco)
    candidatos = [p.strip() for p in partes[1:] if (p or "").strip()]

    resultado = _melhor_candidato(candidatos) if candidatos else ""

    # Fallback: descrição não encontrada antes de LOCALIZAÇÃO — alguns PDFs colocam o segundo
    # «Descrição:» depois da seção de localização; buscar no texto completo.
    if not resultado and loc_m:
        partes_full = re.split(r"(?is)Descri[çc][aã]o\s*:?\s*", texto)
        candidatos_full = [p.strip() for p in partes_full[1:] if (p or "").strip()]
        resultado = _melhor_candidato(candidatos_full)

    return resultado


def _cortar_resto_tabela_antes_localizacao_artemig(resto: str) -> str:
    """Tabela linearizada: indicador+patologia vêm antes de MG-050 SH… / BR-265 SH…."""
    if not (resto or "").strip():
        return ""
    r = resto.strip()
    m = re.search(
        r"(?i)\s+(?=(?:MG|BR)[- ]?\d{3}\s+SH\d{2,3}\b)",
        r,
    )
    if m is not None and m.start() >= 10:
        return r[: m.start()].strip()
    m2 = re.search(
        r"(?i)\s+(?=LOCALIZA[ÇC][AÃ]?O\b)",
        r,
    )
    if m2 is not None and m2.start() >= 10:
        return r[: m2.start()].strip()
    return r


def _indicador_patologia_de_resto_artemig(resto: str) -> tuple[str, str]:
    """Indicador | Patologia sem vazar bloco de localização (comportamento uniforme entre PDFs)."""
    r = _cortar_resto_tabela_antes_localizacao_artemig(resto)
    if not r:
        return "", ""
    r = re.sub(r"\s+", " ", r).strip()
    r = re.sub(
        r"(?is)^gerais\s+(?:par[âãa]metros|parametros)\b\s*",
        "Parâmetros Gerais ",
        r,
    )
    m_sep = re.search(
        r"(?i)\s+(?=(?:Buracos|Panelas)\s+e/ou\b)",
        r,
    )
    if m_sep is not None and 4 <= m_sep.start() <= 1200:
        return r[: m_sep.start()].strip()[:2000], r[m_sep.start() :].strip()[:2000]
    if re.search(r"(?i)(?:buracos|panelas)\s+e/ou\b", r) and not re.search(r"\s+/\s+", r):
        return "", r.strip()[:2000]
    partes = [p.strip() for p in re.split(r"\s{2,}", r) if p.strip()]
    if len(partes) >= 2:
        a0, a1 = partes[0], partes[1]
        if re.match(r"(?is)^gerais$", a0) and re.search(r"(?is)par[âa]metros", a1):
            tail = " ".join(partes[2:]).strip()
            base = "Gerais (Parâmetros)"
            merged = f"{base} — {tail}" if tail else base
            return "", merged[:2000]
        return partes[0][:2000], " ".join(partes[1:])[:2000]
    sp = r.split(None, 3)
    if len(sp) >= 2 and re.match(r"(?is)^gerais$", sp[0]) and re.search(
        r"(?is)par[âa]metros", sp[1]
    ):
        tail = " ".join(sp[2:]).strip() if len(sp) > 2 else ""
        base = "Gerais (Parâmetros)"
        merged = f"{base} — {tail}" if tail else base
        return "", merged[:2000]
    if len(sp) >= 3:
        dois = f"{sp[0]} {sp[1]}".strip()
        pat_tail = " ".join(sp[2:]).strip() if len(sp) > 2 else ""
        # Indicadores de 3 palavras: tentar «tres» quando há tokens suficientes
        if len(sp) >= 4:
            tres = f"{sp[0]} {sp[1]} {sp[2]}".strip()
            pat_tail3 = sp[3].strip() if len(sp) > 3 else ""
            if re.match(r"(?i)^trilha\s+de\s+\S+$", tres):
                return tres[:2000], pat_tail3[:2000]
            if re.match(r"(?i)^afundamento\s+\S+\s+\S+", tres):
                return tres[:2000], pat_tail3[:2000]
        if re.match(r"(?i)^(?:par[âãa]metros|parametros)\s+gerais$", dois):
            return dois[:2000], pat_tail[:2000]
        if re.match(r"(?i)^drenagem\s+\S+$", dois):
            return dois[:2000], pat_tail[:2000]
        if re.match(r"(?i)^sinaliza[çc][aã]o\s+\S+$", dois):
            return dois[:2000], pat_tail[:2000]
        if re.match(r"(?i)^deforma[çc][oõ]es\s+\S+$", dois):
            return dois[:2000], pat_tail[:2000]
        if re.match(r"(?i)^afundamento\s+\S+$", dois):
            return dois[:2000], pat_tail[:2000]
        if re.match(r"(?i)^buracos\s+/\s+panelas", dois):
            return dois[:2000], pat_tail[:2000]
        if re.match(
            r"(?i)(?:Parâmetros|Parametros|Drenagem|Pavimento|Sinaliza|Defensa|Limpeza|Seguran|Deforma|Afundamento)",
            dois,
        ) and not re.match(r"(?i)^(?:par[âãa]metros|parametros)\s+gerais$", dois):
            return dois[:2000], pat_tail[:2000]
    if len(sp) == 2:
        dois2 = f"{sp[0]} {sp[1]}"
        if re.match(r"(?i)^(?:par[âãa]metros|parametros)\s+gerais$", dois2):
            return dois2[:2000], ""
    if len(sp) >= 2:
        return sp[0][:2000], " ".join(sp[1:])[:2000]
    return r[:2000], ""


def _merge_gerais_parametros_ind_pat(ind: str, pat: str) -> tuple[str, str]:
    """
    PDF QID: «Gerais» e «(Parâmetros» em colunas separadas geram Nas01 «pat (ind)» ilegível.
    Funde numa única linha de patologia para a col. U.
    """
    i = (ind or "").strip()
    p = (pat or "").strip()
    if re.match(r"(?is)^gerais$", i):
        m = re.match(r"(?is)^\(?\s*par[âa]metros\b[)\s:.-]*\s*(.*)$", p)
        if m:
            resto = (m.group(1) or "").strip()
            base = "Gerais (Parâmetros)"
            novo_p = f"{base} — {resto}" if resto else base
            return "", novo_p[:2000]
        if re.search(r"(?is)par[âa]metros", p):
            return "", f"Gerais — {p}"[:2000]
    if re.match(r"(?is)^gerais$", p) and re.search(r"(?is)par[âa]metros", i):
        resto = re.sub(r"(?is)^\(?\s*par[âa]metros\b[)\s:.-]*\s*", "", i).strip()
        base = "Gerais (Parâmetros)"
        novo = f"{base} — {resto}" if resto else base
        return "", novo[:2000]
    return ind, pat


def _eh_texto_localizacao_resumo_artemig(s: str) -> bool:
    """Linha tipo MG-050 SH06 128,450 … CRESCENTE PISTA — não é atividade."""
    if not s or len(s) < 20:
        return False
    return bool(
        re.search(r"(?i)(MG[- ]?\d{3}|BR[- ]?\d{3})\s+SH\d{2,3}", s)
    ) and bool(re.search(r"(?i)CRESCENTE|DECRESCENTE|AMBOS|PISTA|DOMÍNIO|FX\.", s))


def _extrair_num_consol_artemig(texto: str) -> str:
    """Só a numeração (6–10 dígitos) após marcadores CONSOL; ignora legenda/colação."""
    if not texto:
        return ""
    for pat in (
        r"N[ºoO°]?\s*da\s*CONSOL",
        r"N[ºoO°]?\s*CONSOL",
        r"(?<![A-Za-z0-9])CONSOL\s*(?:da\s*notifica[çc][aã]o\s*)?:?",
    ):
        for m in re.finditer(pat, texto, re.I):
            trecho = texto[m.end() : m.end() + 60]
            dm = re.search(r"\d{6,10}", trecho)
            if dm:
                return dm.group(0)
    return ""


# Artemig: grupo EAF 50 (contrato MG); MAPA próprio em nc_artemig — nunca mapa lote 13.
_GRUPO_FISCALIZACAO_ARTEMIG = "CONSOL"
_GRUPO_EAF_ARTEMIG_ANALISE = 50


def _par_km_rotulo_artemig(blob: str, label_re: str) -> tuple[int, int] | None:
    m = re.search(
        rf"(?is){label_re}\s*:?\s*(\d{{1,4}})\s*[\+\u002c,]\s*(\d{{1,3}})\b",
        blob,
    )
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))


def _par_km_metros_numa_linha_artemig(ln: str) -> tuple[int, int] | None:
    s = (ln or "").strip()
    if not s:
        return None
    m = re.match(r"^(\d{1,4})\s*\+\s*(\d{1,3})$", s) or re.match(r"^(\d{1,4}),(\d{3})$", s)
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))


def _extrair_km_pdf_artemig_vertical_linhas(linhas: list[str]) -> tuple[int, int, int, int] | None:
    """Tabela PDF: valores «km+metros» (653+500 ou 653,400) podem vir antes dos rótulos Km Inicial / Km Final."""
    idx_ini = next(
        (i for i, l in enumerate(linhas) if re.match(r"(?is)^km\.?\s*inicial$", l.strip())),
        None,
    )
    if idx_ini is None:
        return None
    idx_fin = next(
        (i for i, l in enumerate(linhas) if re.match(r"(?is)^km\.?\s*final$", l.strip())),
        None,
    )
    if idx_fin is None:
        idx_fin = idx_ini
    lo = max(0, min(idx_ini, idx_fin) - 30)
    hi = min(idx_ini, idx_fin)
    pairs: list[tuple[int, int]] = []
    for l in linhas[lo:hi]:
        p = _par_km_metros_numa_linha_artemig(l)
        if p:
            pairs.append(p)
    if not pairs:
        return None
    if len(pairs) >= 2:
        a, b = pairs[0], pairs[1]
    else:
        a = b = pairs[0]
    return (a[0], a[1], b[0], b[1])


def _extrair_km_pdf_artemig(bloco_loc: str, texto_completo: str) -> tuple[float, float, str, str] | None:
    """Km Inicial / Km Final no PDF (rótulo na mesma linha ou tabela linearizada). Não usa malha."""
    blob_src = f"{(bloco_loc or '').strip()} {(texto_completo or '').strip()}".strip()
    if not blob_src:
        return None
    try:
        from nc_artemig.texto_pdf import normalizar_texto_extraido_pdf

        blob_lin = normalizar_texto_extraido_pdf(blob_src)
    except Exception:
        blob_lin = blob_src
    linhas = [x.strip() for x in blob_lin.splitlines() if x.strip()]
    blob = re.sub(r"\s+", " ", blob_lin)

    ini = _par_km_rotulo_artemig(blob, r"km\.?\s*inicial")
    fim_l = _par_km_rotulo_artemig(blob, r"km\.?\s*final")
    if ini:
        fim = fim_l or ini
    else:
        vert = _extrair_km_pdf_artemig_vertical_linhas(linhas)
        if not vert:
            return None
        ini = (vert[0], vert[1])
        fim = (vert[2], vert[3])

    def _pair_str(p: tuple[int, int]) -> str:
        return f"{p[0]} + {p[1]}"

    ki_s, kf_s = _pair_str(ini), _pair_str(fim)
    return (_km_para_float(ki_s), _km_para_float(kf_s), ki_s, kf_s)


def _tabela_ind_pat_pre_localizacao_artemig(texto: str) -> tuple[str, str, str, str] | None:
    """Notificação + Indicador/Patologia em colunas (pipes ou linhas) antes do bloco LOCALIZAÇÃO."""
    for pat in (
        r"(?is)Indicador\s*\|\s*Patologia\s*\|\s*(\d{8,10})\s*\|\s*(\d{2}/\d{2}/\d{2,4})\s*\|\s*"
        r"(\d{1,2}:\d{2})\s*\|\s*([\s\S]+?)(?=\s*\|\s*Local\s*\|)",
        r"(?is)Indicador\s*\n\s*Patologia\s*\n\s*(\d{8,10})\s*\n\s*(\d{2}/\d{2}/\d{2,4})\s*\n\s*"
        r"(\d{1,2}:\d{2})\s*\n\s*([\s\S]+?)(?=\n\s*Local\b|\n\s*Tipo\s*:|\n\s*LOCALIZA[ÇC])",
    ):
        m = re.search(pat, texto)
        if not m:
            continue
        resto = m.group(4).strip()
        resto = re.sub(r"\s*\|\s*", " ", resto)
        resto = re.sub(r"\s+", " ", resto).strip()
        return (
            m.group(1).strip(),
            m.group(2).strip(),
            m.group(3).replace(" ", "").strip(),
            resto,
        )
    return None


def _parse_artemig_texto(texto: str) -> NcItem | None:
    """Layout Artemig MG; código NC = coluna Notificação da tabela."""
    if not texto or len(texto.strip()) < 40:
        return None
    try:
        from nc_artemig.texto_pdf import normalizar_texto_extraido_pdf

        texto = normalizar_texto_extraido_pdf(texto)
    except ImportError:
        pass
    if not re.search(r"(NOTIFICA[OÇ]|CONSOL|LOCALIZA[OÇ][AÃ]O)", texto, re.I):
        return None
    if not re.search(r"(MG[- ]?050|BR[- ]?265|BR[- ]?491)", texto, re.I):
        return None

    def _float_br(s: str) -> float:
        try:
            return float((s or "").replace(",", ".").strip())
        except Exception:
            return 0.0

    # PDFs variam («Tipo: Fiscalização», «Outros Tipo: QID», etc.); col. Tipo do relatório e Kcor ficam sempre QID.
    tipo_artemig = "QID"

    num_consol = _extrair_num_consol_artemig(texto)

    notificacao = ""
    data_con = ""
    hora = ""
    indicador = ""
    patologia = ""
    linha_tab = None
    pre_loc = _tabela_ind_pat_pre_localizacao_artemig(texto)
    loc_tab = None
    for m in re.finditer(
        r"(?is)\bLOCALIZA[ÇC][AÃ]?O\b\s*\n\s*(\d{8,10})",
        texto,
    ):
        loc_tab = m
    slice_tab = texto[loc_tab.start(1) :].lstrip() if loc_tab else ""
    if slice_tab:
        linha_tab = re.search(
            r"(?s)^\s*(\d{8,10})\s+(\d{2}/\d{2}/\d{2,4})\s+(\d{1,2}:\d{2})\s+"
            r"([\s\S]+?)(?=\n\s*(?:Tipo\s*:|Notifica[çc][aã]o\s*\n\s*Data\b))",
            slice_tab,
            re.I,
        )
    if not linha_tab:
        linha_tab = re.search(
            r"(?m)^\s*(\d{8,10})\s+(\d{2}/\d{2}/\d{2,4})\s+(\d{1,2}:\d{2})\s+(.+)$",
            texto,
        )
    if not linha_tab:
        linha_tab = re.search(
            r"(?s)(\d{8,10})\s+(\d{2}/\d{2}/\d{2,4})\s+(\d{1,2}:\d{2})\s+(.+?)(?=\n\s*\n|LOCALIZA[ÇC]|Prazo\s+para|N[ºo°]?\s*da\s*CONSOL)",
            texto,
            re.I,
        )
    use_pre = False
    if pre_loc:
        lr = len((linha_tab.group(4) or "").strip()) if linha_tab else 0
        if not linha_tab or len(pre_loc[3]) > lr + 12:
            use_pre = True
    if use_pre and pre_loc:
        notificacao, d_raw, hora, resto = pre_loc
        data_con = _data_artemig_dd_mm_yyyy(d_raw)
        indicador, patologia = _indicador_patologia_de_resto_artemig(resto)
    elif linha_tab:
        notificacao = linha_tab.group(1).strip()
        data_con = _data_artemig_dd_mm_yyyy(linha_tab.group(2).strip())
        hora = linha_tab.group(3).replace(" ", "").strip()
        resto = (linha_tab.group(4) or "").strip()
        indicador, patologia = _indicador_patologia_de_resto_artemig(resto)

    if not notificacao:
        notif_m = re.search(
            r"(?:Notifica[çc][aã]o\s+)?(\d{8,10})\s+(\d{2}/\d{2}/\d{2,4})",
            texto,
            re.I,
        )
        notificacao = notif_m.group(1) if notif_m else ""
        if notif_m and not data_con:
            data_con = _data_artemig_dd_mm_yyyy(notif_m.group(2).strip())
        if not data_con:
            data_m = re.search(r"(\d{2}/\d{2}/\d{2,4})", texto)
            data_con = _data_artemig_dd_mm_yyyy((data_m.group(1).strip() if data_m else ""))

    codigo = (notificacao or (("CE" + num_consol) if num_consol else "") or "").strip()
    if not codigo:
        fallback = re.search(r"\b(\d{9})\b", texto)
        codigo = fallback.group(1) if fallback else "Artemig-1"

    rodovia_m = re.search(r"(MG[- ]?050|BR[- ]?265|BR[- ]?491)", texto, re.I)
    rodovia = (rodovia_m.group(1).replace(" ", " ").strip()) if rodovia_m else ""
    if rodovia and "-" not in rodovia and " " in rodovia:
        rodovia = rodovia.replace(" ", "-", 1)

    loc_m = re.search(r"LOCALIZA[OÇ][AÃ]O", texto, re.I)
    bloco_loc = texto[loc_m.end() : loc_m.end() + 900] if loc_m else texto

    def _sh_de_texto(tx: str) -> str:
        m = re.search(r"(?i)\bSH\s*0*(\d{1,3})\b", tx or "")
        if m:
            n = int(m.group(1), 10)
            return f"SH{n:02d}" if n < 1000 else f"SH{n}"
        m = re.search(r"\b(SH\d{2,4})\b", tx or "", re.I)
        return m.group(1).upper() if m else ""

    sh_artemig = _sh_de_texto(bloco_loc) or _sh_de_texto(texto)

    km_blk = _extrair_km_pdf_artemig(bloco_loc if loc_m else "", texto)
    if km_blk:
        km_ini, km_fim, km_ini_str, km_fim_str = km_blk
    else:
        kms = re.findall(r"\d{2,3}[,.]\d{3}", bloco_loc) or re.findall(
            r"\d{2,3}[,.]\d{3}", texto
        )
        km_ini = _float_br(kms[0]) if len(kms) >= 1 else 0.0
        km_fim = _float_br(kms[1]) if len(kms) >= 2 else km_ini
        km_ini_str = str(km_ini).replace(".", ",")
        km_fim_str = str(km_fim).replace(".", ",")

    sentido_m = re.search(
        r"(?i)(?:Sentido\s*:?\s*)?(CRESCENTE|DECRESCENTE|AMBOS)\b",
        bloco_loc if loc_m else texto,
    ) or re.search(r"(?i)(?:Sentido\s*:?\s*)?(CRESCENTE|DECRESCENTE|AMBOS)\b", texto)
    sentido_raw = (sentido_m.group(1).strip()) if sentido_m else ""
    sentido = _sentido_artemig_normalizado(rodovia, sentido_raw)

    if not hora:
        hm = re.search(r"(\d{1,2}\s*:\s*\d{2}|\d{1,2}:\d{2})", texto)
        hora = (hm.group(1).replace(" ", "").strip()) if hm else ""

    if not indicador and not patologia:
        linhas = [ln.strip() for ln in texto.splitlines() if ln.strip()]
        for i, ln in enumerate(linhas):
            if re.search(r"\d{1,2}\s*:\s*\d{2}|\d{1,2}:\d{2}", ln) and re.search(r"\d{8,10}", ln):
                resto = re.sub(r"^\s*\d{8,10}\s+", "", ln)
                hm2 = re.search(r"(\d{2}/\d{2}/\d{2,4})\s+(\d{1,2}:\d{2})\s+(.+)$", resto)
                if hm2:
                    ind_pat = hm2.group(3).strip()
                    pp = [p.strip() for p in re.split(r"\s{2,}", ind_pat) if p.strip()]
                    if len(pp) >= 2:
                        indicador, patologia = pp[0][:2000], " ".join(pp[1:])[:2000]
                    elif pp:
                        patologia = pp[0][:2000]
                break
            if re.match(r"^\d{1,2}\s*:\s*\d{2}$|^\d{1,2}:\d{2}$", ln):
                if i + 1 < len(linhas):
                    indicador = linhas[i + 1][:2000]
                if i + 2 < len(linhas):
                    patologia = linhas[i + 2][:2000]
                break

    indicador = _limpar_legendas_campo_artemig(indicador)
    patologia = _limpar_legendas_campo_artemig(patologia)
    indicador, patologia = _merge_gerais_parametros_ind_pat(indicador, patologia)

    descricao = _extrair_descricao_atividade_artemig(texto)

    nome_rt = ""
    for pat in (
        r"Respons[aá]vel\s+T[ée]cnico\s*:?\s*([^\n\r]+?)(?:\n\s*\n|EAF\s*[:\s]|$)",
        r"Respons[aá]vel\s+T[ée]cnico\s*:?\s*\n\s*([^\n\r]+)",
        r"(?:R\.?\s*T\.?\s*Respons[aá]vel|Resp\.?\s*T[ée]cnico)\s*:?\s*([^\n\r]+)",
    ):
        rm = re.search(pat, texto, re.I | re.DOTALL)
        if rm:
            nome_rt = re.sub(r"\s+", " ", rm.group(1).strip())[:120]
            break

    def _sem_consol_em_campos(
        *partes: str, consol: str
    ) -> tuple[str, str, str]:
        """Evita repetir o Nº CONSOL em tipo/grupo/atividade (CONSOL só em observação)."""
        c = (consol or "").strip()
        out: list[str] = []
        for p in partes:
            s = (p or "").strip()
            if c and s == c:
                s = ""
            elif c:
                s = re.sub(rf"\b{re.escape(c)}\b", "", s)
                s = re.sub(r"\s{2,}", " ", s).strip(" ,.;—-")
            out.append(s)
        return (
            out[0] if len(out) > 0 else "",
            out[1] if len(out) > 1 else "",
            out[2] if len(out) > 2 else "",
        )

    atv_desc = descricao
    blob_pi = f"{patologia or ''} {indicador or ''}".strip()
    tipo_a, grp_a, atv_a = _sem_consol_em_campos(
        patologia or indicador,
        indicador or patologia,
        descricao or (patologia + " " + (indicador or "")).strip(),
        consol=num_consol,
    )
    if atv_desc:
        _, _, atv_a = _sem_consol_em_campos("", "", atv_desc, consol=num_consol)
    elif _texto_e_bloco_legenda_atividade_artemig(atv_a) or _texto_e_bloco_legenda_atividade_artemig(
        blob_pi
    ):
        atv_a = _sem_consol_em_campos(
            "", "", descricao or "", consol=num_consol
        )[2]
    if _eh_texto_localizacao_resumo_artemig(atv_a):
        atv_a = (atv_desc or descricao or "").strip()
    if len((atv_a or "").strip()) < 18 or re.fullmatch(
        r"(?i)qid\s*:?\s*",
        (atv_a or "").strip(),
    ):
        atv_a = (atv_desc or descricao or "").strip()
    if _eh_texto_localizacao_resumo_artemig(atv_a) or re.fullmatch(
        r"(?i)qid\s*:?\s*",
        (atv_a or "").strip(),
    ):
        atv_a = ""
    tipo_a = _limpar_legendas_campo_artemig(tipo_a)
    grp_a = _limpar_legendas_campo_artemig(grp_a)
    atv_a = _limpar_legendas_campo_artemig(atv_a)
    if _eh_texto_localizacao_resumo_artemig(tipo_a):
        tipo_a = indicador or tipo_a
    if _eh_texto_localizacao_resumo_artemig(grp_a):
        grp_a = patologia or indicador or grp_a

    prazo_str, prazo_dias, emerg_p = _prazo_artemig(texto, data_con)
    prazo_str_limpa = _prazo_str_valido_artemig(prazo_str)
    if prazo_str_limpa:
        prazo_str = prazo_str_limpa
    else:
        prazo_str = ""
        if prazo_dias is not None and data_con:
            d0p = _parse_data(data_con)
            if d0p:
                prazo_str = (d0p.date() + timedelta(days=int(prazo_dias))).strftime("%d/%m/%Y")
    if prazo_dias is None:
        emerg_p = False

    obs_parts: list[str] = []
    if num_consol:
        obs_parts.append(num_consol)
    obs_m = re.search(r"Observ[aã][çc][aã]o\s*:?\s*([^\n]+)", texto, re.I)
    if obs_m:
        ox = re.sub(
            r"(?i)N[ºoO°]?\s*(?:da\s*)?CONSOL\s*:?\s*\d{0,12}\s*",
            "",
            obs_m.group(1).strip(),
        ).strip(" ,;—-|")
        if ox:
            obs_parts.append(_limpar_legendas_campo_artemig(ox[:300]))
    observacao = _limpar_legendas_campo_artemig(" | ".join(obs_parts))[:500]

    _blob_pav = " ".join(
        x for x in (patologia, indicador, tipo_a, grp_a, atv_a) if (x or "").strip()
    )
    _tipo_panela_art = _is_panela(_blob_pav)

    try:
        from nc_artemig.texto_pdf import colapsar_espacos_pdf as _cx
    except Exception:

        def _cx(s, multiline=False):
            return re.sub(r"\s+", " ", (s or "").strip())

    return NcItem(
        codigo=codigo,
        data_con=data_con,
        horario_fiscalizacao=hora,
        rodovia=rodovia,
        concessionaria="CONSOL",
        lote="50",
        km_ini=km_ini,
        km_fim=km_fim,
        km_ini_str=km_ini_str,
        km_fim_str=km_fim_str,
        sentido=_cx(sentido or "", multiline=False),
        tipo_atividade=_cx(tipo_a or "", multiline=False),
        grupo_atividade=_cx(grp_a or "", multiline=False),
        atividade=_cx(atv_a or "", multiline=False),
        observacao=_cx(observacao or "", multiline=False),
        prazo_str=prazo_str,
        prazo_dias=prazo_dias,
        emergencial=emerg_p,
        tipo_panela=_tipo_panela_art,
        empresa=_GRUPO_FISCALIZACAO_ARTEMIG,
        grupo=_GRUPO_EAF_ARTEMIG_ANALISE,
        nome_fiscal=_cx(nome_rt or "", multiline=False),
        tipo_artemig=tipo_artemig,
        sh_artemig=_cx(sh_artemig or "", multiline=False),
        num_consol=num_consol,
        patologia_artemig=_cx(patologia or "", multiline=False)[:2000],
        indicador_artemig=_cx(indicador or "", multiline=False)[:2000],
    )


def _stem_pdf_upload(nome_arquivo: str) -> str:
    s = (nome_arquivo or "").strip().replace("\\", "/")
    base = s.rsplit("/", 1)[-1]
    return Path(base).stem if base else ""


def _ordem_sufixo_nc_jpg_kcor(nome: str) -> int:
    """Chave de ordenação: sem sufixo = 0, sufixo _(N) ou (N) = N.
    Suporta ambos os formatos gerados por _nome_unico:
      nc (COD).jpg          → 0
      nc (COD) (1).jpg      → 1   (formato gerado por _nome_unico)
      nc (COD)_1.jpg        → 1   (formato legado)
    """
    m = re.match(
        r"(?i)nc \(([^)]+)\)(?:\s*\((\d+)\)|_(\d+))?\.(jpg|jpeg)$",
        (nome or "").strip(),
    )
    if not m:
        return 9999
    return int(m.group(2) or m.group(3) or 0)


def _norm_cod_de_nome_nc_jpg(cod_bruto: str) -> str:
    from nc_artemig.texto_pdf import identificador_pdf_sem_whitespace

    c = identificador_pdf_sem_whitespace((cod_bruto or "").strip())
    if not c:
        return ""
    if re.fullmatch(r"\d{6,14}", c):
        return c
    m = re.search(r"\b(\d{8,10})\b", c)
    return identificador_pdf_sem_whitespace(m.group(1)) if m else c


def _artemig_enriquecer_nomes_jpg_kcor_extracao(
    pdf_bytes: bytes, parcial: list[NcItem], nome_fonte_pdf: str
) -> None:
    """Preenche ``artemig_kcor_nomes_arquivos`` a partir da mesma extração que gera o ZIP (várias fotos por código)."""
    if not parcial or not FITZ_OK:
        return
    import tempfile
    from pathlib import Path as _P

    try:
        from nc_artesp import pdf_extractor as _pex
    except ImportError:
        return
    nome_pdf = (nome_fonte_pdf or "doc.pdf").strip()
    if not nome_pdf.lower().endswith(".pdf"):
        nome_pdf = f"{nome_pdf}.pdf"
    try:
        with tempfile.TemporaryDirectory(prefix="am_jpg_kcor_") as td:
            p = _P(td) / "upload.pdf"
            p.write_bytes(pdf_bytes)
            out = _P(td) / "out"
            out.mkdir(parents=True, exist_ok=True)
            salvos, _ = _pex.extrair_arquivo_pdf_para_pasta(
                p,
                out,
                dpi=None,
                nomear_por_indice_fiscalizacao=False,
                pasta_unica=True,
                raiz_unica_sem_subpastas=False,
                nome_pdf_original=nome_pdf,
            )
            por_cod: dict[str, list[str]] = {}
            for fstr in salvos:
                fp = _P(fstr)
                if not fp.is_file() or fp.suffix.lower() not in (".jpg", ".jpeg"):
                    continue
                name = fp.name
                if not name.lower().startswith("nc ("):
                    continue
                m = re.match(
                    r"(?i)nc \(([^)]+)\)(?:\s*\((\d+)\)|_(\d+))?\.(jpg|jpeg)$", name
                )
                if not m:
                    continue
                ck = _norm_cod_de_nome_nc_jpg(m.group(1))
                if not ck:
                    continue
                por_cod.setdefault(ck, []).append(name)
            for lst in por_cod.values():
                lst.sort(key=lambda n: (_ordem_sufixo_nc_jpg_kcor(n), n.lower()))
            from nc_artemig.exportar_kcor_planilha import _codigo_fiscalizacao_arquivos

            for nc in parcial:
                ck = _codigo_fiscalizacao_arquivos(nc)
                if not ck:
                    ck = _norm_cod_de_nome_nc_jpg((nc.codigo or "").strip())
                nomes = list(por_cod.get(ck, []))
                if nomes:
                    nc.artemig_kcor_nomes_arquivos = nomes
    except Exception as ex:
        logger.debug("artemig_kcor_nomes_arquivos: %s", ex)


def _artemig_paginas_foto_kcor(pdf_bytes: bytes) -> list[int]:
    """Índices 1-based das páginas com foto real (sufixo _N.jpg no Kcor, p.ex. _3,_4,_5).

    Ignora páginas cujas únicas imagens sejam o logo de cabeçalho
    (banner largo no topo: y0 < 60pt e largura ≥ 55 % da página).
    """
    if not FITZ_OK:
        return []
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    except Exception:
        return []

    def _tem_foto_real(page: "fitz.Page") -> bool:
        """True se a página tiver ao menos uma imagem que não seja logo de cabeçalho.
        Logo Artemig: ~57×51 pt = área < 10 000 pt², próximo ao topo (y0 < 100 pt).
        Fotos reais: ~244×150 pt = área > 36 000 pt².
        """
        for img in page.get_images():
            xref = img[0]
            for r in page.get_image_rects(xref):
                if r.width <= 50 or r.height <= 50:
                    continue
                if r.y0 < 100 and r.width * r.height < 10_000:
                    continue   # ícone/logo de cabeçalho — ignorar
                return True
        return False

    try:
        com_foto: list[int] = []
        for i in range(len(doc)):
            if _tem_foto_real(doc[i]):
                com_foto.append(i + 1)
        jpg_pages = [p for p in com_foto if p >= 3]
        if not jpg_pages and len(com_foto) >= 3:
            jpg_pages = com_foto[2:]
        elif not jpg_pages and len(com_foto) >= 2:
            jpg_pages = com_foto[1:]
        elif not jpg_pages and com_foto:
            jpg_pages = com_foto
        return jpg_pages
    finally:
        doc.close()


def parse_pdf_artemig(pdf_bytes: bytes) -> list[NcItem]:
    """NCs do PDF Artemig (MG): tenta 1 NC por página; senão, texto inteiro."""
    if not FITZ_OK:
        texto = _extrair_texto_pdf(pdf_bytes)
        nc = _parse_artemig_texto(texto or "")
        return [nc] if nc else []
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    ncs: list[NcItem] = []
    try:
        for page in doc:
            t = (page.get_text() or "").strip()
            if len(t) < 40:
                continue
            nc = _parse_artemig_texto(t)
            if nc:
                ncs.append(nc)
        if not ncs:
            full = "\n\f\n".join((doc[i].get_text() or "") for i in range(len(doc)))
            nc = _parse_artemig_texto(full)
            if nc:
                ncs = [nc]
    finally:
        doc.close()
    vistos: set[str] = set()
    out: list[NcItem] = []
    for nc in ncs:
        ch = (nc.codigo or "") + "|" + (nc.num_consol or "") + "|" + str(nc.km_ini)
        if ch in vistos:
            continue
        vistos.add(ch)
        out.append(nc)
    return out


# ANÁLISE DE SEQUÊNCIA DE KM

def _trecho_do_grupo_para_nc(nc: NcItem, mapa_eaf: list[dict]) -> tuple[float, float] | None:
    """
    Retorna (km_ini, km_fim) do trecho do MAPA_EAF que contém a NC (rodovia + km_ini).
    Só considera o grupo da própria NC. Retorna None se grupo não tiver trecho ou NC fora do trecho.
    """
    from nc_artesp.utils.helpers import normalizar_rodovia_para_busca
    g = nc.grupo or 0
    if not g or not mapa_eaf:
        return None
    rod_nc = normalizar_rodovia_para_busca(nc.rodovia)
    km = nc.km_ini
    for entry in mapa_eaf:
        if entry.get("grupo") != g:
            continue
        for trecho in entry.get("trechos", []):
            rod_t = normalizar_rodovia_para_busca(trecho.get("rodovia", ""))
            if not rod_t:
                continue
            if rod_t == rod_nc or rod_t in rod_nc or rod_nc in rod_t:
                ki = trecho.get("km_ini", 0.0)
                kf = trecho.get("km_fim", 9999.0)
                if ki <= km <= kf:
                    return (ki, kf)
    return None


def analisar_gaps(ncs: list[NcItem], limiar_km: float = LIMIAR_GAP_KM,
                  mapa_eaf: list[dict] | None = None) -> list[GapAlerta]:
    """
    Detecta saltos de KM acima do limiar apenas entre NCs do mesmo grupo, mesma data
    de constatação e mesmo trecho (rodovia + intervalo km) daquele grupo. Não mistura
    grupos, datas nem trechos. NCs de Meio Ambiente são excluídas.
    Retorna lista de GapAlerta ordenada por grupo → rodovia → sentido → km.
    """
    if not ncs:
        return []
    ncs = [nc for nc in ncs if not getattr(nc, "origem_ma", False)]
    mapa = mapa_eaf if mapa_eaf is not None else _MAPA_EAF_PADRAO

    com_trecho: list[tuple[NcItem, tuple[float, float]]] = []
    for nc in ncs:
        t = _trecho_do_grupo_para_nc(nc, mapa)
        if t is not None:
            com_trecho.append((nc, t))

    buckets: dict[tuple, list[NcItem]] = {}
    for nc, (trecho_ini, trecho_fim) in com_trecho:
        data_con = (nc.data_con or "").strip()
        chave = (nc.grupo or 0, data_con, nc.rodovia, nc.sentido, trecho_ini, trecho_fim)
        buckets.setdefault(chave, []).append(nc)

    alertas: list[GapAlerta] = []
    for (grupo_num, data_con, rodovia, sentido, _ti, _tf), bucket in sorted(buckets.items()):
        if len(bucket) < 2:
            continue
        bucket_ord = sorted(bucket, key=lambda n: n.km_ini)
        empresa = bucket_ord[0].empresa if bucket_ord else ""
        for i in range(len(bucket_ord) - 1):
            a = bucket_ord[i]
            b = bucket_ord[i + 1]
            gap = b.km_ini - a.km_fim
            if gap >= limiar_km:
                alertas.append(GapAlerta(
                    grupo=grupo_num,
                    empresa=empresa,
                    rodovia=rodovia,
                    sentido=sentido,
                    km_antes=a.km_fim,
                    km_depois=b.km_ini,
                    gap_km=round(gap, 3),
                    nc_antes=a.codigo,
                    nc_depois=b.codigo,
                ))

    return alertas


def analisar_sequencia_codigos(ncs: list[NcItem]) -> list[CodigoGapAlerta]:
    """
    [Somente regime ARTESP] Saltos na numeração do Código Fiscalização por grupo EAF.
    Na ARTESP o código é sequencial; lacunas sugerem NC gerada e não entregue (em geral
    retida só buraco/panela). No Artemig (lote 50) essa lógica não se aplica — todas as
    constatações do PDF entram no relatório; não chamar para Artemig (retornar []).
    """
    if not ncs:
        return []
    ncs = [nc for nc in ncs if not getattr(nc, "origem_ma", False)]

    def _para_int(codigo: str) -> Optional[int]:
        digits = re.sub(r'\D', '', str(codigo))
        return int(digits) if digits else None

    buckets: dict[int, list[NcItem]] = {}
    for nc in ncs:
        g = nc.grupo or 0
        buckets.setdefault(g, []).append(nc)

    alertas: list[CodigoGapAlerta] = []
    for grupo_num, bucket in sorted(buckets.items()):
        com_num = [(n, _para_int(n.codigo)) for n in bucket]
        com_num = [(nc, num) for nc, num in com_num if num is not None]
        if len(com_num) < 2:
            continue
        com_num.sort(key=lambda x: x[1])
        empresa = com_num[0][0].empresa or ""

        for i in range(len(com_num) - 1):
            nc_a, num_a = com_num[i]
            nc_b, num_b = com_num[i + 1]
            diff = num_b - num_a
            if diff <= 1:
                continue
            faltantes = [str(num_a + j) for j in range(1, min(diff, 11))]
            alertas.append(CodigoGapAlerta(
                grupo=grupo_num,
                empresa=empresa,
                codigo_antes=nc_a.codigo,
                codigo_depois=nc_b.codigo,
                n_faltantes=diff - 1,
                codigos_faltantes=faltantes,
            ))

    return alertas


def _rotulo_tipo_resumo_artemig(nc: NcItem) -> str:
    """Artemig: indicador/patologia; evita ' / ' com metade vazia (ex.: '/ Panelas e / Buracos')."""

    def _lim(s: str) -> str:
        s = (s or "").strip()
        s = re.sub(r"^\s*/+\s*|\s*/+\s*$", "", s).strip()
        return s

    ta, ga, at = _lim(nc.tipo_atividade), _lim(nc.grupo_atividade), _lim(nc.atividade)
    for bad in ("/", "-", ".", "e", "E"):
        if ta == bad:
            ta = ""
        if ga == bad:
            ga = ""
    if ta and ga:
        if ta.lower() in ga.lower() or ga.lower() in ta.lower():
            lab = ga if len(ga) >= len(ta) else ta
        else:
            lab = f"{ta} / {ga}"
    elif ta:
        lab = ta
    elif ga:
        lab = ga
    elif at:
        lab = at
    else:
        tid = (getattr(nc, "tipo_artemig", None) or "").strip()
        sh = (getattr(nc, "sh_artemig", None) or "").strip()
        lab = " | ".join(x for x in (tid, sh) if x) or "(sem classificacao no PDF)"
    lab = re.sub(r"\s*/\s*/\s*", " / ", lab)
    lab = re.sub(r"^[\s/\-]+", "", lab).strip()
    lab = re.sub(r"\s*/\s*$", "", lab).strip()
    return lab[:200]


def resumo_estatistico(ncs: list[NcItem]) -> dict:
    """Retorna dicionário com métricas resumidas para o cabeçalho do relatório."""
    if not ncs:
        return {}
    eh_artemig_50 = any((getattr(n, "lote", None) or "").strip() == "50" for n in ncs)
    tipos: dict[str, int] = {}
    for nc in ncs:
        if getattr(nc, "origem_ma", False):
            k = (_atividade_exibicao_relatorio(nc) or (nc.atividade or "").strip() or "Meio Ambiente")[:200]
        elif eh_artemig_50:
            k = _rotulo_tipo_resumo_artemig(nc)
        else:
            k = (_atividade_exibicao_relatorio(nc) or "(sem atividade)")[:200]
        tipos[k] = tipos.get(k, 0) + 1
    emergenciais = [nc for nc in ncs if nc.emergencial]
    panelas_poss = [nc for nc in ncs if _is_panela_artemig_nc(nc)]
    rodovias = sorted(set(nc.rodovia for nc in ncs if nc.rodovia))
    if any((getattr(n, "lote", None) or "").strip() == "50" for n in ncs):
        def _sig_rod(r: str) -> str:
            return re.sub(r"[\s-]", "", (r or "").upper())

        padrao = ["MG-050", "BR-265", "BR-491"]
        seen = {_sig_rod(p) for p in padrao}
        out = list(padrao)
        for r in sorted(set(nc.rodovia for nc in ncs if nc.rodovia)):
            if _sig_rod(r) not in seen:
                out.append(r.strip())
                seen.add(_sig_rod(r))
        rodovias = out
    data_con     = ncs[0].data_con if ncs else ""
    grupos: dict[int, dict] = {}
    for nc in ncs:
        if getattr(nc, "origem_ma", False):
            g = -1  # Meio Ambiente (exibido separado no resumo)
            emp = "Meio Ambiente"
        else:
            g = nc.grupo or 0
            emp = nc.empresa
        if g not in grupos:
            grupos[g] = {"grupo": g, "empresa": emp, "total": 0, "emergenciais": 0}
        grupos[g]["total"] += 1
        if nc.emergencial:
            grupos[g]["emergenciais"] += 1
    return {
        "total":          len(ncs),
        "tipos":          tipos,
        "n_tipos":        len(tipos),
        "emergenciais":   emergenciais,
        "panelas":        panelas_poss,
        "rodovias":       rodovias,
        "data_con":       data_con,
        "lote":           ncs[0].lote if ncs else "",
        "grupos":         dict(sorted(grupos.items())),
    }


# GERAÇÃO DO RELATÓRIO PDF (ReportLab)

# Paleta ARTESP
COR_HEADER    = colors.HexColor("#1e3a5f")   # azul escuro
COR_ALERTA    = colors.HexColor("#c0392b")   # vermelho NC
COR_EMERG     = colors.HexColor("#e74c3c")   # vermelho emergencial
COR_OK        = colors.HexColor("#27ae60")   # verde ok
COR_AVISO     = colors.HexColor("#e67e22")   # laranja aviso
COR_LINHAR    = colors.HexColor("#2c3e50")   # cabeçalho tabela
COR_LINHA_ALT = colors.HexColor("#f0f4f8")   # fundo linha par
COR_EMERG_BG  = colors.HexColor("#fdecea")   # fundo linha emergencial


def _estilos():
    ss = getSampleStyleSheet()
    extra = {
        "titulo": ParagraphStyle("titulo",
            fontName="Helvetica-Bold", fontSize=16,
            textColor=COR_HEADER, spaceAfter=4, alignment=TA_CENTER),
        "subtitulo": ParagraphStyle("subtitulo",
            fontName="Helvetica-Bold", fontSize=12,
            textColor=COR_HEADER, spaceAfter=2, alignment=TA_CENTER),
        "secao": ParagraphStyle("secao",
            fontName="Helvetica-Bold", fontSize=11,
            textColor=colors.white, spaceAfter=0, leading=16),
        "corpo": ParagraphStyle("corpo",
            fontName="Helvetica", fontSize=9,
            textColor=colors.HexColor("#2c3e50"), spaceAfter=2, leading=13),
        "alerta": ParagraphStyle("alerta",
            fontName="Helvetica-Bold", fontSize=9,
            textColor=COR_ALERTA, spaceAfter=2, leading=13),
        "emerg": ParagraphStyle("emerg",
            fontName="Helvetica-Bold", fontSize=9,
            textColor=COR_EMERG, spaceAfter=2, leading=13),
        "tabcab": ParagraphStyle("tabcab",
            fontName="Helvetica-Bold", fontSize=8,
            textColor=colors.white, alignment=TA_CENTER),
        "tabcel": ParagraphStyle("tabcel",
            fontName="Helvetica", fontSize=8,
            textColor=colors.HexColor("#2c3e50"), alignment=TA_LEFT),
        "tabcel_qtd": ParagraphStyle("tabcel_qtd",
            fontName="Helvetica-Bold", fontSize=8,
            textColor=colors.HexColor("#2c3e50"), alignment=TA_CENTER),
        "tabcel_art_tipo": ParagraphStyle("tabcel_art_tipo",
            fontName="Helvetica", fontSize=7, leading=8,
            textColor=colors.HexColor("#2c3e50"), alignment=TA_LEFT),
        "tabcel_emerg": ParagraphStyle("tabcel_emerg",
            fontName="Helvetica-Bold", fontSize=8,
            textColor=COR_EMERG, alignment=TA_LEFT),
        "rodape": ParagraphStyle("rodape",
            fontName="Helvetica", fontSize=7,
            textColor=colors.grey, alignment=TA_CENTER),
        "celula": ParagraphStyle("celula",
            fontName="Helvetica", fontSize=9,
            textColor=colors.HexColor("#2c3e50"), spaceAfter=0, leading=11,
            alignment=TA_LEFT),
    }
    return extra


def _km_fmt(v: float) -> str:
    km  = int(v)
    met = round((v - km) * 1000)
    return f"{km}+{met:03d}"


def _safe_latin1(s: str) -> str:
    """Latin-1 para Helvetica: troca travessão/emoji; mantém ç/ã se já couberem em Latin-1."""
    if not s:
        return s
    for a, b in (
        ("\u2014", "-"),
        ("\u2013", "-"),
        ("\u2212", "-"),
        ("\u2264", "<="),
        ("\u2265", ">="),
        ("\u2022", "*"),
    ):
        s = s.replace(a, b)
    s = re.sub(
        r"[\U0001F300-\U0001FAFF\U00002700-\U000027BF\U00002600-\U000027BF]",
        "",
        s,
    )
    try:
        return s.encode("latin-1").decode("latin-1")
    except UnicodeEncodeError:
        nfd = unicodedata.normalize("NFD", s)
        sem_comb = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
        return sem_comb.encode("latin-1", "replace").decode("latin-1")


def _banner(texto: str, cor: colors.Color, estilos: dict):
    """Faixa colorida com texto branco — usada como cabeçalho de seção."""
    return Table(
        [[Paragraph(_safe_latin1(texto), estilos["secao"])]],
        colWidths=[170 * mm],
        style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), cor),
            ("LEFTPADDING",  (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING",   (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
        ]),
    )


def _pdf_eh_artemig_lote50(ncs_ref: list[NcItem]) -> bool:
    return bool(ncs_ref and any((getattr(n, "lote", None) or "").strip() == "50" for n in ncs_ref))


def _pdf_regime_e_artesp(ncs_ref: list[NcItem]) -> bool:
    """True = ARTESP (salto de código, retidas). False = Artemig lote 50."""
    return not _pdf_eh_artemig_lote50(ncs_ref)


def _tabela_ncs(ncs_tipo: list[NcItem], estilos: dict) -> Table:
    """Tabela com dados das NCs; Artemig inclui col. Tipo (QID) e SH."""
    es_artemig = _pdf_eh_artemig_lote50(ncs_tipo)
    col_obs = "Nº CONSOL" if es_artemig else "Obs"
    if es_artemig:
        cabecalho = [
            "Cód.", "Tipo", "SH", "Grp", "KM Ini", "KM Fim", "Sentido",
            "Data", "Prazo", col_obs,
        ]
        colunas_w = [17 * mm, 30 * mm, 13 * mm, 7 * mm, 13 * mm, 13 * mm,
                     22 * mm, 13 * mm, 11 * mm, 22 * mm]
    else:
        cabecalho = ["Cód. Fiscal.", "Grp", "KM Inicial", "KM Final", "Sentido",
                     "Data Const.", "Prazo", col_obs]
        colunas_w = [26 * mm, 10 * mm, 20 * mm, 20 * mm, 26 * mm, 20 * mm, 20 * mm, 28 * mm]

    def _esc(s: str) -> str:
        return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    linhas = [[Paragraph(_safe_latin1(c), estilos["tabcab"]) for c in cabecalho]]

    for nc in sorted(ncs_tipo, key=lambda n: n.km_ini):
        est = estilos["tabcel_emerg"] if nc.emergencial else estilos["tabcel"]
        prazo_txt = nc.prazo_str + (" !" if nc.emergencial else "")
        grupo_txt = str(nc.grupo) if nc.grupo else ("-" if not es_artemig else str(_GRUPO_EAF_ARTEMIG_ANALISE))
        if es_artemig:
            tid = _esc((getattr(nc, "tipo_artemig", None) or "-").strip() or "-")
            sh = _esc((getattr(nc, "sh_artemig", None) or "-").strip() or "-")
            est_tipo = estilos.get("tabcel_art_tipo", estilos["tabcel"])
            linhas.append([
                Paragraph(_safe_latin1(nc.codigo or ""), est),
                Paragraph(_safe_latin1(tid[:40]), est_tipo),
                Paragraph(_safe_latin1(sh[:14]), est),
                Paragraph(grupo_txt, est),
                Paragraph(_km_fmt(nc.km_ini) if nc.km_ini else nc.km_ini_str, est),
                Paragraph(_km_fmt(nc.km_fim) if nc.km_fim else nc.km_fim_str, est),
                Paragraph(_safe_latin1(nc.sentido or ""), est),
                Paragraph(_safe_latin1(nc.data_con or ""), est),
                Paragraph(_safe_latin1(prazo_txt), est),
                Paragraph(_safe_latin1((nc.observacao or "-")[:400]), est),
            ])
        else:
            linhas.append([
                Paragraph(_safe_latin1(nc.codigo or ""), est),
                Paragraph(grupo_txt, est),
                Paragraph(_km_fmt(nc.km_ini) if nc.km_ini else nc.km_ini_str, est),
                Paragraph(_km_fmt(nc.km_fim) if nc.km_fim else nc.km_fim_str, est),
                Paragraph(_safe_latin1(nc.sentido or ""), est),
                Paragraph(_safe_latin1(nc.data_con or ""), est),
                Paragraph(_safe_latin1(prazo_txt), est),
                Paragraph(_safe_latin1((nc.observacao or "-")[:800]), est),
            ])

    ts = TableStyle([
        # Cabeçalho
        ("BACKGROUND",   (0, 0), (-1, 0), COR_LINHAR),
        ("GRID",         (0, 0), (-1, -1), 0.3, colors.HexColor("#bdc3c7")),
        ("TOPPADDING",   (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
        ("LEFTPADDING",  (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
    ])
    # Fundo alternado e destaque emergencial
    for i, nc in enumerate(ncs_tipo, start=1):
        row = i
        if nc.emergencial:
            ts.add("BACKGROUND", (0, row), (-1, row), COR_EMERG_BG)
        elif i % 2 == 0:
            ts.add("BACKGROUND", (0, row), (-1, row), COR_LINHA_ALT)

    return Table(linhas, colWidths=colunas_w, style=ts, repeatRows=1)


def _km_para_partes(km: float) -> tuple[int, int]:
    """Converte km decimal em (km_int, metros). Ex.: 103.5 → (103, 500)."""
    if not km and km != 0.0:
        return 0, 0
    k = int(km)
    m = round((km - k) * 1000)
    if m >= 1000:
        m = 0
    return k, m


def _lote_num_do_pdf(nc: "NcItem") -> str | None:
    """Extrai o número do lote do item (ex.: '13', '21', '50'). Só retorna se for lote conhecido (evita confundir com código da fiscalização)."""
    try:
        from nc_artesp.config import _LOTE_CONCESSIONARIA
        lotes_conhecidos = set(_LOTE_CONCESSIONARIA.keys()) if _LOTE_CONCESSIONARIA else set()
    except Exception:
        lotes_conhecidos = set()
    for s in [(nc.lote or "").strip(), (nc.concessionaria or "").strip()]:
        if not s:
            continue
        if re.match(r"^\d+$", s) and s in lotes_conhecidos:
            return s
        for num in re.findall(r"\d+", s):
            if num in lotes_conhecidos:
                return num
    return None


def _concessionaria_por_lote(lote_ou_concessionaria: str) -> str:
    """Retorna 'Lote N Nome da Concessionária' a partir do número do lote (ex.: '13' → 'Lote 13 Rodovias das Colinas')."""
    try:
        from nc_artesp.config import _LOTE_CONCESSIONARIA
        s = (lote_ou_concessionaria or "").strip()
        n = re.search(r"\d+", s)
        if n:
            num = n.group(0)
            nome = _LOTE_CONCESSIONARIA.get(num)
            if nome:
                return f"Lote {num} {nome}"
    except Exception:
        pass
    return (lote_ou_concessionaria or "").strip()


def _caminho_template_relatorio_xlsx(lote_selecionado: str | None = None) -> Path:
    """Lote 50: template Artemig (A/B/V); demais: ARTESP_TEMPLATE_RELATORIO."""
    if _norm_lote_numero(lote_selecionado) == "50":
        try:
            from nc_artemig.config import ASSETS_DIR, TEMPLATE_RELATORIO_ANALISE_PDF

            candidatos: list[Path] = []
            for base in (Path(TEMPLATE_RELATORIO_ANALISE_PDF),):
                if base not in candidatos:
                    candidatos.append(base)
            tpl_root = ASSETS_DIR / "Template"
            for sub in ("templates",):
                d = tpl_root / sub
                if d.is_dir():
                    q = d / "Template_EAF_artemig.xlsx"
                    if q not in candidatos:
                        candidatos.append(q)
            raiz = tpl_root / "Template_EAF_artemig.xlsx"
            if raiz not in candidatos:
                candidatos.append(raiz)
            for p in candidatos:
                if p.is_file():
                    return p.resolve()
        except Exception:
            pass
    from nc_artesp.config import TEMPLATE_RELATORIO_XLSX
    return Path(TEMPLATE_RELATORIO_XLSX).resolve()


def rotulo_e_slug_lote_para_saida(lote: str | None) -> tuple[str, str]:
    """(rótulo para cabeçalhos, pasta segura para ZIP)."""
    try:
        from nc_artesp.config import LOTES_MENU_ANALISE, _LOTE_CONCESSIONARIA
    except Exception:
        LOTES_MENU_ANALISE, _LOTE_CONCESSIONARIA = [], {}
    num_m = re.search(r"\d+", (lote or "").strip() or "13")
    num = num_m.group(0) if num_m else "13"
    rotulo = ""
    for k, lab in LOTES_MENU_ANALISE:
        if k == num:
            rotulo = lab
            break
    if not rotulo:
        nome = _LOTE_CONCESSIONARIA.get(num, "")
        rotulo = f"Lote {num}" + (f" — {nome}" if nome else "")
    slug_map = {
        "13": "Lote13_Rodovias_Colinas",
        "21": "Lote21_Rodovias_Tiete",
        "26": "Lote26_SP_Serra",
        "50": "Nascentes_ARTEMIG_kriaKcor",
    }
    slug = slug_map.get(num, f"Lote{num}_Analise")
    return rotulo, slug


# Colunas do template de saída (relatório XLSX). km_ini: col 8+9; km_fim: 10+11.
COLUNAS_TEMPLATE_SAIDA: dict[str, int] = {
    "codigo": 3, "data_con": 4, "horario_fiscalizacao": 5, "rodovia": 6, "concessionaria": 7,
    "km_ini_str": 8, "km_fim_str": 10, "sentido": 12, "tipo_atividade": 15, "grupo_atividade": 16,
    "atividade": 17, "prazo_str": 19, "empresa": 20, "nome_fiscal": 21,
}


def _detectar_colunas_saida_template(ws, cabecalho_fim: int = 4) -> dict[str, int]:
    """Detecta colunas no cabeçalho: Responsável Técnico; col. U (obs. Kcor), não a Q do relatório EAF Artemig."""
    col_map = dict(COLUNAS_TEMPLATE_SAIDA)
    for r in range(1, min(cabecalho_fim + 1, ws.max_row + 1)):
        for c in range(1, min(ws.max_column + 1, 30)):
            v = ws.cell(row=r, column=c).value
            s = (str(v or "")).strip().lower().replace("é", "e").replace("á", "a")
            if v and ("observa" in s or "observac" in s) and "gestor" not in s:
                if "fiscaliza" in s and c <= 18:
                    continue
                col_map["obs_linha_kcor"] = c
            if v and ("responsavel" in s or "responsável" in s or ("respons" in s and "tecnico" in s)):
                col_map["nome_fiscal"] = c
    return col_map


def _data_sem_hora_celula(s) -> str:
    """Normaliza valor de célula de data do relatório XLSX para dd/mm/aaaa."""
    from datetime import date, datetime

    if s is None:
        return ""
    if isinstance(s, datetime):
        return s.strftime("%d/%m/%Y")
    try:
        if isinstance(s, date):
            return s.strftime("%d/%m/%Y")
    except Exception:
        pass
    if isinstance(s, (int, float)) and not isinstance(s, bool):
        try:
            from openpyxl.utils.datetime import from_excel

            return from_excel(float(s)).strftime("%d/%m/%Y")
        except Exception:
            pass
    t = str(s).strip()
    if not t:
        return ""
    if " " in t:
        t = t.split()[0].strip()
    return t[:10] if len(t) >= 10 else t


def _coluna_data_reparo_relatorio(val: str) -> str:
    v = (val or "").strip()
    if not v:
        return ""
    if re.match(r"(?i)^em\s+at[eé]?\s+", v):
        return v
    return _data_sem_hora_celula(v)


def _aplicar_borda_fina_linha_relatorio(ws, row: int, col_fim: int) -> None:
    from openpyxl.styles import Border, Side

    thin = Side(style="thin", color="000000")
    b = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, col_fim + 1):
        ws.cell(row=row, column=col).border = b


def _copiar_estilo_linha_relatorio_template(ws, row_orig: int, row_dst: int, col_fim: int) -> None:
    """Duplica estilo da linha modelo e uniformiza bordas na faixa."""
    from copy import copy

    col_ate = max(1, int(col_fim))
    for col in range(1, col_ate + 1):
        src = ws.cell(row=row_orig, column=col)
        dst = ws.cell(row=row_dst, column=col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)
    _aplicar_borda_fina_linha_relatorio(ws, row_dst, col_fim)


def _ajustar_linhas_dados_relatorio_xlsx(ws, lin_ref: int, n_ncs: int, col_fim: int) -> None:
    """Ajusta quantidade de linhas de dados sem apagar o bloco inteiro (preserva formatação do template)."""
    col_fim = max(int(col_fim or 22), 1)
    lin_ref = int(lin_ref)
    n_ncs = max(0, int(n_ncs or 0))
    if n_ncs <= 0:
        if ws.max_row >= lin_ref:
            ws.delete_rows(lin_ref, ws.max_row - lin_ref + 1)
        return
    if ws.max_row < lin_ref:
        ws.insert_rows(ws.max_row + 1, lin_ref - ws.max_row)
        _aplicar_borda_fina_linha_relatorio(ws, lin_ref, col_fim)
    ult = lin_ref + n_ncs - 1
    mr = ws.max_row
    if mr > ult:
        ws.delete_rows(ult + 1, mr - ult)
        mr = ws.max_row
    if mr < ult:
        need = ult - mr
        ws.insert_rows(mr + 1, need)
        for r in range(mr + 1, ult + 1):
            _copiar_estilo_linha_relatorio_template(ws, lin_ref, r, col_fim)


def gerar_relatorio_xlsx(
    ncs: list[NcItem],
    lote_selecionado: str | None = None,
    rotulo_lote_analise: str = "",
) -> bytes:
    """Preenche o template XLSX a partir de NcItem. Coluna fica vazia só quando a informação não está nos documentos lidos."""
    if not OPENPYXL_OK:
        raise ImportError("openpyxl não instalado: pip install openpyxl")
    template_path = _caminho_template_relatorio_xlsx(lote_selecionado)
    try:
        from nc_artemig.sanear_pipeline import relatorio_deve_tratar_artemig, sanear_ncs_lote50_consol

        if relatorio_deve_tratar_artemig(lote_selecionado, ncs):
            sanear_ncs_lote50_consol(ncs, forcar_todas=True)
    except Exception:
        pass
    lo50 = _norm_lote_numero(lote_selecionado) == "50" or any(
        _norm_lote_numero(getattr(n, "lote", None) or "") == "50" for n in ncs
    )
    _colapsar_pdf = None
    if lo50:
        try:
            from nc_artemig.texto_pdf import colapsar_espacos_pdf as _colapsar_pdf
        except ImportError:
            _colapsar_pdf = None
    CABECALHO_FIM = 4
    PRIMEIRA_LINHA_DADOS = 5
    insere_linha_lote = bool((rotulo_lote_analise or "").strip())
    wb = None
    ws = None

    if template_path.is_file():
        logger.info("Relatório XLSX: usando template %s", template_path)
        try:
            if template_path.suffix.lower() == ".xls":
                import xlrd
                raw = template_path.read_bytes()
                rb = xlrd.open_workbook(file_contents=raw)
                sh = rb.sheet_by_index(0)
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Relatório"
                row_off = 1
                if insere_linha_lote:
                    ws.cell(row=1, column=1, value=f"Lote em análise: {rotulo_lote_analise.strip()}")
                    row_off = 2
                    CABECALHO_FIM = 5
                    PRIMEIRA_LINHA_DADOS = 6
                for r in range(min(4, sh.nrows)):
                    for c in range(min(sh.ncols, 30)):
                        v = sh.cell_value(r, c)
                        ws.cell(row=r + row_off, column=c + 1, value=v)
            else:
                import shutil
                import tempfile
                fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
                try:
                    os.close(fd)
                    shutil.copy2(str(template_path), tmp_path)
                    wb = openpyxl.load_workbook(tmp_path)
                    ws = wb.active
                    if insere_linha_lote:
                        # Linha 1 nos templates XLSX costuma estar vazia; não usar insert_rows(1),
                        # pois desloca merges (A2:V2, B3:B4…) e corrompe o cabeçalho no Excel.
                        ult = min(max(ws.max_column or 21, 22), 25)
                        for rng in list(ws.merged_cells.ranges):
                            if rng.min_row <= 1 <= rng.max_row:
                                ws.unmerge_cells(str(rng))
                        if ult >= 2:
                            ws.merge_cells(
                                start_row=1, start_column=1, end_row=1, end_column=ult
                            )
                        ws.cell(
                            row=1,
                            column=1,
                            value=f"Lote em análise: {rotulo_lote_analise.strip()}",
                        )
                    col_fim_tpl = min(max(ws.max_column or 22, 22), 30)
                    _ajustar_linhas_dados_relatorio_xlsx(
                        ws, PRIMEIRA_LINHA_DADOS, len(ncs), col_fim_tpl
                    )
                finally:
                    try:
                        Path(tmp_path).unlink(missing_ok=True)
                    except Exception:
                        pass
        except Exception as e:
            logger.exception("Template relatório %s: %s", template_path, e)
            raise FileNotFoundError(
                f"Template do relatório não carregou: {template_path}\nErro: {e}"
            ) from e
    else:
        logger.warning(
            "Template relatório não encontrado (path=%s); modo fallback com cabeçalho padrão.",
            template_path,
        )

    if wb is None or ws is None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório"
        r0 = 1
        if insere_linha_lote:
            ws.cell(row=1, column=1, value=f"Lote em análise: {rotulo_lote_analise.strip()}")
            r0 = 2
            CABECALHO_FIM = 5
            PRIMEIRA_LINHA_DADOS = 6
        for col, val in enumerate([
            "", "", "Cód. Fiscalização", "Data Fiscalização", "Horário da Fiscalização",
            "Rodovia", "Concessionária Lote", "Trecho", "", "", "", "Sentido",
            "Data Retorno", "Status Retorno", "Tipo de Atividade", "Grupo de Atividade",
            "Atividade", "Data Envio", "Data Reparo", "EAF", "Responsável Técnico",
        ], start=1):
            if col <= 21 and val:
                ws.cell(row=r0 + 1, column=col, value=val)
        for col, val in enumerate(["", ""] + [""] * 5 + [" Km Inicial", "m", "Km Final", "m", ""] + [""] * 8 + ["Responsável Técnico"], start=1):
            if col <= 21 and val:
                ws.cell(row=r0 + 2, column=col, value=val)

    col_map = COLUNAS_TEMPLATE_SAIDA.copy()
    if wb is not None and ws is not None and hasattr(ws, "max_column"):
        try:
            col_map = _detectar_colunas_saida_template(ws, CABECALHO_FIM)
        except Exception:
            pass

    for row_idx, nc in enumerate(ncs, start=PRIMEIRA_LINHA_DADOS):
        km_ini_k, km_ini_m = _km_para_partes(nc.km_ini)
        km_fim_k, km_fim_m = _km_para_partes(nc.km_fim)
        if lo50:
            try:
                from nc_artemig.config import RELATORIO_FISCALIZACAO_COL_G_CONCESSIONARIA

                conc_val = RELATORIO_FISCALIZACAO_COL_G_CONCESSIONARIA
            except ImportError:
                conc_val = "Nascentes das Gerais"
        else:
            conc_val = (
                _concessionaria_por_lote(lote_selecionado)
                if lote_selecionado
                else _concessionaria_por_lote(nc.concessionaria or nc.lote)
                or (nc.concessionaria or nc.lote or "").strip()
            )
        # Preencher quando a informação existe nos documentos; vazio só quando não presente
        ws.cell(row=row_idx, column=col_map["codigo"], value=(nc.codigo or "").strip())
        ws.cell(row=row_idx, column=col_map["data_con"], value=_data_sem_hora_celula(nc.data_con or ""))
        ws.cell(row=row_idx, column=col_map["horario_fiscalizacao"], value=(nc.horario_fiscalizacao or "").strip())
        ws.cell(row=row_idx, column=col_map["rodovia"], value=(nc.rodovia or "").strip())
        ws.cell(row=row_idx, column=col_map["concessionaria"], value=conc_val)
        ws.cell(row=row_idx, column=col_map["km_ini_str"], value=km_ini_k)
        ws.cell(row=row_idx, column=col_map["km_ini_str"] + 1, value=km_ini_m)
        ws.cell(row=row_idx, column=col_map["km_fim_str"], value=km_fim_k)
        ws.cell(row=row_idx, column=col_map["km_fim_str"] + 1, value=km_fim_m)
        sent_out = (nc.sentido or "").strip()
        if lo50:
            try:
                from nc_artemig.sentido_kcor import sentido_artemig_para_kcor
                sent_out = sentido_artemig_para_kcor(nc.rodovia or "", nc.sentido or "")
            except Exception:
                pass
        ws.cell(row=row_idx, column=col_map["sentido"], value=sent_out)
        ws.cell(row=row_idx, column=13, value="")
        ws.cell(row=row_idx, column=14, value="")
        ws.cell(row=row_idx, column=col_map["tipo_atividade"], value=(nc.tipo_atividade or "").strip())
        ws.cell(row=row_idx, column=col_map["grupo_atividade"], value=(nc.grupo_atividade or "").strip())
        ws.cell(row=row_idx, column=col_map["atividade"], value=_atividade_exibicao_relatorio(nc))
        ws.cell(row=row_idx, column=18, value=_data_sem_hora_celula(nc.data_con or ""))
        pz_out = (nc.prazo_str or "").strip()
        if lo50:
            pz_out = _prazo_str_valido_artemig(pz_out)
        ws.cell(row=row_idx, column=col_map["prazo_str"], value=_coluna_data_reparo_relatorio(pz_out))
        ws.cell(row=row_idx, column=col_map["empresa"], value=(nc.empresa or "").strip())
        ws.cell(row=row_idx, column=col_map["nome_fiscal"], value=(nc.nome_fiscal or "").strip())
        if lo50:
            # Artemig: col. A do template é sempre QID (PDF pode trazer «Fiscalização» ou texto com NBSP).
            ws.cell(row=row_idx, column=1, value="QID")
            shv = (getattr(nc, "sh_artemig", None) or "").strip()
            if _colapsar_pdf:
                shv = _colapsar_pdf(shv, multiline=False)
            ws.cell(row=row_idx, column=2, value=shv)
            consol_v = (getattr(nc, "num_consol", None) or "").strip()
            if not consol_v:
                cod = (nc.codigo or "").strip()
                if cod.upper().startswith("CE") and len(cod) > 2:
                    consol_v = cod[2:].strip()
            if _colapsar_pdf:
                consol_v = _colapsar_pdf(consol_v, multiline=False)
            ws.cell(row=row_idx, column=22, value=(consol_v or "")[:120])
            try:
                from nc_artemig.exportar_kcor_planilha import (
                    _escapar_inicio_formula_excel,
                    _excel_valor_texto_ou_none,
                    _montar_v_w_kcor,
                )

                _, w_arq = _montar_v_w_kcor(nc)
                w_xlsx = _excel_valor_texto_ou_none(
                    _escapar_inicio_formula_excel(w_arq) if w_arq else ""
                )
                ws.cell(row=row_idx, column=23, value=w_xlsx)
            except Exception:
                pass
            oc = col_map.get("obs_linha_kcor")
            if oc:
                try:
                    from nc_artemig.exportar_kcor_planilha import _texto_observacoes_nas01

                    u_txt = (_texto_observacoes_nas01(nc) or "")[:32700]
                    if _colapsar_pdf:
                        u_txt = _colapsar_pdf(u_txt, multiline=True)
                    ws.cell(row=row_idx, column=oc, value=u_txt)
                except Exception:
                    pass

    if ncs and wb is not None and ws is not None and template_path.is_file() and template_path.suffix.lower() == ".xlsx":
        ult = PRIMEIRA_LINHA_DADOS + len(ncs) - 1
        col_fim_rel = max(23 if lo50 else 22, min(ws.max_column or (23 if lo50 else 22), 25))
        if lo50:
            from openpyxl.styles import Alignment

            al_left_top = Alignment(horizontal="left", vertical="top", wrap_text=False)
            al_left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
            al_center_top = Alignment(horizontal="center", vertical="top", wrap_text=False)
            oc_art = col_map.get("obs_linha_kcor")
            wrap_cols = {
                col_map.get("tipo_atividade"),
                col_map.get("grupo_atividade"),
                col_map.get("atividade"),
                col_map.get("rodovia"),
                col_map.get("sentido"),
                col_map.get("concessionaria"),
                col_map.get("empresa"),
                col_map.get("nome_fiscal"),
                22,
                23,
            }
            if oc_art:
                wrap_cols.add(oc_art)
            wrap_cols.discard(None)
            km_cols = {
                col_map["km_ini_str"],
                col_map["km_ini_str"] + 1,
                col_map["km_fim_str"],
                col_map["km_fim_str"] + 1,
            }
            date_center_cols = {
                col_map["data_con"],
                18,
                col_map["prazo_str"],
                col_map.get("horario_fiscalizacao"),
            }
            date_center_cols.discard(None)
            centro_cod_sh = {1, 2, col_map["codigo"]}
        for r in range(PRIMEIRA_LINHA_DADOS, ult + 1):
            _aplicar_borda_fina_linha_relatorio(ws, r, col_fim_rel)
            for c in {col_map["data_con"], 18, 19}:
                cell = ws.cell(row=r, column=c)
                if cell.value is not None and str(cell.value).strip():
                    cell.number_format = "@"
            if lo50:
                for c in range(1, col_fim_rel + 1):
                    ce = ws.cell(row=r, column=c)
                    if c in km_cols:
                        ce.alignment = al_center_top
                    elif c in date_center_cols:
                        ce.alignment = al_center_top
                    elif c in centro_cod_sh:
                        ce.alignment = al_center_top
                    elif c in wrap_cols:
                        ce.alignment = al_left_wrap
                    else:
                        ce.alignment = al_left_top
                w23 = ws.cell(row=r, column=23)
                if w23.value is not None and str(w23.value).strip():
                    w23.number_format = "@"
                ws.row_dimensions[r].height = None
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _tabela_indicadores_rodovia(ncs: list[NcItem]) -> list[list[str]]:
    conserv = [n for n in ncs if not getattr(n, "origem_ma", False)]
    by_rod: dict[str, list[NcItem]] = {}
    for n in conserv:
        r = (n.rodovia or "").strip() or "(sem rodovia)"
        by_rod.setdefault(r, []).append(n)
    out: list[list[str]] = []
    for r in sorted(by_rod.keys()):
        lst = by_rod[r]
        kms = [n.km_ini for n in lst if n.km_ini is not None]
        k0, k1 = (min(kms), max(kms)) if kms else (None, None)
        n_pan = sum(1 for n in lst if _is_panela_artemig_nc(n))
        out.append([
            r,
            str(len(lst)),
            _km_fmt(k0) if k0 is not None else "-",
            _km_fmt(k1) if k1 is not None else "-",
            str(n_pan),
        ])
    return out


def gerar_relatorio_pdf(ncs: list[NcItem],
                        alertas_km: list[GapAlerta],
                        alertas_codigo: list[CodigoGapAlerta],
                        limiar_km: float = LIMIAR_GAP_KM,
                        mapa_eaf: list | None = None,
                        rotulo_lote_analise: str = "",
                        forcar_alertas_pdf: bool = False) -> bytes:
    """PDF de analise: mesmo layout ARTESP/Artemig; alertas de codigo so ARTESP."""
    if not REPORTLAB_OK:
        raise ImportError("reportlab não instalado: pip install reportlab")
    mapa_uso = (mapa_eaf or _MAPA_EAF_PADRAO)

    def _rodape_pagina(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        w, _ = doc.pagesize
        canvas.drawCentredString(w / 2, 8 * mm, "Desenvolvedor Ozeias Engler")
        canvas.restoreState()

    est  = _estilos()
    buf  = io.BytesIO()
    doc  = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm,
        topMargin=12*mm,  bottomMargin=12*mm,
        onFirstPage=_rodape_pagina,
        onLaterPages=_rodape_pagina,
    )
    story = []
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    try:
        data_relatorio = datetime.strptime(agora, "%d/%m/%Y %H:%M").date()
    except ValueError:
        data_relatorio = date.today()

    res    = resumo_estatistico(ncs)
    emerg  = res.get("emergenciais", [])
    pdf_artemig_50 = _pdf_eh_artemig_lote50(ncs)
    regime_artesp = _pdf_regime_e_artesp(ncs)
    rodovs = ", ".join(res.get("rodovias", [])) or "-"
    data_c = res.get("data_con", "")
    n_panela = sum(
        1 for n in ncs if not getattr(n, "origem_ma", False) and _is_panela_artemig_nc(n)
    )
    n_ma = sum(1 for n in ncs if getattr(n, "origem_ma", False))
    data_con_dt = _parse_data(data_c)
    data_con_date = data_con_dt.date() if data_con_dt else None
    # Só no dia da constatação, exceto teste_local (PDF antigo no dev).
    relatorio_hoje = forcar_alertas_pdf or (
        data_con_date is not None and data_relatorio == data_con_date
    )
    lote   = res.get("lote", "")

    story.append(Spacer(1, 4*mm))
    if pdf_artemig_50:
        story.append(Paragraph("Artemig (MG) - Relatorio de analise de NCs", est["titulo"]))
        story.append(Paragraph("Notificacao / CONSOL - Conservacao", est["subtitulo"]))
    else:
        story.append(Paragraph("ARTESP - Relatorio de analise de NCs", est["titulo"]))
        story.append(Paragraph("Conservacao de Rotina", est["subtitulo"]))
    story.append(Spacer(1, 3*mm))
    story.append(HRFlowable(width="100%", thickness=2, color=COR_HEADER))
    story.append(Spacer(1, 3*mm))

    # Metadados — Paragraph em cada célula para quebra de texto e evitar overflow
    def _cel(s: str, bold: bool = False) -> Paragraph:
        style = ParagraphStyle("cel", parent=est["celula"], fontName="Helvetica-Bold" if bold else "Helvetica")
        t = _safe_latin1(s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        return Paragraph(t, style)
    meta = []
    if (rotulo_lote_analise or "").strip():
        meta.append(
            [_cel("Lote em análise:", True), _cel(rotulo_lote_analise.strip(), False), _cel("", False), _cel("", False)]
        )
    meta.extend([
        [_cel("Rodovia(s):", True), _cel(rodovs), _cel("Lote (PDF):", True), _cel(lote)],
        [_cel("Data constatacao:", True), _cel(data_c), _cel("Emitido em:", True), _cel(agora)],
        [_cel("Total de NCs:", True), _cel(str(res.get("total", 0))),
         _cel("Tipos atividade:", True), _cel(str(res.get("n_tipos", 0)))],
        [_cel("NCs emerg. (24h):", True), _cel(str(len(emerg))),
         _cel("NCs buraco/panela:", True), _cel(str(n_panela))],
        [_cel("Meio ambiente (se houver):", True), _cel(str(n_ma)),
         _cel("Alertas gap KM:", True), _cel(str(len(alertas_km)))],
    ])
    if regime_artesp:
        tot_cod = sum(a.n_faltantes for a in alertas_codigo)
        meta.append([
            _cel("Saltos codigo (poss. retidos):", True), _cel(str(tot_cod)),
            _cel("Grupos c/ salto codigo:", True), _cel(str(len(alertas_codigo))),
        ])
    else:
        n_qid = len({(getattr(n, "tipo_artemig", None) or "").strip() for n in ncs if not getattr(n, "origem_ma", False)})
        n_sh = len({(getattr(n, "sh_artemig", None) or "").strip() for n in ncs if not getattr(n, "origem_ma", False)})
        meta.append([
            _cel("Regime relatorio:", True), _cel("Artemig (integral)"),
            _cel("Tipos QID / SH dist.:", True), _cel(f"{n_qid} / {n_sh}"),
        ])
    meta_t = Table(meta, colWidths=[48*mm, 48*mm, 38*mm, 38*mm],
        style=TableStyle([
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#bdc3c7")),
            ("BACKGROUND", (0, 0), (-1, -1), COR_LINHA_ALT),
        ]))
    story.append(meta_t)
    story.append(Spacer(1, 6*mm))

    # Resumo por Grupo EAF (equipes por trecho km — Contatos EAFs)
    def _trecho_resumo(grupo_num: int) -> str:
        """Texto rodovia e trecho coberto para o grupo (ex.: SP 75 km 15 > 50)."""
        if pdf_artemig_50 and grupo_num == _GRUPO_EAF_ARTEMIG_ANALISE:
            return "MG-050 · BR-265 · BR-491"
        if grupo_num is None or grupo_num == -1 or grupo_num == 0 or grupo_num == 999:
            return "-"
        for entry in mapa_uso:
            if entry.get("grupo") == grupo_num:
                partes = []
                for t in entry.get("trechos", []):
                    rod = t.get("rodovia", "").strip()
                    ki = t.get("km_ini", 0.0)
                    kf = t.get("km_fim", 0.0)
                    if rod:
                        km_ini_int = int(ki) if ki == int(ki) else ki
                        km_fim_int = int(kf) if kf == int(kf) else kf
                        partes.append(f"{rod} km {km_ini_int} > {km_fim_int}")
                return " | ".join(partes) if partes else "-"
        return "-"

    story.append(_banner("RESUMO POR GRUPO DE FISCALIZAÇÃO", COR_HEADER, est))
    story.append(Spacer(1, 2*mm))
    grupos_res = res.get("grupos", {})
    grp_dados = [[Paragraph("Grupo", est["tabcab"]), Paragraph("Rodovia / Trecho", est["tabcab"]),
                  Paragraph("Empresa", est["tabcab"]), Paragraph("NCs", est["tabcab"]), Paragraph("Emergenciais", est["tabcab"])]]
    if grupos_res:
        for g_num, g_info in sorted(grupos_res.items()):
            if g_num == -1:
                label = "Meio Ambiente"
            elif g_num:
                label = str(g_num)
            else:
                label = "Não ident."
            # Escapar só < e > para Paragraph; não escapar & para não quebrar ">" (evita &amp;gt; no PDF)
            trecho_txt = _trecho_resumo(g_num).replace("<", "&lt;").replace(">", "&gt;")
            emp = (g_info.get("empresa") or "-").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            grp_dados.append([
                Paragraph(label, est["celula"]),
                Paragraph(trecho_txt, est["celula"]),
                Paragraph(emp, est["celula"]),
                Paragraph(str(g_info.get("total", 0)), est["celula"]),
                Paragraph(str(g_info.get("emergenciais", 0)) or "0", est["celula"]),
            ])
    else:
        grp_dados.append([Paragraph("-", est["celula"]), Paragraph("-", est["celula"]),
                          Paragraph("Nenhum grupo identificado", est["celula"]),
                          Paragraph("0", est["celula"]), Paragraph("0", est["celula"])])
    grp_t = Table(grp_dados, colWidths=[28*mm, 52*mm, 52*mm, 16*mm, 26*mm],
        style=TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), COR_LINHAR),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#bdc3c7")),
            *[("BACKGROUND", (0, i), (-1, i), COR_LINHA_ALT)
              for i in range(2, len(grp_dados), 2)],
        ]))
    story.append(grp_t)
    story.append(Spacer(1, 4*mm))

    # Resumo de tipos
    story.append(_banner("RESUMO POR TIPO DE NC", COR_HEADER, est))
    story.append(Spacer(1, 2*mm))
    tipos_sorted = sorted(res.get("tipos", {}).items(), key=lambda x: -x[1])
    col_tipo = "Tipo (indicador/patologia)" if pdf_artemig_50 else "Atividade"
    tipo_dados = [[Paragraph(_safe_latin1(col_tipo), est["tabcab"]), Paragraph("Qtd", est["tabcab"])]]
    for tipo, qtd in tipos_sorted:
        t_esc = _safe_latin1((tipo or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))
        tipo_dados.append([Paragraph(t_esc, est["celula"]), Paragraph(str(qtd), est["celula"])])
    tipo_t = Table(tipo_dados, colWidths=[142*mm, 22*mm],
        style=TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), COR_LINHAR),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#bdc3c7")),
            *[("BACKGROUND", (0, i), (-1, i), COR_LINHA_ALT)
              for i in range(2, len(tipo_dados), 2)],
        ]))
    story.append(tipo_t)

    story.append(Spacer(1, 4 * mm))
    story.append(_banner("INDICADORES POR RODOVIA (conservacao)", COR_HEADER, est))
    story.append(Spacer(1, 2 * mm))
    ind_rod = [["Rodovia", "NCs", "KM min", "KM max", "Buraco/panela"]]
    ind_rod.extend(_tabela_indicadores_rodovia(ncs))
    ind_linhas = [[Paragraph(_safe_latin1(c), est["tabcab"]) for c in ind_rod[0]]]
    for row in ind_rod[1:]:
        ind_linhas.append([Paragraph(_safe_latin1(x), est["celula"]) for x in row])
    story.append(Table(
        ind_linhas,
        colWidths=[38 * mm, 18 * mm, 28 * mm, 28 * mm, 32 * mm],
        style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), COR_LINHAR),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#bdc3c7")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            *[("BACKGROUND", (0, i), (-1, i), COR_LINHA_ALT) for i in range(2, len(ind_linhas), 2)],
        ]),
    ))

    if pdf_artemig_50:
        conserv = [n for n in ncs if not getattr(n, "origem_ma", False)]
        c_tid = Counter()
        c_sh = Counter()
        for n in conserv:
            t = (getattr(n, "tipo_artemig", None) or "").strip()
            c_tid[t if t else "(sem tipo no PDF)"] += 1
            s = (getattr(n, "sh_artemig", None) or "").strip()
            c_sh[s if s else "(sem SH no PDF)"] += 1
        story.append(Spacer(1, 4 * mm))
        story.append(_banner("DISTRIBUICAO QID E SUBTRECHO (SH)", COR_HEADER, est))
        story.append(Spacer(1, 2 * mm))
        est_sub = ParagraphStyle(
            "subdist", parent=est["corpo"], fontName="Helvetica-Bold", fontSize=9,
        )
        story.append(Paragraph(_safe_latin1("Por tipo (col. A template / PDF):"), est_sub))
        qid_data = [("Tipo (QID)", "Qtd")] + [(k, str(v)) for k, v in c_tid.most_common()]
        qid_rows = []
        for i, (a, b) in enumerate(qid_data):
            st = est["tabcab"] if i == 0 else est["tabcel"]
            st_q = est["tabcab"] if i == 0 else est["tabcel_qtd"]
            qid_rows.append([
                Paragraph(_safe_latin1(str(a)), st),
                Paragraph(_safe_latin1(str(b)), st_q),
            ])
        story.append(Table(
            qid_rows,
            colWidths=[118 * mm, 22 * mm],
            style=TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), COR_LINHAR),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("ALIGN", (1, 0), (1, -1), "CENTER"),
                *[("BACKGROUND", (0, i), (-1, i), COR_LINHA_ALT) for i in range(2, len(qid_rows), 2)],
            ]),
        ))
        story.append(Spacer(1, 3 * mm))
        story.append(Paragraph(_safe_latin1("Por subtrecho SH:"), est_sub))
        sh_data = [("SH", "Qtd")] + [(k, str(v)) for k, v in c_sh.most_common()]
        sh_rows = []
        for i, (a, b) in enumerate(sh_data):
            st = est["tabcab"] if i == 0 else est["tabcel"]
            st_q = est["tabcab"] if i == 0 else est["tabcel_qtd"]
            sh_rows.append([
                Paragraph(_safe_latin1(str(a)), st),
                Paragraph(_safe_latin1(str(b)), st_q),
            ])
        story.append(Table(
            sh_rows,
            colWidths=[118 * mm, 22 * mm],
            style=TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), COR_LINHAR),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("ALIGN", (1, 0), (1, -1), "CENTER"),
                *[("BACKGROUND", (0, i), (-1, i), COR_LINHA_ALT) for i in range(2, len(sh_rows), 2)],
            ]),
        ))

    if relatorio_hoje:
        if regime_artesp and alertas_codigo:
            story.append(Spacer(1, 6*mm))
            story.append(_banner(
                "* APONTAMENTOS NAO ENTREGUES - SALTO NA NUMERACAO DO CODIGO",
                COR_EMERG, est
            ))
            story.append(Spacer(1, 2*mm))
            story.append(Paragraph(
                "O Código Fiscalização é atribuído <b>sequencialmente</b> pelo sistema ARTESP. "
                "Um salto na sequência indica apontamentos que foram gerados mas "
                "<b>NÃO foram entregues</b> à concessionária. "
                "Somente <b>buracos/panelas na pista</b> sao retidos (prazo 24 h) - "
                "todos os demais tipos sao sempre entregues.",
                est["corpo"]
            ))
            story.append(Spacer(1, 2*mm))

            total_ocultos = sum(a.n_faltantes for a in alertas_codigo)
            story.append(Paragraph(
                f"<font color='#e74c3c'><b>Total de apontamentos não entregues: "
                f"{total_ocultos}</b></font>",
                est["emerg"]
            ))
            story.append(Spacer(1, 3*mm))

            for i, ca in enumerate(alertas_codigo, 1):
                label_grp = f"Grupo {ca.grupo} - {ca.empresa}" if ca.grupo else "Grupo não identificado"
                faltantes_str = ", ".join(ca.codigos_faltantes)
                if ca.n_faltantes > 10:
                    faltantes_str += f" ... (+{ca.n_faltantes - 10} ocultos)"
                story.append(KeepTogether([
                    Paragraph(
                        f"<b>Ocorrência {i}</b> | {label_grp}",
                        est["emerg"]
                    ),
                    Paragraph(
                        f"Ultimo entregue: <b>{ca.codigo_antes}</b> -> "
                        f"Proximo entregue: <b>{ca.codigo_depois}</b> "
                        f"<font color='#e74c3c'><b>({ca.n_faltantes} não entregue(s))</b></font>",
                        est["corpo"]
                    ),
                    Paragraph(
                        f"Códigos ausentes: {faltantes_str}",
                        est["corpo"]
                    ),
                    Paragraph(
                        "! Esses apontamentos sao potencialmente BURACOS NA PISTA "
                        "(único tipo com prazo de 24 h que pode ser retido).",
                        est["emerg"]
                    ),
                    Spacer(1, 4*mm),
                ]))

        if alertas_km:
            story.append(Spacer(1, 4*mm))
            story.append(_banner(
                f"! ALERTAS DE SALTO DE KM - TRECHO SEM APONTAMENTO (limiar: {limiar_km:.1f} km)",
                COR_AVISO, est
            ))
            story.append(Spacer(1, 2*mm))
            txt_gap_km = (
                "Os trechos abaixo apresentam intervalo superior ao limiar sem apontamentos. "
                "Verifique se ha buracos/panelas nao registrados (ARTESP: possivel lacuna de entrega)."
                if regime_artesp else
                "Os trechos abaixo apresentam intervalo superior ao limiar sem apontamentos; "
                "conferir cobertura da fiscalizacao no trecho (Artemig: todas as NCs do PDF constam no relatorio)."
            )
            story.append(Paragraph(_safe_latin1(txt_gap_km), est["corpo"]))
            story.append(Spacer(1, 2*mm))

            for i, ga in enumerate(alertas_km, 1):
                label_grp = f"Grupo {ga.grupo} - {ga.empresa}" if ga.grupo else "Grupo não identificado"
                story.append(KeepTogether([
                    Paragraph(
                        f"<b>Alerta {i}</b> | {label_grp} | {ga.rodovia} - Sentido: {ga.sentido}",
                        est["alerta"]
                    ),
                    Paragraph(
                        f"Trecho sem NC: km <b>{_km_fmt(ga.km_antes)}</b> -> "
                        f"km <b>{_km_fmt(ga.km_depois)}</b> "
                        f"<font color='#c0392b'><b>(gap: {ga.gap_km:.3f} km)</b></font>",
                        est["corpo"]
                    ),
                    Paragraph(
                        f"NC antes: {ga.nc_antes} | NC depois: {ga.nc_depois}",
                        est["corpo"]
                    ),
                    Spacer(1, 3*mm),
                ]))

        if emerg:
            story.append(Spacer(1, 4*mm))
            story.append(_banner("NCs EMERGENCIAIS - PRAZO <= 24 h", COR_EMERG, est))
            story.append(Spacer(1, 2*mm))
            story.append(Paragraph(
                "As NCs abaixo têm prazo igual ou anterior a 1 dia após a constatação. "
                "Data e hora do prazo: até 23:59 do dia indicado.",
                est["corpo"]
            ))
            story.append(Spacer(1, 2*mm))
            for nc in sorted(emerg, key=lambda n: (n.rodovia, n.km_ini)):
                prazo_data_hora = f"{nc.prazo_str} 23:59" if nc.prazo_str else "-"
                story.append(KeepTogether([
                    Paragraph(
                        f"<b>Cód. {nc.codigo}</b> | {nc.rodovia} | "
                        f"km {_km_fmt(nc.km_ini)} -> {_km_fmt(nc.km_fim)} | "
                        f"Sentido: {nc.sentido}",
                        est["emerg"]
                    ),
                    Paragraph(
                        f"Atividade: <b>{_atividade_exibicao_relatorio(nc)}</b>",
                        est["corpo"]
                    ),
                    Paragraph(
                        f"Prazo: <font color='#e74c3c'><b>{prazo_data_hora}</b></font> "
                        f"(24 h após constatação - {'MESMO DIA' if nc.prazo_dias == 0 else str(nc.prazo_dias) + ' dia(s)'})",
                        est["corpo"]
                    ),
                    *(
                        [Paragraph(
                            ("Nº da CONSOL: " if pdf_artemig_50 else "Obs: ")
                            + (nc.observacao or ""),
                            est["corpo"],
                        )]
                        if nc.observacao else []
                    ),
                    Spacer(1, 3*mm),
                ]))

    if pdf_artemig_50:
        story.append(Spacer(1, 5 * mm))
    else:
        story.append(PageBreak())
    story.append(_banner(
        "NCs POR GRUPO / TIPO / KM (dados originais)",
        COR_HEADER, est
    ))
    story.append(Spacer(1, 2*mm))
    nota_det = (
        "Dados conforme PDF. Artemig: colunas Tipo e SH do layout de notificacao."
        if pdf_artemig_50 else
        "Dados conforme PDF ARTESP. Agrupamento por EAF, atividade e KM."
    )
    story.append(Paragraph(_safe_latin1(nota_det), est["corpo"]))
    story.append(Spacer(1, 4*mm))

    # Agrupar: Conservação por (grupo_num, empresa); Meio Ambiente em bloco à parte
    # Lista (label, grupo_num ou None para MA, por_tipo)
    blocos: list[tuple[str, int | None, dict[str, list[NcItem]]]] = []
    conservacao: dict[tuple, dict[str, list[NcItem]]] = {}
    meio_ambiente: dict[str, list[NcItem]] = {}
    for nc in ncs:
        if getattr(nc, "origem_ma", False):
            meio_ambiente.setdefault(
                _atividade_exibicao_relatorio(nc) or (nc.atividade or "").strip() or "SEM ATIVIDADE",
                [],
            ).append(nc)
        else:
            chave = (nc.grupo or 999, nc.empresa or "Sem EAF identificada")
            conservacao.setdefault(chave, {})
            tipo_k = (
                _rotulo_tipo_resumo_artemig(nc)
                if pdf_artemig_50
                else (_atividade_exibicao_relatorio(nc) or "SEM ATIVIDADE")
            )
            conservacao[chave].setdefault(tipo_k, []).append(nc)
    for (grupo_num, empresa), por_tipo in sorted(conservacao.items()):
        label = f"GRUPO {grupo_num} - {empresa}" if grupo_num != 999 else "GRUPO NAO IDENTIFICADO"
        blocos.append((label, grupo_num, por_tipo))
    if meio_ambiente:
        blocos.append(("MEIO AMBIENTE", None, meio_ambiente))

    for label_grupo, grupo_num, por_tipo in blocos:
        # Banner de grupo
        story.append(Table(
            [[Paragraph(label_grupo, ParagraphStyle(
                "ghdr", fontName="Helvetica-Bold", fontSize=13,
                textColor=colors.white, leading=18,
            ))]],
            colWidths=[170 * mm],
            style=TableStyle([
                ("BACKGROUND",    (0, 0), (-1, -1), COR_HEADER),
                ("TOPPADDING",    (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("LEFTPADDING",   (0, 0), (-1, -1), 10),
            ])
        ))
        story.append(Spacer(1, 3*mm))

        # Trechos fiscalizados (só Conservação; Meio Ambiente não usa MAPA_EAF)
        if grupo_num is not None and grupo_num != 999:
            mostrou = False
            for entry in mapa_uso:
                if entry.get("grupo") == grupo_num:
                    partes = [
                        f"{t['rodovia']} km {t['km_ini']:.3f}-{t['km_fim']:.3f}"
                        for t in entry.get("trechos", [])
                    ]
                    story.append(Paragraph(
                        "<b>Trechos fiscalizados:</b> " + " | ".join(partes),
                        est["corpo"]
                    ))
                    mostrou = True
                    break
            if not mostrou and pdf_artemig_50 and grupo_num == _GRUPO_EAF_ARTEMIG_ANALISE:
                story.append(Paragraph(
                    "<b>Trechos fiscalizados:</b> MG-050, BR-265, BR-491 (CONSOL).",
                    est["corpo"]
                ))
        story.append(Spacer(1, 2*mm))

        # NCs por tipo dentro do grupo
        for idx_t, (tipo, grupo_ncs) in enumerate(sorted(por_tipo.items()), 1):
            tem_emerg = any(n.emergencial for n in grupo_ncs)
            cor_tipo  = COR_EMERG if tem_emerg else COR_ALERTA
            label_emg = "  [EMERG]" if tem_emerg else ""
            tit_ban = re.sub(r"^[\s/\-]+", "", (tipo or "")).strip().upper()
            story.append(_banner(
                f"{idx_t}. {tit_ban} - {len(grupo_ncs)} NC(s){label_emg}",
                cor_tipo, est,
            ))
            story.append(Spacer(1, 2 * mm))
            story.append(_tabela_ncs(grupo_ncs, est))
            story.append(Spacer(1, 5 * mm))
        story.append(Spacer(1, 4*mm))

    # Rodapé final
    story.append(Spacer(1, 4*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
    rod = (
        f"Gerado em {agora}. Fonte: PDF original. Regime: {'ARTESP' if regime_artesp else 'Artemig (MG)'}. "
        "Desenvolvedor Ozeias Engler."
    )
    story.append(Paragraph(_safe_latin1(rod), est["rodape"]))

    doc.build(story)
    return buf.getvalue()


# PONTO DE ENTRADA PARA O ROUTER

def _montar_resumo_serializavel(ncs: list[NcItem],
                                alertas_km: list[GapAlerta],
                                alertas_codigo: list[CodigoGapAlerta]) -> dict:
    """Constrói resumo JSON-serializável."""
    res = resumo_estatistico(ncs)
    res["alertas_gap"] = [
        {
            "grupo":     a.grupo,
            "empresa":   a.empresa,
            "rodovia":   a.rodovia,
            "sentido":   a.sentido,
            "km_antes":  _km_fmt(a.km_antes),
            "km_depois": _km_fmt(a.km_depois),
            "gap_km":    a.gap_km,
        }
        for a in alertas_km
    ]
    res["alertas_codigo"] = [
        {
            "grupo":            a.grupo,
            "empresa":          a.empresa,
            "codigo_antes":     a.codigo_antes,
            "codigo_depois":    a.codigo_depois,
            "n_faltantes":      a.n_faltantes,
            "codigos_faltantes": a.codigos_faltantes,
        }
        for a in alertas_codigo
    ]
    res["total_ocultos"] = sum(a.n_faltantes for a in alertas_codigo)
    res["emergenciais_lista"] = [
        {
            "codigo":     nc.codigo,
            "rodovia":    nc.rodovia,
            "km":         _km_fmt(nc.km_ini),
            "atividade":  _atividade_exibicao_relatorio(nc),
            "prazo":      nc.prazo_str,
            "prazo_dias": nc.prazo_dias,
        }
        for nc in res.get("emergenciais", [])
    ]
    res.pop("emergenciais", None)
    res.pop("panelas", None)
    return res


def _excel_valor_para_horario(val) -> str:
    """Converte valor da célula Excel (serial, datetime.time/datetime ou string) para HH:MM ou HH:MM:SS."""
    if val is None:
        return ""
    try:
        from datetime import time, datetime
        if isinstance(val, time):
            return val.strftime("%H:%M:%S") if val.second else val.strftime("%H:%M")
        if isinstance(val, datetime):
            return val.strftime("%H:%M:%S") if val.second else val.strftime("%H:%M")
    except Exception:
        pass
    if isinstance(val, (int, float)):
        v = float(val)
        if v >= 1:
            v = v % 1
        if 0 <= v < 1:
            h = int(v * 24) % 24
            m = int((v * 24 * 60) % 60)
            s = int((v * 24 * 3600) % 60)
            if s:
                return f"{h:02d}:{m:02d}:{s:02d}"
            return f"{h:02d}:{m:02d}"
        if 0 <= v < 24 and v == int(v):
            return f"{int(v):02d}:00"
    s = str(val).strip()
    if re.match(r"^\d{1,2}:\d{2}", s):
        return s
    return s


# Template: lista de campos exigidos; leitura (Excel/PDF) mapeia colunas/blocos para esses campos.
CAMPOS_TEMPLATE_LIST: tuple[str, ...] = (
    "codigo", "data_con", "horario_fiscalizacao", "rodovia", "concessionaria",
    "km_ini_str", "km_fim_str", "sentido", "tipo_atividade", "grupo_atividade",
    "atividade", "prazo_str", "empresa", "nome_fiscal",
)


def _excel_complemento_pode_mesclar_campo(lote_val: str, campo: str) -> bool:
    """Lote 50 Artemig: planilha complementar não substitui texto de NC usado no Kcor (fonte = PDF)."""
    if _norm_lote_numero(lote_val) == "50" and campo in (
        "tipo_atividade",
        "grupo_atividade",
        "atividade",
        "prazo_str",
        "km_ini_str",
        "km_fim_str",
    ):
        return False
    return True


# (campo, termos no cabeçalho para detectar coluna, termos que desqualificam). Primeiro match ganha.
MAPEAMENTO_EXCEL_PARA_TEMPLATE: list[tuple[str, list[str], list[str]]] = [
    ("codigo", ["cod", "fiscal", "codigo"], []),
    ("codigo", ["cod", "fiscal"], []),
    ("data_con", ["data", "constata"], []),
    ("data_con", ["data", "fiscaliz"], []),
    ("horario_fiscalizacao", ["horario", "hora"], []),
    ("rodovia", ["rodovia", "sp "], []),
    ("rodovia", ["rodovia"], []),
    ("concessionaria", ["concessionaria", "concessionária", "lote"], []),
    ("concessionaria", ["concessionaria", "lote"], []),
    ("km_ini_str", ["km", "inicial", "km+m"], []),
    ("km_ini_str", ["inicial"], []),
    ("km_fim_str", ["km", "final", "km+m"], []),
    ("km_fim_str", ["final"], []),
    ("sentido", ["sentido"], []),
    ("tipo_atividade", ["tipo", "atividade"], []),
    ("grupo_atividade", ["grupo", "atividade"], []),
    ("atividade", ["atividade"], ["tipo", "grupo"]),
    ("atividade", ["evento"], []),
    ("prazo_str", ["prazo", "data limite", "data programada", "termino"], []),
    ("prazo_str", ["data", "termino"], []),
    ("empresa", ["empresa", "fiscalizadora"], []),
    ("nome_fiscal", ["responsavel", "responsável", "tecnico"], []),
    ("nome_fiscal", ["responsavel", "tecnico"], []),
]


def _cel_val(ws_or_sh, is_xlrd: bool, row: int, col: int):
    if is_xlrd:
        if row - 1 >= getattr(ws_or_sh, "nrows", 0) or col - 1 >= getattr(ws_or_sh, "ncols", 0):
            return None
        return ws_or_sh.cell_value(row - 1, col - 1)
    return ws_or_sh.cell(row=row, column=col).value


def _detectar_colunas_template_excel(ws_or_sh, is_xlrd: bool, max_col: int = 30) -> dict[str, Optional[int]]:
    """Cabeçalho (linhas 1–4) → número da coluna por campo do template. None se não encontrado."""
    out: dict[str, Optional[int]] = {attr: None for attr in CAMPOS_TEMPLATE_LIST}
    for row_idx in (2, 3, 1, 4):
        for c in range(1, max_col + 1):
            val = _cel_val(ws_or_sh, is_xlrd, row_idx, c)
            s = (str(val or "")).strip().lower()
            for char, repl in (("é", "e"), ("á", "a"), ("ó", "o"), ("í", "i"), ("ú", "u"), ("ã", "a"), ("õ", "o"), ("ç", "c")):
                s = s.replace(char, repl)
            if not s:
                continue
            for attr, termos, exclude in MAPEAMENTO_EXCEL_PARA_TEMPLATE:
                if out[attr] is not None:
                    continue
                if exclude and any(ex in s for ex in exclude):
                    continue
                if all(t in s for t in termos) or (len(termos) == 1 and termos[0] in s):
                    out[attr] = c
                    break
    return out


def _detectar_colunas_cabecalho(ws_or_sh, is_xlrd: bool, max_col: int = 25) -> tuple[Optional[int], Optional[int], Optional[int], Optional[int]]:
    """
    Detecta índices das colunas Código (C), Horário (E), Tipo (O), Grupo (P) pelo cabeçalho (linhas 1–4).
    Retorna None para colunas não encontradas — planilhas com menos colunas preenchem só o que existir no template.
    """
    cols = _detectar_colunas_template_excel(ws_or_sh, is_xlrd, max_col=max_col)
    return (
        cols.get("codigo"),
        cols.get("horario_fiscalizacao"),
        cols.get("tipo_atividade"),
        cols.get("grupo_atividade"),
    )


def _colunas_disponiveis_no_arquivo(col_map: dict[str, Optional[int]], ncols: int) -> set[str]:
    """Campos do template com coluna identificada neste arquivo."""
    return {
        attr for attr in CAMPOS_TEMPLATE_LIST
        if col_map.get(attr) is not None and col_map[attr] <= ncols
    }


def _ler_excel_complementar(excel_bytes: bytes) -> list[dict]:
    """Lê linhas do Excel; colunas identificadas pelo cabeçalho (MAPEAMENTO_EXCEL_PARA_TEMPLATE). Chaves = CAMPOS_TEMPLATE_LIST."""
    PRIMEIRA_LINHA_DADOS = 5
    out: list[dict] = []
    if not excel_bytes or len(excel_bytes) < 100:
        return out
    try:
        buf = io.BytesIO(excel_bytes)
        is_xls = len(excel_bytes) >= 8 and excel_bytes[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'
        if is_xls:
            import xlrd
            rb = xlrd.open_workbook(file_contents=excel_bytes)
            sh = rb.sheet_by_index(0)
            ncols = getattr(sh, "ncols", 0)
            col_map = _detectar_colunas_template_excel(sh, is_xlrd=True, max_col=min(30, ncols))
            disponiveis = _colunas_disponiveis_no_arquivo(col_map, ncols)
            col_c = col_map.get("codigo") or 1
            if col_c - 1 >= ncols:
                return out
            for r in range(PRIMEIRA_LINHA_DADOS - 1, sh.nrows):
                cod = sh.cell_value(r, col_c - 1)
                if cod is None or (isinstance(cod, str) and not str(cod).strip()):
                    continue
                row_dict: dict[str, str] = {attr: "" for attr in CAMPOS_TEMPLATE_LIST}
                row_dict["codigo"] = str(cod).strip()
                for attr in disponiveis:
                    if attr == "codigo":
                        continue
                    col = col_map[attr]
                    if col is None or col - 1 >= ncols:
                        continue
                    val = sh.cell_value(r, col - 1)
                    if val is None and attr != "codigo":
                        continue
                    s = str(val or "").strip()
                    if attr == "horario_fiscalizacao":
                        s = _excel_valor_para_horario(val) or s
                    if s:
                        row_dict[attr] = s[:200] if attr in ("tipo_atividade", "atividade") else s[:100] if attr == "grupo_atividade" else s
                ci = col_map.get("km_ini_str")
                if ci and ci <= ncols:
                    try:
                        v1 = sh.cell_value(r, ci - 1)
                        v2 = sh.cell_value(r, ci) if ci < ncols else None
                        if v1 is not None and v2 is not None:
                            row_dict["km_ini_str"] = "{} + {}".format(int(float(v1)), int(float(v2)))
                    except (ValueError, TypeError):
                        pass
                cf = col_map.get("km_fim_str")
                if cf and cf <= ncols:
                    try:
                        v1 = sh.cell_value(r, cf - 1)
                        v2 = sh.cell_value(r, cf) if cf < ncols else None
                        if v1 is not None and v2 is not None:
                            row_dict["km_fim_str"] = "{} + {}".format(int(float(v1)), int(float(v2)))
                    except (ValueError, TypeError):
                        pass
                out.append(row_dict)
        else:
            if not OPENPYXL_OK:
                return out
            wb = openpyxl.load_workbook(buf, read_only=False, data_only=True)
            ws = wb.active
            max_col_ws = getattr(ws, "max_column", None) or 30
            col_map = _detectar_colunas_template_excel(ws, is_xlrd=False, max_col=min(30, max_col_ws))
            disponiveis = _colunas_disponiveis_no_arquivo(col_map, max_col_ws)
            col_c = col_map.get("codigo") or 1
            max_row = ws.max_row or 0
            for r in range(PRIMEIRA_LINHA_DADOS, max_row + 1):
                cod = ws.cell(row=r, column=col_c).value
                if cod is None or (isinstance(cod, str) and not cod.strip()):
                    continue
                row_dict = {attr: "" for attr in CAMPOS_TEMPLATE_LIST}
                row_dict["codigo"] = str(cod).strip()
                for attr in disponiveis:
                    if attr == "codigo":
                        continue
                    col = col_map.get(attr)
                    if col is None:
                        continue
                    val = ws.cell(row=r, column=col).value
                    if val is None:
                        continue
                    s = str(val or "").strip()
                    if attr == "horario_fiscalizacao":
                        s = _excel_valor_para_horario(val) or s
                    if s:
                        row_dict[attr] = s[:200] if attr in ("tipo_atividade", "atividade") else s[:100] if attr == "grupo_atividade" else s
                ci = col_map.get("km_ini_str")
                if ci and ci < max_col_ws:
                    try:
                        v1 = ws.cell(row=r, column=ci).value
                        v2 = ws.cell(row=r, column=ci + 1).value
                        if v1 is not None and v2 is not None:
                            row_dict["km_ini_str"] = "{} + {}".format(int(float(v1)), int(float(v2)))
                    except (ValueError, TypeError):
                        pass
                cf = col_map.get("km_fim_str")
                if cf and cf < max_col_ws:
                    try:
                        v1 = ws.cell(row=r, column=cf).value
                        v2 = ws.cell(row=r, column=cf + 1).value
                        if v1 is not None and v2 is not None:
                            row_dict["km_fim_str"] = "{} + {}".format(int(float(v1)), int(float(v2)))
                    except (ValueError, TypeError):
                        pass
                out.append(row_dict)
            wb.close()
    except Exception as e:
        logger.warning("Excel complementar não pôde ser lido: %s", e)
    return out


def analisar_e_gerar_pdf(pdf_bytes: bytes,
                          limiar_km: float = LIMIAR_GAP_KM) -> tuple[bytes, bytes, dict]:
    """Pipeline completo para um único PDF. Retorna (pdf_bytes, xlsx_bytes, resumo_dict)."""
    return analisar_e_gerar_pdf_multi([pdf_bytes], limiar_km=limiar_km)


def _norm_codigo_fiscalizacao(c) -> str:
    """Normaliza código da fiscalização para cruzamento PDF–Excel (ex.: 906290.0 → '906290')."""
    if c is None:
        return ""
    s = str(c).strip()
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
        return s
    except ValueError:
        return s


def _nc_item_desde_excel(row: dict) -> NcItem:
    """Cria NcItem a partir de uma linha do Excel (PDF irmão: código existe só no Excel)."""
    cod = (row.get("codigo") or "").strip()
    km_ini_str = (row.get("km_ini_str") or "").strip()
    km_fim_str = (row.get("km_fim_str") or "").strip()
    km_ini = _km_para_float(km_ini_str)
    km_fim = _km_para_float(km_fim_str)
    nc = NcItem(
        codigo=cod,
        data_con=(row.get("data_con") or "").strip()[:50],
        horario_fiscalizacao=(row.get("horario_fiscalizacao") or "").strip()[:50],
        km_ini_str=km_ini_str[:50],
        km_fim_str=km_fim_str[:50] or km_ini_str[:50],
        km_ini=km_ini,
        km_fim=km_fim if km_fim else km_ini,
        sentido=(row.get("sentido") or "").strip()[:50],
        atividade=(row.get("atividade") or "").strip()[:200],
        tipo_atividade=(row.get("tipo_atividade") or "").strip()[:100],
        grupo_atividade=(row.get("grupo_atividade") or "").strip()[:100],
        observacao="",
        rodovia=(row.get("rodovia") or "").strip()[:50],
        rodovia_nome="",
        lote=(row.get("concessionaria") or "").strip()[:50],
        concessionaria=(row.get("concessionaria") or "").strip()[:50],
        prazo_str=(row.get("prazo_str") or "").strip()[:50],
        prazo_dias=None,
        emergencial=False,
        tipo_panela=False,
        grupo=0,
        empresa="",
        nome_fiscal=(row.get("nome_fiscal") or "").strip()[:100],
        origem_ma=False,
    )
    if (nc.km_ini_str or nc.km_ini) and (not nc.km_fim_str or nc.km_fim == 0.0):
        nc.km_fim_str = nc.km_ini_str or ""
        nc.km_fim = nc.km_ini
    return nc


def _ncs_ma_para_nc_items(ncs_ma: list) -> list[NcItem]:
    """Converte NCs de Meio Ambiente (NcItemMA) para NcItem para o relatório de análise."""
    from .analisar_pdf_ma import NcItemMA
    out: list[NcItem] = []
    for a in ncs_ma:
        if not isinstance(a, NcItemMA):
            continue
        nc = NcItem(
            codigo=(a.codigo_fiscalizacao or a.codigo or "").strip(),
            data_con=a.data_con or "",
            horario_fiscalizacao=getattr(a, "horario_fiscalizacao", "") or "",
            km_ini_str=a.km_ini_str or "",
            km_fim_str=a.km_fim_str or "",
            km_ini=a.km_ini,
            km_fim=a.km_fim,
            sentido=a.sentido or "",
            atividade=a.atividade or "",
            tipo_atividade=getattr(a, "tipo_atividade", "") or "",
            grupo_atividade=getattr(a, "grupo_atividade", "") or "",
            observacao=(a.complemento or a.relatorio or "").strip()[:200],
            rodovia=a.rodovia or "",
            rodovia_nome="",
            lote="",
            concessionaria="",
            prazo_str=a.prazo_str or "",
            prazo_dias=a.prazo_dias,
            emergencial=(a.prazo_dias is not None and a.prazo_dias <= PRAZO_EMERG_MAX),
            tipo_panela=_is_panela(a.atividade or ""),
            grupo=a.grupo,
            empresa=a.empresa or "",
            nome_fiscal=getattr(a, "nome_fiscal", "") or "",
            origem_ma=True,
        )
        if (nc.km_ini_str or nc.km_ini) and (not nc.km_fim_str or nc.km_fim == 0.0):
            nc.km_fim_str = nc.km_ini_str or ""
            nc.km_fim = nc.km_ini
        out.append(nc)
    return out


def analisar_e_gerar_pdf_multi(pdfs_bytes: list[bytes],
                                limiar_km: float = LIMIAR_GAP_KM,
                                nomes: list[str] | None = None,
                                lote: str | None = None,
                                excel_bytes: bytes | list[bytes] | None = None,
                                teste_local: bool = False) -> tuple[bytes, bytes, dict]:
    """
    Pipeline completo para múltiplos PDFs.
    Se excel_bytes for informado (Excel que acompanha os PDFs), preenche lacunas do PDF
    (horário, km, sentido, etc.) por correspondência de código — exceto lote 50 Artemig:
    tipo, grupo, atividade e prazo ficam só do PDF (export Kcor/Kria).
    Retorna (pdf_relatorio_bytes, xlsx_bytes, resumo_dict).
    """
    from .analisar_pdf_ma import parse_pdf_ma
    # Mapa EAF e responsáveis do lote selecionado (13, 21, 26). Lote 13 = config atual; 21/26 quando preenchidos.
    try:
        from nc_artesp.config import get_mapa_eaf, get_mapa_responsavel_tecnico
    except ImportError:
        try:
            from nc_artesp.config import MAPA_RESPONSAVEL_TECNICO as _RT_PADRAO
        except ImportError:
            _RT_PADRAO = {}
        get_mapa_eaf = lambda l: _MAPA_EAF_PADRAO
        get_mapa_responsavel_tecnico = lambda l: _RT_PADRAO
    lote_num = _norm_lote_numero(lote) or "13"
    mapa_eaf_lote = get_mapa_eaf(lote_num)
    if lote_num == "50":
        if not mapa_eaf_lote:
            try:
                from nc_artemig import config as _cfg_am
                mapa_eaf_lote = list(_cfg_am.MAPA_EAF_POR_LOTE.get("50") or [])
            except ImportError:
                mapa_eaf_lote = []
        if not mapa_eaf_lote:
            mapa_eaf_lote = [
                {
                    "grupo": _GRUPO_EAF_ARTEMIG_ANALISE,
                    "empresa": "CONSOL",
                    "trechos": [
                        {"rodovia": "MG 050", "km_ini": 57.6, "km_fim": 402.0},
                        {"rodovia": "BR 265", "km_ini": 637.2, "km_fim": 659.5},
                        {"rodovia": "BR 491", "km_ini": 0.0, "km_fim": 4.7},
                    ],
                }
            ]
    else:
        mapa_eaf_lote = mapa_eaf_lote or _MAPA_EAF_PADRAO
    mapa_responsavel_lote = get_mapa_responsavel_tecnico(lote_num) or {}

    ncs_total: list[NcItem] = []
    pdfs_list = list(pdfs_bytes)
    bloques: list[tuple[str, str, list[NcItem]]] = []
    for i, pdf_bytes in enumerate(pdfs_list):
        src = (nomes[i] if nomes and i < len(nomes) else f"PDF {i + 1}")
        texto_pdf = _extrair_texto_pdf(pdf_bytes) or ""
        parcial: list[NcItem] = []
        if lote_num == "50":
            parcial = parse_pdf_artemig(pdf_bytes)
        if not parcial:
            parcial = parse_pdf_nc(pdf_bytes)
        if not parcial:
            parcial_ma = parse_pdf_ma(pdf_bytes)
            if parcial_ma:
                parcial = _ncs_ma_para_nc_items(parcial_ma)
        if lote_num == "50" and parcial:
            stem_u = _stem_pdf_upload(src)
            pags = _artemig_paginas_foto_kcor(pdf_bytes)
            for nc in parcial:
                if stem_u:
                    nc.artemig_pdf_stem = stem_u
                nc.artemig_kcor_paginas_jpg = list(pags)
            fn_pdf = src if str(src).lower().endswith(".pdf") else f"{src}.pdf"
            _artemig_enriquecer_nomes_jpg_kcor_extracao(pdf_bytes, parcial, fn_pdf)
        for nc in parcial:
            setattr(nc, "_origem", src)
        bloques.append((src, texto_pdf, parcial))
        ncs_total.extend(parcial)

    if bloques:
        _validar_lotes_pdf_vs_selecionado(bloques, lote_num)

    ncs_total = _expandir_eventos_panela_nascentes(ncs_total, lote_num)

    # Atribui EAF/Grupo (para NCs de Conservação; MA já vem com grupo/empresa)
    for nc in ncs_total:
        _atribuir_grupo(nc, mapa_eaf_lote)

    # Ordena: Grupo → Rodovia → Sentido → KM
    ncs_total.sort(key=lambda n: (n.grupo or 999, n.rodovia, n.sentido, n.km_ini))

    vistos: dict[str, NcItem] = {}
    for idx, nc in enumerate(ncs_total):
        cod = (nc.codigo or "").strip()
        if not cod:
            vistos[f"_idx_{len(vistos)}"] = nc
            continue
        chave_cod = cod
        if (lote_num or "").strip() == "50" and _is_panela_artemig_nc(nc):
            prazo_key = nc.prazo_dias if nc.prazo_dias is not None else ""
            tipo_key = (nc.tipo_atividade or "").strip().lower()
            chave_cod = f"{cod}|panela|{prazo_key}|{tipo_key}|{idx}"
        if chave_cod not in vistos:
            vistos[chave_cod] = nc
        else:
            existente = vistos[chave_cod]
            if not existente.horario_fiscalizacao and nc.horario_fiscalizacao:
                existente.horario_fiscalizacao = nc.horario_fiscalizacao
            if not existente.tipo_atividade and nc.tipo_atividade:
                existente.tipo_atividade = nc.tipo_atividade
            if not existente.grupo_atividade and nc.grupo_atividade:
                existente.grupo_atividade = nc.grupo_atividade
            if not (existente.patologia_artemig or "").strip() and (nc.patologia_artemig or "").strip():
                existente.patologia_artemig = nc.patologia_artemig
            if not (existente.indicador_artemig or "").strip() and (nc.indicador_artemig or "").strip():
                existente.indicador_artemig = nc.indicador_artemig
            if not (getattr(existente, "artemig_pdf_stem", None) or "").strip() and (
                getattr(nc, "artemig_pdf_stem", None) or ""
            ).strip():
                existente.artemig_pdf_stem = nc.artemig_pdf_stem
            if not getattr(existente, "artemig_kcor_paginas_jpg", None) and getattr(
                nc, "artemig_kcor_paginas_jpg", None
            ):
                existente.artemig_kcor_paginas_jpg = list(nc.artemig_kcor_paginas_jpg)
            nn_arq = list(getattr(nc, "artemig_kcor_nomes_arquivos", None) or [])
            if nn_arq:
                ex_arq = list(getattr(existente, "artemig_kcor_nomes_arquivos", None) or [])
                visto_arq: set[str] = set()
                merged_arq: list[str] = []
                for x in ex_arq + nn_arq:
                    if x not in visto_arq:
                        visto_arq.add(x)
                        merged_arq.append(x)
                merged_arq.sort(key=lambda n: (_ordem_sufixo_nc_jpg_kcor(n), n.lower()))
                existente.artemig_kcor_nomes_arquivos = merged_arq
            # Trecho (rodovia/km) define EAF; preencher lacunas ou preferir NEP/EBP 22 sobre Autoroutes
            if not (existente.rodovia or "").strip() and (nc.rodovia or "").strip():
                existente.rodovia = nc.rodovia
                if nc.km_ini is not None:
                    existente.km_ini = nc.km_ini
                if getattr(nc, "km_fim", None) is not None:
                    existente.km_fim = nc.km_fim
            elif existente.km_ini is None and nc.km_ini is not None and (nc.rodovia or "").strip():
                existente.rodovia = existente.rodovia or nc.rodovia
                existente.km_ini = nc.km_ini
                if getattr(nc, "km_fim", None) is not None:
                    existente.km_fim = nc.km_fim
            elif (existente.rodovia or "").strip() and (nc.rodovia or "").strip() and existente.km_ini is not None and nc.km_ini is not None:
                from nc_artesp.utils.helpers import obter_grupo_empresa_por_trecho
                _ge, _ee = obter_grupo_empresa_por_trecho(existente.rodovia, existente.km_ini, mapa_eaf_lote)
                _gn, _en = obter_grupo_empresa_por_trecho(nc.rodovia, nc.km_ini, mapa_eaf_lote)
                if _ee == "Autoroutes" and _en in ("NEP", "EBP 22"):
                    existente.rodovia = nc.rodovia
                    existente.km_ini = nc.km_ini
                    if getattr(nc, "km_fim", None) is not None:
                        existente.km_fim = nc.km_fim
    ncs_total = list(vistos.values())
    ncs_total.sort(key=lambda n: (n.grupo or 999, n.rodovia, n.sentido, n.km_ini))

    # Cruzamento por código: uma NC por código; lacunas do PDF preenchidas do Excel. EAF/grupo só por trecho.
    if excel_bytes:
        if isinstance(excel_bytes, bytes):
            excel_bytes = [excel_bytes]
        linhas_excel: list[dict] = []
        for buf in excel_bytes:
            linhas_excel.extend(_ler_excel_complementar(buf))
        if linhas_excel:
            por_codigo_excel = {
                _norm_codigo_fiscalizacao(r.get("codigo")): r
                for r in linhas_excel
                if r.get("codigo") and _norm_codigo_fiscalizacao(r.get("codigo"))
            }
            # Só Excel (sem PDF): construir todas as NCs a partir do Excel
            if not ncs_total:
                for cod_excel, row in por_codigo_excel.items():
                    if cod_excel:
                        nc_excel = _nc_item_desde_excel(row)
                        setattr(nc_excel, "_origem", "Excel")
                        ncs_total.append(nc_excel)
                ncs_total.sort(key=lambda n: (n.grupo or 999, n.rodovia, n.sentido, n.km_ini))
            else:
                codigos_pdf = {_norm_codigo_fiscalizacao(nc.codigo) for nc in ncs_total if (nc.codigo or "").strip()}

                for nc in ncs_total:
                    cod = _norm_codigo_fiscalizacao(nc.codigo)
                    if not cod:
                        continue
                    row = por_codigo_excel.get(cod)
                    if not row:
                        continue
                    for attr in CAMPOS_TEMPLATE_LIST:
                        if attr == "codigo" or attr == "empresa":
                            continue
                        if not _excel_complemento_pode_mesclar_campo(lote_num, attr):
                            continue
                        val_nc = (getattr(nc, attr, None) or "") if hasattr(nc, attr) else ""
                        val_excel = (row.get(attr) or "").strip()
                        if not (val_nc or "").strip() and val_excel:
                            if attr in ("tipo_atividade", "atividade"):
                                setattr(nc, attr, val_excel[:200])
                            elif attr == "grupo_atividade":
                                setattr(nc, attr, val_excel[:100])
                            else:
                                setattr(nc, attr, val_excel[:100])
                            if attr in ("km_ini_str", "km_fim_str") and hasattr(nc, "km_ini"):
                                try:
                                    k = _km_para_float(val_excel)
                                    if attr == "km_ini_str":
                                        nc.km_ini = k
                                    else:
                                        nc.km_fim = k
                                except Exception:
                                    pass

                # Códigos que estão só no Excel: criar NC a partir da linha do Excel (PDF irmão)
                for cod_excel, row in por_codigo_excel.items():
                    if cod_excel and cod_excel not in codigos_pdf:
                        nc_excel = _nc_item_desde_excel(row)
                        setattr(nc_excel, "_origem", "Excel")
                        ncs_total.append(nc_excel)
                ncs_total.sort(key=lambda n: (n.grupo or 999, n.rodovia, n.sentido, n.km_ini))

    # EAF/grupo por trecho (rodovia+km). Trava: se EAF_PERMITIDAS estiver preenchida no config,
    # qualquer empresa fora da lista tem grupo e empresa zerados (ex.: só Autoroutes permitida).
    # EAF_PERMITIDAS vazia = todas as EAFs do MAPA_EAF são aceitas.
    try:
        from nc_artesp.config import EAF_PERMITIDAS
        permitidas = set((e or "").strip() for e in EAF_PERMITIDAS if (e or "").strip())
    except Exception:
        permitidas = set()
    for nc in ncs_total:
        if (nc.rodovia or "").strip() and nc.km_ini is not None:
            _atribuir_grupo(nc, mapa_eaf_lote)
        if permitidas and (nc.empresa or "").strip() not in permitidas:
            nc.grupo = 0
            nc.empresa = ""

    for nc in ncs_total:
        if (not (nc.grupo_atividade or "").strip() or not (nc.tipo_atividade or "").strip()) and (nc.atividade or "").strip():
            grupo_inf, tipo_inf = _inferir_grupo_tipo_da_atividade(nc.atividade)
            if grupo_inf and not (nc.grupo_atividade or "").strip():
                nc.grupo_atividade = grupo_inf
            if tipo_inf and not (nc.tipo_atividade or "").strip():
                nc.tipo_atividade = tipo_inf

    # Lote 50 (Artemig): QID + espaços — um só módulo (`nc_artemig.sanear_pipeline`) antes de relatórios.
    if lote_num == "50" and ncs_total:
        try:
            from nc_artemig.sanear_pipeline import sanear_ncs_lote50_consol

            sanear_ncs_lote50_consol(ncs_total, forcar_todas=True)
        except Exception as ex:
            logger.warning("sanear_ncs_lote50_consol: %s", ex)
            for nc in ncs_total:
                nc.lote = "50"
                nc.tipo_artemig = "QID"

    alertas_km = analisar_gaps(ncs_total, limiar_km=limiar_km, mapa_eaf=mapa_eaf_lote)
    # Regra ARTESP: salto de código ~ apontamento não entregue. Artemig: relatório integral.
    eh_artemig_50 = _pdf_eh_artemig_lote50(ncs_total)
    alertas_codigo = [] if eh_artemig_50 else analisar_sequencia_codigos(ncs_total)

    res = _montar_resumo_serializavel(ncs_total, alertas_km, alertas_codigo)
    res["n_arquivos"] = len(pdfs_list)
    data_c = res.get("data_con", "")
    data_con_dt = _parse_data(data_c)
    data_con_date = data_con_dt.date() if data_con_dt else None
    data_relatorio = date.today()
    res["relatorio_hoje"] = bool(teste_local) or (
        data_con_date is not None and data_relatorio == data_con_date
    )

    # Responsável técnico: zera só se o nome não for do mapa do lote (outra EAF).
    try:
        def _to_nome_list(v: str) -> list[str]:
            s = (v or "").strip()
            if not s:
                return []
            parts = [p.strip() for p in re.split(r"[;,]", s) if p and p.strip()]
            return parts or [s]

        empresa_para_nomes: dict[str, list[str]] = {
            (emp or "").strip(): _to_nome_list(nome_map)
            for emp, nome_map in (mapa_responsavel_lote or {}).items()
        }

        def _mapear_emp_para_chave(empresa: str) -> str:
            e = (empresa or "").strip()
            if not e:
                return ""
            e_norm = re.sub(r"\s+", " ", e)
            e_norm = re.sub(r"\bG\s*\d+\b", "", e_norm, flags=re.IGNORECASE).strip()
            if e_norm in empresa_para_nomes:
                return e_norm
            if e in empresa_para_nomes:
                return e
            e_up = e_norm.upper()
            for k in empresa_para_nomes.keys():
                kk = (k or "").strip()
                if not kk:
                    continue
                kk_up = kk.upper()
                if kk_up in e_up or e_up in kk_up:
                    return k
            return e_norm  # fallback: não encontrado

        # União (apenas para decidir se o nome existe em algum lugar da regra)
        responsaveis_regra = set()
        for nomes in empresa_para_nomes.values():
            for n in nomes:
                nn = (n or "").strip()
                if nn:
                    responsaveis_regra.add(nn)

        def _nome_na_lista_por_substring(nome: str, lista_nomes: list[str]) -> bool:
            n = (nome or "").strip()
            if not n:
                return False
            for r in (lista_nomes or []):
                rr = (r or "").strip()
                if not rr:
                    continue
                if n == rr:
                    return True
                if n in rr or rr in n:
                    return True
            return False

        for nc in ncs_total:
            emp_raw = (nc.empresa or "").strip()
            emp = _mapear_emp_para_chave(emp_raw)
            if not (nc.nome_fiscal or "").strip():
                nc.nome_fiscal = (mapa_responsavel_lote.get(emp) or "").strip()

            nome = (nc.nome_fiscal or "").strip()
            if not nome:
                continue

            nomes_ok = empresa_para_nomes.get(emp, [])
            if _nome_na_lista_por_substring(nome, nomes_ok):
                continue
            # Nome não está na lista desta EAF. Só zera se o nome pertencer a OUTRA EAF (conflito).
            # Se não estiver em nenhum mapa, mantém empresa do trecho (não zera por nome desconhecido).
            nome_em_outra = False
            for outra_emp, lista_outra in empresa_para_nomes.items():
                if outra_emp == emp or not lista_outra:
                    continue
                if _nome_na_lista_por_substring(nome, lista_outra):
                    nome_em_outra = True
                    break
            if nome_em_outra:
                nc.grupo = 0
                nc.empresa = ""
    except Exception:
        pass
    rotulo_lo, slug_zip = rotulo_e_slug_lote_para_saida(lote_num)
    res["rotulo_lote_analise"] = rotulo_lo
    res["slug_zip"] = slug_zip
    pdf_rel = gerar_relatorio_pdf(
        ncs_total, alertas_km, alertas_codigo,
        limiar_km=limiar_km, mapa_eaf=mapa_eaf_lote,
        rotulo_lote_analise=rotulo_lo,
        forcar_alertas_pdf=bool(teste_local),
    )
    lote_ok = (lote or "").strip() or None
    xlsx_bytes = gerar_relatorio_xlsx(
        ncs_total, lote_selecionado=lote_ok, rotulo_lote_analise=rotulo_lo
    )
    res["exportar_kcor_xlsx"] = None
    res["exportar_kcor_nome"] = None
    if lote_num == "50" and ncs_total:
        try:
            from nc_artemig.exportar_kcor_planilha import gerar_exportar_kcor_xlsx_bytes
            from nc_artemig.config import nome_saida_excel_kcor

            kcor_b, kcor_meta = gerar_exportar_kcor_xlsx_bytes(ncs_total)
            res["exportar_kcor_meta"] = kcor_meta
            if kcor_b:
                res["exportar_kcor_xlsx"] = kcor_b
                res["exportar_kcor_nome"] = nome_saida_excel_kcor()
            else:
                res["exportar_kcor_nao_gerado"] = True
                logger.warning(
                    "Lote 50: Exportar Kcor não gerado: %s",
                    kcor_meta.get("motivo") or kcor_meta,
                )
        except Exception as ex:
            res["exportar_kcor_nao_gerado"] = True
            res["exportar_kcor_meta"] = {"ok": False, "motivo": str(ex), "modelo_minimo_gerado": False}
            logger.warning("Lote 50: Exportar Kcor não gerado: %s", ex)
    return pdf_rel, xlsx_bytes, res
