r"""
modulos/exportar_calendario.py
────────────────────────────────────────────────────────────────────────────
Equivalente VBA: Art_06_EAF_Rot_Exportar_Calend

Para cada linha da planilha acumulada (col A=seq ... Y=nr Kria) cria um
AppointmentItem no calendário Outlook (subpasta 'Exportar'):

  Assunto : {TipoNC} - {Rodovia} {KMi} {Sentido} - Kria: {NumKria}
  Corpo   : {ObsGestor}\n\n - Data Constatacao: {DataSolicitacao[:10]}\n\n{Observacoes}
  Inicio  : data extraida de col U (ultimos 10 chars)  -> AllDayEvent=True
  Anexo 1 : {Diretorio}\{arquivo.jpg}   (primeira parte do col W, antes do ';')
  Anexo 2 : {Diretorio}\{pdf(N).jpg}    (segunda parte do col W, apos o ';')

Apos criar os eventos, chama automaticamente o modulo 08 (salvar_imagem).

Requer: Windows + Outlook instalado + pywin32
"""

import logging
from pathlib import Path

from openpyxl import load_workbook

from config import M06_PASTA_OUTLOOK
from utils.helpers import (
    escrever_bytes_caminho,
    garantir_pasta,
    resolver_path_ficheiro_ci,
    str_caminho_io_windows,
)

logger = logging.getLogger(__name__)

_E  = 5   # Tipo NC
_F  = 6   # Rodovia
_G  = 7   # KM inicial
_I  = 9   # Sentido
_M  = 13  # Data Solicitação (data constatação)
_T  = 20  # Obs Gestor
_U  = 21  # Observações / Data superação (últimos 10 chars = data início)
_V  = 22  # Diretório das imagens
_W  = 23  # Arquivos "arq.jpg;pdf(N).jpg"
_Y  = 25  # Número Kria


def _cell(ws, row: int, col: int) -> str:
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""


def _data_inicio(obs_gestor: str) -> str:
    """Extrai data do campo Obs Gestor (col U) — últimos 10 chars."""
    s = obs_gestor.strip()
    return s[-10:] if len(s) >= 10 else s


def _criar_via_win32com(arquivo_acumulado: Path,
                         pasta_outlook: str,
                         callback_progresso=None) -> int:
    """
    Cria os AppointmentItems via pywin32 / Outlook COM.
    Retorna número de eventos criados.
    """
    try:
        import win32com.client as win32
    except ImportError:
        raise ImportError(
            "pywin32 não instalado. Execute: pip install pywin32\n"
            "Requer Windows + Outlook instalado."
        )

    logger.info("Conectando ao Outlook...")
    outlook    = win32.Dispatch("Outlook.Application")
    namespace  = outlook.GetNamespace("MAPI")
    cal_folder = namespace.GetDefaultFolder(9)  # olFolderCalendar = 9

    # Subpasta "Exportar"
    try:
        folder = cal_folder.Folders[pasta_outlook]
    except Exception:
        folder = cal_folder.Folders.Add(pasta_outlook)
        logger.info(f"  Subpasta '{pasta_outlook}' criada no calendário.")

    wb = load_workbook(str_caminho_io_windows(arquivo_acumulado), data_only=True)
    ws = wb.active

    ultima = ws.max_row
    for r in range(ultima, 1, -1):
        if ws.cell(row=r, column=1).value is not None:
            ultima = r
            break

    total = ultima - 1
    criados = 0

    for idx, r in enumerate(range(2, ultima + 1)):
        tipo_nc   = _cell(ws, r, _E)
        rodovia   = _cell(ws, r, _F)
        km_i      = _cell(ws, r, _G)
        sentido   = _cell(ws, r, _I)
        dt_sol    = _cell(ws, r, _M)
        obs_gest  = _cell(ws, r, _T)
        obs_geral = _cell(ws, r, _U)
        diretorio = _cell(ws, r, _V)
        arquivos  = _cell(ws, r, _W)
        num_kria  = _cell(ws, r, _Y)

        if callback_progresso:
            callback_progresso(idx + 1, total, f"Criando evento {idx+1}/{total}...")

        if not tipo_nc:
            continue

        assunto = f"{tipo_nc} - {rodovia} {km_i} {sentido} - Kria: {num_kria}"
        data_con_str = dt_sol[:10] if len(dt_sol) >= 10 else dt_sol
        corpo = f"{obs_gest}\n\n - Data Constatação: {data_con_str}\n\n{obs_geral}"
        data_inicio = _data_inicio(obs_geral)

        try:
            appt = folder.Items.Add()
            appt.Subject     = assunto
            appt.Body        = corpo
            appt.AllDayEvent = True

            if data_inicio:
                import win32com.client
                try:
                    appt.Start = data_inicio
                except Exception:
                    pass  # data inválida — ignora e ainda salva o evento

            if arquivos and diretorio:
                partes = arquivos.split(";")
                arq1 = partes[0].strip()
                arq2 = partes[1].strip() if len(partes) > 1 else ""

                p1 = Path(diretorio) / arq1
                if p1.exists():
                    appt.Attachments.Add(str(p1))
                else:
                    logger.warning(f"  Anexo 1 não encontrado: {p1}")

                if arq2:
                    p2 = Path(diretorio) / arq2
                    if p2.exists():
                        appt.Attachments.Add(str(p2))
                    else:
                        logger.warning(f"  Anexo 2 não encontrado: {p2}")

            appt.Save()
            criados += 1
            logger.debug(f"  Evento criado: {assunto[:60]}")

        except Exception as e:
            logger.error(f"  Erro ao criar evento linha {r}: {e}")

    wb.close()
    return criados


def _criar_via_ical(arquivo_acumulado: Path,
                    pasta_saida: Path,
                    callback_progresso=None) -> Path:
    """
    Gera .ics sem Outlook, usando o mesmo motor que ``gerar_ics_bytes`` (RFC 5545, sem icalendar).
    Evita falhas no Render quando o pacote opcional ``icalendar`` falta ou incompatibilidade de versão.
    """
    pio = str_caminho_io_windows(arquivo_acumulado)
    with open(pio, "rb") as f:
        raw = f.read()
    if not raw:
        raise ValueError("Acumulado.xlsx está vazio.")
    if raw.startswith(b"version https://git-lfs"):
        raise ValueError(
            "Acumulado.xlsx no disco é ponteiro Git LFS (bytes insuficientes). "
            "Garanta Git LFS no build ou regenere o ficheiro no servidor."
        )
    if callback_progresso:
        callback_progresso(1, 1, "Gerando .ics (RFC 5545)...")
    ics_bytes, adicionados = gerar_ics_bytes(raw)
    garantir_pasta(pasta_saida)
    from utils.helpers import timestamp_agora

    destino = pasta_saida / f"{timestamp_agora()} - eventos_kartado.ics"
    escrever_bytes_caminho(destino, ics_bytes)
    logger.info("  .ics gerado: %s (%s evento(s))", destino.name, adicionados)
    return destino


def _escape_ics(s: str) -> str:
    """Escapa caracteres especiais para texto ICS."""
    return (
        s.replace("\\", "\\\\")
         .replace(";", "\\;")
         .replace(",", "\\,")
         .replace("\n", "\\n")
         .replace("\r", "")
    )


def _fold_ics(line: str) -> str:
    """Dobra linhas longas conforme RFC 5545 (máx 75 octetos por linha)."""
    if len(line.encode("utf-8")) <= 75:
        return line
    parts = []
    encoded = line.encode("utf-8")
    pos = 0
    first = True
    while pos < len(encoded):
        chunk_size = 75 if first else 74
        chunk = encoded[pos: pos + chunk_size]
        # Não quebrar no meio de um multi-byte UTF-8
        while len(chunk) and (chunk[-1] & 0xC0) == 0x80:
            chunk = chunk[:-1]
        parts.append((" " if not first else "") + chunk.decode("utf-8", errors="replace"))
        pos += len(chunk)
        first = False
    return "\r\n".join(parts)


def _data_ics(data_str: str) -> str:
    """Converte 'DD/MM/YYYY' ou 'MM/DD/YYYY' para 'YYYYMMDD' do ICS."""
    s = data_str.strip()
    # Tenta parse via helpers
    try:
        from utils.helpers import parse_data
        dt = parse_data(s)
        if dt:
            return dt.strftime("%Y%m%d")
    except Exception:
        pass
    # Tenta DD/MM/YYYY direto
    import re
    m = re.match(r"(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if d <= 12 and mo <= 12:  # ambíguo → assume DD/MM
            return f"{y:04d}{mo:02d}{d:02d}"
        return f"{y:04d}{mo:02d}{d:02d}"
    return ""


def gerar_ics_bytes(xlsx_bytes: bytes) -> tuple[bytes, int]:
    """
    Lê a planilha acumulada (saída do M05) e gera um arquivo .ics (iCalendar)
    compatível com Outlook, Google Calendar, Apple Calendar etc.
    Funciona sem Outlook e sem bibliotecas externas.
    Retorna (ics_bytes, n_eventos).
    """
    import io
    import uuid
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    ultima = ws.max_row
    for r in range(ultima, 1, -1):
        if ws.cell(row=r, column=1).value is not None:
            ultima = r
            break

    def cell(r: int, c: int) -> str:
        v = ws.cell(row=r, column=c).value
        return str(v).strip() if v is not None else ""

    linhas = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//ARTESP NC Pipeline//PT",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
    ]

    n_eventos = 0
    for r in range(2, ultima + 1):
        tipo_nc  = cell(r, _E)
        rodovia  = cell(r, _F)
        km_i     = cell(r, _G)
        sentido  = cell(r, _I)
        dt_sol   = cell(r, _M)
        obs_gest = cell(r, _T)
        obs_ger  = cell(r, _U)
        num_kria = cell(r, _Y)

        if not tipo_nc:
            continue

        assunto  = f"{tipo_nc} - {rodovia} {km_i} {sentido} - Kria: {num_kria}"
        data_con = dt_sol[:10] if len(dt_sol) >= 10 else dt_sol
        corpo    = f"{obs_gest}\n\n - Data Constatação: {data_con}\n\n{obs_ger}"

        # Data de início: últimos 10 chars de col U (data superação/prazo)
        data_inicio_s = obs_ger.strip()[-10:] if len(obs_ger.strip()) >= 10 else ""
        dtstart = _data_ics(data_inicio_s)
        if not dtstart:
            dtstart = _data_ics(data_con)  # fallback: data constatação

        # DTEND = mesmo dia (AllDayEvent, conforme macro original)
        dtend = dtstart

        uid = str(uuid.uuid4())
        linhas += [
            "BEGIN:VEVENT",
            f"UID:{uid}",
            _fold_ics(f"SUMMARY:{_escape_ics(assunto)}"),
            _fold_ics(f"DESCRIPTION:{_escape_ics(corpo)}"),
        ]
        if dtstart:
            linhas += [
                f"DTSTART;VALUE=DATE:{dtstart}",
                f"DTEND;VALUE=DATE:{dtend}",
            ]
        linhas.append("END:VEVENT")
        n_eventos += 1

    linhas.append("END:VCALENDAR")

    ics_content = "\r\n".join(linhas) + "\r\n"
    return ics_content.encode("utf-8"), n_eventos


def executar(arquivo_acumulado: Path,
             usar_outlook: bool = True,
             pasta_saida_ics: Path | None = None,
             executar_mod08: bool = True,
             callback_progresso=None) -> dict:
    """
    Exporta eventos para Outlook (ou gera .ics como alternativa).

    Parâmetros
    ----------
    arquivo_acumulado : Planilha acumulada com col Y preenchida (após mod 05).
    usar_outlook      : True = Outlook COM (Windows); False = gerar .ics.
    pasta_saida_ics   : Pasta para o .ics (se usar_outlook=False).
    executar_mod08    : Após exportar calendário, chama módulo 08 automaticamente.

    Retorna dict com 'eventos', 'ics' (ou None), 'mod08_imagens'.
    """
    resultado = {"eventos": 0, "ics": None, "mod08_imagens": 0}

    arquivo_acumulado = resolver_path_ficheiro_ci(arquivo_acumulado)

    if usar_outlook:
        logger.info("Módulo 06: Exportando para calendário Outlook...")
        try:
            criados = _criar_via_win32com(
                arquivo_acumulado, M06_PASTA_OUTLOOK, callback_progresso
            )
            resultado["eventos"] = criados
            logger.info(f"Módulo 06 concluído: {criados} evento(s) criado(s).")
        except Exception as e:
            logger.error(f"Módulo 06 (Outlook): {e}")
            raise
    else:
        logger.info("Módulo 06: Gerando arquivo .ics...")
        from config import M04_SAIDA
        pasta_ics = pasta_saida_ics or M04_SAIDA
        ics_path = _criar_via_ical(arquivo_acumulado, pasta_ics, callback_progresso)
        resultado["ics"] = ics_path

    if executar_mod08:
        logger.info("Módulo 06 → chamando Módulo 08 automaticamente...")
        try:
            from modulos.salvar_imagem import executar as salv08
            total_img = salv08(arquivo_acumulado)
            resultado["mod08_imagens"] = total_img
            logger.info(f"Módulo 08 (chamado pelo 06): {total_img} imagem(ns).")
        except Exception as e:
            logger.warning(f"Módulo 08 não executado: {e}")

    if callback_progresso:
        callback_progresso(1, 1, "Módulo 06 concluído.")

    return resultado
