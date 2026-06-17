"""
modulos/gerar_modelo_foto.py
──────────────────────────────────────────────────────────────────────────
Equivalente VBA: Art_022_EAF_Gerar_Mod_Ft_Exc_NC
Desenvolvedor: Ozeias Engler

Para cada XLS individual gerado pelo Módulo 01 (pasta Exportar/), produz
duas saídas (sequência macros Artesp 02): Kria e Resposta.

─── SAÍDA A – Planilha Kria de Abertura de Evento ──────────────────────
  Arquivo: yyyymmdd-hhmm - {nome_sem_extensao}.xlsx
  Pasta:   Arquivos/Arquivo Foto - Conserva/
  Modelo:  Modelo Abertura Evento Kria Conserva Rotina.
  Foto:    nc (N).jpg   ← foto da vistoria de campo

─── SAÍDA B – Relatório de Resposta à Artesp ───────────────────────────
  Arquivo: agrupamento igual ao Exportar (tipo, prazo, rodovia, responsável); nome do ficheiro =
           ``nome_resposta_saida.nome_ficheiro_resposta_artesp_xlsx`` (prazo + código + data const. + **atividade da mãe**).
  Pasta:   _Respostas/_Relatório EAF - NC/Pendentes/
  Modelo:  Modelo Resposta.
  Foto:    PDF (N).jpg  ← foto extraída do PDF
"""

import logging
import math
import re
import unicodedata
from copy import copy
from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from PIL import Image as PILImage

from datetime import timedelta

from utils.image_anchor import patch_add_image, get_merged_bounds

from .nome_resposta_saida import (
    nome_ficheiro_resposta_artesp_xlsx,
    _texto_atividade_exibicao_pendentes,
)
from .separar_nc import (
    _detectar_col_atividade_mae_col_n,
    _detectar_col_tipo_nc,
    _sanitizar_nome_xlsx,
    chave_agrupamento_exportar_rotina_valores,
)
from config import (
    M01_EXPORTAR,
    M02_FOTOS_NC,
    M02_FOTOS_PDF,
    M02_MODELO_KRIA,
    M02_SALVAR_FOTO,
    M02_MODELO_RESP,
    M02_PENDENTES,
    M02_FOTO_W, M02_FOTO_H,
    M02_FOTO_PDF_W, M02_FOTO_PDF_H,
    KCOR_KRIA_DIRETORIO_TEXTO_CONSERVACAO,
    PRAZO_DIAS_APOS_ENVIO,
    RESPOSTA_PENDENTES_HEADER_MAE_ATIVIDADE_Q,
    RESPOSTA_PENDENTES_HEADER_MAE_TIPO_ATIV_O,
    RODOVIAS,
)
from utils.excel_io import xls_to_xlsx
from utils.helpers import (
    garantir_pasta,
    parse_data,
    data_br,
    data_yyyymmdd,
    km_mais_metros,
    km_virgula_metros,
    normalizar_rodovia_eaf,
    sanitizar_nome,
    timestamp_agora,
    timestamp_completo,
    caminho_dentro_limite_windows,
    encontrar_foto_por_codigo_ou_numero,
    resolver_path_ficheiro_ci,
    str_caminho_io_windows,
    detectar_coluna_tipo_de_atividade_eaf,
    ficheiro_em_subpasta_planilha_mae_kartado_exportar,
    preservar_ooxml_planilha_pos_openpyxl,
    replicar_ancoras_sem_foto_por_bloco_em_drawing,
)

logger = logging.getLogger(__name__)

BLOCO      = 5    # linhas por NC no Kria
BLOCO_RESP = 28   # linhas por NC no Relatório de Resposta
_ROW_LEGENDA_FOTO_RESP = 17
_ROW_DATA_EXEC_RESP = 26
# No Kria (Art_03), V/W ficam na linha j+1 com j=8 → linha 9. O modelo Pendentes replica a mesma grelha no 1.º bloco.
_OFFSET_ROW_VW_KCOR_DENTRO_BLOCO_PENDENTES = 8
_TEXTO_LEGENDA_FOTO_RESP = "Registro fotográfico da concessionária:"

# Índices de coluna do XLS individual (Módulo 01)
_C  = 3   # C – código fiscalização
_D  = 4   # D – data constatação
_E  = 5   # E – horário
_F  = 6   # F – rodovia
_H  = 8   # H – km inicial (int)
_I  = 9   # I – metros inicial
_J  = 10  # J – km final
_K  = 11  # K – metros final
_L  = 12  # L – sentido
_O  = 15  # O – tipo atividade (mãe EAF; nome Respostas Pendentes)
_Q  = 17  # Q – tipo NC (Atividade)
_DR = 20  # T – Data Reparo (fallback; detectado dinamicamente)
_V  = 22  # V – nº da NC para foto

_NC_NOME_ARQ = {
    "Não Conformidade": "NC",
    "Advertência":      "ADV",
    "Notificação":      "NOT",
}

# UTILITÁRIOS DE CAMINHO DE FOTO
# Macro: antigo nc (1).jpg; atual nc (00001).jpg (número da NC, 5 dígitos).

def _codigo_estilo_ma(codigo: object) -> bool:
    """True se o código é no padrão MA (ex.: NC.13.1039, HE.13.0112). Só MA tem duas fotos por NC."""
    if codigo is None:
        return False
    s = str(codigo).strip()
    return bool(s) and "." in s and any(c.isalpha() for c in s)


def _variantes_nome_foto(prefixo: str, num: object) -> list:
    """Retorna variantes de nome: nc (1).jpg e nc (00001).jpg (5 dígitos), sem repetir."""
    out = [f"{prefixo} ({num}).jpg"]
    try:
        n = int(num)
        nome_5 = f"{prefixo} ({n:05d}).jpg"
        if nome_5 not in out:
            out.append(nome_5)
    except (TypeError, ValueError):
        pass
    return out


def _numero_opcional_de_celula(num: object) -> int | None:
    if num is None:
        return None
    s = str(num).strip()
    if not s:
        return None
    try:
        f = float(s.replace(",", "."))
        if f != int(f):
            return None
        return int(f)
    except (ValueError, TypeError, OverflowError):
        return None


def _celula_parece_cabecalho_ou_coordenada(cod: str) -> bool:
    """Coluna C com rótulo de planilha (Longitude, etc.) — não é código fiscalização."""
    if not cod:
        return True
    s = str(cod).strip().lower()
    if s in (
        "longitude", "latitude", "lat", "long", "coordenada", "coordenadas",
        "northing", "easting", "utm", "código", "codigo", "cod", "id",
    ):
        return True
    if "longitude" in s or "latitude" in s:
        return True
    return False


def _candidatos_identificador_foto(nc: dict) -> list:
    """
    Ordem de tentativa para casar arquivo no disco: código (col C), foto_id e col V.
    Prioriza o padrão histórico de renomeação por código de fiscalização.
    """
    seen: set[str] = set()
    out: list = []

    def add(v: object) -> None:
        if v is None:
            return
        if isinstance(v, float) and v == int(v):
            v = int(v)
        s = str(v).strip()
        if not s or s == "0":
            return
        key = s.lower()
        if key in seen:
            return
        seen.add(key)
        out.append(v)

    cod = (nc.get("codigo") or "").strip() if nc.get("codigo") is not None else ""
    if cod and not cod.upper().startswith("LOTE") and not _celula_parece_cabecalho_ou_coordenada(cod):
        add(cod)
    add(nc.get("foto_id", nc.get("num_foto")))
    add(nc.get("num_foto"))
    return out


def path_foto_nc(pasta: Path, num: object) -> Path:
    """Foto de campo  →  nc (N).jpg ou nc (00000).jpg   — usada no relatório Kria."""
    if not pasta:
        return Path()
    if not pasta.is_dir():
        return pasta / f"nc ({num}).jpg"
    bases = []
    sub_nc = pasta / "nc"
    if sub_nc.is_dir():
        bases.append(sub_nc)
    bases.append(pasta)
    for nome in _variantes_nome_foto("nc", num):
        for base in bases:
            direto = base / nome
            if direto.exists():
                return direto
    nu = _numero_opcional_de_celula(num)
    cs = str(num).strip() if num is not None else ""
    enr = encontrar_foto_por_codigo_ou_numero(pasta, "nc", codigo=cs or None, numero=nu)
    if enr is not None and enr.is_file():
        return enr
    # Fallback de compatibilidade para prefixo em caixa diferente
    enr2 = encontrar_foto_por_codigo_ou_numero(pasta, "NC", codigo=cs or None, numero=nu)
    if enr2 is not None and enr2.is_file():
        return enr2
    return bases[0] / f"nc ({num}).jpg"

def path_foto_pdf(pasta: Path, num: object) -> Path:
    """Foto extraída do PDF  →  PDF (N).jpg   — usada no relatório Resposta. Não alterar."""
    if not pasta:
        return Path()
    if not pasta.is_dir():
        return pasta / f"PDF ({num}).jpg"
    bases = []
    sub_pdf = pasta / "PDF"
    if sub_pdf.is_dir():
        bases.append(sub_pdf)
    bases.append(pasta)
    nome = f"PDF ({num}).jpg"
    for base in bases:
        direto = base / nome
        if direto.exists():
            return direto
    nu = _numero_opcional_de_celula(num)
    cs = str(num).strip() if num is not None else ""
    enr = encontrar_foto_por_codigo_ou_numero(pasta, "PDF", codigo=cs or None, numero=nu)
    if enr is not None and enr.is_file():
        return enr
    # Fallback de compatibilidade para prefixo em caixa diferente
    enr2 = encontrar_foto_por_codigo_ou_numero(pasta, "pdf", codigo=cs or None, numero=nu)
    if enr2 is not None and enr2.is_file():
        return enr2
    return bases[0] / nome


def path_foto_nc_segunda(pasta: Path, codigo: object) -> Path:
    """Segunda foto: nc (codigo)_1.jpg. MA = col C (ex.: NC.13.1039); conservação = num (ex.: 1 ou 00001)."""
    if not pasta or codigo is None:
        return Path()
    cod = str(codigo).strip()
    if not cod:
        return Path()
    variantes = [f"nc ({cod})_1.jpg"]
    try:
        n = int(cod)
        variantes.append(f"nc ({n:05d})_1.jpg")
    except (TypeError, ValueError):
        pass
    bases = []
    if pasta.is_dir():
        sub_nc = pasta / "nc"
        if sub_nc.is_dir():
            bases.append(sub_nc)
        bases.append(pasta)
    else:
        bases.append(pasta.parent)
    for nome in variantes:
        for base in bases:
            direto = base / nome
            if direto.exists():
                return direto
    return Path()


# EMU: 1 cm = 914400/2.54 no OOXML; 1 px de referência Excel ≈ 96 DPI → 914400/96 EMU/px
_EMU_PER_CM = 914400 / 2.54
_EMU_PER_PX_96 = 914400 / 96
# Tamanho ideal da foto no Kria em cm (extent = exatamente isso no Excel)
_KRIA_FOTO_W_CM = 9.70
_KRIA_FOTO_H_CM = 7.49
_PX_PER_CM = 96 / 2.54

# UTILITÁRIOS DE IMAGEM

def _col_px(ws, col_letter: str) -> float:
    dim = ws.column_dimensions.get(col_letter)
    width = dim.width if (dim and dim.width) else 8.0
    return width * 7 + 5


def _col_px_display(ws, col_letter: str) -> int:
    """Largura da coluna como o Excel desenha (~7 px por unidade). Usado no extent da âncora para a foto caber no merge."""
    dim = ws.column_dimensions.get(col_letter)
    width = dim.width if (dim and dim.width) else 8.0
    return max(1, int((width + 0.5) * 7))


def _row_px(ws, row: int) -> float:
    dim = ws.row_dimensions.get(row)
    height = dim.height if (dim and dim.height) else 15.0
    return height * 4 / 3


def _merged_range_px(ws, cell_addr: str):
    """Retorna (largura_px, altura_px) do merged range que contém cell_addr."""
    from openpyxl.utils import coordinate_to_tuple
    row, col = coordinate_to_tuple(cell_addr)
    for mc in ws.merged_cells.ranges:
        if mc.min_row <= row <= mc.max_row and mc.min_col <= col <= mc.max_col:
            w = sum(_col_px(ws, get_column_letter(c)) for c in range(mc.min_col, mc.max_col + 1))
            h = sum(_row_px(ws, r) for r in range(mc.min_row, mc.max_row + 1))
            return w, h
    col_letter = get_column_letter(col)
    return _col_px(ws, col_letter), _row_px(ws, row)


def _merged_range_px_extent(ws, cell_addr: str) -> tuple:
    """Tamanho do merge como o Excel desenha (_col_px_display), para extent da âncora — foto não ultrapassa o quadro."""
    from openpyxl.utils import coordinate_to_tuple
    row, col = coordinate_to_tuple(cell_addr)
    for mc in ws.merged_cells.ranges:
        if mc.min_row <= row <= mc.max_row and mc.min_col <= col <= mc.max_col:
            w = sum(_col_px_display(ws, get_column_letter(c)) for c in range(mc.min_col, mc.max_col + 1))
            h = sum(int(_row_px(ws, r)) for r in range(mc.min_row, mc.max_row + 1))
            return max(w, 1), max(h, 1)
    col_letter = get_column_letter(col)
    return max(_col_px_display(ws, col_letter), 1), max(int(_row_px(ws, row)), 1)


def _merged_range_px_fill(ws, cell_addr: str) -> tuple[int, int]:
    """Largura/altura em px para bitmap: maior entre estimativa «display» e fórmula clássica (preenche merge no ecrã)."""
    we, he = _merged_range_px_extent(ws, cell_addr)
    wf, hf = _merged_range_px(ws, cell_addr)
    w = max(we, int(math.ceil(wf))) if wf else we
    h = max(he, int(math.ceil(hf))) if hf else he
    return max(w, 1), max(h, 1)


def _is_coluna_c_kria_foto(cell_addr: str) -> bool:
    """Foto Kria está na coluna C (3). ``startswith('C')`` falhava em CE7, CA1, etc."""
    try:
        col_letters, _ = coordinate_from_string(str(cell_addr).replace("$", ""))
    except (ValueError, AttributeError, TypeError):
        return False
    return column_index_from_string(col_letters) == 3

def _log_draft_ram(ident: str, size_before: tuple, size_after: tuple, channels: int = 3) -> None:
    """Log estimativa de RAM economizada por draft() (só em DEBUG)."""
    if not logger.isEnabledFor(logging.DEBUG):
        return
    w0, h0 = size_before
    w1, h1 = size_after
    full_mb = (w0 * h0 * channels) / (1024 * 1024)
    after_mb = (w1 * h1 * channels) / (1024 * 1024)
    saved = max(0.0, full_mb - after_mb)
    logger.debug("[draft] %s: %dx%d → %dx%d | ~%.2f MB RAM economizados", ident, w0, h0, w1, h1, saved)

def _redimensionar_imagem_bytes(img_path: Path, largura: int, altura: int) -> bytes:
    """Redimensiona para miniatura. draft() em JPEG reduz RAM; resize() garante tamanho exato."""
    with PILImage.open(str(img_path)) as im:
        try:
            from PIL import ImageOps

            im = ImageOps.exif_transpose(im)
        except (ImportError, AttributeError, TypeError, ValueError):
            pass
        if getattr(im, "format", None) == "JPEG" and (im.width > largura or im.height > altura):
            before = (im.width, im.height)
            try:
                im.draft("RGB", (int(largura), int(altura)))
                _log_draft_ram(img_path.name, before, (im.width, im.height))
            except (AttributeError, TypeError, ValueError):
                pass
        im = im.convert("RGB")
        im = im.resize((int(largura), int(altura)), PILImage.LANCZOS)
        buf = BytesIO()
        im.save(buf, format="PNG")
        return buf.getvalue()


class _ImageFromBytes(XLImage):
    """Imagem a partir de bytes em memória; openpyxl lê no save(), então mantemos referência aos bytes."""

    def __init__(self, data: bytes):
        super().__init__(BytesIO(data))
        self._bytes_data = data

    def _data(self):
        return self._bytes_data


def _tamanho_foto_kria_px() -> tuple[int, int]:
    """Pixels da imagem para 9,70 × 7,49 cm (96 DPI)."""
    w = max(1, int(_KRIA_FOTO_W_CM * _PX_PER_CM))
    h = max(1, int(_KRIA_FOTO_H_CM * _PX_PER_CM))
    return w, h


def _extent_foto_kria_emu() -> tuple[int, int]:
    """Extent da âncora em EMU para exatamente 9,70 × 7,49 cm no Excel."""
    w_emu = int(_KRIA_FOTO_W_CM * _EMU_PER_CM)
    h_emu = int(_KRIA_FOTO_H_CM * _EMU_PER_CM)
    return w_emu, h_emu


def _inserir_imagem(ws, cell_addr: str, img_path: Path, largura: int, altura: int):
    """Kria (coluna C): extent fixo 9,70×7,49 cm. Resposta e restantes: OneCell + extent EMU = merge (PDF/JPG preenchem o quadro)."""
    from openpyxl.utils import coordinate_to_tuple

    row_num, col_num = coordinate_to_tuple(cell_addr)
    min_col, min_row, max_col, max_row = get_merged_bounds(ws, col_num, row_num)
    if _is_coluna_c_kria_foto(cell_addr):
        w, h = _tamanho_foto_kria_px()
        w_emu, h_emu = _extent_foto_kria_emu()
        data = _redimensionar_imagem_bytes(img_path, w, h)
        xl_img = _ImageFromBytes(data)
        xl_img.width = w
        xl_img.height = h
        anchor = OneCellAnchor()
        anchor._from = AnchorMarker(col=min_col - 1, colOff=0, row=min_row - 1, rowOff=0)
        anchor.ext = XDRPositiveSize2D(cx=w_emu, cy=h_emu)
        xl_img.anchor = anchor
        ws.add_image(xl_img)
        return

    w, h = _merged_range_px_fill(ws, cell_addr)
    if w < 2 or h < 2:
        w, h = max(1, int(largura)), max(1, int(altura))
    data = _redimensionar_imagem_bytes(img_path, w, h)
    xl_img = _ImageFromBytes(data)
    xl_img.width = w
    xl_img.height = h
    cx = max(1, int(w * _EMU_PER_PX_96))
    cy = max(1, int(h * _EMU_PER_PX_96))
    anchor = OneCellAnchor()
    anchor._from = AnchorMarker(col=min_col - 1, colOff=0, row=min_row - 1, rowOff=0)
    anchor.ext = XDRPositiveSize2D(cx=cx, cy=cy)
    xl_img.anchor = anchor
    ws.add_image(xl_img)


def _inserir_duas_imagens_ma(
    ws, cell_addr: str, path1: Path, path2: Path, largura: int, altura: int
) -> None:
    """
    Meio Ambiente: insere duas fotos NC no mesmo quadro do Kria.
    Tamanho fixo ideal 9,70 cm × 7,49 cm. Primeira no canto superior esquerdo, segunda no inferior direito.
    """
    from openpyxl.utils import coordinate_to_tuple

    w_px, h_px = _tamanho_foto_kria_px()
    w_emu, h_emu = _extent_foto_kria_emu()
    row, col = coordinate_to_tuple(cell_addr)
    min_col, min_row, max_col, max_row = get_merged_bounds(ws, col, row)
    half_w = w_emu // 2
    half_h = h_emu // 2
    half_w_px = max(1, w_px // 2)
    half_h_px = max(1, h_px // 2)

    for idx, (img_path, col_off_emu, row_off_emu, ext_w_emu, ext_h_emu, ext_w_px, ext_h_px) in enumerate([
        (path1, 0, 0, half_w, half_h, half_w_px, half_h_px),
        (path2, half_w, half_h, half_w, half_h, half_w_px, half_h_px),
    ]):
        if not img_path or not img_path.is_file():
            continue
        data = _redimensionar_imagem_bytes(img_path, ext_w_px, ext_h_px)
        xl_img = _ImageFromBytes(data)
        xl_img.width = ext_w_px
        xl_img.height = ext_h_px
        anchor = OneCellAnchor()
        anchor._from = AnchorMarker(
            col=min_col - 1, row=min_row - 1,
            colOff=col_off_emu, rowOff=row_off_emu,
        )
        anchor.ext = XDRPositiveSize2D(cx=ext_w_emu, cy=ext_h_emu)
        xl_img.anchor = anchor
        ws.add_image(xl_img)

def _copiar_alturas_linhas(ws, src_start: int, num_linhas: int, dst_start: int):
    for offset in range(num_linhas):
        dim = ws.row_dimensions.get(src_start + offset)
        if dim is not None and dim.height is not None:
            ws.row_dimensions[dst_start + offset].height = dim.height

def _replicar_merged_cells(ws, row_ini_src: int, row_fim_src: int, row_ini_dst: int):
    desloc = row_ini_dst - row_ini_src
    ranges_a_replicar = []
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= row_ini_src and mc.max_row <= row_fim_src:
            ranges_a_replicar.append((mc.min_row, mc.max_row, mc.min_col, mc.max_col))
    for min_r, max_r, min_c, max_c in ranges_a_replicar:
        try:
            ws.merge_cells(
                start_row=min_r + desloc, start_column=min_c,
                end_row=max_r + desloc,   end_column=max_c,
            )
        except Exception:
            pass

# LEITURA DAS NCs

def _detectar_col_data_reparo(ws, fallback: int = _DR) -> int:
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=4, column=col).value
        if val and "reparo" in str(val).lower():
            return col
    return fallback

def _cell(ws, row, col):
    return ws.cell(row=row, column=col).value


def _texto_sem_quebras(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if not s:
        return ""
    s = s.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    return re.sub(r" +", " ", s).strip()


def _nc_sanear_texto_sem_quebras(nc: dict) -> None:
    for k in (
        "codigo",
        "hora",
        "rod_raw",
        "rod_codigo",
        "rod_tag",
        "sentido",
        "tipo_nc",
        "atividade",
        "mae_atividade_q",
        "mae_tipo_atividade_o",
        "km_i",
        "km_f",
        "km_i_virg",
        "km_f_virg",
        "km_i_virgula",
        "km_f_virgula",
        "responsavel",
    ):
        if k not in nc or nc[k] is None:
            continue
        if isinstance(nc[k], str):
            nc[k] = _texto_sem_quebras(nc[k])
    fid = nc.get("foto_id")
    if isinstance(fid, str):
        nc["foto_id"] = _texto_sem_quebras(fid)


def _norm_header(s: str) -> str:
    t = unicodedata.normalize("NFD", str(s or ""))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", t).strip().lower()


def _colunas_por_header(ws) -> dict[str, int]:
    cols: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        k = _norm_header(v)
        if k and k not in cols:
            cols[k] = c
    return cols


def _col_responsavel_desde_header(cols_hdr: dict[str, int]) -> int | None:
    for label in (
        "responsavel tecnico",
        "responsavel tecnico fiscal",
        "responsavel",
        "responsavel fiscal",
        "fiscal responsavel",
    ):
        k = _norm_header(label)
        if k in cols_hdr:
            return cols_hdr[k]
    return None


def _detectar_coluna_atividade_mae_so_m02(ws) -> int | None:
    """
    Deteta coluna «Atividade» só pelo cabeçalho (sem fallback de coluna fixa).
    Exclusivo do M02 / nome Pendentes — não altera ``separar_nc._detectar_col_tipo_nc``.
    """
    melhor_c = None
    melhor_score = -1

    def score_header(h: str) -> int:
        if "tipo" in h and "atividade" in h and "tipo de atividade" in h:
            return 1
        if "grupo" in h and "atividade" in h:
            return 1
        if h == "atividade":
            return 6
        if h.endswith("atividade") and "tipo" not in h and "grupo" not in h:
            return 5
        if h == "evento":
            return 3
        if "evento" in h:
            return 2
        if "tipo nc" in h:
            return 2
        if "servico" in h or "serviço" in h:
            return 2
        if "atividade" in h:
            return 2
        return 0

    for r in range(1, min(ws.max_row + 1, 8) + 1):
        for c in range(1, min(ws.max_column + 1, 60) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            h = _norm_header(str(v))
            if not h:
                continue
            sc = score_header(h)
            if sc > melhor_score:
                melhor_score = sc
                melhor_c = c

    return melhor_c if melhor_c is not None and melhor_score >= 3 else None


def _dedupe_int_cols(seq: list[int]) -> list[int]:
    out: list[int] = []
    seen: set[int] = set()
    for c in seq:
        if isinstance(c, int) and c > 0 and c not in seen:
            seen.add(c)
            out.append(c)
    return out


def _valor_atividade_mae_plausivel(s: str) -> bool:
    t = _texto_sem_quebras(str(s or "")).strip()
    if len(t) < 5:
        return False
    tl = t.casefold()
    if tl in (
        "artesp",
        "kartado",
        "origem",
        "soluciona",
        "pendente",
        "sim",
        "nao",
        "não",
        "s/n",
    ):
        return False
    return True


def _colunas_candidatas_atividade_mae(ws, cols_hdr: dict[str, int]) -> list[int]:
    cols: list[int] = []
    col_n = _detectar_col_atividade_mae_col_n(ws)
    if col_n:
        cols.append(col_n)
    c_det = _detectar_coluna_atividade_mae_so_m02(ws)
    if c_det is not None:
        cols.append(c_det)
    for lbl in (RESPOSTA_PENDENTES_HEADER_MAE_ATIVIDADE_Q, "Atividade"):
        k = _norm_header(lbl)
        if k in cols_hdr:
            cols.append(cols_hdr[k])
    col_tipo = _detectar_col_tipo_nc(ws)
    if col_tipo:
        cols.append(col_tipo)
    cols.append(_Q)
    return _dedupe_int_cols(cols)


def _primeiro_valor_atividade_mae_linha(ws, row: int, colunas: list[int]) -> str:
    for c in colunas:
        v = _texto_sem_quebras(str(_cell(ws, row, c) or ""))
        if _valor_atividade_mae_plausivel(v):
            return v
    return ""


def _extrair_ref_foto_de_nome(nome_foto: str) -> str:
    s = str(nome_foto or "").strip()
    if not s:
        return ""
    m = re.search(r"\(([^)]+)\)", s)
    return (m.group(1).strip() if m else s)


def _parse_km_para_partes(v) -> tuple[object, str]:
    s = str(v or "").strip()
    if not s:
        return "", ""
    if "+" in s:
        a, b = s.split("+", 1)
        return a.strip(), b.strip()
    return s, ""

def _detectar_linha_inicio_dados(ws, max_busca: int = 15) -> int:
    """Detecta a primeira linha com código na coluna C (não é cabeçalho). Fallback 5."""
    for r in range(1, min(max_busca + 1, ws.max_row + 1)):
        val = _cell(ws, r, _C)
        if val is None:
            continue
        s = str(val).strip()
        if not s:
            continue
        # Pular linhas de cabeçalho
        if s.upper().startswith("CÓDIGO") or "fiscalização" in s.lower() or s.lower() == "cod":
            continue
        if _celula_parece_cabecalho_ou_coordenada(s):
            continue
        return r
    return 5


def _atividade_para_nome_resposta_pendentes(
    descricao_txt: str, mae_atividade_q: str, tipo_nc_classe: str
) -> str:
    """Alinha ao e-mail (.eml): Kartado usa trecho antes de ``-->`` em «Descrição»; senão Atividade / Classe."""
    d = _texto_sem_quebras(descricao_txt or "")
    if d:
        if "-->" in d:
            ativ = d.split("-->", 1)[0].strip()
        else:
            ln = d.splitlines()
            ativ = (ln[0].strip() if ln else "").strip()
    else:
        ativ = ""
    if not ativ:
        ativ = (mae_atividade_q or tipo_nc_classe or "").strip()
    return _texto_sem_quebras(ativ)


def _ler_ncs(ws, linha_inicio: int = 5) -> list:
    cols_hdr = _colunas_por_header(ws)
    col_resp_hdr = _col_responsavel_desde_header(cols_hdr)
    _h_mae_o = _norm_header(RESPOSTA_PENDENTES_HEADER_MAE_TIPO_ATIV_O)
    col_cod_kart = cols_hdr.get("codigo de fiscalizacao") or cols_hdr.get("codigo fiscalizacao")
    is_layout_kartado = col_cod_kart is not None

    if is_layout_kartado:
        linha_inicio = 2
        col_data_reparo = cols_hdr.get("prazo")
    else:
        linha_inicio = _detectar_linha_inicio_dados(ws) if linha_inicio == 5 else linha_inicio
        col_data_reparo = _detectar_col_data_reparo(ws, fallback=_DR)
    col_tip_o_detectada = detectar_coluna_tipo_de_atividade_eaf(ws, _O) if not is_layout_kartado else None
    colunas_atividade_mae = _colunas_candidatas_atividade_mae(ws, cols_hdr)
    ultima = ws.max_row
    for r in range(ultima, linha_inicio - 1, -1):
        if ws.cell(row=r, column=_D).value:
            ultima = r
            break

    ncs = []
    for r in range(linha_inicio, ultima + 1):
        if is_layout_kartado:
            cod = _texto_sem_quebras(str(_cell(ws, r, col_cod_kart) or ""))
            data_con = parse_data(_cell(ws, r, cols_hdr.get("encontrado em") or 0))
            hora = ""
            rod_raw = _texto_sem_quebras(str(_cell(ws, r, cols_hdr.get("rodovia") or 0) or "").replace("-", " "))
            km_i_src = _cell(ws, r, cols_hdr.get("km") or 0)
            km_f_src = _cell(ws, r, cols_hdr.get("km final") or 0)
            km_i_int, km_i_met = _parse_km_para_partes(km_i_src)
            km_f_int, km_f_met = _parse_km_para_partes(km_f_src)
            sentido = _texto_sem_quebras(str(_cell(ws, r, cols_hdr.get("sentido") or 0) or ""))
            tipo_nc = _texto_sem_quebras(str(_cell(ws, r, cols_hdr.get("classe") or 0) or ""))
            c_mae_o = cols_hdr.get(_h_mae_o)
            mae_atividade_q = _primeiro_valor_atividade_mae_linha(ws, r, colunas_atividade_mae)
            mae_tipo_atividade_o = _texto_sem_quebras(str(_cell(ws, r, c_mae_o) or "")) if c_mae_o else ""
            col_desc = cols_hdr.get("descricao")
            desc_raw = str(_cell(ws, r, col_desc) or "") if col_desc else ""
            atividade = _atividade_para_nome_resposta_pendentes(desc_raw, mae_atividade_q, tipo_nc)
            foto_1_nome = _texto_sem_quebras(str(_cell(ws, r, cols_hdr.get("foto_1") or 0) or ""))
            num_foto = _extrair_ref_foto_de_nome(foto_1_nome) or cod
            data_rep = parse_data(_cell(ws, r, col_data_reparo or 0))
            responsavel = (
                _texto_sem_quebras(str(_cell(ws, r, col_resp_hdr) or ""))
                if col_resp_hdr
                else ""
            )
        else:
            cod = _texto_sem_quebras(str(_cell(ws, r, _C) or ""))
            data_con = parse_data(_cell(ws, r, _D))
            hora = _texto_sem_quebras(str(_cell(ws, r, _E) or ""))
            rod_raw = _texto_sem_quebras(str(_cell(ws, r, _F) or ""))
            km_i_int = _cell(ws, r, _H)
            km_i_met = _texto_sem_quebras(str(_cell(ws, r, _I) or ""))
            km_f_int = _cell(ws, r, _J)
            km_f_met = _texto_sem_quebras(str(_cell(ws, r, _K) or ""))
            sentido = _texto_sem_quebras(str(_cell(ws, r, _L) or ""))
            mae_atividade_q = _primeiro_valor_atividade_mae_linha(ws, r, colunas_atividade_mae)
            tipo_nc = mae_atividade_q or _texto_sem_quebras(str(_cell(ws, r, _Q) or ""))
            mae_tipo_atividade_o = _texto_sem_quebras(
                str(_cell(ws, r, col_tip_o_detectada or _O) or "")
            )
            atividade = _atividade_para_nome_resposta_pendentes("", mae_atividade_q, tipo_nc)
            num_foto = _cell(ws, r, _V)
            data_rep = parse_data(_cell(ws, r, col_data_reparo))
            responsavel = (
                _texto_sem_quebras(str(_cell(ws, r, col_resp_hdr) or ""))
                if col_resp_hdr
                else ""
            )

        if not cod:
            continue
        if _celula_parece_cabecalho_ou_coordenada(cod):
            continue

        rod_info = normalizar_rodovia_eaf(rod_raw, RODOVIAS)
        try:
            num_foto_val = int(num_foto) if num_foto is not None and str(num_foto).strip() else 0
        except (TypeError, ValueError):
            num_foto_val = 0
        # Todas as fotos seguem identificação por código (col C), conforme fluxo histórico.
        # Fallback para col V apenas quando o código estiver inválido/ausente.
        cod_ok = (cod and str(cod).strip() and not str(cod).strip().upper().startswith("LOTE"))
        if cod_ok and not _celula_parece_cabecalho_ou_coordenada(cod):
            foto_id = cod
        else:
            foto_id = num_foto_val
        # Prazo em dias = data vencimento (reparo) − data constatação (para coluna L do Kria)
        prazo_dias = None
        if data_con is not None and data_rep is not None:
            try:
                prazo_dias = (data_rep - data_con).days
            except (TypeError, AttributeError):
                pass

        nc_row = {
            "codigo":      cod,
            "data_con":    data_con,
            "hora":        hora,
            "rod_raw":     rod_raw,
            "rod_codigo":  rod_info["codigo"],
            "rod_tag":     rod_info["tag"],
            "rod_n":       rod_info["n"],
            "km_i":        km_mais_metros(km_i_int, km_i_met),
            "km_i_virg":   km_virgula_metros(km_i_int, km_i_met),
            "km_f":        km_mais_metros(km_f_int, km_f_met),
            "km_f_virg":   km_virgula_metros(km_f_int, km_f_met),
            "sentido":     sentido,
            "tipo_nc":     tipo_nc,
            "atividade":   atividade,
            "mae_atividade_q": mae_atividade_q,
            "mae_tipo_atividade_o": mae_tipo_atividade_o,
            "num_foto":    num_foto_val,
            "foto_id":     foto_id,
            "prazo_dias":  prazo_dias,
            "data_reparo": data_rep,
            "layout_kartado": bool(is_layout_kartado),
            "responsavel": responsavel,
        }
        _nc_sanear_texto_sem_quebras(nc_row)
        ncs.append(nc_row)

    return ncs


def listar_imagens_referenciadas_por_ncs(
    ncs: list,
    pasta_fotos_nc: Path | None,
    pasta_fotos_pdf: Path | None,
) -> list[Path]:
    """
    Caminhos existentes de imagens usadas no modelo Kria (nc/PDF) para a lista de NCs.
    Usado para empacotar Excel + fotos num ZIP (entrega Kartado).
    """
    pn = pasta_fotos_nc if pasta_fotos_nc and pasta_fotos_nc.is_dir() else None
    pp = pasta_fotos_pdf if pasta_fotos_pdf and pasta_fotos_pdf.is_dir() else None
    paths: list[Path] = []
    for nc in ncs:
        if pn:
            for cand in _candidatos_identificador_foto(nc):
                p = path_foto_nc(pn, cand)
                if p.is_file():
                    paths.append(p)
                    break
            p2 = path_foto_nc_segunda(pn, nc["codigo"])
            if p2.is_file():
                paths.append(p2)
        if pp:
            for cand in _candidatos_identificador_foto(nc):
                p3 = path_foto_pdf(pp, cand)
                if p3.is_file():
                    paths.append(p3)
                    break
    seen: set[str] = set()
    out: list[Path] = []
    for p in paths:
        try:
            k = str(p.resolve())
        except OSError:
            k = str(p)
        if k not in seen:
            seen.add(k)
            out.append(p)
    return out


def _kartado_art03_zip_stem(ncs: list, arquivo_stem: str) -> str:
    """
    Stem do .zip alinhado à macro Kartado Art_03_KTD_Inserir_NC_Rot22:
    ``ArquivoZip = … & serv(1) & ".zip"``.

    - Entrada **layout Kartado** (cabeçalho linha 1): ``tipo_nc`` em ``ncs`` vem da coluna «Classe»
      (texto tipo «Pav. - Panela_Buraco …») — equivalente ao serv(1) do Kria no Kartado.
    - Entrada **layout EAF** (Art_022): ``tipo_nc`` é o texto da coluna Q da mãe; o nome do ficheiro
      costuma seguir o padrão Art_011 ``… (RODOVIA - ABREV) - Prazo - …``; nesse caso usa-se o fallback
      pelo parêntesis no ``arquivo_stem``.
    """
    if ncs:
        t = (ncs[0].get("tipo_nc") or "").strip()
        if t:
            return sanitizar_nome(t, max_len=200).strip() or "pacote"
    stem = (arquivo_stem or "").strip()
    if stem:
        m = re.search(r"\(([^)]+)\)", stem)
        if m:
            interior = m.group(1).strip()
            if " - " in interior:
                fb = interior.rsplit(" - ", 1)[-1].strip()
                if fb:
                    return sanitizar_nome(fb, max_len=200).strip() or "pacote"
    s = sanitizar_nome(stem, max_len=200).strip() if stem else ""
    return s or "pacote"


def _alinhar_texto_sem_quebra(ws, row: int, col: int, *, kartado: bool = False) -> None:
    if kartado:
        ws.cell(row=row, column=col).alignment = Alignment(
            wrap_text=False,
            horizontal="center",
            vertical="center",
            shrink_to_fit=False,
        )
    else:
        ws.cell(row=row, column=col).alignment = Alignment(
            wrap_text=False,
            vertical="top",
            horizontal="left",
            shrink_to_fit=False,
        )


def _alinhar_kria_bloco_texto_sem_quebra(ws, j: int, *, kartado: bool = False) -> None:
    for r, co in (
        (j - 2, 2),
        (j - 2, 3),
        (j - 1, 7),
        (j, 4),
        (j, 6),
        (j, 7),
        (j + 1, 4),
        (j + 1, 6),
        (j + 1, 8),
        (j + 1, 12),
        (j + 1, 22),
        (j + 1, 23),
        (j + 2, 3),
        (j + 2, 4),
        (j + 2, 6),
        (j + 2, 8),
        (j + 2, 12),
    ):
        _alinhar_texto_sem_quebra(ws, r, co, kartado=kartado)


def _celula_topo_esquerdo_merge(ws, row: int, col: int):
    for m in ws.merged_cells.ranges:
        if m.min_row <= row <= m.max_row and m.min_col <= col <= m.max_col:
            return ws.cell(row=m.min_row, column=m.min_col)
    return ws.cell(row=row, column=col)


def _definir_valor_preservando_estilo(ws, row: int, col: int, value) -> None:
    c = _celula_topo_esquerdo_merge(ws, row, col)
    saved: dict = {}
    if getattr(c, "has_style", False) and c.has_style:
        for name in ("font", "border", "fill", "alignment", "protection"):
            obj = getattr(c, name, None)
            if obj is not None:
                saved[name] = copy(obj)
        saved["number_format"] = c.number_format
    c.value = value
    for k, v in saved.items():
        if k == "number_format":
            c.number_format = v
        else:
            setattr(c, k, v)


def _preencher_legenda_bloco_resposta(ws, bloco_top: int) -> None:
    r_leg = bloco_top + _ROW_LEGENDA_FOTO_RESP - 1
    r_de = bloco_top + _ROW_DATA_EXEC_RESP - 1
    _definir_valor_preservando_estilo(ws, r_leg, 2, _TEXTO_LEGENDA_FOTO_RESP)
    _definir_valor_preservando_estilo(ws, r_de, 2, None)


def _preencher_colunas_v_w_kcor_resposta_bloco(
    ws,
    bloco_top: int,
    nc: dict,
    pasta_fotos_pdf: Path,
) -> None:
    """Col. V (22) = texto diretório macro Kcor-Kria; W (23) = ``nc.jpg;pdf (ref).jpg`` — não coluna U (fiscal)."""
    r_vw = bloco_top + _OFFSET_ROW_VW_KCOR_DENTRO_BLOCO_PENDENTES
    _definir_valor_preservando_estilo(ws, r_vw, 22, KCOR_KRIA_DIRETORIO_TEXTO_CONSERVACAO)
    foto_ref = nc.get("foto_id", nc.get("num_foto"))
    pdf_part = (
        f"pdf ({foto_ref}).jpg"
        if foto_ref not in (None, "", 0)
        else ""
    )
    foto_pdf = Path()
    for cand in _candidatos_identificador_foto(nc):
        foto_pdf = path_foto_pdf(pasta_fotos_pdf, cand)
        if foto_pdf.is_file():
            break
    if not foto_pdf.is_file():
        foto_pdf = path_foto_pdf(pasta_fotos_pdf, nc.get("foto_id", nc.get("num_foto")))
    arq_part = foto_pdf.name if foto_pdf.is_file() else ""
    if arq_part and pdf_part:
        w_val = f"{arq_part};{pdf_part}"
    else:
        w_val = arq_part or pdf_part
    _definir_valor_preservando_estilo(ws, r_vw, 23, w_val or "")


def _linha_acima_quadro_resposta_pendentes(nc: dict) -> str:
    rod_raw = str(nc.get("rod_codigo") or nc.get("rod_tag") or "").strip()
    rod_c = re.sub(r"\s+", "", rod_raw) or "—"
    km_v = str(nc.get("km_i_virg") or nc.get("km_i_virgula") or "").strip()
    if not km_v:
        ki = str(nc.get("km_i") or "").strip()
        km_v = ki.replace(" + ", ",").replace("+", ",").strip() or "—"
    sent = str(nc.get("sentido") or "").strip() or "—"
    dc = nc.get("data_con")
    dr = nc.get("data_reparo")
    dc_s = data_br(dc) if dc else "—"
    dr_s = data_br(dr) if dr else "—"
    tipo = _texto_atividade_exibicao_pendentes(nc) or "—"
    cod = str(nc.get("codigo") or "").strip() or "—"
    return (
        f"{rod_c} - km {km_v} {sent} - Const: {dc_s} - Prazo: {dr_s} - {tipo} - Cod. Fisc.: {cod}"
    )


# SAÍDA A – Planilha Kria  →  foto: nc (N).jpg

def _gerar_kria(
    ncs: list, nome_base: str,
    modelo: "Path | bytes", pasta_saida: Path,
    pasta_fotos_nc: Path,          # ← fonte principal: nc (N).jpg
    relatorio: str,
    pasta_fotos_pdf_fallback: "Path | None" = None,  # ← fallback (ZIP Extrair PDF tem nc + PDF na mesma pasta)
) -> "Path | None":
    """
    Preenche o modelo Kria com as NCs lidas e insere a foto de campo.
    modelo: Path do .xlsx ou bytes já carregados (reutilização mais rápida).

    Regra de foto (SAÍDA A):
        Usa nc (N).jpg em pasta_fotos_nc; se não existir, tenta pasta_fotos_pdf_fallback
        (ZIP do Extrair PDF coloca nc (CODIGO).jpg e PDF (CODIGO).jpg na mesma pasta).

    Estrutura do bloco (5 linhas, âncora j=8 para a 1ª NC):
      j-2: B=seq,  C=tipo_nc
      j-1: C=âncora foto nc (275×210),  G=data_envio (data constatação) — M03 lê como embasamento e grava na col Data Envio do Kcor-Kria
      j:   D=rodovia,  F=sentido,  G=tipo_nc
      j+1: D=km_i,  F=km_f,  H=codigo,  L=num_foto
      j+2: C="Vencimento",  D=data_reparo,  F=data_con,  H=relatorio,  L=prazo
      j+1: V=KCOR_KRIA_DIRETORIO_TEXTO_CONSERVACAO (rede, macro Art_03), W=«jpg»;«pdf (ref).jpg»
    """
    if isinstance(modelo, Path):
        if not Path(str_caminho_io_windows(modelo)).exists():
            logger.error(f"Modelo Kria não encontrado: {modelo}")
            return None
        modelo_src = str_caminho_io_windows(modelo)
    else:
        modelo_src = BytesIO(modelo)

    garantir_pasta(pasta_saida)
    nc_nome = ncs[0] if ncs else {}
    prefixo_venc = _prefixo_nome_por_vencimento(
        nc_nome.get("data_reparo") if isinstance(nc_nome, dict) else None
    )
    rod_nome = sanitizar_nome(
        str(
            (nc_nome.get("rod_codigo") if isinstance(nc_nome, dict) else "")
            or (nc_nome.get("rod_tag") if isinstance(nc_nome, dict) else "")
            or "sem-rodovia"
        ),
        max_len=96,
    ) or "sem-rodovia"
    tipo_nome = sanitizar_nome(
        str(
            (
                _texto_atividade_exibicao_pendentes(nc_nome)
                if isinstance(nc_nome, dict)
                else ""
            )
            or "sem-tipo"
        ),
        max_len=140,
    ) or "sem-tipo"
    nome_arq = f"{prefixo_venc} - {rod_nome} - {tipo_nome}.xlsx"
    destino  = caminho_dentro_limite_windows(pasta_saida / nome_arq)
    garantir_pasta(destino.parent)

    wb = load_workbook(modelo_src)
    ws = wb.active
    patch_add_image(ws)

    n_ncs = len(ncs)
    if n_ncs == 0:
        logger.warning("Nenhuma NC para inserir na planilha Kria.")
        wb.save(str_caminho_io_windows(destino))
        wb.close()
        preservar_ooxml_planilha_pos_openpyxl(modelo, destino)
        return destino

    kartado_layout = bool(ncs[0].get("layout_kartado"))

    J_INICIO  = 8               # âncora da 1ª NC
    SRC_START = J_INICIO - 2   # linha 6 = início do bloco modelo

    # Expandir blocos extras para NC 2 em diante
    for extra in range(1, n_ncs):
        dst_start = SRC_START + extra * BLOCO
        ws.insert_rows(dst_start, BLOCO)
        for offset in range(BLOCO):
            for col in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=SRC_START + offset, column=col)
                dst_cell = ws.cell(row=dst_start  + offset, column=col)
                dst_cell.value = src_cell.value
                if src_cell.has_style:
                    dst_cell.font          = copy(src_cell.font)
                    dst_cell.border        = copy(src_cell.border)
                    dst_cell.fill          = copy(src_cell.fill)
                    dst_cell.number_format = src_cell.number_format
                    dst_cell.alignment     = copy(src_cell.alignment)
        _copiar_alturas_linhas(ws, SRC_START, BLOCO, dst_start)
        _replicar_merged_cells(ws, SRC_START, SRC_START + BLOCO - 1, dst_start)

    # Preencher cada NC em ordem direta (NC[0] = topo)
    for idx, nc in enumerate(ncs):
        j  = J_INICIO + idx * BLOCO
        dr = nc["data_reparo"]
        dc = nc["data_con"]
        # Data do reparo = data do envio + 10 dias quando não informada
        if not dr and dc:
            dr = dc + timedelta(days=PRAZO_DIAS_APOS_ENVIO)

        # Coluna L (12) linha j+2 (ex.: 10): dias de prazo = vencimento − data constatação (ex.: "7 dias")
        prazo_dias = nc.get("prazo_dias")
        if prazo_dias is None and dr is not None and dc is not None:
            try:
                prazo_dias = (dr - dc).days
            except (TypeError, AttributeError):
                pass
        prazo_l = f"{prazo_dias} dias" if prazo_dias is not None else ""

        ws.cell(row=j - 2, column=2).value  = idx + 1
        ws.cell(row=j - 2, column=3).value  = _texto_atividade_exibicao_pendentes(nc)
        ws.cell(row=j - 1, column=7).value  = data_br(dc) if dc else ""  # G = data envio (constatação); M03 usa para col Data Envio do Kcor-Kria
        ws.cell(row=j,     column=4).value  = nc["rod_codigo"]
        ws.cell(row=j,     column=6).value  = nc["sentido"]
        ws.cell(row=j,     column=7).value  = _texto_atividade_exibicao_pendentes(nc)
        ws.cell(row=j + 1, column=4).value  = nc["km_i"]
        ws.cell(row=j + 1, column=6).value  = nc["km_f"]
        ws.cell(row=j + 1, column=8).value  = nc["codigo"]
        ws.cell(row=j + 1, column=12).value = nc.get("foto_id", nc["num_foto"])
        ws.cell(row=j + 2, column=3).value  = "Vencimento"
        ws.cell(row=j + 2, column=4).value  = data_br(dr) if dr else ""
        ws.cell(row=j + 2, column=6).value  = data_br(dc) if dc else ""
        ws.cell(row=j + 2, column=8).value  = relatorio
        ws.cell(row=j + 2, column=12).value = prazo_l

        # ── Foto: tenta foto_id, depois col V, depois col C (extrator pode nomear por nº ou por código) ──
        cell_foto = f"C{j - 1}"
        foto_path = Path()
        for cand in _candidatos_identificador_foto(nc):
            foto_path = path_foto_nc(pasta_fotos_nc, cand)
            if foto_path.is_file():
                break
            if pasta_fotos_pdf_fallback and pasta_fotos_pdf_fallback.is_dir():
                foto_path = path_foto_nc(pasta_fotos_pdf_fallback, cand)
                if foto_path.is_file():
                    break
        if not foto_path.is_file():
            foto_path = path_foto_nc(pasta_fotos_nc, nc.get("foto_id", nc["num_foto"]))
        foto2_path = path_foto_nc_segunda(pasta_fotos_nc, nc["codigo"])
        if not foto2_path.is_file() and pasta_fotos_pdf_fallback and pasta_fotos_pdf_fallback.is_dir():
            foto2_path = path_foto_nc_segunda(pasta_fotos_pdf_fallback, nc["codigo"])
        if foto_path.is_file() and foto2_path.is_file():
            _inserir_duas_imagens_ma(ws, cell_foto, foto_path, foto2_path, M02_FOTO_W, M02_FOTO_H)
            logger.debug(f"  [Kria] Duas fotos nc inseridas (MA): {foto_path.name} + {foto2_path.name} → {cell_foto}")
        elif foto_path.is_file():
            _inserir_imagem(ws, cell_foto, foto_path, M02_FOTO_W, M02_FOTO_H)
            logger.debug(f"  [Kria] Foto nc inserida: {foto_path.name} → {cell_foto}")
        else:
            logger.warning(
                "  [Kria] Foto não encontrada (tentado %s): nc/PDF por identificador",
                ", ".join(repr(c) for c in _candidatos_identificador_foto(nc)) or "(vazio)",
            )

        # V/W — padrão Kcor-Kria / Art_03 (Diretório + Arquivos), linha dos km/código/foto
        ws.cell(row=j + 1, column=22).value = KCOR_KRIA_DIRETORIO_TEXTO_CONSERVACAO
        foto_ref_w = nc.get("foto_id", nc.get("num_foto"))
        pdf_part = (
            f"pdf ({foto_ref_w}).jpg"
            if foto_ref_w not in (None, "", 0)
            else ""
        )
        arq_part = foto_path.name if foto_path.is_file() else ""
        if arq_part and pdf_part:
            w_val = f"{arq_part};{pdf_part}"
        else:
            w_val = arq_part or pdf_part
        ws.cell(row=j + 1, column=23).value = w_val

        _alinhar_kria_bloco_texto_sem_quebra(ws, j, kartado=kartado_layout)

    wb.save(str_caminho_io_windows(destino))
    wb.close()
    preservar_ooxml_planilha_pos_openpyxl(modelo, destino)
    logger.info(f"Saída A (Kria) salva: {destino.name}")
    return destino


# SAÍDA B – Relatório de Resposta à Artesp  →  foto: PDF (N).jpg

def _tipo_nc_para_chave_agrupamento(nc: dict) -> str:
    for key in ("atividade", "mae_atividade_q", "tipo_nc"):
        t = _texto_sem_quebras(str(nc.get(key) or "").strip())
        if t and parse_data(t) is None:
            return t
    c = _texto_sem_quebras(str(nc.get("codigo") or ""))
    return c or "NC"


def _grupos_ncs_resposta_alinhados_exportar(ncs: list[dict]) -> list[list[dict]]:
    ordem: list[tuple[str, ...]] = []
    buckets: dict[tuple[str, ...], list[dict]] = {}
    for nc in ncs:
        tipo_k = _tipo_nc_para_chave_agrupamento(nc)
        k = chave_agrupamento_exportar_rotina_valores(
            tipo_k,
            nc.get("data_reparo"),
            nc.get("rod_raw"),
            nc.get("responsavel"),
            codigo_fallback=None,
        )
        if k not in buckets:
            ordem.append(k)
            buckets[k] = []
        buckets[k].append(nc)
    return [buckets[key] for key in ordem]


def _alocar_nome_resposta_pendentes_unico(
    nome_preferido: str,
    ocupados: set[str],
) -> str:
    nome_s = _sanitizar_nome_xlsx(nome_preferido) if nome_preferido else ""
    if not nome_s.lower().endswith(".xlsx"):
        nome_s = _sanitizar_nome_xlsx(f"{nome_s}.xlsx")
    cand = nome_s
    stem = Path(cand).stem
    n = 2
    while cand in ocupados:
        cand = _sanitizar_nome_xlsx(f"{stem} - {n}.xlsx")
        n += 1
    ocupados.add(cand)
    return cand


def _gerar_resposta(
    ncs: list, modelo: "Path | bytes",
    pasta_saida: Path,
    pasta_fotos_pdf: Path,
    nome_ficheiro_destino: str | None = None,
) -> "Path | None":
    """
    Preenche o modelo de resposta (28 linhas por NC) com cabeçalho e fotos.
    modelo: Path do .xlsx ou bytes já carregados (reutilização mais rápida).

    Regra de foto (SAÍDA B):
        Usa EXCLUSIVAMENTE  PDF (N).jpg  da pasta_fotos_pdf.
        NÃO utiliza foto nc aqui.

    Linha 1  (B1): linha compacta (rodovia, km, Const/Prazo, tipo, código)
    Linha 2  (B2): cabeçalho curto / âncora da foto PDF (tamanho ao merge do modelo)
    NC 2 em diante: duplica bloco 1→28 ABAIXO do anterior e repete o padrão.
    Texto em B: grava valor e repõe fonte/borda/preenchimento/alinhamento/formato/protecção já definidos no modelo.
    """
    if isinstance(modelo, Path):
        if not Path(str_caminho_io_windows(modelo)).exists():
            logger.error(f"Modelo de resposta não encontrado: {modelo}")
            return None
        modelo_src = str_caminho_io_windows(modelo)
    else:
        modelo_src = BytesIO(modelo)
    if not ncs:
        return None

    garantir_pasta(pasta_saida)

    nc1 = ncs[0]
    dr = nc1["data_reparo"]
    dc = nc1["data_con"]

    dr_str = dr.strftime("%d-%m-%Y") if dr else "00-00-0000"
    dc_str = dc.strftime("%d-%m-%Y") if dc else "00-00-0000"

    if nome_ficheiro_destino:
        nome_arq = nome_ficheiro_destino.strip()
        if not nome_arq.lower().endswith(".xlsx"):
            nome_arq = f"{nome_arq}.xlsx"
        nome_arq = _sanitizar_nome_xlsx(nome_arq)
    else:
        nome_arq = nome_ficheiro_resposta_artesp_xlsx(nc1)
    destino = caminho_dentro_limite_windows(pasta_saida / nome_arq)

    wb = load_workbook(modelo_src)
    ws = wb.active
    patch_add_image(ws)

    SRC_START = 1  # bloco modelo começa na linha 1

    def _cabecalho_curto(nc: dict, dr_: str, dc_: str) -> str:
        ativ_txt = _texto_atividade_exibicao_pendentes(nc)
        return (
            f"{dc_} - {nc['rod_tag']} - {nc['km_i']} - "
            f"{nc['sentido']} - {dr_} - {ativ_txt} - {nc['codigo']}"
        )

    _definir_valor_preservando_estilo(ws, 1, 2, _linha_acima_quadro_resposta_pendentes(nc1))
    _definir_valor_preservando_estilo(ws, 2, 2, _cabecalho_curto(nc1, dr_str, dc_str))

    # Foto PDF: tenta foto_id, col V, col C (nomes PDF (N).jpg ou PDF (código).jpg)
    foto1 = Path()
    for cand in _candidatos_identificador_foto(nc1):
        foto1 = path_foto_pdf(pasta_fotos_pdf, cand)
        if foto1.is_file():
            break
    if not foto1.is_file():
        foto1 = path_foto_pdf(pasta_fotos_pdf, nc1.get("foto_id", nc1["num_foto"]))
    if foto1.exists():
        _inserir_imagem(ws, "B2", foto1, M02_FOTO_PDF_W, M02_FOTO_PDF_H)
        logger.debug(f"  [Resposta] Foto PDF inserida: {foto1.name} → B2")
    else:
        logger.warning(
            "  [Resposta] Foto PDF não encontrada (tentado %s)",
            ", ".join(repr(c) for c in _candidatos_identificador_foto(nc1)) or "(vazio)",
        )
    _preencher_legenda_bloco_resposta(ws, 1)
    _preencher_colunas_v_w_kcor_resposta_bloco(ws, 1, nc1, pasta_fotos_pdf)

    linha = 1  # ponteiro para o início do bloco atual

    for nc in ncs[1:]:
        dst_start = linha + BLOCO_RESP  # sempre abaixo → ordem direta

        ws.insert_rows(dst_start, BLOCO_RESP)

        for offset in range(BLOCO_RESP):
            for col in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=SRC_START + offset, column=col)
                dst_cell = ws.cell(row=dst_start  + offset, column=col)
                dst_cell.value = src_cell.value
                # Copiar sempre o estilo completo evita herança de estilos das linhas
                # deslocadas pelo insert_rows, mantendo o bloco réplica idêntico ao base.
                dst_cell._style = copy(src_cell._style)

        _copiar_alturas_linhas(ws, SRC_START, BLOCO_RESP, dst_start)
        _replicar_merged_cells(ws, SRC_START, SRC_START + BLOCO_RESP - 1, dst_start)

        nc_dr    = nc["data_reparo"]
        nc_dc    = nc["data_con"]
        nc_dr_s  = nc_dr.strftime("%d-%m-%Y") if nc_dr else "00-00-0000"
        nc_dc_s  = nc_dc.strftime("%d-%m-%Y") if nc_dc else "00-00-0000"
        _definir_valor_preservando_estilo(ws, dst_start, 2, _linha_acima_quadro_resposta_pendentes(nc))
        _definir_valor_preservando_estilo(ws, dst_start + 1, 2, _cabecalho_curto(nc, nc_dr_s, nc_dc_s))

        foto_v = Path()
        for cand in _candidatos_identificador_foto(nc):
            foto_v = path_foto_pdf(pasta_fotos_pdf, cand)
            if foto_v.is_file():
                break
        if not foto_v.is_file():
            foto_v = path_foto_pdf(pasta_fotos_pdf, nc.get("foto_id", nc["num_foto"]))
        cell_foto = f"B{dst_start + 1}"
        if foto_v.exists():
            _inserir_imagem(ws, cell_foto, foto_v, M02_FOTO_PDF_W, M02_FOTO_PDF_H)
            logger.debug(f"  [Resposta] Foto PDF inserida: {foto_v.name} → {cell_foto}")
        else:
            logger.warning(
                "  [Resposta] Foto PDF não encontrada (tentado %s)",
                ", ".join(repr(c) for c in _candidatos_identificador_foto(nc)) or "(vazio)",
            )
        _preencher_legenda_bloco_resposta(ws, dst_start)
        _preencher_colunas_v_w_kcor_resposta_bloco(ws, dst_start, nc, pasta_fotos_pdf)

        linha = dst_start  # avança ponteiro

    wb.save(str_caminho_io_windows(destino))
    wb.close()
    preservar_ooxml_planilha_pos_openpyxl(modelo, destino)
    replicar_ancoras_sem_foto_por_bloco_em_drawing(
        destino, BLOCO_RESP, max(0, len(ncs) - 1)
    )
    logger.info(f"Saída B (Resposta) salva: {destino.name}")
    return destino


def _prefixo_nome_por_vencimento(valor_data) -> str:
    dt = parse_data(valor_data)
    hhmm = datetime.now().strftime("%H%M")
    if dt:
        return f"{dt.strftime('%d-%m-%Y')}-{hhmm}"
    return f"sem-vencimento-{hhmm}"


# FUNÇÃO PRINCIPAL

def executar(
    pasta_xls: "Path | None" = None,
    modelo_kria: "Path | None" = None,
    pasta_saida_kria: "Path | None" = None,
    modelo_resposta: "Path | None" = None,
    pasta_saida_resp: "Path | None" = None,
    pasta_fotos_nc: "Path | None" = None,
    pasta_fotos_pdf: "Path | None" = None,
    callback_progresso=None,
) -> dict:
    """
    Processa todos os XLS da pasta de entrada e gera:
      - Kria    (Saída A) com fotos  nc (N).jpg
      - Resposta (Saída B) com fotos  PDF (N).jpg

    Retorna dict: { 'kria': [...], 'resposta': [...], 'erros': [...],
                    'kartado_pacotes': [ { 'kria': Path, 'imagens': [Path,...], 'zip_stem': str }, ... ] }
    """
    pasta_xls        = pasta_xls        or M01_EXPORTAR
    modelo_kria      = resolver_path_ficheiro_ci(modelo_kria or M02_MODELO_KRIA)
    pasta_saida_kria = pasta_saida_kria or M02_SALVAR_FOTO
    modelo_resposta  = resolver_path_ficheiro_ci(modelo_resposta or M02_MODELO_RESP)
    pasta_saida_resp = pasta_saida_resp or M02_PENDENTES
    pasta_fotos_nc   = pasta_fotos_nc   or M02_FOTOS_NC    # nc (N).jpg
    pasta_fotos_pdf  = pasta_fotos_pdf  or M02_FOTOS_PDF   # PDF (N).jpg

    garantir_pasta(pasta_saida_kria)
    garantir_pasta(pasta_saida_resp)

    # Buscar em pasta_xls e em subpastas (ZIP pode ter Exportar/arquivo.xlsx)
    arquivos = sorted([
        f for f in pasta_xls.rglob("*.xls*")
        if f.is_file()
        and not f.name.startswith("~")
        and not f.name.startswith("_")
        and not ficheiro_em_subpasta_planilha_mae_kartado_exportar(f, pasta_xls)
    ])

    if not arquivos:
        logger.warning(f"Nenhum XLS encontrado em: {pasta_xls}")
        return {"kria": [], "resposta": [], "erros": [], "kartado_pacotes": []}

    logger.info(f"Módulo 02: {len(arquivos)} arquivo(s) para processar.")
    resultados: dict = {"kria": [], "resposta": [], "erros": [], "kartado_pacotes": []}
    nomes_resp_ocupados: set[str] = set()

    # Carregar modelos uma vez em memória (evita N leituras em disco — acelera muito)
    modelo_kria_bytes = None
    modelo_resp_bytes = None
    if modelo_kria and Path(modelo_kria).exists():
        modelo_kria_bytes = Path(modelo_kria).read_bytes()
    if modelo_resposta and Path(modelo_resposta).exists():
        modelo_resp_bytes = Path(modelo_resposta).read_bytes()

    for idx, arq in enumerate(arquivos):
        if callback_progresso:
            callback_progresso(idx + 1, len(arquivos), f"Processando: {arq.name[:60]}")

        logger.info(f"\n── Processando [{idx+1}/{len(arquivos)}]: {arq.name}")

        try:
            if arq.suffix.lower() == ".xls":
                path_para_abrir = xls_to_xlsx(arq, dest=None)
                remover_depois  = True
            else:
                path_para_abrir = arq
                remover_depois  = False

            wb = load_workbook(str(path_para_abrir), data_only=True)
            try:
                ws  = wb.active
                ncs = _ler_ncs(ws)
            finally:
                wb.close()
                if remover_depois and path_para_abrir.exists():
                    path_para_abrir.unlink(missing_ok=True)

            if not ncs:
                logger.warning(f"  Nenhuma NC encontrada em {arq.name}, pulando.")
                continue

            logger.info(f"  {len(ncs)} NC(s) lida(s).")

            nome_base = arq.stem
            relatorio = nome_base[:8]

            modelo_kria_eff = modelo_kria_bytes if modelo_kria_bytes is not None else modelo_kria
            arq_kria = _gerar_kria(
                ncs, nome_base,
                modelo_kria_eff, pasta_saida_kria,
                pasta_fotos_nc,
                relatorio,
                pasta_fotos_pdf_fallback=pasta_fotos_pdf,
            )
            if arq_kria:
                resultados["kria"].append(arq_kria)
                imgs = listar_imagens_referenciadas_por_ncs(
                    ncs, pasta_fotos_nc, pasta_fotos_pdf
                )
                zip_stem = _kartado_art03_zip_stem(ncs, arq.stem)
                resultados["kartado_pacotes"].append(
                    {"kria": arq_kria, "imagens": imgs, "zip_stem": zip_stem}
                )

            modelo_resp_eff = modelo_resp_bytes if modelo_resp_bytes is not None else modelo_resposta
            for grupo in _grupos_ncs_resposta_alinhados_exportar(ncs):
                nc0 = grupo[0]
                nome_pref = nome_ficheiro_resposta_artesp_xlsx(nc0)
                nome_final = _alocar_nome_resposta_pendentes_unico(
                    nome_pref,
                    nomes_resp_ocupados,
                )
                arq_resp = _gerar_resposta(
                    grupo,
                    modelo_resp_eff,
                    pasta_saida_resp,
                    pasta_fotos_pdf,
                    nome_ficheiro_destino=nome_final,
                )
                if arq_resp:
                    resultados["resposta"].append(arq_resp)

        except Exception as e:
            logger.error(f"  ERRO em {arq.name}: {e}", exc_info=True)
            resultados["erros"].append(arq.name)

    total_k = len(resultados["kria"])
    total_r = len(resultados["resposta"])
    total_e = len(resultados["erros"])
    logger.info(
        f"\nMódulo 02 concluído: "
        f"{total_k} Kria, {total_r} resposta(s), {total_e} erro(s)."
    )
    if callback_progresso:
        callback_progresso(len(arquivos), len(arquivos), "Módulo 02 concluído.")

    return resultados


def executar_kria_resposta_de_lista(
    ncs: list,
    nome_base: str,
    relatorio: str,
    modelo_kria: "Path | None" = None,
    pasta_saida_kria: "Path | None" = None,
    modelo_resposta: "Path | None" = None,
    pasta_saida_resp: "Path | None" = None,
    pasta_fotos_nc: "Path | None" = None,
    pasta_fotos_pdf: "Path | None" = None,
) -> dict:
    """
    Gera Kria (Saída A) e Resposta (Saída B) a partir de uma lista de NCs
    (dict com codigo, data_con, data_reparo, tipo_nc, rod_codigo, rod_tag,
    sentido, km_i, km_f, num_foto, prazo_dias). Usado pelo pipeline Meio Ambiente
    (equivalente M2 a partir do PDF, sem planilha EAF).
    Retorna {"kria": Path | None, "resposta": list[Path]}.
    """
    modelo_kria = resolver_path_ficheiro_ci(modelo_kria or M02_MODELO_KRIA)
    pasta_saida_kria = pasta_saida_kria or M02_SALVAR_FOTO
    modelo_resposta = resolver_path_ficheiro_ci(modelo_resposta or M02_MODELO_RESP)
    pasta_saida_resp = pasta_saida_resp or M02_PENDENTES
    pasta_fotos_nc = pasta_fotos_nc or M02_FOTOS_NC
    pasta_fotos_pdf = pasta_fotos_pdf or M02_FOTOS_PDF

    if not ncs:
        logger.warning("executar_kria_resposta_de_lista: lista de NCs vazia.")
        return {"kria": None, "resposta": []}

    for nc in ncs:
        if isinstance(nc, dict):
            _nc_sanear_texto_sem_quebras(nc)

    arq_kria = _gerar_kria(
        ncs, nome_base,
        modelo_kria, pasta_saida_kria,
        pasta_fotos_nc,
        relatorio,
        pasta_fotos_pdf_fallback=pasta_fotos_pdf,
    )
    respostas: list[Path] = []
    ocupados: set[str] = set()
    for grupo in _grupos_ncs_resposta_alinhados_exportar(ncs):
        nc0 = grupo[0]
        nome_pref = nome_ficheiro_resposta_artesp_xlsx(nc0)
        nome_final = _alocar_nome_resposta_pendentes_unico(nome_pref, ocupados)
        arq_resp = _gerar_resposta(
            grupo,
            modelo_resposta,
            pasta_saida_resp,
            pasta_fotos_pdf,
            nome_ficheiro_destino=nome_final,
        )
        if arq_resp:
            respostas.append(arq_resp)
    return {"kria": arq_kria, "resposta": respostas}
