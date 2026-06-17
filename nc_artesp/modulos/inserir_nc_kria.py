"""
modulos/inserir_nc_kria.py — Equivalente VBA:
  Art_03_Inserir_NC_Rot_Salva_apo (conservação) | Kria2_Inserir_NC_MA_Salvar_Img (MA).

Formulário foto (.xlsx), bloco de M03_BLOCO linhas por NC (âncora M03_LINHA_INICIO) → planilha Kcor-Kria (A–Y).
Colunas: CABECALHO_KCOR_KRIA em config. Colunas T e U gravadas sem quebra de linha na célula.
Conservação: col W = «arquivo.jpg;pdf (ref).jpg» como na macro Art_03.
"""

import logging
from pathlib import Path
from datetime import datetime, timedelta
import re

from openpyxl import load_workbook
from openpyxl.styles import Border, Side

from config import (
    M03_ENTRADA, M03_IMAGENS, M03_MODELO_KCOR, M03_SAIDA,
    M03_LINHA_INICIO, M03_BLOCO,
    M07_ENTRADA, M07_IMAGENS, M07_MODELO_KCOR, M07_SAIDA,
    KCOR_KRIA_DIRETORIO_TEXTO_CONSERVACAO,
    KCOR_KRIA_DIRETORIO_TEXTO_MEIO_AMBIENTE,
    M02_FOTOS_NC, M02_FOTOS_PDF,
    PRAZO_DIAS_APOS_ENVIO,
    SERVICO_NC, RODOVIAS,
    CONCESSIONARIA_NOME,
)
from utils.helpers import (
    garantir_pasta,
    caminho_dentro_limite_windows,
    copiar_arquivo,
    renomear_arquivo,
    escrever_bytes_caminho,
    resolver_path_ficheiro_ci,
    str_caminho_io_windows,
    parse_data,
    encontrar_foto_por_codigo_ou_numero,
    path_foto_nc,
    path_foto_pdf,
    sanitizar_nome,
    timestamp_agora,
    timestamp_completo,
    km_formato_arquivo,
    formatar_numero,
    pad_metros,
)
from utils.onedrive_local import processar_com_copia_local
from utils.captura_celulas import (
    exportar_range_como_imagem,
    TAMANHO_CONSERVACAO,
    TAMANHO_MA,
)

logger = logging.getLogger(__name__)


def _metros_de_km_t(km_t: str) -> str:
    """Extrai a parte de metros de kminicial_t/kmfinal_t (ex.: '143+800' → '800') para EAF col I/K."""
    s = (km_t or "").strip()
    m = re.search(r"\+(\d+)", s)
    if m:
        return pad_metros(m.group(1))
    m = re.search(r",\s*(\d+)", s)
    if m:
        return pad_metros(m.group(1))
    if s:
        return pad_metros(s[-3:] if len(s) >= 3 else s)
    return "000"


def gerar_eaf_desde_pdf_ma(
    pdf_bytes: bytes,
    pasta_saida: Path,
    nome_arquivo: str = "EAF_MA_desde_PDF.xlsx",
) -> Path | None:
    """
    Gera a planilha do passo 1 (EAF – planilha-mãe) a partir do PDF de Meio Ambiente.
    Extrai as informações em TEXTO do PDF (código, rodovia, km, data, atividade, etc.) via
    analisar_pdf_ma e preenche o Excel no formato esperado pelo Separar NC. Não é extração
    de imagem (essa fica no fluxo Extrair PDF). Retorna o Path do EAF gerado ou None.
    """
    from .analisar_pdf_ma import parse_pdf_ma_para_registros
    from .separar_nc import (
        _caminho_template_eaf,
        PRIMEIRA_LINHA_DADOS,
        COL_CODIGO,
        COL_DATA_CON,
        COL_RODOVIA,
        COL_KM_I_M,
        COL_KM_F_M,
        COL_TIPO_NC,
        COL_SEQ_FOTO,
        COL_DATA_NC,
        COL_CONCESSIONARIA,
        COL_KM_I_FULL,
        COL_KM_F_FULL,
        COL_SENTIDO,
        COL_TIPO_ATIV,
        COL_GRUPO_ATIV,
        COL_RESPONSAVEL,
    )

    registros = parse_pdf_ma_para_registros(pdf_bytes)
    if not registros:
        logger.warning("gerar_eaf_desde_pdf_ma: nenhuma NC extraída do PDF.")
        return None

    template_eaf = _caminho_template_eaf()
    if not template_eaf.is_file():
        logger.error("gerar_eaf_desde_pdf_ma: Template EAF não encontrado: %s", template_eaf)
        return None

    garantir_pasta(pasta_saida)
    destino = caminho_dentro_limite_windows(pasta_saida / nome_arquivo)
    import shutil
    shutil.copy2(template_eaf, destino)
    wb = load_workbook(str_caminho_io_windows(destino))
    ws = wb.active

    while ws.max_row >= PRIMEIRA_LINHA_DADOS:
        ws.delete_rows(ws.max_row, 1)

    col_data_reparo = COL_DATA_NC
    for r in range(1, min(6, ws.max_row + 1)):
        for c in range(1, min(ws.max_column + 1, 25)):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip().lower() == "data reparo":
                col_data_reparo = c
                break
        else:
            continue
        break

    for seq, reg in enumerate(registros, start=1):
        row = PRIMEIRA_LINHA_DADOS + seq - 1
        cf = (reg.get("codigo_fiscalizacao") or "").strip()
        ws.cell(row=row, column=COL_CODIGO).value = (cf if cf and not cf.upper().startswith("LOTE") else reg.get("codigo") or "")
        ws.cell(row=row, column=COL_DATA_CON).value = reg.get("data_raw") or ""
        ws.cell(row=row, column=COL_RODOVIA).value = reg.get("rod_raw") or reg.get("rod_cod") or ""
        ws.cell(row=row, column=COL_KM_I_M).value = _metros_de_km_t(reg.get("kminicial_t") or "")
        ws.cell(row=row, column=COL_KM_F_M).value = _metros_de_km_t(reg.get("kmfinal_t") or reg.get("kminicial_t") or "")
        ws.cell(row=row, column=COL_TIPO_NC).value = (reg.get("resumo") or reg.get("texto") or "")[:255]
        ws.cell(row=row, column=COL_SEQ_FOTO).value = reg.get("numero") or seq
        data_rep = (reg.get("embasamento") or reg.get("prazo") or reg.get("data_raw") or "")
        if hasattr(data_rep, "strftime"):
            data_rep = data_rep.strftime("%d/%m/%Y") if data_rep else ""
        ws.cell(row=row, column=col_data_reparo).value = str(data_rep)[:50]
        ws.cell(row=row, column=COL_CONCESSIONARIA).value = (CONCESSIONARIA_NOME or "")[:100]
        ws.cell(row=row, column=COL_KM_I_FULL).value = (reg.get("kminicial_t") or "")[:50]
        ws.cell(row=row, column=COL_KM_F_FULL).value = (reg.get("kmfinal_t") or reg.get("kminicial_t") or "")[:50]
        ws.cell(row=row, column=COL_SENTIDO).value = (reg.get("sentido") or "")[:50]
        ws.cell(row=row, column=COL_TIPO_ATIV).value = (reg.get("classifica") or "")[:100]
        grupo = reg.get("grupo")
        try:
            _g = int(grupo) if grupo is not None and str(grupo).strip() != "" else None
        except (TypeError, ValueError):
            _g = None
        ws.cell(row=row, column=COL_GRUPO_ATIV).value = _g if _g is not None else ""
        ws.cell(row=row, column=COL_RESPONSAVEL).value = (reg.get("nome_fiscal") or "")[:100]

    wb.save(str_caminho_io_windows(destino))
    wb.close()
    logger.info("Planilha passo 1 (EAF) gerada desde PDF MA: %s (%s NCs)", destino.name, len(registros))
    return destino


def gerar_eaf_desde_pdfs_ma(
    list_pdf_bytes: list[bytes],
    pasta_saida: Path,
    nome_arquivo: str = "EAF_MA_desde_PDF.xlsx",
) -> Path | None:
    """
    Gera a planilha do passo 1 (EAF) a partir de um ou mais PDFs de Meio Ambiente.

    Diferente das demais (conservação/rotina), no MA os dados vêm do PDF e não do Excel:
    não há planilha pré-preenchida; tudo (Código da Fiscalização, Num. da NC, rodovia, km, etc.)
    é extraído do texto do PDF. Por isso a coluna C da EAF usa codigo_fiscalizacao quando
    presente (ex.: 902531), que é o identificador das fotos; Num. da NC (ex.: HE.13.0112)
    é outro campo do PDF.

    Cada PDF pode conter 1 NC; todas as NCs são reunidas numa única planilha EAF.
    Retorna o Path do EAF gerado ou None se nenhuma NC for extraída.
    """
    from .analisar_pdf_ma import parse_pdf_ma_para_registros

    registros = []
    for pdf_bytes in list_pdf_bytes:
        registros.extend(parse_pdf_ma_para_registros(pdf_bytes))
    if not registros:
        logger.warning("gerar_eaf_desde_pdfs_ma: nenhuma NC extraída dos PDFs.")
        return None

    from .separar_nc import (
        _caminho_template_eaf,
        PRIMEIRA_LINHA_DADOS,
        COL_CODIGO,
        COL_DATA_CON,
        COL_RODOVIA,
        COL_KM_I_M,
        COL_KM_F_M,
        COL_TIPO_NC,
        COL_SEQ_FOTO,
        COL_DATA_NC,
        COL_CONCESSIONARIA,
        COL_KM_I_FULL,
        COL_KM_F_FULL,
        COL_SENTIDO,
        COL_TIPO_ATIV,
        COL_GRUPO_ATIV,
        COL_RESPONSAVEL,
    )
    import shutil

    template_eaf = _caminho_template_eaf()
    if not template_eaf.is_file():
        logger.error("gerar_eaf_desde_pdfs_ma: Template EAF não encontrado: %s", template_eaf)
        return None

    garantir_pasta(pasta_saida)
    destino = caminho_dentro_limite_windows(pasta_saida / nome_arquivo)
    shutil.copy2(template_eaf, destino)
    wb = load_workbook(str_caminho_io_windows(destino))
    ws = wb.active

    while ws.max_row >= PRIMEIRA_LINHA_DADOS:
        ws.delete_rows(ws.max_row, 1)

    col_data_reparo = COL_DATA_NC
    col_obs_gestor = None
    col_observacoes = None
    col_prazo = None
    for r in range(1, min(6, ws.max_row + 1)):
        for c in range(1, min(ws.max_column + 1, 30)):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v).strip().lower()
            if s == "data reparo":
                col_data_reparo = c
            elif s in ("observação do gestor", "obsgestor", "obs gestor"):
                col_obs_gestor = c
            elif s in ("observações", "observacao"):
                col_observacoes = c
            elif s == "prazo":
                col_prazo = c

    for seq, reg in enumerate(registros, start=1):
        row = PRIMEIRA_LINHA_DADOS + seq - 1
        # Coluna C = código da fiscalização (extrator); Lote nunca é usado
        cf = (reg.get("codigo_fiscalizacao") or "").strip()
        codigo = (cf if cf and not cf.upper().startswith("LOTE") else reg.get("codigo") or f"NC-MA-{seq}").strip()
        ws.cell(row=row, column=COL_CODIGO).value = codigo or f"NC-MA-{seq}"
        ws.cell(row=row, column=COL_DATA_CON).value = reg.get("data_raw") or ""
        ws.cell(row=row, column=COL_RODOVIA).value = reg.get("rod_raw") or reg.get("rod_cod") or ""
        ws.cell(row=row, column=COL_KM_I_M).value = _metros_de_km_t(reg.get("kminicial_t") or "")
        ws.cell(row=row, column=COL_KM_F_M).value = _metros_de_km_t(reg.get("kmfinal_t") or reg.get("kminicial_t") or "")
        # Fallback quando COL_TIPO_NC vazia
        cf_tipo = (reg.get("codigo_fiscalizacao") or "").strip()
        tipo_nc = (reg.get("resumo") or reg.get("texto") or (cf_tipo if cf_tipo and not cf_tipo.upper().startswith("LOTE") else None) or reg.get("codigo") or "NC Meio Ambiente")
        ws.cell(row=row, column=COL_TIPO_NC).value = (tipo_nc or "NC Meio Ambiente")[:255]
        ws.cell(row=row, column=COL_SEQ_FOTO).value = reg.get("numero") or seq
        data_rep = (reg.get("embasamento") or reg.get("prazo") or reg.get("data_raw") or "")
        if hasattr(data_rep, "strftime"):
            data_rep = data_rep.strftime("%d/%m/%Y") if data_rep else ""
        ws.cell(row=row, column=col_data_reparo).value = str(data_rep)[:50]
        ws.cell(row=row, column=COL_KM_I_FULL).value = (reg.get("kminicial_t") or "")[:50]
        ws.cell(row=row, column=COL_KM_F_FULL).value = (reg.get("kmfinal_t") or reg.get("kminicial_t") or "")[:50]
        ws.cell(row=row, column=COL_SENTIDO).value = (reg.get("sentido") or "")[:50]
        ws.cell(row=row, column=COL_TIPO_ATIV).value = (reg.get("classifica") or "")[:100]
        grupo = reg.get("grupo")
        try:
            _g = int(grupo) if grupo is not None and str(grupo).strip() != "" else None
        except (TypeError, ValueError):
            _g = None
        ws.cell(row=row, column=COL_GRUPO_ATIV).value = _g if _g is not None else ""
        # G = Concessionária (config); U = Responsável (só nome do fiscal, regra MAPA_RESPONSAVEL_TECNICO)
        ws.cell(row=row, column=COL_CONCESSIONARIA).value = (CONCESSIONARIA_NOME or "")[:100]
        ws.cell(row=row, column=COL_RESPONSAVEL).value = (reg.get("nome_fiscal") or "")[:100]
        # Observação do gestor (T): código da fiscalização (extrator insere aqui); Lote nunca é usado
        if col_obs_gestor:
            cod_fisc = (reg.get("codigo_fiscalizacao") or "").strip()
            if not cod_fisc or cod_fisc.upper().startswith("LOTE"):
                cod_fisc = str(seq).zfill(5)
            obs_gestor_val = cod_fisc
            if obs_gestor_val:
                obs_gestor_val = f"Código da fiscalização: {obs_gestor_val}"
            if (reg.get("resumo") or reg.get("texto") or "").strip():
                obs_gestor_val = f"{obs_gestor_val}\n\n{(reg.get('resumo') or reg.get('texto') or '').strip()}"[:500]
            ws.cell(row=row, column=col_obs_gestor).value = (obs_gestor_val or "")[:500]
        if col_observacoes:
            texto_u = " ".join((reg.get("texto") or "").split())
            compl_u = " ".join((reg.get("complemento") or "").split())
            emb_u = str(reg.get("embasamento") or reg.get("prazo") or "").strip()
            obs_u = _observacoes(texto_u, compl_u, emb_u)[:2000]
            ws.cell(row=row, column=col_observacoes).value = obs_u
        if col_prazo is not None:
            prazo_val = reg.get("prazo") or reg.get("embasamento") or ""
            if hasattr(prazo_val, "strftime"):
                prazo_val = prazo_val.strftime("%d/%m/%Y") if prazo_val else ""
            ws.cell(row=row, column=col_prazo).value = str(prazo_val)[:50]

    wb.save(str_caminho_io_windows(destino))
    wb.close()
    logger.info("Planilha passo 1 (EAF) gerada desde %s PDF(s) MA: %s (%s NCs)", len(list_pdf_bytes), destino.name, len(registros))
    return destino


def _processar_pdf_meio_ambiente(
    pdf_bytes: bytes,
    pasta_imagens: Path,
    modelo_kcor: Path,
    pasta_saida: Path,
    nome_origem: str = "PDF MA",
) -> Path | None:
    """
    Processa PDF de Meio Ambiente: extrai todo o texto, parseia as NCs e monta
    a planilha Kcor-Kria (M07). Fluxo de coleta e preenchimento segue a macro MA:
    - Coleta: mesmos campos que a macro lê da planilha (H, G, D, F, L, B em analisar_pdf_ma).
    - Preenchimento: template _Planilha Modelo Kcor-Kria.xlsx, colunas A–Y como na macro.
    As imagens são extraídas do próprio PDF quando possível.
    Retorna o Path do arquivo Kcor-Kria gerado ou None.
    """
    from .analisar_pdf_ma import (
        parse_pdf_ma_para_registros,
        extrair_texto_pdf,
        extrair_texto_pdf_por_blocos,
        extrair_texto_pdf_pdfplumber,
        _sentido_para_texto,
    )

    registros = parse_pdf_ma_para_registros(pdf_bytes)
    if not registros:
        logger.warning("  Nenhuma NC extraída do PDF de Meio Ambiente.")
        try:
            t1 = extrair_texto_pdf(pdf_bytes)
            t2 = extrair_texto_pdf_por_blocos(pdf_bytes)
            t3 = extrair_texto_pdf_pdfplumber(pdf_bytes)
            debug_path = pasta_saida / "debug_texto_extraido_ma.txt"
            with open(debug_path, "w", encoding="utf-8") as f:
                f.write("=== TEXTO PyMuPDF (get_text('text')) ===\n\n")
                f.write(t1[:15000] + ("\n... (truncado)" if len(t1) > 15000 else ""))
                f.write("\n\n=== TEXTO PyMuPDF (get_text('blocks')) ===\n\n")
                f.write(t2[:15000] + ("\n... (truncado)" if len(t2) > 15000 else ""))
                f.write("\n\n=== TEXTO pdfplumber (extract_text) ===\n\n")
                f.write(t3[:15000] + ("\n... (truncado)" if len(t3) > 15000 else ""))
            logger.info("  Texto extraído salvo em: %s (envie esse arquivo para ajustar o parser)", debug_path)
        except Exception as e:
            logger.warning("  Não foi possível salvar texto para debug: %s", e)
        return None

    logger.info(f"  {len(registros)} NC(s) extraída(s) do texto do PDF.")

    # Opcional: extrair imagens do PDF (uso principal é texto; extração de fotos é no fluxo Extrair PDF)
    garantir_pasta(pasta_imagens)
    import tempfile
    try:
        from ..pdf_extractor import extrair_imagens_pdf
        with tempfile.TemporaryDirectory() as tmpdir:
            pdf_path = Path(tmpdir) / "meio_ambiente.pdf"
            escrever_bytes_caminho(pdf_path, pdf_bytes)
            extrair_imagens_pdf(
                str(pdf_path),
                pasta_saida=str(pasta_imagens),
                pasta_saida_nc=str(pasta_imagens),
                pasta_saida_pdf=str(pasta_imagens),
            )
    except Exception:
        pass

    nomes_arquivo = []
    for seq, reg in enumerate(registros, start=1):
        f = encontrar_foto_por_codigo_ou_numero(
            pasta_imagens, "PDF", codigo=reg.get("codigo"), numero=seq
        )
        if not f and pasta_imagens.is_dir():
            f = encontrar_foto_por_codigo_ou_numero(
                pasta_imagens, "nc", codigo=reg.get("codigo"), numero=seq
            )
        nomes_arquivo.append(f.name if f and f.exists() else "")

    if not modelo_kcor.exists():
        logger.error(f"Modelo Kcor-Kria não encontrado: {modelo_kcor}")
        return None

    garantir_pasta(pasta_saida)
    wb_kcor = load_workbook(str_caminho_io_windows(modelo_kcor))
    ws_out = None
    for sheet in wb_kcor.worksheets:
        a1 = sheet.cell(row=1, column=1).value
        if a1 is not None and "numitem" in str(a1).strip().lower():
            ws_out = sheet
            break
    if ws_out is None:
        ws_out = wb_kcor.active
    wb_kcor.active = ws_out
    col_data_envio, col_data_reparo = _detectar_colunas_data_kcor(ws_out)

    diretorio_relatorio_macro = KCOR_KRIA_DIRETORIO_TEXTO_MEIO_AMBIENTE
    for seq, (reg, nome_arq) in enumerate(zip(registros, nomes_arquivo), start=1):
        r = seq + 1
        dt = reg["dt"]
        emb = parse_data(reg["embasamento"])
        dt_str = dt.strftime("%d/%m/%Y") if dt else ""
        emb_str = emb.strftime("%d/%m/%Y") if emb else ""

        ws_out.cell(row=r, column=_K_A).value = seq
        ws_out.cell(row=r, column=_K_B).value = "Artesp"
        ws_out.cell(row=r, column=_K_C).value = "2"
        ws_out.cell(row=r, column=_K_D).value = reg.get("classifica") or ""
        ws_out.cell(row=r, column=_K_E).value = reg.get("serv_nc") or ""
        ws_out.cell(row=r, column=_K_F).value = reg.get("rod_tag") or ""
        ws_out.cell(row=r, column=_K_G).value = reg.get("kminicial_t") or ""
        ws_out.cell(row=r, column=_K_H).value = reg.get("kmfinal_t") or reg.get("kminicial_t") or ""
        ws_out.cell(row=r, column=_K_I).value = _sentido_para_texto(reg.get("sentido") or "") or ""
        ws_out.cell(row=r, column=_K_J).value = ""
        ws_out.cell(row=r, column=_K_K).value = "Conservação"
        ws_out.cell(row=r, column=_K_L).value = ""
        ws_out.cell(row=r, column=col_data_envio).value = dt_str
        ws_out.cell(row=r, column=_K_N).value = ""
        ws_out.cell(row=r, column=_K_O).value = dt_str
        ws_out.cell(row=r, column=col_data_reparo).value = emb_str
        _desfazer_merge_colunas_linha(ws_out, r, _K_Q, _K_T)
        ws_out.cell(row=r, column=_K_Q).value = ""
        ws_out.cell(row=r, column=_K_R).value = ""
        ws_out.cell(row=r, column=_K_S).value = _prazo_para_numero(reg.get("prazo"))
        obs_txt = _obs_gestor(reg["relatorio"], reg["codigo"], reg.get("resumo"), codigo_fiscalizacao=reg.get("codigo_fiscalizacao"))
        ws_out.cell(row=r, column=_K_T).value = _texto_uma_linha(obs_txt)
        texto_u = " ".join((reg.get("texto") or "").split())
        compl_u = " ".join((reg.get("complemento") or "").split())
        emb_u = " ".join((reg.get("embasamento") or "").split())
        ws_out.cell(row=r, column=_K_U).value = _texto_uma_linha(
            _observacoes(texto_u, compl_u, emb_u)
        )
        _copiar_estilo_linha(ws_out, row_origem=2, row_destino=r, col_fim=25)
        # MA: manter diretório canônico do fluxo macro.
        ws_out.cell(row=r, column=_K_V).value = diretorio_relatorio_macro
        ws_out.cell(row=r, column=_K_W).value = nome_arq
        ws_out.cell(row=r, column=_K_X).value = ""
        ws_out.cell(row=r, column=_K_Y).value = ""
        _forcar_texto_so_data_kcor_cols_m_r(ws_out, r)
        _aplicar_bordas_linha(ws_out, r, 25)

    nome_saida = f"{timestamp_agora()} - {nome_origem}.xlsx"
    destino_kcor = caminho_dentro_limite_windows(pasta_saida / nome_saida)
    garantir_pasta(destino_kcor.parent)
    wb_kcor.save(str_caminho_io_windows(destino_kcor))
    wb_kcor.close()
    logger.info(f"  Kcor-Kria (PDF MA) salvo: {destino_kcor.name}")
    return destino_kcor

_COL_B  = 2   # número NC / âncora para seleção
_COL_D  = 4   # rodovia (y-1), km inicial (y)
_COL_F  = 6   # sentido (y-1), km final (y), data (y+1)
_COL_G  = 7   # embasamento (y-2), texto/descrição (y-1)
_COL_H  = 8   # código (y), relatório (y+1)
_COL_L  = 12  # complemento/nº foto (y), prazo dias (y+2)
_COL_C  = 3   # usada apenas para detectar última linha com dado

# Saída Kcor-Kria: colunas 1..25 (A–Y), contíguas para importação.
_K_A  = 1
_K_B  = 2
_K_C  = 3
_K_D  = 4
_K_E  = 5
_K_F  = 6
_K_G  = 7
_K_H  = 8
_K_I  = 9
_K_J  = 10
_K_K  = 11
_K_L  = 12
_K_M  = 13
_K_N  = 14
_K_O  = 15
_K_P  = 16
_K_Q  = 17
_K_R  = 18
_K_S  = 19
_K_T  = 20
_K_U  = 21
_K_V  = 22
_K_W  = 23
_K_X  = 24
_K_Y  = 25


def _cell(ws, row: int, col: int):
    v = ws.cell(row=row, column=col).value
    return v if v is not None else ""


def _detectar_colunas_data_kcor(ws, max_row_cabecalho: int = 2, max_col: int = 25) -> tuple[int, int]:
    """
    Detecta no template Kcor-Kria as colunas 'Data Envio' (ou 'Data Solicitação') e 'Data Reparo' (ou 'DtFim_Prog').
    Retorna (col_data_envio, col_data_reparo). Fallback: (_K_M, _K_P).
    """
    col_envio, col_reparo = None, None
    for row in range(1, max_row_cabecalho + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=row, column=c).value
            if v is None:
                continue
            s = str(v).strip().lower()
            if not s:
                continue
            if "envio" in s or "solicita" in s:
                col_envio = c
            if "reparo" in s or "dtfim" in s.replace(" ", ""):
                col_reparo = c
        if col_envio is not None and col_reparo is not None:
            break
    return (col_envio if col_envio is not None else _K_M, col_reparo if col_reparo is not None else _K_P)


# Borda padrão para reaplicar linhas e bordas nas células preenchidas
_SIDE_THIN = Side(style="thin", color="000000")
_BORDA_CELULA = Border(
    left=_SIDE_THIN, right=_SIDE_THIN, top=_SIDE_THIN, bottom=_SIDE_THIN
)


def _aplicar_bordas_linha(ws, row: int, col_fim: int = 25):
    """Bordas na linha, colunas 1..col_fim."""
    for col in range(1, col_fim + 1):
        ws.cell(row=row, column=col).border = _BORDA_CELULA


def _forcar_texto_so_data_kcor_cols_m_r(ws, row: int) -> None:
    """M–R como texto dd/mm/aaaa (formato @ evita hora herdada do template)."""
    from datetime import date, datetime

    for c in range(_K_M, _K_R + 1):
        cell = ws.cell(row=row, column=c)
        v = cell.value
        if v is None or v == "":
            cell.number_format = "@"
            continue
        if isinstance(v, (datetime, date)):
            cell.value = v.strftime("%d/%m/%Y")
        elif isinstance(v, (int, float)) and not isinstance(v, bool):
            try:
                from openpyxl.utils.datetime import from_excel

                cell.value = from_excel(float(v)).strftime("%d/%m/%Y")
            except Exception:
                pass
        else:
            s = str(v).strip()
            if re.match(r"^\d{1,2}/\d{1,2}/\d{4}", s) and " " in s:
                cell.value = s.split()[0].strip()[:10]
        cell.number_format = "@"


def _desfazer_merge_colunas_linha(ws, row: int, col_ini: int, col_fim: int):
    """Desfaz merges na faixa (template pode mesclar Q:Y; senão só o âncora recebe valor nas células seguintes)."""
    from openpyxl.utils.cell import range_boundaries
    to_unmerge = []
    for mc in list(ws.merged_cells.ranges):
        try:
            min_col, min_row, max_col, max_row = range_boundaries(str(mc))
        except Exception:
            continue
        if row >= min_row and row <= max_row and not (col_fim < min_col or col_ini > max_col):
            to_unmerge.append(mc)
    for mc in to_unmerge:
        try:
            ws.unmerge_cells(str(mc))
        except Exception:
            pass


def _copiar_estilo_linha(ws, row_origem: int, row_destino: int, col_fim: int = 25):
    """Estilo 1–U; bordas até col_fim (V–Y preenchidas depois, sem copiar estilo do modelo)."""
    col_ate_u = min(21, col_fim)
    for col in range(1, col_ate_u + 1):
        src = ws.cell(row=row_origem, column=col)
        dst = ws.cell(row=row_destino, column=col)
        if src.has_style:
            dst.font = src.font.copy()
            dst.border = src.border.copy()
            dst.fill = src.fill.copy()
            dst.number_format = src.number_format
            dst.alignment = src.alignment.copy()
    _aplicar_bordas_linha(ws, row_destino, col_fim)


def _str(v) -> str:
    return str(v).strip() if v is not None else ""


def _normalizar_rodovia_formulario(rod_raw: str) -> tuple[str, str, int]:
    """
    Recebe valor da col D(y-1) do formulário — já no formato SP-075, SPI-102/300 etc.
    Retorna (tag, codigo, n).
    """
    rod = rod_raw.strip()
    mapa = {
        "SP-075":     ("SP075",     "SP075",     1),
        "SP-127":     ("SP127",     "SP127",     2),
        "SP-280":     ("SP280",     "SP280",     3),
        "SP-300":     ("SP300",     "SP300",     4),
        "SPI-102/300":("SPI102_300","SPI102/300",5),
        "CP-127_147": ("CP-127_147","FORA",      6),
        "CP-127_308": ("CP-127_308","FORA",      7),
    }
    if rod in mapa:
        return mapa[rod]
    # fallback: limpeza básica
    tag = sanitizar_nome(rod).replace("-", "").replace("/", "_")
    return (tag, rod, 0)


def _data_nome_yyyymmdd(data_str: str, nome_origem: str = "") -> str:
    """
    Resolve a data usada no nome do JPG.
    Prioridade:
      1) data_str válida (vinda da planilha)
      2) data no nome do arquivo de origem (preferindo a 2ª ocorrência yyyyMMdd)
      3) data atual
    """
    dt = parse_data(data_str)
    if dt:
        return dt.strftime("%Y%m%d")

    datas_nome = re.findall(r"\b(20\d{6})\b", nome_origem or "")
    if datas_nome:
        return datas_nome[1] if len(datas_nome) >= 2 else datas_nome[0]

    return datetime.now().strftime("%Y%m%d")


def _nome_arquivo_jpg(data_str: str, n: int, numero: str,
                      rodoviat: str, kminicial: str, sentido: str,
                      nome_origem: str = "") -> str:
    """
    Monta o nome do JPG exportado.
    Formato VBA: yyyymmdd - hhmmss - n_Roti-NNNNNN-RODOVIA km,metro SENTIDO.jpg
    """
    yyyymmdd = _data_nome_yyyymmdd(data_str, nome_origem=nome_origem)
    hhmmss = datetime.now().strftime("%H%M%S")
    return sanitizar_nome(
        f"{yyyymmdd} - {hhmmss} - "
        f"{n}_Roti-{numero}-{rodoviat} {kminicial} {sentido}.jpg"
    )


def _obs_gestor(relatorio: str, codigo: str, resumo_tecnico: str | None = None, codigo_fiscalizacao: str | None = None) -> str:
    """ObsGestor (col T); texto montado em _obs_gestor; célula final sem CRLF (_texto_uma_linha)."""
    rel = (relatorio or "").strip()
    cod = (codigo or "").strip()
    if resumo_tecnico and resumo_tecnico.strip():
        return resumo_tecnico.strip()
    return f"--> Relatório EAF Conservação Rotina nº: {rel}\n--> Código NC: {cod}"


def _texto_uma_linha(v: str) -> str:
    raw = str(v or "")
    if not raw:
        return ""
    t = raw.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    return " ".join(t.split())


def _observacoes(texto: str, complemento: str, embasamento: str) -> str:
    """Observações (col U); blocos internos; célula final sem CRLF (_texto_uma_linha)."""
    if complemento:
        return (
            f"{texto}\n"
            f"- Complemento ----> {complemento}\n\n"
            f"- Embasamento ----> {embasamento}"
        )
    return f"{texto}\n- Data Superação Artesp ----> {embasamento}"


def _prazo_para_numero(prazo) -> str | int:
    """Converte valor de prazo (L10) para número de dias na col S. Aceita '7 dias', 7, ou string com número."""
    if prazo is None:
        return ""
    if isinstance(prazo, (int, float)) and not isinstance(prazo, bool):
        return int(prazo)
    s = str(prazo).strip()
    if not s:
        return ""
    import re
    m = re.search(r"(\d+)\s*dias?", s, re.I)
    if m:
        return int(m.group(1))
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return s[:20]


# PROCESSAMENTO DE UM ÚNICO ARQUIVO

def _processar_arquivo(arq_path: Path,
                        modo: str,
                        pasta_imagens: Path,
                        pasta_fotos_pdf: Path | None,
                        pasta_fotos_nc: Path | None,
                        modelo_kcor: Path,
                        pasta_saida: Path,
                        forcar_fallback: bool = False,
                        regime_artemig: bool = False) -> Path | None:
    """
    Processa um único arquivo XLSX de formulário de foto.
    Retorna o Path do Kcor-Kria gerado, ou None em caso de erro.
    """
    from .analisar_pdf_ma import _sentido_para_texto
    eh_conservacao = (modo == "conservacao")
    tamanho_img = TAMANHO_CONSERVACAO if eh_conservacao else TAMANHO_MA
    largura, altura = tamanho_img

    logger.info(f"Processando [{modo}]: {arq_path.name}")

    wb_form = load_workbook(str_caminho_io_windows(arq_path), data_only=True)
    ws      = wb_form.active

    ultima = ws.max_row
    for r in range(ultima, M03_LINHA_INICIO - 1, -1):
        if ws.cell(row=r, column=_COL_C).value is not None:
            ultima = r
            break

    registros = []
    y = M03_LINHA_INICIO

    while y <= ultima + (M03_BLOCO - 2):
        if not _str(_cell(ws, y, _COL_D)) and not _str(_cell(ws, y, _COL_H)):
            break

        relatorio   = _str(_cell(ws, y + 1, _COL_H))
        codigo      = _str(_cell(ws, y,     _COL_H))
        complemento = _str(_cell(ws, y,     _COL_L))
        embasamento = _str(_cell(ws, y - 2, _COL_G))
        data_reparo_raw = _str(_cell(ws, y + 1, _COL_D))
        rod_raw     = _str(_cell(ws, y - 1, _COL_D))
        texto       = _str(_cell(ws, y - 1, _COL_G))
        sentido     = _str(_cell(ws, y - 1, _COL_F))
        kminicial_t = _str(_cell(ws, y,     _COL_D))
        kmfinal_t   = _str(_cell(ws, y,     _COL_F))
        data_raw    = _str(_cell(ws, y + 1, _COL_F))
        prazo       = _cell(ws, y + 1, _COL_L)
        numero_raw  = _cell(ws, y - 3, _COL_B)

        foto_id = None
        try:
            foto_id = int(float(str(complemento))) if complemento else None
            complemento = ""
        except (ValueError, TypeError):
            pass

        rod_tag, rod_cod, n = _normalizar_rodovia_formulario(rod_raw)

        kminicial_arq = km_formato_arquivo(kminicial_t)

        numero = formatar_numero(numero_raw or 1, 6)

        dt = parse_data(data_raw)
        data_str_raw = data_raw
        data_reparo_dt = parse_data(data_reparo_raw)

        if eh_conservacao:
            descricao = texto
            desc_s = (descricao or "").strip()
            nc_info = SERVICO_NC.get(desc_s) or SERVICO_NC.get(descricao)
            if nc_info:
                serv_nc, classifica_srv, executor = nc_info
                classifica = (str(classifica_srv).strip() if classifica_srv else "") or "Conservação Rotina"
            else:
                serv_nc = ""
                classifica = "Conservação Rotina"
                executor = "Soluciona - Conserva"
        else:
            serv_nc    = "Reclassificar"
            classifica = "Conservação Rotina"
            executor   = ""

        registros.append({
            "y":           y,
            "relatorio":   relatorio,
            "codigo":      codigo,
            "complemento": complemento,
            "embasamento": embasamento,
            "data_reparo_dt": data_reparo_dt,
            "data_reparo_raw": data_reparo_raw,
            "rod_raw":     rod_raw,
            "rod_tag":     rod_tag,
            "rod_cod":     rod_cod,
            "n":           n,
            "texto":       texto,
            "sentido":     sentido,
            "kminicial_t": kminicial_t,
            "kmfinal_t":   kmfinal_t,
            "kminicial_arq": kminicial_arq,
            "data_raw":    data_str_raw,
            "dt":          dt,
            "prazo":       prazo,
            "numero":      numero,
            "foto_id":     foto_id,
            "serv_nc":     serv_nc,
            "classifica":  classifica,
            "executor":    executor,
        })
        y += M03_BLOCO

    wb_form.close()

    if not registros:
        logger.warning(f"  Nenhum registro encontrado em {arq_path.name}")
        return None

    if regime_artemig and eh_conservacao:
        try:
            from nc_artemig.sentido_kcor import sentido_artemig_para_kcor
            for reg in registros:
                rod = (reg.get("rod_raw") or reg.get("rod_tag") or "").strip()
                reg["sentido"] = sentido_artemig_para_kcor(rod, reg.get("sentido") or "")
        except Exception:
            pass

    logger.info(f"  {len(registros)} NC(s) lida(s).")

    garantir_pasta(pasta_imagens)
    nomes_arquivo = []

    for reg in registros:
        y_anc = reg["y"]

        row_ini = y_anc - 3
        col_ini = 3
        row_fim = y_anc + 1
        col_fim = 6

        nome_jpg = _nome_arquivo_jpg(
            reg["data_raw"], reg["n"], reg["numero"],
            reg["rod_tag"], reg["kminicial_arq"], reg["sentido"],
            nome_origem=arq_path.name,
        )
        destino_jpg = pasta_imagens / nome_jpg

        foto_real_existe = False
        foto_pdf = foto_nc = None
        if eh_conservacao and pasta_fotos_pdf and (reg.get("codigo") or reg.get("foto_id") is not None):
            codigo = reg.get("codigo")
            foto_id = reg.get("foto_id")
            foto_pdf = encontrar_foto_por_codigo_ou_numero(
                pasta_fotos_pdf, "PDF", codigo=codigo, numero=foto_id
            )
            if pasta_fotos_nc:
                foto_nc = encontrar_foto_por_codigo_ou_numero(
                    pasta_fotos_nc, "nc", codigo=codigo, numero=foto_id
                )
            foto_real_existe = (foto_pdf is not None) or (foto_nc is not None)

        if eh_conservacao:
            if pasta_fotos_pdf and foto_real_existe:
                origem_foto = None
                if foto_pdf and foto_pdf.exists():
                    origem_foto = foto_pdf
                elif foto_nc and foto_nc.exists():
                    origem_foto = foto_nc
                if origem_foto:
                    copiar_arquivo(origem_foto, destino_jpg, sobrescrever=True)
                    logger.debug(f"  Foto real usada: {origem_foto.name} → {nome_jpg}")
                else:
                    nome_jpg = ""
            else:
                nome_jpg = ""
                logger.debug(
                    f"  NC y={y_anc}: foto real indisponível (foto_id={reg['foto_id']}), "
                    "não exportando print da planilha."
                )
        else:
            ok = exportar_range_como_imagem(
                wb_path      = arq_path,
                sheet_index  = 0,
                row_ini      = row_ini,
                col_ini      = col_ini,
                row_fim      = row_fim,
                col_fim      = col_fim,
                destino      = destino_jpg,
                largura      = largura,
                altura       = altura,
                forcar_fallback = forcar_fallback,
            )
            if ok:
                logger.debug(f"  JPG gerado: {nome_jpg}")
            else:
                logger.warning(f"  Falha ao exportar JPG para NC y={y_anc}")
                nome_jpg = ""

        nomes_arquivo.append(nome_jpg)

    if not modelo_kcor.exists():
        logger.error(f"Modelo Kcor-Kria não encontrado: {modelo_kcor}")
        return None

    garantir_pasta(pasta_saida)

    wb_kcor = load_workbook(str_caminho_io_windows(modelo_kcor))
    ws_out = None
    for sheet in wb_kcor.worksheets:
        a1 = sheet.cell(row=1, column=1).value
        if a1 is not None and "numitem" in str(a1).strip().lower():
            ws_out = sheet
            break
    if ws_out is None:
        ws_out = wb_kcor.active
    wb_kcor.active = ws_out
    col_data_envio, col_data_reparo = _detectar_colunas_data_kcor(ws_out)
    logger.debug(f"Kcor-Kria: col Data Envio={col_data_envio}, col Data Reparo={col_data_reparo}")

    diretorio_relatorio_macro = (
        KCOR_KRIA_DIRETORIO_TEXTO_CONSERVACAO
        if eh_conservacao
        else KCOR_KRIA_DIRETORIO_TEXTO_MEIO_AMBIENTE
    )
    for seq, (reg, nome_arq) in enumerate(zip(registros, nomes_arquivo), start=1):
        r = seq + 1  # linha de saída (linha 2 em diante)

        dt     = reg["dt"]
        emb    = parse_data(reg["embasamento"])
        dt_str = dt.strftime("%d/%m/%Y") if dt else ""
        emb_str = emb.strftime("%d/%m/%Y") if emb else ""
        data_rep_dt = reg.get("data_reparo_dt")
        if data_rep_dt:
            data_reparo_str = data_rep_dt.strftime("%d/%m/%Y")
        elif dt:
            data_reparo_str = (dt + timedelta(days=PRAZO_DIAS_APOS_ENVIO)).strftime("%d/%m/%Y")
        else:
            data_reparo_str = emb_str
        if reg.get("data_reparo_raw") and _str(reg.get("embasamento")) == _str(reg.get("data_reparo_raw")):
            envio_str = dt_str
        else:
            envio_str = emb_str

        ws_out.cell(row=r, column=_K_A).value = seq
        ws_out.cell(row=r, column=_K_B).value = "Artesp"
        ws_out.cell(row=r, column=_K_C).value = "2"
        ws_out.cell(row=r, column=_K_D).value = reg["classifica"]
        ws_out.cell(row=r, column=_K_E).value = (reg.get("texto") or "").strip() or reg["serv_nc"]
        ws_out.cell(row=r, column=_K_F).value = reg["rod_tag"]
        ws_out.cell(row=r, column=_K_G).value = reg["kminicial_t"]
        ws_out.cell(row=r, column=_K_H).value = reg.get("kmfinal_t") or reg.get("kminicial_t") or ""
        if regime_artemig and eh_conservacao:
            sentido_i = (reg.get("sentido") or "").strip()
        else:
            sentido_i = _sentido_para_texto(reg.get("sentido") or "") or ""
        ws_out.cell(row=r, column=_K_I).value = sentido_i[:120]
        ws_out.cell(row=r, column=_K_J).value = ""
        ws_out.cell(row=r, column=_K_K).value = "Conservação"

        if eh_conservacao:
            ws_out.cell(row=r, column=_K_L).value = reg["executor"]
            ws_out.cell(row=r, column=_K_N).value = ""
            ws_out.cell(row=r, column=_K_O).value = dt_str
            ws_out.cell(row=r, column=_K_Q).value = ""
            ws_out.cell(row=r, column=_K_R).value = ""
        else:
            ws_out.cell(row=r, column=_K_N).value = ""
            ws_out.cell(row=r, column=_K_O).value = dt_str
            ws_out.cell(row=r, column=_K_Q).value = ""
            ws_out.cell(row=r, column=_K_R).value = ""

        ws_out.cell(row=r, column=_K_M).value = dt_str
        ws_out.cell(row=r, column=_K_P).value = envio_str

        _desfazer_merge_colunas_linha(ws_out, r, _K_Q, _K_T)
        obs_gestor_txt = _obs_gestor(
            reg["relatorio"], reg["codigo"], codigo_fiscalizacao=reg.get("codigo_fiscalizacao")
        )
        emb_u = reg.get("embasamento")
        observacoes_u = _observacoes(
            reg["texto"], reg["complemento"], str(emb_u) if emb_u is not None else "",
        )
        ws_out.cell(row=r, column=_K_S).value = _prazo_para_numero(reg.get("prazo"))
        ws_out.cell(row=r, column=_K_T).value = _texto_uma_linha(obs_gestor_txt)
        ws_out.cell(row=r, column=_K_U).value = _texto_uma_linha(observacoes_u)
        _copiar_estilo_linha(ws_out, row_origem=2, row_destino=r, col_fim=25)
        # Coluna "Diretório" deve seguir o caminho padrão das macros (config),
        # e não a pasta temporária/local da execução atual.
        ws_out.cell(row=r, column=_K_V).value = diretorio_relatorio_macro
        # Macro Art_03: coluna W = arquivo(i) & ";pdf (" & foto(i) & ").jpg"
        # onde foto(i) vem da célula L da planilha de entrada (não usa código fiscal).
        foto_ref = reg.get("foto_id")
        if foto_ref is None:
            foto_ref = _str(reg.get("complemento"))
        pdf_nome = f"pdf ({foto_ref}).jpg" if foto_ref not in (None, "") else ""
        if nome_arq and pdf_nome:
            arquivos_w = f"{nome_arq};{pdf_nome}"
        else:
            arquivos_w = nome_arq or pdf_nome
        ws_out.cell(row=r, column=_K_W).value = arquivos_w
        ws_out.cell(row=r, column=_K_X).value = ""
        ws_out.cell(row=r, column=_K_Y).value = ""
        _forcar_texto_so_data_kcor_cols_m_r(ws_out, r)
        _aplicar_bordas_linha(ws_out, r, 25)

    nome_saida = f"{timestamp_agora()} - {arq_path.name}"
    destino_kcor = caminho_dentro_limite_windows(pasta_saida / nome_saida)
    garantir_pasta(destino_kcor.parent)
    wb_kcor.save(str_caminho_io_windows(destino_kcor))
    wb_kcor.close()
    logger.info(f"  Kcor-Kria salvo: {destino_kcor.name}")

    novo_nome = arq_path.parent / f"_Processado - {arq_path.name}"
    renomear_arquivo(arq_path, novo_nome)

    return destino_kcor


def executar_conservacao(pasta_entrada: Path | None = None,
                          pasta_imagens: Path | None = None,
                          modelo_kcor:   Path | None = None,
                          pasta_saida:   Path | None = None,
                          pasta_fotos_pdf: Path | None = None,
                          pasta_fotos_nc: Path | None = None,
                          forcar_fallback: bool = False,
                          callback_progresso=None,
                          regime_artemig: bool = False) -> list[Path]:
    """M03 conservação: .xlsx formulário → .xlsx Kcor-Kria (1 linha/NC)."""
    pasta_entrada   = pasta_entrada   or M03_ENTRADA
    pasta_imagens   = pasta_imagens   or M03_IMAGENS
    modelo_kcor     = resolver_path_ficheiro_ci(modelo_kcor or M03_MODELO_KCOR)
    pasta_saida     = pasta_saida     or M03_SAIDA
    pasta_fotos_pdf = pasta_fotos_pdf or M02_FOTOS_PDF
    pasta_fotos_nc  = pasta_fotos_nc  or M02_FOTOS_NC

    return _executar_em_pasta(
        pasta_entrada, "conservacao",
        pasta_imagens, pasta_fotos_pdf, pasta_fotos_nc,
        modelo_kcor, pasta_saida,
        forcar_fallback, callback_progresso,
        regime_artemig=regime_artemig,
    )


def executar_meio_ambiente(pasta_entrada: Path | None = None,
                            pasta_imagens: Path | None = None,
                            modelo_kcor:   Path | None = None,
                            pasta_saida:   Path | None = None,
                            pasta_fotos_pdf: Path | None = None,
                            pasta_fotos_nc: Path | None = None,
                            forcar_fallback: bool = False,
                            callback_progresso=None) -> list[Path]:
    """M07 MA: pasta .xlsx → Kcor-Kria (modo meio_ambiente)."""
    pasta_entrada = pasta_entrada or M07_ENTRADA
    pasta_imagens = pasta_imagens or M07_IMAGENS
    modelo_kcor   = resolver_path_ficheiro_ci(modelo_kcor or M03_MODELO_KCOR)
    pasta_saida   = pasta_saida   or M07_SAIDA

    return _executar_em_pasta(
        pasta_entrada, "meio_ambiente",
        pasta_imagens, None, None,
        modelo_kcor, pasta_saida,
        forcar_fallback, callback_progresso,
        regime_artemig=False,
    )


def executar_meio_ambiente_pdf(
    pdf_bytes: bytes,
    pasta_imagens: Path | None = None,
    modelo_kcor: Path | None = None,
    pasta_saida: Path | None = None,
    nome_origem: str = "PDF MA",
) -> list[Path]:
    """M07: PDF MA → Kcor-Kria; imagens do PDF quando possível."""
    pasta_imagens = pasta_imagens or M07_IMAGENS
    modelo_kcor   = resolver_path_ficheiro_ci(modelo_kcor or M07_MODELO_KCOR)
    pasta_saida   = pasta_saida   or M07_SAIDA

    destino = _processar_pdf_meio_ambiente(
        pdf_bytes, pasta_imagens, modelo_kcor, pasta_saida, nome_origem=nome_origem
    )
    return [destino] if destino else []


def executar_pipeline_meio_ambiente_pdf(
    pdf_bytes: bytes,
    pasta_imagens: Path | None = None,
    pasta_saida_kria: Path | None = None,
    pasta_saida_resp: Path | None = None,
    pasta_saida_eaf: Path | None = None,
    pasta_saida_separar_nc: Path | None = None,
    modelo_kria: Path | None = None,
    modelo_resposta: Path | None = None,
    modelo_kcor: Path | None = None,
    pasta_saida_kcor: Path | None = None,
    nome_origem: str = "PDF MA",
) -> dict:
    """
    Pipeline Meio Ambiente equivalente a M1 + M2 + M3 a partir do PDF.

    M1 equivalente: extrai todo o texto do PDF e parseia as NCs (parse_pdf_ma).
    Gera também a planilha EAF (template do Separar NC) para uso no fluxo padrão.
    M2 equivalente: gera Kria e Resposta. Se pasta_saida_separar_nc for informada e o EAF
    existir, executa Separar NC no EAF e depois Gerar Modelo Foto (M02) nos arquivos
    individuais — assim o modelo Kria vem preenchido pelo mesmo fluxo do Separar NC.
    Caso contrário, gera um único Kria/Resposta a partir da lista de NCs (executar_kria_resposta_de_lista).
    M3 equivalente: gera Kcor-Kria e imagens em Meio Ambiente.

    Retorna dict com: kria, resposta, kcor (list), eaf (Path | None).
    """
    from .analisar_pdf_ma import parse_pdf_ma, ncs_ma_para_dict_m2
    from . import gerar_modelo_foto

    pasta_imagens    = pasta_imagens or M07_IMAGENS
    pasta_saida_kria = pasta_saida_kria or M07_ENTRADA
    modelo_kcor      = resolver_path_ficheiro_ci(modelo_kcor or M07_MODELO_KCOR)
    pasta_saida_kcor = pasta_saida_kcor or M07_SAIDA

    garantir_pasta(pasta_imagens)
    garantir_pasta(pasta_saida_kria)

    ncs_ma = parse_pdf_ma(pdf_bytes)
    if not ncs_ma:
        logger.warning("Pipeline MA: nenhuma NC extraída do PDF.")
        return {"kria": None, "resposta": None, "kcor": [], "eaf": None}

    logger.info(f"Pipeline MA (M1): {len(ncs_ma)} NC(s) extraída(s) do PDF.")

    # ── EAF (template Separar NC): planilha-mãe no formato do M01 ─────────
    eaf_path = None
    if pasta_saida_eaf is not None:
        eaf_path = gerar_eaf_desde_pdf_ma(
            pdf_bytes,
            pasta_saida=pasta_saida_eaf,
            nome_arquivo="EAF_MA_desde_PDF.xlsx",
        )
        if eaf_path:
            logger.info(f"Pipeline MA: EAF (template Separar NC) gerado: {eaf_path.name}")

    import tempfile
    try:
        from ..pdf_extractor import extrair_imagens_pdf
        with tempfile.TemporaryDirectory() as tmpdir:
            pdf_path = Path(tmpdir) / "meio_ambiente.pdf"
            escrever_bytes_caminho(pdf_path, pdf_bytes)
            extrair_imagens_pdf(
                str(pdf_path),
                pasta_saida=str(pasta_imagens),
                pasta_saida_nc=str(pasta_imagens),
                pasta_saida_pdf=str(pasta_imagens),
            )
    except Exception as e:
        logger.warning(f"Pipeline MA: extração de imagens falhou: {e}")

    # MA coleta dados do PDF (não de Excel). Sempre gerar Kria/Resposta a partir dos dados do PDF.
    resultado_m2 = {"kria": [], "resposta": []}
    try:
        ncs_dict = ncs_ma_para_dict_m2(ncs_ma)
        relatorio = (ncs_ma[0].relatorio or nome_origem)[:8] if ncs_ma else nome_origem[:8]
        res = gerar_modelo_foto.executar_kria_resposta_de_lista(
            ncs_dict,
            nome_base=nome_origem,
            relatorio=relatorio,
            modelo_kria=modelo_kria,
            pasta_saida_kria=pasta_saida_kria,
            modelo_resposta=modelo_resposta,
            pasta_saida_resp=pasta_saida_resp,
            pasta_fotos_nc=pasta_imagens,
            pasta_fotos_pdf=pasta_imagens,
        )
        kria = res.get("kria")
        resp = res.get("resposta")
        if isinstance(kria, list):
            resultado_m2["kria"] = [p.resolve() if hasattr(p, "resolve") else p for p in kria if p]
        elif kria:
            resultado_m2["kria"] = [kria.resolve() if hasattr(kria, "resolve") else kria]
        if isinstance(resp, list):
            resultado_m2["resposta"] = [p.resolve() if hasattr(p, "resolve") else p for p in resp if p]
        elif resp:
            resultado_m2["resposta"] = [resp.resolve() if hasattr(resp, "resolve") else resp]
    except Exception as e_m2:
        logger.exception("Pipeline MA: M2 (Kria/Resposta desde PDF) falhou: %s", e_m2)

    # Opcional: fluxo Excel (Separar NC + M02) — no MA o EAF foi gerado do PDF; pode rodar para ter arquivos no formato do M01
    if pasta_saida_separar_nc is not None and eaf_path is not None and eaf_path.is_file():
        try:
            from . import separar_nc
            garantir_pasta(pasta_saida_separar_nc)
            arqs_separar = separar_nc.executar(arquivo_mae=eaf_path, pasta_destino=pasta_saida_separar_nc, sobrescrever=True)
            if arqs_separar:
                logger.info(f"Pipeline MA: Separar NC (desde EAF do PDF) gerou {len(arqs_separar)} arquivo(s).")
                res_m02 = gerar_modelo_foto.executar(
                    pasta_xls=pasta_saida_separar_nc,
                    modelo_kria=modelo_kria,
                    pasta_saida_kria=pasta_saida_kria,
                    modelo_resposta=modelo_resposta,
                    pasta_saida_resp=pasta_saida_resp,
                    pasta_fotos_nc=pasta_imagens,
                    pasta_fotos_pdf=pasta_imagens,
                )
                # Incluir também os Kria/Resposta gerados a partir do Excel (EAF) na lista de saída
                if res_m02.get("kria"):
                    resultado_m2["kria"] = list(resultado_m2.get("kria") or []) + list(res_m02["kria"])
                if res_m02.get("resposta"):
                    resultado_m2["resposta"] = list(resultado_m2.get("resposta") or []) + list(res_m02["resposta"])
        except Exception as e_m2:
            logger.warning("Pipeline MA: Separar NC + M02 (fluxo Excel) falhou (%s). Kria/Resposta já gerados do PDF.", e_m2)

    kcor_lista = executar_meio_ambiente_pdf(
        pdf_bytes,
        pasta_imagens=pasta_imagens,
        modelo_kcor=modelo_kcor,
        pasta_saida=pasta_saida_kcor,
        nome_origem=nome_origem,
    )

    # Paths absolutos para o ZIP
    kcor_resolved = [p.resolve() if hasattr(p, "resolve") else p for p in (kcor_lista or []) if p]
    eaf_resolved = eaf_path.resolve() if eaf_path and hasattr(eaf_path, "resolve") and eaf_path.is_file() else eaf_path
    return {
        "kria": resultado_m2.get("kria") or [],
        "resposta": resultado_m2.get("resposta") or [],
        "kcor": kcor_resolved,
        "eaf": eaf_resolved,
    }


def _executar_em_pasta(pasta_entrada: Path, modo: str,
                        pasta_imagens: Path,
                        pasta_fotos_pdf: Path | None,
                        pasta_fotos_nc: Path | None,
                        modelo_kcor: Path,
                        pasta_saida: Path,
                        forcar_fallback: bool,
                        callback_progresso,
                        regime_artemig: bool = False) -> list[Path]:
    """Engine comum para os dois modos."""
    arquivos = sorted([
        f for f in pasta_entrada.glob("*.xlsx")
        if not f.name.startswith("~")
        and not f.name.startswith("_Processado")
        and not f.name.startswith("_")
        and "Relatório Fotográfico" not in f.name
    ])

    if not arquivos:
        logger.warning(f"Nenhum .xlsx encontrado em: {pasta_entrada}")
        return []

    label = "Conservação" if modo == "conservacao" else "Meio Ambiente"
    logger.info(f"Módulo 03/07 [{label}]: {len(arquivos)} arquivo(s).")

    gerados: list[Path] = []
    for idx, arq in enumerate(arquivos):
        if callback_progresso:
            callback_progresso(idx + 1, len(arquivos), f"[{label}] {arq.name[:55]}")

        gerado_kcor: list[Path | None] = [None]

        def _processar_com_copia(path_local: str):
            p = Path(path_local)
            res = _processar_arquivo(
                arq_path        = p,
                modo            = modo,
                pasta_imagens   = pasta_imagens,
                pasta_fotos_pdf = pasta_fotos_pdf or Path("."),
                pasta_fotos_nc  = pasta_fotos_nc,
                modelo_kcor     = modelo_kcor,
                pasta_saida     = pasta_saida,
                forcar_fallback = forcar_fallback,
                regime_artemig  = regime_artemig,
            )
            gerado_kcor[0] = res
            # Arquivo de entrada é renomeado para _Processado - nome; copiar esse de volta
            path_renomeado = p.parent / f"_Processado - {p.name}"
            return str(path_renomeado) if path_renomeado.exists() else path_local

        try:
            processar_com_copia_local(arq, _processar_com_copia)
            if gerado_kcor[0]:
                gerados.append(gerado_kcor[0])
        except KeyError as e:
            if "Content_Types" in str(e):
                logger.error(f"Arquivo corrompido ou incompleto: {arq.name}. Remova e regenere pelo Módulo 02.")
            else:
                logger.error(f"Erro em {arq.name}: {e}", exc_info=True)
        except Exception as e:
            logger.error(f"Erro em {arq.name}: {e}", exc_info=True)

    logger.info(f"Módulo 03/07 [{label}] concluído: {len(gerados)} arquivo(s) gerado(s).")
    if callback_progresso:
        callback_progresso(len(arquivos), len(arquivos),
                           f"[{label}] Concluído: {len(gerados)} arquivo(s).")
    return gerados
