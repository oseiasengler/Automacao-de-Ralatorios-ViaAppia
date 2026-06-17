"""
modulos/inserir_numero_kria.py
────────────────────────────────────────────────────────────────────────────
Equivalente VBA: Art_05_EAF_Rot_Ins_Num_Kria_Plan_Padrao
Desenvolvedor: Ozeias Engler

Recebe o número inicial do evento Kria e preenche a coluna Y com a sequência:
  N24, (N+1)24, (N+2)24, ...

O sufixo "24" é literal — representa o ano do evento no sistema Kria.
"""

import logging
from pathlib import Path

from openpyxl import load_workbook

from config import M05_COL_NUMERO, M05_SUFIXO
from utils.helpers import garantir_pasta, str_caminho_io_windows

logger = logging.getLogger(__name__)

# Mapeamento letra → número de coluna
_COL_MAP = {c: i + 1 for i, c in enumerate("ABCDEFGHIJKLMNOPQRSTUVWXY")}


def _col_idx(col_letra: str) -> int:
    return _COL_MAP.get(col_letra.upper(), 25)


def executar(arquivo: Path, numero_inicial: int,
             coluna: str | None = None,
             sufixo: str | None = None,
             callback_progresso=None) -> int:
    """
    Preenche a coluna com a sequência de números de evento Kria.

    Parâmetros
    ----------
    arquivo         : Path para o .xlsx da planilha acumulada.
    numero_inicial  : Número do primeiro evento (int).
    coluna          : Letra da coluna (padrão: config.M05_COL_NUMERO = 'Y').
    sufixo          : Sufixo a concatenar (padrão: config.M05_SUFIXO = '24').

    Retorna
    -------
    Número de linhas preenchidas.
    """
    coluna = (coluna or M05_COL_NUMERO).upper()
    sufixo = sufixo if sufixo is not None else M05_SUFIXO
    col_idx = _col_idx(coluna)

    logger.info(f"Abrindo: {arquivo.name}")
    wb = load_workbook(str_caminho_io_windows(arquivo))
    ws = wb.active

    # Encontrar última linha com dados (col A)
    ultima = ws.max_row
    for r in range(ultima, 1, -1):
        if ws.cell(row=r, column=1).value is not None:
            ultima = r
            break

    total = ultima - 1  # linhas de dados (linha 2 em diante)
    logger.info(f"Preenchendo coluna {coluna} nas linhas 2 a {ultima} "
                f"({total} registros). Início: {numero_inicial}{sufixo}")

    numero = numero_inicial
    for idx, r in enumerate(range(2, ultima + 1)):
        ws.cell(row=r, column=col_idx).value = f"{numero}{sufixo}"
        numero += 1
        if callback_progresso and idx % 10 == 0:
            callback_progresso(idx + 1, total, f"Numerando linha {r}...")

    wb.save(str_caminho_io_windows(arquivo))
    wb.close()

    logger.info(f"Módulo 05 concluído. {total} evento(s) numerado(s).")
    if callback_progresso:
        callback_progresso(total, total, "Módulo 05 concluído.")

    return total
