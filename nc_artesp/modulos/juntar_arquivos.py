"""
modulos/juntar_arquivos.py — Equivalente VBA: Art_04_EAF_Rot_Juntar_Arquivo_Exportar_Kria.

Consolida .xlsx Kcor-Kria (saída M03) no acumulado A–Y. Coluna A = sequência.
Ao gravar, normaliza texto nas colunas T e U (sem quebra de linha na célula).
"""

import logging
import re
import shutil
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

from config import (
    M01_LINHA_INICIO,
    KCOR_KRIA_DIRETORIO_TEXTO_CONSERVACAO,
    M04_ACUMULADO,
    M04_ENTRADA,
    M04_NOME_SAIDA,
    M04_SAIDA,
    NUM_COLUNAS_KCOR_KRIA,
    PRAZO_DIAS_APOS_ENVIO,
    RODOVIA_NOME_SEPARAR,
    RODOVIAS,
    SERVICO_NC,
    CABECALHO_KCOR_KRIA,
    resolver_template_acumulado_kcor_kria,
)
from utils.helpers import (
    EXPORTAR_KARTADO_MAE_SUBDIR,
    formatar_numero,
    garantir_pasta,
    km_formato_arquivo,
    km_mais_metros,
    normalizar_rodovia_eaf,
    parse_data,
    preservar_ooxml_planilha_pos_openpyxl,
    resolver_path_ficheiro_ci,
    str_caminho_io_windows,
)

logger = logging.getLogger(__name__)

NUM_COLUNAS = NUM_COLUNAS_KCOR_KRIA
_CABECALHO_ORDEM = tuple(CABECALHO_KCOR_KRIA)
_COL_DATA_SOLICITACAO = 13  # col M, data para sufixo do nome (macro Art_04)

_SIDE_THIN = Side(style="thin", color="000000")
_BORDA_PADRAO = Border(
    left=_SIDE_THIN, right=_SIDE_THIN, top=_SIDE_THIN, bottom=_SIDE_THIN
)


def _texto_sem_quebra_linha(val):
    """T/U: uma linha (remove \\r\\n / \\n na gravação do acumulado)."""
    if val is None:
        return None
    if not isinstance(val, str):
        return val
    t = val.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    return " ".join(t.split())


def _texto_parece_referencia_pdf(s: str) -> bool:
    """
    Valor da col. sequência foto: número/código curto → entra em «pdf (ref).jpg».
    Nome próprio (ex.: fiscal na col. errada) → não usar como ref; fallback código fiscalização.
    """
    t = (s or "").strip()
    if not t or len(t) > 96:
        return False
    partes = [p for p in t.split() if p]
    if len(partes) >= 3 and not re.search(r"\d", t):
        return False
    if (
        len(partes) == 2
        and all(re.fullmatch(r"[A-Za-zÀ-ÿ]{2,}", p) for p in partes)
        and not re.search(r"\d", t)
    ):
        return False
    return True


def _normalizar_header(s: str) -> str:
    """Normaliza nome de coluna para comparação (minúsculo, sem acentos)."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    for old, new in [("ã", "a"), ("á", "a"), ("ç", "c"), ("é", "e"), ("ê", "e"), ("í", "i"), ("ó", "o"), ("ô", "o"), ("ú", "u")]:
        s = s.replace(old, new)
    return s


def _mapear_colunas_pelo_cabecalho(ws) -> list[int]:
    """
    Entrada: ws (planilha), linha 1 = cabeçalhos.
    Saída: lista de 25 int (1-based), índice i = coluna física do i-ésimo cabeçalho canônico (A=0..Y=24).
    Fallback: se cabeçalho não encontrado, usa posição i+1. Aliases: Executor/Executores, Data Envio/Data Solicitação, Arquivo/Arquivos.
    Porquê: arquivos de entrada (M03) podem ter colunas em ordem diferente ou mescladas; saída do M04 deve ser sempre ordem canônica A–Y.
    """
    mapa = {}  # nome_normalizado -> col (1-based)
    for c in range(1, min(ws.max_column + 1, 50)):
        v = _valor_celula(ws, 1, c, preencher_se_merge=True)
        if v is None:
            continue
        n = _normalizar_header(str(v))
        if n and n not in mapa:
            mapa[n] = c
    # Aliases: template pode ter "Executores"/"Executor", "Data Envio"/"Data Solicitação", "Arquivo"/"Arquivos"
    for n, col in list(mapa.items()):
        if n == "executores" and "executor" not in mapa:
            mapa["executor"] = col
        if n == "executor" and "executores" not in mapa:
            mapa["executores"] = col
        if n == "data envio" and "data solicitacao" not in mapa:
            mapa["data solicitacao"] = col
        if n == "data solicitacao" and "data envio" not in mapa:
            mapa["data envio"] = col
        if n == "arquivo" and "arquivos" not in mapa:
            mapa["arquivos"] = col
        if n == "arquivos" and "arquivo" not in mapa:
            mapa["arquivo"] = col
    out = []
    for i, nome in enumerate(_CABECALHO_ORDEM):
        n = _normalizar_header(nome)
        col = mapa.get(n)
        if col is None:
            col = i + 1
        out.append(col)
    return out


def _nome_saida_macro(todos_registros: list, nome_base: str = M04_NOME_SAIDA) -> str:
    """
    Nome do arquivo de saída igual à macro Art_04 (linhas 224-229):
      dia = Left(Data_Solicitação(g - 1), 2)
      mes = Right(Left(Data_Solicitação(g - 1), 5), 2)
      ano = Right(Left(Data_Solicitação(g - 1), 10), 4)
      NameFile = ano & mes & dia & " - " & Format(Now, "hhmmss") & " - Eventos Acumulado Artesp para Exportar Kria.xlsx"
    Ou seja: YYYYMMDD (da data do último registro) - hhmmss (hora atual) - nome_base
    """
    ano, mes, dia = None, None, None
    if todos_registros:
        ultimo = todos_registros[-1]
        if len(ultimo) >= _COL_DATA_SOLICITACAO and ultimo[_COL_DATA_SOLICITACAO - 1] is not None:
            s = str(ultimo[_COL_DATA_SOLICITACAO - 1]).strip()
            # Macro: dia=Left(s,2), mes=Right(Left(s,5),2), ano=Right(Left(s,10),4) → formato DD/MM/YYYY ou DD-MM-YYYY
            if len(s) >= 10:
                dia = s[:2]
                mes = s[3:5]
                ano = s[6:10]
    if ano is None or mes is None or dia is None:
        now = datetime.now()
        ano = now.strftime("%Y")
        mes = now.strftime("%m")
        dia = now.strftime("%d")
    hhmmss = datetime.now().strftime("%H%M%S")
    return f"{ano}{mes}{dia} - {hhmmss} - {nome_base}"


def criar_base_acumulado(caminho: Path) -> None:
    """Planilha acumulada mínima (só cabeçalho); usada quando não há base enviada."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for c, val in enumerate(CABECALHO_KCOR_KRIA, start=1):
        ws.cell(row=1, column=c).value = val
    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str_caminho_io_windows(caminho))
    wb.close()


def _valor_celula(ws, row: int, col: int, preencher_se_merge: bool = False):
    """Valor da célula. Em merge, openpyxl devolve None fora do canto; com preencher_se_merge repete o valor do âncora (útil em W–Y do M03)."""
    for merged_range in ws.merged_cells.ranges:
        if row < merged_range.min_row or row > merged_range.max_row:
            continue
        if col < merged_range.min_col or col > merged_range.max_col:
            continue
        if row != merged_range.min_row or col != merged_range.min_col:
            if preencher_se_merge:
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
            return None
        break
    return ws.cell(row=row, column=col).value


def _aplicar_bordas_linha(ws, row: int, col_fim: int = NUM_COLUNAS):
    """Borda em células da linha 1..col_fim."""
    for col in range(1, col_fim + 1):
        ws.cell(row=row, column=col).border = _BORDA_PADRAO


def _copiar_bordas_linha(ws, row_origem: int, row_destino: int, col_fim: int = NUM_COLUNAS):
    """Copia borda da linha origem para a destino (preserva formatação do template)."""
    for col in range(1, col_fim + 1):
        src = ws.cell(row=row_origem, column=col)
        dst = ws.cell(row=row_destino, column=col)
        if src.border and getattr(src.border, "left", None) is not None:
            dst.border = src.border.copy()
        else:
            dst.border = _BORDA_PADRAO


def _celula_preenchida(val) -> bool:
    """True se o valor existe e não é string vazia (para detectar última linha em A)."""
    if val is None:
        return False
    if isinstance(val, str) and not val.strip():
        return False
    return True


def _ultima_linha_col_a(ws, max_row: int) -> int:
    """
    Última linha com dado na coluna A, igual à macro: Cells(65536, 1).End(xlUp).Row.
    """
    for r in range(max_row, 0, -1):
        if _celula_preenchida(ws.cell(row=r, column=1).value):
            return r
    return 1


def _ultima_linha_abc(ws, max_row: int) -> int:
    """Última linha com dado em A, B ou C (fallback quando A vem vazio, ex.: fórmulas com data_only=True)."""
    for r in range(max_row, 0, -1):
        for col in (1, 2, 3):
            if _celula_preenchida(ws.cell(row=r, column=col).value):
                return r
    return 1


def _eh_cabecalho(linha: list) -> bool:
    """True se a linha for o cabeçalho (A = 'NumItem')."""
    if not linha:
        return False
    a = linha[0]
    return a is not None and str(a).strip().upper() == "NUMITEM"


def _ultima_linha_qualquer_col(ws, max_row: int, colunas: tuple = (1, 2, 3, 4, 5)) -> int:
    """Última linha com dado em qualquer uma das colunas (ex.: A–E)."""
    for r in range(max_row, 0, -1):
        for col in colunas:
            if _celula_preenchida(ws.cell(row=r, column=col).value):
                return r
    return 1


def _obter_planilha_e_ultima(wb):
    """
    Retorna (ws, ultima): planilha e última linha a ler.
    Igual à macro: ultimalinhaprov = Cells(65536, 1).End(xlUp).Row — só coluna A.
    Se ativa der ultima<=1, tenta aba 'Dados'; se ainda 1, usa max_row.
    """
    ws = wb.active
    max_row = ws.max_row
    ultima = _ultima_linha_col_a(ws, max_row)
    if ultima <= 1 and len(wb.worksheets) > 1:
        for sheet in wb.worksheets:
            if sheet.title and "dados" in sheet.title.lower():
                mr = sheet.max_row
                u = _ultima_linha_col_a(sheet, mr)
                if u > 1:
                    return sheet, u
                if mr >= 2:
                    return sheet, mr
    if ultima <= 1 and max_row >= 2:
        ultima = max_row
    return ws, ultima


def _ler_arquivo(caminho: Path) -> list[list]:
    """
    Lê .xlsx: linhas 2 até última com dado na coluna A.
    Colunas são lidas pelo CABEÇALHO (linha 1), não pela posição fixa, para não
    remontar dados em colunas erradas quando o arquivo tem ordem diferente ou mescladas.
    Cada linha vira um registro na ordem canônica A–Y (25 colunas).
    """
    wb = load_workbook(str_caminho_io_windows(caminho), data_only=True)
    ws, ultima = _obter_planilha_e_ultima(wb)
    col_map = _mapear_colunas_pelo_cabecalho(ws)  # [col_A, col_B, ...] 1-based

    linhas = []
    for r in range(2, ultima + 1):
        # Ler cada valor na coluna correta pelo nome do cabeçalho (ordem canônica)
        # W,X,Y: preencher_se_merge (fontes M03 podem ter merge)
        linha = []
        for i in range(NUM_COLUNAS):
            col = col_map[i] if i < len(col_map) else (i + 1)
            preencher = i >= 22  # Arquivos, Indicador, Unidade
            linha.append(_valor_celula(ws, r, col, preencher_se_merge=preencher))
        if _eh_cabecalho(linha):
            continue
        while len(linha) < NUM_COLUNAS:
            linha.append(None)
        linhas.append(linha[:NUM_COLUNAS])
    wb.close()
    return linhas


def executar(pasta_entrada: Path | None = None,
             arquivo_acumulado: Path | None = None,
             pasta_saida: Path | None = None,
             nome_saida: str | None = None,
             nome_arquivo_completo: str | None = None,
             callback_progresso=None,
             arquivos_entrada: list[Path] | None = None) -> Path | None:
    """
    M04 Juntar: consolida .xlsx Kcor-Kria individuais numa planilha acumulada (uma linha por registro, A–Y).
    Entrada: pasta_entrada (ou arquivos_entrada) com .xlsx; arquivo_acumulado = base existente (cabeçalho + dados); pasta_saida.
    Saída: Path do .xlsx gerado em pasta_saida. nome_arquivo_completo sobrescreve nome gerado por data.
    Retorno None: nenhum .xlsx em entrada, ou arquivo_acumulado não existe (quando obrigatório).
    """
    pasta_entrada     = pasta_entrada     or M04_ENTRADA
    arquivo_acumulado = arquivo_acumulado or M04_ACUMULADO
    pasta_saida       = pasta_saida       or M04_SAIDA
    garantir_pasta(pasta_saida)

    if arquivos_entrada is not None:
        arquivos = sorted([
            Path(f) for f in arquivos_entrada
            if Path(f).exists()
            and Path(f).suffix.lower() == ".xlsx"
            and not Path(f).name.startswith("~")
            and "Acumulado" not in Path(f).name
            and not Path(f).name.startswith("_")
        ])
    else:
        arquivos = sorted([
            f for f in pasta_entrada.glob("*.xlsx")
            if not f.name.startswith("~")
            and "Acumulado" not in f.name
            and not f.name.startswith("_")
        ])

    if not arquivos:
        logger.warning(f"Nenhum .xlsx encontrado em: {pasta_entrada}")
        return None

    logger.info(f"Encontrados {len(arquivos)} arquivo(s) para consolidar.")
    todos_registros: list[list] = []
    for idx, arq in enumerate(arquivos):
        if callback_progresso:
            callback_progresso(idx + 1, len(arquivos), f"Lendo: {arq.name[:60]}")
        logger.info(f"Lendo: {arq.name}")
        registros = _ler_arquivo(arq)
        if not registros:
            logger.warning("  %s: 0 registro(s) (verifique se a planilha tem dados na linha 2+, coluna A ou B/C)", arq.name)
        todos_registros.extend(registros)
        logger.info(f"  {len(registros)} registro(s) lido(s).")

    if not todos_registros:
        logger.warning("Nenhum registro encontrado nos arquivos.")
        return None

    logger.info(f"Total de registros a consolidar: {len(todos_registros)}")
    if not arquivo_acumulado.exists():
        logger.warning("Acumulado não informado. Envie o arquivo acumulado (relatório da rede) para consolidar.")
        return None

    wb_acum = load_workbook(str_caminho_io_windows(arquivo_acumulado))
    ws_acum = None
    for sheet in wb_acum.worksheets:
        a1 = sheet.cell(row=1, column=1).value
        if a1 is not None and "numitem" in str(a1).strip().lower():
            ws_acum = sheet
            break
    if ws_acum is None:
        ws_acum = wb_acum.worksheets[0]
        logger.debug("Usando primeira planilha (cabeçalho 'NumItem' não encontrado em A1).")

    logger.info(
        f"Acumulado: planilha '{ws_acum.title}'. "
        f"Gravando {len(todos_registros)} registro(s) nas colunas A–Y a partir da linha 2."
    )
    max_row_acum = ws_acum.max_row
    N = len(todos_registros)

    from . import inserir_nc_kria as ink_exec

    for idx, registro in enumerate(todos_registros):
        row = 2 + idx
        ink_exec._desfazer_merge_colunas_linha(ws_acum, row, 17, 25)
        ws_acum.cell(row=row, column=1).value = idx + 1  # A = contagem
        for col in range(2, NUM_COLUNAS + 1):
            val = registro[col - 1] if (col - 1) < len(registro) else None
            if col in (20, 21):
                val = _texto_sem_quebra_linha(val)
            ws_acum.cell(row=row, column=col).value = val
        ink_exec._forcar_texto_so_data_kcor_cols_m_r(ws_acum, row)
    for row in range(2, 2 + N):
        _aplicar_bordas_linha(ws_acum, row)
    for r in range(2 + N, max_row_acum + 1):
        for c in range(1, NUM_COLUNAS + 1):
            ws_acum.cell(row=r, column=c).value = None
        _aplicar_bordas_linha(ws_acum, r)

    # Nome saída: macro Art_04 (YYYYMMDD - hhmmss - Eventos Acumulado...)
    if nome_arquivo_completo and nome_arquivo_completo.strip():
        nome_arq_saida = nome_arquivo_completo.strip()
        if not nome_arq_saida.lower().endswith(".xlsx"):
            nome_arq_saida += ".xlsx"
    else:
        nome_base = nome_saida if nome_saida else M04_NOME_SAIDA
        nome_arq_saida = _nome_saida_macro(todos_registros, nome_base)
    destino = pasta_saida / nome_arq_saida
    garantir_pasta(destino.parent)

    wb_acum.active = ws_acum
    wb_acum.save(str_caminho_io_windows(destino))
    wb_acum.close()
    preservar_ooxml_planilha_pos_openpyxl(arquivo_acumulado, destino)
    logger.info(f"Módulo 04 concluído. Acumulado salvo: {destino.name}")

    if callback_progresso:
        callback_progresso(len(arquivos), len(arquivos), "Módulo 04 concluído.")

    return destino


def _ultima_linha_dados_eaf_codigo(ws, linha_inicio: int, col_codigo: int = 3) -> int:
    for r in range(ws.max_row, linha_inicio - 1, -1):
        if ws.cell(row=r, column=col_codigo).value:
            return r
    return linha_inicio - 1


def _linha_inicio_dados_por_ficheiro(ws, col_codigo: int) -> int:
    """
    Planilha-mãe EAF: dados a partir de M01_LINHA_INICIO (5).
    Saída M01 Kartado (1 NC por ficheiro): código col C costuma estar na linha 2.
    """
    v5 = ws.cell(row=M01_LINHA_INICIO, column=col_codigo).value
    if v5 is not None and str(v5).strip():
        return M01_LINHA_INICIO
    for cand in (2, 3, 4):
        v = ws.cell(row=cand, column=col_codigo).value
        if v is not None and str(v).strip():
            return cand
    return M01_LINHA_INICIO


def _iter_xlsx_xls_em_pasta(pasta: Path, ex_ok: tuple[str, ...]) -> list[Path]:
    """Ficheiros na raiz e um nível de subpastas (Exportar/, input/ com subdiretórios)."""
    seen: set[str] = set()
    out: list[Path] = []

    def add(f: Path) -> None:
        if not f.is_file() or f.name.startswith("~"):
            return
        if f.suffix.lower() not in ex_ok:
            return
        try:
            k = str(f.resolve())
        except OSError:
            k = str(f)
        if k not in seen:
            seen.add(k)
            out.append(f)

    if not pasta.is_dir():
        return []
    try:
        for f in pasta.iterdir():
            add(f)
            if f.is_dir():
                if f.name.casefold() == EXPORTAR_KARTADO_MAE_SUBDIR:
                    continue
                try:
                    for g in f.iterdir():
                        add(g)
                except OSError:
                    pass
    except OSError:
        return []
    return sorted(out, key=lambda p: p.as_posix().lower())


def _str_eaf(v) -> str:
    return str(v).strip() if v is not None else ""


def _km_celulas_eaf(ws, row: int, col_km: int, col_m: int) -> str:
    km = ws.cell(row=row, column=col_km).value
    m = ws.cell(row=row, column=col_m).value
    if km is None and (m is None or m == ""):
        return ""
    try:
        return km_mais_metros(km, m).replace(" ", "")
    except Exception:
        return str(km or "").replace(" ", "")


def _eaf_linha_para_registro_kcor(
    ws,
    row: int,
    *,
    col_codigo: int | None = None,
    col_rodovia: int | None = None,
    col_km_i_full: int | None = None,
    col_km_i_m: int | None = None,
    col_km_f_full: int | None = None,
    col_km_f_m: int | None = None,
    col_sentido: int | None = None,
    col_data_con: int | None = None,
    col_data_reparo: int,
    col_data_envio: int,
    col_tipo_nc: int,
    col_seq_foto: int,
) -> list | None:
    """
    Uma linha da planilha-mãe EAF → 25 valores (ordem CABECALHO_KCOR_KRIA).
    Coluna A do acumulado é sequencial no M04.
    V/W seguem o padrão Art_03: Diretório = KCOR_KRIA_DIRETORIO_TEXTO_CONSERVACAO (rede, não Path local);
    Arquivos = «jpg gerado»;«pdf (ref).jpg»
    (ref = coluna sequência foto na EAF/Exportar, por cabeçalho; se vazia ou ≤0, código fiscalização ou número Roti).
    """
    from . import inserir_nc_kria as ink
    from .analisar_pdf_ma import _sentido_para_texto
    from . import separar_nc as sn

    col_codigo = col_codigo or sn.COL_CODIGO
    col_rodovia = col_rodovia or sn.COL_RODOVIA
    col_km_i_full = col_km_i_full or sn.COL_KM_I_FULL
    col_km_i_m = col_km_i_m or sn.COL_KM_I_M
    col_km_f_full = col_km_f_full or sn.COL_KM_F_FULL
    col_km_f_m = col_km_f_m or sn.COL_KM_F_M
    col_sentido = col_sentido or sn.COL_SENTIDO
    col_data_con = col_data_con or sn.COL_DATA_CON

    if not sn._cell(ws, row, col_codigo):
        return None

    descricao = sn._valor_tipo_nc(ws, row, col_tipo_nc)
    desc_s = _str_eaf(descricao)
    nc_info = SERVICO_NC.get(desc_s, None)
    if nc_info:
        serv_nc, classifica_srv, executor = nc_info
        classifica = (str(classifica_srv).strip() if classifica_srv else "") or "Conservação Rotina"
    else:
        serv_nc = ""
        classifica = "Conservação Rotina"
        executor = "Soluciona - Conserva"

    rod_raw = _str_eaf(sn._cell(ws, row, col_rodovia))
    rod_info = normalizar_rodovia_eaf(rod_raw, RODOVIAS)
    if rod_info.get("tag") and rod_info["tag"] != "FORA":
        rod_display = RODOVIA_NOME_SEPARAR.get(rod_info["tag"], rod_info.get("codigo") or rod_raw.strip()).strip()
        rod_for_form = (rod_info.get("codigo") or rod_raw.strip() or " ").strip()
    else:
        rod_display = rod_raw.strip() or " "
        rod_for_form = rod_display
    rod_tag, _rod_cod, _n = ink._normalizar_rodovia_formulario(rod_for_form or " ")

    kmi = _km_celulas_eaf(ws, row, col_km_i_full, col_km_i_m or sn.COL_KM_I_M)
    kmf = _km_celulas_eaf(ws, row, col_km_f_full, col_km_f_m or sn.COL_KM_F_M) or kmi

    sentido_raw = sn._cell(ws, row, col_sentido)
    sentido_txt = _sentido_para_texto(_str_eaf(sentido_raw)) or ""

    dt = parse_data(sn._cell(ws, row, col_data_con))
    dt_str = dt.strftime("%d/%m/%Y") if dt else ""

    data_reparo_raw = sn._cell(ws, row, col_data_reparo)
    data_reparo_dt = parse_data(data_reparo_raw)
    emb_raw = sn._cell(ws, row, col_data_envio)
    emb = parse_data(emb_raw)
    emb_str = emb.strftime("%d/%m/%Y") if emb else ""

    if data_reparo_dt:
        data_reparo_str = data_reparo_dt.strftime("%d/%m/%Y")
    elif dt:
        data_reparo_str = (dt + timedelta(days=PRAZO_DIAS_APOS_ENVIO)).strftime("%d/%m/%Y")
    else:
        data_reparo_str = emb_str

    if data_reparo_raw is not None and emb_raw is not None and _str_eaf(emb_raw) == _str_eaf(data_reparo_raw):
        envio_str = dt_str
    else:
        envio_str = emb_str

    prazo_val = ""
    if data_reparo_dt and dt:
        try:
            prazo_val = (data_reparo_dt - dt).days
        except TypeError:
            prazo_val = ""

    comp_cell = sn._cell(ws, row, col_seq_foto)
    complemento = ""
    foto_ref_pdf = None
    try:
        foto_ref_pdf = int(float(str(comp_cell).replace(",", ".").strip()))
    except (ValueError, TypeError):
        complemento = _str_eaf(comp_cell)
        if complemento and _texto_parece_referencia_pdf(complemento):
            foto_ref_pdf = complemento
        elif complemento:
            foto_ref_pdf = None

    codigo = _str_eaf(sn._cell(ws, row, col_codigo))
    relatorio = ""
    obs_gestor_txt = ink._obs_gestor(relatorio, codigo)
    observacoes_u = ink._observacoes(desc_s, complemento, str(emb_raw) if emb_raw is not None else "")

    # V/W — alinhado a inserir_nc_kria._processar_arquivo (macro Art_03)
    data_raw_str = _str_eaf(sn._cell(ws, row, col_data_con))
    num_roti = (
        foto_ref_pdf
        if isinstance(foto_ref_pdf, int) and foto_ref_pdf > 0
        else 1
    )
    numero_6 = formatar_numero(num_roti, 6)
    nome_jpg = ink._nome_arquivo_jpg(
        data_raw_str,
        _n,
        numero_6,
        rod_tag,
        km_formato_arquivo(kmi),
        _str_eaf(sentido_raw),
        nome_origem="",
    )
    pdf_ref = foto_ref_pdf
    if pdf_ref is None or pdf_ref == "" or (isinstance(pdf_ref, int) and pdf_ref <= 0):
        pdf_ref = codigo if codigo else num_roti
    pdf_nome = f"pdf ({pdf_ref}).jpg" if pdf_ref not in (None, "") else ""
    if nome_jpg and pdf_nome:
        arquivos_w = f"{nome_jpg};{pdf_nome}"
    else:
        arquivos_w = nome_jpg or pdf_nome
    diretorio_v = KCOR_KRIA_DIRETORIO_TEXTO_CONSERVACAO

    linha = [None] * NUM_COLUNAS
    linha[0] = None
    linha[1] = "Artesp"
    linha[2] = "2"
    linha[3] = classifica
    linha[4] = desc_s or serv_nc
    linha[5] = rod_display
    linha[6] = kmi
    linha[7] = kmf
    linha[8] = sentido_txt[:120]
    linha[9] = ""
    linha[10] = "Conservação"
    linha[11] = executor
    linha[12] = dt_str
    linha[13] = ""
    linha[14] = dt_str
    linha[15] = envio_str
    linha[16] = ""
    linha[17] = ""
    linha[18] = ink._prazo_para_numero(prazo_val)
    linha[19] = obs_gestor_txt
    linha[20] = observacoes_u
    linha[21] = diretorio_v
    linha[22] = arquivos_w or ""
    linha[23] = ""
    linha[24] = ""
    return linha


def _chave_dedupe_registro_kcor(reg: list) -> tuple:
    """Identifica linhas equivalentes (ex.: mesma NC no consolidado e no .xlsx unitário). Inclui código (col C)."""
    if not reg or len(reg) < 23:
        return ()
    r = reg[:NUM_COLUNAS] if len(reg) > NUM_COLUNAS else reg
    cod_f = str(reg[NUM_COLUNAS]).strip().casefold() if len(reg) > NUM_COLUNAS else ""
    return (
        cod_f,
        str(r[4] or "").strip().casefold(),
        str(r[5] or "").strip().casefold(),
        str(r[6] or "").strip().casefold(),
        str(r[7] or "").strip().casefold(),
        str(r[8] or "").strip().casefold(),
        str(r[12] or "").strip().casefold(),
        str(r[15] or "").strip().casefold(),
        str(r[22] or "").strip().casefold(),
    )


def _eaf_workbook_para_registros(fpath: Path, sn) -> list[list]:
    """Lê um EAF/Kartado .xlsx/.xls e devolve as linhas Kcor-Kria (lista de 25 colunas)."""
    out: list[list] = []
    p = sn._converter_xls_para_xlsx(fpath) if fpath.suffix.lower() == ".xls" else fpath
    wb = load_workbook(str_caminho_io_windows(p), data_only=True)
    try:
        ws = wb.active
        col_codigo = sn._detectar_col_codigo_fiscalizacao(ws, fallback=sn.COL_CODIGO)
        linha_ini = _linha_inicio_dados_por_ficheiro(ws, col_codigo)
        ultima = _ultima_linha_dados_eaf_codigo(ws, linha_ini, col_codigo)
        if ultima < linha_ini:
            return out
        col_data_con = sn._detectar_col_data_con(ws, fallback=sn.COL_DATA_CON)
        col_data_reparo = sn._detectar_col_data_reparo(ws, fallback=sn.COL_DATA_NC)
        col_data_envio = sn._detectar_col_data_envio(ws, fallback=19)
        col_tipo_nc = sn._detectar_col_tipo_nc(ws, fallback=sn.COL_TIPO_NC)
        col_rodovia = sn._detectar_col_rodovia(ws, fallback=sn.COL_RODOVIA)
        col_km_i_full = sn._detectar_col_km_inicial(ws, fallback=sn.COL_KM_I_FULL)
        col_km_i_m = sn._detectar_col_km_i_metros(ws, fallback=sn.COL_KM_I_M)
        col_km_f_full = sn._detectar_col_km_final(ws, fallback=sn.COL_KM_F_FULL)
        col_km_f_m = sn._detectar_col_km_f_metros(ws, fallback=sn.COL_KM_F_M)
        col_sentido = sn._detectar_col_sentido(ws, fallback=sn.COL_SENTIDO)
        col_seq_foto = sn._detectar_col_seq_foto(ws, fallback=sn.COL_SEQ_FOTO)
        col_resp = sn._detectar_col_responsavel(ws, fallback=sn.COL_RESPONSAVEL)
        if col_seq_foto == col_resp:
            col_seq_foto = sn.COL_SEQ_FOTO
        for r in range(linha_ini, ultima + 1):
            reg = _eaf_linha_para_registro_kcor(
                ws,
                r,
                col_codigo=col_codigo,
                col_rodovia=col_rodovia,
                col_km_i_full=col_km_i_full,
                col_km_i_m=col_km_i_m,
                col_km_f_full=col_km_f_full,
                col_km_f_m=col_km_f_m,
                col_sentido=col_sentido,
                col_data_con=col_data_con,
                col_data_reparo=col_data_reparo,
                col_data_envio=col_data_envio,
                col_tipo_nc=col_tipo_nc,
                col_seq_foto=col_seq_foto,
            )
            if reg:
                cod = str(sn._cell(ws, r, col_codigo) or "").strip()
                out.append(list(reg) + [cod])
        return out
    finally:
        wb.close()


def _ficheiro_e_ponteiro_git_lfs(path: Path) -> bool:
    """True se o ficheiro no disco é só o ponteiro LFS (build sem `git lfs pull`)."""
    try:
        with open(path, "rb") as f:
            head = f.read(120)
        return head.startswith(b"version https://git-lfs")
    except OSError:
        return False


def gerar_acumulado_kcor_kria_desde_pasta_eaf(
    pasta_eaf: Path,
    out_path: Path,
    caminho_template: Path | None = None,
) -> bool:
    """
    Gera acumulado Kcor-Kria a partir de EAF em pasta_eaf (dados a partir de M01_LINHA_INICIO, col C).
    Usado pelo pipeline web quando não há M03.
    Se existir planilha com várias NCs e muitos ficheiros de 1 NC (mesmo volume), ignora os unitários
    e remove linhas duplicadas só entre **vários** ficheiros (código fiscalização + colunas-chave).
    """
    from . import separar_nc as sn

    if not pasta_eaf.is_dir():
        return False

    tpl = None
    if caminho_template:
        tpl = resolver_path_ficheiro_ci(caminho_template)
        if not tpl.is_file():
            tpl = None
    if tpl is None:
        tpl = resolver_template_acumulado_kcor_kria()
    if tpl is not None:
        tpl = resolver_path_ficheiro_ci(tpl)
    if tpl is None or not tpl.is_file():
        logger.warning(
            "Template acumulado Kcor-Kria não encontrado. "
            "Coloque Acumulado.xlsx ou _Planilha Modelo Kcor-Kria.xlsx em nc_artesp/assets/templates/ "
            "ou defina ARTESP_M04_TEMPLATE_ACUMULADO_KCOR_KRIA."
        )
        return False
    if _ficheiro_e_ponteiro_git_lfs(tpl):
        logger.warning(
            "Template Kcor-Kria em %s é ponteiro Git LFS (conteúdo real não está no disco). "
            "Ative Git LFS no deploy ou defina ARTESP_M04_TEMPLATE_ACUMULADO_KCOR_KRIA para um .xlsx real.",
            tpl,
        )
        return False

    ex_ok = (".xlsx", ".xlsm", ".xls")
    arquivos = _iter_xlsx_xls_em_pasta(pasta_eaf, ex_ok)
    if not arquivos:
        logger.warning("Acumulado Kcor-Kria: nenhum .xlsx/.xls em %s", pasta_eaf)
        return False

    blocos: list[tuple[Path, list[list]]] = []
    for fpath in arquivos:
        try:
            regs = _eaf_workbook_para_registros(fpath, sn)
            if regs:
                blocos.append((fpath, regs))
        except Exception as exc:
            logger.warning("Acumulado EAF→Kcor: ignorar %s (%s)", fpath.name, exc)
            continue

    if not blocos:
        logger.warning(
            "Acumulado Kcor-Kria: %s ficheiro(s) lidos em %s, mas nenhuma linha com código (col C) válido.",
            len(arquivos),
            pasta_eaf,
        )
        return False

    any_multi = any(len(regs) > 1 for _, regs in blocos)
    any_single = any(len(regs) == 1 for _, regs in blocos)
    max_rows = max(len(regs) for _, regs in blocos)
    n_single_files = sum(1 for _, regs in blocos if len(regs) == 1)
    if (
        any_multi
        and any_single
        and max_rows >= 5
        and n_single_files >= max_rows
    ):
        n_skip = n_single_files
        blocos = [(p, r) for p, r in blocos if len(r) > 1]
        logger.info(
            "Acumulado Kcor-Kria: ignorados %d ficheiro(s) de 1 NC — há planilha com %d linhas "
            "(padrão consolidado + mesmas NCs em ficheiros separados).",
            n_skip,
            max_rows,
        )

    todos: list[list] = []
    for _, regs in blocos:
        todos.extend(regs)

    n_antes = len(todos)
    if len(blocos) > 1:
        visto: set[tuple] = set()
        todos_dedup: list[list] = []
        for r in todos:
            k = _chave_dedupe_registro_kcor(r)
            if k in visto:
                continue
            visto.add(k)
            todos_dedup.append(r)
        todos = todos_dedup
        if len(todos) < n_antes:
            logger.info(
                "Acumulado Kcor-Kria: removidas %d linha(s) duplicada(s) (código fiscalização + colunas-chave).",
                n_antes - len(todos),
            )

    if not todos:
        logger.warning(
            "Acumulado Kcor-Kria: %s ficheiro(s) lidos em %s, mas nenhuma linha com código (col C) válido.",
            len(arquivos),
            pasta_eaf,
        )
        return False

    garantir_pasta(out_path.parent)
    shutil.copy2(tpl, out_path)

    from . import inserir_nc_kria as ink

    wb_acum = load_workbook(str_caminho_io_windows(out_path))
    ws_acum = None
    for sheet in wb_acum.worksheets:
        a1 = sheet.cell(row=1, column=1).value
        if a1 is not None and "numitem" in str(a1).strip().lower():
            ws_acum = sheet
            break
    if ws_acum is None:
        ws_acum = wb_acum.worksheets[0]
        logger.debug("Acumulado EAF: primeira aba (A1 sem NumItem).")

    wb_acum.active = ws_acum
    max_row_acum = ws_acum.max_row
    N = len(todos)

    for idx, registro in enumerate(todos):
        row = 2 + idx
        reg_write = registro[:NUM_COLUNAS] if len(registro) > NUM_COLUNAS else registro
        ink._desfazer_merge_colunas_linha(ws_acum, row, 17, 25)
        ws_acum.cell(row=row, column=1).value = idx + 1
        for col in range(2, NUM_COLUNAS + 1):
            val = reg_write[col - 1] if (col - 1) < len(reg_write) else None
            if col in (20, 21):
                val = _texto_sem_quebra_linha(val)
            ws_acum.cell(row=row, column=col).value = val
        ink._forcar_texto_so_data_kcor_cols_m_r(ws_acum, row)
        _aplicar_bordas_linha(ws_acum, row)

    for r in range(2 + N, max_row_acum + 1):
        for c in range(1, NUM_COLUNAS + 1):
            ws_acum.cell(row=r, column=c).value = None
        _aplicar_bordas_linha(ws_acum, r)

    wb_acum.save(str_caminho_io_windows(out_path))
    wb_acum.close()
    preservar_ooxml_planilha_pos_openpyxl(tpl, out_path)
    logger.info("Acumulado Kcor-Kria (EAF) gravado: %s (%s linha(s))", out_path.name, N)
    return True
