r"""
modulos/salvar_imagem.py
────────────────────────────────────────────────────────────────────────────
Equivalente VBA: Salvar_IMG_NC_Artesp_Pasta_Sep

Para cada linha da planilha acumulada:
  - Le tipo NC (col E), rodovia (col F), KM (col G), sentido (col I),
    data solicitacao (col M), prazo (col P), evento (col T), nr Kria (col Y),
    e nome do arquivo de imagem (col W, primeira parte antes do ';').
  - Copia o JPG para:
      D:\Apontamentos NC Artesp - Imagens Classificadas\{Tipo}\{nome}.jpg
  - Se o tipo for Pav. - Depressao ou Pav. - Pano de Rolamento,
    copia tambem para a subpasta _Exportar\.

Nome do arquivo de destino:
  rodovia - Sentido - km,metro - yyyymmdd - ddmmaaaa - evento.jpg
"""

import logging
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from config import (
    M08_IMAGENS_SRC,
    M08_DESTINO,
    M08_EXPORTAR,
    M08_TIPOS_EXPORTAR,
    M08_TIPO_NOME_PASTA,
)
from utils.helpers import (
    garantir_pasta,
    copiar_arquivo,
    sanitizar_nome,
    parse_data,
    data_yyyymmdd,
)

logger = logging.getLogger(__name__)

import io
import zipfile as _zipfile

# Índices de coluna (1-based)
COL_TIPO    = 5   # E
COL_RODOVIA = 6   # F
COL_KMI     = 7   # G
COL_SENTIDO = 9   # I
COL_DT_SOL  = 13  # M
COL_PRAZO   = 16  # P
COL_EVENTO  = 20  # T
COL_ARQUIVO = 23  # W
COL_NUMERO  = 25  # Y


def _cell(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""


def _km_para_virgula(km_str: str) -> str:
    """Substitui '+' por ',' no KM."""
    return km_str.replace("+", ",")


def _tipo_para_pasta(tipo: str) -> str:
    """Nome da pasta por tipo (VBA: mapeamentos específicos + / → _)."""
    return M08_TIPO_NOME_PASTA.get(tipo, tipo.replace("/", "_"))


def _extrair_evento(texto_evento: str, numero_kria: str) -> str:
    """
    Extrai o número/descrição do evento do campo T.
    Formato esperado: '... NC: XXXXXX ...'
    Concatena com número Kria (col Y).
    """
    match = re.search(r"NC:\s*(.+)", texto_evento, re.IGNORECASE)
    nc_parte = match.group(1).strip() if match else texto_evento.strip()
    if numero_kria:
        return f"{nc_parte} - {numero_kria}"
    return nc_parte


def organizar_imagens_bytes(xlsx_bytes: bytes, zip_imagens_bytes: bytes) -> tuple[bytes, int]:
    """
    Versão web do M08: recebe a planilha acumulada e um ZIP com as imagens.
    Organiza as imagens em subpastas por tipo de NC e retorna um novo ZIP.
    Replica exatamente o comportamento da macro Salvar_IMG_NC_Artesp_Pasta_Sep:
      • Subpasta por tipo de NC
      • Nome do arquivo: rodovia - sentido - km,metro - yyyymmdd - ddmmaaaa - evento.jpg
      • Tipos em M08_TIPOS_EXPORTAR também vão para a pasta _Exportar/
    Retorna (zip_bytes, n_copiadas).
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    ultima = ws.max_row
    for r in range(ultima, 1, -1):
        if ws.cell(row=r, column=1).value is not None:
            ultima = r
            break

    def cell(r: int, c: int) -> str:
        v = ws.cell(row=r, column=c).value
        return str(v).strip() if v is not None else ""

    # Indexa o ZIP de imagens por nome de arquivo (case-insensitive)
    imagens: dict[str, bytes] = {}
    with _zipfile.ZipFile(io.BytesIO(zip_imagens_bytes)) as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            nome = Path(info.filename).name
            imagens[nome.lower()] = zf.read(info)

    saida = _zipfile.ZipFile(io.BytesIO(), "w", _zipfile.ZIP_DEFLATED)
    buf   = io.BytesIO()
    saida = _zipfile.ZipFile(buf, "w", _zipfile.ZIP_DEFLATED)
    copiadas  = 0
    seen: set[str] = set()

    def _dest_unico(path: str) -> str:
        base = path
        n = 1
        while base in seen:
            stem, ext = path.rsplit(".", 1) if "." in path else (path, "")
            base = f"{stem} ({n}).{ext}" if ext else f"{stem} ({n})"
            n += 1
        seen.add(base)
        return base

    for r in range(2, ultima + 1):
        tipo     = cell(r, COL_TIPO)
        rodovia  = cell(r, COL_RODOVIA)
        km_i     = cell(r, COL_KMI)
        sentido  = cell(r, COL_SENTIDO)
        dt_sol   = cell(r, COL_DT_SOL)
        prazo    = cell(r, COL_PRAZO)
        evento_t = cell(r, COL_EVENTO)
        arq_w    = cell(r, COL_ARQUIVO)
        num_kria = cell(r, COL_NUMERO)

        if not arq_w:
            continue

        nome_img = arq_w.split(";")[0].strip()
        if not nome_img:
            continue

        dados = imagens.get(nome_img.lower())
        if dados is None:
            logger.warning(f"Linha {r}: imagem não encontrada no ZIP: {nome_img}")
            continue

        # Normalizar rodovia
        rod_arq = rodovia
        if rod_arq in ("SPI102/300", "SPI-102/300"):
            rod_arq = "SP102"

        # Nome do arquivo destino (igual à macro)
        km_fmt   = _km_para_virgula(km_i)
        dt_sol_o = parse_data(dt_sol)
        dt_pra_o = parse_data(prazo)
        from datetime import datetime as _dt
        if not dt_sol_o: dt_sol_o = _dt.now()
        if not dt_pra_o: dt_pra_o = _dt.now()
        data_sol_s = data_yyyymmdd(dt_sol_o)
        prazo_s    = dt_pra_o.strftime("%d%m%Y")
        evento_s   = _extrair_evento(evento_t, num_kria)
        tipo_pasta = _tipo_para_pasta(tipo)

        nome_dest = sanitizar_nome(
            f"{rod_arq} - {sentido} - {km_fmt} - {data_sol_s} - {prazo_s} - {evento_s}.jpg"
        )

        # Pasta por tipo
        arc_path = _dest_unico(f"{sanitizar_nome(tipo_pasta)}/{nome_dest}")
        saida.writestr(arc_path, dados)
        copiadas += 1

        # Cópia extra para _Exportar (tipos de pavimento)
        if tipo in M08_TIPOS_EXPORTAR:
            arc_exp = _dest_unico(f"_Exportar/{nome_dest}")
            saida.writestr(arc_exp, dados)

    saida.close()
    return buf.getvalue(), copiadas


def executar(arquivo_acumulado: Path,
             pasta_imagens: Path | None = None,
             pasta_destino: Path | None = None,
             pasta_exportar: Path | None = None,
             callback_progresso=None) -> int:
    """
    Processa a planilha acumulada e copia imagens para as pastas classificadas.

    Retorna número de imagens copiadas.
    """
    pasta_imagens  = pasta_imagens  or M08_IMAGENS_SRC
    pasta_destino  = pasta_destino  or M08_DESTINO
    pasta_exportar = pasta_exportar or M08_EXPORTAR

    garantir_pasta(pasta_exportar)

    logger.info(f"Abrindo acumulado: {arquivo_acumulado.name}")
    wb = load_workbook(str(arquivo_acumulado), data_only=True)
    ws = wb.active

    ultima = ws.max_row
    for r in range(ultima, 1, -1):
        if ws.cell(row=r, column=1).value is not None:
            ultima = r
            break

    total = ultima - 1
    copiadas = 0

    for idx, r in enumerate(range(2, ultima + 1)):
        try:
            tipo     = _cell(ws, r, COL_TIPO)
            rodovia  = _cell(ws, r, COL_RODOVIA)
            km_i     = _cell(ws, r, COL_KMI)
            sentido  = _cell(ws, r, COL_SENTIDO)
            dt_sol   = _cell(ws, r, COL_DT_SOL)
            prazo    = _cell(ws, r, COL_PRAZO)
            evento_t = _cell(ws, r, COL_EVENTO)
            arq_w    = _cell(ws, r, COL_ARQUIVO)
            num_kria = _cell(ws, r, COL_NUMERO)

            if callback_progresso:
                callback_progresso(idx + 1, total, f"Processando linha {r}...")

            if not arq_w:
                logger.warning(f"Linha {r}: coluna W vazia, pulando.")
                continue

            # Primeiro arquivo antes do ';'
            nome_img = arq_w.split(";")[0].strip()
            if not nome_img:
                logger.warning(f"Linha {r}: primeiro arquivo vazio (sem foto do PDF), pulando.")
                continue

            origem   = pasta_imagens / nome_img

            if not origem.exists():
                logger.warning(f"Imagem não encontrada: {origem}")
                continue

            rod_arq = rodovia
            if rod_arq in ("SPI102/300", "SPI-102/300"):
                rod_arq = "SP102"

            km_formatado = _km_para_virgula(km_i)

            dt_sol_obj = parse_data(dt_sol)
            dt_pra_obj = parse_data(prazo)
            if not dt_sol_obj:
                dt_sol_obj = datetime.now()
            data_sol_s = data_yyyymmdd(dt_sol_obj)

            # Prazo: ddmmaaaa (sem separadores, formato original do VBA)
            if not dt_pra_obj:
                dt_pra_obj = datetime.now()
            prazo_s = dt_pra_obj.strftime("%d%m%Y")

            evento_s = _extrair_evento(evento_t, num_kria)
            tipo_pasta = _tipo_para_pasta(tipo)

            nome_destino = sanitizar_nome(
                f"{rod_arq} - {sentido} - {km_formatado} - {data_sol_s} - {prazo_s} - {evento_s}.jpg"
            )

            pasta_tipo = pasta_destino / sanitizar_nome(tipo_pasta)
            garantir_pasta(pasta_tipo)
            destino = pasta_tipo / nome_destino

            if copiar_arquivo(origem, destino):
                copiadas += 1

            if tipo in M08_TIPOS_EXPORTAR:
                destino_exp = pasta_exportar / nome_destino
                copiar_arquivo(origem, destino_exp)
        except (OSError, PermissionError) as e:
            logger.warning(f"Linha {r}: permissão negada ou erro de acesso (pasta pode estar em uso). Pulando: {e}")
            continue

    wb.close()
    logger.info(f"Módulo 08 concluído. {copiadas} imagem(ns) copiada(s).")
    if callback_progresso:
        callback_progresso(total, total, "Módulo 08 concluído.")
    return copiadas
