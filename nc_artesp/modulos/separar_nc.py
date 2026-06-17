"""
modulos/01_separar_nc.py
────────────────────────────────────────────────────────────────────────────
Equivalente VBA: Art_011_EAF_Separar_Mod_Exc_NC
Desenvolvedor: Ozeias Engler

A partir da planilha-mãe EAF (única, com todas as NCs do período),
gera ficheiros XLS individuais. Por defeito, só junta várias linhas no mesmo Excel
se **todas** as células da linha (colunas A–última) forem iguais; data de reparo/execução
diferente implica ficheiros separados. Com `um_arquivo_por_nc=True`, uma linha por ficheiro (não consolidado).
No modo Art_011 / Template EAF, por omissão gera-se **um único** .xlsx com todas as NCs ordenadas por rodovia,
atividade e código (desativar com ``unico_arquivo_organizado=False`` ou ``um_arquivo_por_nc=True``).

Fluxo:
  Com M01_COPIA_PLANILHA_MAE=False (Kartado): usa **um** template consolidado
  (``Template - Geral - 4 e 5 - Final.xlsx``, ou ``Template - geral.xlsx`` se só esse existir),
  apaga linhas de dados, preenche **todas** as NCs num único .xlsx (por defeito; desativar com
  ``unico_arquivo_organizado=False`` ou ``um_arquivo_por_nc=True``). A coluna «Classe» do consolidado
  só recebe textos da lista canónica Kartado: cruza M, N, O (tipicamente cols 13–15) e a coluna «Atividade»
  detetada (cabeçalho) com ``ART03_ATIVIDADE_PARA_SERVICO_KARTADO`` e ``KARTADO_RELATORIO_SERVICOS_TODOS``
  (match exacto, compostos M+N+O, depois fuzzy por tokens); sem match seguro → fallback FD com aviso no log.
  Com M01_COPIA_PLANILHA_MAE=True ou ``executar(..., copia_planilha_mae=True)``: macro Art_011 — base ``Template_EAF.xlsx``
  (cabeçalho 1–4, sem dados); cada grupo recebe um ficheiro com cópia **literal** das linhas da mãe
  (valores e estilos de célula, por coluna); só a agregação por rodovia/atividade muda quais linhas vão juntas.
  A mãe continua a ser padronizada (I, K, V) antes da extração.
  API web: ``copia_planilha_mae = not m01_kartado`` (``m01_kartado=false`` → cópia mãe Art_011; ``true`` → templates Kartado).
  PDF de constatação (``<mãe>.pdf`` junto ao Excel ou ``executar(..., pdf_constatacao=…)``): o M01 Kartado classifica
  o texto de «Observação» (``utils/kartado_observacao_pdf.py``): léxico de **Localização Tipo** → col. **Y**;
  **Localização Pista** → col. **X**; restante (texto livre) → coluna **AA** «Observações» (só esse cabeçalho;
  não misturar com **ObsGestor** / observação do gestor). Ex.: «fora de plataforma»
  após «Drenagem» não dispara o rótulo «Fora de Plataforma» (falso positivo evitado).
"""

from __future__ import annotations

import logging
import re
import shutil
import tempfile
import unicodedata
import uuid
from copy import copy
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from functools import lru_cache
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import xlrd

from config import (
    ART03_ATIVIDADE_PARA_SERVICO_KARTADO,
    KARTADO_RELATORIO_SERVICOS_TODOS,
    M01_COPIA_PLANILHA_MAE,
    M01_DICAS_PALAVRA_TEMPLATE_KARTADO,
    M01_EXPORTAR,
    M01_LINHA_INICIO,
    M01_LOTE,
    M01_MAPA_ATIVIDADE_TEMPLATE_KARTADO,
    m01_servico_abrev_art011_lookup,
    M01_TEMPLATE_EAF,
    PRAZO_DIAS_APOS_ENVIO,
    RESPOSTA_PENDENTES_HEADER_MAE_ATIVIDADE_Q,
    RESPOSTA_PENDENTES_HEADER_MAE_TIPO_ATIV_O,
    RODOVIA_NOME_SEPARAR,
    SERVICO_ABREV,
    SERVICO_NC,
    RODOVIAS,
    TEMPLATE_EXPORTAR_ROTINA,
)
from utils.kartado_observacao_pdf import (
    normalizar_observacao_extraida_pdf,
    rotear_observacao_pdf_para_kartado,
)
from utils.helpers import (
    EXPORTAR_KARTADO_MAE_SUBDIR,
    pad_metros,
    parse_data,
    data_yyyymmdd,
    km_formato_arquivo,
    normalizar_rodovia_eaf,
    garantir_pasta,
    encurtar_nome_em_pasta,
    preservar_ooxml_planilha_pos_openpyxl,
    sanitizar_nome,
    str_caminho_io_windows,
    truncar_nome_preservando_sufixo_prazo_m01,
    detectar_coluna_tipo_de_atividade_eaf,
)

logger = logging.getLogger(__name__)


def _str_linha_eaf(val) -> str:
    return str(val).strip() if val is not None else ""


def _sanitizar_nome_xlsx(nome: str, max_stem: int = 380) -> str:
    """sanitizar_nome sem cortar a extensão .xlsx nem o sufixo « - Prazo - data» (limite só no stem)."""
    nome = (nome or "").strip()
    if not nome:
        return ""
    ext = Path(nome).suffix
    if ext.lower() == ".xlsx" and nome.lower().endswith(".xlsx"):
        stem = nome[: -len(ext)]
    else:
        stem, ext = Path(nome).stem, Path(nome).suffix
    # Remove duplicado estilo Windows «…~1» no stem (não entra no padrão macro das constatações).
    stem = re.sub(r"~\d+$", "", stem).rstrip(" -.")
    return sanitizar_nome(stem, max_len=max_stem) + (ext if ext else ".xlsx")


@contextmanager
def abrir_workbook(path: Path, **kwargs):
    """Context manager para garantir fechamento do workbook."""
    wb = load_workbook(str_caminho_io_windows(path), **kwargs)
    try:
        yield wb
    finally:
        wb.close()


class ValidadorArquivoEAF:
    """Valida arquivo de entrada da etapa M01."""

    EXTENSOES_VALIDAS = {".xls", ".xlsx", ".xlsm", ".xltx", ".xltm"}

    @staticmethod
    def validar(arquivo: Path) -> None:
        if not arquivo.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {arquivo}")
        if arquivo.is_dir():
            raise ValueError(
                "O caminho informado é uma pasta, não um arquivo. "
                "Selecione o arquivo Excel da planilha EAF."
            )
        if arquivo.suffix.lower() == ".pdf":
            raise ValueError(
                "O arquivo selecionado é um PDF. O passo [1/6] Separando NCs exige a PLANILHA EXCEL (EAF), não o PDF."
            )


# TEMPLATE EAF — planilha base para os arquivos gerados (cabeçalho 1–4; dados a partir da 5)
# Ordem: ARTESP_TEMPLATE_EAF (se definido) → nc_artesp/assets/templates/
# Aceita: Template_EAF.xlsx ou Template_EAF.xlsx.xlsx
def _caminho_template_eaf() -> Path:
    """Retorna o Path do template EAF. Procura em nc_artesp/assets/templates/."""
    # 1. ARTESP_TEMPLATE_EAF (ficheiro ou pasta)
    if M01_TEMPLATE_EAF.is_file():
        return M01_TEMPLATE_EAF
    if M01_TEMPLATE_EAF.is_dir():
        for n in ("Template_EAF.xlsx", "Template_EAF.xlsx.xlsx"):
            c = M01_TEMPLATE_EAF / n
            if c.is_file():
                return c
    # 2. Pacote nc_artesp
    _nc = Path(__file__).resolve().parent.parent
    pasta = _nc / "assets" / "templates"
    for nome in ("Template_EAF.xlsx", "Template_EAF.xlsx.xlsx"):
        candidato = pasta / nome
        if candidato.is_file():
            return candidato
    return M01_TEMPLATE_EAF  # usado na mensagem de erro se não encontrar nenhum


def _caminho_template_geral_final() -> Path | None:
    """Template único Kartado (relatório consolidado): preferência «Final», depois «geral»."""
    pasta = Path(__file__).resolve().parent.parent / "assets" / "templates"
    for nome in ("Template - Geral - 4 e 5 - Final.xlsx", "Template - geral.xlsx"):
        candidato = pasta / nome
        if candidato.is_file():
            return candidato
    return None


def _norm_stem_comparar(s: str) -> str:
    """Igual ao critério de fotos_campo.core para nomes de .xlsx (Kartado / macros)."""
    t = unicodedata.normalize("NFC", s or "")
    for u in ("\u2013", "\u2014", "\u2212"):
        t = t.replace(u, "-")
    return t.casefold()


def _repo_root() -> Path:
    return Path(__file__).resolve().parent.parent.parent


def _kartado_repo_template_roots() -> tuple[Path, ...]:
    """Raiz do repo: Kartado/Planilhas Padrão - Templates (e variante sem acento), se existirem."""
    kd = _repo_root() / "Kartado"
    if not kd.is_dir():
        return ()
    out: list[Path] = []
    for name in ("Planilhas Padrão - Templates", "Planilhas Padrao - Templates"):
        p = kd / name
        if p.is_dir():
            out.append(p)
    return tuple(out)


def _iter_xlsx_kartado_repo_extra() -> list[Path]:
    found: list[Path] = []
    for root in _kartado_repo_template_roots():
        try:
            for f in root.rglob("*.xlsx"):
                if _deve_excluir_xlsx_template_m01(f):
                    continue
                found.append(f)
        except OSError:
            continue
    return found


# Modelos de pipeline (Kria/Kcor/acumulado) — não usar como base do M01 ao listar/fazer match fuzzy.
_M01_EXCLUIR_NOME_XLSX_NORM: frozenset[str] = frozenset(
    _norm_stem_comparar(n)
    for n in (
        "Modelo Abertura Evento Kria Conserva Rotina.xlsx",
        "Modelo.xlsx",
        "_Planilha Modelo Kcor-Kria.xlsx",
        "Acumulado.xlsx",
        "Eventos Acumulado Artesp para Exportar Kria.xlsx",
    )
)


def _deve_excluir_xlsx_template_m01(f: Path) -> bool:
    if not f.is_file() or f.name.startswith("~$"):
        return True
    if "Template_EAF" in f.name:
        return True
    if "Planilha Modelo Conservação" in f.name and "Foto 2 Lados" in f.name:
        return True
    return _norm_stem_comparar(f.name) in _M01_EXCLUIR_NOME_XLSX_NORM


@lru_cache(maxsize=1024)
def _xlsx_parece_layout_kartado(path_str: str) -> bool:
    """
    Valida se o XLSX parece template/planilha Kartado (cabeçalho na linha 1).
    Evita selecionar modelos de outros fluxos (ex.: Kria/Resposta) no M01 Kartado.
    """
    p = Path(path_str)
    try:
        wb = load_workbook(str(p), read_only=True, data_only=True)
        try:
            ws = wb.active
            max_c = min(int(ws.max_column or 0), 120)
            hdr: set[str] = set()
            for c in range(1, max_c + 1):
                v = ws.cell(row=1, column=c).value
                if v is None:
                    continue
                k = _norm_header(str(v))
                if k:
                    hdr.add(k)
            # Mínimo para considerar layout Kartado válido.
            return (
                ("rodovia" in hdr)
                and ("classe" in hdr)
                and (("codigo de fiscalizacao" in hdr) or ("codigo fiscalizacao" in hdr))
            )
        finally:
            wb.close()
    except Exception:
        return False


def _iter_nc_assets_xlsx_kartado_candidatos() -> list[Path]:
    """.xlsx Kartado apenas em nc_artesp/assets/templates/ (recursivo)."""
    d = Path(__file__).resolve().parent.parent / "assets" / "templates"
    if not d.is_dir():
        return []
    try:
        return list(d.rglob("*.xlsx"))
    except OSError:
        return []


def _resolver_ficheiro_xlsx_por_nome_em_repo(nome: str) -> Path | None:
    """Localiza `nome` em nc_artesp/assets/templates, fotos_campo/assets/templates ou repo Kartado/."""
    if not (nome or "").strip():
        return None
    alvo = _norm_stem_comparar(nome)
    try:
        for f in _iter_nc_assets_xlsx_kartado_candidatos():
            if _deve_excluir_xlsx_template_m01(f):
                continue
            if _norm_stem_comparar(f.name) != alvo:
                continue
            if _xlsx_parece_layout_kartado(str(f.resolve())):
                return f
    except OSError:
        pass
    try:
        from fotos_campo.core import (
            _ficheiro_xlsx_bundled_por_nome,
            _ficheiro_xlsx_por_nome_em_assets,
        )
    except ImportError:
        return None
    p = _ficheiro_xlsx_por_nome_em_assets(nome)
    if p is not None and p.is_file() and _xlsx_parece_layout_kartado(str(p.resolve())):
        return p
    p = _ficheiro_xlsx_bundled_por_nome(nome)
    if p is not None and p.is_file() and _xlsx_parece_layout_kartado(str(p.resolve())):
        return p
    for root in _kartado_repo_template_roots():
        try:
            for f in root.rglob("*.xlsx"):
                if _deve_excluir_xlsx_template_m01(f):
                    continue
                if _norm_stem_comparar(f.name) == alvo:
                    if _xlsx_parece_layout_kartado(str(f.resolve())):
                        return f
        except OSError:
            continue
    return None


@lru_cache(maxsize=32)
def _listar_candidatos_templates_kartado_cache() -> tuple[Path, ...]:
    """
    .xlsx em nc_artesp/assets/templates e fotos_campo/assets/templates (recursivo),
    e em repo/Kartado/Planilhas Padrão - Templates — exceto EAF, Kria/Kcor/acumulado e Foto 2 Lados.
    """
    repo = _repo_root()
    dirs = [
        Path(__file__).resolve().parent.parent / "assets" / "templates",
        repo / "fotos_campo" / "assets" / "templates",
    ]
    out: list[Path] = []
    for d in dirs:
        if not d.is_dir():
            continue
        try:
            for f in d.rglob("*.xlsx"):
                if _deve_excluir_xlsx_template_m01(f):
                    continue
                if not _xlsx_parece_layout_kartado(str(f.resolve())):
                    continue
                out.append(f)
        except OSError:
            continue
    out.extend(_iter_xlsx_kartado_repo_extra())
    # únicos por caminho resolvido
    seen: set[str] = set()
    uniq: list[Path] = []
    for p in out:
        try:
            k = str(p.resolve())
        except OSError:
            k = str(p)
        if k not in seen:
            seen.add(k)
            uniq.append(p)
    return tuple(uniq)


def _listar_candidatos_templates_kartado() -> list[Path]:
    """Wrapper compatível que devolve list a partir do cache."""
    return list(_listar_candidatos_templates_kartado_cache())


def _tokens_atividade(s: str) -> set[str]:
    return {t for t in re.split(r"[^\w]+", (s or "").lower()) if len(t) >= 3}


def _norm_key_template_lookup(s: str) -> str:
    """
    Normaliza texto para lookup de mapa de templates:
    - remove acentos
    - remove pontuação (parênteses, ponto final, etc.)
    - colapsa espaços
    """
    t = unicodedata.normalize("NFD", str(s or ""))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    t = t.lower()
    t = re.sub(r"[^0-9a-zA-Z]+", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


_M01_MAPA_ATIVIDADE_TEMPLATE_KARTADO_NORM: dict[str, str] = {
    _norm_key_template_lookup(k): v for k, v in M01_MAPA_ATIVIDADE_TEMPLATE_KARTADO.items()
}

_ART03_CLASSE_POR_ATIVIDADE_NORM: dict[str, str] = {
    _norm_key_template_lookup(k): v for k, v in ART03_ATIVIDADE_PARA_SERVICO_KARTADO.items()
}

_CLASSES_KARTADO_PERMITIDAS: frozenset[str] = frozenset(
    {*ART03_ATIVIDADE_PARA_SERVICO_KARTADO.values(), *KARTADO_RELATORIO_SERVICOS_TODOS}
)

COL_EAF_GRUPO_MACRO_M = 13
COL_EAF_ATIVIDADE_N = 14


def _tokens_para_match_classe(s: str) -> set[str]:
    return {t for t in re.split(r"[^\w]+", _norm_key_template_lookup(s)) if len(t) >= 2}


def _remover_prefixo_codigo_apontamento(s: str) -> str:
    """
    Remove prefixo de catálogo no formato "<sigla>-<subitem>-", ex. "XX - a.1.3 - ...",
    mantendo só o apontamento descritivo para mapear classe ARTESP.
    """
    txt = _str_linha_eaf(s)
    if not txt:
        return txt
    out = re.sub(r"^\s*[A-Za-z]{1,4}\s*-\s*[a-z]\.\d+(?:\.\d+)?\s*-\s*", "", txt, flags=re.I)
    return out.strip() or txt


def _classe_por_codigo_apontamento(s: str) -> str | None:
    txt = _str_linha_eaf(s)
    if not txt:
        return None
    m = re.match(r"^\s*[A-Za-z]{1,4}\s*-\s*([a-z]\.\d+(?:\.\d+)?)\s*-\s*(.+)$", txt, flags=re.I)
    if not m:
        return None
    codigo = m.group(1).lower()
    desc = _norm_key_template_lookup(m.group(2))
    if codigo in {"b.7.1", "b.9.1"} and "limpeza" in desc:
        return "FD - Parada onibus - Limpeza/Pintura"
    return None


def _classe_kartado_permitida_no_contexto(classe: str) -> bool:
    s = _str_linha_eaf(classe)
    if not s:
        return False
    return s in _CLASSES_KARTADO_PERMITIDAS


def _clasificar_texto_art03_ou_servico_nc_bruto(txt: str) -> str | None:
    """Devolve texto de serviço Kartado apenas se pertencer à lista canónica (ou mapeável por ART03)."""
    s = _str_linha_eaf(txt)
    if not s:
        return None
    s_sem_codigo = _remover_prefixo_codigo_apontamento(s)
    classe_codigo = _classe_por_codigo_apontamento(s)
    if classe_codigo and _classe_kartado_permitida_no_contexto(classe_codigo):
        return classe_codigo
    candidatos_txt = [s]
    if s_sem_codigo != s:
        candidatos_txt.append(s_sem_codigo)
    v = next(
        (ART03_ATIVIDADE_PARA_SERVICO_KARTADO.get(c) for c in candidatos_txt if ART03_ATIVIDADE_PARA_SERVICO_KARTADO.get(c)),
        None,
    )
    if v and _classe_kartado_permitida_no_contexto(v):
        return v
    vn = next(
        (
            _ART03_CLASSE_POR_ATIVIDADE_NORM.get(_norm_key_template_lookup(c))
            for c in candidatos_txt
            if _ART03_CLASSE_POR_ATIVIDADE_NORM.get(_norm_key_template_lookup(c))
        ),
        None,
    )
    if vn and _classe_kartado_permitida_no_contexto(vn):
        return vn
    tup = next((SERVICO_NC.get(c) for c in candidatos_txt if SERVICO_NC.get(c)), None)
    if tup and len(tup) > 1:
        svc = str(tup[1])
        if _classe_kartado_permitida_no_contexto(svc):
            return svc
    if _classe_kartado_permitida_no_contexto(s):
        return s

    # Regra ARTESP: apontamento da mãe pode vir sem prefixo Kartado; tenta "FD - <apontamento>".
    candidatos_apontamento = [s]
    if not _norm_key_template_lookup(s).startswith("fd "):
        candidatos_apontamento.append(f"FD - {s}")

    for cand in candidatos_apontamento:
        if _classe_kartado_permitida_no_contexto(cand):
            return cand

    norm_cands = {_norm_key_template_lookup(c) for c in candidatos_apontamento}
    for perm in _CLASSES_KARTADO_PERMITIDAS:
        if _norm_key_template_lookup(perm) in norm_cands:
            if _classe_kartado_permitida_no_contexto(perm):
                return perm

    nk = _norm_key_template_lookup(s)
    for perm in _CLASSES_KARTADO_PERMITIDAS:
        if _norm_key_template_lookup(perm) == nk:
            if _classe_kartado_permitida_no_contexto(perm):
                return perm
    return None


def _melhor_classe_kartado_fuzzy(texto: str) -> tuple[str, float] | None:
    if not _str_linha_eaf(texto):
        return None
    nt = _norm_key_template_lookup(texto)
    tt = _tokens_para_match_classe(texto)
    best: str | None = None
    best_sc = 0.0
    for cl in _CLASSES_KARTADO_PERMITIDAS:
        nc = _norm_key_template_lookup(cl)
        if nt and nt == nc:
            return (cl, 1.0)
        if nt and len(nt) >= 6 and (nt in nc or nc in nt):
            sc = 0.9
        else:
            ct = _tokens_para_match_classe(cl)
            if not tt or not ct:
                continue
            inter = len(tt & ct)
            uni = len(tt | ct)
            sc = inter / uni if uni else 0.0
        if sc > best_sc:
            best_sc, best = sc, cl
    if best is not None and best_sc >= 0.22:
        return (best, best_sc)
    return None


def _partir_celula_grupo_atividade_mae_combinada(s: str) -> tuple[str, str]:
    """
    Célula da mãe com rótulos «Grupo de atividade» … «Atividade» na mesma linha de texto
    (ex.: ``Grupo de atividade; Prédios e Pátios - Atividade - Cobertura / Forro``).
    Devolve ``(trecho_grupo, trecho_atividade)`` ou ``("", "")`` se o formato não for reconhecido.
    """
    t = _str_linha_eaf(s)
    if not t:
        return "", ""
    m = re.search(
        r"(?is)grupo\s+de\s+atividade\s*[;:]\s*(.+?)\s*[-–]\s*atividade\s*[-–]\s*(.+)\s*$",
        t,
    )
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return "", ""


def _resolver_classe_kartado_inteligente_mno(
    *,
    texto_m: str,
    texto_n: str,
    texto_o: str,
    tipo_txt_detetado: str,
) -> str:
    """
    Cruza macrogrupo (M), atividade (N), tipo de atividade (O) e o texto da coluna «Atividade»/ART03
    com ``ART03_ATIVIDADE_PARA_SERVICO_KARTADO`` e a lista canónica Kartado.
    A coluna «Classe» do consolidado só recebe valores presentes nessa lista.
    """
    m, n, o = _str_linha_eaf(texto_m), _str_linha_eaf(texto_n), _str_linha_eaf(texto_o)
    tq = _str_linha_eaf(tipo_txt_detetado)
    individuais = _uniq_textos_classe([n, tq, o, m])
    compostos: list[str] = []
    if m and n:
        compostos.extend((f"{m} - {n}", f"{m}. {n}", f"{m} / {n}"))
    if n and o:
        compostos.extend((f"{n} - {o}", f"{n} ({o})"))
    if m and n and o:
        compostos.extend((f"{m} - {n} - {o}", f"{n} - {o} — {m}"))

    for cand in individuais + _uniq_textos_classe(compostos):
        hit = _clasificar_texto_art03_ou_servico_nc_bruto(cand)
        if hit:
            return hit

    fuzzy_pool = individuais + _uniq_textos_classe(compostos)
    if m and n and o:
        fuzzy_pool.append(f"{m} {n} {o}")
    best_pair: tuple[str, float] | None = None
    for cand in fuzzy_pool:
        pair = _melhor_classe_kartado_fuzzy(cand)
        if pair and (best_pair is None or pair[1] > best_pair[1]):
            best_pair = pair
    if best_pair is not None:
        return best_pair[0]

    for fb in ("FD - Prédio e Pátio", "FD - Conformação Lateral", "FD - Lixo/Entulho"):
        if _classe_kartado_permitida_no_contexto(fb):
            logger.warning(
                "Classe Kartado sem match forte (M=%r N=%r O=%r Q-det=%r) — usando fallback %r",
                m[:80], n[:80], o[:80], tq[:80], fb,
            )
            return fb

    pick = next(
        (
            c
            for c in sorted(_CLASSES_KARTADO_PERMITIDAS)
            if _classe_kartado_permitida_no_contexto(c)
        ),
        sorted(_CLASSES_KARTADO_PERMITIDAS)[0],
    )
    logger.warning(
        "Classe Kartado: fallback extremo (%r) para M/N/O/Q=%r/%r/%r/%r",
        pick, m[:40], n[:40], o[:40], tq[:40],
    )
    return pick


def _uniq_textos_classe(seq: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for x in seq:
        s = _str_linha_eaf(x)
        if not s:
            continue
        k = s.casefold()
        if k in seen:
            continue
        seen.add(k)
        out.append(s)
    return out


def _detectar_col_macrogrupo_mae(ws, fallback: int = COL_EAF_GRUPO_MACRO_M) -> int:
    """Coluna do **macrogrupo** (ex.: segurança rodoviária), não a coluna «Grupo de atividade» descritiva."""
    melhor_c: int | None = None
    melhor_score = -1
    for r in range(1, min(ws.max_row + 1, 8) + 1):
        for c in range(1, min(ws.max_column + 1, 22) + 1):
            h = _norm_header(str(ws.cell(row=r, column=c).value or ""))
            if not h:
                continue
            if "tipo" in h and "atividade" in h:
                continue
            if "grupo" in h and "atividade" in h:
                continue
            sc = 0
            if h in {"grupo macro", "macrogrupo", "macro grupo", "macrogrupo rodoviario"}:
                sc = 9
            elif "macro" in h and "grupo" in h:
                sc = 8
            elif h == "grupo" and c <= 14:
                sc = 4
            if sc > melhor_score:
                melhor_score, melhor_c = sc, c
    return melhor_c if melhor_c is not None and melhor_score >= 7 else fallback


def _detectar_col_grupo_atividade_descricao_mae(ws) -> int | None:
    """Coluna cujo cabeçalho é «Grupo de atividade» (texto descritivo), distinta do macrogrupo e da «Atividade»."""
    melhor_c: int | None = None
    melhor_score = -1
    for r in range(1, min(ws.max_row + 1, 8) + 1):
        for c in range(1, min(ws.max_column + 1, 40) + 1):
            h = _norm_header(str(ws.cell(row=r, column=c).value or ""))
            if not h or ("tipo" in h and "atividade" in h):
                continue
            if "grupo" in h and "atividade" in h:
                sc = 10 if "grupo de atividade" in h else 9
            else:
                continue
            if sc > melhor_score:
                melhor_score, melhor_c = sc, c
    return melhor_c if melhor_c is not None and melhor_score >= 9 else None


def _detectar_col_grupo_macro_mae(ws, fallback: int = COL_EAF_GRUPO_MACRO_M) -> int:
    """Compatível: macrogrupo; se não houver cabeçalho explícito, coincide com coluna M padrão."""
    return _detectar_col_macrogrupo_mae(ws, fallback=fallback)


def _detectar_col_atividade_mae_col_n(ws, fallback: int = COL_EAF_ATIVIDADE_N) -> int:
    melhor_c: int | None = None
    melhor_score = -1
    for r in range(1, min(ws.max_row + 1, 8) + 1):
        for c in range(1, min(ws.max_column + 1, 40) + 1):
            h = _norm_header(str(ws.cell(row=r, column=c).value or ""))
            if not h:
                continue
            if h == "tipo de atividade" or h == "tipo atividade":
                continue
            if "grupo" in h and "atividade" in h:
                continue
            sc = 0
            if h == "atividade":
                sc = 8
            elif h.endswith("atividade") and "tipo" not in h:
                sc = 6
            elif "atividade" in h and "tipo" not in h:
                sc = 4
            if sc > melhor_score:
                melhor_score, melhor_c = sc, c
    if melhor_c is not None and melhor_score >= 6:
        return melhor_c
    return fallback


def _resolver_template_kartado_para_atividade(tipo_atividade: str, fallback: Path | None = None) -> Path | None:
    """
    Usado quando M01_COPIA_PLANILHA_MAE=False: escolhe o .xlsx Kartado antes de colar linhas do grupo.

    Ordem:
      1. Mapa exato M01_MAPA_ATIVIDADE_TEMPLATE_KARTADO (texto da coluna de atividade na mãe, alinhado ao Art_03).
      2. Dicas por palavra-chave → ficheiro em assets/templates.
      3. Maior sobreposição de tokens entre atividade e stem do ficheiro.
      4. fallback (opcional).
    """
    tipo = (tipo_atividade or "").strip()
    if not tipo:
        return fallback

    nome_mapa = M01_MAPA_ATIVIDADE_TEMPLATE_KARTADO.get(tipo)
    if not nome_mapa:
        # Lookup normalizado (tolerante a pontuação/acentos/espacos finais).
        nome_mapa = _M01_MAPA_ATIVIDADE_TEMPLATE_KARTADO_NORM.get(_norm_key_template_lookup(tipo))
    if nome_mapa:
        p = _resolver_ficheiro_xlsx_por_nome_em_repo(nome_mapa)
        if p is not None:
            return p

    candidatos = _listar_candidatos_templates_kartado()
    if not candidatos:
        return fallback

    tipo_l = tipo.lower()
    for palavra, frag in M01_DICAS_PALAVRA_TEMPLATE_KARTADO:
        if palavra not in tipo_l:
            continue
        frag_n = _norm_stem_comparar(frag)
        for f in candidatos:
            if frag_n in _norm_stem_comparar(f.stem):
                return f

    tt = _tokens_atividade(tipo)
    best: Path | None = None
    best_score = 0
    for f in candidatos:
        st = _tokens_atividade(f.stem)
        sc = len(tt & st)
        if sc > best_score:
            best_score = sc
            best = f
    if best is not None and best_score >= 1:
        return best

    # Fallback operacional: se não houver match por nome/tokens, usar qualquer template
    # Kartado válido disponível para não interromper o M01.
    if candidatos:
        return candidatos[0]

    return fallback


# ESTRUTURA — planilha-mãe EAF e Template_EAF.xlsx (fallback do M01):
#   Linhas 1 a 4 = cabeçalho; dados a partir da linha 5 (ver M01_LINHA_INICIO).
LINHA_CABECALHO_FIM = 4   # última linha do cabeçalho EAF (1–4)
PRIMEIRA_LINHA_DADOS = 5  # primeira linha de dados na mãe / Template_EAF
COL_SCAN_MAX_CABECALHO_DATAS = 120  # mães Kartado / export largos: «Prazo» / «DtFim_Prog» podem passar da col Z

# Templates Kartado (nc_artesp/assets/templates/): só a linha 1 é cabeçalho — não apagar; dados a partir da 2.
PRIMEIRA_LINHA_DADOS_TEMPLATE_KARTADO = 2


def _limpar_linhas_dados_eaf_no_sheet(ws, primeira_linha: int) -> None:
    """Remove linhas de dados a partir de ``primeira_linha`` (mantém cabeçalho EAF 1..primeira_linha-1)."""
    while ws.max_row >= primeira_linha:
        ws.delete_rows(ws.max_row, 1)


def _copiar_linha_mae_para_template_eaf(ws_mae, linha_mae: int, ws_tpl, linha_tpl: int, ultima_coluna: int) -> None:
    """Cópia coluna a coluna (valor + estilo) de uma linha da mãe para o template — sem remapear colunas."""
    if ultima_coluna < 1:
        return
    for col in range(1, ultima_coluna + 1):
        src = ws_mae.cell(row=linha_mae, column=col)
        dst = ws_tpl.cell(row=linha_tpl, column=col)
        dst.value = src.value
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)
    if linha_mae in ws_mae.row_dimensions and ws_mae.row_dimensions[linha_mae].height is not None:
        ws_tpl.row_dimensions[linha_tpl].height = ws_mae.row_dimensions[linha_mae].height

# CONSTANTES DE COLUNAS – planilha-mãe EAF (índice 1 = col A)
# Colunas fixas (iguais em todas as versões do EAF):
#   C=código, D=data constatação, F=rodovia, I=m_ini, K=m_fim, Atividade em N ou Q conforme modelo, V=nº foto
# Coluna variável detectada dinamicamente no cabeçalho:
#   "Data Reparo" → T(20) no template manual | S(19) nos exports do sistema ARTESP
# Demais colunas (para preenchimento completo pelo módulo MA):
#   E=horário fiscalização, G=concessionária/EAF, H=km inicial (formato 143+800), J=km final, L=sentido, O=tipo atividade, P=grupo, U=responsável
COL_KM_I_M   = 9   # I – metros inicial
COL_KM_F_M   = 11  # K – metros final
COL_CODIGO   = 3   # C – código fiscalização / número da NC
COL_SEQ_FOTO = 22  # V – número da NC para foto (código da col C, ou sequencial)
COL_DATA_NC  = 20  # T – data reparo/prazo (fallback; detectado dinamicamente em executar())
COL_RODOVIA  = 6   # F – rodovia
COL_TIPO_NC  = 17  # fallback Q — «Atividade»; detetada via cabeçalho (em vários EAF é col N)
COL_DATA_CON = 4   # D – data da constatação
COL_HORA_FISC = 5  # E – horário da fiscalização (mãe EAF padrão)
COL_CONCESSIONARIA = 7   # G – concessionária / EAF
COL_KM_I_FULL     = 8   # H – km inicial (formato 143+800)
COL_KM_F_FULL     = 10  # J – km final (formato 143+800)
COL_SENTIDO       = 12  # L – sentido
COL_TIPO_ATIV     = 15  # O – tipo de atividade (macro; não é o texto ART03 da «Classe»)
COL_GRUPO_ATIV    = 16  # P – grupo (fiscalização)
COL_RESPONSAVEL   = 21  # U – responsável (fiscal)


@dataclass
class ColunasDetectadas:
    """Estrutura para colunas de datas detectadas no cabeçalho."""
    data_reparo: int
    data_envio: int | None = None
    data_reparo_detectada: bool = False


def _detectar_colunas_datas(ws, fallback_reparo: int = 20, fallback_envio: int | None = 19) -> ColunasDetectadas:
    """
    Detecta colunas de datas no cabeçalho (linhas 1-5) em uma única passada.
    - data_reparo: «Data Reparo» / «Data do Reparo», sinónimos Kartado («Prazo», «DtFim_Prog»), etc.
    - data_envio: cabeçalho normalizado contendo 'data' e 'envio'
    """
    col_envio = None
    best_rep_score = 0
    best_rep_col: int | None = None
    lim_c = min(ws.max_column or 0, COL_SCAN_MAX_CABECALHO_DATAS)
    for r in range(1, 6):
        for c in range(1, lim_c + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            h = _norm_header(str(v))
            sr = _score_header_data_reparo(h)
            if sr > 0 and (
                best_rep_col is None
                or sr > best_rep_score
                or (sr == best_rep_score and c < best_rep_col)
            ):
                best_rep_score = sr
                best_rep_col = c
            if col_envio is None and "data" in h and "envio" in h:
                col_envio = c
        if best_rep_col is not None and col_envio is not None:
            break
    if best_rep_col is not None:
        return ColunasDetectadas(
            data_reparo=best_rep_col,
            data_envio=col_envio if col_envio is not None else fallback_envio,
            data_reparo_detectada=True,
        )
    col_reparo = fallback_reparo
    if col_envio is None:
        col_envio = fallback_envio
    return ColunasDetectadas(
        data_reparo=col_reparo,
        data_envio=col_envio,
        data_reparo_detectada=False,
    )


def _detectar_col_data_reparo(ws, fallback: int = 20) -> int:
    """
    Detecta a coluna 'Data Reparo' lendo o cabeçalho da planilha (linhas 1-5).
    Compatível com:
      - Template manual (_Planilha Modelo nc lote 13.xls): col T(20) = 'Data Reparo'
      - Exports do sistema ARTESP: col S(19) = 'Data Reparo'
      - Kartado / consolidados: «Prazo», «DtFim_Prog», «Data do Reparo» (até col ~120)
    Retorna o índice 1-based encontrado, ou `fallback` se não localizar.
    """
    cols = _detectar_colunas_datas(ws, fallback_reparo=fallback, fallback_envio=None)
    if not cols.data_reparo_detectada:
        logger.warning(f"Coluna 'Data Reparo' nao encontrada no cabecalho — usando fallback col {fallback}")
    else:
        logger.debug(f"Coluna 'Data Reparo' detectada: col {cols.data_reparo}")
    return cols.data_reparo


def _detectar_col_data_envio(ws, fallback: int = 19) -> int:
    """
    Detecta a coluna 'Data do envio' ou 'Data envio' no cabeçalho (linhas 1-5).
    Usado no template EAF para mapear data da fiscalização (constatação) para a coluna correta.
    Retorna o índice 1-based; se não encontrar, retorna fallback.
    """
    cols = _detectar_colunas_datas(ws, fallback_reparo=COL_DATA_NC, fallback_envio=fallback)
    return cols.data_envio if cols.data_envio is not None else fallback


def _detectar_col_tipo_nc(ws, fallback: int = COL_TIPO_NC) -> int:
    """
    Detecta a coluna de "Atividade" no cabeçalho.
    Objetivo: retornar a coluna que contém o valor que vira template Kartado (ex.: "Defesa metálica (manutenção ou substituição)"),
    e NÃO "Tipo de Atividade" (ex.: "Segurança Rodoviária") nem colunas de data.
    """
    # Maior score = maior prioridade.
    melhor_c = None
    melhor_score = -1

    # Normaliza para evitar "Atividade", "Atividade " etc.
    def score_header(h: str) -> int:
        # Evitar pegar "Tipo de Atividade" como se fosse "Atividade".
        if "tipo" in h and "atividade" in h and "tipo de atividade" in h:
            return 1
        if "grupo" in h and "atividade" in h:
            return 1

        # Preferir explicitamente a coluna "Atividade".
        if h == "atividade":
            return 6
        if h.endswith("atividade") and "tipo" not in h and "grupo" not in h:
            return 5

        # Legado: "Evento".
        if h == "evento":
            return 3
        if "evento" in h:
            return 2

        # Outros casos possíveis (Kria/NC templates).
        if "tipo nc" in h:
            return 2
        if "servico" in h or "serviço" in h:
            return 2
        if "atividade" in h:
            # Se sobrou algo com "atividade" mas não bateu, fica abaixo.
            return 2
        return 0

    # Cabeçalho costuma estar nas primeiras linhas (1..6).
    for r in range(1, min(ws.max_row + 1, 8) + 1):
        for c in range(1, min(ws.max_column + 1, 60) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            h = _norm_header(str(v))
            if not h:
                continue
            sc = score_header(h)
            if sc > melhor_score:
                melhor_score = sc
                melhor_c = c

    # Se não achou coluna confiável, volta ao fallback.
    return melhor_c if melhor_c is not None and melhor_score >= 3 else fallback


def _detectar_linha_inicio_dados(ws, col_codigo: int, max_busca: int = 15) -> int:
    """
    Detecta a primeira linha com um valor numérico plausível como código de fiscalização.
    Garante que nunca desce abaixo de M01_LINHA_INICIO para não quebrar planilhas padrão.
    """
    limite = min(int(ws.max_row or M01_LINHA_INICIO), max_busca)
    for r in range(1, limite + 1):
        v = ws.cell(row=r, column=col_codigo).value
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        # Rejeitar cabeçalhos (texto) e horários (HH:MM:SS)
        if re.fullmatch(r"\d{1,2}:\d{2}(:\d{2})?", s):
            continue
        if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{2,4}", s):
            continue
        # Aceitar como primeiro dado se for numérico puro (código fiscalização) ou alfanumérico com dígitos
        if re.search(r"\d{3,}", s):
            return r
    return M01_LINHA_INICIO


def _detectar_col_rodovia(ws, fallback: int = COL_RODOVIA) -> int:
    """Deteta a coluna de rodovia no cabeçalho (linhas 1-8)."""
    for r in range(1, min(ws.max_row + 1, 8) + 1):
        for c in range(1, min(ws.max_column + 1, 30) + 1):
            h = _norm_header(str(ws.cell(row=r, column=c).value or ""))
            if h == "rodovia":
                return c
    return fallback


def _detectar_col_km_inicial(ws, fallback: int = COL_KM_I_FULL) -> int:
    """Deteta a coluna de km inicial (formato 143+800) no cabeçalho (linhas 1-8)."""
    melhor_c = None
    melhor_score = -1
    for r in range(1, min(ws.max_row + 1, 8) + 1):
        for c in range(1, min(ws.max_column + 1, 40) + 1):
            h = _norm_header(str(ws.cell(row=r, column=c).value or ""))
            if not h:
                continue
            score = 0
            if h in {"km inicial", "km inicio", "km ini"}:
                score = 6
            elif "km" in h and ("inicial" in h or "inicio" in h):
                score = 5
            elif h == "km":
                score = 4
            if "projeto" in h:
                score = 0
            if score > melhor_score:
                melhor_score = score
                melhor_c = c
    return melhor_c if melhor_c is not None and melhor_score >= 4 else fallback


def _detectar_col_km_final(ws, fallback: int = COL_KM_F_FULL) -> int:
    """Deteta a coluna de km final (formato 143+800) no cabeçalho (linhas 1-8)."""
    melhor_c = None
    melhor_score = -1
    for r in range(1, min(ws.max_row + 1, 8) + 1):
        for c in range(1, min(ws.max_column + 1, 40) + 1):
            h = _norm_header(str(ws.cell(row=r, column=c).value or ""))
            if not h:
                continue
            score = 0
            if h in {"km final", "km fim"}:
                score = 6
            elif "km" in h and ("final" in h or "fim" in h):
                score = 5
            if "projeto" in h:
                score = 0
            if score > melhor_score:
                melhor_score = score
                melhor_c = c
    return melhor_c if melhor_c is not None and melhor_score >= 5 else fallback


def _detectar_col_sentido(ws, fallback: int = COL_SENTIDO) -> int:
    """Deteta a coluna de sentido no cabeçalho (linhas 1-8)."""
    for r in range(1, min(ws.max_row + 1, 8) + 1):
        for c in range(1, min(ws.max_column + 1, 40) + 1):
            h = _norm_header(str(ws.cell(row=r, column=c).value or ""))
            if h == "sentido":
                return c
    return fallback


def _detectar_col_km_i_metros(ws, fallback: int | None = COL_KM_I_M) -> int | None:
    """Deteta a coluna de metros do km inicial no cabeçalho (linhas 1-8)."""
    for r in range(1, min(ws.max_row + 1, 8) + 1):
        for c in range(1, min(ws.max_column + 1, 50) + 1):
            h = _norm_header(str(ws.cell(row=r, column=c).value or ""))
            if h in {"m inicial", "metros inicial", "m ini", "m inicio"}:
                return c
            if ("metro" in h or h.startswith("m ")) and ("inicial" in h or "inicio" in h):
                return c
    return fallback


def _detectar_col_km_f_metros(ws, fallback: int | None = COL_KM_F_M) -> int | None:
    """Deteta a coluna de metros do km final no cabeçalho (linhas 1-8)."""
    for r in range(1, min(ws.max_row + 1, 8) + 1):
        for c in range(1, min(ws.max_column + 1, 50) + 1):
            h = _norm_header(str(ws.cell(row=r, column=c).value or ""))
            if h in {"m final", "metros final", "m fim"}:
                return c
            if ("metro" in h or h.startswith("m ")) and ("final" in h or "fim" in h):
                return c
    return fallback


def _refinar_mapa_tpl_metros_km(
    ws_m,
    hdr_mae: dict[str, int],
    hdr_tpl: dict[str, int],
    mapa_tpl_para_mae: dict[int, int],
) -> None:
    aliases_ini = (
        "km inicial | m",
        "km inicio | m",
        "km inicial | metro",
        "km inicial | metros",
        "metros inicial",
        "m inicial",
        "m ini",
        "m inicio",
    )
    aliases_fim = (
        "km final | m",
        "km fim | m",
        "km final | metro",
        "km final | metros",
        "metros final",
        "m final",
        "m fim",
    )

    def pick_col(keys: tuple[str, ...]) -> int | None:
        for k in keys:
            c = hdr_mae.get(k)
            if c:
                return int(c)
        return None

    for tpl_k, col_tpl in hdr_tpl.items():
        if col_tpl in mapa_tpl_para_mae:
            continue
        lk = tpl_k.casefold()
        need_ini = tpl_k in aliases_ini or (
            lk.endswith("| m") and ("inicial" in lk or "inicio" in lk)
        )
        need_fim = tpl_k in aliases_fim or (
            lk.endswith("| m") and ("final" in lk or "fim" in lk)
        )
        if need_ini:
            cm = pick_col(aliases_ini)
            if cm is None:
                cm = _detectar_col_km_i_metros(ws_m, fallback=None)
            if cm is None:
                ck = hdr_mae.get("km inicial") or hdr_mae.get("km inicio")
                if ck:
                    cm = int(ck) + 1
            if cm:
                mapa_tpl_para_mae[col_tpl] = cm
        elif need_fim:
            cm = pick_col(aliases_fim)
            if cm is None:
                cm = _detectar_col_km_f_metros(ws_m, fallback=None)
            if cm is None:
                ck = hdr_mae.get("km final") or hdr_mae.get("km fim")
                if ck:
                    cm = int(ck) + 1
            if cm:
                mapa_tpl_para_mae[col_tpl] = cm


def _detectar_col_seq_foto(ws, fallback: int = COL_SEQ_FOTO) -> int:
    """Deteta a coluna de número/sequência de foto no cabeçalho (linhas 1-8)."""
    melhor_c = None
    melhor_score = -1
    for r in range(1, min(ws.max_row + 1, 8) + 1):
        for c in range(1, min(ws.max_column + 1, 80) + 1):
            h = _norm_header(str(ws.cell(row=r, column=c).value or ""))
            if not h:
                continue
            if "responsavel" in h:
                continue
            if "fiscal" in h and "foto" not in h:
                continue
            score = 0
            if h in {
                "n foto",
                "numero foto",
                "n da foto",
                "sequencia foto",
                "seq foto",
                "seq. foto",
                "foto sequencia",
            }:
                score = 6
            elif ("foto" in h and ("numero" in h or "sequencia" in h or h.startswith("n "))):
                score = 5
            if score > melhor_score:
                melhor_score = score
                melhor_c = c
    return melhor_c if melhor_c is not None and melhor_score >= 5 else fallback


def _detectar_col_responsavel(ws, fallback: int = COL_RESPONSAVEL) -> int:
    """Deteta a coluna de responsável/fiscal no cabeçalho (linhas 1-8)."""
    for r in range(1, min(ws.max_row + 1, 8) + 1):
        for c in range(1, min(ws.max_column + 1, 80) + 1):
            h = _norm_header(str(ws.cell(row=r, column=c).value or ""))
            if h in {"responsavel", "fiscal", "responsavel fiscal"}:
                return c
            if "responsavel" in h or ("fiscal" in h and "codigo" not in h):
                return c
    return fallback


def _parece_km(s: str) -> bool:
    t = _limpar_str(s).replace(" ", "")
    return bool(re.fullmatch(r"\d{1,4}\+\d{1,3}", t))


def _normalizar_km_final_e_sentido(km_f, sentido):
    """
    Corrige linhas com layout inconsistente:
    - se km final vier como sentido textual (Norte/Sul/Marginal...) e sentido estiver vazio,
      move o valor para sentido e deixa km final vazio.
    """
    kmf_s = _limpar_str(km_f)
    sen_s = _limpar_str(sentido)
    if sen_s:
        return km_f, sentido
    if not kmf_s:
        return km_f, sentido
    kmf_norm = _norm_header(kmf_s)
    sentidos_txt = {
        "norte",
        "sul",
        "leste",
        "oeste",
        "marginal norte",
        "marginal sul",
        "norte/sul",
        "sul/norte",
    }
    if kmf_norm in sentidos_txt:
        return "", kmf_s
    return km_f, sentido


def _detectar_col_data_con(ws, fallback: int = COL_DATA_CON) -> int:
    """Deteta a coluna de data da constatação/fiscalização no cabeçalho (linhas 1-8)."""
    alvos = {"data fiscalizacao", "data da fiscalizacao", "data constatacao", "data da constatacao", "data envio", "data do envio"}
    melhor_c = None
    melhor_score = -1
    for r in range(1, min(ws.max_row + 1, 8) + 1):
        for c in range(1, min(ws.max_column + 1, 30) + 1):
            h = _norm_header(str(ws.cell(row=r, column=c).value or ""))
            if not h:
                continue
            score = 0
            if h in alvos:
                score = 5
            elif "data" in h and ("fiscal" in h or "constat" in h):
                score = 4
            elif "data" in h and "envio" in h:
                score = 3
            if score > melhor_score:
                melhor_score = score
                melhor_c = c
    return melhor_c if melhor_c is not None and melhor_score >= 3 else fallback


def _detectar_col_hora_fiscalizacao(ws, fallback: int = COL_HORA_FISC) -> int:
    """Deteta a coluna do horário da fiscalização no cabeçalho (linhas 1–8). Padrão col E."""
    alvos = {
        "horario",
        "horario da fiscalizacao",
        "horario fiscalizacao",
        "hora",
        "hora da fiscalizacao",
        "hora fiscalizacao",
    }
    melhor_c = None
    melhor_score = -1
    for r in range(1, min(ws.max_row + 1, 8) + 1):
        for c in range(1, min(ws.max_column + 1, 40) + 1):
            h = _norm_header(str(ws.cell(row=r, column=c).value or ""))
            if not h:
                continue
            score = 0
            if h in alvos:
                score = 5
            elif "horario" in h and "data" not in h:
                score = 4
            elif h == "hora" or (h.startswith("hora ") and "data" not in h):
                score = 4
            if score > melhor_score:
                melhor_score = score
                melhor_c = c
    return melhor_c if melhor_c is not None and melhor_score >= 4 else fallback


def _parse_hora_celula_mae(val_hora) -> tuple[int, int, int]:
    """Parte hora da célula mãe → (hh, mm, ss); vazio → meia-noite."""
    if val_hora is None:
        return (0, 0, 0)
    if isinstance(val_hora, datetime):
        return (val_hora.hour, val_hora.minute, val_hora.second)
    if isinstance(val_hora, time):
        return (val_hora.hour, val_hora.minute, val_hora.second)
    s = str(val_hora).strip()
    if not s or s.lower() in ("none", "nan"):
        return (0, 0, 0)
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            t = datetime.strptime(s, fmt).time()
            return (t.hour, t.minute, t.second)
        except ValueError:
            pass
    if isinstance(val_hora, (int, float)):
        v = float(val_hora)
        if 0 <= v < 1.0:
            total_seconds = int(round(v * 24 * 3600))
            if total_seconds >= 24 * 3600:
                total_seconds = min(total_seconds, 24 * 3600 - 1)
            hh = total_seconds // 3600
            mm = (total_seconds % 3600) // 60
            ss = total_seconds % 60
            return (hh, mm, ss)
        try:
            from openpyxl.utils.datetime import from_excel

            dtx = from_excel(v)
            if isinstance(dtx, datetime):
                return (dtx.hour, dtx.minute, dtx.second)
        except Exception:
            pass
    return (0, 0, 0)


def _combinar_data_e_hora_fiscalizacao(val_data, val_hora) -> datetime | None:
    """
    Data (constatação ou prazo) + horário col E da mãe → datetime no Kartado.
    Sem hora na mãe → 00:00 na data indicada (não inventa outra hora).
    """
    if isinstance(val_data, datetime):
        base = val_data.replace(microsecond=0)
    elif isinstance(val_data, date):
        base = datetime(val_data.year, val_data.month, val_data.day)
    else:
        base = parse_data(val_data)
    if not base:
        return None
    hh, mm, ss = _parse_hora_celula_mae(val_hora)
    return base.replace(hour=hh, minute=mm, second=ss, microsecond=0)


def _detectar_col_codigo_fiscalizacao(ws, fallback: int = COL_CODIGO) -> int:
    """Deteta a coluna de código de fiscalização no cabeçalho (linhas 1-8)."""
    candidatos = {
        "codigo de fiscalizacao",
        "codigo fiscalizacao",
        "codigo da fiscalizacao",
        "cod. fiscalizacao",
        "cod fiscalizacao",
        "codigo fisc",
        "numero da nc",
        "numero nc",
        "n da nc",
    }
    melhor_c = None
    melhor_score = -1
    for r in range(1, min(ws.max_row + 1, 8) + 1):
        for c in range(1, min(ws.max_column + 1, 80) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            h = _norm_header(str(v))
            if not h:
                continue
            score = 0
            if h in candidatos:
                score = 5
            elif "fiscalizacao" in h and ("codigo" in h or h.startswith("cod")):
                score = 4
            elif "nc" in h and ("numero" in h or h.startswith("n ")):
                score = 3
            elif h == "codigo":
                score = 2
            if score > melhor_score:
                melhor_score = score
                melhor_c = c
    return melhor_c if melhor_c is not None and melhor_score >= 3 else fallback


def _detectar_col_concessionaria(ws, fallback: int = COL_CONCESSIONARIA) -> int:
    """Deteta a coluna de concessionária (linhas 1-8)."""
    melhor_c = None
    melhor_score = -1
    for r in range(1, min(ws.max_row + 1, 8) + 1):
        for c in range(1, min(ws.max_column + 1, 80) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            h = _norm_header(str(v))
            if not h:
                continue
            score = 0
            if "concessionaria" in h:
                score = 5
            elif "lote" in h and ("concessionaria" in h or "concess" in h):
                score = 4
            elif "concess" in h:
                score = 3
            if score > melhor_score:
                melhor_score = score
                melhor_c = c
    return melhor_c if melhor_c is not None and melhor_score >= 3 else fallback


def _ler_cabecalhos_planilha_hierarquico(
    ws,
    linha_principal: int,
    linha_sub: int | None = None,
    max_col: int | None = None,
) -> dict[str, int]:
    """
    Lê cabeçalho em uma ou duas linhas (principal + sub) e devolve {nome_normalizado: coluna}.

    - Quando há sub-cabeçalho na coluna (linha_sub), a chave fica "<principal> | <sub>".
    - Quando não há sub, a chave é só o cabeçalho principal.
    - Cabeçalhos idênticos repetidos verticalmente contam como UMA coluna.
    """
    max_c = max_col or int(ws.max_column or 0)
    out: dict[str, int] = {}
    if linha_sub is None or linha_sub == linha_principal:
        for c in range(1, max_c + 1):
            v = ws.cell(row=linha_principal, column=c).value
            if v is None:
                continue
            k = _norm_header(str(v))
            if k and k not in out:
                out[k] = c
        return out

    # Propaga cabeçalho principal pela direita (cobre células mescladas)
    principal_propagado: list[str] = [""] * (max_c + 1)
    ultimo = ""
    for c in range(1, max_c + 1):
        v = ws.cell(row=linha_principal, column=c).value
        if v is not None and str(v).strip():
            ultimo = _norm_header(str(v))
        principal_propagado[c] = ultimo

    for c in range(1, max_c + 1):
        v_p = ws.cell(row=linha_principal, column=c).value
        v_s = ws.cell(row=linha_sub, column=c).value
        nome_principal = _norm_header(str(v_p)) if v_p is not None else ""
        nome_sub = _norm_header(str(v_s)) if v_s is not None else ""

        if nome_principal and nome_sub and nome_principal == nome_sub:
            chave = nome_principal
        elif not nome_principal and nome_sub:
            esq = principal_propagado[c]
            for c_left in range(c - 1, 0, -1):
                v_left_sub = ws.cell(row=linha_sub, column=c_left).value
                if v_left_sub is None or not str(v_left_sub).strip():
                    continue
                h_left = _norm_header(str(v_left_sub))
                if not h_left:
                    continue
                if any(x in h_left for x in ("final", "fim", "inicial", "inicio")):
                    esq = h_left
                elif ("km" in h_left or "metro" in h_left) and len(h_left) > 3:
                    esq = h_left
                break
            chave = f"{esq} | {nome_sub}" if esq else nome_sub
        elif nome_principal and not nome_sub:
            chave = nome_principal
        elif nome_principal and nome_sub:
            chave = f"{nome_principal} | {nome_sub}"
        else:
            continue

        if chave and chave not in out:
            out[chave] = c
    return out


def _detectar_cabecalho_template(ws) -> tuple[int, int | None]:
    """
    Retorna (linha_principal, linha_sub_ou_None) do template.
    """
    max_r = min(int(ws.max_row or 0), 8)
    max_c = int(ws.max_column or 0)
    linha_principal = None
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            h = _norm_header(str(ws.cell(row=r, column=c).value or ""))
            if h in ("rodovia", "codigo de fiscalizacao", "cod fiscalizacao", "cod. fiscalizacao", "codigo fiscalizacao", "codigo"):
                linha_principal = r
                break
        if linha_principal:
            break
    if linha_principal is None:
        return (4, None)

    linha_sub = linha_principal + 1
    if linha_sub > max_r:
        return (linha_principal, None)
    tem_sub = False
    for c in range(1, max_c + 1):
        h = _norm_header(str(ws.cell(row=linha_sub, column=c).value or ""))
        if h:
            tem_sub = True
            break
    return (linha_principal, linha_sub if tem_sub else None)


def _detectar_linha_inicio_dados_mae(ws, col_codigo_mae: int) -> int:
    """
    Primeira linha com código plausível na coluna de código da mãe.
    """
    limite = min(int(ws.max_row or 0), 20)
    for r in range(1, limite + 1):
        v = ws.cell(row=r, column=col_codigo_mae).value
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        if re.fullmatch(r"\d{1,2}:\d{2}(:\d{2})?", s):
            continue
        if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{2,4}", s):
            continue
        if re.search(r"\d{3,}", s):
            return r
    return M01_LINHA_INICIO


def _valor_tipo_nc(ws, row: int, col_tipo_nc: int):
    """Lê tipo NC evitando usar acidentalmente coluna de data."""
    v = _cell(ws, row, col_tipo_nc)
    if v and parse_data(v) is None:
        return v
    # Fallback comum no layout emergencial: coluna P (Evento)
    evento = _cell(ws, row, 16)
    if evento and parse_data(evento) is None:
        return evento
    return v


def _detectar_colunas_data_no_template(ws_template) -> tuple[int | None, int | None]:
    """
    Detecta no template EAF (cabeçalho linhas 1-5) as colunas 'Data do envio' e 'Data do reparo'.
    Retorna (col_data_envio, col_data_reparo). Se alguma não for encontrada, retorna None para essa.
    Fallback: se só 'Data Reparo' for encontrada em T(20), assume S(19) = Data do envio.
    """
    col_envio = None
    best_rep_score = 0
    best_rep_col: int | None = None
    lim_c = min(ws_template.max_column or 0, COL_SCAN_MAX_CABECALHO_DATAS)
    for r in range(1, 6):
        for c in range(1, lim_c + 1):
            v = ws_template.cell(row=r, column=c).value
            if v is None:
                continue
            h = _norm_header(str(v))
            sr = _score_header_data_reparo(h)
            if sr > 0 and (
                best_rep_col is None
                or sr > best_rep_score
                or (sr == best_rep_score and c < best_rep_col)
            ):
                best_rep_score = sr
                best_rep_col = c
            if col_envio is None and "data" in h and "envio" in h:
                col_envio = c
        if best_rep_col is not None and col_envio is not None:
            break
    col_reparo = best_rep_col
    if col_reparo is not None and col_envio is None and col_reparo == 20:
        col_envio = 19
    return (col_envio, col_reparo)


def _converter_xls_para_xlsx(path_xls: Path) -> Path:
    """
    Lê um arquivo .xls (formato antigo) com xlrd e grava um .xlsx equivalente
    com openpyxl no mesmo diretório. Cabeçalho (linhas 1 a M01_LINHA_INICIO-1)
    e demais linhas são copiadas só com valores; não inventa cabeçalho genérico.
    Retorna o Path do arquivo .xlsx gerado.
    """
    path_xlsx = path_xls.with_suffix(".xlsx")
    if path_xlsx == path_xls:
        path_xlsx = path_xls.parent / (path_xls.stem + "_convertido.xlsx")

    book = xlrd.open_workbook(str(path_xls))
    sheet = book.sheet_by_index(0)

    wb = openpyxl.Workbook()
    ws = wb.active
    if sheet.name:
        ws.title = sheet.name[:31]  # limite de 31 caracteres no Excel

    # Copiar só os valores do .xls (cabeçalho e dados); não gravar nada genérico.
    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            cell = sheet.cell(row, col)
            if cell.ctype == xlrd.XL_CELL_EMPTY:
                continue
            excel_cell = ws.cell(row=row + 1, column=col + 1)
            if cell.ctype == xlrd.XL_CELL_NUMBER:
                excel_cell.value = cell.value
            elif cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    dt = xlrd.xldate.xldate_as_datetime(cell.value, book.datemode)
                    excel_cell.value = dt
                except (ValueError, OverflowError):
                    excel_cell.value = cell.value
            elif cell.ctype == xlrd.XL_CELL_TEXT:
                excel_cell.value = cell.value
            elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                excel_cell.value = bool(cell.value)
            else:
                excel_cell.value = cell.value

    wb.save(str_caminho_io_windows(path_xlsx))
    logger.info(f"Arquivo .xls convertido para: {path_xlsx.name}")
    return path_xlsx


def _cell(ws, row: int, col: int):
    """Retorna valor da célula (row, col) ou string vazia."""
    v = ws.cell(row=row, column=col).value
    return v if v is not None else ""


def _padronizar_colunas_km(ws, row: int) -> None:
    """Padroniza colunas I e K (metros) em uma linha."""
    for col in (COL_KM_I_M, COL_KM_F_M):
        cell = ws.cell(row=row, column=col)
        cell.number_format = "@"
        cell.value = pad_metros(cell.value)


def _limpar_str(val) -> str:
    """Normaliza valor para comparação (string limpa)."""
    return str(val).strip() if val is not None else ""


def _strip_descricao_kartado_excel(s: str) -> str:
    """
    Descrição Kartado numa única linha (sem \\n na célula): remove control chars,
    quebras → espaço, colapsa espaços (evita «buracos» e spill no Excel).
    """
    if not s:
        return ""
    t = s.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    t = t.replace("\u2028", " ").replace("\u2029", " ")
    out: list[str] = []
    for ch in t:
        o = ord(ch)
        if ch == "\t":
            out.append(" ")
        elif o < 32:
            continue
        elif o == 0x7F:
            continue
        else:
            out.append(ch)
    return re.sub(r"\s+", " ", "".join(out)).strip()


# Excel: tipo data/hora (ordem dia/mês/ano, BR).
_FMT_DATA_EXCEL_KARTADO = "dd/mm/yyyy hh:mm:ss"


def _kartado_data_sem_hora_celula(val) -> str | None:
    """Texto DD/MM/AAAA para descrições e contextos onde o Excel não deve serializar data."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, date):
        return val.strftime("%d/%m/%Y")
    dt = parse_data(val)
    if dt:
        return dt.strftime("%d/%m/%Y")
    s = str(val).strip()
    return s if s else None


def _aplicar_celula_data_excel(ws, row: int, col: int, val, *, preservar_hora: bool = False) -> None:
    """``datetime`` + formato ``dd/mm/yyyy hh:mm``. Sem ``preservar_hora``, normaliza à meia-noite (só data)."""
    if col is None or col < 1:
        return
    cell = ws.cell(row=row, column=col)
    dt = None
    if val is None or (isinstance(val, str) and not str(val).strip()):
        cell.value = None
        return
    if isinstance(val, datetime):
        dt = val.replace(microsecond=0)
    else:
        dt = parse_data(val)
        if dt:
            dt = dt.replace(microsecond=0)
    if dt is not None and not preservar_hora:
        dt = dt.replace(hour=0, minute=0, second=0, microsecond=0)
    cell.value = dt
    if dt is not None:
        cell.number_format = _FMT_DATA_EXCEL_KARTADO


def _foto_ref_numerica(val) -> str:
    """
    Referência de foto válida para nome PDF (macro usa nº/foto da linha).
    Aceita apenas numérico para evitar usar campos textuais (ex.: responsável técnico).
    """
    s = _limpar_str(val)
    if not s:
        return ""
    try:
        f = float(s.replace(",", "."))
        if f == int(f):
            return str(int(f))
    except (ValueError, TypeError):
        return ""
    return ""


def _norm_header(s: str) -> str:
    t = unicodedata.normalize("NFD", str(s or ""))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", t).strip().lower()


def _score_header_data_reparo(h: str) -> int:
    if not h:
        return 0
    if h == "data reparo":
        return 100
    if h == "data do reparo":
        return 99
    if "data" in h and "reparo" in h and "envio" not in h:
        return 85
    if h in ("dtfim_prog", "dt fim prog"):
        return 80
    if h == "prazo":
        return 70
    return 0


_KARTADO_CANON_OBSERVACOES_K = _norm_header("Observações")
_KARTADO_HDR_OBSERVACOES_AA_SINONIMOS: tuple[str, ...] = (
    _KARTADO_CANON_OBSERVACOES_K,
    _norm_header("Observação"),
)


def _kartado_primeira_coluna_observacoes_por_hdr(hdr: dict[str, int]) -> int | None:
    """Coluna **AA** «Observações» (texto PDF / localização); não confundir com ObsGestor / observação do gestor."""
    for k in _KARTADO_HDR_OBSERVACOES_AA_SINONIMOS:
        if "gestor" in k:
            continue
        c = hdr.get(k)
        if c is not None and int(c) > 0:
            return int(c)
    return None


def _kartado_hdr_tem_coluna_observacoes(hdr: dict[str, int]) -> bool:
    return _kartado_primeira_coluna_observacoes_por_hdr(hdr) is not None


def _kartado_cols_tpl_unificar_cabecalho_observacoes(cols_tpl: dict[str, int]) -> None:
    """Uma só chave canónica «Observações» (AA); não altera ObsGestor nem outras colunas."""
    c = _kartado_primeira_coluna_observacoes_por_hdr(cols_tpl)
    if c is None:
        return
    for syn in _KARTADO_HDR_OBSERVACOES_AA_SINONIMOS:
        if "gestor" in syn:
            continue
        cols_tpl.pop(syn, None)
    cols_tpl[_KARTADO_CANON_OBSERVACOES_K] = int(c)


_KARTADO_HDR_DATA_EXACT = frozenset(
    _norm_header(x)
    for x in (
        "Encontrado em",
        "Prazo",
        "Data Solicitação",
        "Data Suspensão",
        "DtInicio_Prog",
        "DtFim_Prog",
        "DtInicio_Exec",
        "DtFim_Exec",
    )
)


def _kartado_hdr_e_coluna_data(header_norm: str) -> bool:
    if not header_norm:
        return False
    if header_norm in _KARTADO_HDR_DATA_EXACT:
        return True
    if header_norm.startswith("data "):
        return True
    if len(header_norm) >= 4 and header_norm.startswith("dt") and header_norm[2] != " ":
        return True
    return False


def _kartado_preservar_hora_celula_data(val) -> bool:
    if isinstance(val, datetime):
        return bool(val.hour or val.minute or val.second)
    if isinstance(val, date) and not isinstance(val, datetime):
        return False
    dt = parse_data(val)
    if dt:
        return bool(dt.hour or dt.minute or dt.second)
    return False


def _colunas_kartado_por_header(ws) -> dict[str, int]:
    """Mapa de cabeçalho (linha 1) -> índice de coluna no template Kartado."""
    out: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        k = _norm_header(str(v))
        if k and k not in out:
            out[k] = c
    return out


def _kartado_garantir_coluna_texto_mae_resposta(ws, cols_tpl: dict[str, int], header_label: str) -> int:
    k = _norm_header(header_label)
    if k in cols_tpl:
        return cols_tpl[k]
    base = max(int(ws.max_column or 0), max(cols_tpl.values(), default=0))
    c = base + 1
    ws.cell(row=1, column=c).value = header_label
    cols_tpl[k] = c
    return c


def _set_if_header(ws, row: int, cols: dict[str, int], header: str, valor) -> None:
    c = cols.get(_norm_header(header))
    if c:
        ws.cell(row=row, column=c).value = valor


def _set_data_if_header(
    ws, row: int, cols: dict[str, int], header: str, val, *, preservar_hora: bool = False
) -> None:
    c = cols.get(_norm_header(header))
    if c:
        _aplicar_celula_data_excel(ws, row, c, val, preservar_hora=preservar_hora)


def _rodovia_chave(s: str) -> str:
    t = _norm_header(s or "")
    t = t.replace("-", "").replace(" ", "").replace("/", "")
    return t


def _sentido_chave(s: str) -> str:
    return _norm_header(s or "").replace("-", " ").strip()


def _rodovia_fmt_eaf_para_kartado(rodovia_raw: str) -> str:
    """Texto único da coluna Rodovia no Exportar/Kartado (ex.: SPI, SPI 102/300 → «SPI-102/300»)."""
    raw = _limpar_str(rodovia_raw)
    if not raw:
        return ""
    info = normalizar_rodovia_eaf(raw, RODOVIAS)
    if info.get("tag") and info["tag"] != "FORA":
        return RODOVIA_NOME_SEPARAR.get(info["tag"], info.get("codigo") or raw).strip()
    return raw


def consolidar_kartados_em_unico_excel(
    arqs: list[Path],
    pasta_destino: Path,
    nome_saida: str | None = None,
) -> Path | None:
    """
    Junta todas as linhas de dados de vários Excel Kartado (layout 1 cabeçalho + N linhas)
    num único ficheiro baseado no template único.

    Ordena por Rodovia → Classe → Código Fiscalização.
    Devolve o Path do ficheiro gerado, ou None se não houver linhas.
    """
    from openpyxl import load_workbook as _lw

    tpl = _caminho_template_geral_final()
    if tpl is None or not tpl.is_file():
        raise FileNotFoundError(
            "Template consolidado Kartado não encontrado: "
            "'Template - Geral - 4 e 5 - Final.xlsx' ou 'Template - geral.xlsx' em assets/templates."
        )

    # 1. Ler cabeçalho do template
    wb_tpl = _lw(str_caminho_io_windows(tpl), data_only=True)
    ws_tpl = wb_tpl.active
    hdr_tpl: dict[str, int] = {}
    for c in range(1, int(ws_tpl.max_column or 0) + 1):
        v = ws_tpl.cell(row=1, column=c).value
        if v is not None:
            k = _norm_header(str(v))
            if k and k not in hdr_tpl:
                hdr_tpl[k] = c
    max_col_tpl = int(ws_tpl.max_column or 0)
    max_idx_tpl = max(max(hdr_tpl.values()) if hdr_tpl else 0, max_col_tpl)
    wb_tpl.close()

    _hq_mae = _norm_header(RESPOSTA_PENDENTES_HEADER_MAE_ATIVIDADE_Q)
    _ho_mae = _norm_header(RESPOSTA_PENDENTES_HEADER_MAE_TIPO_ATIV_O)
    _obs_kartado_k = _norm_header("Observações")
    _labels_mae_extra = {
        _hq_mae: RESPOSTA_PENDENTES_HEADER_MAE_ATIVIDADE_Q,
        _ho_mae: RESPOSTA_PENDENTES_HEADER_MAE_TIPO_ATIV_O,
    }
    for arq in arqs:
        if not arq.is_file():
            continue
        try:
            wb0 = _lw(str_caminho_io_windows(arq), data_only=True)
            ws0 = wb0.active
            hdr0: dict[str, int] = {}
            for c in range(1, int(ws0.max_column or 0) + 1):
                v0 = ws0.cell(row=1, column=c).value
                if v0 is not None:
                    k0 = _norm_header(str(v0))
                    if k0 and k0 not in hdr0:
                        hdr0[k0] = c
            wb0.close()
            for kx in (_hq_mae, _ho_mae):
                if kx in hdr0 and kx not in hdr_tpl:
                    max_idx_tpl += 1
                    hdr_tpl[kx] = max_idx_tpl
            if _kartado_hdr_tem_coluna_observacoes(hdr0) and _obs_kartado_k not in hdr_tpl:
                max_idx_tpl += 1
                hdr_tpl[_obs_kartado_k] = max_idx_tpl
        except Exception:
            pass

    if _obs_kartado_k not in hdr_tpl:
        max_idx_tpl += 1
        hdr_tpl[_obs_kartado_k] = max_idx_tpl

    cols_tpl_data = {
        int(c_tpl)
        for h, c_tpl in hdr_tpl.items()
        if _kartado_hdr_e_coluna_data(h)
    }

    # 2. Agregar todas as linhas de todos os Excels
    todas_linhas: list[tuple[list, tuple]] = []  # (valores, chave_ordem)
    for arq in arqs:
        if not arq.is_file():
            continue
        try:
            wb = _lw(str_caminho_io_windows(arq), data_only=True)
            ws = wb.active
            hdr_arq: dict[str, int] = {}
            for c in range(1, int(ws.max_column or 0) + 1):
                v = ws.cell(row=1, column=c).value
                if v is not None:
                    k = _norm_header(str(v))
                    if k and k not in hdr_arq:
                        hdr_arq[k] = c
            col_cod_arq = hdr_arq.get("codigo fiscalizacao") or hdr_arq.get("codigo de fiscalizacao") or hdr_arq.get("cod. fiscalizacao")
            col_rod_arq = hdr_arq.get("rodovia")
            col_cls_arq = hdr_arq.get("classe")
            for r in range(2, int(ws.max_row or 0) + 1):
                # Verificar se linha tem dado
                tem = any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, int(ws.max_column or 0) + 1))
                if not tem:
                    continue
                # Mapear colunas do ficheiro origem → colunas do template (por nome de cabeçalho)
                valores = [None] * (max_idx_tpl + 1)
                for h_nome, c_tpl in hdr_tpl.items():
                    if h_nome in ("latitude", "longitude"):
                        continue
                    if h_nome == _obs_kartado_k:
                        c_arq = _kartado_primeira_coluna_observacoes_por_hdr(hdr_arq)
                    else:
                        c_arq = hdr_arq.get(h_nome)
                    if c_arq and 0 < int(c_tpl) <= max_idx_tpl:
                        valores[int(c_tpl)] = ws.cell(row=r, column=int(c_arq)).value
                cod = _limpar_str(ws.cell(row=r, column=col_cod_arq).value) if col_cod_arq else ""
                rod = _limpar_str(ws.cell(row=r, column=col_rod_arq).value) if col_rod_arq else ""
                cls = _limpar_str(ws.cell(row=r, column=col_cls_arq).value) if col_cls_arq else ""
                chave = (rod.casefold(), cls.casefold(), cod)
                todas_linhas.append((valores, chave))
            wb.close()
        except Exception as e:
            logger.warning("consolidar_kartados: erro a ler %s — %s", arq.name, e)

    if not todas_linhas:
        return None

    todas_linhas.sort(key=lambda x: x[1])

    # 3. Criar ficheiro final baseado no template
    garantir_pasta(pasta_destino)
    if not nome_saida:
        from datetime import datetime as _dt
        nome_saida = _sanitizar_nome_xlsx(
            f"{_dt.now().strftime('%Y%m%d')} - CONSTATAÇÕES NC {M01_LOTE} - Kartado Consolidado.xlsx"
        )
    destino = encurtar_nome_em_pasta(pasta_destino, nome_saida)
    shutil.copy2(str_caminho_io_windows(tpl), str_caminho_io_windows(destino))

    wb_out = _lw(str_caminho_io_windows(destino))
    ws_out = wb_out.active
    # Apagar linhas de dados do template (mantém linha 1 = cabeçalho)
    while ws_out.max_row >= 2:
        ws_out.delete_rows(ws_out.max_row, 1)

    for nk, lab in _labels_mae_extra.items():
        if nk not in hdr_tpl:
            continue
        cw = hdr_tpl[nk]
        if ws_out.cell(row=1, column=cw).value != lab:
            ws_out.cell(row=1, column=cw).value = lab
    if _obs_kartado_k in hdr_tpl:
        cw_o = hdr_tpl[_obs_kartado_k]
        if ws_out.cell(row=1, column=cw_o).value != "Observações":
            ws_out.cell(row=1, column=cw_o).value = "Observações"

    for seq, (valores, _) in enumerate(todas_linhas, start=2):
        for c_tpl in range(1, len(valores)):
            val = valores[c_tpl]
            if val is None:
                continue
            if c_tpl in cols_tpl_data:
                _aplicar_celula_data_excel(
                    ws_out,
                    seq,
                    c_tpl,
                    val,
                    preservar_hora=_kartado_preservar_hora_celula_data(val),
                )
            else:
                ws_out.cell(row=seq, column=c_tpl).value = val

    destino_xls = destino.with_suffix(".xls")
    if destino_xls.exists():
        destino_xls.unlink()
    wb_out.save(str_caminho_io_windows(destino))
    wb_out.close()
    logger.info("Kartado consolidado: %d linhas → %s", len(todas_linhas), destino.name)
    return destino


def _fingerprint_linha_mae(
    ws, row: int, col_max: int, *, forcar_linha_unica: bool
) -> tuple[str, ...]:
    """
    Tuplo de todas as células da linha (1..col_max) normalizadas.
    Linhas só partilham o mesmo Excel se o tuplo for idêntico (incl. data reparo, km, código, etc.).
    Com forcar_linha_unica=True (um_arquivo_por_nc), acrescenta o nº da linha para nunca agrupar.
    """
    cells = tuple(_limpar_str(ws.cell(row=row, column=c).value) for c in range(1, col_max + 1))
    return cells + (str(row),) if forcar_linha_unica else cells


def _copiar_linha_com_estilo(ws_src, row_src: int, ws_dst, row_dst: int, max_col: int) -> None:
    """
    Copia uma linha inteira (valor + estilo) de ws_src para ws_dst.
    Mesma regra do gerador de modelo/foto: preserva font, border, fill, alignment, number_format.
    """
    for col in range(1, max_col + 1):
        src_cell = ws_src.cell(row=row_src, column=col)
        dst_cell = ws_dst.cell(row=row_dst, column=col)
        dst_cell.value = src_cell.value
        if src_cell.has_style:
            dst_cell.font = src_cell.font.copy()
            dst_cell.border = src_cell.border.copy()
            dst_cell.fill = src_cell.fill.copy()
            dst_cell.number_format = src_cell.number_format
            dst_cell.alignment = src_cell.alignment.copy()


def _copiar_valores_linha_com_offset_colunas(
    ws_src,
    row_src: int,
    ws_dst,
    row_dst: int,
    max_col_dst: int,
    *,
    offset_col_src: int = 0,
) -> None:
    """
    Copia apenas valores, alinhando colunas do destino com offset na origem.
    Ex.: offset_col_src=2 => dst C recebe src E (compensa 2 colunas vazias em A/B na mãe).
    """
    for c_dst in range(1, max_col_dst + 1):
        c_src = c_dst + offset_col_src
        if c_src <= 0:
            ws_dst.cell(row=row_dst, column=c_dst).value = None
            continue
        ws_dst.cell(row=row_dst, column=c_dst).value = ws_src.cell(row=row_src, column=c_src).value


def _copiar_alturas_linhas(ws_src, ws_dst, src_start: int, num_linhas: int, dst_start: int) -> None:
    """Copia as alturas das linhas do bloco origem para o bloco destino (como em gerar_modelo_foto)."""
    for offset in range(num_linhas):
        dim = ws_src.row_dimensions.get(src_start + offset)
        if dim is not None and dim.height is not None:
            ws_dst.row_dimensions[dst_start + offset].height = dim.height


def _replicar_merges_linhas_com_offset(
    ws_src,
    ws_dst,
    row_max_src: int,
    row_offset: int,
    *,
    row_min_src: int = 1,
) -> None:
    """Replica mesclagens da mãe (linhas ``row_min_src``..``row_max_src``) no destino com deslocamento de linha."""
    for mc in list(ws_src.merged_cells.ranges):
        if mc.max_row < row_min_src or mc.min_row > row_max_src:
            continue
        if mc.min_row < row_min_src:
            continue
        r1 = mc.min_row + row_offset
        r2 = mc.max_row + row_offset
        try:
            ws_dst.merge_cells(
                start_row=r1,
                start_column=mc.min_col,
                end_row=r2,
                end_column=mc.max_col,
            )
        except Exception:
            pass


def _snapshot_template_linhas_1_a_n(ws, linha_fim: int) -> tuple[dict, list[tuple[int, int, int, int]], dict[int, float], int]:
    """
    Captura valores/estilos das linhas 1..linha_fim.

    Mesclagens: todas as áreas que **intercetam** 1..linha_fim (ex.: Exportar.xlsx usa C3:C4 para rótulos);
    ao restauro, cada mescla é cortada ao intervalo das linhas 1–linha_fim para não atravessar linha dados.
    """
    max_c = max(int(ws.max_column or 1), 1)
    for mc in ws.merged_cells.ranges:
        if mc.min_row <= linha_fim:
            max_c = max(max_c, mc.max_col)
    cells: dict[tuple[int, int], tuple] = {}
    for r in range(1, linha_fim + 1):
        for c in range(1, max_c + 1):
            cell = ws.cell(row=r, column=c)
            cells[(r, c)] = (
                cell.value,
                copy(cell.font) if cell.font else None,
                copy(cell.border) if cell.border else None,
                copy(cell.fill) if cell.fill else None,
                cell.number_format,
                copy(cell.alignment) if cell.alignment else None,
                bool(cell.has_style),
            )
    merges: list[tuple[int, int, int, int]] = []
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row > linha_fim or mc.max_row < 1:
            continue
        mr = max(mc.min_row, 1)
        xr = min(mc.max_row, linha_fim)
        mc_r = mc.min_col
        xc_c = mc.max_col
        if mr <= xr and not (mr == xr and mc_r == xc_c):
            merges.append((mr, mc_r, xr, xc_c))
    heights: dict[int, float] = {}
    for r in range(1, linha_fim + 1):
        dim = ws.row_dimensions.get(r)
        if dim is not None and dim.height is not None:
            heights[r] = dim.height
    return cells, merges, heights, max_c


def _template_limpar_linhas_sem_delete_rows(ws, linha_ini: int, ate_linha_inclusive: int, max_col: int) -> None:
    """Esconde apenas valores na faixa para evitar ''insert_rows'' efectivos ou estados estranhos de ``delete_rows``."""
    tpl_mc = max(int(ws.max_column or 0), max_col, 1)
    ult = max(ate_linha_inclusive, int(ws.max_row or linha_ini))
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= linha_ini and mc.max_row <= ult:
            try:
                ws.unmerge_cells(str(mc))
            except Exception:
                pass
    for r in range(linha_ini, ult + 1):
        for c in range(1, tpl_mc + 1):
            ws.cell(row=r, column=c).value = None


def _preencher_corpo_template_com_valores(
    ws_mae,
    ws_template,
    *,
    linha_ini_mae: int,
    linha_ini_template: int,
    offset_col_mae: int = 0,
    col_map: dict[int, int | None] | None = None,
) -> None:
    """
    Preenche só valores da mãe no corpo do template, preservando layout/estilo/mesclas.
    """
    max_r_mae = int(ws_mae.max_row or 0)
    # O mapeamento é dirigido pelo template: preencher apenas colunas existentes nele.
    max_c = max(int(ws_template.max_column or 0), 1)
    if max_r_mae < linha_ini_mae:
        return
    row_offset = linha_ini_template - linha_ini_mae
    # Modelo de alinhamento da 1.ª linha do corpo do template (ex.: linha 5 no Exportar.xlsx).
    alinh_modelo: dict[int, object] = {}
    borda_modelo: dict[int, object] = {}
    for c in range(1, max_c + 1):
        src_tpl = ws_template.cell(row=linha_ini_template, column=c)
        if isinstance(src_tpl, MergedCell):
            continue
        alinh_modelo[c] = copy(src_tpl.alignment) if src_tpl.alignment is not None else None
        borda_modelo[c] = copy(src_tpl.border) if src_tpl.border is not None else None
    ultima_linha_template = linha_ini_template + (max_r_mae - linha_ini_mae)
    for r_t in range(linha_ini_template, ultima_linha_template + 1):
        for c in range(1, max_c + 1):
            dst = ws_template.cell(row=r_t, column=c)
            if isinstance(dst, MergedCell):
                continue
            dst.value = None
            if c in alinh_modelo and alinh_modelo[c] is not None:
                dst.alignment = copy(alinh_modelo[c])
            if c in borda_modelo and borda_modelo[c] is not None:
                dst.border = copy(borda_modelo[c])
    for r_m in range(linha_ini_mae, max_r_mae + 1):
        r_t = r_m + row_offset
        for c in range(1, max_c + 1):
            dst = ws_template.cell(row=r_t, column=c)
            if isinstance(dst, MergedCell):
                continue
            # Regra do fluxo Exportar: coluna R deve permanecer vazia.
            if c == 18:
                dst.value = None
                continue
            if col_map is not None and c in col_map:
                c_src = int(col_map[c]) if col_map[c] is not None else 0
            else:
                c_src = c + offset_col_mae
            if c_src <= 0:
                dst.value = None
                continue
            dst.value = ws_mae.cell(row=r_m, column=c_src).value


def _tem_valor(v) -> bool:
    return v is not None and str(v).strip() != ""


def _pontuar_offset_colunas_mae(ws, linha_ini: int, offset: int) -> int:
    """
    Pontua quão coerente fica o mapeamento para o template Exportar:
    C=código, D=data, E=hora, F=rodovia. Maior score = melhor offset.
    """
    max_r = int(ws.max_row or 0)
    if max_r < linha_ini:
        return -10**6
    score = 0
    usados = 0
    for r in range(linha_ini, max_r + 1):
        c_cod = COL_CODIGO + offset
        c_data = COL_DATA_CON + offset
        c_hora = COL_HORA_FISC + offset
        c_rod = COL_RODOVIA + offset
        if min(c_cod, c_data, c_hora, c_rod) <= 0:
            continue
        v_cod = ws.cell(row=r, column=c_cod).value
        v_data = ws.cell(row=r, column=c_data).value
        v_hora = ws.cell(row=r, column=c_hora).value
        v_rod = ws.cell(row=r, column=c_rod).value
        if not any(_tem_valor(v) for v in (v_cod, v_data, v_hora, v_rod)):
            continue
        usados += 1
        s_cod = str(v_cod or "").strip()
        s_hora = str(v_hora or "").strip()
        s_rod = str(v_rod or "").strip().upper()
        # código: tipicamente numérico/string sem ':'
        if _tem_valor(v_cod) and ":" not in s_cod and any(ch.isdigit() for ch in s_cod):
            score += 3
        elif _tem_valor(v_cod):
            score -= 2
        # data constatação
        if parse_data(v_data) is not None:
            score += 3
        elif _tem_valor(v_data):
            score -= 2
        # hora
        if ":" in s_hora:
            score += 2
        elif _tem_valor(v_hora) and parse_data(v_hora) is not None:
            score += 1
        elif _tem_valor(v_hora):
            score -= 1
        # rodovia
        if s_rod.startswith(("SP", "SPI", "SPA", "R", "BR")):
            score += 3
        elif _tem_valor(v_rod):
            score -= 2
        if usados >= 8:
            break
    # sem linhas úteis -> forte penalização
    if usados == 0:
        return -10**6
    return score


def _dedup_merged_ranges(sorted_merges: list[tuple[int, int, int, int]]) -> list[tuple[int, int, int, int]]:
    seen: set[tuple[int, int, int, int]] = set()
    out: list[tuple[int, int, int, int]] = []
    for tup in sorted_merges:
        if tup in seen:
            continue
        seen.add(tup)
        out.append(tup)
    return out


def _restaurar_template_linhas_1_a_n(
    ws,
    linha_fim: int,
    cells: dict,
    merges: list[tuple[int, int, int, int]],
    heights: dict[int, float],
) -> None:
    """Reaplica snapshot das linhas 1..linha_fim por cima de qualquer alteração anterior."""
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row <= linha_fim:
            try:
                ws.unmerge_cells(str(mc))
            except Exception:
                pass
    for (r, c), pack in cells.items():
        val, font, border, fill, nf, alignment, has_style = pack
        dst = ws.cell(row=r, column=c)
        dst.value = val
        if has_style:
            if font is not None:
                dst.font = copy(font)
            if border is not None:
                dst.border = copy(border)
            if fill is not None:
                dst.fill = copy(fill)
            dst.number_format = nf
            if alignment is not None:
                dst.alignment = copy(alignment)
    merged_sorted = sorted(merges, key=lambda q: ((q[2] - q[0]) * (q[3] - q[1])))
    for min_r, min_c, max_r, max_c in _dedup_merged_ranges(merged_sorted):
        if max_r < min_r or max_c < min_c:
            continue
        if min_r == max_r and min_c == max_c:
            continue
        try:
            ws.merge_cells(
                start_row=min_r,
                start_column=min_c,
                end_row=max_r,
                end_column=max_c,
            )
        except Exception:
            pass
    for r, h in heights.items():
        ws.row_dimensions[r].height = h


def _reaplicar_cabecalho_template_1a3(ws_dest, ws_tpl, linha_fim: int = 4) -> None:
    """
    Reaplica exatamente as linhas 1..N do template no destino (valores + estilos + alturas),
    preservando-as contra qualquer efeito colateral do preenchimento da área de dados.
    """
    max_c = max(int(ws_tpl.max_column or 1), int(ws_dest.max_column or 1), 1)
    for mc in list(ws_dest.merged_cells.ranges):
        if mc.min_row <= linha_fim:
            try:
                ws_dest.unmerge_cells(str(mc))
            except Exception:
                pass
    for r in range(1, linha_fim + 1):
        _copiar_linha_com_estilo(ws_tpl, r, ws_dest, r, max_c)
        dim = ws_tpl.row_dimensions.get(r)
        if dim is not None and dim.height is not None:
            ws_dest.row_dimensions[r].height = dim.height
    for mc in list(ws_tpl.merged_cells.ranges):
        if mc.max_row <= linha_fim:
            try:
                ws_dest.merge_cells(
                    start_row=mc.min_row,
                    start_column=mc.min_col,
                    end_row=mc.max_row,
                    end_column=mc.max_col,
                )
            except Exception:
                pass


def _replicar_merged_cells_header(ws_src, ws_dst, row_ini: int, row_fim: int) -> None:
    """
    Replica no sheet destino as células mescladas que estão inteiras no range [row_ini, row_fim] do origem.
    Destino usa as mesmas coordenadas (cabeçalho 1:row_fim).
    """
    for mc in list(ws_src.merged_cells.ranges):
        if mc.min_row >= row_ini and mc.max_row <= row_fim:
            try:
                ws_dst.merge_cells(
                    start_row=mc.min_row,
                    start_column=mc.min_col,
                    end_row=mc.max_row,
                    end_column=mc.max_col,
                )
            except Exception:
                pass


def atualizar_col_v_indice_global(arqs: list[Path], start_index: int) -> int:
    """
    Atualiza a coluna V (número da foto) nos XLS individuais com índice global
    sequencial (start_index+1, start_index+2, ...), na ordem dos arquivos e das linhas.
    Usado após a extração de fotos do PDF para que M02 encontre PDF (1).jpg, PDF (2).jpg, etc.
    Retorna o número de linhas atualizadas.
    """
    from openpyxl import load_workbook
    idx = start_index
    total = 0
    for path in arqs:
        if not path.exists():
            continue
        wb = load_workbook(str(path))
        ws = wb.active
        ultima = ws.max_row
        # Encontrar última linha com dados na col C
        for r in range(ultima, M01_LINHA_INICIO - 1, -1):
            if ws.cell(row=r, column=COL_CODIGO).value:
                ultima = r
                break
        for r in range(M01_LINHA_INICIO, ultima + 1):
            if ws.cell(row=r, column=COL_CODIGO).value is None:
                continue
            idx += 1
            ws.cell(row=r, column=COL_SEQ_FOTO).value = idx
            total += 1
        wb.save(str_caminho_io_windows(path))
        wb.close()
    return total


def _desfazer_merges_que_incluem_coluna(ws, col_1based: int) -> None:
    """Evita falhas do openpyxl ao apagar/inserir colunas dentro de áreas mescladas."""
    for mrange in list(ws.merged_cells.ranges):
        try:
            if mrange.min_col <= col_1based <= mrange.max_col:
                ws.unmerge_cells(str(mrange))
        except (ValueError, KeyError, TypeError):
            continue


def aplicar_layout_exportacao_planilha_mae_pos_coleta_kartado(ws) -> None:
    """
    Só fluxo Kartado (cópia da mãe para entrega): remove col R; coluna vazia entre O e P;
    duas colunas vazias em A. Ordem de aplicação pedida para o layout final.
    """
    col_r = 18
    _desfazer_merges_que_incluem_coluna(ws, col_r)
    ws.delete_cols(col_r)
    _desfazer_merges_que_incluem_coluna(ws, 16)
    ws.insert_cols(16, 1)
    _desfazer_merges_que_incluem_coluna(ws, 1)
    ws.insert_cols(1, 2)


def _pick_hdr_col(hdr_mae: dict[str, int], *keys: str) -> int | None:
    for k in keys:
        c = hdr_mae.get(k)
        if c:
            return int(c)
    return None


def _resolver_cols_metadados_agrupamento_exportar(hdr_mae: dict[str, int]) -> dict[str, int | None]:
    return {
        "tipo_nc": _pick_hdr_col(
            hdr_mae,
            "tipo | nc",
            "tipo nc",
            "atividade",
            "evento",
            "servico",
            "serviço",
            "classe kartado",
        ),
        "vencimento": _pick_hdr_col(
            hdr_mae,
            "prazo",
            "data do prazo",
            "data prazo",
            "data reparo",
            "data do reparo",
            "vencimento",
            "data vencimento",
            "data nc",
        ),
        "rodovia": _pick_hdr_col(hdr_mae, "rodovia"),
        "codigo": _pick_hdr_col(hdr_mae, "codigo de fiscalizacao", "codigo fiscalizacao", "codigo"),
        "grupo": _pick_hdr_col(hdr_mae, "grupo", "grupo atividade", "grupo de atividade"),
        "km_i": _pick_hdr_col(hdr_mae, "km inicial | km", "km inicial", "km inicio"),
        "km_i_m": _pick_hdr_col(hdr_mae, "km inicial | m", "metros inicial", "m inicial", "m inicio"),
        "km_f": _pick_hdr_col(hdr_mae, "km final | km", "km final", "km fim"),
        "km_f_m": _pick_hdr_col(hdr_mae, "km final | m", "metros final", "m final"),
        "sentido": _pick_hdr_col(hdr_mae, "sentido"),
        "data_con": _pick_hdr_col(
            hdr_mae,
            "data da fiscalizacao",
            "data fiscalizacao",
            "data constatacao",
            "data da constatacao",
        ),
        "responsavel": _pick_hdr_col(
            hdr_mae,
            "responsavel tecnico",
            "responsavel tecnico fiscal",
            "responsavel",
            "responsavel fiscal",
            "fiscal responsavel",
        ),
    }


def chave_agrupamento_exportar_rotina_valores(
    tipo_nc: str | None,
    vencimento_raw,
    rodovia: str | None,
    responsavel: str | None,
    *,
    codigo_fallback: str | None = None,
) -> tuple[str, ...]:
    """Mesma chave que `_chave_agrupamento_linha_exportar`, com valores já lidos (ex.: lista NC do M02)."""
    tipo_raw = tipo_nc
    if not (tipo_raw and str(tipo_raw).strip()):
        tipo_raw = codigo_fallback or "NC"
    t = _limpar_str(tipo_raw).casefold()

    v_norm = ""
    if vencimento_raw is not None and str(vencimento_raw).strip() != "":
        dt = parse_data(vencimento_raw)
        if dt:
            v_norm = dt.strftime("%Y-%m-%d")
        else:
            v_norm = _limpar_str(vencimento_raw).casefold()

    return (
        t,
        v_norm,
        _limpar_str(rodovia).casefold(),
        _limpar_str(responsavel).casefold(),
    )


def _chave_agrupamento_linha_exportar(
    ws,
    row: int,
    cols_meta: dict[str, int | None],
) -> tuple[str, ...]:
    col_tipo = cols_meta.get("tipo_nc") or COL_TIPO_NC
    tipo_nc = _valor_tipo_nc(ws, row, col_tipo)
    if not tipo_nc or not str(tipo_nc).strip():
        cod_c = cols_meta.get("codigo")
        tipo_nc = (_cell(ws, row, cod_c) if cod_c else "") or "NC"

    def cel(col_key: str) -> str:
        c = cols_meta.get(col_key)
        if not c:
            return ""
        return _limpar_str(ws.cell(row=row, column=c).value)

    ven_c = cols_meta.get("vencimento")
    v_raw = ws.cell(row=row, column=ven_c).value if ven_c else None

    return chave_agrupamento_exportar_rotina_valores(
        tipo_nc,
        v_raw,
        cel("rodovia"),
        cel("responsavel"),
    )


def _agrupar_linhas_exportar_rotina(
    ws,
    linha_ini: int,
    ultima: int,
    cols_meta: dict[str, int | None],
) -> list[list[int]]:
    ordem: list[tuple[str, ...]] = []
    buckets: dict[tuple[str, ...], list[int]] = {}
    for r in range(linha_ini, ultima + 1):
        chave = _chave_agrupamento_linha_exportar(ws, r, cols_meta)
        if chave not in buckets:
            ordem.append(chave)
            buckets[chave] = []
        buckets[chave].append(r)
    return [buckets[k] for k in ordem]


def _texto_parece_apenas_data_ou_serial_excel(s: str) -> bool:
    """Evita usar data/prazo (ou serial numérico) como «serv» no nome Art_011."""
    t = _limpar_str(s)
    if not t or len(t) > 48:
        return False
    if re.fullmatch(r"\d{1,2}[_/\-]\d{1,2}[_/\-]\d{2,4}", t):
        return True
    if re.fullmatch(r"\d{4,5}(?:\.\d+)?", t.replace(",", ".")):
        try:
            fv = float(t.replace(",", "."))
        except ValueError:
            return False
        return 30000 <= fv <= 600000
    dt = parse_data(t)
    if not dt:
        return False
    if re.search(r"[A-Za-zÀ-ÿ]{4,}", t):
        return False
    return True


def _tipo_nc_texto_para_nome_exportar_art011(
    ws, row: int, col_tipo_fallback: int | None, col_codigo: int | None
) -> str:
    """Art_011 / Exportar: texto da **Atividade** (coluna detetada, tipicamente N), não «Grupo de atividade» (Q/P). Q (17) só se a detetada estiver vazia ou for data/serial."""
    cfb = int(col_tipo_fallback) if col_tipo_fallback else COL_TIPO_NC
    if cfb and cfb > 0:
        t = _limpar_str(_cell(ws, row, cfb))
        if t and not _texto_parece_apenas_data_ou_serial_excel(t):
            return t
    if cfb != COL_TIPO_NC:
        t = _limpar_str(_cell(ws, row, COL_TIPO_NC))
        if t and not _texto_parece_apenas_data_ou_serial_excel(t):
            return t
    if col_codigo:
        t = _limpar_str(_cell(ws, row, int(col_codigo)))
        if t and not _texto_parece_apenas_data_ou_serial_excel(t):
            return t
    return "NC"


def _nome_arquivo_linha_exportar(ws, row: int, cols_meta: dict[str, int | None]) -> str:
    col_q_fb = cols_meta.get("tipo_nc") or COL_TIPO_NC
    cod_c = cols_meta.get("codigo")
    tipo_raw = _tipo_nc_texto_para_nome_exportar_art011(ws, row, col_q_fb, cod_c)
    cr = cols_meta.get("rodovia")
    rod = _cell(ws, row, cr) if cr else ""
    dc_c = cols_meta.get("data_con")
    data_con = _cell(ws, row, dc_c) if dc_c else ""
    dv_c = cols_meta.get("vencimento")
    data_venc = _cell(ws, row, dv_c) if dv_c else ""
    return _nome_arquivo(rod, tipo_raw, data_con, data_venc)


def _limitar_nome_ficheiro_exportar_windows(
    pasta: Path,
    nome: str,
    *,
    max_caminho: int = 240,
    max_basename: int = 180,
) -> str:
    """
    Encurta o nome do .xlsx para caminhos curtos (Explorer, ZIP, APIs sem ``\\\\?\\``).
    Preserva o sufixo « - Prazo - dd-mm-aaaa» quando existir.
    """
    nome = (nome or "").strip()
    if not nome:
        return "export.xlsx"
    if not nome.lower().endswith(".xlsx"):
        nome = _sanitizar_nome_xlsx(f"{Path(nome).stem}.xlsx")
    pasta_s = str(Path(pasta))
    por_caminho = max(40, max_caminho - len(pasta_s) - 1)
    limite = min(max_basename, por_caminho)
    return truncar_nome_preservando_sufixo_prazo_m01(nome, limite)


def _alocar_nome_exportar_unico(ocupados: set[str], pasta: Path, nome_preferido: str) -> str:
    nome_s = _limitar_nome_ficheiro_exportar_windows(pasta, _sanitizar_nome_xlsx(nome_preferido))
    if not nome_s:
        nome_s = "export.xlsx"
    cand = nome_s
    n = 2
    stem, ext = Path(cand).stem, Path(cand).suffix or ".xlsx"
    while cand in ocupados:
        cand = _limitar_nome_ficheiro_exportar_windows(
            pasta,
            _sanitizar_nome_xlsx(f"{stem} - {n}{ext}"),
        )
        n += 1
    ocupados.add(cand)
    return cand


def gravar_planilha_mae_kartado_pasta_exportar(
    arquivo_mae: Path,
    pasta_m01: Path,
    nome_ficheiro: str,
) -> list[Path]:
    """
    Mapeia cabeçalho da mãe -> cabeçalho do template (mesmos nomes, posições variáveis).
    Vários ficheiros quando a chave de agrupamento difere (**atividade**, prazo, rodovia,
    responsável técnico**; linhas com essa quadra iguais ficam no mesmo .xlsx). O **nome** de cada ficheiro segue a Art_011 (texto da coluna **Atividade** detetada na mãe — não «Grupo de atividade» na Q — para a cadeia de abreviações ``serv``).
    Coluna V (sequência foto): 1..n por ficheiro (sobrescreve cópia da mãe nessa coluna).
    Com um único grupo e ``nome_ficheiro`` preenchido, usa esse nome (compatível com o fluxo M01).
    """
    sem_saida: list[Path] = []
    if not arquivo_mae.is_file():
        return sem_saida
    tpl = Path(TEMPLATE_EXPORTAR_ROTINA)
    if not tpl.is_file():
        logger.error("Template Exportar Rotina não encontrado: %s", tpl)
        return sem_saida
    pasta_out = Path(pasta_m01) / EXPORTAR_KARTADO_MAE_SUBDIR
    garantir_pasta(pasta_out)
    nome_ok = _sanitizar_nome_xlsx(nome_ficheiro) if nome_ficheiro else ""

    COL_R = 18

    try:
        hdr_tpl: dict[str, int] = {}
        max_c_tpl = 0
        alinh_modelo: dict[int, object] = {}
        borda_modelo: dict[int, object] = {}

        with abrir_workbook(tpl, read_only=True, data_only=True) as wb_hdr:
            ws_hdr = wb_hdr.active
            linha_hdr_tpl, linha_sub_tpl = _detectar_cabecalho_template(ws_hdr)
            linha_ini_dados_tpl = (linha_sub_tpl or linha_hdr_tpl) + 1
            hdr_tpl = _ler_cabecalhos_planilha_hierarquico(
                ws_hdr, linha_hdr_tpl, linha_sub_tpl
            )
            max_c_tpl = int(ws_hdr.max_column or 0)
            for c in range(1, max_c_tpl + 1):
                src_tpl = ws_hdr.cell(row=linha_ini_dados_tpl, column=c)
                if isinstance(src_tpl, MergedCell):
                    continue
                alinh_modelo[c] = copy(src_tpl.alignment) if src_tpl.alignment is not None else None
                borda_modelo[c] = copy(src_tpl.border) if src_tpl.border is not None else None

        if not hdr_tpl:
            logger.error("Template sem cabeçalhos detectáveis: %s", tpl.name)
            return sem_saida

        mapa_tpl_para_mae: dict[int, int] = {}
        linha_ini_dados_mae = M01_LINHA_INICIO
        ultima_linha_mae = linha_ini_dados_mae - 1
        col_codigo_mae = COL_CODIGO
        cols_meta: dict[str, int | None] = {}
        grupos_linhas: list[list[int]] = []

        with abrir_workbook(arquivo_mae, read_only=True, data_only=True) as wb_leitura:
            ws_lv = wb_leitura.active
            linha_hdr_mae, linha_sub_mae = _detectar_cabecalho_template(ws_lv)
            hdr_mae = _ler_cabecalhos_planilha_hierarquico(
                ws_lv, linha_hdr_mae, linha_sub_mae
            )
            if not hdr_mae:
                melhor: dict[str, int] = {}
                melhor_tpl = (linha_hdr_mae, linha_sub_mae)
                limite_hdr = min(int(ws_lv.max_row or 0), 8)
                for r in range(1, limite_hdr + 1):
                    cand_sem_sub = _ler_cabecalhos_planilha_hierarquico(ws_lv, r, None)
                    if len(cand_sem_sub) > len(melhor):
                        melhor = cand_sem_sub
                        melhor_tpl = (r, None)
                    if r + 1 <= limite_hdr:
                        cand_com_sub = _ler_cabecalhos_planilha_hierarquico(ws_lv, r, r + 1)
                        if len(cand_com_sub) > len(melhor):
                            melhor = cand_com_sub
                            melhor_tpl = (r, r + 1)
                if melhor:
                    linha_hdr_mae, linha_sub_mae = melhor_tpl
                    hdr_mae = melhor
            if not hdr_mae:
                logger.error("Planilha-mãe sem cabeçalhos detectáveis: %s", arquivo_mae.name)
                return sem_saida

            faltando_no_mae: list[str] = []
            for header_norm, col_tpl in hdr_tpl.items():
                col_mae = hdr_mae.get(header_norm)
                if col_mae is not None:
                    mapa_tpl_para_mae[col_tpl] = col_mae
                else:
                    faltando_no_mae.append(header_norm)
            if faltando_no_mae:
                logger.info(
                    "Cabeçalhos do template ausentes na mãe (deixados vazios): %s",
                    ", ".join(sorted(faltando_no_mae)),
                )
            _refinar_mapa_tpl_metros_km(ws_lv, hdr_mae, hdr_tpl, mapa_tpl_para_mae)

            col_codigo_mae = (
                hdr_mae.get("codigo de fiscalizacao")
                or hdr_mae.get("codigo fiscalizacao")
                or hdr_mae.get("codigo")
                or COL_CODIGO
            )
            if not col_codigo_mae:
                logger.error("Mãe sem coluna de código para detectar início dos dados.")
                return sem_saida
            if not mapa_tpl_para_mae:
                offset = max(int(col_codigo_mae) - COL_CODIGO, 0)
                max_c_tpl_local = max_c_tpl
                for col_tpl in range(3, min(max_c_tpl_local, 21) + 1):
                    if col_tpl == COL_R:
                        continue
                    mapa_tpl_para_mae[col_tpl] = col_tpl + offset

            linha_ini_dados_mae = _detectar_linha_inicio_dados_mae(ws_lv, col_codigo_mae)
            ultima_linha_mae = linha_ini_dados_mae - 1
            colunas_mapeadas_mae = sorted(
                {c for c in mapa_tpl_para_mae.values() if c is not None and c > 0}
            )
            for r in range(int(ws_lv.max_row or 0), linha_ini_dados_mae - 1, -1):
                if ws_lv.cell(row=r, column=col_codigo_mae).value not in (None, ""):
                    ultima_linha_mae = r
                    break
                if any(ws_lv.cell(row=r, column=c).value not in (None, "") for c in colunas_mapeadas_mae):
                    ultima_linha_mae = r
                    break
            if ultima_linha_mae < linha_ini_dados_mae:
                logger.warning("Mãe sem linhas de dados: %s", arquivo_mae.name)
                return sem_saida

            cols_meta = _resolver_cols_metadados_agrupamento_exportar(hdr_mae)
            col_tipo_det = _detectar_col_tipo_nc(
                ws_lv, fallback=int(cols_meta.get("tipo_nc") or COL_TIPO_NC)
            )
            cols_meta["tipo_nc"] = col_tipo_det
            grupos_linhas = _agrupar_linhas_exportar_rotina(
                ws_lv, linha_ini_dados_mae, ultima_linha_mae, cols_meta
            )

        paths_out: list[Path] = []
        ocupados_nomes: set[str] = set()

        for linhas_grupo in grupos_linhas:
            if not linhas_grupo:
                continue
            destino: Path | None = None
            try:
                with abrir_workbook(arquivo_mae, read_only=True, data_only=True) as wb_mae:
                    ws_m = wb_mae.active
                    if len(grupos_linhas) == 1 and nome_ok:
                        nome_dest = _alocar_nome_exportar_unico(ocupados_nomes, pasta_out, nome_ok)
                    else:
                        nome_dest = _alocar_nome_exportar_unico(
                            ocupados_nomes,
                            pasta_out,
                            _nome_arquivo_linha_exportar(ws_m, linhas_grupo[0], cols_meta),
                        )

                    destino = encurtar_nome_em_pasta(pasta_out, nome_dest)
                    shutil.copy2(str_caminho_io_windows(tpl), str_caminho_io_windows(destino))

                    with abrir_workbook(destino, read_only=False, data_only=False) as wb_out:
                        ws_t = wb_out.active
                        n_lin = len(linhas_grupo)
                        ultimo_tpl_necessario = linha_ini_dados_tpl + n_lin - 1
                        limpar_ate = max(int(ws_t.max_row or 0), ultimo_tpl_necessario)
                        for r in range(linha_ini_dados_tpl, limpar_ate + 1):
                            for c in range(1, max_c_tpl + 1):
                                cell = ws_t.cell(row=r, column=c)
                                if isinstance(cell, MergedCell):
                                    continue
                                cell.value = None
                                if c in alinh_modelo and alinh_modelo[c] is not None:
                                    cell.alignment = copy(alinh_modelo[c])
                                if c in borda_modelo and borda_modelo[c] is not None:
                                    cell.border = copy(borda_modelo[c])

                        for offset, r_mae in enumerate(linhas_grupo):
                            r_tpl = linha_ini_dados_tpl + offset
                            seq = offset + 1
                            for col_tpl in range(1, max_c_tpl + 1):
                                if col_tpl == COL_R or col_tpl == COL_SEQ_FOTO:
                                    continue
                                col_mm = mapa_tpl_para_mae.get(col_tpl)
                                if col_mm is None:
                                    continue
                                dst = ws_t.cell(row=r_tpl, column=col_tpl)
                                if isinstance(dst, MergedCell):
                                    continue
                                dst.value = ws_m.cell(row=r_mae, column=col_mm).value
                                if col_tpl in alinh_modelo and alinh_modelo[col_tpl] is not None:
                                    dst.alignment = copy(alinh_modelo[col_tpl])

                            dst_v = ws_t.cell(row=r_tpl, column=COL_SEQ_FOTO)
                            if not isinstance(dst_v, MergedCell):
                                dst_v.value = seq
                                if COL_SEQ_FOTO in alinh_modelo and alinh_modelo[COL_SEQ_FOTO] is not None:
                                    dst_v.alignment = copy(alinh_modelo[COL_SEQ_FOTO])

                        wb_out.save(str_caminho_io_windows(destino))
                    preservar_ooxml_planilha_pos_openpyxl(tpl, destino)
            except Exception:
                logger.exception(
                    "gravar_planilha_mae_kartado_pasta_exportar: falha ao gravar %s",
                    destino.name if destino else "(sem destino)",
                )
                if destino:
                    try:
                        destino.unlink(missing_ok=True)
                    except OSError:
                        pass
                continue

            if destino is not None and destino.is_file():
                paths_out.append(destino)

        return paths_out
    except Exception:
        logger.exception(
            "gravar_planilha_mae_kartado_pasta_exportar: falha ao preparar exportar (%s)",
            arquivo_mae.name,
        )
        return sem_saida


def _chaves_codigo_fiscalizacao_pdf(cod: str) -> frozenset[str]:
    s = _limpar_str(cod)
    if not s:
        return frozenset()
    keys: set[str] = {s, s.casefold()}
    try:
        f = float(s.replace(",", "."))
        if abs(f - round(f)) < 1e-9:
            si = str(int(round(f)))
            keys.add(si)
            keys.add(si.casefold())
    except (ValueError, TypeError, OverflowError):
        pass
    if "-" in s:
        b = s.split("-")[-1].strip()
        if b:
            keys.add(b)
            keys.add(b.casefold())
            try:
                fb = float(b.replace(",", "."))
                if abs(fb - round(fb)) < 1e-9:
                    bi = str(int(round(fb)))
                    keys.add(bi)
                    keys.add(bi.casefold())
            except (ValueError, TypeError, OverflowError):
                pass
    return frozenset(k for k in keys if k)


def _mapa_observacoes_desde_ncs_pdf(ncs) -> dict[str, str]:
    out: dict[str, str] = {}
    for nc in ncs:
        obs = str(getattr(nc, "observacao", None) or "").strip()
        if not obs:
            continue
        cod = str(getattr(nc, "codigo", None) or "").strip()
        if not cod:
            continue
        for k in _chaves_codigo_fiscalizacao_pdf(cod):
            out[k] = obs
    return out


def _observacao_pdf_para_codigo_mae(mapa: dict[str, str], codigo_mae: str) -> str:
    if not mapa:
        return ""
    for k in _chaves_codigo_fiscalizacao_pdf(_limpar_str(codigo_mae)):
        v = mapa.get(k)
        if v:
            return v
    return ""


def _carregar_mapa_observacao_pdf(arquivo_mae: Path, pdf_constatacao: Path | None) -> dict[str, str]:
    path = pdf_constatacao if (pdf_constatacao and pdf_constatacao.is_file()) else None
    if path is None:
        cand = arquivo_mae.with_suffix(".pdf")
        if cand.is_file():
            path = cand
    if path is None or not path.is_file():
        return {}
    try:
        from .analisar_pdf_nc import parse_pdf_nc
    except ImportError:
        from modulos.analisar_pdf_nc import parse_pdf_nc
    try:
        raw = path.read_bytes()
        ncs = parse_pdf_nc(raw)
        m = _mapa_observacoes_desde_ncs_pdf(ncs)
        if m:
            logger.info(
                "M01 Kartado: %d chave(s) com «Observação» extraída(s) do PDF %s",
                len(m),
                path.name,
            )
        elif ncs:
            logger.debug(
                "M01 Kartado: %d NC(s) parseada(s) de %s mas nenhuma tem campo «Observação» "
                "(relatório resumido — col. AA ficará vazia)",
                len(ncs),
                path.name,
            )
        else:
            logger.debug(
                "M01 Kartado: %s — sem blocos de NC reconhecíveis no PDF; col. AA ficará vazia",
                path.name,
            )
        return m
    except Exception as e:
        logger.warning("M01 Kartado: não foi possível ler observações do PDF %s — %s", path.name, e)
        return {}


def _nome_arquivo(rodovia_raw: str, tipo_nc: str,
                  data_constatacao, data_prazo) -> str:
    """
    Nome na pasta Exportar = ``Art_011`` / ``sfile_2`` (``01 - Art_011_EAF_Separar_Mod_Exc_NC.bas`` ~228):
    ``yyyymmdd`` a partir da data constatação (col D), ``rod`` = ``Left(F,6)`` + SPI 10→102-300,
    ``serv`` = abreviatura pelo texto da **coluna Q** (17), como ``Range("Q")`` na macro (``M01_SERVICO_ABREV_ART011`` + lookup tolerante),
    sufixo ``- Prazo - dd-mm-aaaa`` a partir da coluna prazo (T). Extensão ``.xlsx`` no pipeline Python.
    """
    rod_info = normalizar_rodovia_eaf(rodovia_raw, RODOVIAS)
    # VBA: rod = Left(F,6); se "SPI 10" então "SPI 102-300"; senão usa rod
    rod_6    = str(rodovia_raw).strip()[:6]
    rod_nome = RODOVIA_NOME_SEPARAR.get(rod_info["tag"], rod_6) if rod_info["tag"] != "FORA" else sanitizar_nome(str(rodovia_raw)[:10])

    # Abreviação do serviço: Art_011 (``M01_SERVICO_ABREV_ART011``); se não houver match, ``SERVICO_ABREV`` legado; último recurso texto sanitizado.
    tipo_st = _limpar_str(tipo_nc)
    serv_abrev = m01_servico_abrev_art011_lookup(tipo_st)
    if serv_abrev:
        serv_abrev = sanitizar_nome(serv_abrev)
    else:
        serv_abrev = SERVICO_ABREV.get(tipo_st, sanitizar_nome(tipo_nc[:30]))

    # Datas: evita 00000000/00-00-0000 no nome usando data de hoje como fallback
    dt_con  = parse_data(data_constatacao)
    dt_praz = parse_data(data_prazo)
    if not dt_con:
        dt_con = datetime.now()
    if not dt_praz:
        dt_praz = datetime.now()
    yyyymmdd = data_yyyymmdd(dt_con)
    # Sufixo como na macro Art_011: " - Prazo - dd-mm-aaaa" (hífens; evita '/' no nome)
    prazo_s = dt_praz.strftime("%d-%m-%Y") if dt_praz else datetime.now().strftime("%d-%m-%Y")

    nome = (
        f"{yyyymmdd} - CONSTATAÇÕES NC {M01_LOTE} "
        f"({rod_nome} - {serv_abrev}) - Prazo - {prazo_s}.xlsx"
    )
    return _sanitizar_nome_xlsx(nome)


def _nome_arquivo_consolidado_eaf(
    linhas_info: list,
    linhas_rows: list[int],
    linha_inicio: int,
    arquivo_mae: Path | None = None,
) -> str:
    """
    Um único Excel: mesmo padrão Art_011 que ``_nome_arquivo`` (data, rodovia, **serv** da coluna Q),
    com sufixo ``- Consolidado`` e, se houver ficheiro-mãe, o stem da mãe para desambiguar jobs na mesma pasta.
    """
    if not linhas_rows:
        dt_ref = datetime.now()
        yyyymmdd = data_yyyymmdd(dt_ref)
        stem_mae = (
            sanitizar_nome(Path(arquivo_mae).stem, max_len=60).strip(" -.")
            if arquivo_mae
            else ""
        )
        if stem_mae:
            return _sanitizar_nome_xlsx(
                f"{yyyymmdd} - CONSTATAÇÕES NC {M01_LOTE} - Consolidado - {stem_mae}.xlsx"
            )
        return _sanitizar_nome_xlsx(f"{yyyymmdd} - CONSTATAÇÕES NC {M01_LOTE} - Consolidado.xlsx")

    r0 = linhas_rows[0]
    idx0 = r0 - linha_inicio
    if not (0 <= idx0 < len(linhas_info)):
        dt_ref = datetime.now()
        yyyymmdd = data_yyyymmdd(dt_ref)
        stem_mae = (
            sanitizar_nome(Path(arquivo_mae).stem, max_len=60).strip(" -.")
            if arquivo_mae
            else ""
        )
        if stem_mae:
            return _sanitizar_nome_xlsx(
                f"{yyyymmdd} - CONSTATAÇÕES NC {M01_LOTE} - Consolidado - {stem_mae}.xlsx"
            )
        return _sanitizar_nome_xlsx(f"{yyyymmdd} - CONSTATAÇÕES NC {M01_LOTE} - Consolidado.xlsx")

    data_reparo, rodov_raw, tipo_nc, data_const_0 = linhas_info[idx0]
    datas_const: list[datetime] = []
    for r in linhas_rows:
        idx = r - linha_inicio
        if 0 <= idx < len(linhas_info):
            *_, dc = linhas_info[idx]
            dt = parse_data(dc)
            if dt:
                datas_const.append(dt)
    data_con_arg = min(datas_const) if datas_const else data_const_0
    base = Path(_nome_arquivo(rodov_raw, tipo_nc, data_con_arg, data_reparo))
    stem_mae = ""
    if arquivo_mae:
        stem_mae = sanitizar_nome(Path(arquivo_mae).stem, max_len=60).strip(" -.")
    if stem_mae and stem_mae.lower() not in base.stem.lower():
        nome = f"{base.stem} - Consolidado - {stem_mae}{base.suffix}"
    else:
        nome = f"{base.stem} - Consolidado{base.suffix}"
    return _sanitizar_nome_xlsx(nome)


def executar(arquivo_mae: Path, pasta_destino: Path | None = None,
             callback_progresso=None, sobrescrever: bool = False,
             um_arquivo_por_nc: bool = False,
             copia_planilha_mae: bool | None = None,
             unico_arquivo_organizado: bool | None = None,
             pdf_constatacao: Path | None = None) -> list[Path]:
    """
    Processa a planilha-mãe EAF e gera os arquivos individuais de NC.
    sobrescrever: se True, regrava arquivos que já existem (útil em testes locais).
    um_arquivo_por_nc: se True, um Excel por linha; senão agrupa só linhas com **todas** as colunas iguais
    (incl. datas de reparo/execução — valores diferentes → ficheiros distintos). Ignorado quando a saída
    consolidada em único ficheiro está ativa (ver ``unico_arquivo_organizado``).

    copia_planilha_mae: None → usa ``M01_COPIA_PLANILHA_MAE`` (env ``ARTESP_M01_COPIA_PLANILHA_MAE``);
    True → fluxo Art_011 (``Template_EAF.xlsx`` + linhas da mãe coladas sem alterar valores); False → templates Kartado por atividade.

    unico_arquivo_organizado: None → **um** .xlsx com todas as NCs (Art_011 e Kartado), ordenadas por
    rodovia, atividade e código. False força vários ficheiros por grupo de linhas idênticas (comportamento antigo).

    Com modo cópia mãe, a planilha-mãe é gravada no disco após padronizar I, K e V.

    pdf_constatacao: PDF de constatação (opcional). Só no fluxo Kartado (``copia_planilha_mae=False``): o campo
    «Observação» de cada NC no PDF (``parse_pdf_nc``) é classificada no template Kartado (Localização Pista **X**,
    Localização Tipo **Y**, texto livre em **AA** «Observações»), casando pelo código de fiscalização. Se ``None``,
    tenta-se ``<mesmo stem que a mãe>.pdf`` ao lado do ficheiro Excel.

    Parâmetros
    ----------
    arquivo_mae      : Path para o .xls/.xlsx mãe (planilha EAF completa).
    pasta_destino    : Pasta onde os XLS serão salvos (padrão: M01_EXPORTAR).
    callback_progresso : função(atual, total, msg) para atualizar GUI.
    um_arquivo_por_nc : se True, um arquivo por linha da EAF (uma NC por Excel).
    copia_planilha_mae : força o modo M01 (None = configuração global).
    unico_arquivo_organizado : None = automático (**um** ficheiro consolidado; Art_011 ou Kartado).
    pdf_constatacao : PDF para enriquecer «Observações» (AA) e localizações X/Y no Kartado; ver acima.

    Retorna
    -------
    Lista de Path dos arquivos gerados.
    """
    pasta_destino = Path(pasta_destino) if pasta_destino else M01_EXPORTAR
    garantir_pasta(pasta_destino)
    # Evita reuso stale de cache de templates entre execuções consecutivas.
    try:
        _listar_candidatos_templates_kartado_cache.cache_clear()
    except Exception:
        pass
    usar_copia_mae = M01_COPIA_PLANILHA_MAE if copia_planilha_mae is None else copia_planilha_mae
    if um_arquivo_por_nc:
        consolidar_um_ficheiro = False
    elif unico_arquivo_organizado is None:
        # Com template único Kartado, consolidar por defeito mesmo sem cópia mãe
        consolidar_um_ficheiro = True
    else:
        consolidar_um_ficheiro = bool(unico_arquivo_organizado)

    ValidadorArquivoEAF.validar(arquivo_mae)
    suff = arquivo_mae.suffix.lower()

    # openpyxl só lê .xlsx/.xlsm; converter .xls quando necessário
    if suff == ".xls":
        arquivo_mae = _converter_xls_para_xlsx(arquivo_mae)
    elif suff not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        # Sem extensão ou extensão desconhecida: tentar .xls (xlrd) primeiro
        try:
            xlrd.open_workbook(str(arquivo_mae))
            arquivo_mae = _converter_xls_para_xlsx(arquivo_mae)
        except xlrd.biffh.XLRDError:
            # Pode ser .xlsx salvo sem extensão – deixar load_workbook tentar
            pass

    obs_pdf_por_codigo: dict[str, str] = {}
    if not usar_copia_mae:
        obs_pdf_por_codigo = _carregar_mapa_observacao_pdf(arquivo_mae, pdf_constatacao)
        if not obs_pdf_por_codigo:
            p_try = (
                pdf_constatacao
                if (pdf_constatacao is not None and Path(pdf_constatacao).is_file())
                else arquivo_mae.with_suffix(".pdf")
            )
            if not Path(p_try).is_file():
                logger.info(
                    "M01 Kartado: sem ficheiro %s ao lado da mãe; AA/X/Y só com observação do PDF se o gravar aí ou passar pdf_constatacao.",
                    arquivo_mae.with_suffix(".pdf").name,
                )

    logger.info("M01 ARTESP: classes especiais desativadas.")

    # Abrir só para ler: detectar colunas, última linha e lista de (data, rodovia, tipo) por linha.
    # Modo cópia mãe (Art_011): saídas usam Template_EAF.xlsx + linhas coladas da mãe (não cópia binária da mãe).
    logger.info(f"Abrindo planilha-mãe (somente leitura): {arquivo_mae.name}")
    with abrir_workbook(arquivo_mae, read_only=False) as wb_mae:
        ws = wb_mae.active

        col_data_reparo = _detectar_col_data_reparo(ws, fallback=COL_DATA_NC)
        col_tipo_nc = _detectar_col_tipo_nc(ws, fallback=COL_TIPO_NC)
        col_tipo_ativ = detectar_coluna_tipo_de_atividade_eaf(ws, COL_TIPO_ATIV)
        col_macro_grupo_m = _detectar_col_macrogrupo_mae(ws, fallback=COL_EAF_GRUPO_MACRO_M)
        col_grupo_atividade_desc = _detectar_col_grupo_atividade_descricao_mae(ws)
        col_atividade_mae_n = _detectar_col_atividade_mae_col_n(ws, fallback=COL_EAF_ATIVIDADE_N)
        col_codigo = _detectar_col_codigo_fiscalizacao(ws, fallback=COL_CODIGO)
        col_rodovia = _detectar_col_rodovia(ws, fallback=COL_RODOVIA)
        col_data_con = _detectar_col_data_con(ws, fallback=COL_DATA_CON)
        col_hora_fisc = _detectar_col_hora_fiscalizacao(ws, fallback=COL_HORA_FISC)
        col_km_i_full = _detectar_col_km_inicial(ws, fallback=COL_KM_I_FULL)
        col_km_f_full = _detectar_col_km_final(ws, fallback=COL_KM_F_FULL)
        col_km_i_m = _detectar_col_km_i_metros(ws, fallback=COL_KM_I_M)
        col_km_f_m = _detectar_col_km_f_metros(ws, fallback=COL_KM_F_M)
        col_sentido = _detectar_col_sentido(ws, fallback=COL_SENTIDO)
        col_seq_foto = _detectar_col_seq_foto(ws, fallback=COL_SEQ_FOTO)
        col_responsavel = _detectar_col_responsavel(ws, fallback=COL_RESPONSAVEL)
        linha_inicio = _detectar_linha_inicio_dados(ws, col_codigo)
        logger.info(f"Coluna 'Data Reparo': {col_data_reparo}")
        logger.info(f"Coluna 'Tipo NC/Atividade': {col_tipo_nc}")
        logger.info(f"Coluna 'Tipo de Atividade' (mãe): {col_tipo_ativ}")
        logger.info(f"Coluna macrogrupo (M / Classe Kartado): {col_macro_grupo_m}")
        logger.info(
            f"Coluna «Grupo de atividade» (texto descritivo na mãe): {col_grupo_atividade_desc!r}"
        )
        logger.info(f"Coluna atividade (N / Classe Kartado): {col_atividade_mae_n}")
        logger.info(f"Coluna 'Código Fiscalização': {col_codigo}")
        logger.info(f"Coluna 'Rodovia': {col_rodovia}")
        logger.info(f"Coluna 'Data Constatação': {col_data_con}")
        logger.info(f"Coluna 'Horário fiscalização': {col_hora_fisc}")
        logger.info(f"Coluna 'km inicial': {col_km_i_full}")
        logger.info(f"Coluna 'km final': {col_km_f_full}")
        logger.info(f"Coluna 'm inicial': {col_km_i_m}")
        logger.info(f"Coluna 'm final': {col_km_f_m}")
        logger.info(f"Coluna 'Sentido': {col_sentido}")
        logger.info(f"Coluna 'Seq. foto': {col_seq_foto}")
        logger.info(f"Coluna 'Responsável': {col_responsavel}")
        logger.info(f"Linha início dados: {linha_inicio}")

        ultima_linha = linha_inicio - 1
        for r in range(ws.max_row, linha_inicio - 1, -1):
            if ws.cell(row=r, column=col_codigo).value:
                ultima_linha = r
                break

        total_linhas = ultima_linha - linha_inicio + 1
        logger.info(f"Linhas de dados: {total_linhas} (L{linha_inicio}–L{ultima_linha})")

        linhas_info = []
        for r in range(linha_inicio, ultima_linha + 1):
            tipo_para_nome = _tipo_nc_texto_para_nome_exportar_art011(ws, r, col_tipo_nc, col_codigo)
            linhas_info.append((
                _cell(ws, r, col_data_reparo),
                _cell(ws, r, col_rodovia),
                tipo_para_nome,
                _cell(ws, r, col_data_con),
            ))

        fallback_tpl = _caminho_template_eaf()
        max_col = ws.max_column

        if usar_copia_mae:
            for r in range(linha_inicio, ultima_linha + 1):
                _padronizar_colunas_km(ws, r)
            qseq = 1
            for r in range(linha_inicio, ultima_linha + 1):
                ws.cell(row=r, column=col_seq_foto).value = qseq
                qseq += 1
            wb_mae.save(str_caminho_io_windows(arquivo_mae))
            logger.info("Planilha-mãe gravada (I, K, V) — modo cópia mãe (Art_011).")

        if consolidar_um_ficheiro:
            candidatos: list[int] = []
            for r in range(linha_inicio, ultima_linha + 1):
                tipo_nc = _valor_tipo_nc(ws, r, col_tipo_nc)
                if not tipo_nc or not str(tipo_nc).strip():
                    tipo_nc = _cell(ws, r, col_codigo) or "NC"
                if not tipo_nc or not str(tipo_nc).strip():
                    continue
                candidatos.append(r)

            def _chave_ordem_consolidado(rr: int) -> tuple:
                return (
                    _limpar_str(_cell(ws, rr, col_rodovia)).casefold(),
                    _limpar_str(_valor_tipo_nc(ws, rr, col_tipo_nc)).casefold(),
                    _limpar_str(_cell(ws, rr, col_codigo)),
                    rr,
                )

            if not candidatos:
                grupos_ord = []
            else:
                linhas_ord = sorted(candidatos, key=_chave_ordem_consolidado)
                grupos_ord = [(tuple(), linhas_ord)]
        else:
            index_fp: dict[tuple[str, ...], int] = {}
            grupos_ord = []
            for r in range(linha_inicio, ultima_linha + 1):
                tipo_nc = _valor_tipo_nc(ws, r, col_tipo_nc)
                if not tipo_nc or not str(tipo_nc).strip():
                    tipo_nc = _cell(ws, r, col_codigo) or "NC"
                if not tipo_nc or not str(tipo_nc).strip():
                    continue
                fp = _fingerprint_linha_mae(ws, r, max_col, forcar_linha_unica=um_arquivo_por_nc)
                if fp not in index_fp:
                    index_fp[fp] = len(grupos_ord)
                    grupos_ord.append((fp, []))
                grupos_ord[index_fp[fp]][1].append(r)

    # Gerar ficheiros (reabre a mãe para ``ws`` válido durante o loop).
    with abrir_workbook(arquivo_mae, read_only=False, data_only=False) as wb_gen:
        ws = wb_gen.active
        arquivos_gerados: list[Path] = []
        processadas: set[str] = set()
        nomes_emitidos: set[str] = set()

        for fp, linhas_do_grupo in grupos_ord:
            r0 = linhas_do_grupo[0]
            idx0 = r0 - linha_inicio
            data_nc, rodov_raw, tipo_nc, data_con = linhas_info[idx0]

            if not tipo_nc:
                continue

            if consolidar_um_ficheiro:
                nome_arq = _nome_arquivo_consolidado_eaf(
                    linhas_info, linhas_do_grupo, linha_inicio, arquivo_mae
                )
            else:
                nome_arq = _nome_arquivo(rodov_raw, tipo_nc, data_con, data_nc)
            if um_arquivo_por_nc and not consolidar_um_ficheiro:
                codigo = _cell(ws, r0, col_codigo)
                codigo_safe = sanitizar_nome(str(codigo).strip())[:80] if codigo and str(codigo).strip() else f"NC-{r0}"
                stem, ext = Path(nome_arq).stem, Path(nome_arq).suffix
                nome_base = f"{stem} - {codigo_safe}{ext}"
                nome_arq = _sanitizar_nome_xlsx(nome_base)
                n = 1
                while nome_arq in processadas:
                    nome_arq = _sanitizar_nome_xlsx(f"{stem} - {codigo_safe} ({n}){ext}")
                    n += 1
                processadas.add(nome_arq)

            destino = encurtar_nome_em_pasta(pasta_destino, nome_arq)
            garantir_pasta(pasta_destino)

            if um_arquivo_por_nc and not consolidar_um_ficheiro:
                cont = 1
                while destino.exists() and not sobrescrever:
                    stem, ext = Path(nome_arq).stem, Path(nome_arq).suffix
                    nome_arq = _sanitizar_nome_xlsx(f"{stem} ({cont}){ext}")
                    processadas.add(nome_arq)
                    destino = encurtar_nome_em_pasta(pasta_destino, nome_arq)
                    cont += 1
            elif not consolidar_um_ficheiro and destino.exists() and not sobrescrever:
                cont = 1
                while destino.exists() and not sobrescrever:
                    stem, ext = Path(nome_arq).stem, Path(nome_arq).suffix
                    nome_arq = _sanitizar_nome_xlsx(f"{stem} ({cont}){ext}")
                    destino = encurtar_nome_em_pasta(pasta_destino, nome_arq)
                    cont += 1
                if cont > 1:
                    logger.warning(
                        "Colisão de nome (modo agrupado); a gravar variante numerada: %s (%d linhas no grupo)",
                        nome_arq,
                        len(linhas_do_grupo),
                    )

            if not consolidar_um_ficheiro and not um_arquivo_por_nc:
                nomes_emitidos.add(nome_arq)

            tipo_str = _limpar_str(tipo_nc)

            logger.info(
                "Gerando arquivo individual",
                extra={
                    "nome_arquivo": nome_arq,
                    "num_linhas": len(linhas_do_grupo),
                    "tipo_nc": str(tipo_str),
                    "rodovia": str(rodov_raw),
                    "modo": "copia_mae" if usar_copia_mae else "template",
                },
            )

            template_src = _caminho_template_geral_final()
            if template_src is None and usar_copia_mae:
                template_src = fallback_tpl
            if template_src is None or not template_src.is_file():
                raise FileNotFoundError(
                    "Template único Kartado não encontrado: "
                    "'Template - Geral - 4 e 5 - Final.xlsx'. "
                    "Coloque o arquivo em nc_artesp/assets/templates/."
                )

            if usar_copia_mae and (template_src == fallback_tpl or (fallback_tpl.is_file() and template_src.resolve() == fallback_tpl.resolve())):
                shutil.copy2(str_caminho_io_windows(template_src), str_caminho_io_windows(destino))
                with abrir_workbook(arquivo_mae, read_only=False, data_only=False) as wb_mae_linhas:
                    ws_mae_linhas = wb_mae_linhas.active
                    ultima_col = max(int(ws_mae_linhas.max_column or 0), 1)
                    with abrir_workbook(destino) as wb_out:
                        ws_out = wb_out.active
                        _limpar_linhas_dados_eaf_no_sheet(ws_out, M01_LINHA_INICIO)
                        for seq, row_orig in enumerate(sorted(linhas_do_grupo), start=0):
                            row_dest = M01_LINHA_INICIO + seq
                            _copiar_linha_mae_para_template_eaf(
                                ws_mae_linhas, row_orig, ws_out, row_dest, ultima_col
                            )
                        destino_xls = destino.with_suffix(".xls")
                        if destino_xls.exists():
                            destino_xls.unlink()
                            logger.debug("Removido .xls antigo: %s", destino_xls.name)
                        wb_out.save(str_caminho_io_windows(destino))
                arquivos_gerados.append(destino)
                logger.info("  ✓ Salvo (Template EAF + linhas da mãe): %s", destino.name)
                continue

            logger.info(f"  Template base (único): {template_src.name}")
    
            # 1) Cópia binária do template Kartado ou fallback EAF (cabeçalho / formatação)
            shutil.copy2(str_caminho_io_windows(template_src), str_caminho_io_windows(destino))
    
            with abrir_workbook(destino) as wb_copia:
                ws_copia = wb_copia.active
                cols_tpl = _colunas_kartado_por_header(ws_copia)
                _kartado_cols_tpl_unificar_cabecalho_observacoes(cols_tpl)
                col_envio_tpl, col_reparo_tpl = _detectar_colunas_data_no_template(ws_copia)
    
                try:
                    mesmo_que_template_eaf = template_src.resolve() == fallback_tpl.resolve()
                except OSError:
                    mesmo_que_template_eaf = False
                primeira_linha_dados = (
                    PRIMEIRA_LINHA_DADOS if mesmo_que_template_eaf else PRIMEIRA_LINHA_DADOS_TEMPLATE_KARTADO
                )
    
                # 2) Apagar só linhas de dados: Kartado preserva cabeçalho na linha 1; Template_EAF preserva linhas 1–4
                while ws_copia.max_row >= primeira_linha_dados:
                    ws_copia.delete_rows(ws_copia.max_row, 1)

                col_resposta_mae_q = _kartado_garantir_coluna_texto_mae_resposta(
                    ws_copia, cols_tpl, RESPOSTA_PENDENTES_HEADER_MAE_ATIVIDADE_Q
                )
                col_resposta_mae_o = _kartado_garantir_coluna_texto_mae_resposta(
                    ws_copia, cols_tpl, RESPOSTA_PENDENTES_HEADER_MAE_TIPO_ATIV_O
                )
                _kartado_garantir_coluna_texto_mae_resposta(ws_copia, cols_tpl, "Observações")

                # 3) Preencher linhas por cabeçalho Kartado + manter colunas técnicas para M02/M03
                for seq, row_origem in enumerate(linhas_do_grupo, start=1):
                    row_dest = primeira_linha_dados + seq - 1
                    val_envio = ws.cell(row=row_origem, column=col_data_con).value
                    val_reparo = ws.cell(row=row_origem, column=col_data_reparo).value
                    val_hora = ws.cell(row=row_origem, column=col_hora_fisc).value

                    dt_envio = parse_data(val_envio)
                    dt_reparo = parse_data(val_reparo)
                    if dt_reparo is None and dt_envio is not None:
                        dt_reparo = dt_envio + timedelta(days=PRAZO_DIAS_APOS_ENVIO)

                    val_encontrado_dt = _combinar_data_e_hora_fiscalizacao(val_envio, val_hora)
                    val_prazo_dt = _combinar_data_e_hora_fiscalizacao(
                        dt_reparo if dt_reparo else val_reparo,
                        val_hora,
                    )

                    if col_envio_tpl is not None:
                        _aplicar_celula_data_excel(
                            ws_copia,
                            row_dest,
                            col_envio_tpl,
                            val_encontrado_dt if val_encontrado_dt is not None else val_envio,
                            preservar_hora=val_encontrado_dt is not None,
                        )
                    if col_reparo_tpl is not None:
                        if val_prazo_dt is not None:
                            _aplicar_celula_data_excel(
                                ws_copia,
                                row_dest,
                                col_reparo_tpl,
                                val_prazo_dt,
                                preservar_hora=True,
                            )
                        else:
                            _aplicar_celula_data_excel(ws_copia, row_dest, col_reparo_tpl, val_reparo)

                    codigo = _limpar_str(ws.cell(row=row_origem, column=col_codigo).value)
                    codigo_cel = codigo if codigo else None
                    macro_m = _str_linha_eaf(_cell(ws, row_origem, col_macro_grupo_m))
                    ativ_n = _str_linha_eaf(_cell(ws, row_origem, col_atividade_mae_n))
                    desc_ga = (
                        _str_linha_eaf(_cell(ws, row_origem, col_grupo_atividade_desc))
                        if col_grupo_atividade_desc
                        else ""
                    )
                    g_split, a_split = _partir_celula_grupo_atividade_mae_combinada(desc_ga)
                    texto_m_linha = macro_m or _str_linha_eaf(g_split)
                    texto_n_linha = ativ_n or _str_linha_eaf(a_split)
                    tipo_txt = _limpar_str(_valor_tipo_nc(ws, row_origem, col_tipo_nc))
                    atividade_mae_literal = (
                        ativ_n or _str_linha_eaf(a_split) or _limpar_str(_cell(ws, row_origem, col_tipo_nc))
                    )
                    if mesmo_que_template_eaf:
                        classifica = _strip_descricao_kartado_excel(
                            str(atividade_mae_literal or tipo_txt or "")
                        )
                    else:
                        classifica = _resolver_classe_kartado_inteligente_mno(
                            texto_m=texto_m_linha,
                            texto_n=texto_n_linha,
                            texto_o=_str_linha_eaf(_cell(ws, row_origem, col_tipo_ativ)),
                            tipo_txt_detetado=tipo_txt,
                        )
                    rodovia_fmt = _rodovia_fmt_eaf_para_kartado(ws.cell(row=row_origem, column=col_rodovia).value)
                    km_i_int = ws.cell(row=row_origem, column=col_km_i_full).value
                    km_f_int = ws.cell(row=row_origem, column=col_km_f_full).value
                    km_i_m = ws.cell(row=row_origem, column=col_km_i_m).value
                    km_f_m = ws.cell(row=row_origem, column=col_km_f_m).value
                    # Padrão VBA: coluna "km" no Kartado recebe km+metros (ex.: '68+100').
                    km_i_formato = km_formato_arquivo(km_i_int, km_i_m)
                    km_f_formato = km_formato_arquivo(km_f_int, km_f_m)
                    sentido = ws.cell(row=row_origem, column=col_sentido).value
                    km_f_formato, sentido = _normalizar_km_final_e_sentido(km_f_formato, sentido)

                    # Texto só do EAF; uma linha na célula (sem \\n — evita confusão no Excel).
                    relatorio_ref = _limpar_str(ws.cell(row=row_origem, column=col_data_con).value)
                    descricao_kartado = _strip_descricao_kartado_excel(
                        f"{atividade_mae_literal or tipo_txt} --> Relatório EAF Conservação Rotina nº: {relatorio_ref} "
                        f"--> Código NC: {codigo}"
                    )
                    obs_pdf = _observacao_pdf_para_codigo_mae(obs_pdf_por_codigo, codigo)
                    if obs_pdf:
                        obs_n = normalizar_observacao_extraida_pdf(obs_pdf)
                        rota = rotear_observacao_pdf_para_kartado(obs_n)
                        if rota.texto_livre_observacoes:
                            obs_u = _strip_descricao_kartado_excel(rota.texto_livre_observacoes)[:32700]
                            _set_if_header(ws_copia, row_dest, cols_tpl, "Observações", obs_u)
                        if rota.localizacao_pista_x:
                            _set_if_header(
                                ws_copia,
                                row_dest,
                                cols_tpl,
                                "Localização Pista",
                                _strip_descricao_kartado_excel(rota.localizacao_pista_x)[:500],
                            )
                        if rota.localizacao_tipo_y:
                            _set_if_header(
                                ws_copia,
                                row_dest,
                                cols_tpl,
                                "Localização Tipo",
                                _strip_descricao_kartado_excel(rota.localizacao_tipo_y)[:500],
                            )

                    foto_seq = _foto_ref_numerica(ws.cell(row=row_origem, column=col_seq_foto).value)
                    foto_ref_nc = codigo or foto_seq
                    foto_ref_pdf = foto_seq or codigo or ""
                    foto_1 = f"nc ({foto_ref_nc}).jpg" if foto_ref_nc else ""
                    # Mesmo padrão dos ficheiros em disco (pdf_extractor): «PDF (COD).jpg», não «pdf» minúsculo.
                    foto_2 = f"PDF ({foto_ref_pdf}).jpg" if foto_ref_pdf else ""
    
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Origem", "Artesp")
                    _set_if_header(
                        ws_copia,
                        row_dest,
                        cols_tpl,
                        "Classe",
                        _strip_descricao_kartado_excel(str(classifica or tipo_txt or "")),
                    )
                    val_tipo_ativ = _limpar_str(ws.cell(row=row_origem, column=col_tipo_ativ).value)
                    ws_copia.cell(row=row_dest, column=col_resposta_mae_q).value = _strip_descricao_kartado_excel(
                        str(atividade_mae_literal or tipo_txt or "")
                    )
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Atividade", _strip_descricao_kartado_excel(str(atividade_mae_literal or tipo_txt or "")))
                    ws_copia.cell(row=row_dest, column=col_resposta_mae_o).value = val_tipo_ativ
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Rodovia", rodovia_fmt)
                    # km e km de Projeto vêm da EAF (apontamento) — mesmo valor nas duas colunas.
                    _set_if_header(ws_copia, row_dest, cols_tpl, "km", km_i_formato)
                    _set_if_header(ws_copia, row_dest, cols_tpl, "km final", km_f_formato)
                    _set_if_header(ws_copia, row_dest, cols_tpl, "km de Projeto", km_i_formato)
                    _set_if_header(ws_copia, row_dest, cols_tpl, "km final de Projeto", km_f_formato)
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Sentido", sentido)
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Status", "Solicitado")
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Equipe", "Sala Técnica - Soluciona")
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Menu", "Não Conformidades")
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Faixa", "Não se aplica")
                    _set_data_if_header(
                        ws_copia,
                        row_dest,
                        cols_tpl,
                        "Encontrado em",
                        val_encontrado_dt if val_encontrado_dt is not None else val_envio,
                        preservar_hora=val_encontrado_dt is not None,
                    )
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Código de Fiscalização", codigo_cel)
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Código Fiscalização", codigo_cel)
                    _set_data_if_header(
                        ws_copia,
                        row_dest,
                        cols_tpl,
                        "Prazo",
                        val_prazo_dt if val_prazo_dt is not None else (dt_reparo if dt_reparo else val_reparo),
                        preservar_hora=val_prazo_dt is not None,
                    )
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Descrição", descricao_kartado)
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Disciplina", "Conservação de Rotina")
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Doc. Origem", "EAF - ROTINA")
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Motivo", "Não Conformidade")
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Foto_1", foto_1)
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Tipo Foto_1", "Antes")
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Descrição Foto_1", "Imagem - Informações Padrão")
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Foto_2", foto_2)
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Tipo Foto_2", "Antes")
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Descrição Foto_2", "Print PDF do Apontamento")
    
                    # Compatibilidade técnica só quando o template for EAF.
                    # Em template Kartado puro, escrever por índice fixo (C..V) corrompe colunas de negócio.
                    if mesmo_que_template_eaf:
                        ws_copia.cell(row=row_dest, column=COL_CODIGO).value = codigo_cel
                        _aplicar_celula_data_excel(ws_copia, row_dest, COL_DATA_CON, val_envio, preservar_hora=False)
                        ws_copia.cell(row=row_dest, column=COL_RODOVIA).value = _rodovia_fmt_eaf_para_kartado(
                            ws.cell(row=row_origem, column=col_rodovia).value
                        )
                        ws_copia.cell(row=row_dest, column=COL_KM_I_FULL).value = km_i_int
                        ws_copia.cell(row=row_dest, column=COL_KM_F_FULL).value = km_f_int
                        ws_copia.cell(row=row_dest, column=COL_KM_I_M).value = _cell(ws, row_origem, col_km_i_m)
                        ws_copia.cell(row=row_dest, column=COL_KM_F_M).value = _cell(ws, row_origem, col_km_f_m)
                        _padronizar_colunas_km(ws_copia, row_dest)
                        ws_copia.cell(row=row_dest, column=COL_SENTIDO).value = sentido
                        ws_copia.cell(row=row_dest, column=COL_TIPO_NC).value = atividade_mae_literal or tipo_txt
                        _aplicar_celula_data_excel(
                            ws_copia,
                            row_dest,
                            COL_DATA_NC,
                            val_prazo_dt if val_prazo_dt is not None else (dt_reparo if dt_reparo else val_reparo),
                            preservar_hora=val_prazo_dt is not None,
                        )
                        ws_copia.cell(row=row_dest, column=COL_RESPONSAVEL).value = ws.cell(row=row_origem, column=col_responsavel).value
                        ws_copia.cell(row=row_dest, column=COL_SEQ_FOTO).value = foto_ref_pdf
    
                destino_xls = destino.with_suffix(".xls")
                if destino_xls.exists():
                    destino_xls.unlink()
                    logger.debug(f"Removido .xls antigo: {destino_xls.name}")
    
                wb_copia.save(str_caminho_io_windows(destino))
            arquivos_gerados.append(destino)
            logger.info(f"  ✓ Salvo: {destino.name}")

    if not usar_copia_mae and arquivos_gerados and grupos_ord:
        linhas_primeiro = grupos_ord[0][1]
        if linhas_primeiro:
            r0 = linhas_primeiro[0]
            idx0 = r0 - linha_inicio
            if 0 <= idx0 < len(linhas_info):
                data_nc, rodov_raw, tipo_nc, data_con = linhas_info[idx0]
                nome_mae_k = _nome_arquivo(rodov_raw, tipo_nc, data_con, data_nc)
                paths_mae_exp = gravar_planilha_mae_kartado_pasta_exportar(
                    arquivo_mae, pasta_destino, nome_mae_k
                )
                for p_mae in paths_mae_exp:
                    logger.info("Planilha-mãe Kartado (pasta exportar): %s", p_mae.name)

    logger.info(f"Módulo 01 concluído. {len(arquivos_gerados)} arquivo(s) gerado(s).")
    if callback_progresso:
        callback_progresso(total_linhas, total_linhas, "Módulo 01 concluído.")
    return arquivos_gerados
