"""
nc_artesp/utils/excel_io.py
Conversão de .xls (xlrd) → .xlsx (openpyxl) e utilitários de formatação.
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from .helpers import str_caminho_io_windows


def aplicar_bordas(
    caminho_arquivo: "str | Path",
    nome_aba: Optional[str] = None,
    estilo: str = "thin",
) -> None:
    """
    Aplica bordas finas em todas as células da aba do Excel.
    Útil após exportar com pandas to_excel ou gerar planilhas programaticamente.

    Args:
        caminho_arquivo: Caminho do .xlsx (ex.: 'Relatorio_Consolidado_Lote.xlsx').
        nome_aba: Nome da aba (ex.: 'Planilha1'). Se None, usa a primeira aba.
        estilo: Estilo da linha ('thin', 'medium', 'thick'). Default 'thin'.
    """
    import openpyxl
    from openpyxl.styles import Border, Side

    caminho_arquivo = Path(caminho_arquivo)
    if not Path(str_caminho_io_windows(caminho_arquivo)).exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho_arquivo}")

    wb = openpyxl.load_workbook(str_caminho_io_windows(caminho_arquivo))
    ws = wb[nome_aba] if nome_aba else wb.active

    side = Side(style=estilo)
    thin_border = Border(left=side, right=side, top=side, bottom=side)

    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border

    wb.save(str_caminho_io_windows(caminho_arquivo))


def xls_to_xlsx(path_xls: Path, dest: "Path | None" = None) -> Path:
    """
    Lê um .xls com xlrd e grava um .xlsx equivalente.
    dest: caminho de destino opcional; se None, usa o mesmo diretório com sufixo .xlsx.
    Retorna o Path do arquivo .xlsx gerado.
    Se o arquivo já for .xlsx, retorna-o sem conversão.
    """
    from pathlib import Path as _Path
    path_xls = _Path(path_xls)
    if path_xls.suffix.lower() in (".xlsx", ".xlsm"):
        return path_xls

    if dest is not None:
        path_xlsx = _Path(dest)
    else:
        path_xlsx = path_xls.with_suffix(".xlsx")

    if path_xls.stem.endswith("_convertido") and dest is None:
        path_xlsx = path_xls.parent / (path_xls.stem + "_conv2.xlsx")

    import xlrd
    import openpyxl

    book  = xlrd.open_workbook(str_caminho_io_windows(path_xls))
    sheet = book.sheet_by_index(0)

    wb = openpyxl.Workbook()
    ws = wb.active
    if sheet.name:
        ws.title = sheet.name[:31]

    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            cell = sheet.cell(row, col)
            if cell.ctype == xlrd.XL_CELL_EMPTY:
                continue
            xl_cell = ws.cell(row=row + 1, column=col + 1)
            if cell.ctype == xlrd.XL_CELL_NUMBER:
                xl_cell.value = cell.value
            elif cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    import xlrd.xldate as xld
                    from datetime import datetime
                    dt = xld.xldate_as_datetime(cell.value, book.datemode)
                    xl_cell.value = dt
                except Exception:
                    xl_cell.value = cell.value
            elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                xl_cell.value = bool(cell.value)
            else:
                xl_cell.value = cell.value

    wb.save(str_caminho_io_windows(path_xlsx))
    return path_xlsx
