"""
nc_artesp/utils/helpers.py
────────────────────────────────────────────────────────────────────────────
Funções utilitárias do pipeline NC ARTESP.
"""

from __future__ import annotations

import logging
import os
import posixpath
import re
import shutil
import unicodedata
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Optional

import warnings

warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    message=r".*Data Validation extension is not supported.*",
)


def configurar_log(nivel: int = logging.INFO,
                   arquivo: "Path | None" = None) -> None:
    """Configura o logging raiz com formatação padrão."""
    handlers = [logging.StreamHandler()]
    if arquivo:
        try:
            handlers.append(logging.FileHandler(str(arquivo), encoding="utf-8"))
        except Exception:
            pass
    logging.basicConfig(
        level=nivel,
        format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=handlers,
        force=True,
    )


# DATAS

def parse_data(valor) -> Optional[datetime]:
    """Parseia datas em vários formatos; retorna datetime ou None."""
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor
    # date → datetime
    try:
        from datetime import date
        if isinstance(valor, date):
            return datetime(valor.year, valor.month, valor.day)
    except Exception:
        pass
    # número serial Excel (float/int)
    if isinstance(valor, (int, float)):
        try:
            import xlrd
            return datetime(*xlrd.xldate_as_tuple(float(valor), 0)[:6])
        except Exception:
            pass
        try:
            from openpyxl.utils.datetime import from_excel
            return from_excel(valor)
        except Exception:
            pass
    s = str(valor).strip()
    if not s or s.lower() in ("none", "nan", ""):
        return None
    s_norm = re.sub(r" +", " ", s)
    for fmt in (
        "%d/%m/%Y", "%d/%m/%y",
        "%Y-%m-%d", "%Y%m%d",
        "%d-%m-%Y", "%d-%m-%y",
        "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(s_norm, fmt)
        except ValueError:
            pass
    m_serial = re.fullmatch(r"(\d{5,6})(?:\.0+)?", s_norm)
    if m_serial:
        try:
            n = float(m_serial.group(1))
            if 29500 <= n <= 65000:
                from openpyxl.utils.datetime import from_excel

                return from_excel(n)
        except Exception:
            pass
    return None


def data_yyyymmdd(dt: Optional[datetime]) -> str:
    """datetime → 'YYYYMMDD'. Retorna '00000000' se None."""
    if not dt:
        return "00000000"
    return dt.strftime("%Y%m%d")


def data_ddmmaaaa(dt: Optional[datetime]) -> str:
    """datetime → 'DD/MM/AAAA'. Retorna '' se None."""
    if not dt:
        return ""
    return dt.strftime("%d/%m/%Y")


def data_br(dt: Optional[datetime]) -> str:
    """datetime → 'DD/MM/YYYY'. Alias de data_ddmmaaaa."""
    return data_ddmmaaaa(dt)


def timestamp_agora() -> str:
    """Retorna 'YYYYMMDD-HHMM'."""
    return datetime.now().strftime("%Y%m%d-%H%M")


def timestamp_completo() -> str:
    """Retorna 'YYYYMMDD - HHMMSS'."""
    return datetime.now().strftime("%Y%m%d - %H%M%S")


# KM E METROS

def pad_metros(valor) -> str:
    """Normaliza metros para 3 dígitos ('50' → '050', '1000' → '000')."""
    if valor is None:
        return "000"
    s = str(valor).strip()
    # Remove parte decimal se presente
    s = s.split(".")[0].split(",")[0]
    # Mantém só dígitos
    s = re.sub(r"\D", "", s)
    if not s:
        return "000"
    # Trunca para os últimos 3 dígitos (metro 1000 → 000)
    return s[-3:].zfill(3)


def km_mais_metros(km, metros) -> str:
    """'50 + 950' a partir de km=50 e metros='950'."""
    try:
        km_s = str(int(float(str(km).replace(",", "."))))
    except Exception:
        km_s = str(km)
    met_s = pad_metros(metros)
    return f"{km_s} + {met_s}"


def km_virgula_metros(km, metros) -> str:
    """'50,950'."""
    try:
        km_s = str(int(float(str(km).replace(",", "."))))
    except Exception:
        km_s = str(km)
    met_s = pad_metros(metros)
    return f"{km_s},{met_s}"


def km_formato_arquivo(km, metros=None) -> str:
    """
    Retorna KM no formato '50+950' (sem espaços) para nomes de arquivo.
    Aceita:
      - km_formato_arquivo(50, 950)         → '50+950'
      - km_formato_arquivo('50 + 950')      → '50+950'
      - km_formato_arquivo('50+950')        → '50+950'
      - km_formato_arquivo(50.950)          → '50+950'
    """
    if metros is not None:
        return km_mais_metros(km, metros).replace(" ", "")
    # Argumento único: pode ser string formatada ou float
    s = str(km).strip()
    # Se já tem '+' ou ',', limpa espaços
    if '+' in s or ',' in s:
        return s.replace(" ", "").replace(",", "+")
    # Float → converte
    try:
        v = float(s.replace(",", "."))
        km_int = int(v)
        met = round((v - km_int) * 1000)
        return f"{km_int}+{met:03d}"
    except Exception:
        return s.replace(" ", "")


def formatar_numero(n, largura_ou_decimais: int = 3) -> str:
    """
    Formata número. Quando usado no pipeline NC ARTESP, formata como zero-padded int
    (ex: formatar_numero(1, 6) → '000001').
    Com valor grande (>= 1000) ou float, formata com casas decimais.
    """
    try:
        v = float(n)
        # Se o argumento é <= 9 provavelmente é largura (padrão do pipeline)
        if largura_ou_decimais <= 9 and v == int(v):
            return str(int(v)).zfill(largura_ou_decimais)
        return f"{v:.{largura_ou_decimais}f}"
    except Exception:
        return str(n)


# NOMES DE ARQUIVO E PASTAS

_CHARS_INVALIDOS = re.compile(r'[\\/:*?"<>|\x00-\x1f]')


def sanitizar_nome(s: str, max_len: int = 200) -> str:
    """Remove caracteres inválidos para nome de arquivo Windows; «_» na saída torna-se espaço."""
    if not s or not isinstance(s, str):
        return ""
    s = _CHARS_INVALIDOS.sub(" ", s)
    s = s.replace("_", " ")
    s = re.sub(r" +", " ", s).strip(". ")
    return s[:max_len]


_M01_EXPORTAR_STEM_RE = re.compile(
    r"^(?P<head>\d{8} - .+? \([^)]+ - )(?P<serv>.+?)(?P<tail>\) - Prazo - \d{1,2}-\d{1,2}-\d{4})",
    re.IGNORECASE,
)


def truncar_nome_preservando_sufixo_prazo_m01(nome: str, max_chars: int) -> str:
    """
    Encurta o nome do ficheiro (com extensão) para no máximo ``max_chars`` caracteres,
    preservando o sufixo Art_011 / M01 `` - Prazo - dd-mm-aaaa`` antes da extensão quando existir.
    Nomes ``yyyymmdd - CONSTATAÇÕES NC … (rodovia - serviço) - Prazo - data``: encurta só ``serviço``.
    """
    nome = (nome or "").strip()
    if not nome or len(nome) <= max_chars:
        return nome
    ext = Path(nome).suffix
    stem = Path(nome).stem
    room = max(8, max_chars - len(ext))
    if len(stem) <= room:
        return nome
    m01 = _M01_EXPORTAR_STEM_RE.match(stem)
    if m01:
        head, serv, tail = m01.group("head"), m01.group("serv"), m01.group("tail")
        need = len(head) + len(serv) + len(tail)
        if need <= room:
            return nome
        budget_serv = room - len(head) - len(tail)
        if budget_serv >= 4:
            serv_t = serv[:budget_serv].rstrip(" -,.")
            if serv_t:
                return head + serv_t + tail + ext
        if len(tail) + len(ext) <= max_chars:
            return (head.rstrip() + tail + ext)[:max_chars]
    tail = ""
    m = re.search(r"( - Prazo - \d{1,2}-\d{1,2}-\d{4})$", stem)
    if m:
        tail = m.group(1)
    else:
        m = re.search(r"( - Prazo - .+)$", stem)
        if m:
            tail = m.group(1)
    if tail and len(tail) <= room:
        head_budget = room - len(tail)
        if head_budget > 0:
            head = stem[:head_budget].rstrip(" -")
            if not head:
                head = stem[:head_budget]
        else:
            head = ""
        return head + tail + ext
    if tail and len(tail) + len(ext) <= max_chars:
        return tail.strip() + ext
    return stem[:room].rstrip(" -.") + ext


def resolver_path_ficheiro_ci(path: Path | str) -> Path:
    """
    Resolve o nome real no disco ignorando maiúsculas/minúsculas no **último** componente.

    No Linux (ex.: Render), ``Acumulado.xlsx`` e ``Acumulado.XLSX`` são ficheiros distintos; se o Git
    tiver ``Acumulado.XLSX`` mas o código construir ``Acumulado.xlsx``, ``Path.is_file()`` falha.
    Esta função, quando o pai existe, procura no diretório um ficheiro com o mesmo nome por ``casefold()``.
    """
    p = Path(path)
    if not p.name:
        return p
    try:
        if p.is_file():
            return p
    except OSError:
        return p
    parent = p.parent
    try:
        if not parent.is_dir():
            return p
    except OSError:
        return p
    alvo = p.name.casefold()
    try:
        for c in parent.iterdir():
            if c.is_file() and c.name.casefold() == alvo:
                return c
    except OSError:
        pass
    return p


def str_caminho_io_windows(caminho) -> str:
    """
    Caminho absoluto para ``open()``, ``shutil``, ``ZipFile.extractall``, openpyxl, etc. no Windows.

    Prefixa **sempre** ``\\\\?\\`` (ou ``\\\\?\\UNC\\`` em partilhas de rede) no caminho absoluto
    resolvido, permitindo até ~32 767 caracteres sem «LongPathsEnabled» e sem falhas perto do
    limite clássico de 260. Em outros SO devolve ``str(Path.resolve())``.
    """
    if os.name != "nt":
        p = Path(caminho)
        try:
            return str(p.resolve(strict=False))
        except (OSError, RuntimeError):
            return str(p)
    p = Path(caminho)
    try:
        abs_s = str(p.resolve(strict=False))
    except (OSError, RuntimeError):
        abs_s = str(p if p.is_absolute() else Path.cwd() / p)
    abs_s = os.path.normpath(abs_s)
    if abs_s.startswith("\\\\?\\"):
        return abs_s
    if abs_s.startswith("\\\\"):
        return "\\\\?\\UNC\\" + abs_s[2:].lstrip("\\")
    return "\\\\?\\" + abs_s


def str_caminho_outlook_mapi(caminho) -> str:
    """
    Caminho para ``Outlook.Application`` / ``Attachments.Add`` e outras APIs MAPI/COM.

    O Outlook costuma **falhar** com o prefixo ``\\\\?\\``; usa-se caminho clássico (sem prefixo)
    quando o comprimento absoluto ≤ 259. Acima disso volta a ``str_caminho_io_windows`` (pode ainda
    falhar no COM — nesse caso copiar o ficheiro para pasta curta).
    """
    if os.name != "nt":
        p = Path(caminho)
        try:
            return str(p.resolve(strict=False))
        except (OSError, RuntimeError):
            return str(p)
    p = Path(caminho)
    try:
        s = str(p.resolve(strict=False))
    except (OSError, RuntimeError):
        s = str(p if p.is_absolute() else Path.cwd() / p)
    s = os.path.normpath(s)
    if s.startswith("\\\\?\\"):
        return s
    if len(s) <= 259:
        return s
    return str_caminho_io_windows(p)


def extrair_zipfile_para_pasta(zf, destino) -> None:
    """
    ``ZipFile.extractall`` com pasta de destino criada via caminho estendido no Windows
    (``\\\\?\\``), para extrações profundas no servidor/pipeline.
    """
    import zipfile as _zipfile

    if not isinstance(zf, _zipfile.ZipFile):
        raise TypeError("extrair_zipfile_para_pasta espera zipfile.ZipFile")
    p = Path(destino)
    if os.name != "nt":
        p.mkdir(parents=True, exist_ok=True)
        zf.extractall(str(p))
        return
    dest_s = str_caminho_io_windows(p)
    os.makedirs(dest_s, exist_ok=True)
    zf.extractall(dest_s)


def garantir_pasta(caminho) -> Path:
    """Cria o diretório se não existir. Retorna Path. No Windows usa caminho longo se preciso."""
    p = Path(caminho)
    if os.name == "nt":
        os.makedirs(str_caminho_io_windows(p), exist_ok=True)
        return p
    p.mkdir(parents=True, exist_ok=True)
    return p


def escrever_bytes_caminho(caminho, data: bytes) -> Path:
    """Grava bytes no ficheiro; cria pastas e usa caminho longo no Windows quando necessário."""
    p = Path(caminho)
    garantir_pasta(p.parent)
    with open(str_caminho_io_windows(p), "wb") as f:
        f.write(data)
    return p


def caminho_dentro_limite_windows(caminho, max_len: int = 248) -> Path:
    """
    Caminho total até ``max_len`` caracteres (trunca o stem ou usa hash curto no nome).
    Omissão 248: margem abaixo de 260 para Explorador / cópia Shell (0x80010135 «caminho muito longo»).
    """
    import hashlib

    p = Path(caminho)
    parent_s = str(p.parent)
    ext = p.suffix or ""
    base = p.stem
    sep = 1

    def _total(stem: str) -> int:
        return len(parent_s) + sep + len(stem) + len(ext)

    if len(str(p)) <= max_len:
        return p
    room = max_len - len(parent_s) - sep - len(ext)
    if room < 1:
        digest = hashlib.sha1(str(p).encode("utf-8", errors="replace")).hexdigest()
        max_digest = max_len - len(parent_s) - sep - len(ext)
        max_digest = min(32, max(0, max_digest))
        while max_digest >= 1:
            cand = p.parent / f"{digest[:max_digest]}{ext}"
            if len(str(cand)) <= max_len:
                return cand
            max_digest -= 1
        stub = (base[:1] if base else "x") + ext
        cand = p.parent / stub
        if len(str(cand)) <= max_len:
            return cand
        return p
    return p.parent / (base[:room] + ext)


def encurtar_nome_em_pasta(pasta: Path, nome: str, max_path: int = 259) -> Path:
    """
    Retorna Path(pasta/nome) encurtando o nome se o caminho total exceder max_path.
    Preserva o sufixo M01 « - Prazo - dd-mm-aaaa» antes de .xlsx (ver ``truncar_nome_preservando_sufixo_prazo_m01``).
    """
    destino = pasta / nome
    if len(str(destino)) <= max_path:
        return destino
    pasta_s = str(pasta)
    max_nome = max_path - len(pasta_s) - 1
    if max_nome < 12:
        max_nome = 12
    nome_curto = truncar_nome_preservando_sufixo_prazo_m01(nome, max_nome)
    return pasta / nome_curto


def copiar_arquivo(src, dst, sobrescrever: bool = True) -> Path:
    """Copia arquivo src → dst. Cria pasta de destino se necessário."""
    src = Path(src);  dst = Path(dst)
    if not sobrescrever and dst.exists():
        return dst
    garantir_pasta(dst.parent)
    shutil.copy2(str_caminho_io_windows(src), str_caminho_io_windows(dst))
    return dst


def renomear_arquivo(src, dst) -> Path:
    """Renomeia/move src → dst. Se o destino já existir (reprocessamento), remove ou substitui.
    No Windows/OneDrive o destino pode persistir após unlink; usa os.replace como fallback."""
    src = Path(src)
    dst = Path(dst)
    garantir_pasta(dst.parent)
    if dst.exists():
        try:
            dst.unlink()
        except OSError:
            pass
    try:
        src.rename(dst)
    except FileExistsError:
        # WinError 183 / OneDrive: destino ainda existe; substituir explicitamente
        os.replace(str_caminho_io_windows(src), str_caminho_io_windows(dst))
    return dst


# RODOVIAS E MAPA EAF (grupos por trecho km — Contatos EAFs)

def normalizar_rodovia_para_busca(rodovia: str) -> str:
    """
    Normaliza o nome da rodovia para comparação com MAPA_EAF e RODOVIAS.
    Aceita: "SP 075", "SP075", "SP-075", "SP 127", "SPI 102-300", "SPI102/300", etc.
    SPI102/300 pertence à Autoroutes (trecho SPI 102-300 no MAPA_EAF).
    """
    s = str(rodovia or "").strip().upper()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("-", " ").replace("_", " ").replace("/", " ")
    s = re.sub(r"\bSP(\d)", r"SP \1", s)
    s = re.sub(r"\bSPI(\d)", r"SPI \1", s)
    return s


def obter_grupo_empresa_por_trecho(rodovia: str, km: float, mapa_eaf: list) -> tuple[int, str]:
    """
    Retorna (grupo, empresa) para uma NC com base em rodovia + km, usando MAPA_EAF.
    mapa_eaf: list[dict] com keys grupo, empresa, trechos (lista de {rodovia, km_ini, km_fim}).
    Retorna (0, "") se não houver trecho correspondente.
    """
    if not mapa_eaf or km is None:
        return 0, ""
    rod_nc = normalizar_rodovia_para_busca(rodovia)
    if not rod_nc:
        return 0, ""
    # Tolerancia mínima para contornar erros de arredondamento em floats
    # (ex.: 43.000 pode chegar como 42.999999).
    # Em PDFs, o km frequentemente vem com arredondamento/representação float.
    # Usar eps maior para não "perder" a borda exata (ex.: 43.000 lido como 42.999).
    eps = 1e-3

    def _rodovias_equivalentes(rod_trecho: str, rod_nc: str) -> bool:
        """
        Considera equivalentes rodovias como:
        - "SP 075" == "SP 75" (zeros à esquerda no número da SP)
        - "SPI 102-300" / "SPI 102 300" / "SPI 102300" (mesma rodovia)
        """
        a = normalizar_rodovia_para_busca(rod_trecho or "")
        b = normalizar_rodovia_para_busca(rod_nc or "")
        if not a or not b:
            return False
        if a == b:
            return True
        m1 = re.match(r"^(SP)\s*(\d+)$", a)
        m2 = re.match(r"^(SP)\s*(\d+)$", b)
        if m1 and m2 and m1.group(1) == m2.group(1):
            try:
                return int(m1.group(2)) == int(m2.group(2))
            except Exception:
                return False
        # SPI: comparar por sequência de dígitos (ex.: "SPI 102 300" e "SPI 102300")
        if a.startswith("SPI") and b.startswith("SPI"):
            dig_a = "".join(re.findall(r"\d+", a))
            dig_b = "".join(re.findall(r"\d+", b))
            return dig_a == dig_b
        if a.startswith("MG") and b.startswith("MG"):
            dig_a = "".join(re.findall(r"\d+", a))
            dig_b = "".join(re.findall(r"\d+", b))
            return bool(dig_a and dig_b and dig_a == dig_b)
        if a.startswith("BR") and b.startswith("BR"):
            dig_a = "".join(re.findall(r"\d+", a))
            dig_b = "".join(re.findall(r"\d+", b))
            return bool(dig_a and dig_b and dig_a == dig_b)
        return False
    # Pode haver múltiplas EAFs na mesma rodovia (e até trechos próximos).
    # Então, em vez de retornar a "primeira" correspondência, coletamos todas
    # as que contêm o km e escolhemos a mais específica.
    candidatos: list[tuple[int, str, float]] = []  # (grupo, empresa, km_ini do trecho)

    for entry in mapa_eaf:
        for trecho in entry.get("trechos", []):
            rod_t = normalizar_rodovia_para_busca(trecho.get("rodovia", ""))
            if not rod_t:
                continue
            # Rodovia deve casar por equivalência (ex.: SP 075 vs SP 75).
            if _rodovias_equivalentes(rod_t, rod_nc):
                ki = trecho.get("km_ini", 0.0)
                kf = trecho.get("km_fim", 9999.0)
                if (ki - eps) <= km <= (kf + eps):
                    candidatos.append((entry.get("grupo", 0), entry.get("empresa", ""), float(ki)))

    if not candidatos:
        return 0, ""

    # Mais específica = maior km_ini (trecho mais "tarde" na mesma rodovia).
    candidatos.sort(key=lambda x: x[2], reverse=True)
    return candidatos[0][0], candidatos[0][1]


def normalizar_rodovia_eaf(rodovia_raw: str, rodovias: dict) -> dict:
    """
    Normaliza o nome de rodovia da EAF buscando em `rodovias` (config.RODOVIAS).
    Retorna dict com keys 'tag', 'nome', 'sentidos', 'codigo', 'n'.
    Se não encontrar, retorna tag='FORA' com o raw como nome.
    """
    raw = str(rodovia_raw or "").strip()

    def _completar(info: dict, chave: str) -> dict:
        tag = info.get("tag", chave)
        # codigo: forma exibição (ex. SP-075); se chave é "SP 075" -> "SP-075"
        codigo = info.get("codigo") or chave.replace(" ", "-") if chave else tag
        n = info.get("n", 0)
        return {**info, "tag": tag, "codigo": codigo, "n": n}

    # Busca exata
    if raw in rodovias:
        return _completar(rodovias[raw].copy(), raw)
    # Busca por prefixo (primeiros 6 chars)
    prefixo = raw[:6]
    for chave, info in rodovias.items():
        if chave.startswith(prefixo) or prefixo.startswith(chave[:6]):
            return _completar(info.copy(), chave)
    # Busca case-insensitive
    raw_up = raw.upper()
    for chave, info in rodovias.items():
        if chave.upper() in raw_up or raw_up in chave.upper():
            return _completar(info.copy(), chave)
    return {"tag": "FORA", "nome": raw or "Desconhecida", "sentidos": [], "codigo": raw or "FORA", "n": 0}


# CAMINHOS DE FOTOS (usados por gerar_modelo_foto e inserir_nc_kria)

def path_foto_nc(pasta_nc, numero: "int | str") -> Path:
    """Retorna Path para 'nc (N).jpg' na pasta. N = número ou código (ex: HE.13.0111 para MA)."""
    return Path(pasta_nc) / f"nc ({numero}).jpg"


def path_foto_pdf(pasta_pdf, numero: "int | str") -> Path:
    """Retorna Path para 'PDF (N).jpg' (subpasta PDF/ se existir)."""
    p = Path(pasta_pdf)
    sub = p / "PDF"
    if sub.is_dir():
        return sub / f"PDF ({numero}).jpg"
    return p / f"PDF ({numero}).jpg"


def _pastas_busca_foto_extracao(pasta: "Path", prefixo: str) -> list:
    """Pastas candidatas para fotos extraídas (raiz, subpastas e ZIP com pasta-base)."""
    pasta = Path(pasta)
    sub = "PDF" if (prefixo or "").strip().upper() == "PDF" else "nc"
    out: list[Path] = []

    def _add(p: Path) -> None:
        if p.is_dir() and p not in out:
            out.append(p)

    if pasta.is_dir():
        # Estrutura direta (legado e alguns fluxos): raiz + raiz/sub
        _add(pasta / sub)
        _add(pasta)

        # Estrutura comum do ZIP web: pasta/lote_.../{arquivos} e/ou pasta/lote_.../sub
        try:
            for child in sorted(pasta.iterdir()):
                if not child.is_dir():
                    continue
                _add(child / sub)
                _add(child)
        except OSError:
            pass

    return out or [pasta]


_FOTO_INDEX_CACHE: dict[tuple[str, str], tuple[dict[str, Path], dict[str, Path]]] = {}
_FOTO_INDEX_RECURSIVO_CACHE: dict[tuple[str, str], tuple[dict[str, Path], dict[str, Path]]] = {}


def limpar_cache_indices_foto() -> None:
    """Limpa cache global de índices de imagens (evita paths stale após purge/extract)."""
    _FOTO_INDEX_CACHE.clear()
    _FOTO_INDEX_RECURSIVO_CACHE.clear()


def _indexar_fotos_base(base: Path, prefixo: str) -> tuple[dict[str, Path], dict[str, Path]]:
    """
    Indexa 1x os JPG de uma pasta:
      - exato: nome lower -> Path
      - mid: valor dentro de '(...)' -> Path (case-insensitive)
    """
    key = (str(base), (prefixo or "").strip().lower())
    cached = _FOTO_INDEX_CACHE.get(key)
    if cached is not None:
        return cached

    exato: dict[str, Path] = {}
    mid: dict[str, Path] = {}
    pref_l = (prefixo or "").strip().lower()
    start_l = f"{pref_l} ("

    try:
        for f in base.iterdir():
            if not f.is_file():
                continue
            name = f.name
            low = name.lower()
            if not low.endswith(".jpg"):
                continue
            exato[low] = f
            if not low.startswith(start_l):
                continue
            rest = name[len(prefixo) + 2 :]  # após "PREFIXO ("
            if ")" not in rest:
                continue
            mid_raw = rest.split(")", 1)[0].strip()
            if not mid_raw:
                continue
            mid_l = mid_raw.lower()
            # Primeiro encontrado vence para manter determinismo.
            if mid_l not in mid:
                mid[mid_l] = f
            if "_" in mid_l:
                base_mid = mid_l.split("_", 1)[0].strip()
                if base_mid and base_mid not in mid:
                    mid[base_mid] = f
    except OSError:
        pass

    _FOTO_INDEX_CACHE[key] = (exato, mid)
    return exato, mid


def _indexar_fotos_recursivo(pasta: Path, prefixo: str) -> tuple[dict[str, Path], dict[str, Path]]:
    """Indexa JPG recursivamente (fallback robusto para ZIPs com pastas profundas)."""
    key = (str(pasta), (prefixo or "").strip().lower())
    cached = _FOTO_INDEX_RECURSIVO_CACHE.get(key)
    if cached is not None:
        return cached

    exato: dict[str, Path] = {}
    mid: dict[str, Path] = {}
    pref_l = (prefixo or "").strip().lower()
    start_l = f"{pref_l} ("

    try:
        for f in pasta.rglob("*"):
            if not f.is_file():
                continue
            name = f.name
            low = name.lower()
            if not low.endswith(".jpg"):
                continue
            if low not in exato:
                exato[low] = f
            if not low.startswith(start_l):
                continue
            rest = name[len(prefixo) + 2 :]
            if ")" not in rest:
                continue
            mid_raw = rest.split(")", 1)[0].strip()
            if not mid_raw:
                continue
            mid_l = mid_raw.lower()
            if mid_l not in mid:
                mid[mid_l] = f
            if "_" in mid_l:
                base_mid = mid_l.split("_", 1)[0].strip()
                if base_mid and base_mid not in mid:
                    mid[base_mid] = f
    except OSError:
        pass

    _FOTO_INDEX_RECURSIVO_CACHE[key] = (exato, mid)
    return exato, mid


def encontrar_foto_por_codigo_ou_numero(
    pasta: "Path",
    prefixo: str,
    codigo: str | int | None = None,
    numero: int | None = None,
) -> "Path | None":
    """
    Encontra arquivo de foto por código ou número.
    prefixo: "nc" ou "PDF". Procura em pasta/nc/ e pasta/PDF/ (ZIP extração) e na raiz.
    """
    pasta = Path(pasta)
    if not pasta.is_dir():
        return None
    prefix = f"{prefixo} ("
    suffix = ").jpg"
    bases = _pastas_busca_foto_extracao(pasta, prefixo)

    def _buscar_por_codigo_str(cod: str) -> "Path | None":
        cod = (cod or "").strip()
        if not cod:
            return None
        cod_l = cod.lower()
        for base in bases:
            exato, mid = _indexar_fotos_base(base, prefixo)
            hit = exato.get(f"{prefix}{cod}{suffix}".lower())
            if hit is not None:
                return hit
            hit = mid.get(cod_l)
            if hit is not None:
                return hit
        exato_r, mid_r = _indexar_fotos_recursivo(pasta, prefixo)
        hit = exato_r.get(f"{prefix}{cod}{suffix}".lower())
        if hit is not None:
            return hit
        hit = mid_r.get(cod_l)
        if hit is not None:
            return hit
        return None

    # Código alfanumérico (ex: HE.13.0111)
    if codigo is not None and isinstance(codigo, str) and codigo.strip():
        try:
            int(float(codigo.strip()))
        except (ValueError, TypeError):
            r = _buscar_por_codigo_str(codigo)
            if r is not None:
                return r

    for valor in (codigo, numero):
        if valor is None:
            continue
        try:
            n = int(float(str(valor).strip()))
        except (ValueError, TypeError):
            continue
        for cod in (str(n), str(n).zfill(5)):
            cod_l = cod.lower()
            for base in bases:
                exato, mid = _indexar_fotos_base(base, prefixo)
                hit = exato.get(f"{prefix}{cod}{suffix}".lower())
                if hit is not None:
                    return hit
                hit = mid.get(cod_l)
                if hit is not None:
                    return hit
            exato_r, mid_r = _indexar_fotos_recursivo(pasta, prefixo)
            hit = exato_r.get(f"{prefix}{cod}{suffix}".lower())
            if hit is not None:
                return hit
            hit = mid_r.get(cod_l)
            if hit is not None:
                return hit
    return None


def _norm_header_celula_eaf(s: str) -> str:
    t = unicodedata.normalize("NFD", str(s or ""))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", t).strip().lower()


def detectar_coluna_tipo_de_atividade_eaf(ws, fallback: int = 15) -> int:
    """«Tipo de Atividade» na mãe EAF (tipicamente O), distinta da coluna «Atividade» (Q)."""
    melhor_c = None
    melhor_sc = -1
    max_r = min(int(ws.max_row or 0), 8)
    max_c = min(int(ws.max_column or 0), 40)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            h = _norm_header_celula_eaf(str(v))
            sc = 0
            if h == "tipo de atividade":
                sc = 10
            elif h == "tipo atividade":
                sc = 9
            elif "tipo" in h and "atividade" in h and "grupo" not in h:
                if "tipo de" in h or h.startswith("tipo "):
                    sc = 8
            if sc > melhor_sc:
                melhor_sc, melhor_c = sc, c
    return melhor_c if melhor_c is not None and melhor_sc >= 8 else fallback


EXPORTAR_KARTADO_MAE_SUBDIR = "exportar"

_REL_NS_PKG = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}
_REL_T_PRINTER = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/printerSettings"
)

_OOXML_XML_DECL = b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'


def _ooxml_bytes_com_declaracao_padrao(corpo_sem_decl: bytes) -> bytes:
    corpo = corpo_sem_decl.lstrip(b"\xef\xbb\xbf").lstrip()
    return _OOXML_XML_DECL + corpo


def _ooxml_normalizar_declaracao_xml(part: bytes) -> bytes:
    part = part.lstrip(b"\xef\xbb\xbf")
    if not part.startswith(b"<?xml"):
        return _ooxml_bytes_com_declaracao_padrao(part)
    fim = part.find(b"?>")
    if fim == -1:
        return _ooxml_bytes_com_declaracao_padrao(part)
    resto = part[fim + 2 :].lstrip(b"\r\n")
    return _ooxml_bytes_com_declaracao_padrao(resto)


def _norm_part_from_rels(rels_member: str, target: str) -> str:
    if not target or target.startswith("http"):
        return ""
    base = rels_member.split("/_rels/", 1)[0] if "/_rels/" in rels_member else ""
    if not base:
        return ""
    return posixpath.normpath(f"{base}/{target}").replace("\\", "/")


def _rels_targets_normalized(rels_member: str, rels_xml: bytes) -> set[str]:
    try:
        root = ET.fromstring(rels_xml)
    except ET.ParseError:
        return set()
    out: set[str] = set()
    for el in root.findall("rel:Relationship", _REL_NS_PKG):
        tgt = el.get("Target")
        if not tgt:
            continue
        n = _norm_part_from_rels(rels_member, tgt)
        if n:
            out.add(n)
    return out


def _rels_max_rid_index(rels_xml: bytes) -> int:
    try:
        root = ET.fromstring(rels_xml)
    except ET.ParseError:
        return 0
    m = 0
    for el in root.findall("rel:Relationship", _REL_NS_PKG):
        rid = el.get("Id") or ""
        mm = re.match(r"rId(\d+)$", rid, re.I)
        if mm:
            m = max(m, int(mm.group(1)))
    return m


def _merge_worksheet_rels_complement_printer_settings(
    rels_member: str, tpl_rels: bytes, out_rels: bytes
) -> bytes | None:
    """
    Se a saída do openpyxl só referencia o drawing mas o template tinha impressora,
    acrescenta o Relationship de ``printerSettings`` com novo ``rId`` e o mesmo ``Target``.
    Inserção em texto para não re-serializar o .rels com prefixos ``ns0:`` (Excel rejeita).
    """
    try:
        t_root = ET.fromstring(tpl_rels)
    except ET.ParseError:
        return None
    have = _rels_targets_normalized(rels_member, out_rels)
    next_n = _rels_max_rid_index(out_rels)
    to_add: list[tuple[str, str]] = []
    for el in t_root.findall("rel:Relationship", _REL_NS_PKG):
        typ = el.get("Type") or ""
        if typ != _REL_T_PRINTER:
            continue
        tgt = el.get("Target")
        if not tgt:
            continue
        norm = _norm_part_from_rels(rels_member, tgt)
        if not norm or norm in have:
            continue
        next_n += 1
        to_add.append((f"rId{next_n}", tgt))
        have.add(norm)
    if not to_add:
        return None
    try:
        s = out_rels.decode("utf-8")
    except UnicodeDecodeError:
        return None
    idx = s.rfind("</Relationships>")
    if idx == -1:
        return None
    frag = "".join(
        f'<Relationship Id="{nid}" Type="{_REL_T_PRINTER}" Target="{tgt}"/>'
        for nid, tgt in to_add
    )
    return (s[:idx] + frag + s[idx:]).encode("utf-8")


def _rels_relationship_type_for_id(rels_xml: bytes | None, rid: str) -> str | None:
    if not rels_xml or not rid:
        return None
    try:
        root = ET.fromstring(rels_xml)
    except ET.ParseError:
        return None
    for el in root.findall("rel:Relationship", _REL_NS_PKG):
        if el.get("Id") == rid:
            return el.get("Type") or ""
    return None


def _primeira_rid_printer_em_rels(rels_xml: bytes | None) -> str | None:
    if not rels_xml:
        return None
    try:
        root = ET.fromstring(rels_xml)
    except ET.ParseError:
        return None
    for el in root.findall("rel:Relationship", _REL_NS_PKG):
        typ = el.get("Type") or ""
        if typ == _REL_T_PRINTER:
            rid = el.get("Id")
            if rid:
                return rid
    return None


def _fundir_desenhos_drawing_xml_no_mapa(tpl_bytes: bytes, out_map: dict[str, bytes]) -> None:
    """
    O openpyxl reescreve ``xl/drawings/drawingN.xml`` só com a foto; o template pode ter formas
    (caixa «Data Execução», etc.) no mesmo ficheiro. Junta âncoras sem ``pic`` do template com
    âncoras com ``pic`` da saída.
    """
    SD_URI = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    A_URI = "http://schemas.openxmlformats.org/drawingml/2006/main"
    R_URI = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    A16_URI = "http://schemas.microsoft.com/office/drawing/2014/main"
    ET.register_namespace("xdr", SD_URI)
    ET.register_namespace("a", A_URI)
    ET.register_namespace("r", R_URI)
    ET.register_namespace("a16", A16_URI)

    SD = f"{{{SD_URI}}}"
    _ANCHORS = frozenset(
        {"twoCellAnchor", "oneCellAnchor", "absoluteAnchor", "absoluteCellAnchor"}
    )

    def _local(tag: str) -> str:
        return tag.split("}")[-1] if "}" in tag else tag

    def _is_anchor(el: ET.Element) -> bool:
        return _local(el.tag) in _ANCHORS

    def _anchor_has_pic(anchor: ET.Element) -> bool:
        return anchor.find(f"{SD}pic") is not None

    def _clone(el: ET.Element) -> ET.Element:
        return ET.fromstring(ET.tostring(el, encoding="utf-8"))

    with zipfile.ZipFile(BytesIO(tpl_bytes), "r") as zt:
        for name in sorted(zt.namelist()):
            if not re.match(r"^xl/drawings/drawing\d+\.xml$", name):
                continue
            if name not in out_map:
                continue
            try:
                tpl_xml = zt.read(name)
                out_xml = out_map[name]
                t_root = ET.fromstring(tpl_xml)
                o_root = ET.fromstring(out_xml)
            except (ET.ParseError, KeyError):
                continue
            if _local(t_root.tag) != "wsDr" or _local(o_root.tag) != "wsDr":
                continue
            from_tpl = [
                _clone(ch)
                for ch in t_root
                if _is_anchor(ch) and not _anchor_has_pic(ch)
            ]
            from_out = [
                _clone(ch)
                for ch in o_root
                if _is_anchor(ch) and _anchor_has_pic(ch)
            ]
            if not from_tpl or not from_out:
                continue
            merged = ET.Element(f"{SD}wsDr")
            for ch in from_tpl:
                merged.append(ch)
            for ch in from_out:
                merged.append(ch)
            out_map[name] = _ooxml_normalizar_declaracao_xml(
                ET.tostring(merged, encoding="utf-8", xml_declaration=False)
            )


def preservar_ooxml_planilha_pos_openpyxl(template: Path | bytes, output_path: Path) -> None:
    """
    O openpyxl descarta ao gravar várias partes OOXML (desenhos/formas, impressora, rels da folha).
    Copia do template as partes em falta no ficheiro gerado, repõe referências na folha e **funde**
    ``drawing*.xml`` quando a saída só contém a foto e o template tinha formas no mesmo ficheiro.
    """
    out_path = Path(output_path)
    if not out_path.is_file():
        return
    if isinstance(template, Path):
        tpl_path = template
        if not tpl_path.is_file():
            return
        tpl_bytes = tpl_path.read_bytes()
        tpl_log = str(tpl_path)
    else:
        tpl_bytes = bytes(template)
        tpl_log = "(template em bytes)"

    REL_NS = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}
    CT_NS = {"ct": "http://schemas.openxmlformats.org/package/2006/content-types"}

    def _base_dir_for_rels(rels_member: str) -> str:
        if "/_rels/" not in rels_member:
            return ""
        return rels_member.split("/_rels/", 1)[0]

    def _resolve_target(rels_member: str, target: str) -> str:
        if not target or target.startswith("http"):
            return ""
        base = _base_dir_for_rels(rels_member)
        if not base:
            return ""
        return posixpath.normpath(f"{base}/{target}").replace("\\", "/")

    def _targets_from_rels(rels_member: str, rels_xml: bytes) -> list[str]:
        try:
            root = ET.fromstring(rels_xml)
        except ET.ParseError:
            return []
        out: list[str] = []
        for el in root.findall("rel:Relationship", REL_NS):
            tgt = el.get("Target")
            if not tgt:
                continue
            r = _resolve_target(rels_member, tgt)
            if r:
                out.append(r)
        return out

    def _merge_sheet_xml_ooxml_refs(
        out_sheet: str, sheet_rels_out: bytes | None
    ) -> str:
        if not sheet_rels_out:
            return out_sheet
        page_setup_ganhou_rid = False
        m_ps = re.search(r"<pageSetup[^>]*/>", out_sheet)
        if m_ps:
            tag = m_ps.group(0)
            mrid = re.search(r'\sr:id="(rId\d+)"', tag)
            if mrid:
                rid = mrid.group(1)
                tipo = _rels_relationship_type_for_id(sheet_rels_out, rid) or ""
                if "printerSettings" not in tipo:
                    new_tag = re.sub(r'\s+r:id="rId\d+"', "", tag)
                    if new_tag != tag:
                        out_sheet = out_sheet.replace(tag, new_tag, 1)
        m_ps2 = re.search(r"<pageSetup[^>]*/>", out_sheet)
        if m_ps2 and "r:id=" not in m_ps2.group(0):
            pid = _primeira_rid_printer_em_rels(sheet_rels_out)
            if pid:
                inner = m_ps2.group(0)
                idx = inner.rfind("/>")
                if idx != -1:
                    new_el = inner[:idx] + f' r:id="{pid}"' + inner[idx:]
                    out_sheet = out_sheet.replace(inner, new_el, 1)
                    page_setup_ganhou_rid = True
        if page_setup_ganhou_rid:
            mws = re.search(r"<worksheet[^>]*>", out_sheet)
            if mws and "xmlns:r=" not in mws.group(0):
                old = mws.group(0)
                new = (
                    old[:-1]
                    + ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
                )
                out_sheet = out_sheet.replace(old, new, 1)
        return out_sheet

    try:
        out_bytes = out_path.read_bytes()
    except OSError:
        return

    try:
        with zipfile.ZipFile(BytesIO(tpl_bytes), "r") as zt, zipfile.ZipFile(BytesIO(out_bytes), "r") as zo:
            tpl_names = set(zt.namelist())
            out_map: dict[str, bytes] = {n: zo.read(n) for n in zo.namelist()}
        added: set[str] = set()
        queue: list[str] = []
        seen_q: set[str] = set()

        with zipfile.ZipFile(BytesIO(tpl_bytes), "r") as zt:
            for name in sorted(tpl_names):
                if not (
                    name.startswith("xl/worksheets/_rels/sheet")
                    and name.endswith(".xml.rels")
                ):
                    continue
                if name not in out_map:
                    rel_data = zt.read(name)
                    out_map[name] = rel_data
                    added.add(name)
                    for t in _targets_from_rels(name, rel_data):
                        if t not in seen_q:
                            seen_q.add(t)
                            queue.append(t)
                else:
                    try:
                        merged_rs = _merge_worksheet_rels_complement_printer_settings(
                            name, zt.read(name), out_map[name]
                        )
                    except KeyError:
                        merged_rs = None
                    if merged_rs is not None:
                        out_map[name] = merged_rs
                        for t in _targets_from_rels(name, merged_rs):
                            if t not in seen_q:
                                seen_q.add(t)
                                queue.append(t)

            for name in sorted(tpl_names):
                if name in out_map:
                    continue
                if name.startswith(("xl/drawings/", "xl/charts/", "xl/printerSettings/")):
                    out_map[name] = zt.read(name)
                    added.add(name)
                    if name.endswith(".rels"):
                        for t in _targets_from_rels(name, out_map[name]):
                            if t not in seen_q:
                                seen_q.add(t)
                                queue.append(t)

            while queue:
                part = queue.pop(0)
                if part in out_map:
                    continue
                if part not in tpl_names:
                    continue
                try:
                    pdata = zt.read(part)
                except KeyError:
                    continue
                out_map[part] = pdata
                added.add(part)
                if part.endswith(".rels"):
                    for t in _targets_from_rels(part, pdata):
                        if t not in seen_q:
                            seen_q.add(t)
                            queue.append(t)

        ct_path = "[Content_Types].xml"
        if ct_path in out_map and ct_path in tpl_names:
            try:
                tpl_root = ET.fromstring(
                    zipfile.ZipFile(BytesIO(tpl_bytes)).read(ct_path)
                )
                out_root = ET.fromstring(out_map[ct_path])
                existing_pn = {
                    ov.get("PartName")
                    for ov in out_root.findall("ct:Override", CT_NS)
                    if ov.get("PartName")
                }
                ct_alterado = False
                existing_def_ext = {
                    (d.get("Extension") or "").lower()
                    for d in out_root.findall("ct:Default", CT_NS)
                    if d.get("Extension")
                }
                for de in tpl_root.findall("ct:Default", CT_NS):
                    ext = (de.get("Extension") or "").lower()
                    if not ext or ext in existing_def_ext:
                        continue
                    out_root.append(de)
                    existing_def_ext.add(ext)
                    ct_alterado = True
                for ov in tpl_root.findall("ct:Override", CT_NS):
                    pn = ov.get("PartName") or ""
                    norm = pn.lstrip("/")
                    if pn in existing_pn or norm not in added:
                        continue
                    out_root.append(ov)
                    existing_pn.add(pn)
                    ct_alterado = True
                if ct_alterado:
                    out_map[ct_path] = _ooxml_normalizar_declaracao_xml(
                        ET.tostring(out_root, encoding="utf-8", xml_declaration=False)
                    )
            except ET.ParseError:
                pass

        ct_path2 = "[Content_Types].xml"
        if ct_path2 in out_map:
            raw_ct = out_map[ct_path2].lstrip(b"\xef\xbb\xbf")
            if raw_ct.startswith(b"<Types") and not raw_ct.startswith(b"<?xml"):
                out_map[ct_path2] = _ooxml_bytes_com_declaracao_padrao(raw_ct)

        for _dn in list(out_map):
            if re.match(r"^xl/drawings/drawing\d+\.xml$", _dn):
                out_map[_dn] = _ooxml_normalizar_declaracao_xml(out_map[_dn])

        with zipfile.ZipFile(BytesIO(tpl_bytes), "r") as zt:
            for sn in sorted(
                n for n in tpl_names if re.match(r"^xl/worksheets/sheet\d+\.xml$", n)
            ):
                if sn not in out_map:
                    continue
                try:
                    o_xml = out_map[sn].decode("utf-8")
                except (UnicodeDecodeError, KeyError):
                    continue
                mnum = re.search(r"sheet(\d+)\.xml$", sn)
                rels_path = (
                    f"xl/worksheets/_rels/sheet{mnum.group(1)}.xml.rels"
                    if mnum
                    else ""
                )
                rels_out = out_map.get(rels_path) if rels_path else None
                merged = _merge_sheet_xml_ooxml_refs(o_xml, rels_out)
                if merged != o_xml:
                    out_map[sn] = merged.encode("utf-8")

        _fundir_desenhos_drawing_xml_no_mapa(tpl_bytes, out_map)

        buf = BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zw:
            for name in sorted(out_map):
                zw.writestr(name, out_map[name])
        out_path.write_bytes(buf.getvalue())
    except Exception:
        logging.getLogger(__name__).exception(
            "preservar_ooxml_planilha_pos_openpyxl: falha ao mesclar %s (template %s)",
            out_path,
            tpl_log,
        )


def replicar_ancoras_sem_foto_por_bloco_em_drawing(
    output_path: Path, bloco_linhas: int, repeticoes: int
) -> None:
    """
    Duplica âncoras de desenho sem ``pic`` (formas/textos do template) para blocos seguintes.
    Usado no Relatório de Resposta: o modelo tem shape(s) no 1.º bloco que o openpyxl não replica
    ao inserir linhas; aqui replicamos no OOXML final com deslocamento de linhas por bloco.
    """
    if repeticoes <= 0 or bloco_linhas <= 0:
        return
    out_path = Path(output_path)
    if not out_path.is_file():
        return
    try:
        out_bytes = out_path.read_bytes()
    except OSError:
        return
    SD_URI = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    A_URI = "http://schemas.openxmlformats.org/drawingml/2006/main"
    R_URI = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    A16_URI = "http://schemas.microsoft.com/office/drawing/2014/main"
    ET.register_namespace("xdr", SD_URI)
    ET.register_namespace("a", A_URI)
    ET.register_namespace("r", R_URI)
    ET.register_namespace("a16", A16_URI)
    ns = {"xdr": SD_URI}

    def _local(tag: str) -> str:
        return tag.split("}")[-1] if "}" in tag else tag

    def _is_anchor(el: ET.Element) -> bool:
        return _local(el.tag) in {
            "twoCellAnchor",
            "oneCellAnchor",
            "absoluteAnchor",
            "absoluteCellAnchor",
        }

    def _anchor_has_pic(anchor: ET.Element) -> bool:
        return anchor.find("xdr:pic", ns) is not None

    def _clone(el: ET.Element) -> ET.Element:
        return ET.fromstring(ET.tostring(el, encoding="utf-8"))

    def _shift_anchor_rows(anchor: ET.Element, delta: int) -> None:
        for pt in ("from", "to"):
            base = anchor.find(f"xdr:{pt}", ns)
            if base is None:
                continue
            row_el = base.find("xdr:row", ns)
            if row_el is None or not (row_el.text or "").strip():
                continue
            try:
                row_el.text = str(int(row_el.text) + delta)
            except (TypeError, ValueError):
                continue

    try:
        with zipfile.ZipFile(BytesIO(out_bytes), "r") as zo:
            out_map: dict[str, bytes] = {n: zo.read(n) for n in zo.namelist()}
        alterado = False
        for name in sorted(out_map):
            if not re.match(r"^xl/drawings/drawing\d+\.xml$", name):
                continue
            try:
                root = ET.fromstring(out_map[name])
            except ET.ParseError:
                continue
            if _local(root.tag) != "wsDr":
                continue
            anchors_base = [ch for ch in list(root) if _is_anchor(ch) and not _anchor_has_pic(ch)]
            if not anchors_base:
                continue
            max_id = 0
            for cnp in root.findall(".//xdr:cNvPr", ns):
                try:
                    max_id = max(max_id, int(cnp.get("id") or "0"))
                except (TypeError, ValueError):
                    continue
            for rep in range(1, repeticoes + 1):
                delta = rep * bloco_linhas
                for anc in anchors_base:
                    neo = _clone(anc)
                    _shift_anchor_rows(neo, delta)
                    cnp = neo.find(".//xdr:cNvPr", ns)
                    if cnp is not None:
                        max_id += 1
                        cnp.set("id", str(max_id))
                        nm = cnp.get("name") or "Shape"
                        cnp.set("name", f"{nm} (bloco {rep + 1})")
                    root.append(neo)
            out_map[name] = _ooxml_normalizar_declaracao_xml(
                ET.tostring(root, encoding="utf-8", xml_declaration=False)
            )
            alterado = True
        if not alterado:
            return
        buf = BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zw:
            for name in sorted(out_map):
                zw.writestr(name, out_map[name])
        out_path.write_bytes(buf.getvalue())
    except Exception:
        logging.getLogger(__name__).exception(
            "replicar_ancoras_sem_foto_por_bloco_em_drawing: falha em %s",
            out_path,
        )


def ficheiro_em_subpasta_planilha_mae_kartado_exportar(ficheiro: Path, pasta_xls: Path) -> bool:
    """
    True se ``ficheiro`` está em ``pasta_xls/exportar/`` (cópia da planilha-mãe pós-M01 Kartado).
    Excluir do M02, e-mail e leituras de EAF em pasta Exportar para não misturar com XLS de NC.
    """
    try:
        rel = ficheiro.resolve().relative_to(Path(pasta_xls).resolve())
    except ValueError:
        return False
    return len(rel.parts) >= 2 and rel.parts[0].casefold() == EXPORTAR_KARTADO_MAE_SUBDIR
