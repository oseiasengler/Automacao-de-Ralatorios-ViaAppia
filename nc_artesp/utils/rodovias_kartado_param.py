"""Carrega siglas do Excel Kartado «Rodovias e Municipios» e funde no mapa RODOVIAS."""

from __future__ import annotations

import logging
import re
import unicodedata
from pathlib import Path

logger = logging.getLogger(__name__)

_SKIP_DESC_KARTADO = frozenset({
    "rodovias principais",
    "rodovias interligacao",
    "rodovias de acesso",
    "rodovias vicinais",
})


def _norm_txt(val) -> str:
    if val is None:
        return ""
    s = str(val).strip().lower()
    if not s:
        return ""
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def _sheet_rodoivas_principal(wb) -> object | None:
    for ws in wb.worksheets:
        sn = _norm_txt(ws.title).strip()
        if sn.startswith("parametrizacao rodo"):
            return ws
    prefer = ("rodovia", "parametriz")
    best = None
    best_score = -1
    for ws in wb.worksheets:
        sn = _norm_txt(ws.title)
        score = sum(1 for p in prefer if p in sn)
        if score > best_score:
            best_score = score
            best = ws
    return best


def _linha_e_cols_cabecalho(ws) -> tuple[int, dict[str, int]] | None:
    for r in range(1, min(ws.max_row or 0, 25) + 1):
        cols: dict[str, int] = {}
        lim_c = min(ws.max_column or 0, 40)
        for c in range(1, lim_c + 1):
            h = _norm_txt(ws.cell(row=r, column=c).value)
            if not h:
                continue
            if h == "rodovia":
                cols["rodovia"] = c
            elif "descri" in h or h.startswith("descricao"):
                cols["descricao"] = c
        if "rodovia" in cols:
            return r, cols
    return None


def _variantes_chave_lookup(tok: str) -> list[str]:
    t = str(tok).strip()
    if not t:
        return []
    vs: set[str] = set()
    vs.add(t)
    vs.add(t.upper())
    compact = re.sub(r"\s+", "", t)
    if compact:
        vs.add(compact)
        vs.add(compact.upper())
    hy_sp = re.sub(r"\s*-\s*", " ", t)
    if hy_sp != t:
        vs.add(hy_sp.strip())
        vs.add(re.sub(r"\s+", "", hy_sp))
    return [x for x in vs if x]


def carregar_rodoivas_extra_kartado_xlsx(path: Path) -> dict[str, dict]:
    """Só adiciona chaves que ainda não existem no mapa principal (não sobrescreve core)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}

    try:
        from nc_artesp.utils.helpers import rodovia_info_heuristica_de_sigla
    except ImportError:
        from utils.helpers import rodovia_info_heuristica_de_sigla

    path = Path(path)
    if not path.is_file():
        return {}

    out: dict[str, dict] = {}
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = _sheet_rodoivas_principal(wb)
        if ws is None:
            return {}
        hdr = _linha_e_cols_cabecalho(ws)
        if not hdr:
            logger.warning("Kartado rodovias: cabecalho «Rodovia» nao encontrado em %s", path.name)
            return {}
        row_h, cols = hdr
        c_rod = cols["rodovia"]
        n_lin = 0
        for r in range(row_h + 1, (ws.max_row or 0) + 1):
            cel = ws.cell(row=r, column=c_rod).value
            if cel is None or str(cel).strip() == "":
                continue
            tok = str(cel).strip()
            info = rodovia_info_heuristica_de_sigla(tok)
            if not info:
                continue
            info = {**info}
            if "descricao" in cols:
                dcel = ws.cell(row=r, column=cols["descricao"]).value
                desc_txt = str(dcel).strip() if dcel else ""
                dn = _norm_txt(desc_txt)
                if desc_txt and dn not in _SKIP_DESC_KARTADO:
                    info["nome"] = desc_txt[:400]
            base = info
            for k in _variantes_chave_lookup(tok):
                if k not in out:
                    out[k] = base
            n_lin += 1
        logger.info("Kartado rodovias: %d linha(s) lidas de %s (%d chaves alias)", n_lin, path.name, len(out))
    finally:
        wb.close()
    return out
