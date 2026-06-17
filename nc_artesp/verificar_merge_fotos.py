"""
Script para verificar os merged ranges dos templates e dimensões calculadas
para as células de foto (Kria: C7, Resposta: B2).
Compara com as constantes do config (M02_FOTO_W/H e M02_FOTO_PDF_W/H).
"""
import sys
from pathlib import Path

# Permitir import do config e utils
sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from config import (
    M02_MODELO_KRIA,
    M02_MODELO_RESP,
    M02_FOTO_W, M02_FOTO_H,
    M02_FOTO_PDF_W, M02_FOTO_PDF_H,
)


def _col_px(ws, col_letter: str) -> float:
    dim = ws.column_dimensions.get(col_letter)
    width = dim.width if (dim and dim.width) else 8.0
    return width * 7 + 5


def _row_px(ws, row: int) -> float:
    dim = ws.row_dimensions.get(row)
    height = dim.height if (dim and dim.height) else 15.0
    return height * 4 / 3


def merged_range_px(ws, cell_addr: str):
    """Retorna (largura_px, altura_px) do merged range que contém cell_addr."""
    from openpyxl.utils import coordinate_to_tuple
    row, col = coordinate_to_tuple(cell_addr)
    for mc in ws.merged_cells.ranges:
        if mc.min_row <= row <= mc.max_row and mc.min_col <= col <= mc.max_col:
            w = sum(_col_px(ws, get_column_letter(c)) for c in range(mc.min_col, mc.max_col + 1))
            h = sum(_row_px(ws, r) for r in range(mc.min_row, mc.max_row + 1))
            return w, h, (mc.min_col, mc.min_row, mc.max_col, mc.max_row)
    col_letter = get_column_letter(col)
    return _col_px(ws, col_letter), _row_px(ws, row), (col, row, col, row)


def main():
    base = Path(__file__).resolve().parent
    kria_path = base / "assets" / "templates" / "Modelo Abertura Evento Kria Conserva Rotina.xlsx"
    resp_path = base / "assets" / "templates" / "Modelo.xlsx"

    print("=" * 70)
    print("VERIFICAÇÃO DE MERGE E TAMANHO DAS FOTOS NOS TEMPLATES")
    print("=" * 70)
    print(f"\nConstantes no config:")
    print(f"  Kria (nc):   {M02_FOTO_W} x {M02_FOTO_H} px  (9,86 cm × 7,51 cm)")
    print(f"  Resposta:    {M02_FOTO_PDF_W} x {M02_FOTO_PDF_H} px  (PDF)")

    # ─── Kria ───
    if not kria_path.exists():
        print(f"\n[ERRO] Template Kria não encontrado: {kria_path}")
    else:
        print(f"\n--- Template Kria: {kria_path.name} ---")
        wb = load_workbook(str(kria_path), read_only=False, data_only=True)
        ws = wb.active
        cell_foto_kria = "C7"  # j=8 → j-1=7
        w_px, h_px, bounds = merged_range_px(ws, cell_foto_kria)
        w_px, h_px = int(w_px), int(h_px)
        print(f"  Célula da foto: {cell_foto_kria}")
        print(f"  Merge (bounds): min_col={bounds[0]}, min_row={bounds[1]}, max_col={bounds[2]}, max_row={bounds[3]}")
        print(f"  Dimensão calculada do merge: {w_px} x {h_px} px")
        print(f"  Config (M02_FOTO_W/H):       {M02_FOTO_W} x {M02_FOTO_H} px")
        if (w_px, h_px) != (M02_FOTO_W, M02_FOTO_H):
            print(f"  >> DIFERENCA: merge do template nao coincide com o config.")
            print(f"     Para a foto preencher exatamente o espaco, o codigo usa o tamanho do merge.")
            print(f"     Para o config refletir o template: ARTESP_M02_FOTO_W={w_px} ARTESP_M02_FOTO_H={h_px}")
        else:
            print(f"  >> OK: merge coincide com o config.")
        wb.close()

    # ─── Resposta ───
    if not resp_path.exists():
        print(f"\n[ERRO] Template Resposta não encontrado: {resp_path}")
    else:
        print(f"\n--- Template Resposta: {resp_path.name} ---")
        wb = load_workbook(str(resp_path), read_only=False, data_only=True)
        ws = wb.active
        cell_foto_resp = "B2"
        w_px, h_px, bounds = merged_range_px(ws, cell_foto_resp)
        w_px, h_px = int(w_px), int(h_px)
        print(f"  Célula da foto: {cell_foto_resp}")
        print(f"  Merge (bounds): min_col={bounds[0]}, min_row={bounds[1]}, max_col={bounds[2]}, max_row={bounds[3]}")
        print(f"  Dimensão calculada do merge: {w_px} x {h_px} px")
        print(f"  Config (M02_FOTO_PDF_W/H):   {M02_FOTO_PDF_W} x {M02_FOTO_PDF_H} px")
        if (w_px, h_px) != (M02_FOTO_PDF_W, M02_FOTO_PDF_H):
            print(f"  >> DIFERENCA: merge do template nao coincide com o config.")
            print(f"     O codigo redimensiona a imagem para o tamanho do merge.")
            print(f"     Para alinhar config ao template: ARTESP_M02_FOTO_PDF_W={w_px} ARTESP_M02_FOTO_PDF_H={h_px}")
        else:
            print(f"  >> OK: merge coincide com o config.")
        wb.close()

    print("\n" + "=" * 70)
    print("Nota: Em gerar_modelo_foto.py, _inserir_imagem() usa o merge do template para")
    print("Resposta/Pendentes (não-C); só Kria (C*) usa tamanho fixo 9,70×7,49 cm.")
    print("M02_FOTO_PDF_* no config é fallback se o merge não tiver dimensão detetável.")
    print("=" * 70)


if __name__ == "__main__":
    main()
