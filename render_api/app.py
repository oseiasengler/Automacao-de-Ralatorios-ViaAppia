import asyncio
import logging
from contextlib import asynccontextmanager
from io import BytesIO
import os
import re
import sys
import json
import base64
import secrets
import hashlib
import tempfile
import zipfile
import datetime
import platform
from zoneinfo import ZoneInfo
import time
import subprocess
import shutil
import calendar
from pathlib import Path
from queue import Empty, Queue
from collections import defaultdict
from typing import Any, Dict, Literal, Optional

import pandas as pd
import jsonschema
import jwt
import threading
from fastapi import (
    FastAPI, Request, HTTPException, UploadFile, File,
    Header, Cookie, Depends, Body, Form, Query,
)
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, StreamingResponse, RedirectResponse
from starlette.responses import PlainTextResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer
from pydantic import BaseModel, Field

try:
    from .auth_crypto import gerar_hash_senha, verificar_senha
except ImportError:
    from auth_crypto import gerar_hash_senha, verificar_senha

# ═══════════════════════════════════════════════════════════════
#  CAMINHOS BASE — seguros para qualquer ambiente (local e Render)
# ═══════════════════════════════════════════════════════════════
BASE_DIR   = Path(__file__).resolve().parent.parent
WEB_DIR    = Path(__file__).resolve().parent / "web"
STATIC_DIR = Path(__file__).resolve().parent / "web-static"
# Schema e assets na pasta do projeto (c:\GeradorARTESP\assets\schema)
SCHEMA_PATH = BASE_DIR / "assets" / "schema"

_env_output = (os.getenv("ARTESP_OUTPUT_DIR") or "").strip()
if _env_output:
    OUTPUT_PATH = Path(_env_output).resolve()
else:
    OUTPUT_PATH = (
        # /tmp tem ~400 GB no Render (disco raiz efêmero) — usado para outputs temporários.
        # /data tem apenas 1 GB (disco persistente) — reservado para users.json e metrics.json.
        Path("/tmp/outputs") if platform.system() == "Linux"
        else Path(__file__).resolve().parent.parent / "outputs"
    ).resolve()
OUTPUT_PATH.mkdir(parents=True, exist_ok=True)

# ═══════════════════════════════════════════════════════════════
#  COFRE / DISCO PERSISTENTE — users.json e metrics.json
#  Sempre usar caminho absoluto para evitar conflito cwd (ex.: /app) vs disco em /data.
#  Env ARTESP_DATA_DIR tem prioridade (ex.: /data no Render). Linux sem env = /data.
#  Windows sem env = pasta do projeto (BASE_DIR/data) para cenário local.
# ═══════════════════════════════════════════════════════════════
_env_data = (os.getenv("ARTESP_DATA_DIR") or "").strip()
if _env_data:
    DATA_DIR = Path(_env_data).resolve()
else:
    DATA_DIR = (
        Path("/data") if platform.system() == "Linux"
        else (BASE_DIR / "data").resolve()
    )
# Arquivo legado para migrar para DATA_DIR (raiz do projeto ou cwd)
OLD_USER_DB = (BASE_DIR / "users.json").resolve()

DATA_DIR.mkdir(parents=True, exist_ok=True)
USER_DB_PATH = (DATA_DIR / "users.json").resolve()
METRICS_PATH = (DATA_DIR / "metrics.json").resolve()

# Log do caminho de usuários ao iniciar (diagnóstico pós-deploy)
logging.info("Banco de usuários: %s (existe: %s)", USER_DB_PATH, USER_DB_PATH.is_file())


def _executar_migracao_users():
    """Move users.json do disco efêmero para o disco persistente (Render)."""
    if OLD_USER_DB.is_file() and not USER_DB_PATH.is_file():
        try:
            shutil.copy2(str(OLD_USER_DB), str(USER_DB_PATH))
            logging.info("Usuários migrados para %s", USER_DB_PATH)
        except OSError as e:
            logging.warning("Erro na migração de users.json: %s", e)


_executar_migracao_users()

SCHEMA_MAP = {
    "conserva": "conserva.schema.r0.json",
    "obras": "obras.schema.r0.json",
}

# ═══════════════════════════════════════════════════════════════
#  AUTO-LIMPEZA — remove arquivos antigos para não lotar o disco
# ═══════════════════════════════════════════════════════════════
def _limpar_arquivos_antigos(horas: int = 2) -> dict:
    """
    Remove arquivos em OUTPUT_PATH com mais de X horas e aciona limpeza
    dos diretórios de jobs NC expirados (OUTPUT_PATH/nc/<job_id>/).
    Retorna {"removidos": N, "erros": M, "bytes_liberados": B}.
    """
    if horas < 1:
        horas = 1
    limite = time.time() - (horas * 3600)
    removidos = 0
    erros = 0
    bytes_liberados = 0
    try:
        if not OUTPUT_PATH.is_dir():
            return {"removidos": 0, "erros": 0, "bytes_liberados": 0}
        for f in OUTPUT_PATH.iterdir():
            if not f.is_file():
                continue
            try:
                mtime = f.stat().st_mtime
                if mtime < limite:
                    size = f.stat().st_size
                    f.unlink()
                    removidos += 1
                    bytes_liberados += size
            except OSError:
                erros += 1
        if removidos:
            logging.info(
                "Auto-limpeza: %d arquivo(s) removido(s), ~%.1f MB liberados",
                removidos,
                bytes_liberados / (1024 * 1024),
            )
    except Exception as e:
        logging.warning("Auto-limpeza falhou: %s", e)
        erros += 1

    # Limpar diretórios de jobs NC expirados (OUTPUT_PATH/nc/<job_id>/)
    # Esses diretórios nunca eram alcançados pelo loop acima (is_file() os pulava).
    nc_removidos = _limpar_jobs_nc()
    removidos += nc_removidos

    # Limpar ZIPs de fotos (/data/fotos_downloads/) — fora de OUTPUT_PATH, nunca limpa pelo loop acima.
    removidos += _limpar_fotos_downloads()

    return {"removidos": removidos, "erros": erros, "bytes_liberados": bytes_liberados}


def _limpar_jobs_nc() -> int:
    """
    Remove jobs NC expirados em OUTPUT_PATH/nc/.
    Delega para nc_router._cleanup_expired_jobs se disponível;
    caso contrário usa lógica local (mtime > 2h).
    """
    try:
        try:
            from render_api.nc_router import _cleanup_expired_jobs
        except ImportError:
            from nc_router import _cleanup_expired_jobs  # type: ignore
        return _cleanup_expired_jobs()
    except Exception:
        pass
    # Fallback: remoção por mtime quando nc_router não está disponível
    nc_base = OUTPUT_PATH / "nc"
    if not nc_base.is_dir():
        return 0
    limite = time.time() - 2 * 3600
    removed = 0
    for job_dir in list(nc_base.iterdir()):
        if not job_dir.is_dir():
            continue
        try:
            if job_dir.stat().st_mtime < limite:
                import shutil as _shutil
                _shutil.rmtree(job_dir, ignore_errors=True)
                removed += 1
        except OSError:
            pass
    return removed


def _limpar_fotos_downloads() -> int:
    """
    Remove ZIPs de fotos antigos em /data/fotos_downloads/.
    Essa pasta fica FORA de OUTPUT_PATH, por isso nunca era alcançada
    pelo cleanup geral — acumulava indefinidamente entre deploys.
    TTL: 30 minutos (arquivos criados por sessão de download).
    """
    _env_data = (os.getenv("ARTESP_OUTPUT_DIR") or os.getenv("ARTESP_DATA_DIR") or "").strip()
    if _env_data:
        fotos_dir = Path(_env_data).resolve() / "fotos_downloads"
    elif platform.system() == "Linux":
        fotos_dir = Path("/tmp/outputs/fotos_downloads").resolve()
    else:
        return 0  # local: não limpar
    if not fotos_dir.is_dir():
        return 0
    limite = time.time() - 30 * 60  # 30 minutos (arquivos de sessão de download)
    removed = 0
    for f in fotos_dir.iterdir():
        if not f.is_file():
            continue
        try:
            if f.stat().st_mtime < limite:
                f.unlink(missing_ok=True)
                removed += 1
        except OSError:
            pass
    if removed:
        logging.info("Limpeza fotos_downloads: %d arquivo(s) removido(s)", removed)
    return removed


def _ler_int_env(nome, padrao):
    raw = (os.getenv(nome) or "").strip()
    if not raw:
        return padrao
    match = re.search(r"\d+", raw)
    if not match:
        return padrao
    try:
        return int(match.group(0))
    except Exception:
        return padrao


# Horas para auto-limpeza (arquivos > X h são removidos). Env: ARTESP_LIMPEZA_HORAS
LIMPEZA_HORAS = max(1, _ler_int_env("ARTESP_LIMPEZA_HORAS", 2))


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Startup: limpeza de disco PRIMEIRO (evita ENOSPC), depois sincroniza banco e inicia loop periódico."""
    loop = asyncio.get_event_loop()
    # Limpeza ANTES de qualquer escrita em disco (users.json, metrics.json etc.)
    await loop.run_in_executor(None, lambda: _limpar_arquivos_antigos(LIMPEZA_HORAS))
    await loop.run_in_executor(None, sincronizar_banco_usuarios)

    intervalo = _ler_int_env("ARTESP_LIMPEZA_INTERVALO_SEG", 1 * 3600)  # loop a cada 1h
    if intervalo > 0:
        horas = LIMPEZA_HORAS

        async def _loop_limpeza():
            while True:
                await asyncio.sleep(intervalo)
                await asyncio.get_event_loop().run_in_executor(
                    None, lambda h=horas: _limpar_arquivos_antigos(h)
                )

        asyncio.create_task(_loop_limpeza())

    yield
    # shutdown — nada a fazer por ora


app = FastAPI(title="API Gerador ARTESP", lifespan=lifespan)

try:
    from render_api.fotos_router import router as fotos_router
    app.include_router(fotos_router)
    logging.info("Router /fotos carregado com sucesso.")
except ImportError:
    try:
        from fotos_router import router as fotos_router
        app.include_router(fotos_router)
        logging.info("Router /fotos carregado (modo local).")
    except ImportError as _e:
        logging.warning("Router /fotos não carregado: %s", _e)

try:
    from render_api.nc_router import router as nc_router
    app.include_router(nc_router)
    logging.info("Router /nc carregado com sucesso.")
except ImportError:
    try:
        from nc_router import router as nc_router
        app.include_router(nc_router)
        logging.info("Router /nc carregado (modo local).")
    except ImportError as _e:
        logging.warning("Router /nc não carregado: %s", _e)

# CORS: para carregar GeoJSON/API em mapa em site externo (ex.: Locaweb), inclua a origem
# em ARTESP_CORS_ORIGINS. Com allow_credentials=True não dá para usar "*"; use lista explícita
# ou ARTESP_CORS_ORIGINS=* para refletir qualquer Origin (menos seguro, útil para testes).
_cors_raw = (os.getenv("ARTESP_CORS_ORIGINS") or "").strip()
_CORS_REFLECT_ANY = _cors_raw.strip().lower() == "*"
_CORS_ORIGINS: list[str] = []
if _CORS_REFLECT_ANY:
    _CORS_ORIGINS = []  # middleware customizado reflete o Origin
else:
    _CORS_ORIGINS = [o.strip() for o in _cors_raw.split(",") if o.strip()] if _cors_raw else [
        "http://localhost:8000",
        "http://127.0.0.1:8000",
        "http://localhost:3000",
    ]

if _CORS_REFLECT_ANY:

    @app.middleware("http")
    async def cors_reflect_origin(request: Request, call_next):
        origin = request.headers.get("origin")
        if request.method == "OPTIONS":
            r = Response(status_code=204)
            if origin:
                r.headers["Access-Control-Allow-Origin"] = origin
                r.headers["Access-Control-Allow-Credentials"] = "true"
                r.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, PATCH, DELETE, OPTIONS"
                r.headers["Access-Control-Allow-Headers"] = "*"
                r.headers["Access-Control-Max-Age"] = "86400"
            return r
        response = await call_next(request)
        if origin:
            response.headers["Access-Control-Allow-Origin"] = origin
            response.headers["Access-Control-Allow-Credentials"] = "true"
            response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, PATCH, DELETE, OPTIONS"
            response.headers["Access-Control-Allow-Headers"] = "*"
            response.headers["Access-Control-Expose-Headers"] = "*"
        return response

else:
    app.add_middleware(
        CORSMiddleware,
        allow_origins=_CORS_ORIGINS,
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
        expose_headers=["Content-Disposition"],
    )

# Rate limit global (slowapi): 50 req/10s por IP — protege contra scanners/curiosos
_slowapi_limit = (os.getenv("ARTESP_RATE_GLOBAL") or "50/10seconds").strip()
def _slowapi_key(request: Request) -> str:
    fwd = (request.headers.get("x-forwarded-for") or "").split(",")
    if fwd and fwd[0].strip():
        return fwd[0].strip()
    if request.headers.get("x-real-ip"):
        return request.headers.get("x-real-ip").strip()
    return request.client.host if request.client else "unknown"
try:
    from slowapi import Limiter, _rate_limit_exceeded_handler
    from slowapi.middleware import SlowAPIMiddleware
    from slowapi.errors import RateLimitExceeded
    _slowapi_limiter = Limiter(key_func=_slowapi_key, default_limits=[_slowapi_limit])
    app.state.limiter = _slowapi_limiter
    app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
    app.add_middleware(SlowAPIMiddleware)
    logging.info("slowapi: rate limit global %s por IP", _slowapi_limit)
except ImportError:
    logging.warning("slowapi não instalado — rate limit global desativado. pip install slowapi")

if STATIC_DIR.is_dir():
    app.mount("/web-static", StaticFiles(directory=str(STATIC_DIR)), name="web_static")


_ROBOTS_TXT = """# Bloqueia varredura da API por crawlers.
User-agent: *
Disallow: /api/
Disallow: /auth/
Disallow: /fotos/
Disallow: /nc/
"""

@app.get("/robots.txt", include_in_schema=False)
async def robots_txt():
    """robots.txt na raiz para crawlers."""
    return PlainTextResponse(_ROBOTS_TXT)

@app.get("/favicon.ico", include_in_schema=False)
async def favicon():
    """Favicon na raiz (evita 404 nos logs)."""
    for name, media in (("favicon.ico", "image/x-icon"), ("favicon.png", "image/png"), ("favicon.svg", "image/svg+xml")):
        path = STATIC_DIR / name
        if path.is_file():
            return FileResponse(str(path), media_type=media)
    return Response(status_code=204)


# Rotas de probe (/.env, wp-login, etc.): 404 + log + rate limit
@app.get("/.env", include_in_schema=False)
async def probe_env(request: Request):
    return await _probe_404(request, ".env")

@app.get("/wp-login.php", include_in_schema=False)
async def probe_wp_login(request: Request):
    return await _probe_404(request, "wp-login.php")

@app.get("/wp-admin", include_in_schema=False)
async def probe_wp_admin(request: Request):
    return await _probe_404(request, "wp-admin")

@app.get("/wp-admin/", include_in_schema=False)
async def probe_wp_admin_slash(request: Request):
    return await _probe_404(request, "wp-admin/")

@app.get("/.git/config", include_in_schema=False)
async def probe_git_config(request: Request):
    return await _probe_404(request, ".git/config")

@app.get("/config.php", include_in_schema=False)
async def probe_config_php(request: Request):
    return await _probe_404(request, "config.php")


@app.exception_handler(Exception)
async def global_exception_handler(request, exc: Exception):
    """Garante que qualquer exceção não tratada retorne JSON, nunca HTML."""
    logging.exception("Exceção não tratada em %s: %s", request.url.path, exc)
    return JSONResponse(
        status_code=500,
        content={
            "detail": str(exc),
            "type": type(exc).__name__,
        },
    )


AUTH_BEARER = HTTPBearer(auto_error=False)
TOKENS: Dict[str, Dict[str, Any]] = {}
JWT_SECRET = os.getenv("ARTESP_JWT_SECRET") or os.getenv("ARTESP_WEB_JWT_SECRET") or "artesp-gerador-secret-change-in-production"
JWT_ALGORITHM = "HS256"
TOKENS_REVOGADOS: set = set()


# TTL da sessão (segundos). Env: ARTESP_WEB_TOKEN_TTL_SECONDS. Default 24h.
TOKEN_TTL_SECONDS = max(60, _ler_int_env("ARTESP_WEB_TOKEN_TTL_SECONDS", 86400))

# Proprietário (admin master): não pode ser removido, rebaixado nem bloqueado. Env: ARTESP_OWNER_EMAIL.
OWNER_EMAIL = (os.getenv("ARTESP_OWNER_EMAIL") or "").strip().lower()


def _eh_proprietario(email: str) -> bool:
    """Indica se o e-mail é o do proprietário (imutável nas ações admin)."""
    return bool(OWNER_EMAIL and (email or "").strip().lower() == OWNER_EMAIL)


# ═══════════════════════════════════════════════════════════════
#  [P2] RATE LIMITER — proteção contra abuso / DoS
# ═══════════════════════════════════════════════════════════════
class RateLimiter:
    def __init__(self):
        self._requests: Dict[str, list] = defaultdict(list)
        self._last_cleanup: float = time.time()
        self._cleanup_interval: float = 300.0

    def _limpar_antigos(self, janela_max: float = 3600.0):
        agora = time.time()
        if agora - self._last_cleanup < self._cleanup_interval:
            return
        chaves_remover = []
        for chave, timestamps in self._requests.items():
            self._requests[chave] = [t for t in timestamps if agora - t < janela_max]
            if not self._requests[chave]:
                chaves_remover.append(chave)
        for chave in chaves_remover:
            del self._requests[chave]
        self._last_cleanup = agora

    def verificar(self, chave: str, limite: int, janela_segundos: float) -> bool:
        self._limpar_antigos()
        agora = time.time()
        self._requests[chave] = [t for t in self._requests[chave] if agora - t < janela_segundos]
        if len(self._requests[chave]) >= limite:
            return False
        self._requests[chave].append(agora)
        return True

    def tempo_restante(self, chave: str, janela_segundos: float) -> float:
        agora = time.time()
        validos = [t for t in self._requests.get(chave, []) if agora - t < janela_segundos]
        if not validos:
            return 0.0
        return max(0.0, janela_segundos - (agora - min(validos)))


_rate_limiter = RateLimiter()

RATE_LIMIT_LOGIN_MAX = _ler_int_env("ARTESP_RATE_LOGIN_MAX", 5)
RATE_LIMIT_LOGIN_JANELA = _ler_int_env("ARTESP_RATE_LOGIN_JANELA", 60)
RATE_LIMIT_GERAR_MAX = _ler_int_env("ARTESP_RATE_GERAR_MAX", 10)
RATE_LIMIT_GERAR_JANELA = _ler_int_env("ARTESP_RATE_GERAR_JANELA", 300)
RATE_LIMIT_SIMULAR_MAX = _ler_int_env("ARTESP_RATE_SIMULAR_MAX", 20)
RATE_LIMIT_SIMULAR_JANELA = _ler_int_env("ARTESP_RATE_SIMULAR_JANELA", 300)
RATE_LIMIT_DOWNLOAD_MAX = _ler_int_env("ARTESP_RATE_DOWNLOAD_MAX", 50)
RATE_LIMIT_DOWNLOAD_JANELA = _ler_int_env("ARTESP_RATE_DOWNLOAD_JANELA", 300)
RATE_LIMIT_PROBE_MAX = _ler_int_env("ARTESP_RATE_PROBE_MAX", 30)
RATE_LIMIT_PROBE_JANELA = _ler_int_env("ARTESP_RATE_PROBE_JANELA", 60)


def _get_client_ip(request: Request) -> str:
    forwarded = (request.headers.get("x-forwarded-for") or "").split(",")
    if forwarded and forwarded[0].strip():
        return forwarded[0].strip()
    real_ip = request.headers.get("x-real-ip")
    if real_ip:
        return real_ip.strip()
    return request.client.host if request.client else "unknown"


def _check_rate_limit(request: Request, chave_extra: str, limite: int, janela: float):
    ip = _get_client_ip(request)
    chave = f"{ip}:{chave_extra}"
    if not _rate_limiter.verificar(chave, limite, janela):
        tempo = _rate_limiter.tempo_restante(chave, janela)
        raise HTTPException(
            status_code=429,
            detail=f"Muitas requisições. Aguarde {int(tempo)} segundos.",
            headers={"Retry-After": str(int(tempo))},
        )


async def _probe_404(request: Request, path: str):
    """Rate limit + log para rotas de probe (/.env, wp-login, etc.); retorna 404."""
    _check_rate_limit(request, "probe", RATE_LIMIT_PROBE_MAX, RATE_LIMIT_PROBE_JANELA)
    ip = _get_client_ip(request)
    logging.warning("Probe bloqueado: path=%s client_ip=%s", path, ip)
    return Response(status_code=404)


# ═══════════════════════════════════════════════════════════════
#  [P2] LIMITE DE UPLOAD — proteção contra arquivos gigantes
# ═══════════════════════════════════════════════════════════════
MAX_UPLOAD_BYTES = _ler_int_env("ARTESP_MAX_UPLOAD_MB", 50) * 1024 * 1024
_NOME_ARQUIVO_REGEX = re.compile(r"[^a-zA-Z0-9_\-.]")


def _validar_upload(arquivo: UploadFile, max_bytes: int = None) -> None:
    """
    Valida o arquivo enviado:
    - Extensão permitida (.xlsx, .xls)
    - Tamanho máximo
    - Nome seguro (sem path traversal)
    Levanta HTTPException se inválido.
    """
    max_bytes = max_bytes or MAX_UPLOAD_BYTES

    nome = (arquivo.filename or "").strip()
    if not nome:
        raise HTTPException(status_code=400, detail="Nome do arquivo ausente.")
    ext = nome.lower().rsplit(".", 1)[-1] if "." in nome else ""
    if ext not in ("xlsx", "xls"):
        raise HTTPException(status_code=400, detail="Apenas arquivos .xlsx ou .xls são aceitos.")

    if ".." in nome or "/" in nome or "\\" in nome:
        raise HTTPException(status_code=400, detail="Nome de arquivo inválido.")

    if hasattr(arquivo, "size") and arquivo.size is not None:
        if arquivo.size > max_bytes:
            mb = max_bytes / (1024 * 1024)
            raise HTTPException(
                status_code=413,
                detail=f"Arquivo muito grande. Máximo: {mb:.0f}MB.",
            )


async def _ler_upload_com_limite(arquivo: UploadFile, max_bytes: int = None) -> bytes:
    """
    Lê o conteúdo do upload com limite de tamanho.
    Lê em chunks para não alocar memória desnecessária.
    Levanta HTTPException 413 se exceder o limite.
    """
    max_bytes = max_bytes or MAX_UPLOAD_BYTES
    chunks = []
    total = 0
    chunk_size = 64 * 1024  # 64KB por vez

    while True:
        chunk = await arquivo.read(chunk_size)
        if not chunk:
            break
        total += len(chunk)
        if total > max_bytes:
            mb = max_bytes / (1024 * 1024)
            raise HTTPException(
                status_code=413,
                detail=f"Arquivo excede o limite de {mb:.0f}MB.",
            )
        chunks.append(chunk)

    return b"".join(chunks)


def _nome_arquivo_seguro(nome: str) -> str:
    """
    Sanitiza o nome do arquivo para evitar path traversal e caracteres perigosos.
    """
    nome = (nome or "upload").strip()
    nome = nome.replace("\\", "/")
    nome = nome.split("/")[-1]
    base, _, ext = nome.rpartition(".")
    if not base:
        base = nome
        ext = ""
    base = _NOME_ARQUIVO_REGEX.sub("_", base)[:100]
    ext = _NOME_ARQUIVO_REGEX.sub("", ext)[:10]
    return f"{base}.{ext}" if ext else base


# ═══════════════════════════════════════════════════════════════
#  [P3] ESTRUTURA DE PASTAS — compatível com auditoria ARTESP
# ═══════════════════════════════════════════════════════════════
_MESES_NOME_COMPLETO = {
    1: "janeiro", 2: "fevereiro", 3: "marco", 4: "abril",
    5: "maio", 6: "junho", 7: "julho", 8: "agosto",
    9: "setembro", 10: "outubro", 11: "novembro", 12: "dezembro",
}


def _classificar_versao(versao_key: str) -> str:
    """
    Retorna o tipo de período com base na versão.
    'A' → 'Anual', 'M' → 'Mensal', 'E' → 'Executado'.
    """
    v = (versao_key or "").strip().upper()
    mapa = {"A": "Anual", "M": "Mensal", "E": "Executado"}
    return mapa.get(v, "Anual")


def _construir_caminho_artesp(
    lote: str,
    ano: int,
    versao_key: str,
    mes: int = None,
) -> str:
    """
    [P3] Monta o caminho relativo dentro do ZIP (uma única pasta).
    Ex: "Lote_21 Anual", "Lote_13 Mensal 03", "Lote_26 Executado 06".
    """
    tipo = _classificar_versao(versao_key)
    lote_num = (lote or "").replace("L", "").strip() or "0"

    if tipo in ("Mensal", "Executado") and mes:
        mes_int = int(mes) if isinstance(mes, (int, float, str)) else 0
        nome_mes = _MESES_NOME_COMPLETO.get(mes_int, f"{mes_int:02d}")
        return f"Lote_{lote_num} {tipo} {nome_mes}"

    return f"Lote_{lote_num} {tipo}"


def _construir_nome_zip_artesp(
    lote: str,
    modalidade: str,
    versao_key: str,
    ano: int,
    mes: int = None,
) -> str:
    """
    [P3] Nome do ZIP no padrão ARTESP.
    """
    tipo = _classificar_versao(versao_key)
    mod_label = "conservacao" if str(modalidade).lower() in ("1", "conserva", "conservacao") else "obras"

    if tipo == "Mensal" and mes:
        return f"{lote}_{mod_label}_programado_mes_{int(mes):02d}_{ano}.zip"
    elif tipo == "Executado" and mes:
        return f"{lote}_{mod_label}_executado_mes_{int(mes):02d}_{ano}.zip"
    else:
        return f"{lote}_{mod_label}_{ano}_Anual.zip"


class ProcessarRelatorioPayload(BaseModel):
    geojson: Dict[str, Any]
    template_schema: Literal["conserva", "obras"] = Field(default="conserva", alias="schema")
    lote: Optional[str] = Field(default=None, pattern=r"^L\d{2}$")
    assinar: bool = True
    nome_arquivo: Optional[str] = None


class LoginPayload(BaseModel):
    """Aceita email + senha (formulários PT) ou email + password (compatível com referência)."""
    email: str
    senha: Optional[str] = None
    password: Optional[str] = None


class NovoUsuarioPayload(BaseModel):
    """Payload para adicionar novo usuário via rota admin."""

    email: str = Field(..., min_length=5, description="E-mail válido")
    senha: str = Field(..., min_length=1, description="Senha")
    role: str = "user"  # Pode ser 'user' ou 'admin'


class TrocarSenhaPayload(BaseModel):
    """Troca de senha (primeiro acesso ou alteração)."""
    senha_atual: str = Field(..., min_length=1)
    nova_senha: str = Field(..., min_length=6, description="Mínimo 6 caracteres")


class RedefinirSenhaAdminPayload(BaseModel):
    """Admin redefine senha de um usuário (senha temporária; usuário deve trocar no próximo login)."""
    email: str = Field(..., min_length=5)
    nova_senha_temporaria: str = Field(..., min_length=6)


class AlterarRolePayload(BaseModel):
    """Admin altera perfil do usuário (user ↔ admin)."""
    email: str = Field(..., min_length=5)
    role: str = Field(..., description="admin ou user")


class BloquearUsuarioPayload(BaseModel):
    """Admin bloqueia ou desbloqueia usuário (disabled)."""
    email: str = Field(..., min_length=5)
    bloquear: bool = True


class RemoverUsuarioPayload(BaseModel):
    """Admin remove usuário do banco (apenas users.json)."""
    email: str = Field(..., min_length=5)


USERS_JSON_PATH = USER_DB_PATH
TZ_BRASILIA = ZoneInfo("America/Sao_Paulo")


def _agora_brasilia() -> datetime.datetime:
    """Data/hora atual em Brasília (arquivos e relatórios)."""
    return datetime.datetime.now(TZ_BRASILIA)


def _normalizar_banco_usuarios(usuarios: dict) -> dict:
    """Chaves em lowercase para evitar duplicata por caixa no cadastro."""
    if not isinstance(usuarios, dict):
        return {}
    return {(str(k).strip().lower()): v for k, v in usuarios.items() if (k or "").strip()}


# ═══════════════════════════════════════════════════════════════
#  [P2] FILE LOCK — proteção contra race condition no users.json
# ═══════════════════════════════════════════════════════════════
class FileLock:
    """Lock por arquivo para leitura+escrita atômica do users.json."""

    def __init__(self):
        self._locks: Dict[str, threading.Lock] = {}
        self._global_lock = threading.Lock()

    def get_lock(self, path: str) -> threading.Lock:
        with self._global_lock:
            if path not in self._locks:
                self._locks[path] = threading.Lock()
            return self._locks[path]


_file_locks = FileLock()


def carregar_banco_usuarios() -> dict:
    """
    [P2] Carrega usuários do users.json com lock de leitura.
    Retorna {} se o arquivo não existir ou for inválido.
    """
    lock = _file_locks.get_lock(str(USERS_JSON_PATH))
    with lock:
        try:
            if USERS_JSON_PATH.is_file():
                with open(USERS_JSON_PATH, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    usuarios = data if isinstance(data, dict) else {}
                    return _normalizar_banco_usuarios(usuarios)
        except (OSError, json.JSONDecodeError) as e:
            logging.warning("[P2] Falha ao carregar users.json (%s): %s", USERS_JSON_PATH, e)
    return {}


def salvar_banco_usuarios(usuarios: dict) -> None:
    """
    [P2] Persiste o dicionário de usuários em users.json com escrita atômica.
    Usa arquivo temporário + os.replace para evitar corrupção.
    """
    lock = _file_locks.get_lock(str(USERS_JSON_PATH))
    with lock:
        try:
            USERS_JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
            temp_path = str(USERS_JSON_PATH) + ".tmp"
            with open(temp_path, "w", encoding="utf-8") as f:
                json.dump(usuarios, f, indent=2, ensure_ascii=False)
                f.flush()
                os.fsync(f.fileno())
            os.replace(temp_path, str(USERS_JSON_PATH))
        except OSError as e:
            logging.error("[P2] Falha ao salvar users.json: %s", e)
            try:
                temp_path = str(USERS_JSON_PATH) + ".tmp"
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
            except OSError:
                pass
            raise HTTPException(status_code=500, detail="Falha ao persistir usuários.")


def _modificar_banco_usuarios(fn) -> dict:
    """
    [P2] Lê, modifica e salva users.json de forma atômica.
    fn: callable que recebe o dict de usuários e retorna o dict modificado.
    Retorna o dict modificado.
    """
    lock = _file_locks.get_lock(str(USERS_JSON_PATH))
    with lock:
        usuarios = {}
        try:
            if USERS_JSON_PATH.is_file():
                with open(USERS_JSON_PATH, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    usuarios = _normalizar_banco_usuarios(data if isinstance(data, dict) else {})
        except (OSError, json.JSONDecodeError) as e:
            logging.warning("[P2] Falha ao ler users.json para modificação: %s", e)

        usuarios = fn(usuarios)

        temp_path = str(USERS_JSON_PATH) + ".tmp"
        try:
            USERS_JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
            with open(temp_path, "w", encoding="utf-8") as f:
                json.dump(usuarios, f, indent=2, ensure_ascii=False)
                f.flush()
                os.fsync(f.fileno())
            os.replace(temp_path, str(USERS_JSON_PATH))
        except OSError as e:
            logging.error("[P2] Falha ao salvar users.json após modificação: %s", e)
            try:
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
            except OSError:
                pass
            raise HTTPException(status_code=500, detail="Falha ao persistir usuários.")

    return usuarios


def sincronizar_banco_usuarios():
    """
    [P2] Sincroniza banco de usuários com lock atômico.
    Lógica do Cadeado: só adiciona ou atualiza, nunca limpa.
    """

    def _sincronizar(usuarios):
        admin_raw = os.environ.get("ADMIN_PASSWORD_MIGRATION", "").strip()
        if not admin_raw:
            return usuarios

        try:
            dados = json.loads(admin_raw)
        except json.JSONDecodeError:
            logging.warning("[P2] ADMIN_PASSWORD_MIGRATION não é JSON válido.")
            return usuarios

        if not isinstance(dados, dict):
            return usuarios

        alterados = 0
        for email_key, valor in dados.items():
            email_key = email_key.strip().lower()
            if not email_key:
                continue

            if isinstance(valor, dict):
                hashed = valor.get("hashed_password", "")
                role = valor.get("role", "admin")
            elif isinstance(valor, str):
                hashed = valor
                role = "admin"
            else:
                continue

            if not hashed:
                continue

            existente = usuarios.get(email_key)
            if isinstance(existente, dict) and existente.get("hashed_password") == hashed:
                continue

            usuarios[email_key] = {
                "hashed_password": hashed,
                "disabled": False,
                "role": role,
            }
            alterados += 1

        if alterados > 0:
            logging.info("[P2] Sincronização: %d usuário(s) atualizado(s).", alterados)

        return usuarios

    try:
        _modificar_banco_usuarios(_sincronizar)
    except Exception as e:
        logging.error("[P2] Falha na sincronização de usuários: %s", e)


def _carregar_usuarios_web():
    """
    Carrega usuarios de ARTESP_WEB_USERS.
    Formatos aceitos:
      1) JSON objeto: {"email@dom.com":"senha"}
      2) JSON lista: [{"email":"...","senha":"..."}, ...]
      3) CSV: email1@dom.com:senha1,email2@dom.com:senha2
    Fallback:
      ARTESP_WEB_ADMIN_EMAIL + ARTESP_WEB_ADMIN_PASSWORD
    """
    raw = (os.getenv("ARTESP_WEB_USERS") or "").strip()
    usuarios: Dict[str, str] = {}

    if raw:
        if raw.startswith("{") or raw.startswith("["):
            try:
                parsed = json.loads(raw)
                if isinstance(parsed, dict):
                    for email, senha in parsed.items():
                        if str(email).strip():
                            usuarios[str(email).strip().lower()] = str(senha or "")
                elif isinstance(parsed, list):
                    for item in parsed:
                        if not isinstance(item, dict):
                            continue
                        email = str(item.get("email", "")).strip().lower()
                        senha = str(item.get("senha", ""))
                        if email:
                            usuarios[email] = senha
            except json.JSONDecodeError:
                pass
        else:
            for part in raw.split(","):
                part = part.strip()
                if ":" not in part:
                    continue
                email, senha = part.split(":", 1)
                email = email.strip().lower()
                if email:
                    usuarios[email] = senha

    if not usuarios:
        admin_email = (os.getenv("ARTESP_WEB_ADMIN_EMAIL") or "").strip().lower()
        admin_password = os.getenv("ARTESP_WEB_ADMIN_PASSWORD") or ""
        if admin_email and admin_password:
            usuarios[admin_email] = admin_password

    return usuarios


USUARIOS_WEB = _carregar_usuarios_web()


def _limpar_tokens_expirados():
    agora = datetime.datetime.now(datetime.timezone.utc).timestamp()
    expirados = [token for token, data in TOKENS.items() if data.get("exp", 0) <= agora]
    for token in expirados:
        TOKENS.pop(token, None)


def _criar_access_token(data: dict) -> str:
    """JWT com sub (email) e exp."""
    payload = dict(data)
    now = datetime.datetime.now(datetime.timezone.utc).timestamp()
    payload.setdefault("exp", now + TOKEN_TTL_SECONDS)
    payload.setdefault("iat", now)
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def _gerar_token(email: str):
    """JWT com sub=email."""
    token_data = {"sub": email}
    return _criar_access_token(token_data)


def _validar_token(token: str):
    """Valida JWT; retorna email (sub) ou None."""
    if token in TOKENS_REVOGADOS:
        return None
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload.get("sub") or None
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None


def _get_usuario_autenticado(
    authorization: Optional[str] = Header(None),
    artesp_session: Optional[str] = Cookie(None),
) -> str:
    """Autentica por cookie httpOnly (navegador) ou Bearer (API/scripts)."""
    token = None
    if artesp_session:
        token = artesp_session
    if not token and authorization:
        scheme, _, credentials = (authorization or "").partition(" ")
        if scheme.lower() == "bearer" and credentials.strip():
            token = credentials.strip()

    if not token:
        raise HTTPException(status_code=401, detail="Não autenticado. Faça login.")

    if token in TOKENS_REVOGADOS:
        raise HTTPException(status_code=401, detail="Sessão revogada. Faça login novamente.")

    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        email = payload.get("sub", "")
        if not email:
            raise HTTPException(status_code=401, detail="Token inválido.")
        return email
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Sessão expirada. Faça login novamente.")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido.")


def _get_admin_emails() -> set:
    """Emails com role admin (env + users.json, excl. disabled)."""
    admins = set()
    raw = (os.getenv("ARTESP_ADMIN_EMAILS") or "").strip()
    if raw:
        for e in raw.split(","):
            e = e.strip().lower()
            if e:
                admins.add(e)
    admin_email = (os.getenv("ARTESP_WEB_ADMIN_EMAIL") or "").strip().lower()
    if admin_email:
        admins.add(admin_email)
    banco = carregar_banco_usuarios()
    for email, obj in (banco.items() if isinstance(banco, dict) else []):
        if isinstance(obj, dict) and (obj.get("role") or "").strip().lower() == "admin":
            if not obj.get("disabled"):
                admins.add((email or "").strip().lower())
    return admins


def _numero_admins(usuarios: dict) -> int:
    """Conta admins (env + dict) para garantir ao menos um ao rebaixar/bloquear/remover."""
    n = 0
    raw = (os.getenv("ARTESP_ADMIN_EMAILS") or "").strip()
    if raw:
        n += sum(1 for e in raw.split(",") if (e or "").strip().lower())
    if (os.getenv("ARTESP_WEB_ADMIN_EMAIL") or "").strip().lower():
        n += 1
    for _email, obj in (usuarios.items() if isinstance(usuarios, dict) else []):
        if isinstance(obj, dict) and (obj.get("role") or "").strip().lower() == "admin" and not obj.get("disabled"):
            n += 1
    return n


def _get_user_role(email: str) -> str:
    """'admin' ou 'user' conforme users.json / admins."""
    email = (email or "").strip().lower()
    banco = carregar_banco_usuarios()
    obj = banco.get(email)
    if isinstance(obj, dict) and isinstance(obj.get("role"), str):
        r = (obj.get("role") or "user").strip().lower()
        return r if r in ("admin", "user") else "user"
    admins = _get_admin_emails()
    return "admin" if email in admins else "user"


def _get_usuario_admin(usuario_email: str = Depends(_get_usuario_autenticado)):
    """Dependency: exige admin."""
    admins = _get_admin_emails()
    if not admins:
        raise HTTPException(status_code=403, detail="Nenhum admin configurado. Defina ARTESP_ADMIN_EMAILS ou ARTESP_WEB_ADMIN_EMAIL.")
    if (usuario_email or "").strip().lower() not in admins:
        raise HTTPException(status_code=403, detail="Acesso restrito a administradores.")
    return usuario_email


@app.get("/", status_code=200)
@app.head("/", status_code=200)
def read_root():
    """Página inicial — redireciona para /web/geojson. HEAD aceito para health check."""
    html = """<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<meta http-equiv="refresh" content="0;url=/web/geojson">
<title>Gerador ARTESP — GeoJSON / Relatórios</title>
<style>body{font-family:system-ui,sans-serif;background:#0f172a;color:#e2e8f0;min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;margin:0;padding:1rem;}
h1{font-size:1.25rem;margin:0 0 .5rem;} a{color:#3b82f6;} .muted{font-size:.875rem;color:#94a3b8;}</style></head>
<body>
<h1>Gerador ARTESP</h1>
<p class="muted">Projeto: Conservação e Obras — GeoJSON e relatórios</p>
<p>Redirecionando para o <a href="/web/geojson">Gerador GeoJSON</a>&hellip;</p>
</body></html>"""
    return HTMLResponse(html)


@app.get("/web")
def web_index():
    """Página inicial com os 3 sistemas."""
    p = WEB_DIR / "home.html"
    if not p.is_file():
        p = WEB_DIR / "index.html"
    if not p.is_file():
        raise HTTPException(status_code=404, detail="Interface web não encontrada.")
    return FileResponse(str(p), media_type="text/html")


@app.get("/web/geojson")
@app.get("/web/geojson/")
def web_geojson():
    """GeoJSON ARTESP — upload e geração de relatórios."""
    p = WEB_DIR / "index.html"
    if not p.is_file():
        # Fallback: redireciona para home em vez de 404 (ex.: deploy sem pasta web)
        return RedirectResponse(url="/web", status_code=302)
    return FileResponse(str(p), media_type="text/html")


@app.get("/web/home")
def web_home():
    p = WEB_DIR / "home.html"
    if not p.is_file():
        raise HTTPException(status_code=404, detail="home.html não encontrado.")
    return FileResponse(str(p), media_type="text/html")


@app.get("/web/fotos")
def web_fotos():
    p = WEB_DIR / "fotos.html"
    if not p.is_file():
        raise HTTPException(status_code=404, detail="fotos.html não encontrado.")
    return FileResponse(str(p), media_type="text/html")


@app.get("/web/nc")
def web_nc():
    p = WEB_DIR / "nc.html"
    if not p.is_file():
        raise HTTPException(status_code=404, detail="nc.html não encontrado.")
    return FileResponse(str(p), media_type="text/html")


@app.get("/web/admin")
def web_admin():
    """Página administrativa: cadastro de analistas (área restrita a administradores)."""
    admin_path = WEB_DIR / "admin.html"
    if not admin_path.is_file():
        raise HTTPException(status_code=404, detail="Página admin não encontrada.")
    return FileResponse(str(admin_path), media_type="text/html")


# Config para UI web (igual à aplicação local)
CONFIG_LOTES = [
    {"key": "13", "label": "Lote 13", "sigla": "L13"},
    {"key": "21", "label": "Lote 21", "sigla": "L21"},
    {"key": "26", "label": "Lote 26", "sigla": "L26"},
]
CONFIG_MODALIDADES = [
    {"key": "1", "label": "Conservação", "chave": "conserva"},
    {"key": "2", "label": "Obras", "chave": "obras"},
]
CONFIG_VERSOES = [
    {"key": "r0", "label": "R0 — Programação Anual", "tipo": "ANUAL"},
    {"key": "r1", "label": "R1 — Revisão 1", "tipo": "ANUAL"},
    {"key": "r2", "label": "R2 — Revisão 2", "tipo": "ANUAL"},
    {"key": "r3", "label": "R3 — Revisão 3", "tipo": "ANUAL"},
    {"key": "m", "label": "M — Programação Mensal", "tipo": "MENSAL"},
    {"key": "e", "label": "E — Executado", "tipo": "EXECUTADO"},
]
MESES_OPCOES = [
    (1, "01 - Janeiro"), (2, "02 - Fevereiro"), (3, "03 - Março"),
    (4, "04 - Abril"), (5, "05 - Maio"), (6, "06 - Junho"),
    (7, "07 - Julho"), (8, "08 - Agosto"), (9, "09 - Setembro"),
    (10, "10 - Outubro"), (11, "11 - Novembro"), (12, "12 - Dezembro"),
]


def _get_desenvolvedor_nome():
    # 1. Variável de ambiente (para Render, sem alterar o core)
    env_nome = (os.getenv("ARTESP_DESENVOLVEDOR_NOME") or "").strip()
    if env_nome:
        return env_nome
    # 2. Core (igual à aplicação local)
    try:
        if str(BASE_DIR) not in sys.path:
            sys.path.insert(0, str(BASE_DIR))
        import gerador_artesp_core as _core
        nome = (getattr(_core, "DESENVOLVEDOR_NOME", None) or "").strip()
        return nome if nome else ""
    except Exception:
        return ""


@app.get("/api/config")
def api_config():
    kartado_fd: list[str] = []
    kartado_todos: list[str] = []
    try:
        from nc_artesp.config import (
            KARTADO_RELATORIO_SERVICOS_FD,
            KARTADO_RELATORIO_SERVICOS_TODOS,
        )

        kartado_fd = list(KARTADO_RELATORIO_SERVICOS_FD)
        kartado_todos = list(KARTADO_RELATORIO_SERVICOS_TODOS)
    except Exception:
        pass
    return {
        "lotes": CONFIG_LOTES,
        "modalidades": CONFIG_MODALIDADES,
        "versoes": CONFIG_VERSOES,
        "meses": [{"valor": v, "rotulo": r} for v, r in MESES_OPCOES],
        "desenvolvedor_nome": _get_desenvolvedor_nome(),
        "kartado_relatorio_servicos_fd": kartado_fd,
        "kartado_relatorio_servicos_todos": kartado_todos,
    }


def _escrever_csv_pendencias_com_resumo(caminho: str, pendencias: list, pendencias_por_motivo: dict) -> None:
    """Grava CSV de pendências com primeira linha de resumo (diagnóstico: motivo → quantidade)."""
    with open(caminho, "w", encoding="utf-8") as f:
        if pendencias_por_motivo:
            resumo_str = "; ".join(
                f"{motivo}: {qtd}" for motivo, qtd in sorted(pendencias_por_motivo.items(), key=lambda x: -x[1])
            )
            f.write(f"# RESUMO_PENDENCIAS; {resumo_str}\n")
        pd.DataFrame(pendencias).to_csv(f, sep=";", index=False)


def _classificar_erro(detail: str) -> dict:
    """Classifica mensagem de erro em tipo e código para tratamento específico."""
    d = (detail or "").strip().lower()
    if "core não disponível" in d or "core nao disponivel" in d:
        return {"code": "CORE_INDISPONIVEL", "type": "config", "detail": detail}
    if "não encontrado" in d or "nao encontrado" in d or "não encontrada" in d:
        if "template" in d or "schema" in d or "malha" in d:
            return {"code": "ARQUIVO_AUSENTE", "type": "validação", "detail": detail}
    if "nenhuma linha válida" in d or "nenhuma feature gerada" in d:
        return {"code": "DADOS_VAZIOS", "type": "validação", "detail": detail}
    if "nao encontrada na malha" in d or "rodovia" in d and "malha" in d:
        return {"code": "RODOVIA_FORA_MALHA", "type": "malha", "detail": detail}
    if "geojson invalido" in d or "schema" in d:
        return {"code": "VALIDACAO_SCHEMA", "type": "validação", "detail": detail}
    if "assinatura" in d or "pfx" in d or "osslsigncode" in d or "signtool" in d:
        return {"code": "ASSINATURA_FALHOU", "type": "assinatura", "detail": detail}
    if "excel" in d or "xlsx" in d:
        return {"code": "EXCEL_INVALIDO", "type": "validação", "detail": detail}
    return {"code": "ERRO_INTERNO", "type": "outros", "detail": detail}


# ═══════════════════════════════════════════════════════════════
#  [P1] PERÍODO MENSAL — usa lógica do core (fonte única de verdade)
# ═══════════════════════════════════════════════════════════════
def _periodo_mensal_por_versao(versao_key: str) -> tuple[Optional[int], Optional[int]]:
    """
    Calcula ano/mês para relatórios mensais com base na versão.
    Versão 'M' (programado) = mês seguinte.
    Versão 'E' (executado)  = mês anterior.
    Retorna (ano, mes) ou (None, None) se não se aplica.

    [P1] Centralizado: mesma lógica do core, sem duplicação.
    Usa dateutil.relativedelta se disponível, senão cálculo manual seguro.
    """
    v = (versao_key or "").strip().lower()
    if v not in ("m", "e"):
        return None, None

    hoje = datetime.date.today()

    try:
        from dateutil.relativedelta import relativedelta
        if v == "m":
            # Programado: mês seguinte
            alvo = hoje + relativedelta(months=1)
        else:
            # Executado: mês anterior
            alvo = hoje - relativedelta(months=1)
        return alvo.year, alvo.month
    except ImportError:
        # Fallback sem dateutil — cálculo manual seguro
        if v == "m":
            # Programado: mês seguinte
            if hoje.month == 12:
                return hoje.year + 1, 1
            return hoje.year, hoje.month + 1
        else:
            # Executado: mês anterior
            if hoje.month == 1:
                return hoje.year - 1, 12
            return hoje.year, hoje.month - 1


def _linha_sobrepoe_mes_relatorio(
    row: Any,
    tipo: str,
    mes: Optional[int],
    ano: int,
    formatar_data_iso,
) -> tuple[bool, Optional[str]]:
    """
    MENSAL/EXECUTADO: só entra no GeoJSON se ``data_inicial``/``data_final`` existirem
    na planilha e o intervalo intersectar o mês alvo (ano/mes).
    Sem datas não se usa mais o atalho «todo o mês» — isso incluía todas as linhas como no anual.
    Retorna (incluir, motivo_pendencia). motivo None + incluir False = fora do mês (omitir sem pendência).
    """
    if tipo not in ("MENSAL", "EXECUTADO") or not mes:
        return True, None
    iso_i = formatar_data_iso(row.get("data_inicial"))
    iso_f = formatar_data_iso(row.get("data_final"))
    if not iso_i or not iso_f:
        return False, "mensal: preencha data_inicial e data_final na planilha"
    try:
        di = datetime.date.fromisoformat(iso_i)
        df_d = datetime.date.fromisoformat(iso_f)
    except ValueError:
        return False, "mensal: data_inicial ou data_final invalida"
    if df_d < di:
        return False, "mensal: data_final anterior a data_inicial"
    ms = datetime.date(ano, mes, 1)
    me = datetime.date(ano, mes, calendar.monthrange(ano, mes)[1])
    if df_d < ms or di > me:
        return False, None
    return True, None


def verificar_consistencia_dados(content: bytes, modalidade: str) -> dict:
    """
    Valida se o Excel é um "clone" do template oficial (Keys do template presentes e Values coerentes).
    modalidade: "1" = Conservação, "2" = Obras (ou "conserva"/"obras").
    Retorna status, criticos, alertas e resumo.
    """
    import gerador_artesp_core as core
    erros: list = []
    alertas: list = []

    if not content or len(content) == 0:
        return {"status": "erro", "criticos": ["Arquivo Excel vazio."], "alertas": [], "resumo": {}}

    # 1. Leitura padrão core: cabeçalho linha 1, dados a partir da linha 6 (skiprows 1–4)
    try:
        df = pd.read_excel(BytesIO(content), header=0, skiprows=range(1, 5))
        if df is None or df.empty:
            return {"status": "erro", "criticos": ["Planilha vazia ou sem dados."], "alertas": [], "resumo": {}}
        df = core.normalizar_colunas_df(df)
    except Exception as e:
        return {"status": "erro", "criticos": [f"Erro na leitura do arquivo: {str(e)}"], "alertas": [], "resumo": {}}

    # 2. Colunas obrigatórias (Keys do template) por modalidade
    colunas_base = ["rodovia", "km_inicial", "km_final", "item", "unidade", "local"]
    if modalidade in ("2", "obras"):
        colunas_base = colunas_base + ["programa", "subitem"]

    colunas_faltantes = [c for c in colunas_base if c not in df.columns]
    if colunas_faltantes:
        erros.append(f"Colunas obrigatórias ausentes para esta modalidade: {', '.join(colunas_faltantes)}")

    if erros:
        return {"status": "erro", "criticos": erros, "alertas": alertas, "resumo": {"total_linhas": len(df)}}

    # 3. Validação de conteúdo (Values): KM vazios
    vazios_km = df[df["km_inicial"].isna() | df["km_final"].isna()]
    if not vazios_km.empty:
        erros.append(f"Existem {len(vazios_km)} linhas com KM Inicial ou Final vazios.")

    try:
        ki_num = pd.to_numeric(df["km_inicial"], errors="coerce")
        kf_num = pd.to_numeric(df["km_final"], errors="coerce")
        ordem_km_invalida = (ki_num.notna()) & (kf_num.notna()) & (kf_num < ki_num)
        n_km = int(ordem_km_invalida.sum())
        if n_km:
            erros.append(f"Existem {n_km} linhas com KM Final menor que KM Inicial.")
    except Exception:
        pass

    if "data_inicial" in df.columns and "data_final" in df.columns:
        n_datas = 0
        for _, row in df.iterrows():
            iso_i = core._formatar_data_iso(row.get("data_inicial"))
            iso_f = core._formatar_data_iso(row.get("data_final"))
            if not iso_i or not iso_f:
                continue
            try:
                di = datetime.date.fromisoformat(iso_i)
                df_dt = datetime.date.fromisoformat(iso_f)
                if df_dt < di:
                    n_datas += 1
            except ValueError:
                continue
        if n_datas:
            erros.append(f"Existem {n_datas} linhas com Data Final anterior à Data Inicial.")

    codigos_rod = core.obter_codigos_rodovias_validos()
    if codigos_rod is not None:
        try:
            norm_rod = df["rodovia"].apply(lambda v: core.normalizar_rodovia(v))
            preench = df["rodovia"].notna() & (df["rodovia"].astype(str).str.strip() != "")
            fora = preench & (norm_rod != "") & ~norm_rod.isin(codigos_rod)
            n_rod = int(fora.sum())
            if n_rod:
                erros.append(
                    f"Existem {n_rod} linhas com código de rodovia que não consta em assets/malha/rodovias.xlsx "
                    "(lista oficial ARTESP)."
                )
        except Exception:
            pass

    # Latitude/Longitude: alerta se vazio, erro se texto/vírgula (core normaliza para "Latitude"/"Longitude")
    lat_col = "Latitude" if "Latitude" in df.columns else ("latitude" if "latitude" in df.columns else None)
    lon_col = "Longitude" if "Longitude" in df.columns else ("longitude" if "longitude" in df.columns else None)
    if lat_col and lon_col:
        for col in [lat_col, lon_col]:
            if df[col].apply(lambda x: isinstance(x, str)).any():
                erros.append(f"A coluna {col} contém texto ou vírgula. Use apenas números com ponto decimal.")
                break
        vazios_geo = df[df[lat_col].isna() | df[lon_col].isna()]
        if not vazios_geo.empty:
            alertas.append(f"{len(vazios_geo)} itens sem coordenadas geográficas (Latitude/Longitude).")

    if erros:
        return {"status": "erro", "criticos": erros, "alertas": alertas, "resumo": {"total_linhas": len(df)}}

    # 4. Alertas de engenharia: saltos de KM > 1 km
    try:
        df_ord = df.sort_values(by="km_inicial").copy()
        df_ord["gap"] = df_ord["km_inicial"].astype(float).diff().abs()
        saltos = df_ord[df_ord["gap"] > 1.0]
        if not saltos.empty:
            alertas.append("Detectados saltos de KM superiores a 1 km entre trechos consecutivos.")
    except Exception:
        pass

    resumo = {
        "total_linhas": len(df),
        "km_inicial": float(df["km_inicial"].min()),
        "km_final": float(df["km_final"].max()),
    }
    return {"status": "sucesso", "criticos": erros, "alertas": alertas, "resumo": resumo}


def _normalizar_nome_arquivo(nome):
    if not nome:
        stamp = _agora_brasilia().strftime("%Y%m%d_%H%M%S")
        return f"artesp_relatorio_{stamp}.geojson"
    limpo = re.sub(r"[^A-Za-z0-9_.-]+", "_", nome).strip("._")
    if not limpo:
        limpo = "artesp_relatorio.geojson"
    if not limpo.lower().endswith(".geojson"):
        limpo += ".geojson"
    return limpo


def _sha256_arquivo(path):
    """Calcula SHA-256 de um arquivo. Retorna hex string ou vazio se inexistente/erro."""
    if not path or not os.path.isfile(path):
        return ""
    try:
        h = hashlib.sha256()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(1024 * 1024), b""):
                h.update(chunk)
        return h.hexdigest()
    except (IOError, OSError):
        return ""


def _safe_output_path(nome_arquivo):
    nome = os.path.basename(nome_arquivo)
    full = OUTPUT_PATH / nome
    full_resolved = full.resolve()
    output_resolved = OUTPUT_PATH.resolve()
    if output_resolved not in full_resolved.parents and full_resolved != output_resolved:
        raise HTTPException(status_code=400, detail="Nome de arquivo inválido.")
    return full_resolved


# ═══════════════════════════════════════════════════════════════
#  [P1] GERAÇÃO DE PDF — relatório de processamento
# ═══════════════════════════════════════════════════════════════
_REPORTLAB_OK = False
_HRFlowable = None
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm, cm
    from reportlab.lib.colors import HexColor
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        Image as RLImage,
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    try:
        from reportlab.platypus.flowables import HRFlowable
        _HRFlowable = HRFlowable
    except ImportError:
        _HRFlowable = None
    _REPORTLAB_OK = True
except ImportError:
    logging.info("[P1] reportlab não disponível — PDF não será gerado. pip install reportlab")


def _hr_line(width="100%", thickness=1, color=None):
    """Linha horizontal: HRFlowable se disponível, senão tabela fina."""
    from reportlab.lib.colors import HexColor as _HexColor
    if color is None:
        color = _HexColor("#0a2540") if _REPORTLAB_OK else None
    if _HRFlowable is not None:
        return _HRFlowable(width=width, thickness=thickness, color=color)
    # Fallback: tabela 1px
    from reportlab.lib.units import mm
    t = Table([[""]], colWidths=[165 * mm], rowHeights=[thickness])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), color or _HexColor("#0a2540")),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    return t


def _gerar_pdf_relatorio(
    output_path: str,
    nome_arquivo: str,
    lote: str,
    modalidade: str,
    versao: str,
    ano: int,
    mes: int = None,
    usuario_email: str = "",
    n_linhas: int = 0,
    n_features: int = 0,
    n_pendencias: int = 0,
    sha256_geojson: str = "",
    sha256_excel: str = "",
    arquivos_gerados: list = None,
    erros_criticos: list = None,
    alertas: list = None,
) -> str:
    """
    Gera PDF de relatório de processamento.
    Retorna o caminho do PDF gerado, ou string vazia se reportlab não disponível.
    """
    if not _REPORTLAB_OK:
        logging.warning("[P1] reportlab ausente — PDF não gerado.")
        return ""

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib import colors as _rlab_colors
    _hex = _rlab_colors.HexColor

    pdf_nome = nome_arquivo.replace(".geojson", "") + "_RELATORIO.pdf"
    pdf_path = os.path.join(output_path, pdf_nome)
    agora = _agora_brasilia()

    try:
        doc = SimpleDocTemplate(
            pdf_path,
            pagesize=A4,
            topMargin=20 * mm,
            bottomMargin=15 * mm,
            leftMargin=15 * mm,
            rightMargin=15 * mm,
        )

        styles = getSampleStyleSheet()

        style_titulo = ParagraphStyle(
            "Titulo",
            parent=styles["Heading1"],
            fontSize=16,
            textColor=_hex("#0a2540"),
            spaceAfter=6 * mm,
            alignment=TA_CENTER,
        )
        style_subtitulo = ParagraphStyle(
            "Subtitulo",
            parent=styles["Heading2"],
            fontSize=12,
            textColor=_hex("#1a3a5c"),
            spaceBefore=4 * mm,
            spaceAfter=2 * mm,
        )
        style_normal = ParagraphStyle(
            "NormalCustom",
            parent=styles["Normal"],
            fontSize=9,
            leading=12,
        )
        style_small = ParagraphStyle(
            "SmallCustom",
            parent=styles["Normal"],
            fontSize=8,
            textColor=_hex("#666666"),
        )
        style_alerta = ParagraphStyle(
            "Alerta",
            parent=styles["Normal"],
            fontSize=9,
            textColor=_hex("#cc6600"),
        )
        style_erro = ParagraphStyle(
            "Erro",
            parent=styles["Normal"],
            fontSize=9,
            textColor=_hex("#cc0000"),
        )

        elements = []

        elements.append(Paragraph("RELATÓRIO DE GERAÇÃO — ARTESP", style_titulo))
        elements.append(_hr_line())
        elements.append(Spacer(1, 4 * mm))

        elements.append(Paragraph("1. Dados do Processamento", style_subtitulo))

        mes_label = f" / Mês {mes:02d}" if mes else ""
        dados_tabela = [
            ["Campo", "Valor"],
            ["Data/Hora", agora.strftime("%d/%m/%Y %H:%M:%S")],
            ["Usuário", usuario_email or "(não informado)"],
            ["Lote", lote],
            ["Modalidade", modalidade],
            ["Versão", versao],
            ["Ano" + (" / Mês" if mes else ""), f"{ano}{mes_label}"],
            ["Linhas no Excel", str(n_linhas)],
            ["Features (trechos)", str(n_features)],
            ["Pendências", str(n_pendencias)],
        ]

        t = Table(dados_tabela, colWidths=[45 * mm, 120 * mm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _hex("#0a2540")),
            ("TEXTCOLOR", (0, 0), (-1, 0), _hex("#ffffff")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (0, 0), (0, -1), "RIGHT"),
            ("ALIGN", (1, 0), (1, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.5, _hex("#cccccc")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_hex("#f8f9fa"), _hex("#ffffff")]),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 4 * mm))

        elements.append(Paragraph("2. Integridade dos Arquivos", style_subtitulo))
        integ_data = [["Arquivo", "SHA-256"]]
        if sha256_geojson:
            integ_data.append(["GeoJSON", sha256_geojson])
        if sha256_excel:
            integ_data.append(["Excel (entrada)", sha256_excel])

        if len(integ_data) > 1:
            t2 = Table(integ_data, colWidths=[35 * mm, 130 * mm])
            t2.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), _hex("#1a3a5c")),
                ("TEXTCOLOR", (0, 0), (-1, 0), _hex("#ffffff")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (1, 1), (1, -1), "Courier"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, _hex("#cccccc")),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]))
            elements.append(t2)
        else:
            elements.append(Paragraph("(hashes não disponíveis)", style_small))
        elements.append(Spacer(1, 4 * mm))

        elements.append(Paragraph("3. Arquivos Gerados", style_subtitulo))
        if arquivos_gerados:
            for i, arq in enumerate(arquivos_gerados, 1):
                elements.append(Paragraph(f"{i}. {arq}", style_normal))
        else:
            elements.append(Paragraph("(nenhum arquivo registrado)", style_small))
        elements.append(Spacer(1, 4 * mm))

        if n_pendencias > 0 or alertas:
            elements.append(Paragraph("4. Alertas e Pendências", style_subtitulo))
            if alertas:
                for a in (alertas or [])[:50]:
                    elements.append(Paragraph(f"⚠ {a}", style_alerta))
            if n_pendencias > 0:
                elements.append(Paragraph(
                    f"Total de pendências: {n_pendencias} — verifique PENDENCIAS_*.csv",
                    style_alerta,
                ))
            elements.append(Spacer(1, 4 * mm))

        if erros_criticos:
            elements.append(Paragraph("5. Erros Críticos", style_subtitulo))
            for err in erros_criticos[:30]:
                elements.append(Paragraph(f"✖ {err}", style_erro))
            elements.append(Spacer(1, 4 * mm))

        elements.append(_hr_line(width="100%", thickness=0.5, color=_hex("#cccccc")))
        elements.append(Spacer(1, 2 * mm))

        dev_nome = os.getenv("ARTESP_DEV_NOME", "").strip()
        dev_email = os.getenv("ARTESP_DEV_EMAIL", "").strip()
        dev_info = " | ".join(x for x in [dev_nome, dev_email] if x)

        rodape_text = f"Gerado automaticamente em {agora.strftime('%d/%m/%Y %H:%M:%S')}"
        if dev_info:
            rodape_text += f" — {dev_info}"

        elements.append(Paragraph(rodape_text, style_small))
        elements.append(Paragraph(
            "Este documento é parte do relatório de geração ARTESP e não substitui a validação oficial.",
            style_small,
        ))

        doc.build(elements)
        logging.info("[P1] PDF gerado: %s", pdf_nome)
        return pdf_nome

    except Exception as e:
        logging.error("[P1] Falha ao gerar PDF: %s", e)
        return ""


# ═══════════════════════════════════════════════════════════════
#  [P1] EXCEL PROTEGIDO — planilha de resumo com proteção de sheet
# ═══════════════════════════════════════════════════════════════
_OPENPYXL_OK = False
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except ImportError:
    logging.info("[P1] openpyxl não disponível — Excel protegido não será gerado. pip install openpyxl")


def _gerar_excel_resumo(
    output_path: str,
    nome_arquivo: str,
    lote: str,
    modalidade: str,
    versao: str,
    ano: int,
    mes: int = None,
    usuario_email: str = "",
    n_linhas: int = 0,
    n_features: int = 0,
    n_pendencias: int = 0,
    sha256_geojson: str = "",
    sha256_excel: str = "",
    arquivos_gerados: list = None,
) -> str:
    """
    Gera planilha Excel de resumo com proteção de sheet.
    Retorna nome do arquivo gerado ou string vazia.
    """
    if not _OPENPYXL_OK:
        logging.warning("[P1] openpyxl ausente — Excel de resumo não gerado.")
        return ""

    xlsx_nome = nome_arquivo.replace(".geojson", "") + "_RESUMO.xlsx"
    xlsx_path = os.path.join(output_path, xlsx_nome)
    agora = _agora_brasilia()

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resumo"

        header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="0A2540", end_color="0A2540", fill_type="solid")
        label_font = Font(name="Calibri", bold=True, size=10, color="1A3A5C")
        value_font = Font(name="Calibri", size=10)
        hash_font = Font(name="Consolas", size=9, color="555555")
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        ws.merge_cells("A1:B1")
        ws["A1"] = "RELATÓRIO DE GERAÇÃO — ARTESP"
        ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="0A2540")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:B2")
        ws["A2"] = f"Gerado em {agora.strftime('%d/%m/%Y %H:%M:%S')}"
        ws["A2"].font = Font(name="Calibri", size=9, color="888888")
        ws["A2"].alignment = Alignment(horizontal="center")

        row = 4
        ws.cell(row=row, column=1, value="Campo").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=2, value="Valor").font = header_font
        ws.cell(row=row, column=2).fill = header_fill

        mes_label = f" / Mês {mes:02d}" if mes else ""
        dados = [
            ("Usuário", usuario_email or "(não informado)"),
            ("Lote", lote),
            ("Modalidade", modalidade),
            ("Versão", versao),
            ("Ano" + (" / Mês" if mes else ""), f"{ano}{mes_label}"),
            ("Linhas no Excel", n_linhas),
            ("Features (trechos)", n_features),
            ("Pendências", n_pendencias),
        ]

        for i, (campo, valor) in enumerate(dados, start=1):
            r = row + i
            c1 = ws.cell(row=r, column=1, value=campo)
            c1.font = label_font
            c1.border = thin_border
            c1.alignment = Alignment(horizontal="right")

            c2 = ws.cell(row=r, column=2, value=valor)
            c2.font = value_font
            c2.border = thin_border

            if i % 2 == 0:
                fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                c1.fill = fill
                c2.fill = fill

        row_hash = row + len(dados) + 2
        ws.cell(row=row_hash, column=1, value="Integridade (SHA-256)").font = label_font
        if sha256_geojson:
            ws.cell(row=row_hash + 1, column=1, value="GeoJSON").font = label_font
            ws.cell(row=row_hash + 1, column=2, value=sha256_geojson).font = hash_font
        if sha256_excel:
            ws.cell(row=row_hash + 2, column=1, value="Excel (entrada)").font = label_font
            ws.cell(row=row_hash + 2, column=2, value=sha256_excel).font = hash_font

        row_arq = row_hash + 4
        ws.cell(row=row_arq, column=1, value="Arquivos Gerados").font = label_font
        if arquivos_gerados:
            for i, arq in enumerate(arquivos_gerados, 1):
                ws.cell(row=row_arq + i, column=1, value=f"{i}.")
                ws.cell(row=row_arq + i, column=2, value=arq).font = value_font

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 70

        senha_protecao = os.environ.get("ARTESP_EXCEL_SENHA", "").strip()
        ws.protection.sheet = True
        ws.protection.enable()
        if senha_protecao:
            ws.protection.set_password(senha_protecao)

        ws.protection.formatCells = False
        ws.protection.formatColumns = False
        ws.protection.formatRows = False
        ws.protection.insertColumns = False
        ws.protection.insertRows = False
        ws.protection.insertHyperlinks = False
        ws.protection.deleteColumns = False
        ws.protection.deleteRows = False
        ws.protection.selectLockedCells = True
        ws.protection.sort = False
        ws.protection.autoFilter = False
        ws.protection.pivotTables = False

        wb.save(xlsx_path)
        logging.info("[P1] Excel protegido gerado: %s", xlsx_nome)
        return xlsx_nome

    except Exception as e:
        logging.error("[P1] Falha ao gerar Excel protegido: %s", e)
        return ""


def _resolver_pfx_path():
    """
    Resolve o PFX por caminho (ARTESP_PFX) ou conteudo Base64 (ARTESP_PFX_CONTENT).
    Retorna: (pfx_path, is_temp)
    """
    pfx_path = (os.getenv("ARTESP_PFX") or "").strip()
    if pfx_path:
        if os.path.isfile(pfx_path):
            return pfx_path, False
        raise HTTPException(status_code=500, detail=f"ARTESP_PFX nao encontrado: {pfx_path}")

    pfx_b64 = (os.getenv("ARTESP_PFX_CONTENT") or "").strip()
    pfx_b64 = re.sub(r"\s+", "", pfx_b64)
    if not pfx_b64:
        raise HTTPException(status_code=500, detail="Configure ARTESP_PFX ou ARTESP_PFX_CONTENT.")

    try:
        pfx_data = base64.b64decode(pfx_b64, validate=True)
        temp_pfx = tempfile.NamedTemporaryFile(delete=False, suffix=".pfx")
        temp_pfx.write(pfx_data)
        temp_pfx.close()
        return temp_pfx.name, True
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Falha ao reconstruir PFX Base64: {e}")


def _validar_geojson_schema(geojson_obj, schema_key, lote=None):
    schema_file = SCHEMA_MAP[schema_key]
    schema_path = SCHEMA_PATH / schema_file
    if not schema_path.is_file():
        raise HTTPException(status_code=500, detail=f"Schema nao encontrado: {schema_path}")

    try:
        with schema_path.open("r", encoding="utf-8") as f:
            schema = json.load(f)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Falha ao carregar schema: {e}")

    # Mesmo ajuste usado no core: permite lote dinamico no enum
    if lote:
        for def_name in ("FeatureConservacao", "FeatureObras"):
            try:
                feat = schema.get("$defs", {}).get(def_name)
                if feat and "properties" in feat:
                    props = feat["properties"].get("properties", {})
                    if isinstance(props, dict) and "properties" in props:
                        lote_prop = props["properties"].get("lote", {})
                        if isinstance(lote_prop.get("enum"), list) and lote not in lote_prop["enum"]:
                            lote_prop["enum"].append(lote)
            except Exception:
                pass

    val_cls = jsonschema.validators.validator_for(schema)
    validator = val_cls(schema)
    erros = sorted(validator.iter_errors(geojson_obj), key=lambda e: list(e.path))

    if erros:
        detalhes = []
        for e in erros[:20]:
            detalhes.append(f"{e.message} (em {list(e.path)})")
        raise HTTPException(
            status_code=422,
            detail={
                "mensagem": "GeoJSON invalido no schema.",
                "schema": schema_file,
                "erros": detalhes,
                "total_erros": len(erros),
            },
        )

    return str(schema_path)


def _levantar_se_rodovia_fora_lista_official(geojson_obj):
    import gerador_artesp_core as core

    codigos = core.obter_codigos_rodovias_validos()
    if codigos is None:
        return
    invalid = []
    for ft in geojson_obj.get("features") or []:
        p = ft.get("properties") or {}
        r = core.normalizar_rodovia(p.get("rodovia"))
        if r and r not in codigos:
            invalid.append(r)
    if not invalid:
        return
    uniq = sorted(set(invalid))[:25]
    extra = len(set(invalid)) - len(uniq)
    suf = f" (+{extra} outras)" if extra > 0 else ""
    raise HTTPException(
        status_code=422,
        detail="Rodovia(s) fora da lista oficial (assets/malha/rodovias.xlsx): " + ", ".join(uniq) + suf,
    )


def assinar_geojson_api(caminho_arquivo, pfx_path, pfx_password):
    sistema = platform.system()
    if sistema == "Windows":
        comando = [
            "signtool",
            "sign",
            "/f",
            pfx_path,
            "/p",
            pfx_password,
            "/v",
            caminho_arquivo,
        ]
    else:
        # Linux (Render): usa osslsigncode com arquivo temporario de saida
        out_tmp = f"{caminho_arquivo}.signed"
        comando = [
            "osslsigncode",
            "sign",
            "-pkcs12",
            pfx_path,
            "-pass",
            pfx_password,
            "-in",
            caminho_arquivo,
            "-out",
            out_tmp,
        ]

    try:
        subprocess.run(comando, check=True, capture_output=True, text=True)
        if sistema != "Windows":
            os.replace(out_tmp, caminho_arquivo)
        return True
    except subprocess.CalledProcessError as e:
        stderr = (e.stderr or "").strip()
        print(f"Erro na assinatura ({sistema}): {stderr or e}")
        return False
    except FileNotFoundError:
        print(f"Ferramenta de assinatura nao encontrada para {sistema}.")
        return False
    finally:
        if sistema != "Windows":
            try:
                if os.path.exists(out_tmp):
                    os.unlink(out_tmp)
            except Exception:
                pass


def _obter_senha_usuario(email: str):
    """
    Retorna a senha (hash ou legado) do usuário, consultando users.json e depois USUARIOS_WEB.
    Para users.json: retorna obj["hashed_password"] se existir e não estiver disabled.
    """
    banco = carregar_banco_usuarios()
    obj = banco.get((email or "").strip().lower())
    if isinstance(obj, dict):
        if obj.get("disabled"):
            return None
        return obj.get("hashed_password")
    if obj is not None:
        return str(obj)
    return USUARIOS_WEB.get((email or "").strip().lower())


# ═══════════════════════════════════════════════════════════════
#  [P0] COOKIE HELPER — cria resposta de login com httpOnly cookie
# ═══════════════════════════════════════════════════════════════
def _is_production() -> bool:
    """Detecta se está rodando em ambiente de produção (Render, etc.)."""
    return bool(
        (os.getenv("RENDER") or "").strip()
        or (os.getenv("PRODUCTION") or "").strip()
        or (os.getenv("ARTESP_PRODUCTION") or "").strip()
    )


def _criar_resposta_login(email: str, token: str, role: str, must_change_password: bool = False) -> JSONResponse:
    """Cria JSONResponse com cookie httpOnly e access_token no body (compatibilidade frontend/testes)."""
    response = JSONResponse(content={
        "ok": True,
        "email": email,
        "role": role,
        "must_change_password": must_change_password,
        "expires_in": TOKEN_TTL_SECONDS,
        "access_token": token,
        "token_type": "bearer",
    })
    response.set_cookie(
        key="artesp_session",
        value=token,
        httponly=True,
        secure=_is_production(),
        samesite="lax",
        max_age=TOKEN_TTL_SECONDS,
        path="/",
    )
    return response


@app.post("/auth/login")
def auth_login(payload: LoginPayload, request: Request):
    """[P2] Rate limit por IP. Login com cookie httpOnly."""
    _check_rate_limit(request, "login", RATE_LIMIT_LOGIN_MAX, RATE_LIMIT_LOGIN_JANELA)
    email = (payload.email or "").strip().lower()
    senha = (payload.senha or payload.password or "").strip()
    if not email or not senha:
        raise HTTPException(status_code=401, detail="E-mail ou senha incorretos.")

    try:
        # O Python acessa o Value através da Key (e-mail)
        usuarios = carregar_banco_usuarios()  # Lê o arquivo /data/users.json
        user_data = usuarios.get(email)  # Tenta buscar a Key

        if user_data and isinstance(user_data, dict):
            if user_data.get("disabled"):
                raise HTTPException(status_code=401, detail="E-mail ou senha incorretos.")
            hashed = user_data.get("hashed_password")
            if hashed and verificar_senha(senha, str(hashed)):
                role = _get_user_role(email)
                token = _gerar_token(email)
                must_change = bool(user_data.get("must_change_password"))
                return _criar_resposta_login(email, token, role, must_change_password=must_change)
            # Se hash falhou, tenta senha do env (admin local)
            senha_env = USUARIOS_WEB.get(email)
            if senha_env and verificar_senha(senha, str(senha_env)):
                role = _get_user_role(email)
                token = _gerar_token(email)
                return _criar_resposta_login(email, token, role, must_change_password=False)
        # Fallback: usuários de env (USUARIOS_WEB)
        senha_cadastrada = _obter_senha_usuario(email)
        if senha_cadastrada and verificar_senha(senha, str(senha_cadastrada)):
            role = _get_user_role(email)
            token = _gerar_token(email)
            return _criar_resposta_login(email, token, role, must_change_password=False)
    except HTTPException:
        raise
    except Exception as e:
        logging.warning("Erro no login (%s): %s", type(e).__name__, e, exc_info=True)
        raise HTTPException(status_code=401, detail="E-mail ou senha incorretos.")

    # Usuário não encontrado no banco nem em USUARIOS_WEB
    logging.info(
        "Login falhou: e-mail não encontrado (path: %s, arquivo existe: %s)",
        USERS_JSON_PATH,
        USERS_JSON_PATH.is_file(),
    )
    if not USUARIOS_WEB and not carregar_banco_usuarios():
        raise HTTPException(
            status_code=500,
            detail="Usuarios web nao configurados. Defina ARTESP_WEB_USERS ou ARTESP_WEB_ADMIN_EMAIL/ARTESP_WEB_ADMIN_PASSWORD.",
        )
    raise HTTPException(status_code=401, detail="E-mail ou senha incorretos.")


@app.get("/auth/status-db")
def auth_status_db():
    """
    Diagnóstico pós-deploy: retorna se o arquivo de usuários existe e o caminho.
    Só retorna dados quando ARTESP_DEBUG_AUTH=1 (evita expor em produção).
    """
    if (os.getenv("ARTESP_DEBUG_AUTH") or "").strip() not in ("1", "true", "sim", "yes"):
        raise HTTPException(status_code=404, detail="Not found")
    usuarios = carregar_banco_usuarios()
    return {
        "users_path": str(USERS_JSON_PATH),
        "users_file_exists": USERS_JSON_PATH.is_file(),
        "users_count": len(usuarios) if isinstance(usuarios, dict) else 0,
    }


@app.post("/auth/logout")
def auth_logout(
    authorization: Optional[str] = Header(None),
    artesp_session: Optional[str] = Cookie(None),
):
    """[P0] Logout: revoga token e limpa cookie httpOnly."""
    token = None
    if artesp_session:
        token = artesp_session
    if not token and authorization:
        scheme, _, credentials = (authorization or "").partition(" ")
        if scheme.lower() == "bearer" and credentials.strip():
            token = credentials.strip()

    if token:
        TOKENS_REVOGADOS.add(token)

    response = JSONResponse(content={"status": "ok", "mensagem": "Logout realizado."})
    response.delete_cookie(
        key="artesp_session",
        path="/",
        httponly=True,
        secure=_is_production(),
        samesite="lax",
    )
    return response


@app.get("/auth/me")
def auth_me(email: str = Depends(_get_usuario_autenticado)):
    """[P0] Verifica sessão e retorna dados do usuário autenticado."""
    email_norm = (email or "").strip().lower()
    banco = carregar_banco_usuarios()
    obj = banco.get(email_norm) if isinstance(banco, dict) else None
    must_change = bool(isinstance(obj, dict) and obj.get("must_change_password"))
    can_change_password = isinstance(obj, dict)
    return {
        "ok": True,
        "email": email,
        "role": _get_user_role(email),
        "must_change_password": must_change,
        "can_change_password": can_change_password,
    }


@app.post("/auth/trocar-senha")
def auth_trocar_senha(
    payload: TrocarSenhaPayload,
    usuario_email: str = Depends(_get_usuario_autenticado),
):
    """
    Troca a senha do usuário autenticado. Exigido no primeiro acesso (must_change_password).
    """
    email = (usuario_email or "").strip().lower()
    senha_atual = (payload.senha_atual or "").strip()
    nova_senha = (payload.nova_senha or "").strip()
    if not senha_atual or not nova_senha:
        raise HTTPException(status_code=400, detail="Preencha senha atual e nova senha.")
    if len(nova_senha) < 6:
        raise HTTPException(status_code=400, detail="Nova senha deve ter no mínimo 6 caracteres.")
    if nova_senha == senha_atual:
        raise HTTPException(status_code=400, detail="A nova senha deve ser diferente da senha atual.")

    def _atualizar(usuarios):
        user_data = usuarios.get(email)
        if not isinstance(user_data, dict):
            raise HTTPException(status_code=400, detail="Usuário não encontrado no banco.")
        hashed = user_data.get("hashed_password")
        if not hashed or not verificar_senha(senha_atual, str(hashed)):
            raise HTTPException(status_code=401, detail="Senha atual incorreta.")
        usuarios[email] = {
            **user_data,
            "hashed_password": gerar_hash_senha(nova_senha),
            "must_change_password": False,
        }
        return usuarios

    _modificar_banco_usuarios(_atualizar)
    return {"message": "Senha alterada com sucesso. Use a nova senha no próximo login."}


@app.post("/geojson/upload")
async def geojson_upload(
    request: Request,
    usuario_email: str = Depends(_get_usuario_autenticado),
    file: UploadFile = File(..., description="Arquivo Excel .xlsx ou .xls"),
):
    """
    Upload simplificado: Excel -> GeoJSON.
    Usa lote L13, Conservação, R0, ano atual.
    Compatível com o botão 'Converter para GeoJSON' do index.html.
    """
    _check_rate_limit(request, f"geojson:{usuario_email}", RATE_LIMIT_GERAR_MAX, RATE_LIMIT_GERAR_JANELA)
    _validar_upload(file)
    ano_int = _agora_brasilia().year
    content = await _ler_upload_com_limite(file)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(content)
        tmp_path = tmp.name
    try:
        resultado, erro = _gerar_relatorio_do_excel(
            excel_path=tmp_path,
            lote_key="13",
            modalidade_key="1",
            versao_key="r0",
            ano=ano_int,
            mes=None,
            correcao_eixo=False,
            dashboard=False,
            assinar=True,
            usuario_email=usuario_email,
            progress_queue=None,
            dry_run=False,
        )
        if erro:
            raise HTTPException(status_code=400, detail=erro)
        nome_principal = resultado.get("principal", "")
        if not nome_principal:
            raise HTTPException(status_code=500, detail="Nenhum arquivo gerado.")
        sha_geojson = _sha256_arquivo(str(OUTPUT_PATH / nome_principal))
        return {
            "status": "sucesso",
            "download_url": f"/outputs/{nome_principal}",
            "arquivo_nome": nome_principal,
            "sha256_geojson": sha_geojson,
            "n_linhas_df": resultado.get("n_linhas_df"),
            "n_features": resultado.get("n_features"),
        }
    finally:
        if tmp_path and os.path.isfile(tmp_path):
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


@app.post("/simular")
async def rota_simular(
    request: Request,
    usuario_email: str = Depends(_get_usuario_autenticado),
    arquivo_excel: UploadFile = File(...),
    lote: str = Form(default=""),
    modalidade: str = Form(...),  # "1" = Conservação, "2" = Obras
):
    """Valida o Excel como clone do template. [P2] Rate limit + upload validado."""
    _check_rate_limit(request, f"simular:{usuario_email}", RATE_LIMIT_SIMULAR_MAX, RATE_LIMIT_SIMULAR_JANELA)
    # [P2] Validação e leitura com limite
    _validar_upload(arquivo_excel)
    content = await _ler_upload_com_limite(arquivo_excel)
    mod = (modalidade or "").strip() or "1"
    return verificar_consistencia_dados(content, modalidade=mod)


@app.post("/processar-relatorio")
def processar_relatorio(payload: ProcessarRelatorioPayload, usuario_email: str = Depends(_get_usuario_autenticado)):
    pfx_path = None
    pfx_is_temp = False
    assinatura_ok = False

    if not isinstance(payload.geojson, dict):
        raise HTTPException(status_code=400, detail="Campo 'geojson' deve ser um objeto JSON.")
    if payload.geojson.get("type") != "FeatureCollection":
        raise HTTPException(status_code=400, detail="GeoJSON deve ter type='FeatureCollection'.")

    geojson_obj = dict(payload.geojson)
    metadata = geojson_obj.setdefault("metadata", {})
    metadata.setdefault("schema_version", "R0")
    metadata.setdefault("data_geracao", _agora_brasilia().isoformat(timespec="seconds"))
    metadata.setdefault("gerador_api", "render_api")
    metadata.setdefault("usuario_email", usuario_email)

    schema_usado = _validar_geojson_schema(geojson_obj, payload.template_schema, payload.lote)
    _levantar_se_rodovia_fora_lista_official(geojson_obj)

    nome_arquivo = _normalizar_nome_arquivo(payload.nome_arquivo)
    caminho_saida = _safe_output_path(nome_arquivo)

    try:
        with caminho_saida.open("w", encoding="utf-8") as f:
            json.dump(geojson_obj, f, ensure_ascii=False, separators=(",", ":"), allow_nan=False)

        if payload.assinar:
            pfx_password = (os.getenv("ARTESP_PFX_PASSWORD") or "").strip()
            if not pfx_password:
                raise HTTPException(status_code=500, detail="Configure ARTESP_PFX_PASSWORD para assinatura.")
            pfx_path, pfx_is_temp = _resolver_pfx_path()
            assinatura_ok = assinar_geojson_api(str(caminho_saida), pfx_path, pfx_password)
            if not assinatura_ok:
                raise HTTPException(status_code=500, detail="Falha ao assinar arquivo no ambiente atual.")

        sha_geojson = _sha256_arquivo(caminho_saida)
        with caminho_saida.open("r", encoding="utf-8") as f:
            geojson_retorno = json.load(f)

        return {
            "status": "ok",
            "usuario_email": usuario_email,
            "schema": payload.template_schema,
            "schema_path": schema_usado,
            "validacao": "aprovado",
            "assinatura": {
                "solicitada": payload.assinar,
                "ok": assinatura_ok if payload.assinar else False,
                "ambiente": platform.system(),
            },
            "arquivo_saida": str(caminho_saida),
            "arquivo_nome": nome_arquivo,
            "download_url": f"/outputs/{nome_arquivo}",
            "sha256_geojson": sha_geojson,
            "geojson": geojson_retorno,
        }
    finally:
        if pfx_is_temp and pfx_path and os.path.isfile(pfx_path):
            try:
                os.unlink(pfx_path)
            except Exception:
                pass


def _extrair_pontos_interesse_geometria_api(geom, ki, kf):
    """Fallback: formato igual ao script mãe {coordinates, km, tipo}."""
    ki_f = float(ki) if ki is not None else 0.0
    kf_f = float(kf) if kf is not None else 0.0

    def snap(v):
        return round(float(v), 3) if v is not None else 0.0

    if not geom or not isinstance(geom, dict):
        return []
    gtype = geom.get("type")
    coords = geom.get("coordinates")
    if not coords:
        return []
    out = []
    if gtype == "Point" and isinstance(coords, (list, tuple)) and len(coords) >= 2:
        out.append({"coordinates": [float(coords[0]), float(coords[1])], "km": snap(ki_f), "tipo": "ponto"})
    elif gtype == "LineString" and isinstance(coords, list) and len(coords) >= 2:
        out.append({"coordinates": list(coords[0])[:2], "km": snap(ki_f), "tipo": "inicial"})
        out.append({"coordinates": list(coords[-1])[:2], "km": snap(kf_f), "tipo": "final"})
    elif gtype == "MultiLineString" and isinstance(coords, list) and coords:
        fl, ll = coords[0], coords[-1]
        if isinstance(fl, (list, tuple)) and fl and isinstance(fl[0], (list, tuple)) and len(fl[0]) >= 2:
            out.append({"coordinates": list(fl[0])[:2], "km": snap(ki_f), "tipo": "inicial"})
        if isinstance(ll, (list, tuple)) and ll and isinstance(ll[-1], (list, tuple)) and len(ll[-1]) >= 2:
            out.append({"coordinates": list(ll[-1])[:2], "km": snap(kf_f), "tipo": "final"})
    elif gtype == "MultiPoint":
        for i, c in enumerate(coords or []):
            if isinstance(c, (list, tuple)) and len(c) >= 2:
                km = snap(ki_f + (kf_f - ki_f) * (i / max(1, len(coords) - 1))) if len(coords) > 1 else snap(ki_f)
                out.append({"coordinates": [float(c[0]), float(c[1])], "km": km, "tipo": "ponto"})
    return out


_CORES_MARCADORES = {"inicial": "#1976D2", "final": "#388E3C", "ponto": "#D32F2F"}


def _expandir_features_com_marcadores_alfinete_api(features):
    """Fallback standalone: formato igual ao script mãe (pontos_interesse com coordinates, tipo, km)."""
    out = []
    for ft in features:
        out.append(ft)
        p = ft.get("properties") or {}
        if p.get("marcador_apenas"):
            continue
        pontos = p.get("pontos_interesse") or []
        fid = p.get("id") or ""
        for i, pt in enumerate(pontos):
            if isinstance(pt, dict):
                coords = pt.get("coordinates")
                tipo = pt.get("tipo", "ponto")
                km_val = pt.get("km")
            elif isinstance(pt, (list, tuple)) and len(pt) >= 2:
                coords = pt
                tipo = "ponto"
                km_val = None
            else:
                continue
            if not isinstance(coords, (list, tuple)) or len(coords) < 2:
                continue
            cor = _CORES_MARCADORES.get(tipo, _CORES_MARCADORES.get("ponto", "#D32F2F"))
            id_marcador = f"{fid}_m{tipo[:1]}{i}"[:50]
            props_marcador = dict(p)
            props_marcador.update({
                "id": id_marcador,
                "marcador_apenas": True,
                "tipo_marcador": tipo,
                "cor_hex": cor,
                "id_feature": fid,
                "km_inicial": km_val if km_val is not None else p.get("km_inicial"),
                "km_final": km_val if km_val is not None else p.get("km_final"),
            })
            out.append({
                "type": "Feature",
                "geometry": {"type": "Point", "coordinates": [float(coords[0]), float(coords[1])]},
                "properties": props_marcador
            })
    return out


def _gerar_relatorio_do_excel(
    excel_path: str,
    lote_key: str,
    modalidade_key: str,
    versao_key: str,
    ano: int,
    mes: Optional[int],
    correcao_eixo: bool,
    dashboard: bool,
    assinar: bool,
    usuario_email: str,
    progress_queue: Optional[Queue] = None,
    dry_run: bool = False,
) -> tuple[dict, Optional[str]]:
    """
    Usa gerador_artesp_core para gerar GeoJSON a partir do Excel.
    Retorna: (resultado_dict ou "", mensagem_erro)
    Se progress_queue for informado, emite {"msg": str, "pct": int} durante a execução.
    """
    def prog(msg: str, pct: int) -> None:
        if progress_queue is not None:
            try:
                progress_queue.put({"msg": msg, "pct": pct})
            except Exception:
                pass

    try:
        if str(BASE_DIR) not in sys.path:
            sys.path.insert(0, str(BASE_DIR))
        import gerador_artesp_core as core
    except ImportError as e:
        return "", f"Core não disponível: {e}"

    LOTES = core.LOTES
    MODALIDADES = core.MODALIDADES
    VERSOES_RELATORIO = core.VERSOES_RELATORIO
    MESES_NOME_COMPLETO = core.MESES_NOME_COMPLETO
    ler_excel = core.ler_excel
    normalizar_colunas_df = core.normalizar_colunas_df
    normalizar_rodovia = core.normalizar_rodovia
    normalizar_item = core.normalizar_item
    normalizar_unidade = core.normalizar_unidade
    processar_local = core.processar_local
    filtrar_dados_por_lote = core.filtrar_dados_por_lote
    carregar_malha = core.carregar_malha
    extrair_ponto_geom = core.extrair_ponto_geom
    extrair_linha_geom = core.extrair_linha_geom
    extrair_sentido = core.extrair_sentido
    _path_asset = core._path_asset
    _path_asset_eixo = core._path_asset_eixo
    _path_asset_template = core._path_asset_template
    basename_saida = core.basename_saida
    salvar_geojson = core.salvar_geojson
    _simplificar_coordenadas_core = getattr(core, "_simplificar_coordenadas", None)
    LINHAS_EXEMPLO_TEMPLATE = core.LINHAS_EXEMPLO_TEMPLATE
    CACHE = core.CACHE
    DADOS_MALHA = core.DADOS_MALHA
    MAPA_LOCAL_PARA_SENTIDO_MALHA = core.MAPA_LOCAL_PARA_SENTIDO_MALHA
    LOTE_SENTIDO_PARA_MALHA = getattr(core, "LOTE_SENTIDO_PARA_MALHA", {})
    RODOVIA_SENTIDO_PARA_MALHA = getattr(core, "RODOVIA_SENTIDO_PARA_MALHA", {})
    UNIDADES_VALIDAS = core.UNIDADES_VALIDAS
    _parse_km_excel = core._parse_km_excel
    _snap_km = core._snap_km
    _to_float = core._to_float
    _to_string_required = core._to_string_required
    _to_string_or_null = core._to_string_or_null
    _normalizar_geoms = core._normalizar_geoms
    _formatar_data_iso = core._formatar_data_iso
    _extrair_pontos_interesse_geometria = getattr(
        core, "_extrair_pontos_interesse_geometria", _extrair_pontos_interesse_geometria_api
    )
    _expandir_features_com_marcadores_alfinete = getattr(
        core, "_expandir_features_com_marcadores_alfinete", _expandir_features_com_marcadores_alfinete_api
    )
    gerar_id = core.gerar_id
    validar_json = core.validar_json
    gerar_pdf_relatorio = getattr(core, "gerar_pdf_relatorio", None)
    gerar_dashboard_artesp = getattr(core, "gerar_dashboard_artesp", None)
    gerar_resumo_rodovia_sentido = getattr(core, "gerar_resumo_rodovia_sentido", None)
    _formatar_km_relatorio = getattr(core, "_formatar_km_relatorio", lambda x: str(x) if x is not None else "")
    _formatar_data_saida_dma = getattr(core, "_formatar_data_saida_dma", lambda x: str(x) if x is not None else "")
    LINHA_INICIO_DADOS = getattr(core, "LINHA_INICIO_DADOS", 6)
    sha256_arquivo_core = getattr(core, "sha256_arquivo", _sha256_arquivo)

    lote = LOTES.get(lote_key)
    mod = MODALIDADES.get(modalidade_key)
    vinfo = VERSOES_RELATORIO.get(versao_key, VERSOES_RELATORIO["r0"])
    tipo = vinfo.get("tipo", "ANUAL")
    is_obras = mod["chave"] == "obras"

    base = str(OUTPUT_PATH)

    TEMPLATE = _path_asset_template(lote["sigla"], mod["chave"], ano, versao_key)
    SCHEMA = _path_asset(*mod["schema_asset"])
    EIXO = _path_asset_eixo(*lote["eixo"])
    for nome, caminho in [("Template", TEMPLATE), ("Schema", SCHEMA), ("Malha", EIXO)]:
        if not os.path.exists(str(caminho)):
            return "", f"{nome} não encontrado: {caminho}"

    setattr(core, "CORRIGIR_EIXO", correcao_eixo)

    prog("Lendo Excel...", 5)
    df, _, _ = ler_excel(excel_path, sheet=0)
    df = normalizar_colunas_df(df)
    set_exemplo = set((l, r, i) for l, r, i in LINHAS_EXEMPLO_TEMPLATE)
    mask = df.apply(
        lambda row: (
            str(row.get("lote", "")).strip(),
            normalizar_rodovia(row.get("rodovia")),
            normalizar_item(row.get("item")),
        )
        not in set_exemplo,
        axis=1,
    )
    df = df[mask].copy()
    if "lote" not in df.columns:
        return "", "Coluna 'lote' não encontrada na planilha. Inclua a coluna 'lote' com valores L13, L21 ou L26 para que as linhas sejam filtradas pelo lote selecionado (evitando milhares de pendências por rodovia fora da malha)."
    df = filtrar_dados_por_lote(df, lote["sigla"])
    if df.empty:
        return "", "Nenhuma linha válida encontrada para este lote."

    prog(f"Excel lido: {len(df)} linhas válidas para {lote['sigla']}", 8)
    prog("Carregando malha de eixo...", 10)
    CACHE.limpar()
    try:
        DADOS_MALHA.clear()
    except Exception:
        pass
    carregar_malha(EIXO, lote["sigla"])
    if not CACHE.dados:
        return "", f"Malha do lote {lote['sigla']} não carregou ou está vazia. Verifique se o arquivo de eixo existe: {EIXO}"

    usar_geom_por_rodovia = getattr(core, "LOTE_GEOMETRIA_POR_RODOVIA_SO", {}).get(lote["sigla"], False)
    usar_alfinete = getattr(core, "LOTE_USAR_ALFINETE_QUANDO_KM_IGUAL", True)
    tol_alf = getattr(core, "TOLERANCIA_KM_ALFINETE", 0.01)

    prog("Gerando features...", 15)
    features = []
    pendencias = []
    feat_seq = 0
    total_linhas = len(df)
    codigos_oficiais = core.obter_codigos_rodovias_validos()
    # Pré-processar: índice e total de itens por grupo (rodovia, km_inicial, km_final).
    # Usado para distribuir alfinetes em posições km distintas dentro do trecho.
    _grp = df.groupby(["rodovia", "km_inicial", "km_final"], sort=False)
    df = df.copy()
    df["_item_idx"] = _grp.cumcount()           # 0, 1, 2, ... por grupo
    df["_item_n"]   = _grp["rodovia"].transform("count")  # total no grupo
    for seq, (idx, row) in enumerate(df.iterrows(), start=1):
        rod = normalizar_rodovia(row.get("rodovia"))
        if not rod:
            pendencias.append({"seq": seq, "rod": row.get("rodovia"), "motivo": "rodovia vazia"})
            continue
        if codigos_oficiais is not None and rod not in codigos_oficiais:
            pendencias.append(
                {"seq": seq, "rod": rod, "motivo": "rodovia nao consta em rodovias.xlsx (lista oficial ARTESP)"}
            )
            continue
        if CACHE.tem_sentidos and not usar_geom_por_rodovia:
            if not CACHE.sentidos_disponiveis(rod):
                pendencias.append({"seq": seq, "rod": rod, "motivo": f"rodovia {rod} nao encontrada na malha"})
                continue
        elif not CACHE.contem(rod):
            pendencias.append({"seq": seq, "rod": rod, "motivo": f"rodovia {rod} nao encontrada na malha"})
            continue

        ki = _parse_km_excel(row.get("km_inicial"))
        kf = _parse_km_excel(row.get("km_final"))
        if ki is None or kf is None:
            pendencias.append({"seq": seq, "rod": rod, "motivo": "km invalido"})
            continue

        unid_raw = row.get("unidade")
        unid = normalizar_unidade(unid_raw)
        if not unid or unid not in UNIDADES_VALIDAS:
            pendencias.append({"seq": seq, "rod": rod, "motivo": f"unidade: {unid_raw}"})
            continue

        locais_raw = row.get("local", [])
        locais = processar_local(locais_raw) if locais_raw is not None else None
        if not locais:
            pendencias.append({"seq": seq, "rod": rod, "motivo": "local vazio"})
            continue

        if is_obras:
            programa_raw = str(row.get("programa", "") or "").strip().upper()
            if programa_raw not in ("REVIT", "CAPEX", "NS"):
                pendencias.append({"seq": seq, "rod": rod, "motivo": f"obras: programa invalido: {row.get('programa')}"})
                continue
            try:
                item_int = int(float(row.get("item") or 0))
            except (TypeError, ValueError):
                pendencias.append({"seq": seq, "rod": rod, "motivo": f"obras: item deve ser inteiro: {row.get('item')}"})
                continue
            try:
                int(float(row.get("subitem") or 0))
            except (TypeError, ValueError):
                pendencias.append({"seq": seq, "rod": rod, "motivo": f"obras: subitem deve ser inteiro: {row.get('subitem')}"})
                continue
            item_norm = str(item_int)
        else:
            item_norm = normalizar_item(row.get("item"))
            if not re.match(r"^[a-z](\.\d+)+$", item_norm or ""):
                # Conservação: aceitar item numérico (1, 2, 3) e converter para a.1, a.2 (mesma planilha que obras)
                raw_item = row.get("item")
                try:
                    n = int(float(raw_item or 0))
                    if n > 0:
                        item_norm = f"a.{n}"
                except (TypeError, ValueError):
                    pass
                if not re.match(r"^[a-z](\.\d+)+$", item_norm or ""):
                    pendencias.append({"seq": seq, "rod": rod, "motivo": f"item invalido: {row.get('item')} (conservacao: use ex. a.1.2 ou numero 1, 2, 3)"})
                    continue

        inc_mes, pend_mes = _linha_sobrepoe_mes_relatorio(row, tipo, mes, ano, _formatar_data_iso)
        if not inc_mes:
            if pend_mes:
                pendencias.append({"seq": seq, "rod": rod, "motivo": pend_mes})
            continue

        lote_sigla = lote.get("sigla", "")

        def _sentido_malha_para_local(loc: str):
            if not CACHE.tem_sentidos or not loc:
                return None
            if loc not in MAPA_LOCAL_PARA_SENTIDO_MALHA:
                return None
            s_display = MAPA_LOCAL_PARA_SENTIDO_MALHA[loc]
            s_rod = (RODOVIA_SENTIDO_PARA_MALHA.get(rod) or {}).get(s_display)
            s_lote = (LOTE_SENTIDO_PARA_MALHA.get(lote_sigla) or {}).get(s_display) or s_display
            for s_try in (s_rod, s_lote):
                if s_try is not None and CACHE.contem(rod, s_try, loc):
                    return s_try
            for s_try in CACHE.sentidos_disponiveis(rod) or []:
                if CACHE.contem(rod, s_try, loc):
                    return s_try
            return None

        detalhe = _to_string_required(
            row.get("detalhamento_servico"),
            "Obra" if is_obras else "Servico de conservacao",
        )
        if mes:
            d1 = f"{ano}-{mes:02d}-01"
            dn = f"{ano}-{mes:02d}-{calendar.monthrange(ano, mes)[1]:02d}"
            if tipo in ("MENSAL", "EXECUTADO"):
                data_ini = _formatar_data_iso(row.get("data_inicial"))
                data_fin = _formatar_data_iso(row.get("data_final"))
            else:
                data_ini = _formatar_data_iso(row.get("data_inicial")) or d1
                data_fin = _formatar_data_iso(row.get("data_final")) or dn
        else:
            data_ini = _formatar_data_iso(row.get("data_inicial")) or f"{ano}-01-01"
            data_fin = _formatar_data_iso(row.get("data_final")) or f"{ano}-12-31"
        obs = _to_string_or_null(row.get("observacoes_gerais"))

        # Cor por categoria de item (reconhecida por geojson.io, GitHub, QGIS data-driven)
        _ITEM_CORES = {
            "a": "#1565C0",  # azul     — pavimento
            "b": "#2E7D32",  # verde    — conservação geral
            "c": "#F57F17",  # laranja  — drenagem/OAE
            "e": "#6A1B9A",  # roxo     — sinalização
            "f": "#B71C1C",  # vermelho — emergencial
        }
        cor_item = _ITEM_CORES.get((item_norm or "x")[0], "#546E7A")

        # Alfinete no km distribuído dentro do trecho — posição distinta por item.
        # Divide [ki, kf] em n_itens fatias; cada item ocupa o centro da sua fatia.
        # Ex.: 34 itens em km 0-32 → espaçamento ~0,94 km; pontos visualmente distintos.
        _item_idx = int(row.get("_item_idx", 0))
        _item_n   = max(int(row.get("_item_n",   1)), 1)
        _step = (kf - ki) / _item_n
        km_alf = ki + _step * _item_idx + _step / 2   # centro da fatia deste item
        km_alf = round(min(kf - 1e-6, max(ki + 1e-6, km_alf)), 6)
        trechos_geom: list[tuple[dict, list]] = []
        if usar_geom_por_rodovia:
            pt = extrair_ponto_geom(rod, km_alf, sentido=None, local=None)
            if pt:
                trechos_geom.append(({"type": "Point", "coordinates": pt}, list(locais)))
        else:
            grupos = {}
            for loc in locais:
                s_g = _sentido_malha_para_local(loc) if CACHE.tem_sentidos else None
                pt = extrair_ponto_geom(rod, km_alf, sentido=s_g, local=loc)
                if not pt:
                    pt = extrair_ponto_geom(rod, km_alf, sentido=s_g, local=None)
                if not pt:
                    continue
                key = (round(float(pt[0]), 5), round(float(pt[1]), 5))
                if key not in grupos:
                    grupos[key] = ([], [float(pt[0]), float(pt[1])])
                grupos[key][0].append(loc)
            for locs_sub, coords in grupos.values():
                trechos_geom.append(({"type": "Point", "coordinates": coords}, locs_sub))

        if not trechos_geom:
            ki_s = _formatar_km_relatorio(ki) or (f"{ki:.3f}" if ki is not None else "?")
            kf_s = _formatar_km_relatorio(kf) or (f"{kf:.3f}" if kf is not None else "?")
            pendencias.append({"seq": seq, "rod": rod, "motivo": f"sem geometria para {rod} km {ki_s}-{kf_s}"})
            continue

        for geom_final, locais_props in trechos_geom:
            sentido_str = extrair_sentido(locais_props) if locais_props else "-"
            feat_seq += 1
            fid = gerar_id(lote["sigla"], rod, item_norm, ki, kf, sentido_str, feat_seq)
            _props_base = {
                "id": fid,
                "lote": lote["sigla"],
                "sentido": sentido_str,
                "rodovia": rod,
                "detalhamento_servico": detalhe,
                "unidade": unid,
                "quantidade": _snap_km(_to_float(row.get("quantidade"))),
                "km_inicial": _snap_km(ki),
                "km_final": _snap_km(kf),
                "local": locais_props,
                "data_inicial": data_ini,
                "data_final": data_fin,
                "observacoes_gerais": obs,
                "marker-color": cor_item,
                "marker-size": "medium",
                "marker-symbol": "",
                "stroke": cor_item,
                "stroke-width": 2,
                "fill": cor_item,
                "fill-opacity": 0.8,
            }
            if is_obras:
                props = {
                    **_props_base,
                    "programa": str(row.get("programa", "") or "").strip().upper(),
                    "item": int(float(row.get("item") or 0)),
                    "subitem": int(float(row.get("subitem") or 0)),
                }
            else:
                props = {**_props_base, "item": item_norm}
            features.append({"type": "Feature", "geometry": geom_final, "properties": props})
        if total_linhas and seq % max(5, total_linhas // 8) == 0:
            pct = min(69, 15 + int(54 * seq / total_linhas))
            prog(f"Gerando features... {feat_seq} trechos de {total_linhas} linhas", pct)

    prog("Salvando GeoJSON...", 70)
    # Breakdown de pendências por motivo (diagnóstico: evita "3239 pendências" sem saber o porquê)
    pendencias_por_motivo = defaultdict(int)
    for p in pendencias:
        m = (p.get("motivo") or "sem_motivo")
        pendencias_por_motivo[m] += 1

    if dry_run:
        pend_path = OUTPUT_PATH / f"PENDENCIAS_{lote['sigla']}.csv"
        try:
            _escrever_csv_pendencias_com_resumo(str(pend_path), pendencias, dict(pendencias_por_motivo))
            nome_csv = pend_path.name
        except Exception:
            nome_csv = ""
        return {
            "principal": "",
            "arquivos": [nome_csv] if nome_csv else [],
            "n_linhas_df": len(df),
            "n_features": 0,
            "n_pendencias": len(pendencias),
            "pendencias_por_motivo": dict(pendencias_por_motivo),
            "dry_run": True,
            "csv_url": f"/outputs/{nome_csv}" if nome_csv else None,
        }, None

    if not features:
        resumo = "; ".join(f"{motivo}: {qtd}" for motivo, qtd in sorted(pendencias_por_motivo.items(), key=lambda x: -x[1])[:5])
        # Diagnóstico: o que a malha carregada contém (rodovias e extensão em km)
        try:
            resumo_malha = getattr(core.CACHE, "resumo_rodovias_km", lambda: [])()
            if resumo_malha:
                txt_malha = "; ".join(f"{r['rodovia']} km {r['km_min']:.3f}-{r['km_max']:.3f}" for r in resumo_malha[:15])
                if len(resumo_malha) > 15:
                    txt_malha += f" (+{len(resumo_malha) - 15} rodovias)"
                return "", (
                    f"Nenhuma feature gerada (todas as linhas em pendência). Principais motivos: {resumo}. "
                    f"Malha do lote {lote['sigla']} contém: {txt_malha}. "
                    "Verifique: lote correto, coluna lote preenchida com L13/L21/L26, arquivo de eixo na pasta assets/malha, "
                    "rodovia/km da planilha dentro do intervalo da malha, local/item/unidade válidos."
                )
        except Exception:
            pass
        return "", (
            f"Nenhuma feature gerada (todas as linhas em pendência). Principais motivos: {resumo}. "
            "Verifique: lote correto, coluna lote preenchida, malha do lote carregada (arquivo de eixo), rodovia/km/local/item/unidade válidos."
        )

    # Alfinetes (marcadores) desativados: não expandir pontos_interesse em features Point extras
    # features_principal = apenas trechos (linhas/pontos de trecho), sem features de alfinete
    # Ordenar por categoria de item: menos comuns primeiro (f, e, c, a) → mais comuns por cima (b).
    # O mapa renderiza em sequência — o último sobrepõe; 'b' (conservação geral) domina visualmente.
    _ORDEM_CATEGORIA = {'f': 0, 'e': 1, 'c': 2, 'a': 3, 'b': 4}
    features_principal = sorted(
        features,
        key=lambda ft: _ORDEM_CATEGORIA.get(
            str(ft.get('properties', {}).get('item') or 'b')[0], 5
        )
    )
    prog(f"Features geradas: {len(features_principal)} trechos de {len(df)} linhas", 70)

    geojson_obj = {
        "type": "FeatureCollection",
        "metadata": {
            "schema_version": "R0",
            "data_geracao": _agora_brasilia().strftime("%Y-%m-%dT%H:%M:%S-03:00"),
            "lote": lote["sigla"],
            "gerador_versao": getattr(core, "VERSAO", "3.8.3"),
            "geometria_por_sentido": not usar_geom_por_rodovia,
            "correcao_eixo_aplicada": correcao_eixo,
        },
        "crs": {"type": "name", "properties": {"name": "urn:ogc:def:crs:EPSG::4674"}},
        "features": features_principal,
    }

    # Normalizar properties.local para o schema (MARGINAL NORTE → MARGINAL_NORTE).
    # Malha por sentido/local: agrupa na mesma feature os locais que caem no mesmo ponto; várias features se houver pontos distintos.
    for f in geojson_obj["features"]:
        p = f.get("properties") or {}
        loc_val = p.get("local")
        if isinstance(loc_val, list):
            p["local"] = [str(x).strip().upper().replace(" ", "_") if x else x for x in loc_val]
        elif isinstance(loc_val, str):
            p["local"] = [loc_val.strip().upper().replace(" ", "_")]

    # Reduzir tamanho do GeoJSON (step=2, 4 decimais) antes de salvar — garante efeito mesmo no render
    _STEP_REDUCAO = 2
    _DECIMAIS_REDUCAO = 4
    geojson_para_salvar = geojson_obj
    if _simplificar_coordenadas_core and geojson_obj.get("features"):

        def _round_coords(coord, dec=_DECIMAIS_REDUCAO):
            if isinstance(coord, (list, tuple)):
                return [_round_coords(c, dec) for c in coord]
            if isinstance(coord, (int, float)):
                if -180 <= coord <= 180:
                    return round(coord, dec)
                return round(coord, 6) if abs(coord) > 20 else round(coord, 3)
            return coord

        features_reduzidas = []
        for ft in geojson_obj["features"]:
            geom = ft.get("geometry")
            if isinstance(geom, dict) and "coordinates" in geom:
                coords = geom["coordinates"]
                coords = _simplificar_coordenadas_core(coords, _STEP_REDUCAO)
                coords = _round_coords(coords)
                geom = {**geom, "coordinates": coords}
            features_reduzidas.append({
                "type": "Feature",
                "geometry": geom,
                "properties": dict(ft.get("properties") or {}),
            })
        geojson_para_salvar = {
            "type": "FeatureCollection",
            "metadata": dict(geojson_obj.get("metadata") or {}),
            "crs": dict(geojson_obj.get("crs") or {}),
            "features": features_reduzidas,
        }

    nome_base = basename_saida(lote["sigla"], mod["chave"], ano, versao_key, mes, tipo)
    nome_arquivo = f"{nome_base}.geojson"
    caminho_saida = OUTPUT_PATH / nome_arquivo
    # Forçar step=1 no core para não reduzir de novo (já reduzido acima)
    _env_step = os.environ.pop("ARTESP_GEOJSON_SIMPLIFY_STEP", None)
    os.environ["ARTESP_GEOJSON_SIMPLIFY_STEP"] = "1"
    try:
        salvar_geojson(str(caminho_saida), geojson_para_salvar)
    finally:
        os.environ.pop("ARTESP_GEOJSON_SIMPLIFY_STEP", None)
        if _env_step is not None:
            os.environ["ARTESP_GEOJSON_SIMPLIFY_STEP"] = _env_step
    prog("Validando schema...", 75)
    val = validar_json(geojson_para_salvar, str(SCHEMA), lote["sigla"])
    if not (val and val.get("ok")):
        return "", f"GeoJSON invalido no schema: {val.get('msg', 'Erro de validacao')}"

    assinatura_ok = False
    if assinar:
        pfx_password = (os.getenv("ARTESP_PFX_PASSWORD") or "").strip()
        if pfx_password:
            try:
                pfx_path, pfx_is_temp = _resolver_pfx_path()
                assinatura_ok = assinar_geojson_api(str(caminho_saida), pfx_path, pfx_password)
                if pfx_is_temp and pfx_path and os.path.isfile(pfx_path):
                    os.unlink(pfx_path)
            except Exception:
                pass

    prog("Exportando Excel...", 80)
    arquivos = [nome_arquivo]
    if pendencias:
        try:
            pend_path = OUTPUT_PATH / f"PENDENCIAS_{lote['sigla']}.csv"
            _escrever_csv_pendencias_com_resumo(str(pend_path), pendencias, dict(pendencias_por_motivo))
            arquivos.append(f"PENDENCIAS_{lote['sigla']}.csv")
        except Exception:
            pass
    n_marcadores = 0  # Alfinetes desativados
    resumo = (gerar_resumo_rodovia_sentido or (lambda x: {}))(features_principal)
    sha_geo = sha256_arquivo_core(str(caminho_saida))
    sha_sch = sha256_arquivo_core(str(SCHEMA))

    prog("Gerando PDF...", 85)
    try:
        import openpyxl
        xls_path = OUTPUT_PATH / f"{nome_base}.xlsx"
        if os.path.exists(str(TEMPLATE)):
            wb = openpyxl.load_workbook(str(TEMPLATE))
            ws = wb.active
            for num, ft in enumerate(features_principal, start=1):
                r = LINHA_INICIO_DADOS + num - 1
                props = ft.get("properties", {})
                loc_str = "; ".join(props.get("local", []))
                ws.cell(r, 1).value = num
                ws.cell(r, 2).value = props.get("lote")
                ws.cell(r, 3).value = props.get("rodovia")
                if is_obras:
                    ws.cell(r, 4).value = props.get("programa")
                    ws.cell(r, 5).value = props.get("item")
                    ws.cell(r, 6).value = props.get("subitem")
                    ws.cell(r, 7).value = props.get("detalhamento_servico")
                    ws.cell(r, 8).value = props.get("unidade")
                    ws.cell(r, 9).value = props.get("quantidade")
                    ws.cell(r, 10).value = _formatar_km_relatorio(props.get("km_inicial"))
                    ws.cell(r, 11).value = _formatar_km_relatorio(props.get("km_final"))
                    ws.cell(r, 12).value = loc_str
                    ws.cell(r, 13).value = _formatar_data_saida_dma(props.get("data_inicial"))
                    ws.cell(r, 14).value = _formatar_data_saida_dma(props.get("data_final"))
                    ws.cell(r, 15).value = props.get("observacoes_gerais")
                else:
                    ws.cell(r, 4).value = props.get("item")
                    ws.cell(r, 5).value = props.get("detalhamento_servico")
                    ws.cell(r, 6).value = props.get("unidade")
                    ws.cell(r, 7).value = props.get("quantidade")
                    ws.cell(r, 8).value = _formatar_km_relatorio(props.get("km_inicial"))
                    ws.cell(r, 9).value = _formatar_km_relatorio(props.get("km_final"))
                    ws.cell(r, 10).value = loc_str
                    ws.cell(r, 11).value = _formatar_data_saida_dma(props.get("data_inicial"))
                    ws.cell(r, 12).value = _formatar_data_saida_dma(props.get("data_final"))
                    ws.cell(r, 13).value = props.get("observacoes_gerais")
            wb.save(str(xls_path))
            arquivos.append(f"{nome_base}.xlsx")
    except Exception:
        pass

    sha_xls = sha256_arquivo_core(str(OUTPUT_PATH / f"{nome_base}.xlsx")) if (OUTPUT_PATH / f"{nome_base}.xlsx").exists() else "N/A"

    if gerar_pdf_relatorio:
        try:
            pdf_path = OUTPUT_PATH / f"{nome_base}.pdf"
            log_data = {
                "versao": getattr(core, "VERSAO", "3.8.3"),
                "versao_key": versao_key,
                "lote_sigla": lote["sigla"],
                "modalidade_rotulo": mod.get("rotulo", mod["chave"]),
                "n_linhas_df": len(df),
                "n_features": len(features_principal),
                "n_marcadores": n_marcadores,
                "n_pendencias": len(pendencias),
                "resumo": resumo,
                "sha_schema": sha_sch,
                "sha_geojson": sha_geo,
                "sha_excel": sha_xls,
                "validacao_resultado": val,
                "correcao_eixo": correcao_eixo,
            }
            gerar_pdf_relatorio(
                features_principal, lote, str(pdf_path), resumo,
                validacao_resultado=val,
                relatorio_log_data=log_data,
            )
            arquivos.append(f"{nome_base}.pdf")
        except Exception:
            pass

    log_path = OUTPUT_PATH / f"{nome_base}_LOG.txt"
    try:
        with open(log_path, "w", encoding="utf-8") as f:
            f.write("=" * 70 + "\n")
            f.write(f"RELATÓRIO DE GERAÇÃO — GEOJSON DO LOTE {lote['sigla']}\n")
            f.write("=" * 70 + "\n\n")
            f.write(f"Data/Hora: {_agora_brasilia().strftime('%d/%m/%Y %H:%M:%S')}\n")
            f.write(f"Lote: {lote['sigla']}\n")
            f.write(f"Modalidade: {mod.get('rotulo', mod['chave'])}\n")
            f.write(f"Versão: {versao_key}\n\n")
            f.write(f"Linhas Excel: {len(df)}\n")
            f.write(f"Features: {len(features_principal)}\n")
            f.write(f"Marcadores: {n_marcadores}\n")
            f.write(f"Pendências: {len(pendencias)}\n")
            if pendencias_por_motivo:
                f.write("\nResumo das pendências (motivo → quantidade):\n")
                for motivo, qtd in sorted(pendencias_por_motivo.items(), key=lambda x: -x[1]):
                    f.write(f"  • {motivo}: {qtd}\n")
                f.write("  (detalhe por linha em PENDENCIAS_{}.csv)\n".format(lote["sigla"]))
            f.write("\nSHA256 GeoJSON: {sha_geo}\n".format(sha_geo=sha_geo))
            f.write(f"SHA256 Excel:   {sha_xls}\n")
            f.write(f"SHA256 Schema:  {sha_sch}\n\n")
            st = "APROVADO" if val and val.get("ok") else "NEGADO"
            f.write(f"Validação: {st}\n")
            if val:
                f.write(f"{val.get('msg', '')}\n")
            f.write("\n" + "=" * 70 + "\nFIM\n" + "=" * 70 + "\n")
        arquivos.append(f"{nome_base}_LOG.txt")
    except Exception:
        pass

    if dashboard and gerar_dashboard_artesp:
        try:
            titulo = f"ARTESP {lote['sigla']} — {nome_base}"
            html_path = gerar_dashboard_artesp(str(caminho_saida), titulo=titulo)
            if html_path and os.path.isfile(html_path):
                arquivos.append(os.path.basename(html_path))
        except Exception:
            pass

    # ─── [P1] Gerar PDF de relatório ───
    pdf_nome = _gerar_pdf_relatorio(
        output_path=str(OUTPUT_PATH),
        nome_arquivo=nome_arquivo,
        lote=lote_key,
        modalidade=modalidade_key,
        versao=versao_key,
        ano=ano,
        mes=mes,
        usuario_email=usuario_email,
        n_linhas=len(df),
        n_features=len(features_principal),
        n_pendencias=len(pendencias),
        sha256_geojson=sha_geo,
        sha256_excel=_sha256_arquivo(excel_path),
        arquivos_gerados=arquivos,
    )
    if pdf_nome:
        arquivos.append(pdf_nome)

    # ─── [P1] Gerar Excel protegido de resumo ───
    xlsx_nome = _gerar_excel_resumo(
        output_path=str(OUTPUT_PATH),
        nome_arquivo=nome_arquivo,
        lote=lote_key,
        modalidade=modalidade_key,
        versao=versao_key,
        ano=ano,
        mes=mes,
        usuario_email=usuario_email,
        n_linhas=len(df),
        n_features=len(features_principal),
        n_pendencias=len(pendencias),
        sha256_geojson=sha_geo,
        sha256_excel=_sha256_arquivo(excel_path),
        arquivos_gerados=arquivos,
    )
    if xlsx_nome:
        arquivos.append(xlsx_nome)

    prog("Concluído", 98)
    resultado = {
        "principal": nome_arquivo,
        "arquivos": arquivos,
        "n_linhas_df": len(df),
        "n_features": len(features_principal),
        "n_pendencias": len(pendencias),
        "pendencias_por_motivo": dict(pendencias_por_motivo),
    }
    # ─── [P3] Metadados de estrutura para auditoria ───
    resultado["_artesp_meta"] = {
        "lote": lote_key,
        "modalidade": modalidade_key,
        "versao": versao_key,
        "ano": ano,
        "mes": mes,
        "caminho_relativo": _construir_caminho_artesp(lote_key, ano, versao_key, mes),
        "nome_zip": _construir_nome_zip_artesp(lote_key, modalidade_key, versao_key, ano, mes),
    }
    return resultado, None


@app.post("/gerar-relatorio-excel")
async def gerar_relatorio_excel(
    request: Request,
    usuario_email: str = Depends(_get_usuario_autenticado),
    arquivo_excel: UploadFile = File(...),
    lote: str = Form(...),
    modalidade: str = Form(...),
    versao: str = Form(...),
    ano: str = Form(...),
    mes: str = Form(default=""),
    correcao_eixo: bool = Form(default=False),
    dashboard: bool = Form(default=False),
    assinar: bool = Form(default=True),
    dry_run: bool = Form(default=False),
    x_stream_progress: Optional[str] = Header(default=None, alias="X-Stream-Progress"),
    x_format: Optional[str] = Header(default=None, alias="X-Format"),
):
    _check_rate_limit(request, f"gerar:{usuario_email}", RATE_LIMIT_GERAR_MAX, RATE_LIMIT_GERAR_JANELA)
    # [P2] Validação de upload (extensão + nome seguro + tamanho)
    _validar_upload(arquivo_excel)
    try:
        ano_int = int(ano.strip())
    except ValueError:
        raise HTTPException(status_code=400, detail="Ano inválido.")
    if not (2020 <= ano_int <= 2035):
        raise HTTPException(status_code=400, detail="Ano fora do intervalo (2020–2035).")
    if not (lote or "").strip():
        raise HTTPException(status_code=400, detail="Selecione o lote.")

    mes_int = None
    if mes and mes.strip():
        try:
            mes_int = int(mes.strip().split("-")[0].split(" ")[0])
        except (ValueError, IndexError):
            pass

    # Relatórios mensais: programado (M) = somente mês seguinte; executado (E) = somente mês anterior
    ano_forcado, mes_forcado = _periodo_mensal_por_versao(versao)
    if ano_forcado is not None and mes_forcado is not None:
        ano_int, mes_int = ano_forcado, mes_forcado

    # [P2] Leitura com limite de tamanho
    content = await _ler_upload_com_limite(arquivo_excel)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    stream_requested = (x_stream_progress or "").strip().lower() in ("1", "true", "sim", "yes")

    if stream_requested:
        progress_queue: Queue = Queue()
        result_holder: list = []

        def worker():
            try:
                r, e = _gerar_relatorio_do_excel(
                    excel_path=tmp_path,
                    lote_key=lote.strip(),
                    modalidade_key=modalidade.strip(),
                    versao_key=versao.strip(),
                    ano=ano_int,
                    mes=mes_int,
                    correcao_eixo=correcao_eixo,
                    dashboard=dashboard,
                    assinar=assinar,
                    usuario_email=usuario_email,
                    progress_queue=progress_queue,
                    dry_run=dry_run,
                )
                result_holder.append((r, e))
            except Exception as ex:
                result_holder.append((None, str(ex)))
            finally:
                progress_queue.put({"done": True})

        threading.Thread(target=worker, daemon=True).start()

        use_sse = (x_format or "").strip().lower() == "sse"

        def _evt(obj):
            j = json.dumps(obj, ensure_ascii=False)
            return f"data: {j}\n\n" if use_sse else j + "\n"

        def stream_gen():
            while True:
                try:
                    item = progress_queue.get(timeout=0.5)
                except Empty:
                    yield _evt({"type": "ping"})
                    continue
                if item.get("done"):
                    break
                yield _evt({"type": "progress", "status": item.get("msg", ""), "progresso": item.get("pct", 0), "msg": item.get("msg", ""), "pct": item.get("pct", 0)})

            r, e = result_holder[0] if result_holder else (None, "Erro interno")
            try:
                if tmp_path and os.path.isfile(tmp_path):
                    os.unlink(tmp_path)
            except Exception:
                pass
            if e:
                err = _classificar_erro(str(e))
                yield _evt({"type": "result", "status": "error", "detail": err["detail"], "error_code": err["code"], "error_type": err["type"]})
            elif isinstance(r, dict) and r.get("dry_run"):
                n_pend = r.get("n_pendencias", 0) or 0
                arquivos_gerados = r.get("arquivos", [])
                yield _evt({
                    "type": "result",
                    "status": "simulacao",
                    "progresso": 100,
                    "pendencias": n_pend,
                    "pendencias_por_motivo": r.get("pendencias_por_motivo", {}),
                    "csv_url": r.get("csv_url"),
                    "arquivos_gerados": arquivos_gerados,
                    "arquivo_nome": arquivos_gerados[0] if arquivos_gerados else "",
                })
            else:
                nome_principal = r.get("principal", "") if isinstance(r, dict) else ""
                arquivos_gerados = r.get("arquivos", [nome_principal]) if isinstance(r, dict) else []
                n_pend = r.get("n_pendencias", 0) or 0 if isinstance(r, dict) else 0
                sha_geojson = _sha256_arquivo(str(OUTPUT_PATH / nome_principal)) if nome_principal else ""
                arquivos = _montar_arquivos_links(arquivos_gerados, nome_principal, n_pend)
                yield _evt({
                    "type": "result",
                    "status": "sucesso",
                    "progresso": 100,
                    "link": f"/outputs/{nome_principal}",
                    "usuario_email": usuario_email,
                    "arquivo_nome": nome_principal,
                    "arquivos_gerados": arquivos_gerados,
                    "arquivos": arquivos,
                    "download_url": f"/outputs/{nome_principal}",
                    "sha256_geojson": sha_geojson,
                    "n_linhas_df": r.get("n_linhas_df") if isinstance(r, dict) else None,
                    "n_features": r.get("n_features") if isinstance(r, dict) else None,
                    "n_pendencias": n_pend,
                    "pendencias_por_motivo": r.get("pendencias_por_motivo", {}) if isinstance(r, dict) else {},
                    "versao": versao.strip(),
                })

        return StreamingResponse(
            stream_gen(),
            media_type="text/event-stream" if use_sse else "application/x-ndjson",
            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
        )

    loop = asyncio.get_event_loop()
    try:
        resultado, erro = await loop.run_in_executor(
            None,
            lambda: _gerar_relatorio_do_excel(
                excel_path=tmp_path,
                lote_key=lote.strip(),
                modalidade_key=modalidade.strip(),
                versao_key=versao.strip(),
                ano=ano_int,
                mes=mes_int,
                correcao_eixo=correcao_eixo,
                dashboard=dashboard,
                assinar=assinar,
                usuario_email=usuario_email,
                dry_run=dry_run,
            ),
        )
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass

    if erro:
        err = _classificar_erro(erro)
        raise HTTPException(status_code=422, detail={"message": err["detail"], "error_code": err["code"], "error_type": err["type"]})

    if resultado.get("dry_run"):
        return {
            "status": "simulacao",
            "pendencias": resultado.get("n_pendencias", 0),
            "pendencias_por_motivo": resultado.get("pendencias_por_motivo", {}),
            "csv_url": resultado.get("csv_url"),
            "arquivos_gerados": resultado.get("arquivos", []),
        }

    nome_principal = resultado["principal"]
    arquivos_gerados = resultado.get("arquivos", [nome_principal])
    n_pendencias = resultado.get("n_pendencias", 0) or 0
    sha_geojson = _sha256_arquivo(str(OUTPUT_PATH / nome_principal))
    arquivos = _montar_arquivos_links(arquivos_gerados, nome_principal, n_pendencias)
    return {
        "status": "sucesso",
        "usuario_email": usuario_email,
        "arquivo_nome": nome_principal,
        "arquivos_gerados": arquivos_gerados,
        "arquivos": arquivos,
        "download_url": f"/outputs/{nome_principal}",
        "download_urls": {os.path.basename(f): f"/outputs/{f}" for f in arquivos_gerados},
        "sha256_geojson": sha_geojson,
        "n_linhas_df": resultado.get("n_linhas_df"),
        "n_features": resultado.get("n_features"),
        "n_pendencias": n_pendencias,
        "pendencias_por_motivo": resultado.get("pendencias_por_motivo", {}),
        "versao": versao.strip(),
    }


@app.post("/gerar-relatorio-progresso")
async def gerar_com_progresso(
    request: Request,
    usuario_email: str = Depends(_get_usuario_autenticado),
    arquivo_excel: UploadFile = File(...),
    lote: str = Form(...),
    modalidade: str = Form(...),
    versao: str = Form(...),
    ano: str = Form(...),
    mes: str = Form(default=""),
    correcao_eixo: bool = Form(default=False),
    dashboard: bool = Form(default=False),
    assinar: bool = Form(default=True),
    dry_run: bool = Form(default=False),
):
    """Rota que retorna stream SSE com status em tempo real. [P2] Rate limit + upload validado."""
    _check_rate_limit(request, f"gerar:{usuario_email}", RATE_LIMIT_GERAR_MAX, RATE_LIMIT_GERAR_JANELA)
    # [P2] Validação de upload (extensão + nome seguro + tamanho)
    _validar_upload(arquivo_excel)
    try:
        ano_int = int(ano.strip())
    except ValueError:
        raise HTTPException(status_code=400, detail="Ano inválido.")
    if not (2020 <= ano_int <= 2035):
        raise HTTPException(status_code=400, detail="Ano fora do intervalo (2020–2035).")
    if not (lote or "").strip():
        raise HTTPException(status_code=400, detail="Selecione o lote.")

    mes_int = None
    if mes and mes.strip():
        try:
            mes_int = int(mes.strip().split("-")[0].split(" ")[0])
        except (ValueError, IndexError):
            pass

    # Relatórios mensais: programado (M) = somente mês seguinte; executado (E) = somente mês anterior
    ano_forcado, mes_forcado = _periodo_mensal_por_versao(versao)
    if ano_forcado is not None and mes_forcado is not None:
        ano_int, mes_int = ano_forcado, mes_forcado

    # [P2] Leitura com limite de tamanho
    content = await _ler_upload_com_limite(arquivo_excel)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    progress_queue: Queue = Queue()
    result_holder: list = []

    def worker():
        try:
            r, e = _gerar_relatorio_do_excel(
                excel_path=tmp_path,
                lote_key=lote.strip(),
                modalidade_key=modalidade.strip(),
                versao_key=versao.strip(),
                ano=ano_int,
                mes=mes_int,
                correcao_eixo=correcao_eixo,
                dashboard=dashboard,
                assinar=assinar,
                usuario_email=usuario_email,
                progress_queue=progress_queue,
                dry_run=dry_run,
            )
            result_holder.append((r, e))
        except Exception as ex:
            result_holder.append((None, str(ex)))
        finally:
            progress_queue.put({"done": True})

    threading.Thread(target=worker, daemon=True).start()

    async def event_generator():
        try:
            yield f"data: {json.dumps({'type': 'progress', 'status': 'Recebendo arquivo Excel...', 'progresso': 5}, ensure_ascii=False)}\n\n"
            await asyncio.sleep(0.05)

            while True:
                try:
                    item = progress_queue.get_nowait()
                except Empty:
                    await asyncio.sleep(0.2)
                    yield f"data: {json.dumps({'type': 'ping', 'status': 'Processando...', 'progresso': None}, ensure_ascii=False)}\n\n"
                    continue
                if item.get("done"):
                    break
                msg = item.get("msg", "Processando...")
                pct = item.get("pct", 0)
                yield f"data: {json.dumps({'type': 'progress', 'status': msg, 'progresso': pct}, ensure_ascii=False)}\n\n"
                await asyncio.sleep(0.02)

            resultado, erro = result_holder[0] if result_holder else (None, "Erro interno")

            if tmp_path and os.path.isfile(tmp_path):
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass

            if erro:
                err = _classificar_erro(erro)
                yield f"data: {json.dumps({'type': 'result', 'status': 'error', 'detail': err['detail'], 'error_code': err['code'], 'error_type': err['type'], 'progresso': 0}, ensure_ascii=False)}\n\n"
                return

            if resultado.get("dry_run"):
                n_pend = resultado.get("n_pendencias", 0) or 0
                arq = resultado.get("arquivos", [])
                _atualizar_metricas_globais(resultado)
                yield f"data: {json.dumps({'type': 'result', 'status': 'simulacao', 'progresso': 100, 'pendencias': n_pend, 'pendencias_por_motivo': resultado.get('pendencias_por_motivo', {}), 'csv_url': resultado.get('csv_url'), 'arquivos_gerados': arq, 'arquivo_nome': arq[0] if arq else ''}, ensure_ascii=False)}\n\n"
                return

            n_pend = resultado.get("n_pendencias", 0) or 0
            yield f"data: {json.dumps({'type': 'progress', 'status': f'Processado: {n_pend} pendência(s) encontrada(s).', 'progresso': 70}, ensure_ascii=False)}\n\n"
            await asyncio.sleep(0.1)

            yield f"data: {json.dumps({'type': 'progress', 'status': 'Aplicando assinatura digital osslsigncode...', 'progresso': 90}, ensure_ascii=False)}\n\n"
            await asyncio.sleep(0.1)

            nome_principal = resultado.get("principal", "")
            arquivos_gerados = resultado.get("arquivos", [nome_principal])
            # [P3] Metadados de estrutura
            artesp_meta = resultado.get("_artesp_meta", {})
            links = _montar_arquivos_links(arquivos_gerados, nome_principal, n_pend)
            _atualizar_metricas_globais(resultado)

            yield f"data: {json.dumps({'type': 'result', 'status': 'sucesso', 'progresso': 100, 'arquivos': links, 'arquivos_gerados': arquivos_gerados, 'arquivo_nome': nome_principal, 'n_linhas_df': resultado.get('n_linhas_df'), 'n_features': resultado.get('n_features'), 'n_pendencias': n_pend, 'pendencias_por_motivo': resultado.get('pendencias_por_motivo', {}), 'artesp_meta': artesp_meta}, ensure_ascii=False)}\n\n"

        except Exception as e:
            if tmp_path and os.path.isfile(tmp_path):
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass
            err = _classificar_erro(str(e))
            yield f"data: {json.dumps({'type': 'result', 'status': 'error', 'detail': err['detail'], 'error_code': err['code'], 'error_type': err['type'], 'progresso': 0}, ensure_ascii=False)}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


def _montar_arquivos_links(arquivos_gerados: list, nome_principal: str, n_pendencias: int) -> dict:
    """Monta objeto com links de download por tipo para o analista."""
    links = {
        "geojson": f"/outputs/{nome_principal}" if nome_principal else None,
        "excel": None,
        "pdf": None,
        "log": None,
        "pendencias": None,
        "dashboard": None,
    }
    for f in (arquivos_gerados or []):
        if f.endswith(".geojson"):
            links["geojson"] = f"/outputs/{f}"
        elif f.endswith(".xlsx"):
            links["excel"] = f"/outputs/{f}"
        elif f.endswith(".pdf"):
            links["pdf"] = f"/outputs/{f}"
        elif f.endswith("_LOG.txt"):
            links["log"] = f"/outputs/{f}"
        elif f.startswith("PENDENCIAS_") and f.endswith(".csv"):
            links["pendencias"] = f"/outputs/{f}"
        elif f.endswith(".html"):
            links["dashboard"] = f"/outputs/{f}"
    return links


HISTORICO_METRICAS_DIAS = 30


def _atualizar_metricas_globais(resultado: dict) -> None:
    """[P2] Atualiza metrics.json com lock para evitar race condition."""
    lock = _file_locks.get_lock(str(METRICS_PATH))
    with lock:
        hoje = _agora_brasilia().strftime("%Y-%m-%d")
        data: dict = {}

        if METRICS_PATH.is_file():
            try:
                with open(METRICS_PATH, "r", encoding="utf-8") as fp:
                    data = json.load(fp)
            except (json.JSONDecodeError, IOError):
                pass

        if data.get("ultima_atualizacao_data") != hoje:
            data["pendencias_hoje"] = 0
        data["ultima_atualizacao_data"] = hoje

        n_pend = resultado.get("n_pendencias", 0) or 0
        data["pendencias_hoje"] = data.get("pendencias_hoje", 0) + n_pend

        total_hoje = 0
        if not resultado.get("dry_run"):
            data["total_gerado"] = data.get("total_gerado", 0) + 1
            total_hoje = 1

        historico = data.get("historico", [])
        entry = next((e for e in historico if e.get("data") == hoje), None)
        if entry:
            entry["total"] = entry.get("total", 0) + total_hoje
            entry["pendencias"] = entry.get("pendencias", 0) + n_pend
        else:
            historico.append({"data": hoje, "total": total_hoje, "pendencias": n_pend})
        historico.sort(key=lambda x: x.get("data", ""), reverse=True)
        data["historico"] = historico[:HISTORICO_METRICAS_DIAS]

        temp_path = str(METRICS_PATH) + ".tmp"
        try:
            with open(temp_path, "w", encoding="utf-8") as fp:
                json.dump(data, fp, indent=2, ensure_ascii=False)
                fp.flush()
                os.fsync(fp.fileno())
            os.replace(temp_path, str(METRICS_PATH))
        except (IOError, OSError) as e:
            logging.warning("[P2] Falha ao gravar metrics.json: %s", e)
            try:
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
            except OSError:
                pass


def _coletar_stats_admin() -> dict:
    """Coleta métricas para o painel admin a partir de OUTPUT_PATH."""
    total_gerado = 0
    lote_counts: Dict[str, int] = {}
    bytes_total = 0
    hoje = _agora_brasilia().date()
    bytes_liberados_hoje = 0
    try:
        if OUTPUT_PATH.is_dir():
            for f in OUTPUT_PATH.iterdir():
                if not f.is_file():
                    continue
                total_gerado += 1 if f.suffix.lower() == ".geojson" else 0
                try:
                    stat = f.stat()
                    bytes_total += stat.st_size
                    mtime = datetime.datetime.fromtimestamp(stat.st_mtime, tz=TZ_BRASILIA).date()
                    if mtime == hoje:
                        bytes_liberados_hoje += stat.st_size
                    # Contar por lote (ex: L13, L21 no nome)
                    for sigla in ["L13", "L21", "L26"]:
                        if sigla in f.name:
                            lote_counts[sigla] = lote_counts.get(sigla, 0) + 1
                            break
                except OSError:
                    pass
        lote_mais_ativo = max(lote_counts, key=lote_counts.get) if lote_counts else "—"
        if METRICS_PATH.is_file():
            try:
                with open(METRICS_PATH, "r", encoding="utf-8") as fp:
                    m = json.load(fp)
                return {
                    "total_gerado": m.get("total_gerado", total_gerado),
                    "pendencias_hoje": m.get("pendencias_hoje", 0),
                    "historico": m.get("historico", []),
                    "concessionaria_lider": m.get("concessionaria_lider", "Via Colinas"),
                    "lote_mais_ativo": m.get("lote_mais_ativo", lote_mais_ativo),
                    "percentual_conformidade": m.get("percentual_conformidade", "94%"),
                    "espaco_disco_liberado_hoje": f"{bytes_liberados_hoje / (1024 * 1024):.1f}MB" if bytes_liberados_hoje else "0MB",
                    "arquivos_outputs": total_gerado,
                    "bytes_outputs": bytes_total,
                }
            except (json.JSONDecodeError, IOError):
                pass
        return {
            "total_gerado": total_gerado,
            "pendencias_hoje": 0,
            "historico": [],
            "concessionaria_lider": "Via Colinas",
            "lote_mais_ativo": lote_mais_ativo,
            "percentual_conformidade": "94%",
            "espaco_disco_liberado_hoje": f"{bytes_liberados_hoje / (1024 * 1024):.1f}MB" if bytes_liberados_hoje else "0MB",
            "arquivos_outputs": total_gerado,
            "bytes_outputs": bytes_total,
        }
    except Exception as e:
        logging.warning("Erro ao coletar stats admin: %s", e)
        return {
            "total_gerado": 0,
            "pendencias_hoje": 0,
            "historico": [],
            "concessionaria_lider": "—",
            "lote_mais_ativo": "—",
            "percentual_conformidade": "—",
            "espaco_disco_liberado_hoje": "0MB",
        }


@app.get("/api/stats")
async def get_api_stats(usuario_email: str = Depends(_get_usuario_autenticado)):
    """Métricas para o Dashboard (usuários autenticados)."""
    _ = usuario_email
    return _coletar_stats_admin()


@app.get("/admin/stats")
async def get_admin_stats(current_user: str = Depends(_get_usuario_admin)):
    """Dashboard de métricas (Painel Admin). Requer permissão de administrador."""
    _ = current_user
    stats = _coletar_stats_admin()
    return stats


@app.get("/admin/check")
async def admin_check(current_user: str = Depends(_get_usuario_admin)):
    """
    Verifica se o usuário autenticado é administrador (proteção dupla).
    Retorna 403 se não houver token válido de admin. Use no frontend para
    esconder/mostrar link de cadastro ou bloquear acesso à página de gestão.
    """
    role = _get_user_role(current_user)
    return {"ok": True, "role": role}


@app.post("/admin/adicionar-usuario")
async def adicionar_usuario(
    payload: NovoUsuarioPayload,
    current_user: str = Depends(_get_usuario_admin),
):
    """
    [P2] Adiciona novo usuário com operação atômica (leitura+escrita com lock).
    """
    _ = current_user
    email = (payload.email or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="E-mail inválido.")

    senha = (payload.senha or "").strip()
    if not senha:
        raise HTTPException(status_code=400, detail="Senha não pode ser vazia.")
    try:
        hashed_password = gerar_hash_senha(senha)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    role = (payload.role or "user").strip().lower() or "user"

    def _adicionar(usuarios):
        if email in usuarios:
            raise HTTPException(status_code=400, detail="Usuário já cadastrado.")
        usuarios[email] = {
            "hashed_password": hashed_password,
            "disabled": False,
            "role": role,
            "must_change_password": True,
        }
        return usuarios

    _modificar_banco_usuarios(_adicionar)
    return {"message": f"Acesso para {email} criado com sucesso!"}


@app.post("/admin/redefinir-senha-usuario")
async def admin_redefinir_senha(
    payload: RedefinirSenhaAdminPayload,
    current_user: str = Depends(_get_usuario_admin),
):
    """
    Admin redefine a senha de um usuário (senha temporária). O usuário deve trocar no próximo login.
    """
    _ = current_user
    email = (payload.email or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="E-mail inválido.")
    nova = (payload.nova_senha_temporaria or "").strip()
    if len(nova) < 6:
        raise HTTPException(status_code=400, detail="Senha temporária deve ter no mínimo 6 caracteres.")
    try:
        hashed = gerar_hash_senha(nova)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    def _redefinir(usuarios):
        if email not in usuarios:
            raise HTTPException(status_code=404, detail="Usuário não encontrado.")
        usuarios[email] = {
            **usuarios[email],
            "hashed_password": hashed,
            "must_change_password": True,
        }
        return usuarios

    _modificar_banco_usuarios(_redefinir)
    return {"message": f"Senha de {email} redefinida. O usuário deve trocar no próximo acesso."}


@app.get("/admin/usuarios")
async def listar_usuarios_admin(current_user: str = Depends(_get_usuario_admin)):
    """Lista usuários do users.json (email, role, disabled). Não expõe senha."""
    _ = current_user
    banco = carregar_banco_usuarios()
    lista = []
    for email, obj in (banco.items() if isinstance(banco, dict) else []):
        if not isinstance(obj, dict):
            continue
        role = (obj.get("role") or "user").strip().lower()
        if role not in ("admin", "user"):
            role = "user"
        lista.append({
            "email": email,
            "role": role,
            "disabled": bool(obj.get("disabled")),
            "owner": _eh_proprietario(email),
        })
    lista.sort(key=lambda x: (x["email"],))
    return {"usuarios": lista}


@app.post("/admin/alterar-role-usuario")
async def alterar_role_usuario(
    payload: AlterarRolePayload,
    current_user: str = Depends(_get_usuario_admin),
):
    """Altera perfil do usuário (user ↔ admin). Não permite rebaixar o último admin nem o proprietário."""
    _ = current_user
    email = (payload.email or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="E-mail inválido.")
    if _eh_proprietario(email):
        raise HTTPException(status_code=403, detail="Não é possível alterar o perfil do proprietário (admin master).")
    role = (payload.role or "user").strip().lower()
    if role not in ("admin", "user"):
        role = "user"

    def _alterar(usuarios):
        if email not in usuarios:
            raise HTTPException(status_code=404, detail="Usuário não encontrado.")
        prev = usuarios[email]
        if not isinstance(prev, dict):
            raise HTTPException(status_code=400, detail="Dados do usuário inválidos.")
        if role == "user" and (prev.get("role") or "").strip().lower() == "admin":
            copia = {k: dict(v) if isinstance(v, dict) else v for k, v in usuarios.items()}
            copia[email] = {**prev, "role": "user"}
            if _numero_admins(copia) < 1:
                raise HTTPException(status_code=400, detail="Não é possível rebaixar o último administrador.")
        usuarios[email] = {**prev, "role": role}
        return usuarios

    _modificar_banco_usuarios(_alterar)
    return {"message": f"Perfil de {email} alterado para {role}."}


@app.post("/admin/bloquear-usuario")
async def bloquear_usuario(
    payload: BloquearUsuarioPayload,
    current_user: str = Depends(_get_usuario_admin),
):
    """Bloqueia ou desbloqueia usuário (campo disabled). Não permite bloquear o último admin nem o proprietário."""
    _ = current_user
    email = (payload.email or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="E-mail inválido.")
    if _eh_proprietario(email) and payload.bloquear:
        raise HTTPException(status_code=403, detail="Não é possível bloquear o proprietário (admin master).")

    def _bloquear(usuarios):
        if email not in usuarios:
            raise HTTPException(status_code=404, detail="Usuário não encontrado.")
        prev = usuarios[email]
        if not isinstance(prev, dict):
            raise HTTPException(status_code=400, detail="Dados do usuário inválidos.")
        if payload.bloquear and (prev.get("role") or "").strip().lower() == "admin":
            copia = {k: dict(v) if isinstance(v, dict) else v for k, v in usuarios.items()}
            copia[email] = {**prev, "disabled": True}
            if _numero_admins(copia) < 1:
                raise HTTPException(status_code=400, detail="Não é possível bloquear o último administrador.")
        usuarios[email] = {**prev, "disabled": payload.bloquear}
        return usuarios

    _modificar_banco_usuarios(_bloquear)
    msg = "bloqueado" if payload.bloquear else "desbloqueado"
    return {"message": f"Usuário {email} {msg}."}


@app.post("/admin/remover-usuario")
async def remover_usuario(
    payload: RemoverUsuarioPayload,
    current_user: str = Depends(_get_usuario_admin),
):
    """Remove usuário do users.json. Não permite remover o último admin nem o proprietário."""
    _ = current_user
    email = (payload.email or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="E-mail inválido.")
    if _eh_proprietario(email):
        raise HTTPException(status_code=403, detail="Não é possível remover o proprietário (admin master).")

    def _remover(usuarios):
        if email not in usuarios:
            raise HTTPException(status_code=404, detail="Usuário não encontrado.")
        prev = usuarios[email]
        if isinstance(prev, dict) and (prev.get("role") or "").strip().lower() == "admin":
            copia = {k: v for k, v in usuarios.items() if k != email}
            if _numero_admins(copia) < 1:
                raise HTTPException(status_code=400, detail="Não é possível remover o último administrador.")
        del usuarios[email]
        return usuarios

    _modificar_banco_usuarios(_remover)
    return {"message": f"Usuário {email} removido."}


def _media_type_output(nome: str) -> str:
    ext = (nome or "").lower().split(".")[-1]
    return {
        "geojson": "application/geo+json",
        "json": "application/json",
        "pdf": "application/pdf",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xls": "application/vnd.ms-excel",
        "html": "text/html",
        "txt": "text/plain",
    }.get(ext, "application/octet-stream")


def _gerar_readme_auditoria(
    meta: dict,
    arquivos: list,
    usuario_email: str = "",
) -> str:
    """
    [P3] Gera texto README para incluir no ZIP de auditoria.
    """
    agora = _agora_brasilia()
    lote = meta.get("lote", "?")
    modalidade = meta.get("modalidade", "?")
    versao = meta.get("versao", "?")
    ano = meta.get("ano", "?")
    mes = meta.get("mes")

    mod_label = "Conservação" if str(modalidade).lower() in ("1", "conserva", "conservacao") else "Obras"
    tipo_label = _classificar_versao(str(versao))
    mes_label = _MESES_NOME_COMPLETO.get(int(mes), str(mes)) if mes else "N/A"

    linhas = [
        "=" * 60,
        "  RELATÓRIO DE GERAÇÃO ARTESP — INFORMAÇÕES DE AUDITORIA",
        "=" * 60,
        "",
        f"  Data/Hora de Geração: {agora.strftime('%d/%m/%Y %H:%M:%S')}",
        f"  Usuário:              {usuario_email or '(não informado)'}",
        f"  Lote:                 {lote}",
        f"  Modalidade:           {mod_label}",
        f"  Versão:               {tipo_label} ({versao})",
        f"  Ano:                  {ano}",
        f"  Mês:                  {mes_label}",
        "",
        "-" * 60,
        "  ESTRUTURA DE PASTAS",
        "-" * 60,
        "",
        f"  {meta.get('caminho_relativo', 'N/A')}",
        "",
        "  Estrutura: uma única pasta por relatório (ex.: Lote_21 Anual),",
        "  contendo GeoJSON, Excel, PDF e demais arquivos gerados.",
        "",
        "-" * 60,
        "  ARQUIVOS INCLUÍDOS",
        "-" * 60,
        "",
    ]

    for i, arq in enumerate(arquivos, 1):
        ext = arq.rsplit(".", 1)[-1].lower() if "." in arq else "?"
        tipo_arq = {
            "geojson": "GeoJSON (dados geográficos)",
            "pdf": "PDF (relatório de geração)",
            "xlsx": "Excel (resumo protegido)",
            "csv": "CSV (pendências)",
            "txt": "TXT (log de processamento)",
            "html": "HTML (dashboard)",
        }.get(ext, f"Arquivo .{ext}")
        linhas.append(f"  {i:2d}. {arq}")
        linhas.append(f"      Tipo: {tipo_arq}")
        linhas.append("")

    linhas.extend([
        "-" * 60,
        "  CONFORMIDADE",
        "-" * 60,
        "",
        "  [x] GeoJSON validado contra schema JSON (conserva.schema.r0.json)",
        "  [x] data_geracao em RFC3339 com timezone (-03:00)",
        "  [x] rodovia: pattern SP/SPM/SPI/SPD 9-10 chars",
        "  [x] item: pattern ^[a-z](\\.\\d+)+$",
        "  [x] local: enum validado contra schema",
        "  [x] quantidade/km_inicial/km_final: multipleOf 0.001",
        "  [x] observacoes_gerais: null ou string (nunca vazio)",
        "  [x] id <= 50 caracteres",
        "  [x] SHA256 dos arquivos no PDF/log para rastreabilidade",
        "  [x] Nomenclatura conforme padrão ARTESP",
        "  [x] Excel protegido contra alteração",
        "",
        "-" * 60,
        "  REFERÊNCIAS",
        "-" * 60,
        "",
        "  Portal de Dados Abertos ARTESP:",
        "    https://dadosabertos.artesp.sp.gov.br/dataset/programacao-de-obras",
        "",
        "  Manual GeoJSON v1.0:",
        "    https://dorettoartesp.github.io/geojson-manual/",
        "",
        "  Schema JSON (Draft 2020-12):",
        "    conserva.schema.r0.json / obras.schema.r0.json",
        "",
        "=" * 60,
        "  Gerado automaticamente pelo Sistema ARTESP Web",
        "=" * 60,
        "",
    ])

    dev_nome = os.getenv("ARTESP_DEV_NOME", "").strip()
    dev_email = os.getenv("ARTESP_DEV_EMAIL", "").strip()
    if dev_nome or dev_email:
        linhas.append(f"  Desenvolvedor: {dev_nome} | {dev_email}")
        linhas.append("")

    return "\n".join(linhas)


class DownloadZipPayload(BaseModel):
    files: list = Field(default_factory=list, description="Lista de nomes de arquivos para incluir no ZIP")
    artesp_meta: Optional[dict] = Field(default=None, description="[P3] Metadados opcionais para estruturar o ZIP")


@app.post("/outputs/zip")
async def download_outputs_zip(
    payload: DownloadZipPayload,
    request: Request,
    usuario_email: str = Depends(_get_usuario_autenticado),
):
    """
    [P3] Gera ZIP com estrutura de pastas ARTESP para auditoria.
    [P2] Rate limit por usuário.
    """
    _check_rate_limit(request, f"download:{usuario_email}", RATE_LIMIT_DOWNLOAD_MAX, RATE_LIMIT_DOWNLOAD_JANELA)
    _ = usuario_email
    arquivos = payload.files or []
    if not arquivos:
        raise HTTPException(status_code=400, detail="Nenhum arquivo selecionado.")

    meta = payload.artesp_meta or {}
    caminho_relativo = meta.get("caminho_relativo", "")
    nome_zip = meta.get("nome_zip", "")

    if not nome_zip:
        stamp = _agora_brasilia().strftime("%Y%m%d_%H%M%S")
        nome_zip = f"artesp_relatorio_{stamp}.zip"
    if not nome_zip.lower().endswith(".zip"):
        nome_zip += ".zip"

    buffer = BytesIO()
    num_arquivos = 0
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
        for nome_arquivo in arquivos:
            nome_limpo = os.path.basename(nome_arquivo)
            if not nome_limpo or ".." in nome_limpo:
                continue

            path_real = OUTPUT_PATH / nome_limpo
            if not path_real.is_file():
                logging.warning("[P3] Arquivo não encontrado para ZIP: %s", nome_limpo)
                continue

            try:
                path_resolved = path_real.resolve()
                output_resolved = OUTPUT_PATH.resolve()
                if output_resolved not in path_resolved.parents and path_resolved != output_resolved:
                    logging.warning("[P3] Path traversal bloqueado: %s", nome_arquivo)
                    continue
            except (ValueError, OSError):
                continue

            if caminho_relativo:
                arcname = os.path.join(caminho_relativo, nome_limpo)
            else:
                arcname = nome_limpo

            zf.write(str(path_real), arcname)
            num_arquivos += 1

        if caminho_relativo:
            readme = _gerar_readme_auditoria(meta, arquivos, usuario_email or "")
            readme_path = os.path.join(caminho_relativo, "README_AUDITORIA.txt")
            zf.writestr(readme_path, readme)

    if num_arquivos == 0:
        raise HTTPException(
            status_code=404,
            detail="Nenhum dos arquivos foi encontrado no servidor. Os arquivos podem ter expirado (limpeza automática). Gere o relatório novamente e baixe o ZIP em seguida.",
        )

    buffer.seek(0)
    conteudo = buffer.getvalue()
    nome_zip_safe = nome_zip.replace('"', "'")
    headers = {
        "Content-Disposition": f'attachment; filename="{nome_zip_safe}"',
        "Content-Length": str(len(conteudo)),
        "X-ARTESP-Estrutura": caminho_relativo or "flat",
    }
    return Response(
        content=conteudo,
        media_type="application/zip",
        headers=headers,
    )


@app.get("/outputs/nc/{job_id}/{subpath:path}")
async def download_nc_job_file(
    job_id: str,
    subpath: str,
    request: Request,
):
    """
    Ficheiros do workspace NC (ex.: final/*.zip, stage1/nc_separados.zip).
    O JSON do pipeline devolve URLs neste formato; não confundir com GET /outputs/{ficheiro_plano}.
    """
    # Fluxo NC permite acesso sem sessão; manter rate-limit por IP para proteção básica.
    _check_rate_limit(request, "download:nc", RATE_LIMIT_DOWNLOAD_MAX, RATE_LIMIT_DOWNLOAD_JANELA)
    if not job_id or ".." in job_id or "/" in job_id or "\\" in job_id:
        raise HTTPException(status_code=400, detail="job_id inválido.")
    if not subpath or not str(subpath).strip():
        raise HTTPException(status_code=400, detail="Caminho inválido.")
    norm = str(subpath).replace("\\", "/").lstrip("/")
    if any(p == ".." for p in norm.split("/")):
        raise HTTPException(status_code=400, detail="Caminho inválido.")
    base = (OUTPUT_PATH / "nc" / job_id).resolve()
    if not base.is_dir():
        raise HTTPException(status_code=404, detail="Not Found")
    target = (base / norm).resolve()
    try:
        target.relative_to(base)
    except ValueError:
        raise HTTPException(status_code=400, detail="Acesso negado.")
    if not target.is_file():
        raise HTTPException(status_code=404, detail="Not Found")
    try:
        from render_api.job_manager import carregar_job_nc as _nc_job_touch

        try:
            _nc_job_touch(job_id, touch=True)
        except HTTPException:
            pass
    except ImportError:
        pass
    nome_limpo = target.name
    ext = nome_limpo.rsplit(".", 1)[-1].lower() if "." in nome_limpo else ""
    content_types = {
        "geojson": "application/geo+json",
        "json": "application/json",
        "pdf": "application/pdf",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "csv": "text/csv; charset=utf-8",
        "txt": "text/plain; charset=utf-8",
        "html": "text/html; charset=utf-8",
        "zip": "application/zip",
    }
    media_type = content_types.get(ext, "application/octet-stream")
    return FileResponse(
        path=str(target),
        filename=nome_limpo,
        media_type=media_type,
        headers={"X-NC-Job-Id": job_id, "X-ARTESP-Arquivo": nome_limpo},
    )


@app.get("/outputs/{nome_arquivo}")
async def download_output(
    nome_arquivo: str,
    request: Request,
    usuario_email: str = Depends(_get_usuario_autenticado),
):
    _check_rate_limit(request, f"download:{usuario_email}", RATE_LIMIT_DOWNLOAD_MAX, RATE_LIMIT_DOWNLOAD_JANELA)
    _ = usuario_email
    nome_limpo = os.path.basename(nome_arquivo)
    if not nome_limpo or ".." in nome_limpo:
        raise HTTPException(status_code=400, detail="Nome de arquivo inválido.")

    path = OUTPUT_PATH / nome_limpo
    if not path.is_file():
        raise HTTPException(status_code=404, detail="Arquivo não encontrado.")

    try:
        path_resolved = path.resolve()
        output_resolved = OUTPUT_PATH.resolve()
        if output_resolved not in path_resolved.parents and path_resolved != output_resolved:
            raise HTTPException(status_code=400, detail="Acesso negado.")
    except (ValueError, OSError):
        raise HTTPException(status_code=400, detail="Acesso negado.")

    ext = nome_limpo.rsplit(".", 1)[-1].lower() if "." in nome_limpo else ""
    content_types = {
        "geojson": "application/geo+json",
        "json": "application/json",
        "pdf": "application/pdf",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "csv": "text/csv; charset=utf-8",
        "txt": "text/plain; charset=utf-8",
        "html": "text/html; charset=utf-8",
        "zip": "application/zip",
    }
    media_type = content_types.get(ext, "application/octet-stream")

    return FileResponse(
        path=str(path),
        filename=nome_limpo,
        media_type=media_type,
        headers={"X-ARTESP-Arquivo": nome_limpo},
    )


@app.get("/health")
def health_check():
    schemas_disponiveis = [f for f in SCHEMA_MAP.values() if (SCHEMA_PATH / f).is_file()]
    if platform.system() == "Windows":
        assinatura_tool = "signtool"
        assinatura_ok = shutil.which("signtool") is not None
    else:
        assinatura_tool = "osslsigncode"
        assinatura_ok = shutil.which("osslsigncode") is not None

    pfx_configurado = bool((os.getenv("ARTESP_PFX") or "").strip() or (os.getenv("ARTESP_PFX_CONTENT") or "").strip())
    return {
        "status": "online",
        "ambiente": platform.system(),
        "auth": {
            "modo": "email_senha_bearer",
            "usuarios_configurados": len(USUARIOS_WEB),
            "tokens_ativos": len(TOKENS),
            "token_ttl_segundos": TOKEN_TTL_SECONDS,
        },
        "schema_disponiveis": schemas_disponiveis,
        "modulo_assinatura": "Configurado" if pfx_configurado else "Ausente",
        "ferramenta_assinatura": {
            "nome": assinatura_tool,
            "disponivel": assinatura_ok,
        },
        "diretorio_raiz": str(BASE_DIR),
    }


# ── INVENTÁRIO DE DRENAGEM ─────────────────────────────────────────────────────
try:
    from render_api.inventario_router import router as _inv_router, setup_inventario as _inv_setup
    app.include_router(_inv_router)
    _inv_setup()
    _inv_dir = Path(__file__).parent.parent / "frontend_inventario"
    if _inv_dir.is_dir():
        app.mount("/inventario", StaticFiles(directory=str(_inv_dir), html=True), name="inventario")
        logging.info("PWA Inventário montado em /inventario")
    else:
        logging.warning("Inventário: pasta frontend_inventario não encontrada em %s", _inv_dir)
except Exception as _e:
    logging.warning("Módulo inventário não carregado: %s", _e)