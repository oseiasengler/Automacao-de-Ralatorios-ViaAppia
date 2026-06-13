"""
render_api/nc_router.py
────────────────────────────────────────────────────────────────────────────
Router FastAPI para o pipeline de Não Conformidades (NC Artesp).

Modos de uso:
  • Etapa isolada (single-shot): sem job_id → cada endpoint cria job novo, processa,
    grava em final/, marca finished (retain 24h), retorna job_id e download_urls.
  • Pipeline completo: primeira chamada sem job_id cria o job; seguintes enviam job_id
    (Form), reutilizam o mesmo workspace (stage1/, stage2/, final/). Só marca finished
    quando created ou finalize=1 (retain 72h); senão status=running e stage=stage1|stage2.

Ordem típica (etapas isoladas): separar → gerar-modelo-foto → inserir-conservacao → juntar →
exportar-calendario; **M05 inserir-número** só se precisar preencher coluna Y manualmente.
**POST /nc/separar** (por omissão `entrega_completa=true`), **POST /nc/completo** e **POST /nc/stage2**
encadeiam M01→e-mail→M02→M04→M06 **sem** M03/M05; o ZIP final agrupa `Kartado_relatorio_fotos/`, `pendentes/`,
`Kartado/` (um ZIP **Kartado NCs Consolidadas** com Excel **layout Kartado** do M01 + imagens),
`acumulado/`, `calendario/`, `emails/` com caminhos internos limitados para extração no Windows (MAX_PATH).
Quando houver PDFs no pedido, as imagens são extraídas no servidor; o ZIP interno dessa extração
é também empacotado em `backup/nc_<job_id>_imagens_extraidas_backup.zip` dentro do ZIP final
(sem depender de ZIP de imagens enviado pelo cliente — opcional só em ``/nc/stage2``).
Com **um** EAF e ``m01_kartado=true``, o(s) mesmo(s) PDF(s) é(são) gravado(s) como ``<EAF>.pdf`` ao lado do Excel
antes do M01 para preencher a coluna **«Observações»** (AA no template Kartado geral) com o texto livre do campo «Observação» do PDF (léxicos X/Y: ``nc_artesp/utils/kartado_observacao_pdf.py``).
Alternativa em 2 chamadas: POST /nc/start (EAF) → POST /nc/stage2 (job_id + opcional ZIP imagens).

Fluxo completo ARTESP (Excel EAF + PDF só para imagens):
  • **Recomendado:** POST /nc/separar ou POST /nc/completo — multipart com EAF + PDF(s) opcional + **lote** (13, 21 ou 26).
    O contrato HTTP por omissão mantém **Art_011** (``m01_kartado`` omitido ou false); **Kartado** é opt-in com
    ``m01_kartado=true`` (lotes 13/21/26; lote 50 nunca usa este Kartado). O ``nc.html`` envia explicitamente ``m01_kartado=true``.
  • Alternativa em 2 chamadas: POST /nc/start (EAF + **lote** opcional) → POST /nc/stage2 (job_id + opcional imagens_pdf_zip).
  Artemig (lote 50) segue com analisar-pdf / fluxos próprios; /nc/completo replica a mesma extração de
  imagens que /nc/extrair-pdf (incl. pasta única lote 50).

Endpoints:
POST /nc/completo               – EAF + PDF(s) opcional → ZIP final (inclui backup de imagens embutido)
  POST /nc/extrair-pdf            – PDF NC Constatação → ZIP com nc(N).jpg e PDF(N).jpg; vários PDFs → Constatacoes_unificadas.pdf
  POST /nc/analisar-pdf           – PDF NC Constatação → ZIP (PDF de análise + XLSX); vários PDFs → Constatacoes_unificadas.pdf
  POST /nc/separar                → M01: EAF → ZIP final no servidor (padrão) ou só XLS (`entrega_completa=false`)
  POST /nc/gerar-modelo-foto      → M02: XLS ZIP + modelos → ZIP Kria + Resposta
  POST /nc/inserir-conservacao    → M03: Kria ZIP → ZIP Kcor-Kria Conservação
  POST /nc/inserir-meio-ambiente  → M07: Kria MA ZIP → ZIP Kcor-Kria MA
  POST /nc/juntar                 → M04: Kcor ZIP → XLSX acumulado
  POST /nc/inserir-numero         → M05 opcional (manual): acumulado + nº inicial → coluna Y
  POST /nc/exportar-calendario    → M06: acumulado → arquivo .ics (iCalendar)
  GET  /nc/                       → status e deps
  (M08 Organizar Imagens existia apenas nas macros VBA — removido do fluxo web.)

Fonte de dados (referência de desenho):
  • **Artemig (ex.: lote 50 / analisar-pdf):** muito do conteúdo estruturado vem do **texto do PDF**
    parseado para preencher relatórios/planilhas (EAF, Exportar Kcor, etc.).
  • **ARTESP (Separar NC + templates Kartado / Kria):** as **informações de negócio** devem vir dos
    **Excel que acompanham cada PDF de apontamento** (planilha-mãe EAF ou export por fiscalização);
    o **PDF** entra sobretudo para **imagens** (Extrair PDF → nc/PDF .jpg), alinhado às macros adaptadas
    em fotos de campo / Separar NC — não como substituto do Excel para dados tabulares.
"""

from __future__ import annotations

import asyncio
import base64
import io
import json
import logging
import os
import platform
import re
import secrets
import shutil
import sys
import tempfile
import time
import traceback
import unicodedata
import zipfile
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Annotated, Any, Dict, List, Optional

from fastapi import APIRouter, Body, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, JSONResponse, Response, StreamingResponse
from starlette.background import BackgroundTask
try:
    from render_api.job_manager import carregar_job_nc as job_manager_carregar
except ImportError:
    job_manager_carregar = None

logger = logging.getLogger(__name__)

MAX_MB = 200
MAX_BYTES = MAX_MB * 1024 * 1024

# Tenta, em ordem:
#   1. Variável de ambiente ARTESP_NC_PROJ (configurável no Render)
#   2. Pasta nc_artesp/ dentro do próprio repositório (deploy Render)
#   3. Caminho Windows local (desenvolvimento desktop)
def _resolver_nc_proj() -> Path:
    # 1. Env var (mais flexível — configure no painel do Render se necessário)
    env = (
        __import__("os").getenv("ARTESP_NC_PROJ") or ""
    ).strip()
    if env:
        p = Path(env)
        if p.exists():
            return p

    # 2. Dentro do repo: GeradorARTESP/nc_artesp/
    repo_path = Path(__file__).resolve().parent.parent / "nc_artesp"
    if repo_path.exists():
        return repo_path

    # 3. Fallback local Windows (só funciona no desktop do dev)
    win_path = Path(r"C:\AUTOMAÇÃO_MACROS\Macros Kcor Ellen\artesp_nc_v2.0")
    return win_path  # pode não existir no Render — _nc_proj_disponivel() verifica


_NC_PROJ   = _resolver_nc_proj()
# Raiz do repositório (para importar nc_artesp como pacote)
_REPO_ROOT = Path(__file__).resolve().parent.parent

# Sempre resolve pelo repositório, independente do caminho do projeto local.
try:
    from nc_artesp.utils.helpers import (
        EXPORTAR_KARTADO_MAE_SUBDIR,
        extrair_zipfile_para_pasta,
        sanitizar_nome,
        truncar_nome_preservando_sufixo_prazo_m01,
    )
except ImportError:  # execução com cwd em nc_artesp
    from utils.helpers import (  # type: ignore[no-redef]
        EXPORTAR_KARTADO_MAE_SUBDIR,
        extrair_zipfile_para_pasta,
        sanitizar_nome,
        truncar_nome_preservando_sufixo_prazo_m01,
    )

def _nc_zip_stage1_nc_separados(
    zf: zipfile.ZipFile,
    stage1_dir: Path,
    arqs_all: list,
    usados: set[str],
) -> None:
    """Empacota saídas M01 (nomes planos) + ``exportar/*.xlsx`` (planilha-mãe pós-Kartado)."""
    for a in arqs_all:
        p = Path(a)
        if p.exists():
            arc = _nc_arcnome_zip_para_extracao_windows(p.name, usados=usados)
            zf.write(p, arc)
    sub = stage1_dir / EXPORTAR_KARTADO_MAE_SUBDIR
    if not sub.is_dir():
        return
    for f in sorted(sub.glob("*.xlsx")):
        if not f.is_file() or f.name.startswith("~"):
            continue
        try:
            rel_arc = f.relative_to(stage1_dir).as_posix()
        except ValueError:
            rel_arc = f"{EXPORTAR_KARTADO_MAE_SUBDIR}/{f.name}"
        arc = _nc_arcnome_zip_para_extracao_windows(rel_arc, usados=usados)
        zf.write(f, arc)


_NC_ASSETS = Path(__file__).resolve().parent.parent / "nc_artesp" / "assets" / "templates"
_FOTOS_ASSETS = Path(__file__).resolve().parent.parent / "fotos_campo" / "assets"
_NC_ARTEMIG_TEMPLATES = Path(__file__).resolve().parent.parent / "nc_artemig" / "assets" / "Template"


def _pastas_artemig_busca() -> list[Path]:
    """XLSX Artemig estão em Template/templates/; mantém Template/ para retrocompat."""
    t = _NC_ARTEMIG_TEMPLATES
    return [p for p in (t / "templates", t) if p.is_dir()]


def _lote_num_texto(lote: Optional[str]) -> str:
    s = (lote or "").strip()
    if not s:
        return ""
    m_art = re.search(r"(?:^|[^0-9])(26|21|13)(?:[^0-9]|$)", s, re.IGNORECASE)
    if m_art:
        return m_art.group(1)
    m = re.search(r"\d+", s)
    return m.group(0) if m else ""


def _m01_kartado_ativo_para_lote(flag_kartado: bool, lote: Optional[str]) -> bool:
    if not flag_kartado:
        return False
    from nc_artesp.config import M01_LOTE, lote_eh_kartado_artesp

    lote_num = _lote_num_texto(lote) or _lote_num_texto(M01_LOTE)
    if lote_num == "50":
        logger.info("M01 Kartado desligado: lote 50 (Artemig).")
        return False
    if lote_eh_kartado_artesp(lote_num):
        return True
    logger.warning(
        "M01 Kartado pedido (m01_kartado=true) mas o lote «%s» não é ARTESP 13, 21 nem 26. "
        "Indique `lote` no pedido ou ARTESP_LOTE / M01_LOTE no ambiente. A usar Art_011 (cópia mãe).",
        lote_num or "vazio",
    )
    return False

# ── Nomes de pastas dos relatórios (alinhados às macros nc_artesp/config.py) ───
DIR_EXPORTAR = "Exportar"
DIR_IMAGENS_PDF = "Imagens Provisórias - PDF"
DIR_KARTADO_RELATORIO_FOTOS = "Kartado_relatorio_fotos"
DIR_RESPOSTAS_KARTADO_FOTOS = "respostas_kartado_fotos"
DIR_RESPOSTAS_PENDENTES = "Respostas Pendentes"
# Pasta de respostas M02 (pendentes) no pacote final (ZIP).
DIR_PENDENTES_ENTREGA = "pendentes"
DIR_IMAGENS_CONSERVACAO = "Imagens Conservação"
DIR_CONSERVACAO = "Conservação"
DIR_IMAGENS_MA = "Imagens Meio Ambiente"
DIR_MA = "Meio Ambiente"
DIR_ACUMULADO = "Acumulado"
DIR_KCOR_CONSERVACAO = "Kcor Conservação"
# ZIP único (Excel + fotos) gerado após M02; copiado para entrega `Kartado/`.
_SUB_PACOTES_KARTADO_M01 = "_pacotes_kartado"
_NC_KARTADO_ZIP_STEM = "Kartado NCs Consolidadas"
# Nome curto do .xlsx **dentro** do ZIP (Explorer 0x80010135); o ficheiro em disco mantém o nome M01/ART011.
_NC_KARTADO_ZIP_INTERNO_XLSX = "Kartado Consolidado.xlsx"
_SUB_PACOTES_KRIA_KTD = "_pacotes_kria_ktd"
# Pasta no ZIP final: ZIPs M02 (modelo Kria + fotos), estilo macro KTD — não misturar com Kartado/
DIR_PACOTES_KRIA_KTD_ENTREGA = "Pacotes_KTD"

# Profundidade típica no ZIP ``_entrega/`` (caminho relativo = segmentos separados por «/»):
#   Kartado_relatorio_fotos | pendentes / ficheiro → 2 níveis
#   Kartado / ficheiro → 2; exportar / … / ficheiro (rglob) → ≥2; acumulado | calendario | emails | backup / ficheiro → 2
# O Explorador (0x80010135) limita o caminho completo na extração; o nome interno no .zip fica ≤ _NC_ZIP_MAX_ARC_TOTAL.
NC_ENTREGA_ZIP_PROFUNDIDADE_RESPOSTAS = 2

# Estratégias: (1) workspace por job + descarte controlado (2) apagar stage1/2 após sucesso
# (3) retenção por estado: running nunca, finished 72h, failed 24h (4) ZIP final único
# Checklist: cada job tem pasta própria | stage intermediário descartável | retenção por estado
#   | ZIP final único | sem duplicação | job.json com log resumido | limpeza automática ativa
# OUTPUT_PATH para NC: mesmo critério do app (ARTESP_OUTPUT_DIR ou defaults).
NC_SUBDIRS = ("input", "stage1", "stage2", "final")


def _nc_output_path() -> Path:
    """Base de outputs para NC — mesmo critério do app (OUTPUT_PATH)."""
    env = (os.getenv("ARTESP_OUTPUT_DIR") or "").strip()
    if env:
        return Path(env).resolve()
    if platform.system() == "Linux":
        return Path("/tmp/outputs").resolve()  # /tmp tem ~400 GB; /data tem só 1 GB
    return (Path(__file__).resolve().parent.parent / "outputs").resolve()


def _safe_nc_job_dir(job_id: str) -> Path:
    """
    Retorna o diretório do job NC garantindo que está sob OUTPUT_PATH/nc/.
    Valida path traversal: job_id não pode conter .., / ou \\. Usa .resolve() e
    relative_to(base) para garantir que job_dir fica dentro de OUTPUT_PATH/nc.
    """
    if not job_id or ".." in job_id or "/" in job_id or "\\" in job_id:
        raise HTTPException(status_code=400, detail="job_id inválido.")
    base = _nc_output_path() / "nc"
    job_dir = (base / job_id).resolve()
    base_resolved = base.resolve()
    try:
        job_dir.relative_to(base_resolved)
    except ValueError:
        raise HTTPException(status_code=400, detail="job_id inválido.")
    return job_dir


@dataclass(frozen=True)
class NCWorkspace:
    """Workspace por execução (job): job_dir + subpastas input/, stage1/, stage2/, final/."""

    job_id: str
    job_dir: Path
    input: Path
    stage1: Path
    stage2: Path
    final: Path

    def ensure_dirs(self) -> None:
        for p in (self.input, self.stage1, self.stage2, self.final):
            p.mkdir(parents=True, exist_ok=True)


def _gerar_job_id() -> str:
    """Identificador único por execução: nc_YYYYMMDD_HHMMSS_<suffix> (auditoria/rastreabilidade)."""
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    suffix = secrets.token_hex(4)
    return f"nc_{ts}_{suffix}"


def create_nc_workspace() -> NCWorkspace:
    """
    Cria um workspace por execução (job_id) sob OUTPUT_PATH/nc/<job_id>/.
    Subpastas: input/, stage1/, stage2/, final/.
    Pipeline stateful: arquivos podem ser guardados entre etapas sem re-upload.
    Remove jobs expirados antes de criar o novo workspace para evitar disco cheio.
    """
    _cleanup_expired_jobs()
    job_id = _gerar_job_id()
    job_dir = _safe_nc_job_dir(job_id)
    job_dir.mkdir(parents=True, exist_ok=True)
    ws = NCWorkspace(
        job_id=job_id,
        job_dir=job_dir,
        input=job_dir / "input",
        stage1=job_dir / "stage1",
        stage2=job_dir / "stage2",
        final=job_dir / "final",
    )
    ws.ensure_dirs()
    _update_job_json(ws, status="created")
    logger.info("NC workspace criado: job_id=%s path=%s", job_id, job_dir)
    return ws


def resolve_nc_workspace(job_id: str) -> NCWorkspace:
    """Resolve workspace existente por job_id (não cria diretórios)."""
    job_dir = _safe_nc_job_dir(job_id)
    return NCWorkspace(
        job_id=job_id,
        job_dir=job_dir,
        input=job_dir / "input",
        stage1=job_dir / "stage1",
        stage2=job_dir / "stage2",
        final=job_dir / "final",
    )


def resolve_workspace(job_id: Optional[str] = None) -> tuple:
    """
    Regra de ouro: etapa isolada cria job / pipeline reutiliza job.
    - job_id None ou "" → cria workspace novo (etapa isolada), retorna (ws, True).
    - job_id preenchido → abre workspace existente (pipeline), touch, retorna (ws, False).
    """
    if job_id and str(job_id).strip():
        j = str(job_id).strip()
        ws = resolve_nc_workspace(j)
        if not ws.job_dir.is_dir():
            raise HTTPException(404, detail="Workspace não encontrado. Execute a primeira etapa (ex.: separar) sem job_id.")
        _update_job_json(ws, status="running")
        return ws, False
    ws = create_nc_workspace()
    _update_job_json(ws, status="running", stage="stage1")
    return ws, True


def _artifacts_for_stage(path: Path, base: Path) -> List[str]:
    """Lista paths relativos (sempre com /) de arquivos em path. Nunca retorna absolutos."""
    if not path.is_dir():
        return []
    out = []
    for p in path.rglob("*"):
        if p.is_file():
            try:
                rel = p.relative_to(base)
                # Normaliza barras (Windows) e evita vazar estrutura do servidor
                out.append(str(rel).replace("\\", "/"))
            except ValueError:
                out.append(p.name)
    return sorted(out)


def _nc_response(
    ws: NCWorkspace,
    stage: str,
    *,
    download_urls: Optional[List[str]] = None,
    final_files: Optional[List[str]] = None,
    artifacts: Optional[Dict[str, List[str]]] = None,
    step_label: Optional[str] = None,
    next_step_label: Optional[str] = None,
) -> Dict[str, Any]:
    """Resposta padrão da API NC: job_id, stage, artifacts, download_urls. step_label/next_step_label = nomes das etapas (iguais aos botões) para exibir no frontend."""
    prefix = f"/outputs/nc/{ws.job_id}"
    if artifacts is None:
        artifacts = {}
    for name, dir_path in [("stage1", ws.stage1), ("stage2", ws.stage2), ("final", ws.final)]:
        if name not in artifacts and dir_path.is_dir():
            artifacts[name] = _artifacts_for_stage(dir_path, ws.job_dir)
    payload: Dict[str, Any] = {
        "ok": True,
        "job_id": ws.job_id,
        "stage": stage,
        "artifacts": artifacts,
    }
    if download_urls:
        payload["download_urls"] = [u if u.startswith("/") else f"{prefix}/{u}".lstrip("/") for u in download_urls]
    if final_files:
        payload["final_files"] = final_files
    if step_label is not None:
        payload["step_label"] = step_label
    if next_step_label is not None:
        payload["next_step_label"] = next_step_label
    return payload


def _list_stage_files(path: Path) -> List[str]:
    """Lista nomes de arquivos em um stage (para job.json)."""
    if not path.is_dir():
        return []
    return sorted(p.name for p in path.iterdir() if p.is_file())


def _update_job_json(
    ws: NCWorkspace,
    status: Optional[str] = None,
    stage: Optional[str] = None,
    stages: Optional[Dict[str, List[str]]] = None,
    log_summary: Optional[Dict[str, Any]] = None,
    retain_hours: Optional[float] = None,
) -> None:
    """
    Atualiza job.json: last_access (limpeza), status, stage, stages, log resumido, retain_until.
    Regra: nunca depender de memória; estado explícito no disco.
    """
    job_json = ws.job_dir / "job.json"
    now = time.time()
    now_iso = datetime.now(timezone.utc).isoformat()
    data: Dict[str, Any] = {
        "job_id": ws.job_id,
        "last_access": now,
        "last_access_iso": now_iso,
    }
    if job_json.is_file():
        try:
            with open(job_json, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (json.JSONDecodeError, OSError):
            pass
    data["last_access"] = now
    data["last_access_iso"] = now_iso
    if "created_at" not in data:
        data["created_at"] = now
    if "created_at_iso" not in data:
        data["created_at_iso"] = datetime.now(timezone.utc).isoformat()
    if status is not None:
        data["status"] = status
    if stage is not None:
        data["stage"] = stage
    if log_summary is not None:
        data["log"] = log_summary
    if retain_hours is not None and retain_hours > 0:
        retain_until = datetime.now(timezone.utc) + timedelta(hours=retain_hours)
        data["retain_until"] = retain_until.isoformat()
        data["retain_until_ts"] = (now + retain_hours * 3600)
    if stages is not None:
        data["stages"] = stages
    else:
        data["stages"] = {
            "input": _list_stage_files(ws.input),
            "stage1": _list_stage_files(ws.stage1),
            "stage2": _list_stage_files(ws.stage2),
            "final": _list_stage_files(ws.final),
        }
    try:
        with open(job_json, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
    except OSError as e:
        logger.warning("Não foi possível atualizar job.json: %s", e)


def _touch_job_access(ws: NCWorkspace) -> None:
    """Atualiza last_access em job.json (compatibilidade; usa _update_job_json)."""
    _update_job_json(ws)


def _safe_input_filename(nome: str) -> str:
    """Nome seguro para salvar em input/ (evita path traversal e caracteres inválidos)."""
    if not nome or ".." in nome or "/" in nome or "\\" in nome:
        return "arquivo.xlsx"
    # mantém só caracteres seguros
    safe = "".join(c for c in nome if c.isalnum() or c in "._- ")
    safe = safe.strip() or "arquivo.xlsx"
    # Sufixo «~1» de duplicado Windows / ZIP antigo — não faz parte do nome macro das constatações.
    p = Path(safe)
    if p.suffix:
        st = re.sub(r"~\d+$", "", p.stem).rstrip(" -.")
        if st:
            safe = f"{st}{p.suffix}"
    return safe


def _purge_dir_contents(path: Path) -> None:
    """Remove todo o conteúdo de um diretório (mantém a pasta). Facilita descarte de stage1/ e stage2/."""
    if not path.is_dir():
        return
    for p in path.iterdir():
        try:
            if p.is_file():
                p.unlink()
            elif p.is_dir():
                shutil.rmtree(p)
        except OSError as e:
            logger.warning("Purge %s: %s", p, e)


# TTL de fallback: jobs sem retain_until ou sem job.json são removidos após este tempo.
_JOB_TTL_FALLBACK_HOURS = 2.0


def _cleanup_expired_jobs() -> int:
    """
    Remove jobs expirados em OUTPUT_PATH/nc/.

    Critério de expiração (qualquer um):
      1. job.json presente e retain_until_ts < now  →  expirou conforme TTL declarado
      2. job.json presente, sem retain_until_ts e last_access > _JOB_TTL_FALLBACK_HOURS  →  expirou por inatividade
      3. job.json ausente e mtime do diretório > _JOB_TTL_FALLBACK_HOURS  →  job abandonado

    Retorna o número de diretórios removidos.
    """
    base = _nc_output_path() / "nc"
    if not base.is_dir():
        return 0
    now = time.time()
    fallback_secs = _JOB_TTL_FALLBACK_HOURS * 3600
    removed = 0
    for job_dir in list(base.iterdir()):
        if not job_dir.is_dir():
            continue
        try:
            job_json = job_dir / "job.json"
            expired = False
            if job_json.is_file():
                try:
                    with open(job_json, "r", encoding="utf-8") as f:
                        meta = json.load(f)
                    retain_ts = meta.get("retain_until_ts")
                    if retain_ts is not None:
                        expired = now > float(retain_ts)
                    else:
                        last = meta.get("last_access", 0)
                        expired = (now - float(last)) > fallback_secs
                except (json.JSONDecodeError, OSError, ValueError):
                    # JSON corrompido — usar mtime do diretório
                    expired = (now - job_dir.stat().st_mtime) > fallback_secs
            else:
                expired = (now - job_dir.stat().st_mtime) > fallback_secs

            if expired:
                shutil.rmtree(job_dir, ignore_errors=True)
                removed += 1
                logger.info("Cleanup: job removido %s", job_dir.name)
        except OSError as e:
            logger.warning("Cleanup: erro ao verificar %s: %s", job_dir, e)
    if removed:
        logger.info("Cleanup NC: %d job(s) expirado(s) removido(s)", removed)
    return removed


def _pasta_tem_imagens_jpg_recursiva(p: Path) -> bool:
    """True se existir pelo menos um ficheiro .jpg/.jpeg (qualquer caixa) sob p."""
    if not p.is_dir():
        return False
    for sub in p.rglob("*"):
        if sub.is_file() and sub.suffix.lower() in (".jpg", ".jpeg"):
            return True
    return False


_EXT_IMAGEM_BACKUP = frozenset(
    {".jpg", ".jpeg", ".jpe", ".png", ".webp", ".gif", ".bmp", ".tif", ".tiff"}
)


def _nc_listar_imagens_extraidas_pdf(pasta_raiz: Path) -> list[Path]:
    """Ficheiros de imagem sob a pasta de extração PDF (estrutura relativa preservada no ZIP backup)."""
    if not pasta_raiz.is_dir():
        return []
    out: list[Path] = []
    for f in pasta_raiz.rglob("*"):
        if f.is_file() and f.suffix.lower() in _EXT_IMAGEM_BACKUP:
            out.append(f)
    return sorted(out)


def _nc_zip_backup_desde_bytes_zip_interno(zip_bytes: bytes, zip_out: Path) -> int:
    """
    Grava ``zip_out`` copiando do ZIP em memória apenas membros que são imagens.
    Usado para o backup embutido no pacote final: o ZIP de entrada é o **gerado no servidor**
    ao extrair os PDFs do fluxo (``_nc_pdf_paths_para_zip_imagens_bytes_sync``), não um upload
    separado de imagens pelo cliente.
    """
    if not zip_bytes or len(zip_bytes) < 22:
        return 0
    try:
        with zipfile.ZipFile(io.BytesIO(zip_bytes), "r") as zin:
            usados: set[str] = set()
            entradas: list[tuple[str, bytes]] = []
            for info in zin.infolist():
                if info.is_dir():
                    continue
                name = (info.filename or "").replace("\\", "/").strip()
                if not name or name.endswith("/"):
                    continue
                if Path(name).suffix.lower() not in _EXT_IMAGEM_BACKUP:
                    continue
                try:
                    data = zin.read(info)
                except (RuntimeError, zipfile.BadZipFile, OSError) as e:
                    logger.warning("Backup imagens (ZIP interno): falha ao ler %s: %s", name, e)
                    continue
                entradas.append((name, data))
            if not entradas:
                return 0
            zip_out.parent.mkdir(parents=True, exist_ok=True)
            with zipfile.ZipFile(zip_out, "w", zipfile.ZIP_DEFLATED) as zout:
                for name, data in entradas:
                    arc = _nc_arcnome_zip_para_extracao_windows(
                        name.replace("\\", "/"),
                        usados=usados,
                    )
                    zout.writestr(arc, data, compress_type=zipfile.ZIP_DEFLATED)
            return len(entradas)
    except zipfile.BadZipFile:
        return 0


def _nc_extrair_zip_para_pasta_seguro(zip_bytes: bytes, destino: Path) -> int:
    """
    Extrai membros de um ZIP para ``destino`` evitando path traversal (equivalente seguro a extractall).
    Retorna o número de ficheiros escritos.
    """
    if not zip_bytes:
        return 0
    destino.mkdir(parents=True, exist_ok=True)
    root = destino.resolve()
    n = 0
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            name = (info.filename or "").replace("\\", "/").strip()
            if not name or name.endswith("/"):
                continue
            parts = tuple(p for p in Path(name).as_posix().split("/") if p and p != ".." and p != ".")
            if not parts:
                continue
            rel = Path(*parts)
            target = (destino / rel).resolve()
            try:
                target.relative_to(root)
            except ValueError:
                logger.warning("ZIP imagens: membro ignorado (caminho fora da pasta destino): %s", name)
                continue
            target.parent.mkdir(parents=True, exist_ok=True)
            with zf.open(info, "r") as src, open(target, "wb") as dst:
                shutil.copyfileobj(src, dst)
            n += 1
    return n


def _nc_zip_imagens_extraidas_backup(
    pasta_work: Path,
    pasta_fotos_pdf: Path,
    zip_path: Path,
    *,
    zip_bytes_pipeline_pdf: Optional[bytes] = None,
) -> int:
    """
    ZIP de backup dentro do pacote total.

    1. **ZIP interno** produzido no servidor ao processar os PDFs do pedido (fluxo típico
       ``/nc/completo`` / ``/nc/separar`` com PDFs — não exige ZIP de imagens enviado pelo cliente).
    2. Ficheiros já extraídos para «Imagens Provisórias - PDF» (M02 / e-mail).
    3. Fallback: imagens em todo ``pasta_work`` quando (2) está vazio mas houve bytes em (1).
    """
    if zip_bytes_pipeline_pdf:
        n_mem = _nc_zip_backup_desde_bytes_zip_interno(zip_bytes_pipeline_pdf, zip_path)
        if n_mem > 0:
            logger.info(
                "Backup imagens: %d ficheiro(s) a partir do ZIP interno (extração dos PDFs no servidor).",
                n_mem,
            )
            return n_mem

    files = _nc_listar_imagens_extraidas_pdf(pasta_fotos_pdf)
    arc_base = pasta_fotos_pdf
    if not files and zip_bytes_pipeline_pdf and pasta_work.is_dir():
        files = _nc_listar_imagens_extraidas_pdf(pasta_work)
        arc_base = pasta_work
        if files:
            logger.info(
                "Backup imagens: %d ficheiro(s) em %s (fallback em disco; ZIP interno sem imagens listáveis).",
                len(files),
                pasta_work,
            )
    if not files:
        if zip_bytes_pipeline_pdf:
            logger.warning(
                "Backup imagens: ZIP interno presente mas sem imagens reconhecidas; pasta PDF=%s existe=%s.",
                pasta_fotos_pdf,
                pasta_fotos_pdf.is_dir(),
            )
        return 0
    zip_path.parent.mkdir(parents=True, exist_ok=True)
    usados: set[str] = set()
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in files:
            try:
                arc = f.relative_to(arc_base).as_posix()
            except ValueError:
                arc = f.name
            arc = _nc_arcnome_zip_para_extracao_windows(arc, usados=usados)
            zf.write(f, arc)
    return len(files)


def _lista_uploads_eaf(
    arquivo: Optional[UploadFile],
    arquivos: Optional[List[UploadFile]],
) -> List[UploadFile]:
    """Junta campo legado `arquivo` com lista `arquivos` (multipart repetido)."""
    out: List[UploadFile] = []
    if arquivo and getattr(arquivo, "filename", None):
        out.append(arquivo)
    if arquivos:
        for u in arquivos:
            if u and getattr(u, "filename", None):
                out.append(u)
    return out


def _nc_norm_header_celula(s: str) -> str:
    """Normaliza texto de cabeçalho (igual ao critério em pacotes Kartado / e-mail NC)."""
    t = unicodedata.normalize("NFD", str(s or ""))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", t).strip().lower()


def _nc_workbook_primeira_linha_eh_layout_kartado(xls: Path) -> bool:
    """
    True se a folha ativa tiver cabeçalho Kartado na linha 1 (código de fiscalização).
    Layout EAF / Art_011 (planilha-mãe) não coincide — nesse caso o pacote Kartado deve vir do M02.
    """
    try:
        from openpyxl import load_workbook  # import local
    except Exception:
        return False
    try:
        wb = load_workbook(str(xls), read_only=True, data_only=True)
        try:
            ws = wb.active
            max_c = min(int(ws.max_column or 0), 256)
            hdr_keys: set[str] = set()
            for c in range(1, max_c + 1):
                v = ws.cell(row=1, column=c).value
                if v is None:
                    continue
                k = _nc_norm_header_celula(v)
                if k:
                    hdr_keys.add(k)
        finally:
            wb.close()
        if ("codigo de fiscalizacao" in hdr_keys) or ("codigo fiscalizacao" in hdr_keys):
            return True
        for k in hdr_keys:
            if "codigo" in k and "fiscal" in k:
                return True
        return False
    except Exception:
        return False


def _nc_exportar_contem_excel_layout_kartado(pasta_xls: Path) -> bool:
    """Há pelo menos um .xlsx de saída M01 no layout Kartado (não planilha-mãe EAF)."""
    if not pasta_xls.is_dir():
        return False
    for f in pasta_xls.rglob("*.xlsx"):
        if f.name.startswith("~") or f.name.startswith("_"):
            continue
        if _nc_workbook_primeira_linha_eh_layout_kartado(f):
            return True
    return False


def _nc_zip_stem_fallback_constatacao(stem: str) -> str:
    """
    Quando a coluna Classe não pôde ser lida: extrai o serviço curto do nome do Excel,
    no padrão Art_011 / M01: «… (RODOVIA - SERVIÇO_ABREV) - Prazo - …» → SERVIÇO_ABREV.
    """
    if not stem:
        return ""
    m = re.search(r"\(([^)]+)\)", stem)
    if not m:
        return ""
    interior = m.group(1).strip()
    if " - " in interior:
        return interior.rsplit(" - ", 1)[-1].strip()
    return interior


def _nc_gravar_pacotes_kria_ktd_zip(pacotes: Any, pasta_pacotes: Path) -> None:
    """
    Para cada saída **M02 (modelo Kria)**, cria um ZIP com o .xlsx Kria e imagens nc/PDF.
    Nome do .zip alinhado à **Art_03_KTD** (fluxo Kria — não confundir com pacotes do Excel Kartado do M01).
    O ``zip_stem`` vem de ``gerar_modelo_foto``; se faltar, usa-se o stem do ficheiro Kria.
    """
    if not pacotes:
        return
    pasta_pacotes.mkdir(parents=True, exist_ok=True)
    for pkg in pacotes:
        kria_p = pkg.get("kria") if isinstance(pkg, dict) else None
        if not kria_p:
            continue
        kria_p = Path(kria_p)
        if not kria_p.is_file():
            continue
        imgs = pkg.get("imagens") if isinstance(pkg, dict) else None
        imgs = imgs or []
        raw_zip = ""
        if isinstance(pkg, dict):
            zs = pkg.get("zip_stem")
            if zs is not None and str(zs).strip():
                raw_zip = str(zs).strip()
        stem = ""
        if raw_zip:
            stem = sanitizar_nome(raw_zip, max_len=180).strip()
            stem = _nc_truncar_nome_zip(stem, 160) if stem else ""
        if not stem:
            stem = _nc_truncar_nome_zip(kria_p.stem, 100)
        if not stem:
            stem = "pacote"
        zip_path = pasta_pacotes / f"{stem}.zip"
        n = 1
        while zip_path.exists():
            zip_path = pasta_pacotes / f"{stem}_{n}.zip"
            n += 1
        inner_used: set[str] = set()
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            arc_x = _nc_arcnome_zip_para_extracao_windows(kria_p.name, usados=inner_used)
            zf.write(kria_p, arc_x)
            for img in imgs:
                ip = Path(img)
                if not ip.is_file():
                    continue
                arc = _nc_arcnome_zip_para_extracao_windows(f"fotos/{ip.name}", usados=inner_used)
                zf.write(ip, arc)


def _nc_m01_kartado_consolidar_multiplos_excels(
    mod: Any,
    arqs_all: list[Path | str],
    pasta_destino: Path,
    *,
    m01_kartado: bool,
) -> list[Path]:
    paths = [Path(a) for a in arqs_all if Path(a).exists()]
    if not m01_kartado or len(paths) <= 1:
        return paths
    if not all(_nc_workbook_primeira_linha_eh_layout_kartado(p) for p in paths):
        return paths
    try:
        merged = mod.consolidar_kartados_em_unico_excel(paths, pasta_destino=pasta_destino)
        if merged and merged.is_file():
            m_res = merged.resolve()
            for p in paths:
                if p.resolve() != m_res:
                    p.unlink(missing_ok=True)
            logger.info("M01 Kartado: consolidado em ficheiro único → %s", merged.name)
            return [merged]
    except Exception as e:
        logger.warning("M01 consolidação Kartado falhou (%s); usando ficheiros individuais.", e)
    return paths


def _nc_gravar_pacotes_kartado_de_m01(
    pasta_xls: Path, pasta_fotos_pdf: Optional[Path] = None
) -> None:
    """
    Após M02: funde Excels **layout Kartado** num único .xlsx se houver mais de um;
    grava ``_pacotes_kartado/{Kartado NCs Consolidadas}.zip`` com esse Excel e as imagens
    referenciadas (Foto_1 / Foto_2) quando existirem em ``pasta_fotos_pdf``.
    """
    if not pasta_xls.is_dir():
        return
    out_dir = pasta_xls / _SUB_PACOTES_KARTADO_M01
    if out_dir.is_dir():
        shutil.rmtree(out_dir, ignore_errors=True)
    out_dir.mkdir(parents=True, exist_ok=True)

    kartado_pre: list[Path] = []
    for x in sorted(pasta_xls.rglob("*.xlsx")):
        if x.name.startswith("~") or x.name.startswith("_"):
            continue
        if _nc_workbook_primeira_linha_eh_layout_kartado(x):
            kartado_pre.append(x)
    if not kartado_pre:
        return

    if len(kartado_pre) > 1:
        try:
            mod_sep = _importar_modulo("separar_nc")
            merged_p = mod_sep.consolidar_kartados_em_unico_excel(
                kartado_pre, pasta_destino=pasta_xls
            )
            if merged_p and merged_p.is_file():
                m_res = merged_p.resolve()
                for p in kartado_pre:
                    if p.resolve() != m_res:
                        p.unlink(missing_ok=True)
                logger.info(
                    "Kartado: %s Excels fundidos em único ficheiro → %s",
                    len(kartado_pre),
                    merged_p.name,
                )
            else:
                logger.warning(
                    "Kartado: fusão de %d Excels não gerou ficheiro único; mantêm-se os .xlsx individuais.",
                    len(kartado_pre),
                )
        except Exception as ex:
            logger.warning("Kartado: fusão multi-xlsx falhou (%s)", ex)

    xs = [
        p
        for p in sorted(pasta_xls.rglob("*.xlsx"))
        if p.is_file()
        and not p.name.startswith("~")
        and not p.name.startswith("_")
        and _nc_workbook_primeira_linha_eh_layout_kartado(p)
    ]
    if not xs:
        return
    prefer = [p for p in xs if "Kartado Consolidado" in p.stem]
    xls_k = prefer[0] if prefer else xs[0]
    if len(xs) > 1 and not prefer:
        logger.warning(
            "Kartado ZIP: %d Excels layout Kartado; a empacotar %s.",
            len(xs),
            xls_k.name,
        )

    fotos_idx: dict[str, Path] = {}
    if pasta_fotos_pdf and pasta_fotos_pdf.is_dir():
        for img in pasta_fotos_pdf.rglob("*"):
            if not img.is_file():
                continue
            if img.suffix.lower() not in (".jpg", ".jpeg"):
                continue
            k = img.name.strip().lower()
            if k and k not in fotos_idx:
                fotos_idx[k] = img

    def _norm_h(s: str) -> str:
        t = unicodedata.normalize("NFD", str(s or ""))
        t = "".join(c for c in t if unicodedata.category(c) != "Mn")
        return re.sub(r"\s+", " ", t).strip().lower()

    fotos_evento: list[Path] = []
    try:
        from openpyxl import load_workbook as _lw
    except Exception:
        _lw = None  # type: ignore[assignment, misc]
    if _lw is not None:
        try:
            wb = _lw(str(xls_k), read_only=False, data_only=True)
            ws = wb.active
            max_c = int(ws.max_column or 0)
            max_r = int(ws.max_row or 0)
            hdr: dict[str, int] = {}
            for c in range(1, max_c + 1):
                v = ws.cell(row=1, column=c).value
                if v is None:
                    continue
                k = _norm_h(str(v))
                if k and k not in hdr:
                    hdr[k] = c
            cf1 = hdr.get("foto_1")
            cf2 = hdr.get("foto_2")
            if cf1 or cf2:
                for r in range(2, max_r + 1):
                    for c in (cf1, cf2):
                        if not c:
                            continue
                        v = ws.cell(row=r, column=c).value
                        nm = str(v or "").strip()
                        if not nm:
                            continue
                        pth = fotos_idx.get(nm.lower())
                        if pth and pth.is_file():
                            fotos_evento.append(pth)
            wb.close()
        except Exception as exc:
            logger.warning("Kartado ZIP: leitura de %s falhou (%s)", xls_k.name, exc)

    zip_name = sanitizar_nome(f"{_NC_KARTADO_ZIP_STEM}.zip", max_len=180)
    if not zip_name.lower().endswith(".zip"):
        zip_name = f"{_NC_KARTADO_ZIP_STEM}.zip"
    zip_path = out_dir / zip_name
    inner_used: set[str] = set()
    try:
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            arc_x = _nc_arcnome_zip_para_extracao_windows(
                _NC_KARTADO_ZIP_INTERNO_XLSX, usados=inner_used
            )
            zf.write(xls_k, arc_x)
            seen_img: set[str] = set()
            for img in fotos_evento:
                try:
                    kimg = str(img.resolve())
                except OSError:
                    kimg = str(img)
                if kimg in seen_img:
                    continue
                seen_img.add(kimg)
                arc_i = _nc_arcnome_zip_para_extracao_windows(
                    img.name, usados=inner_used
                )
                zf.write(img, arc_i)
        logger.info(
            "Kartado: ZIP %s (Excel + %d imagem(ns)).",
            zip_path.name,
            len(seen_img),
        )
    except OSError as exc:
        logger.exception("Kartado: falha ao gravar %s — %s", zip_path.name, exc)


def _nc_gerar_acumulado_xlsx(
    pasta_input_eaf: Path,
    out_path: Path,
    pasta_fallback: Optional[Path] = None,
) -> bool:
    """
    Gera «Acumulado.xlsx» no **layout Kcor-Kria** (template ``Acumulado.xlsx`` ou ``_Planilha Modelo Kcor-Kria``),
    a partir de EAF em ``input/``. Se não houver linhas úteis, tenta ``pasta_fallback`` (ex. Exportar
    extraído do ZIP do M01 — ficheiros Kartado com dados na linha 2).
    """
    mod = _importar_modulo("juntar_arquivos")
    for pasta in (pasta_input_eaf, pasta_fallback):
        if pasta is None or not pasta.is_dir():
            continue
        try:
            if mod.gerar_acumulado_kcor_kria_desde_pasta_eaf(pasta, out_path, None):
                if pasta != pasta_input_eaf:
                    logger.info("Acumulado Kcor-Kria gerado a partir de %s (fallback).", pasta)
                return True
        except HTTPException:
            raise
        except Exception as exc:
            logger.warning("Acumulado Kcor-Kria (EAF) em %s: %s", pasta, exc)
    return False


def _nc_copiar_kartado_para_entrega(origem_exportar: Path, destino: Path) -> None:
    """Copia o ZIP Kartado (Excel + imagens) para ``Kartado/``; se não houver ZIP, o Excel .xlsx."""
    if not origem_exportar.is_dir():
        return
    destino.mkdir(parents=True, exist_ok=True)
    sub = origem_exportar / _SUB_PACOTES_KARTADO_M01
    alvo = sanitizar_nome(f"{_NC_KARTADO_ZIP_STEM}.zip", max_len=180)
    if not alvo.lower().endswith(".zip"):
        alvo = f"{_NC_KARTADO_ZIP_STEM}.zip"
    if sub.is_dir():
        for z in sub.glob("*.zip"):
            if z.name == alvo or z.stem == Path(alvo).stem:
                shutil.copy2(z, destino / z.name)
                return
        zips = sorted(sub.glob("*.zip"), key=lambda p: p.name.lower())
        if zips:
            z = zips[0]
            shutil.copy2(z, destino / z.name)
            return
    xs = [
        p
        for p in sorted(origem_exportar.rglob("*.xlsx"))
        if p.is_file()
        and not p.name.startswith("~")
        and not p.name.startswith("_")
        and _nc_workbook_primeira_linha_eh_layout_kartado(p)
    ]
    if not xs:
        return
    prefer = [p for p in xs if "Kartado Consolidado" in p.stem]
    ficheiro = prefer[0] if prefer else xs[0]
    if len(xs) > 1 and not prefer:
        logger.warning(
            "Kartado entrega: %d ficheiros layout Kartado sem nome «Kartado Consolidado»; a copiar %s.",
            len(xs),
            ficheiro.name,
        )
    shutil.copy2(ficheiro, destino / ficheiro.name)


def _nc_copiar_pacotes_kria_ktd_para_entrega(work: Path, destino_respostas_kria: Path) -> None:
    """ZIPs M02 (modelo abertura + imagens, estilo KTD) → ``respostas_kartado_fotos/Pacotes_KTD/`` — separado de Kartado/."""
    if not work.is_dir():
        return
    sub = work / _SUB_PACOTES_KRIA_KTD
    if not sub.is_dir():
        return
    zips = list(sub.glob("*.zip"))
    if not zips:
        return
    # Regra de entrega: Pacotes_KTD é irmão de Kartado_relatorio_fotos/, nunca subpasta desta.
    base_entrega = destino_respostas_kria
    _n_sub = destino_respostas_kria.name.strip().lower()
    if _n_sub in ("kria", "kartado_relatorio_fotos", DIR_KARTADO_RELATORIO_FOTOS.lower()):
        base_entrega = destino_respostas_kria.parent
    out_p = base_entrega / DIR_PACOTES_KRIA_KTD_ENTREGA
    out_p.mkdir(parents=True, exist_ok=True)

    # Evita duplicar conteúdo: se já existe planilha em respostas_kartado_fotos/Kartado_relatorio_fotos
    # com o mesmo stem do pacote, não copia o ZIP KTD equivalente.
    stems_kria: set[str] = set()
    pasta_kria = base_entrega / DIR_KARTADO_RELATORIO_FOTOS
    if pasta_kria.is_dir():
        for x in pasta_kria.rglob("*.xlsx"):
            stems_kria.add(x.stem.strip().lower())

    for z in zips:
        if z.stem.strip().lower() in stems_kria:
            continue
        shutil.copy2(z, out_p / z.name)


def _nc_copiar_xlsx_de_pasta(
    origem: Path,
    destino: Path,
    *,
    excluir_substr_no_nome: Optional[tuple[str, ...]] = None,
) -> None:
    """Copia .xlsx recursivamente (estrutura relativa); ignora ~$ e nomes que começam por '_'."""
    if not origem.is_dir():
        return
    excl = excluir_substr_no_nome or ()
    for f in origem.rglob("*.xlsx"):
        if f.name.startswith("~") or f.name.startswith("_"):
            continue
        if any(s in f.name for s in excl):
            continue
        rel = f.relative_to(origem)
        out = destino / rel
        out.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(f, out)


# Extração no Explorador falha se (pasta_destino + caminho_no_zip) > ~260 (sem \\?\).
# Nomes M01 Exportar (~80–120 chars) não podem ser cortados a 80: corrompem serviço e «Prazo».
_NC_ZIP_MAX_COMPONENTE = 150
_NC_ZIP_MAX_ARC_TOTAL = 175
_WIN_ZIP_RESERVED_STEMS = frozenset(
    {
        "CON",
        "PRN",
        "AUX",
        "NUL",
        "COM0",
        "COM1",
        "COM2",
        "COM3",
        "COM4",
        "COM5",
        "COM6",
        "COM7",
        "COM8",
        "COM9",
        "LPT0",
        "LPT1",
        "LPT2",
        "LPT3",
        "LPT4",
        "LPT5",
        "LPT6",
        "LPT7",
        "LPT8",
        "LPT9",
    }
)


def _nc_zip_stem_seguro_ficheiro(stem: str, fallback_stem: str) -> str:
    """Evita nome de ZIP vazio, só pontos ou reservado no Windows (CON, PRN, …)."""
    s = (stem or "").strip().strip(". ")
    if not s:
        s = (fallback_stem or "").strip().strip(". ")
    if not s:
        s = "Kartado_M01"
    base = Path(s).stem
    if base.upper() in _WIN_ZIP_RESERVED_STEMS:
        base = f"_{base}_kartado"
    if not base.strip("._- "):
        base = "Kartado_M01"
    return base


def _nc_truncar_nome_zip(nome: str, max_len: int = _NC_ZIP_MAX_COMPONENTE) -> str:
    """Encurta um componente de caminho dentro do ZIP, preservando sufixo M01 « - Prazo - data»."""
    return truncar_nome_preservando_sufixo_prazo_m01((nome or "").strip(), max_len)


def _nc_arcnome_artemig_lote50(stem: str, base_arq: str) -> str:
    """Lote 50 (Artemig): ZIP ↔ col. W ``Exportar Kcor`` coerentes.

    Regra única (single ou multi-PDF):
    - PDF integral: ``{stem}.pdf`` (sem prefixo extra).
    - JPG ``nc (cod).jpg`` e variantes `` (N).jpg``: ``{stem} nc (cod).jpg`` (espaço entre stem e ``nc``).

    Compat.: aceita prefixo antigo ``{stem}_`` ao ler nomes já gerados.

    Quando o ZIP interno ou um merge anterior já devolveu ``{stem}_{stem}.pdf``,
    o ramo de colapso não pode devolver o nome tal qual — reduz a ``{stem}.pdf``.
    """
    base_l = Path((base_arq or "").replace("\\", "/")).name
    if not (stem or "").strip():
        return base_l
    stem = sanitizar_nome(stem.strip(), max_len=200).strip() or stem.strip()
    if base_l.lower().endswith(".pdf"):
        s = sanitizar_nome(Path(base_l).stem, max_len=200).strip() or Path(base_l).stem
        if s == stem:
            return f"{stem}.pdf"
        t = s
        while True:
            if t.startswith(f"{stem}_"):
                t = t[len(stem) + 1 :]
            elif t.startswith(f"{stem} "):
                t = t[len(stem) + 1 :].lstrip()
            else:
                break
            if t == stem:
                return f"{stem}.pdf"
        return sanitizar_nome(base_l, max_len=250).strip() or base_l
    pref_space = f"{stem} "
    pref_us = f"{stem}_"
    if base_l.startswith(pref_space) or base_l.startswith(pref_us):
        return sanitizar_nome(base_l, max_len=250).strip() or base_l
    return sanitizar_nome(f"{pref_space}{base_l}", max_len=250).strip() or f"{pref_space}{base_l}"


def _nc_arcnome_zip_para_extracao_windows(
    arc_posix: str,
    *,
    usados: Optional[set[str]] = None,
) -> str:
    """
    Encurta o caminho relativo gravado no ZIP (forward slashes) para reduzir erros
    «caminho de destino demasiado longo» ao extrair no Explorador de ficheiros.
    """
    arc_posix = (arc_posix or "").replace("\\", "/").strip("/")
    parts = [p for p in arc_posix.split("/") if p]
    if not parts:
        return "ficheiro.bin"
    short_parts = [_nc_truncar_nome_zip(p, _NC_ZIP_MAX_COMPONENTE) for p in parts]
    arc = "/".join(short_parts)
    if len(arc) > _NC_ZIP_MAX_ARC_TOTAL:
        dirs_short = short_parts[:-1]
        leaf_orig = (parts[-1] or "").strip()
        prefix = "/".join(dirs_short)
        sep = 1 if prefix else 0
        budget = max(24, _NC_ZIP_MAX_ARC_TOTAL - len(prefix) - sep)
        leaf = truncar_nome_preservando_sufixo_prazo_m01(leaf_orig, budget)
        arc = f"{prefix}/{leaf}" if prefix else leaf
    if usados is None:
        return arc
    candidate = arc
    parent = Path(candidate).parent
    parent_s = parent.as_posix() if str(parent) not in (".", "") else ""
    base_stem = Path(candidate).stem
    ext = Path(candidate).suffix or ""
    k = 0
    while candidate in usados:
        k += 1
        dup = f" ({k})"
        # Espaço + parênteses (como em duplicados de ZIP no projeto), não «stem~1» (confunde com 8.3).
        budget = max(8, _NC_ZIP_MAX_COMPONENTE - len(dup) - len(ext))
        st = Path(
            truncar_nome_preservando_sufixo_prazo_m01(base_stem + ext, budget + len(ext))
        ).stem
        leaf = f"{st}{dup}{ext}"
        candidate = f"{parent_s}/{leaf}" if parent_s else leaf
    usados.add(candidate)
    return candidate


def _purge_work_if_finished(ws: NCWorkspace) -> None:
    """Remove stage2/_work para economizar disco quando o job foi finalizado."""
    work = ws.stage2 / "_work"
    if work.is_dir():
        try:
            shutil.rmtree(work)
        except OSError as e:
            logger.warning("Purge _work %s: %s", work, e)


def _garantir_path_nc() -> None:
    """
    Garante que o repositório e nc_artesp estejam em sys.path para:
    - 'from nc_artesp ...' (precisa do repo root)
    - 'from config import ...' / 'from utils ...' dentro dos módulos (precisa de nc_artesp).
    """
    repo = str(_REPO_ROOT)
    proj = str(_NC_PROJ)
    if repo not in sys.path:
        sys.path.insert(0, repo)
    if proj not in sys.path:
        sys.path.insert(0, proj)

# Nomes dos templates conforme config.py e arquivos em nc_artesp/assets/templates/
_NOME_MODELO_KRIA = "Modelo Abertura Evento Kria Conserva Rotina.xlsx"
_NOME_MODELO_RESP = "Modelo.xlsx"
_NOME_MODELO_KCOR = "_Planilha Modelo Kcor-Kria.xlsx"


def _pastas_busca_templates_nc() -> list[Path]:
    """Templates Kria/Kcor/EAF: nc_artesp/assets/templates e fotos_campo/assets/templates."""
    return [
        _NC_ASSETS,
        _FOTOS_ASSETS / "templates",
    ]


def _ler_asset(nome: str, pasta_base: Path | None = None) -> bytes:
    """Lê template: pasta_base (ex. Artemig) primeiro; depois nc_artesp/assets/templates e fotos_campo/assets/templates."""
    pastas = ([pasta_base] if pasta_base and pasta_base.is_dir() else []) + _pastas_busca_templates_nc()
    for pasta in pastas:
        if not pasta or not pasta.is_dir():
            continue
        p = pasta / nome
        if p.is_file():
            return p.read_bytes()
        nome_cf = nome.casefold()
        for f in pasta.iterdir():
            if f.is_file() and f.name.casefold() == nome_cf:
                return f.read_bytes()
    raise HTTPException(
        status_code=503,
        detail=(
            f"Template '{nome}' não encontrado. Coloque em nc_artesp/assets/templates/ "
            f"ou fotos_campo/assets/templates/."
        ),
    )


def _carregar_modelo_kria(lote: str | None = None) -> bytes:
    if (lote or "").strip() == "50":
        for pasta in _pastas_artemig_busca():
            try:
                return _ler_asset(_NOME_MODELO_KRIA, pasta)
            except HTTPException:
                continue
    return _ler_asset(_NOME_MODELO_KRIA)


def _carregar_modelo_resp(lote: str | None = None) -> bytes:
    pastas = _pastas_artemig_busca() if (lote or "").strip() == "50" else []
    for nome in (_NOME_MODELO_RESP, "Modelo Resposta.xlsx", "Modelo_Resposta.xlsx"):
        for pasta in pastas:
            try:
                return _ler_asset(nome, pasta)
            except HTTPException:
                continue
        try:
            return _ler_asset(nome)
        except HTTPException:
            continue
    raise HTTPException(
        status_code=503,
        detail=f"Template de Resposta não encontrado (tentou {_NOME_MODELO_RESP} e alternativas).",
    )


def _carregar_modelo_kcor(lote: str | None = None) -> bytes:
    if (lote or "").strip() == "50":
        try:
            from nc_artemig.config import MODELO_KCOR_KRIA

            p = Path(MODELO_KCOR_KRIA)
            if p.is_file():
                return p.read_bytes()
        except Exception:
            pass
        for pasta in _pastas_artemig_busca():
            try:
                return _ler_asset(_NOME_MODELO_KCOR, pasta)
            except HTTPException:
                continue
    return _ler_asset(_NOME_MODELO_KCOR)


def _check_auth(request: Request) -> dict:
    """Auth desabilitado para NC — acesso direto sem login."""
    return {}


def _ler(f: UploadFile) -> bytes:
    data = f.file.read()
    if len(data) > MAX_BYTES:
        raise HTTPException(413, f"Arquivo '{f.filename}' excede {MAX_MB} MB.")
    return data


def _gravar_pdf_observacao_ao_lado_do_eaf_kartado(
    eaf_path: Path,
    uploads_pdf: Optional[List[UploadFile]],
    *,
    m01_kartado_ativo: bool,
    num_eafs: int,
) -> None:
    """Grava ``<eaf>.pdf`` para o M01 Kartado enriquecer AA/X/Y a partir da «Observação» do PDF (``separar_nc`` lê ao lado da mãe)."""
    if not (m01_kartado_ativo and uploads_pdf):
        return
    blobs_pdf: list[bytes] = []
    for pf in uploads_pdf:
        if pf is None or not getattr(pf, "filename", None):
            continue
        blobs_pdf.append(_ler(pf))
        try:
            pf.file.seek(0)
        except Exception:
            pass
    if not blobs_pdf:
        return
    dest_pdf = eaf_path.with_suffix(".pdf")
    try:
        if len(blobs_pdf) == 1:
            dest_pdf.write_bytes(blobs_pdf[0])
        else:
            from nc_artesp.pdf_extractor import FITZ_OK, merge_pdfs_bytes

            if FITZ_OK:
                merged = merge_pdfs_bytes(blobs_pdf)
                dest_pdf.write_bytes(merged if merged else blobs_pdf[0])
            else:
                dest_pdf.write_bytes(blobs_pdf[0])
    except Exception as ex:
        logger.warning("nc: não gravou PDF ao lado do EAF (Descrição Kartado): %s", ex)


async def _nc_gravar_upload_pdf_com_limite(upload: UploadFile, dest: Path) -> None:
    """Grava upload em disco por chunks (evita segurar o PDF inteiro na RAM do event loop)."""
    dest.parent.mkdir(parents=True, exist_ok=True)
    total = 0
    chunk_sz = 1024 * 1024
    with dest.open("wb") as out:
        while True:
            chunk = await upload.read(chunk_sz)
            if not chunk:
                break
            total += len(chunk)
            if total > MAX_BYTES:
                try:
                    dest.unlink(missing_ok=True)
                except OSError:
                    pass
                raise HTTPException(
                    413,
                    detail=f"PDF '{upload.filename}' excede {MAX_MB} MB.",
                )
            out.write(chunk)


def _safe_filename_header(nome: str) -> str:
    """Nome seguro para Content-Disposition (evita erro latin-1 em headers HTTP)."""
    if not nome:
        return nome
    nfd = unicodedata.normalize("NFD", nome)
    sem_comb = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
    try:
        return sem_comb.encode("latin-1").decode("latin-1")
    except UnicodeEncodeError:
        return sem_comb.encode("latin-1", "replace").decode("latin-1")


def _nc_kartado_merge_avisos_payload() -> list[dict[str, str]]:
    try:
        from nc_artesp import config as _nc_cfg

        return list(getattr(_nc_cfg, "KARTADO_SERVICOS_MERGE_AVISOS", ()) or ())
    except Exception:
        return []


def _nc_kartado_merge_avisos_headers() -> dict[str, str]:
    av = _nc_kartado_merge_avisos_payload()
    if not av:
        return {}
    try:
        b64 = base64.b64encode(json.dumps(av, ensure_ascii=False).encode("utf-8")).decode("ascii")
    except Exception:
        return {}
    return {"X-NC-Kartado-Avisos": b64}


def _stream_zip(
    data: bytes,
    nome: str,
    job_id: Optional[str] = None,
    extra_headers: Optional[dict[str, str]] = None,
) -> StreamingResponse:
    """Stream ZIP a partir de bytes (evitar para ficheiros muito grandes — preferir ``_stream_zip_path``)."""
    headers = {"Content-Disposition": f'attachment; filename="{_safe_filename_header(nome)}"'}
    if job_id:
        headers["X-NC-Job-Id"] = job_id
    if extra_headers:
        for k, v in extra_headers.items():
            if k and v is not None:
                headers[str(k)] = str(v)
    return StreamingResponse(
        io.BytesIO(data), media_type="application/zip",
        headers=headers,
    )


def _stream_zip_path(
    zip_path: Path,
    nome: str,
    job_id: Optional[str] = None,
    extra_headers: Optional[dict[str, str]] = None,
) -> FileResponse:
    """Envia ZIP a partir do disco (Starlette/FastAPI leem o ficheiro em streaming; não carregam o ZIP inteiro na RAM)."""
    if not zip_path.is_file():
        raise HTTPException(status_code=500, detail="ZIP não encontrado para download.")
    headers = {"Content-Disposition": f'attachment; filename="{_safe_filename_header(nome)}"'}
    if job_id:
        headers["X-NC-Job-Id"] = job_id
    if extra_headers:
        for k, v in extra_headers.items():
            if k and v is not None:
                headers[str(k)] = str(v)
    return FileResponse(
        str(zip_path.resolve()),
        media_type="application/zip",
        headers=headers,
    )


def _stream_xlsx(data: bytes, nome: str, job_id: Optional[str] = None) -> StreamingResponse:
    """Stream XLSX; se job_id informado, adiciona header X-NC-Job-Id para o front encadear."""
    headers = {"Content-Disposition": f'attachment; filename="{_safe_filename_header(nome)}"'}
    if job_id:
        headers["X-NC-Job-Id"] = job_id
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _nc_proj_disponivel() -> bool:
    return _NC_PROJ.exists()


_modulos_carregados_log: set = set()


def _importar_modulo(nome: str):
    """
    Importa módulo do nc_artesp adicionando ao sys.path se necessário.
    Funciona tanto no Render (nc_artesp/ no repo) quanto no desktop Windows.
    """
    _garantir_path_nc()
    if not _nc_proj_disponivel():
        raise HTTPException(
            503,
            f"Módulos NC não encontrados.\n"
            f"Caminho verificado: {_NC_PROJ}\n"
            f"Opções:\n"
            f"  1. Copie artesp_nc_v2.0/modulos/ para GeradorARTESP/nc_artesp/modulos/\n"
            f"  2. Defina a variável de ambiente ARTESP_NC_PROJ no Render.",
        )
    proj = str(_NC_PROJ)
    if proj not in sys.path:
        sys.path.insert(0, proj)
    try:
        import importlib
        mod = importlib.import_module(f"modulos.{nome}")
        if nome not in _modulos_carregados_log:
            _modulos_carregados_log.add(nome)
            mod_path = getattr(mod, "__file__", "?")
            logger.info("NC módulo carregado: %s → %s | pasta NC: %s", nome, mod_path, _NC_PROJ)
        return mod
    except ImportError as e:
        raise HTTPException(503, f"Módulo '{nome}' não carregado: {e}")


def _limpar_cache_indices_foto() -> None:
    """
    Limpa cache de índices de fotos no módulo helpers.
    Necessário quando a pasta de imagens é purgada e reextraída no mesmo caminho.
    """
    try:
        _garantir_path_nc()
        import importlib
        h = importlib.import_module("utils.helpers")
        fn = getattr(h, "limpar_cache_indices_foto", None)
        if callable(fn):
            fn()
    except Exception as e:
        logger.debug("Não foi possível limpar cache de índices de foto: %s", e)


router = APIRouter(prefix="/nc", tags=["NC Artesp"])


def _importar_analisar_pdf():
    """Importa módulo de análise de PDF de NC."""
    _garantir_path_nc()
    try:
        from nc_artesp.modulos import analisar_pdf_nc
        return analisar_pdf_nc
    except ImportError as e:
        raise HTTPException(
            status_code=503,
            detail=(
                f"Módulo de análise não disponível: {e}\n"
                "Verifique se pymupdf e reportlab estão instalados."
            ),
        )


def _importar_pdf_extractor():
    """Importa pdf_extractor de nc_artesp (raiz do repositório)."""
    _garantir_path_nc()
    try:
        from nc_artesp import pdf_extractor
        return pdf_extractor
    except ImportError as e:
        raise HTTPException(
            status_code=503,
            detail=(
                f"Extração de PDF não disponível: {e}\n"
                "Verifique se pymupdf e pillow estão instalados."
            ),
        )


def _nc_pdf_paths_extrair_para_pasta_sync(
    pdf_entries: List[tuple[Path, str]],
    dest_dir: Path,
    lote: str,
    dpi: Optional[int],
    nomear_por_indice: bool,
) -> None:
    """
    Extrai cada PDF para ``dest_dir`` (nomes achatados, sem ZIP intermédio em RAM).
    Um diretório temporário por PDF libera bitmaps após cada iteração.
    """
    extrator = _importar_pdf_extractor()
    lote_m = re.search(r"\d+", (lote or "").strip() or "13")
    pasta_unica_artemig = (lote_m.group(0) if lote_m else "") == "50"
    # Regra operacional atual: identificação de fotos sempre por CÓDIGO da fiscalização.
    nomear_por_indice = False
    dest_dir.mkdir(parents=True, exist_ok=True)
    nomes_usados: set[str] = set()

    for pdf_path, fname in pdf_entries:
        with tempfile.TemporaryDirectory(prefix="nc_one_pdf_") as td:
            sub_out = Path(td) / "saida"
            sub_out.mkdir(parents=True, exist_ok=True)
            salvos, _ = extrator.extrair_arquivo_pdf_para_pasta(
                pdf_path,
                sub_out,
                dpi=dpi,
                nomear_por_indice_fiscalizacao=nomear_por_indice,
                pasta_unica=pasta_unica_artemig,
                raiz_unica_sem_subpastas=not pasta_unica_artemig,
                nome_pdf_original=fname,
            )
            for f in salvos:
                fp = Path(f)
                if not fp.is_file():
                    continue
                final = fp.name
                n_col = 1
                while final in nomes_usados:
                    pth = Path(final)
                    stem2 = f"{pth.stem}_{n_col}{pth.suffix or '.jpg'}"
                    final = stem2
                    n_col += 1
                nomes_usados.add(final)
                shutil.move(str(fp), str(dest_dir / final))


def _nc_pdf_paths_para_zip_imagens_bytes_sync(
    pdf_entries: List[tuple[Path, str]],
    lote: str,
    dpi: Optional[int],
    nomear_por_indice: bool,
) -> bytes:
    """ZIP em RAM só na fase final; PDFs já estão em disco (caminhos em ``pdf_entries``)."""
    work = Path(tempfile.mkdtemp(prefix="nc_zip_pdf_"))
    try:
        out_dir = work / "out"
        out_dir.mkdir()
        _nc_pdf_paths_extrair_para_pasta_sync(pdf_entries, out_dir, lote, dpi, nomear_por_indice)
        n = len(pdf_entries)
        lote_m = re.search(r"\d+", (lote or "").strip() or "13")
        lote_num = lote_m.group(0) if lote_m else "13"
        pasta_zip = f"lote_{lote_num}_{n}_pdfs_imagens"
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf_out:
            for fp in sorted(out_dir.iterdir(), key=lambda x: x.name):
                if fp.is_file():
                    arc = f"{pasta_zip}/{fp.name}".replace("\\", "/")
                    zf_out.write(fp, arc)
        return buf.getvalue()
    finally:
        shutil.rmtree(work, ignore_errors=True)


async def _nc_pdfs_uploads_para_zip_imagens_bytes(
    pdfs: List[UploadFile],
    lote: str,
    dpi: Optional[int],
    nomear_por_indice: bool,
) -> Optional[bytes]:
    """Grava PDFs em disco por chunks e devolve bytes do ZIP de imagens, ou None se não houver PDFs válidos."""
    if not (lote or "").strip():
        raise HTTPException(400, detail="Selecione o lote ao enviar PDFs.")
    root = Path(tempfile.mkdtemp(prefix="nc_zip_upload_"))
    try:
        pdf_in = root / "_in"
        pdf_in.mkdir()
        pdf_entries: list[tuple[Path, str]] = []
        for i, f in enumerate(pdfs or []):
            if not (f.filename or "").strip():
                continue
            dest = pdf_in / f"{i}.pdf"
            await _nc_gravar_upload_pdf_com_limite(f, dest)
            pdf_entries.append((dest, f.filename or "doc.pdf"))
        if not pdf_entries:
            return None
        return await asyncio.to_thread(
            _nc_pdf_paths_para_zip_imagens_bytes_sync,
            pdf_entries,
            lote,
            dpi,
            nomear_por_indice,
        )
    finally:
        shutil.rmtree(root, ignore_errors=True)


async def _nc_pdfs_uploads_para_arquivos_imagens(
    pdfs: List[UploadFile],
    lote: str,
    dpi: Optional[int],
    nomear_por_indice: bool,
) -> tuple[Optional[Path], Optional[Path]]:
    """
    Grava cada PDF em disco (por chunks), extrai imagens para ``out/`` dentro de um mkdtemp.
    Retorna ``(raiz_tmp, pasta_out)``. Sem PDFs válidos: ``(None, None)``.
    O chamador deve apagar ``raiz_tmp`` com ``shutil.rmtree(..., ignore_errors=True)`` quando terminar.
    """
    entries_files: List[UploadFile] = []
    for f in pdfs or []:
        if (f.filename or "").strip():
            entries_files.append(f)
    if not entries_files:
        return None, None
    if not (lote or "").strip():
        raise HTTPException(400, detail="Selecione o lote ao enviar PDFs.")
    root = Path(tempfile.mkdtemp(prefix="nc_pdf_img_"))
    try:
        pdf_in = root / "_in"
        pdf_in.mkdir()
        pdf_entries: list[tuple[Path, str]] = []
        for i, f in enumerate(entries_files):
            dest = pdf_in / f"{i}.pdf"
            await _nc_gravar_upload_pdf_com_limite(f, dest)
            pdf_entries.append((dest, f.filename or "doc.pdf"))
        out_dir = root / "out"
        out_dir.mkdir()
        await asyncio.to_thread(
            _nc_pdf_paths_extrair_para_pasta_sync,
            pdf_entries,
            out_dir,
            lote,
            dpi,
            nomear_por_indice,
        )
        for p, _ in pdf_entries:
            try:
                p.unlink()
            except OSError:
                pass
        return root, out_dir
    except Exception:
        shutil.rmtree(root, ignore_errors=True)
        raise


def _flag_teste_local(val_form: str) -> bool:
    v = (val_form or "").strip().lower()
    if v in ("0", "false", "no", "off"):
        return False
    if v in ("1", "true", "yes", "on"):
        return True
    return (os.getenv("ARTESP_NC_TESTE_LOCAL") or "").strip().lower() in ("1", "true", "yes", "on")


@router.post(
    "/analisar-pdf",
    summary="Analisar sequência de KMs e tipos de NCs do PDF de Constatação",
    response_description="ZIP com PDF de análise e XLSX; vários PDFs incluem Constatacoes_unificadas.pdf (uma sequência de páginas). Lote 50: Exportar Kcor + extração nc/PDF na pasta.",
)
async def nc_analisar_pdf(
    request: Request,
    pdfs: List[UploadFile] = File(..., description="Um ou mais PDFs de NC Constatação de Rotina Artesp"),
    limiar_km: float = Form(2.0, description="Gap mínimo de KM para gerar alerta (padrão 2 km)"),
    lote: str = Form("", description="Lote para o relatório (13, 21, 26 ou 50 Artemig). Obrigatório."),
    excel: List[UploadFile] = File(default=[], description="Um ou mais Excels que acompanham os PDFs (mesmo layout do relatório). Preenchem col E, O, P."),
    teste_local: str = Form(
        "",
        description="1/true ou env ARTESP_NC_TESTE_LOCAL: força alertas no PDF (útil com PDFs de data ≠ hoje).",
    ),
):
    _check_auth(request)
    if not (lote or "").strip():
        raise HTTPException(400, "Selecione o lote.")
    mod = _importar_analisar_pdf()
    try:
        pdfs_bytes = []
        for i, f in enumerate(pdfs):
            data = await f.read()
            if len(data) > MAX_BYTES:
                raise HTTPException(413, f"Arquivo '{f.filename}' excede {MAX_MB} MB.")
            pdfs_bytes.append(data)
        nomes = [Path(f.filename or f"pdf_{i+1}").stem for i, f in enumerate(pdfs)]
        lote_ok = (lote or "").strip() or None
        excel_list: List[bytes] = []
        for f in excel or []:
            if f and f.filename and (f.filename.lower().endswith(".xlsx") or f.filename.lower().endswith(".xls")):
                data = await f.read()
                if len(data) > MAX_BYTES:
                    raise HTTPException(413, f"Arquivo Excel '{f.filename}' excede o tamanho máximo.")
                excel_list.append(data)
        tl = _flag_teste_local(teste_local)
        pdf_rel, xlsx_bytes, resumo = await asyncio.to_thread(
            mod.analisar_e_gerar_pdf_multi,
            pdfs_bytes,
            limiar_km,
            nomes,
            lote_ok,
            excel_list,
            tl,
        )
        n_arqs = len(pdfs)
        relatorio_hoje = resumo.get("relatorio_hoje", True)
        if relatorio_hoje:
            n_emerg = len(resumo.get("emergenciais_lista", []))
            n_alertas = len(resumo.get("alertas_gap", []))
            n_ocultos = resumo.get("total_ocultos", 0)
        else:
            n_emerg = n_alertas = n_ocultos = 0
            logger.info(
                "nc/analisar-pdf: relatório anterior (data constatação ≠ hoje); "
                "alertas não exibidos no PDF (Total NCs: %s)",
                resumo.get("total", 0),
            )
        lote_slug = (lote or "").strip() or "13"
        try:
            _rotulo, slug = mod.rotulo_e_slug_lote_para_saida(lote_slug)
        except Exception:
            slug = (resumo.get("slug_zip") or "Lote13_Rodovias_Colinas").strip()
        slug = "".join(c if c.isalnum() or c in "_-" else "_" for c in slug) or "Analise"
        pasta = slug
        nome_pdf = f"{pasta}/Analise_NCs_{slug}.pdf"
        nome_xlsx = f"{pasta}/Relatorio_Fiscalizacao_{slug}.xlsx"
        if (lote_slug or "").strip() == "50":
            nome_zip = f"{slug}.zip"
        else:
            nome_zip = f"Relatorio_Analise_NCs_{slug}.zip"
        root_an = Path(tempfile.mkdtemp(prefix="nc_analisar_"))
        zip_disk = root_an / nome_zip
        extrator = None
        if (lote_slug or "").strip() == "50":
            try:
                extrator = _importar_pdf_extractor()
            except HTTPException:
                extrator = None
        try:
            with zipfile.ZipFile(zip_disk, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.writestr(nome_pdf.replace("\\", "/"), pdf_rel)
                zf.writestr(nome_xlsx.replace("\\", "/"), xlsx_bytes)
                if len(pdfs_bytes) > 1:
                    try:
                        from nc_artesp.pdf_extractor import FITZ_OK, merge_pdfs_bytes

                        if FITZ_OK:
                            arc_unificado = f"{pasta}/Constatacoes_unificadas.pdf"
                            if (lote_slug or "").strip() == "50":
                                arc_unificado = f"{pasta}/Consolidado/PDF_Consolidado_Constatacoes.pdf"
                            zf.writestr(
                                arc_unificado.replace("\\", "/"),
                                merge_pdfs_bytes(pdfs_bytes),
                            )
                    except Exception as ex:
                        logger.warning("nc/analisar-pdf: PDFs constatação unificados: %s", ex)
                kcor_b = resumo.get("exportar_kcor_xlsx") or b""
                kcor_nome = (resumo.get("exportar_kcor_nome") or "").strip()
                if kcor_b and kcor_nome and (lote_slug or "").strip() == "50":
                    arc_kcor = f"{pasta}/{kcor_nome}".replace("\\", "/")
                    zf.writestr(arc_kcor, kcor_b)
                if extrator and (lote_slug or "").strip() == "50":
                    usados_stem: set[str] = set()
                    arcs_ja: set[str] = set()
                    for i, data in enumerate(pdfs_bytes):
                        fn = (pdfs[i].filename if i < len(pdfs) else None) or f"pdf_{i + 1}.pdf"
                        try:
                            zip_i, _ = extrator.extrair_pdf_para_zip(
                                data,
                                dpi=None,
                                nomear_por_indice_fiscalizacao=False,
                                pasta_unica=True,
                                raiz_unica_sem_subpastas=False,
                                nome_pdf_original=fn,
                            )
                        except Exception as ex:
                            logger.warning("nc/analisar-pdf lote 50: extração imagens/PDF %s: %s", fn, ex)
                            continue
                        stem0 = Path(fn).stem
                        stem = "".join(c if c.isalnum() or c in "_-" else "_" for c in stem0) or f"pdf{i + 1}"
                        base_stem = stem
                        n_dup = 0
                        while stem in usados_stem:
                            n_dup += 1
                            stem = f"{base_stem}_{n_dup}"
                        usados_stem.add(stem)
                        with zipfile.ZipFile(io.BytesIO(zip_i)) as zf_i:
                            for name in zf_i.namelist():
                                if name.endswith("/") or not name.strip():
                                    continue
                                rel = name.replace("\\", "/").lstrip("/")
                                base_arq = Path(rel).name
                                nome_no_zip = _nc_arcnome_artemig_lote50(stem, base_arq)
                                arc = f"{pasta}/{nome_no_zip}".replace("\\", "/")
                                n_col = 1
                                while arc in arcs_ja:
                                    p = Path(nome_no_zip)
                                    nome_no_zip = f"{p.stem}_{n_col}{p.suffix}"
                                    arc = f"{pasta}/{nome_no_zip}".replace("\\", "/")
                                    n_col += 1
                                arcs_ja.add(arc)
                                zf.writestr(arc, zf_i.read(name))
            hdr_zip: dict[str, str] = {
                "Content-Disposition": f'attachment; filename="{_safe_filename_header(nome_zip)}"',
                "X-NC-Total":           str(resumo.get("total", 0)),
                "X-NC-Emergenciais":    str(n_emerg),
                "X-NC-Alertas":         str(n_alertas),
                "X-NC-Ocultos":         str(n_ocultos),
                "X-NC-Arquivos":        str(n_arqs),
                "X-NC-Relatorio-Dia":   "1" if relatorio_hoje else "0",
                "X-NC-Zip-Slug":        slug,
            }
            if (lote_slug or "").strip() == "50":
                km = resumo.get("exportar_kcor_meta")
                if isinstance(km, dict):
                    hdr_zip["X-NC-Kcor-Ok"] = "1" if km.get("ok") else "0"
                    if km.get("ok"):
                        hdr_zip["X-NC-Kcor-Modelo"] = (
                            "minimo" if km.get("modelo_minimo_gerado") else "ficheiro"
                        )
                else:
                    hdr_zip["X-NC-Kcor-Ok"] = "0"
            return FileResponse(
                str(zip_disk),
                media_type="application/zip",
                headers=hdr_zip,
                background=BackgroundTask(lambda r=root_an: shutil.rmtree(r, ignore_errors=True)),
            )
        except Exception:
            shutil.rmtree(root_an, ignore_errors=True)
            raise
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        logger.error("nc/analisar-pdf: %s", traceback.format_exc())
        raise HTTPException(500, str(e))


@router.post(
    "/extrair-pdf",
    summary="Extrair imagens do PDF de NC Constatação",
    response_description="ZIP: uma pasta lote_*; vários PDFs incluem Constatacoes_unificadas.pdf (páginas em sequência). Lote 50 = nc + PDF integral; demais = JPG na raiz.",
)
async def nc_extrair_pdf(
    request: Request,
    pdfs: List[UploadFile] = File(..., description="Um ou mais PDFs de NC Constatação Artesp"),
    dpi: Optional[int] = Form(
        None,
        description="Resolução PyMuPDF antes do redimensionamento; padrão = ARTESP_M02_EXTRACAO_RENDER_DPI (150).",
    ),
    lote: str = Form("", description="Número do lote (13, 21, 26, 50…). Obrigatório; entra no nome do ZIP."),
    nomear_por_indice_fiscalizacao: bool = Form(
        False,
        description="Se True, nomeia fotos por índice (00001, 00002...) em vez do código do PDF. Use para Meio Ambiente / Modelo Foto Kria.",
    ),
):
    """
    Processa **um ou mais** PDFs de Não Conformidade Artesp e gera dois arquivos JPG por NC:
    - **nc(N).jpg** — apenas a foto
    - **PDF(N).jpg** — texto de cabeçalho + foto

    Se nomear_por_indice_fiscalizacao=True, N = 00001, 00002... (índice = coluna V da EAF).
    Caso contrário, N = código extraído do PDF (ex.: 896643, HE.13.0111).
    **Lote 50 (Artemig):** **nc (COD).jpg** + **.pdf** integral na mesma pasta do ZIP.
    Demais lotes: **nc (...).jpg** e **PDF (...).jpg** na mesma pasta (sem subpastas **nc/** e **PDF/**).
    """
    _check_auth(request)
    if not (lote or "").strip():
        raise HTTPException(400, "Selecione o lote.")
    lote_m = re.search(r"\d+", (lote or "").strip() or "13")
    lote_num = lote_m.group(0) if lote_m else "13"
    # Regra operacional atual: identificação de fotos sempre por CÓDIGO da fiscalização.
    nomear_por_indice_fiscalizacao = False
    try:
        root = Path(tempfile.mkdtemp(prefix="nc_extrair_pdf_"))
        try:
            pdf_in = root / "_in"
            pdf_in.mkdir()
            pdf_entries: list[tuple[Path, str]] = []
            blobs_para_merge: list[bytes] = []
            for i, f in enumerate(pdfs):
                dest = pdf_in / f"{i}.pdf"
                await _nc_gravar_upload_pdf_com_limite(f, dest)
                blobs_para_merge.append(dest.read_bytes())
                pdf_entries.append((dest, f.filename or f"pdf_{i+1}.pdf"))
            out_dir = root / "out"
            out_dir.mkdir()
            await asyncio.to_thread(
                _nc_pdf_paths_extrair_para_pasta_sync,
                pdf_entries,
                out_dir,
                lote,
                dpi,
                nomear_por_indice_fiscalizacao,
            )
            if len(blobs_para_merge) > 1:
                try:
                    from nc_artesp.pdf_extractor import FITZ_OK, merge_pdfs_bytes

                    if FITZ_OK:
                        uni = await asyncio.to_thread(merge_pdfs_bytes, blobs_para_merge)
                        nome_unificado = "Constatacoes_unificadas.pdf"
                        if lote_num == "50":
                            nome_unificado = "PDF_Consolidado_Constatacoes.pdf"
                            out_cons = out_dir / "Consolidado"
                            out_cons.mkdir(parents=True, exist_ok=True)
                            (out_cons / nome_unificado).write_bytes(uni)
                        else:
                            (out_dir / nome_unificado).write_bytes(uni)
                except Exception as ex:
                    logger.warning("nc/extrair-pdf: PDFs constatação unificados: %s", ex)
            for p, _ in pdf_entries:
                try:
                    p.unlink()
                except OSError:
                    pass
            n = len(pdfs)
            pasta_zip = f"lote_{lote_num}_{n}_pdfs_imagens"
            nome_zip = f"{pasta_zip}.zip"
            zip_path = root / nome_zip
            with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf_out:
                for fp in sorted(out_dir.rglob("*"), key=lambda x: str(x).lower()):
                    if fp.is_file():
                        rel = fp.relative_to(out_dir).as_posix()
                        arc = f"{pasta_zip}/{rel}".replace("\\", "/")
                        zf_out.write(fp, arc)
            return FileResponse(
                str(zip_path),
                media_type="application/zip",
                headers={"Content-Disposition": f'attachment; filename="{_safe_filename_header(nome_zip)}"'},
                background=BackgroundTask(lambda r=root: shutil.rmtree(r, ignore_errors=True)),
            )
        except Exception:
            shutil.rmtree(root, ignore_errors=True)
            raise
    except HTTPException:
        raise
    except Exception as e:
        logger.error("nc/extrair-pdf: %s", traceback.format_exc())
        raise HTTPException(500, str(e))


@router.post("/separar", summary="M01 — Separar NC (ZIP XLS ou pipeline completo no servidor)")
async def nc_separar(
    request: Request,
    eafs: List[UploadFile] = File(..., description="Uma ou mais planilhas EAF (.xlsx ou .xls)"),
    job_id: Optional[str] = Form(None),
    finalize: bool = Form(False),
    m01_kartado: bool = Form(
        False,
        description="True: M01 com templates Kartado (lotes ARTESP 13/21/26). False (padrão): Art_011 — não alterar omissão (compat.).",
    ),
    m01_consolidado: bool = Form(
        True,
        description="True (padrão com Art_011): um Excel por EAF com todas as NCs ordenadas. False: vários Excels por grupo.",
    ),
    entrega_completa: bool = Form(
        True,
        description="True (padrão): após M01 executa e-mail, M02, M04 e M06; ZIP com respostas_kartado_fotos/ (incl. Pacotes_KTD se M02), Kartado/ só se M01 Kartado, etc. False: só ZIP dos XLS.",
    ),
    pdfs: Optional[List[UploadFile]] = File(
        None,
        description="PDFs de constatação (opcional). Com entrega_completa: extração de imagens como /nc/completo. "
        "Com um único EAF e m01_kartado=true: também alimenta «Observações» (AA) do Kartado com o texto livre da «Observação» do PDF.",
    ),
    lote: str = Form(
        "",
        description="Lote ARTESP (13, 21 ou 26) para ativar Kartado com m01_kartado=true; obrigatório se enviar PDFs (entrega_completa).",
    ),
    dpi: Optional[int] = Form(
        None,
        description="DPI na extração de imagens dos PDFs; padrão do projeto se omitido.",
    ),
    nomear_por_indice_fiscalizacao: bool = Form(
        False,
        description="Se True, nomes das fotos por índice (como em /nc/extrair-pdf).",
    ),
):
    """
    Por omissão (**entrega_completa**): mesmo fluxo contínuo que ``/nc/completo`` após o M01 — sem novo upload
    de ZIP intermédio; grava ``stage1/nc_separados.zip``, corre o pipeline e devolve o ZIP final.
    Com ``entrega_completa=false``: apenas o ZIP plano dos XLS (legado), com ``finalize``/``job_id`` como antes.
    """
    _check_auth(request)
    mod = _importar_modulo("separar_nc")
    m01_kartado_ativo = _m01_kartado_ativo_para_lote(m01_kartado, lote)
    try:
        ws, created = resolve_workspace(job_id or None)
        stage1_dir = ws.stage1
        stage1_dir.mkdir(parents=True, exist_ok=True)
        _purge_dir_contents(stage1_dir)
        if entrega_completa:
            ws.input.mkdir(parents=True, exist_ok=True)
            _purge_dir_contents(ws.input)

        arqs_all: list[Path] = []
        for i, eaf in enumerate(eafs):
            eaf_bytes = _ler(eaf)
            nome_safe = _safe_input_filename(eaf.filename or "eaf.xlsx")
            if i > 0:
                stem, suff = Path(nome_safe).stem, Path(nome_safe).suffix
                nome_safe = _safe_input_filename(f"{stem}_{i}{suff}")
            eaf_path = ws.input / nome_safe
            ws.input.mkdir(parents=True, exist_ok=True)
            eaf_path.write_bytes(eaf_bytes)
            _gravar_pdf_observacao_ao_lado_do_eaf_kartado(
                eaf_path, pdfs, m01_kartado_ativo=m01_kartado_ativo, num_eafs=len(eafs)
            )
            arqs = mod.executar(
                eaf_path,
                pasta_destino=stage1_dir,
                sobrescrever=True,
                copia_planilha_mae=not m01_kartado_ativo,
                unico_arquivo_organizado=(None if m01_consolidado else False),
            )
            arqs_all.extend(arqs or [])

        if not arqs_all:
            raise HTTPException(500, detail="M01 não gerou arquivos.")

        arqs_all = _nc_m01_kartado_consolidar_multiplos_excels(
            mod, arqs_all, stage1_dir, m01_kartado=m01_kartado_ativo
        )

        zip_path_stage1 = ws.stage1 / "nc_separados.zip"
        _used_s: set[str] = set()
        with zipfile.ZipFile(zip_path_stage1, "w", zipfile.ZIP_DEFLATED) as zf:
            _nc_zip_stage1_nc_separados(zf, ws.stage1, arqs_all, _used_s)

        ws.final.mkdir(parents=True, exist_ok=True)
        zip_final_stage1_dest = ws.final / "nc_separados.zip"
        shutil.copy2(zip_path_stage1, zip_final_stage1_dest)

        if entrega_completa:
            img_tmp_root: Optional[Path] = None
            try:
                img_tmp_root, img_out = await _nc_pdfs_uploads_para_arquivos_imagens(
                    pdfs or [],
                    lote,
                    dpi,
                    nomear_por_indice_fiscalizacao,
                )
                _touch_job_access(ws)
                _update_job_json(ws, status="running")
                try:
                    out = _nc_executar_pipeline_stage2_interno(
                        ws,
                        lote=(lote or "").strip() or None,
                        imagens_pdf_pasta_preparada=img_out,
                    )
                except HTTPException:
                    raise
                except Exception as e:
                    logger.error("nc/separar (pipeline): %s", traceback.format_exc())
                    try:
                        _update_job_json(
                            ws,
                            status="failed",
                            log_summary={"errors": 1, "warnings": 0, "message": str(e)[:200]},
                            retain_hours=2.0,
                        )
                    except Exception:
                        pass
                    raise HTTPException(500, str(e))

                if request.query_params.get("format") == "json":
                    base = _nc_response(
                        ws,
                        "final",
                        download_urls=[f"final/{out['download_zip']}"],
                        final_files=[out["download_zip"]],
                        step_label="Separar NC",
                        next_step_label="—",
                    )
                    base["download_links"] = out.get("download_links")
                    base["download_zip"] = out.get("download_zip")
                    base["kartado_merge_avisos"] = out.get("kartado_merge_avisos") or []
                    return JSONResponse(base)

                zip_final = ws.final / out["download_zip"]
                if not zip_final.is_file():
                    raise HTTPException(500, detail="ZIP final não encontrado após o pipeline.")
                return _stream_zip_path(
                    zip_final,
                    out["download_zip"],
                    ws.job_id,
                    extra_headers=_nc_kartado_merge_avisos_headers(),
                )
            finally:
                if img_tmp_root is not None and img_tmp_root.is_dir():
                    shutil.rmtree(img_tmp_root, ignore_errors=True)

        is_final = created or finalize
        if is_final:
            retain_hours = 2.0  # limpeza automática após 2h
            _update_job_json(ws, status="finished", stage="final", retain_hours=retain_hours)
        else:
            _update_job_json(ws, status="running", stage="stage1")

        if request.query_params.get("format") == "json":
            return JSONResponse(_nc_response(
                ws, "final" if is_final else "stage1",
                download_urls=["final/nc_separados.zip"],
                final_files=["nc_separados.zip"] if is_final else None,
                step_label="Separar NC",
                next_step_label="E-mail, Modelo Foto",
            ))
        return _stream_zip_path(zip_final_stage1_dest, "nc_separados.zip", ws.job_id)
    except HTTPException:
        raise
    except Exception as e:
        logger.error("nc/separar: %s", traceback.format_exc())
        raise HTTPException(500, str(e))


@router.post("/criar-email", summary="Gerar e-mails .eml a partir do ZIP de XLS (saída Separar NC)")
async def nc_criar_email_endpoint(
    request: Request,
    xls_zip: UploadFile = File(..., description="ZIP com XLS individuais (saída de Separar NC)"),
    imagens_pdf_zip: Optional[UploadFile] = File(None, description="ZIP opcional com imagens PDF (N).jpg (saída Extrair PDF) — para embutir fotos nos e-mails"),
):
    """
    Gera arquivos .eml (rascunhos de resposta NC) a partir do ZIP de planilhas XLS.
    Se enviar imagens_pdf_zip, as fotos são embutidas no corpo do e-mail.
    Retorna um ZIP com todos os .eml gerados (pasta emails/).
    """
    _check_auth(request)
    _garantir_path_nc()
    mod = _importar_modulo("nc_criar_email")
    try:
        xls_bytes = await xls_zip.read()
        imagens_bytes = await imagens_pdf_zip.read() if (imagens_pdf_zip and imagens_pdf_zip.filename) else None
        if len(xls_bytes) > MAX_BYTES:
            raise HTTPException(413, f"Arquivo '{xls_zip.filename}' excede {MAX_MB} MB.")
        if imagens_bytes is not None and len(imagens_bytes) > MAX_BYTES:
            raise HTTPException(413, f"Arquivo '{imagens_pdf_zip.filename}' excede {MAX_MB} MB.")
        root = Path(tempfile.mkdtemp(prefix="nc_email_"))
        try:
            tmp_path = root
            pasta_xls = tmp_path / DIR_EXPORTAR
            pasta_xls.mkdir(parents=True, exist_ok=True)
            with zipfile.ZipFile(io.BytesIO(xls_bytes), "r") as zf:
                extrair_zipfile_para_pasta(zf, pasta_xls)

            pasta_fotos_pdf = tmp_path / DIR_IMAGENS_PDF
            if imagens_bytes is not None:
                pasta_fotos_pdf.mkdir(parents=True, exist_ok=True)
                _limpar_cache_indices_foto()
                with zipfile.ZipFile(io.BytesIO(imagens_bytes), "r") as zf:
                    extrair_zipfile_para_pasta(zf, pasta_fotos_pdf)

            pasta_emails = tmp_path / "emails"
            pasta_emails.mkdir(parents=True, exist_ok=True)
            pasta_fotos_extr = (
                pasta_fotos_pdf if (imagens_bytes is not None and pasta_fotos_pdf.is_dir()) else None
            )
            resultado = await asyncio.to_thread(
                mod.executar,
                pasta_xls=pasta_xls,
                pasta_fotos_pdf=pasta_fotos_extr,
                pasta_fotos_nc=pasta_fotos_extr,
                usar_outlook=False,
                pasta_saida_eml=pasta_emails,
            )
            emls = resultado.get("eml") or []
            zip_path = root / "emails_nc.zip"
            with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                for eml in pasta_emails.rglob("*"):
                    if eml.is_file():
                        zf.write(eml, f"emails/{eml.name}")
        except Exception:
            shutil.rmtree(root, ignore_errors=True)
            raise
        if len(emls) == 0:
            logger.warning("criar-email: nenhum .eml gerado. Verifique o ZIP (XLS com colunas C, U, V preenchidas).")
        return FileResponse(
            str(zip_path),
            media_type="application/zip",
            headers={"Content-Disposition": f'attachment; filename="{_safe_filename_header("emails_nc.zip")}"'},
            background=BackgroundTask(lambda r=root: shutil.rmtree(r, ignore_errors=True)),
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error("nc/criar-email: %s", traceback.format_exc())
        raise HTTPException(500, str(e))


@router.post("/gerar-modelo-foto", summary="M02 — Gerar Kria + Resposta")
async def nc_gerar_modelo_foto(
    request: Request,
    xls_zip: UploadFile = File(..., description="ZIP com os XLS individuais (saída M01)"),
    modelo_kria: Optional[UploadFile] = File(None, description="Modelo Kria (.xlsx) — padrão: nc_artesp/assets/templates/Modelo Abertura Evento Kria Conserva Rotina.xlsx"),
    modelo_resp: Optional[UploadFile] = File(None, description="Modelo Resposta (.xlsx) — padrão: nc_artesp/assets/templates/Modelo.xlsx"),
    fotos_pdf_zip: Optional[UploadFile] = File(None, description="ZIP com fotos PDF (N).jpg (opcional — saída do Extrair PDF)"),
    job_id: Optional[str] = Form(None),
    finalize: bool = Form(False),
    lote: Optional[str] = Form(None, description="Lote 50 = ARTEMIG (templates em nc_artemig/assets/Template)"),
):
    """
    Etapa isolada: sem job_id → cria job, grava stage2/ e final/, marca finished.
    Pipeline: com job_id → reutiliza job, grava em stage2/; finalize=1 marca finished.
    """
    _check_auth(request)
    mod = _importar_modulo("gerar_modelo_foto")
    lote_ok = (lote or "").strip() or None
    try:
        ws, created = resolve_workspace(job_id or None)
        work = ws.stage2 / "_work"
        work.mkdir(parents=True, exist_ok=True)

        pasta_xls = work / DIR_EXPORTAR
        pasta_xls.mkdir(parents=True, exist_ok=True)
        _purge_dir_contents(pasta_xls)
        with zipfile.ZipFile(io.BytesIO(_ler(xls_zip))) as zf:
            extrair_zipfile_para_pasta(zf, pasta_xls)

        p_modelo_kria = work / "modelo_kria.xlsx"
        p_modelo_resp = work / "modelo_resp.xlsx"
        p_modelo_kria.write_bytes(_ler(modelo_kria) if modelo_kria else _carregar_modelo_kria(lote_ok))
        p_modelo_resp.write_bytes(_ler(modelo_resp) if modelo_resp else _carregar_modelo_resp(lote_ok))

        pasta_fotos_pdf = None
        pasta_fotos_nc = None
        if fotos_pdf_zip:
            pasta_fotos_pdf = work / DIR_IMAGENS_PDF
            pasta_fotos_pdf.mkdir(parents=True, exist_ok=True)
            _purge_dir_contents(pasta_fotos_pdf)
            _limpar_cache_indices_foto()
            with zipfile.ZipFile(io.BytesIO(_ler(fotos_pdf_zip))) as zf:
                extrair_zipfile_para_pasta(zf, pasta_fotos_pdf)
            # O ZIP do Extrair PDF contém nc (CODIGO).jpg e PDF (CODIGO).jpg na mesma pasta
            pasta_fotos_nc = pasta_fotos_pdf

        # Resolver como absolutos; limpar saídas anteriores para devolver só os arquivos desta requisição
        pasta_kria = (work / DIR_KARTADO_RELATORIO_FOTOS).resolve()
        pasta_resp = (work / DIR_RESPOSTAS_PENDENTES).resolve()
        pasta_kria.mkdir(parents=True, exist_ok=True)
        pasta_resp.mkdir(parents=True, exist_ok=True)
        _purge_dir_contents(pasta_kria)
        _purge_dir_contents(pasta_resp)

        resultado = mod.executar(
            pasta_xls=pasta_xls.resolve() if hasattr(pasta_xls, "resolve") else pasta_xls,
            modelo_kria=p_modelo_kria,
            pasta_saida_kria=pasta_kria,
            modelo_resposta=p_modelo_resp,
            pasta_saida_resp=pasta_resp,
            pasta_fotos_nc=pasta_fotos_nc,
            pasta_fotos_pdf=pasta_fotos_pdf,
        )

        ws.final.mkdir(parents=True, exist_ok=True)
        zip_out = ws.final / "modelos_kria.zip"
        n_arquivos_zip = 0
        with zipfile.ZipFile(zip_out, "w", zipfile.ZIP_DEFLATED) as zf:
            for d in [pasta_kria, pasta_resp]:
                for f in sorted(d.rglob("*.xlsx")):
                    if f.is_file():
                        try:
                            zf.write(f, f"{d.name}/{f.name}")
                            n_arquivos_zip += 1
                        except Exception as ex:
                            logger.warning("gerar-modelo-foto: não foi possível adicionar ao ZIP %s: %s", f, ex)

        # Nunca devolver ZIP vazio: rejeitar por contagem ou por tamanho (ZIP vazio ≈ 22 bytes)
        if n_arquivos_zip == 0 or zip_out.stat().st_size < 100:
            try:
                zip_out.unlink(missing_ok=True)
            except OSError:
                pass
            entradas = list(pasta_xls.rglob("*.xls*"))
            entradas = [f for f in entradas if f.is_file()]
            n_erros = len(resultado.get("erros") or [])
            msg = (
                "Nenhum arquivo gerado (ZIP vazio). "
                "Verifique: 1) O ZIP deve conter planilhas .xlsx ou .xls (saída do M01 Separar NC ou EAF); "
                "2) Planilhas com coluna C (Código) preenchida a partir da linha 5."
            )
            if not entradas:
                msg += f" Nenhuma planilha .xls/.xlsx encontrada na pasta extraída ({pasta_xls})."
            else:
                msg += f" Encontradas {len(entradas)} planilha(s) de entrada; nenhuma NC válida ou erro ao processar."
                if n_erros:
                    msg += f" Erros em {n_erros} arquivo(s): {resultado.get('erros', [])[:3]}."
            msg += " Confira os logs do servidor para detalhes (ex.: Permission denied, modelo não encontrado)."
            logger.warning("gerar-modelo-foto: %s", msg)
            raise HTTPException(status_code=422, detail=msg)

        is_final = created or finalize
        if is_final:
            retain_hours = 2.0  # limpeza automática após 2h
            _update_job_json(ws, status="finished", stage="final", retain_hours=retain_hours)
            _purge_work_if_finished(ws)
        else:
            _update_job_json(ws, status="running", stage="stage2")

        if request.query_params.get("format") == "zip":
            kria_count = len(resultado.get("kria") or [])
            resp_count = len(resultado.get("resposta") or [])
            err_count = len(resultado.get("erros") or [])
            pacote_count = len(resultado.get("kartado_pacotes") or [])
            return _stream_zip_path(
                zip_out,
                "modelos_kria.zip",
                ws.job_id,
                extra_headers={
                    "X-NC-Kria-Count": str(kria_count),
                    "X-NC-Resposta-Count": str(resp_count),
                    "X-NC-Erros-Count": str(err_count),
                    "X-NC-Pacotes-Count": str(pacote_count),
                },
            )
        return JSONResponse(_nc_response(
            ws, "final" if is_final else "stage2",
            download_urls=["final/modelos_kria.zip"],
            final_files=["modelos_kria.zip"] if is_final else None,
            step_label="Modelo Foto",
            next_step_label="Conservação",
        ))
    except HTTPException:
        raise
    except Exception as e:
        logger.error("nc/gerar-modelo-foto: %s", traceback.format_exc())
        raise HTTPException(500, str(e))


@router.post("/inserir-conservacao", summary="M03 — Kcor-Kria Conservação")
async def nc_inserir_conservacao(
    request: Request,
    kria_zip: UploadFile = File(..., description="ZIP com planilhas Kria (saída M02)"),
    modelo_kcor: Optional[UploadFile] = File(None, description="Modelo Kcor-Kria (.xlsx) — padrão: nc_artesp/assets/templates/_Planilha Modelo Kcor-Kria.xlsx"),
    fotos_zip: Optional[UploadFile] = File(None, description="ZIP fotos PDF (opcional)"),
    job_id: Optional[str] = Form(None),
    finalize: bool = Form(False),
    lote: Optional[str] = Form(None, description="Lote 50 = ARTEMIG (template em nc_artemig/assets/Template)"),
):
    _check_auth(request)
    return await _inserir_nc(request, kria_zip, modelo_kcor, fotos_zip, modo="conservacao", job_id=job_id, finalize=finalize, lote=lote)


@router.post("/inserir-meio-ambiente", summary="M07 — Kcor-Kria Meio Ambiente")
async def nc_inserir_ma(
    request: Request,
    kria_zip: UploadFile = File(..., description="ZIP com planilhas Kria MA"),
    modelo_kcor: Optional[UploadFile] = File(None, description="Modelo Kcor-Kria (.xlsx) — padrão: nc_artesp/assets/templates/_Planilha Modelo Kcor-Kria.xlsx"),
    job_id: Optional[str] = Form(None),
    finalize: bool = Form(False),
    lote: Optional[str] = Form(None, description="Lote 50 = ARTEMIG (template em nc_artemig/assets/Template)"),
):
    _check_auth(request)
    return await _inserir_nc(request, kria_zip, modelo_kcor, None, modo="meio_ambiente", job_id=job_id, finalize=finalize, lote=lote)


@router.post("/inserir-meio-ambiente-pdf", summary="M07 — Separar NC Meio Ambiente a partir de PDF(s)")
async def nc_inserir_ma_pdf(
    request: Request,
    pdf: list[UploadFile] = File(..., description="Um ou mais PDFs de Meio Ambiente"),
    job_id: Optional[str] = Form(None),
    finalize: bool = Form(False),
    m01_kartado: bool = Form(
        False,
        description="False (padrão): cópia planilha-mãe no M01. True: templates Kartado por atividade.",
    ),
):
    """Processa um ou mais PDFs de Meio Ambiente: extrai as informações em TEXTO do PDF,
    gera a planilha EAF (passo 1) e executa o Separar NC. Retorna ZIP com EAF + NCs separados."""
    _check_auth(request)
    m01_kartado_ativo = _m01_kartado_ativo_para_lote(m01_kartado, None)
    if not pdf:
        raise HTTPException(400, "Envie pelo menos um PDF.")
    pdf_list = pdf if isinstance(pdf, list) else [pdf]
    mod_kria = _importar_modulo("inserir_nc_kria")
    mod_separar = _importar_modulo("separar_nc")
    try:
        ws, created = resolve_workspace(job_id or None)
        list_pdf_bytes = [await up.read() for up in pdf_list]
        download_urls = []
        ws.input.mkdir(parents=True, exist_ok=True)
        ws.stage1.mkdir(parents=True, exist_ok=True)

        eaf_path = mod_kria.gerar_eaf_desde_pdfs_ma(
            list_pdf_bytes,
            pasta_saida=ws.input,
            nome_arquivo="eaf_ma_desde_pdf.xlsx",
        )
        if not eaf_path or not eaf_path.is_file():
            raise HTTPException(
                422,
                "Não foi possível gerar a planilha EAF a partir dos PDFs. Verifique se o PDF é de Meio Ambiente e contém blocos com data, código ou KM.",
            )
        download_urls.append(f"input/{eaf_path.name}")

        arqs = mod_separar.executar(
            arquivo_mae=eaf_path,
            pasta_destino=ws.stage1,
            um_arquivo_por_nc=True,
            copia_planilha_mae=not m01_kartado_ativo,
        )

        # ZIP de saída MA: nc_separados_ma.zip com pasta "Separar NC MA"
        nome_zip = "nc_separados_ma.zip"
        pasta_raiz = "Separar NC MA"
        ws.final.mkdir(parents=True, exist_ok=True)
        zip_ma_out = ws.final / nome_zip
        with zipfile.ZipFile(zip_ma_out, "w", zipfile.ZIP_DEFLATED) as zf:
            if arqs:
                for a in arqs:
                    p = Path(a)
                    if p.is_file():
                        try:
                            arcname = f"{pasta_raiz}/{p.relative_to(ws.stage1).as_posix()}"
                        except ValueError:
                            arcname = f"{pasta_raiz}/{p.name}"
                        zf.write(p, arcname)
                sub_mae = ws.stage1 / EXPORTAR_KARTADO_MAE_SUBDIR
                if sub_mae.is_dir():
                    for fx in sorted(sub_mae.glob("*.xlsx")):
                        if fx.is_file() and not fx.name.startswith("~"):
                            try:
                                arc_m = f"{pasta_raiz}/{fx.relative_to(ws.stage1).as_posix()}"
                            except ValueError:
                                arc_m = f"{pasta_raiz}/{EXPORTAR_KARTADO_MAE_SUBDIR}/{fx.name}"
                            zf.write(fx, arc_m)
            else:
                zf.write(eaf_path, f"{pasta_raiz}/{eaf_path.name}")
                zf.writestr(
                    f"{pasta_raiz}/LEIA-ME Separar NC.txt",
                    "O Separar NC não gerou arquivos individuais.\n"
                    "A planilha EAF está neste ZIP — abra e verifique se há dados a partir da linha 5 (colunas C e Q).\n"
                    "Se a EAF estiver correta, o problema pode ser o template Template_EAF.xlsx em nc_artesp/assets/templates/.",
                )
        download_urls.append(f"final/{nome_zip}")

        is_final = created or finalize
        if is_final:
            _update_job_json(ws, status="finished", stage="final", retain_hours=2.0)
            _purge_work_if_finished(ws)
        else:
            _update_job_json(ws, status="running", stage="stage1")
        if request.query_params.get("format") == "zip":
            return _stream_zip_path(zip_ma_out, nome_zip, ws.job_id)
        return JSONResponse(_nc_response(
            ws, "final" if is_final else "stage1",
            download_urls=download_urls,
            final_files=[nome_zip] if is_final else None,
            step_label="Meio Ambiente PDF",
            next_step_label="Acumulado",
        ))
    except HTTPException:
        raise
    except Exception as e:
        logger.error("nc/inserir-meio-ambiente-pdf: %s", traceback.format_exc())
        raise HTTPException(500, str(e))


@router.post("/pipeline-meio-ambiente-pdf", summary="M1+M2+M3 — Pipeline Meio Ambiente a partir de PDF")
async def nc_pipeline_ma_pdf(
    request: Request,
    pdf: UploadFile = File(..., description="PDF de Meio Ambiente"),
    imagens_zip: Optional[UploadFile] = File(None, description="ZIP opcional com imagens extraídas (nc (N).jpg, PDF (N).jpg). Se enviado, é extraído no início e usado no M2."),
    job_id: Optional[str] = Form(None),
    finalize: bool = Form(False),
    lote: Optional[str] = Form(None, description="Lote 50 = ARTEMIG (templates em nc_artemig/assets/Template)"),
):
    """
    Executa o equivalente a M1, M2 e M3 a partir do PDF de Meio Ambiente:
    M1 = extrai e parseia NCs do texto do PDF; gera também a planilha EAF (template do Separar NC).
    M2 = gera Kria (Arquivo Foto - MA) e Resposta (Pendentes).
    M3 = gera Kcor-Kria e imagens.
    Opcional: envie imagens_zip com as fotos extraídas (nc (1).jpg, PDF (1).jpg, etc.) para o M2 preencher os modelos.
    Retorna ZIP com EAF MA, Kria MA, Resposta MA, Kcor-Kria Meio Ambiente e Imagens MA.
    """
    _check_auth(request)
    mod = _importar_modulo("inserir_nc_kria")
    try:
        ws, created = resolve_workspace(job_id or None)
        work = ws.stage2 / "_work"
        work.mkdir(parents=True, exist_ok=True)
        work = work.resolve()
        pasta_imagens = (work / DIR_IMAGENS_MA).resolve()
        pasta_kria = (work / "Arquivo Foto MA").resolve()
        pasta_resp = (work / "Resposta MA").resolve()
        pasta_kcor = (work / DIR_MA).resolve()
        pasta_eaf = (work / "EAF MA").resolve()
        pasta_separar_nc = (work / "Separar NC MA").resolve()
        for p in (pasta_imagens, pasta_kria, pasta_resp, pasta_kcor, pasta_eaf, pasta_separar_nc):
            p.mkdir(parents=True, exist_ok=True)
        # Evitar acúmulo (persistência disco entre requisições/job_id reutilizado):
        for p in (pasta_imagens, pasta_kria, pasta_resp, pasta_kcor, pasta_eaf, pasta_separar_nc):
            _purge_dir_contents(p)
        # Opcional: extrair ZIP de imagens no início (nc (N).jpg, PDF (N).jpg) para o M2 usar
        if imagens_zip and imagens_zip.filename and (imagens_zip.filename.lower().endswith(".zip") or getattr(imagens_zip, "content_type", "") == "application/zip"):
            try:
                zip_bytes = await imagens_zip.read()
                if len(zip_bytes) > 0:
                    with zipfile.ZipFile(io.BytesIO(zip_bytes), "r") as zf:
                        extrair_zipfile_para_pasta(zf, pasta_imagens)
                    logger.info("Pipeline MA: ZIP de imagens extraído em %s", pasta_imagens.name)
            except Exception as e_zip:
                logger.warning("Pipeline MA: falha ao extrair ZIP de imagens: %s", e_zip)
        pdf_bytes = await pdf.read()
        nome_origem = (pdf.filename or "PDF MA").replace(".pdf", "").replace(".PDF", "")[:50]
        lote_ok = (lote or "").strip() or None
        p_kria = work / "modelo_kria_ma.xlsx"
        p_resp = work / "modelo_resp_ma.xlsx"
        modelo_kria_ma = None
        modelo_resp_ma = None
        try:
            p_kria.write_bytes(_carregar_modelo_kria(lote_ok))
            modelo_kria_ma = p_kria
        except HTTPException:
            pass
        try:
            p_resp.write_bytes(_carregar_modelo_resp(lote_ok))
            modelo_resp_ma = p_resp
        except HTTPException:
            # Fallback: usar o mesmo modelo do Kria para o Resposta e garantir que o segundo modelo seja gerado
            if modelo_kria_ma and modelo_kria_ma.is_file():
                p_resp.write_bytes(p_kria.read_bytes())
                modelo_resp_ma = p_resp
                logger.warning("Modelo Resposta não encontrado; usando modelo Kria como fallback para gerar o segundo arquivo.")
        # Se ainda faltar modelo Kria (asset não carregou), tentar path do config do nc_artesp
        if modelo_kria_ma is None:
            try:
                _garantir_path_nc()
                import importlib
                cfg = importlib.import_module("config")
                p = getattr(cfg, "M02_MODELO_KRIA", None)
                if p is not None and getattr(p, "is_file", lambda: False)() and getattr(p, "suffix", "") and str(p.suffix).lower() == ".xlsx":
                    modelo_kria_ma = p
                    if not modelo_resp_ma or not getattr(modelo_resp_ma, "is_file", lambda: False)():
                        p_resp.write_bytes(p.read_bytes())
                        modelo_resp_ma = p_resp
                    logger.info("Usando modelo Kria do config nc_artesp para MA.")
            except Exception:
                pass
        if modelo_kria_ma is None or not getattr(modelo_kria_ma, "is_file", lambda: False)():
            logger.warning(
                "Pipeline MA: modelo Kria não disponível. Kria/Resposta podem não ser gerados. "
                "Verifique nc_artesp/assets/templates/ e fotos_campo/assets/templates/."
            )
        resultado = mod.executar_pipeline_meio_ambiente_pdf(
            pdf_bytes,
            pasta_imagens=pasta_imagens,
            pasta_saida_kria=pasta_kria,
            pasta_saida_resp=pasta_resp,
            pasta_saida_eaf=pasta_eaf,
            pasta_saida_separar_nc=pasta_separar_nc,
            modelo_kria=modelo_kria_ma,
            modelo_resposta=modelo_resp_ma,
            modelo_kcor=None,
            pasta_saida_kcor=pasta_kcor,
            nome_origem=nome_origem,
        )
        nome = "pipeline_ma.zip"
        ws.final.mkdir(parents=True, exist_ok=True)
        zip_ma_pipe = ws.final / nome
        raiz_zip_ma = "Pipeline MA"
        adicionados = set()

        def _adicionar_arq(path, subpasta: str) -> None:
            if path is None:
                return
            p = path if isinstance(path, Path) else Path(str(path))
            if not p.is_file():
                return
            arcname = f"{raiz_zip_ma}/{subpasta}/{p.name}"
            if arcname in adicionados:
                return
            adicionados.add(arcname)
            try:
                zf.write(p, arcname)
            except Exception:
                pass

        with zipfile.ZipFile(zip_ma_pipe, "w", zipfile.ZIP_DEFLATED) as zf:
            # Fonte principal: conteúdo das pastas (onde o pipeline grava). Assim o ZIP reflete o que está em disco.
            pastas_zip = [
                (pasta_eaf, "EAF MA", ("*.xlsx",)),
                (pasta_kria, "Kria MA", ("*.xlsx",)),
                (pasta_resp, "Resposta MA", ("*.xlsx",)),
                (pasta_kcor, "Kcor-Kria Meio Ambiente", ("*.xlsx",)),
                (pasta_imagens, "Imagens MA", ("*.jpg", "*.jpeg", "*.png", "*.JPG", "*.JPEG", "*.PNG")),
            ]
            for folder, nome_pasta, extensoes in pastas_zip:
                try:
                    for ext in extensoes:
                        for f in folder.rglob(ext):
                            if f.is_file():
                                arcname = f"{raiz_zip_ma}/{nome_pasta}/{f.name}"
                                if arcname not in adicionados:
                                    adicionados.add(arcname)
                                    try:
                                        zf.write(str(f.resolve()), arcname)
                                    except Exception:
                                        try:
                                            zf.write(f, arcname)
                                        except Exception:
                                            pass
                except Exception:
                    pass
            # Incluir também os paths retornados pelo pipeline (caso estejam fora das pastas)
            _adicionar_arq(resultado.get("eaf"), "EAF MA")
            for path in resultado.get("kria") or []:
                _adicionar_arq(path, "Kria MA")
            for path in resultado.get("resposta") or []:
                _adicionar_arq(path, "Resposta MA")
            for path in resultado.get("kcor") or []:
                _adicionar_arq(path, "Kcor-Kria Meio Ambiente")
        is_final = created or finalize
        if is_final:
            _update_job_json(ws, status="finished", stage="final", retain_hours=2.0)
            _purge_work_if_finished(ws)
        else:
            _update_job_json(ws, status="running", stage="stage2")
        if request.query_params.get("format") == "zip":
            return _stream_zip_path(zip_ma_pipe, nome, ws.job_id)
        return JSONResponse(_nc_response(
            ws, "final" if is_final else "stage2",
            download_urls=[f"final/{nome}"],
            final_files=[nome] if is_final else None,
            step_label="Meio Ambiente",
            next_step_label="Acumulado",
        ))
    except HTTPException:
        raise
    except Exception as e:
        logger.error("nc/pipeline-meio-ambiente-pdf: %s", traceback.format_exc())
        raise HTTPException(500, str(e))


async def _inserir_nc(request, kria_zip, modelo_kcor, fotos_zip, modo, job_id: Optional[str] = None, finalize: bool = False, lote: Optional[str] = None):
    mod = _importar_modulo("inserir_nc_kria")
    lote_ok = (lote or "").strip() or None
    try:
        ws, created = resolve_workspace(job_id or None)
        work = ws.stage2 / "_work"
        work.mkdir(parents=True, exist_ok=True)

        pasta_kria = work / DIR_KARTADO_RELATORIO_FOTOS
        pasta_kria.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(io.BytesIO(_ler(kria_zip))) as zf:
            extrair_zipfile_para_pasta(zf, pasta_kria)
        pasta_entrada = pasta_kria
        for sub in (DIR_KARTADO_RELATORIO_FOTOS, "Kria", "kria"):
            cand = pasta_kria / sub
            if cand.is_dir() and list(cand.glob("*.xlsx")):
                pasta_entrada = cand
                break
        else:
            pasta_entrada = pasta_kria

        p_modelo = work / "modelo.xlsx"
        p_modelo.write_bytes(_ler(modelo_kcor) if modelo_kcor else _carregar_modelo_kcor(lote_ok))

        pasta_fotos = None
        if fotos_zip:
            pasta_fotos = work / DIR_IMAGENS_PDF
            pasta_fotos.mkdir(parents=True, exist_ok=True)
            with zipfile.ZipFile(io.BytesIO(_ler(fotos_zip))) as zf:
                extrair_zipfile_para_pasta(zf, pasta_fotos)

        pasta_imagens = work / DIR_IMAGENS_CONSERVACAO
        pasta_saida   = work / DIR_CONSERVACAO
        pasta_imagens.mkdir(parents=True, exist_ok=True)
        pasta_saida.mkdir(parents=True, exist_ok=True)

        regime_artemig = modo == "conservacao" and (lote_ok or "").strip() == "50"
        if modo == "conservacao":
            mod.executar_conservacao(
                pasta_entrada=pasta_entrada,
                pasta_imagens=pasta_imagens,
                modelo_kcor=p_modelo,
                pasta_saida=pasta_saida,
                pasta_fotos_pdf=pasta_fotos,
                pasta_fotos_nc=pasta_fotos if fotos_zip else None,
                forcar_fallback=True,
                regime_artemig=regime_artemig,
            )
        else:
            mod.executar_meio_ambiente(
                pasta_entrada=pasta_entrada,
                pasta_imagens=pasta_imagens,
                modelo_kcor=p_modelo,
                pasta_saida=pasta_saida,
                pasta_fotos_pdf=pasta_fotos,
                pasta_fotos_nc=pasta_fotos if fotos_zip else None,
                forcar_fallback=True,
            )

        nome = "kcor_conservacao.zip" if modo == "conservacao" else "kcor_ma.zip"
        ws.final.mkdir(parents=True, exist_ok=True)
        zip_kcor_out = ws.final / nome
        pasta_zip_ma = "Kcor-Kria Meio Ambiente"  # pasta identificada dentro do ZIP de MA
        with zipfile.ZipFile(zip_kcor_out, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in pasta_saida.rglob("*.xlsx"):
                arcname = f"{pasta_zip_ma}/{f.name}" if modo == "meio_ambiente" else f.name
                zf.write(f, arcname)

        is_final = created or finalize
        if is_final:
            retain_hours = 2.0  # limpeza automática após 2h
            _update_job_json(ws, status="finished", stage="final", retain_hours=retain_hours)
            _purge_work_if_finished(ws)
        else:
            _update_job_json(ws, status="running", stage="stage2")

        if request.query_params.get("format") == "zip":
            return _stream_zip_path(zip_kcor_out, nome, ws.job_id)
        step_label = "Conservação" if modo == "conservacao" else "Meio Ambiente"
        return JSONResponse(_nc_response(
            ws, "final" if is_final else "stage2",
            download_urls=[f"final/{nome}"],
            final_files=[nome] if is_final else None,
            step_label=step_label,
            next_step_label="Acumulado",
        ))
    except HTTPException:
        raise
    except Exception as e:
        logger.error("nc/inserir-%s: %s", modo, traceback.format_exc())
        raise HTTPException(500, str(e))


@router.post("/juntar", summary="M04 — Consolidar Acumulado")
async def nc_juntar(
    request: Request,
    kcor_zip: UploadFile = File(..., description="ZIP com Kcor-Kria individuais"),
    acumulado: Optional[UploadFile] = File(None, description="Planilha acumulada atual (opcional)"),
    nome_arquivo: Optional[str] = Form(None, description="Nome exato do arquivo de saída (ex.: 20260220-1313 - 20260220-1310 - 20260213 - CONSTATAÇÕES NC LOTE 13...xlsx)"),
    job_id: Optional[str] = Form(None),
    finalize: bool = Form(False),
):
    _check_auth(request)
    mod = _importar_modulo("juntar_arquivos")
    try:
        ws, created = resolve_workspace(job_id or None)
        work = ws.stage2 / "_work"
        work.mkdir(parents=True, exist_ok=True)

        pasta_kcor = work / DIR_KCOR_CONSERVACAO
        pasta_kcor.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(io.BytesIO(_ler(kcor_zip))) as zf:
            extrair_zipfile_para_pasta(zf, pasta_kcor)

        p_acum = work / "acumulado_base.xlsx"
        if acumulado:
            p_acum.write_bytes(_ler(acumulado))
        else:
            mod.criar_base_acumulado(p_acum)

        pasta_saida = work / DIR_ACUMULADO
        pasta_saida.mkdir(parents=True, exist_ok=True)

        # Arquivos xlsx: busca recursiva (ZIP pode ter subpastas)
        xlsx_lista = [f for f in pasta_kcor.rglob("*.xlsx")
                      if not f.name.startswith("~")
                      and "Acumulado" not in f.name
                      and not f.name.startswith("_")]
        arquivos_entrada = sorted(xlsx_lista) if xlsx_lista else None

        resultado = mod.executar(
            pasta_entrada=pasta_kcor,
            arquivo_acumulado=p_acum,  # sempre pasta do job; se vazio, módulo cria base
            pasta_saida=pasta_saida,
            arquivos_entrada=arquivos_entrada,
            nome_arquivo_completo=nome_arquivo,
        )

        xlsx_bytes = None
        nome_final = "acumulado.xlsx"
        if resultado and Path(resultado).exists():
            xlsx_bytes = Path(resultado).read_bytes()
            nome_final = Path(resultado).name
        else:
            for f in pasta_saida.glob("*.xlsx"):
                xlsx_bytes = f.read_bytes()
                nome_final = f.name
                break
        if not xlsx_bytes:
            if arquivos_entrada:
                raise HTTPException(
                    400,
                    "Nenhum registro encontrado nos arquivos .xlsx do ZIP. "
                    "Verifique se as planilhas têm dados a partir da linha 2 (colunas A–Y). "
                    "Envie também o arquivo acumulado (relatório da rede) para consolidar.",
                )
            raise HTTPException(
                400,
                "Envie o arquivo acumulado (relatório da rede) para consolidar. Sem esse arquivo nada é gerado.",
            )
        ws.final.mkdir(parents=True, exist_ok=True)
        (ws.final / nome_final).write_bytes(xlsx_bytes)

        is_final = created or finalize
        if is_final:
            retain_hours = 2.0  # limpeza automática após 2h
            _update_job_json(ws, status="finished", stage="final", retain_hours=retain_hours)
            _purge_work_if_finished(ws)
        else:
            _update_job_json(ws, status="running", stage="stage2")

        if request.query_params.get("format") == "xlsx":
            return _stream_xlsx(xlsx_bytes, nome_final, ws.job_id)
        return JSONResponse(_nc_response(
            ws, "final" if is_final else "stage2",
            download_urls=[f"final/{nome_final}"],
            final_files=[nome_final] if is_final else None,
            step_label="Acumulado",
            next_step_label="Inserir Nº, Calendário",
        ))
    except HTTPException:
        raise
    except Exception as e:
        logger.error("nc/juntar: %s", traceback.format_exc())
        raise HTTPException(500, str(e))


@router.post("/inserir-numero", summary="M05 — Inserir Nº Kria no Acumulado")
async def nc_inserir_numero(
    request: Request,
    acumulado: UploadFile = File(..., description="Planilha acumulada (.xlsx)"),
    numero_inicial: int = Form(1),
    sufixo: str = Form("26"),
    job_id: Optional[str] = Form(None),
    finalize: bool = Form(False),
):
    _check_auth(request)
    mod = _importar_modulo("inserir_numero_kria")
    try:
        ws, created = resolve_workspace(job_id or None)
        work = ws.stage2 / "_work"
        work.mkdir(parents=True, exist_ok=True)

        p = work / "acumulado.xlsx"
        p.write_bytes(_ler(acumulado))

        mod.executar(arquivo=p, numero_inicial=numero_inicial, sufixo=sufixo.strip())

        nome_arquivo = f"acumulado_{numero_inicial}{sufixo}.xlsx"
        xlsx_bytes = p.read_bytes()
        ws.final.mkdir(parents=True, exist_ok=True)
        (ws.final / nome_arquivo).write_bytes(xlsx_bytes)

        is_final = created or finalize
        if is_final:
            retain_hours = 2.0  # limpeza automática após 2h
            _update_job_json(ws, status="finished", stage="final", retain_hours=retain_hours)
            _purge_work_if_finished(ws)
        else:
            _update_job_json(ws, status="running", stage="stage2")

        if request.query_params.get("format") == "xlsx":
            return _stream_xlsx(xlsx_bytes, nome_arquivo, ws.job_id)
        return JSONResponse(_nc_response(
            ws, "final" if is_final else "stage2",
            download_urls=[f"final/{nome_arquivo}"],
            final_files=[nome_arquivo] if is_final else None,
            step_label="Inserir Nº",
            next_step_label="—",
        ))
    except HTTPException:
        raise
    except Exception as e:
        logger.error("nc/inserir-numero: %s", traceback.format_exc())
        raise HTTPException(500, str(e))


@router.post("/exportar-calendario", summary="M06 — Exportar eventos para iCalendar (.ics)")
async def nc_exportar_calendario(
    request: Request,
    acumulado: UploadFile = File(..., description="Planilha acumulada com col Y preenchida (saída M05)"),
):
    """
    Gera um arquivo **.ics** (iCalendar) a partir da planilha acumulada.
    Cria um evento por NC com:
    - **Assunto:** TipoNC - Rodovia KM Sentido - Kria: {nº}
    - **Data:** extraída do campo Observações (col U)
    - **Descrição:** Obs Gestor + Data Constatação + Observações

    O arquivo pode ser importado diretamente no Outlook, Google Calendar ou Apple Calendar.
    Equivalente à macro `Art_06_EAF_Rot_Exportar_Calend` (modo .ics, sem Outlook).
    """
    _check_auth(request)
    mod = _importar_modulo("exportar_calendario")
    try:
        xlsx_bytes = _ler(acumulado)
        ics_bytes, n_eventos = await asyncio.to_thread(mod.gerar_ics_bytes, xlsx_bytes)
        stem = Path(acumulado.filename or "acumulado").stem
        return StreamingResponse(
            io.BytesIO(ics_bytes),
            media_type="text/calendar",
            headers={
                "Content-Disposition": f'attachment; filename="{stem}_eventos.ics"',
                "X-NC-Eventos": str(n_eventos),
            },
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error("nc/exportar-calendario: %s", traceback.format_exc())
        raise HTTPException(500, str(e))


@router.post("/organizar-imagens", summary="M08 — Organizar imagens por tipo de NC")
async def nc_organizar_imagens(
    request: Request,
    acumulado: UploadFile = File(..., description="Planilha acumulada (.xlsx) com colunas E, G, I, M, P, T, W, Y"),
    imagens_zip: UploadFile = File(..., description="ZIP com as imagens geradas no M03 (col W do acumulado)"),
):
    """
    Organiza as imagens em subpastas por tipo de NC.
    Nome de cada arquivo: `rodovia - sentido - km,metro - yyyymmdd - ddmmaaaa - evento.jpg`

    Estrutura do ZIP gerado:
    - `{Tipo NC}/` → imagens daquele tipo
    - `_Exportar/` → cópia extra dos tipos de pavimento (Depressão e Pano de Rolamento)

    Equivalente à macro `Salvar_IMG_NC_Artesp_Pasta_Sep`.
    """
    _check_auth(request)
    mod = _importar_modulo("salvar_imagem")
    try:
        xlsx_bytes = _ler(acumulado)
        zip_bytes  = _ler(imagens_zip)
        zip_saida, n_copiadas = await asyncio.to_thread(
            mod.organizar_imagens_bytes, xlsx_bytes, zip_bytes
        )
        return StreamingResponse(
            io.BytesIO(zip_saida),
            media_type="application/zip",
            headers={
                "Content-Disposition": 'attachment; filename="imagens_classificadas.zip"',
                "X-NC-Imagens": str(n_copiadas),
            },
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error("nc/organizar-imagens: %s", traceback.format_exc())
        raise HTTPException(500, str(e))


@router.post("/job", summary="Criar workspace por execução (job_id)")
async def nc_job_create(request: Request):
    """
    Cria um workspace por execução sob OUTPUT_PATH/nc/<job_id>/ com subpastas:
    input/, stage1/, stage2/, final/. Pipeline stateful: arquivos entre etapas
    sem re-upload. Retorna job_id para uso nos endpoints que aceitarem job_id.
    """
    _check_auth(request)
    ws = create_nc_workspace()
    return {
        "job_id": ws.job_id,
        "job_dir": str(ws.job_dir),
        "paths": {
            "input": str(ws.input),
            "stage1": str(ws.stage1),
            "stage2": str(ws.stage2),
            "final": str(ws.final),
        },
    }


@router.get("/job/{job_id}", summary="Info do workspace (job)")
async def nc_job_info(request: Request, job_id: str):
    """Retorna informações do workspace (existência, job.json com status/stages)."""
    _check_auth(request)
    ws = resolve_nc_workspace(job_id)
    if ws.job_dir.is_dir():
        _update_job_json(ws)
    exists = ws.job_dir.is_dir()
    subdirs = {}
    job_state = None
    if exists:
        for name in NC_SUBDIRS:
            p = getattr(ws, name)
            subdirs[name] = {"path": str(p), "exists": p.is_dir(), "files": _list_stage_files(p)}
        job_json = ws.job_dir / "job.json"
        if job_json.is_file():
            try:
                with open(job_json, "r", encoding="utf-8") as f:
                    job_state = json.load(f)
            except (json.JSONDecodeError, OSError):
                pass
    return {
        "job_id": ws.job_id,
        "job_dir": str(ws.job_dir),
        "exists": exists,
        "paths": {k: str(getattr(ws, k)) for k in NC_SUBDIRS},
        "subdirs": subdirs,
        "job": job_state,
    }


@router.get("/jobs/{job_id}", summary="Status do job com touch (job_manager)")
async def nc_job_status_touch(request: Request, job_id: str):
    """Carrega job via job_manager, atualiza last_access (touch) e devolve { ok, job }."""
    _check_auth(request)
    if not job_manager_carregar:
        raise HTTPException(501, detail="job_manager não disponível.")
    try:
        job = job_manager_carregar(job_id, touch=True)
    except HTTPException:
        raise
    return {"ok": True, "job": job}


@router.post("/start", summary="Etapa 1 — Upload EAF (um ou mais) e extração (M01)")
async def nc_start(
    request: Request,
    arquivo: Annotated[Optional[UploadFile], File(description="Legado: um único EAF")] = None,
    arquivos: Annotated[
        Optional[List[UploadFile]],
        File(description="Um ou mais EAF (.xlsx/.xls) — mesmo campo repetido no formulário"),
    ] = None,
    m01_kartado: bool = Form(
        False,
        description="True: M01 Kartado (lotes 13/21/26). False (padrão): Art_011 — omissão preserva fluxo legado.",
    ),
    m01_consolidado: bool = Form(
        True,
        description="True (padrão com Art_011): um .xlsx por EAF com todas as NCs ordenadas. False: separar por grupos.",
    ),
    lote: str = Form(
        "",
        description="Lote ARTESP (13, 21 ou 26); com ARTESP_LOTE determina Kartado se m01_kartado=true.",
    ),
):
    """
    Recebe **um ou mais** EAF, salva em input/, executa M01 (Separar NC) sobre cada um
    e agrega os XLS em stage1/. Retorna job_id e link para stage1/nc_separados.zip.
    Use o campo `arquivos` (repetido) ou `arquivo` (único) para compatibilidade.
    """
    _check_auth(request)
    eafs = _lista_uploads_eaf(arquivo, arquivos)
    m01_kartado_ativo = _m01_kartado_ativo_para_lote(m01_kartado, lote)
    if not eafs:
        raise HTTPException(400, detail="Envie pelo menos um ficheiro EAF (.xlsx ou .xls).")
    mod = _importar_modulo("separar_nc")
    ws = create_nc_workspace()
    try:
        ws.input.mkdir(parents=True, exist_ok=True)
        arqs_all: list[Path] = []
        nomes_gravados: list[str] = []
        for i, eaf in enumerate(eafs):
            data = _ler(eaf)
            if len(data) > MAX_BYTES:
                raise HTTPException(413, detail=f"EAF {eaf.filename!r} excede {MAX_MB} MB.")
            nome_safe = _safe_input_filename(eaf.filename or "eaf.xlsx")
            if i > 0:
                stem, suff = Path(nome_safe).stem, Path(nome_safe).suffix
                nome_safe = _safe_input_filename(f"{stem}_{i}{suff}")
            input_path = ws.input / nome_safe
            input_path.write_bytes(data)
            nomes_gravados.append(nome_safe)
            arqs = mod.executar(
                input_path,
                pasta_destino=ws.stage1,
                sobrescrever=True,
                copia_planilha_mae=not m01_kartado_ativo,
                unico_arquivo_organizado=(None if m01_consolidado else False),
            )
            arqs_all.extend(arqs or [])
        if not arqs_all:
            raise HTTPException(500, detail="M01 não gerou arquivos.")

        arqs_all = _nc_m01_kartado_consolidar_multiplos_excels(
            mod, arqs_all, ws.stage1, m01_kartado=m01_kartado_ativo
        )

        zip_path = ws.stage1 / "nc_separados.zip"
        _used_s1: set[str] = set()
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            _nc_zip_stage1_nc_separados(zf, ws.stage1, arqs_all, _used_s1)

        _update_job_json(ws, status="stage1")
        prefix = f"/outputs/nc/{ws.job_id}"
        return {
            "job_id": ws.job_id,
            "input_file": nomes_gravados[0] if nomes_gravados else "",
            "input_files": nomes_gravados,
            "stage1_files": ["nc_separados.zip"],
            "download_links": [f"{prefix}/stage1/nc_separados.zip"],
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error("nc/start: %s", traceback.format_exc())
        raise HTTPException(500, str(e))


def _nc_executar_pipeline_stage2_interno(
    ws: NCWorkspace,
    *,
    lote: Optional[str],
    imagens_pdf_zip_bytes: Optional[bytes] = None,
    imagens_pdf_arquivos: Optional[Dict[str, bytes]] = None,
    imagens_pdf_pasta_preparada: Optional[Path] = None,
) -> dict:
    """
    Pipeline único após M01: e-mail (rascunhos) → M02 (Kria + Resposta) →
    M04 (acumulado Kcor-Kria a partir dos EAF em ``input/``) → M06 (`.ics` sem Outlook).
    **M03 (inserir Kria → Kcor)** não faz parte deste pipeline (desativado).
    **Kartado/** no ZIP final: um único Excel **layout Kartado** consolidado (M01 com ``m01_kartado=true``).
    **Pacotes KTD** (M02: modelo abertura + fotos) vão para ``respostas_kartado_fotos/Pacotes_KTD/``, não para Kartado/.
    Se existirem imagens PDF extraídas, grava backup e embute no ZIP final em `backup/`.
    imagens_pdf_zip_bytes: mesmo formato que o ZIP devolvido por POST /nc/extrair-pdf (opcional).
    imagens_pdf_pasta_preparada: diretório com JPG já extraídos (fluxo com baixa retenção em RAM).
    """
    etapa_t0: dict[int, float] = {}

    def _log_etapa(indice: int, total: int, titulo: str, estado: str = "INICIO") -> None:
        if estado == "INICIO":
            etapa_t0[indice] = time.perf_counter()
            logger.info("[PIPELINE][%s][%d/%d] %s", estado, indice, total, titulo)
            return
        t0 = etapa_t0.get(indice)
        if t0 is None:
            logger.info("[PIPELINE][%s][%d/%d] %s", estado, indice, total, titulo)
            return
        dt_s = time.perf_counter() - t0
        logger.info("[PIPELINE][%s][%d/%d] %s (%.2fs)", estado, indice, total, titulo, dt_s)

    total_etapas = 7
    _log_etapa(1, total_etapas, "Preparar workspace stage2 e extrair stage1")
    zip_stage1 = ws.stage1 / "nc_separados.zip"
    if not zip_stage1.is_file():
        raise HTTPException(400, detail="stage1/nc_separados.zip não encontrado.")

    _garantir_path_nc()
    mod_modelo = _importar_modulo("gerar_modelo_foto")
    mod_calendario = _importar_modulo("exportar_calendario")
    mod_criar_email = _importar_modulo("nc_criar_email")

    ws.stage2.mkdir(parents=True, exist_ok=True)
    work = ws.stage2 / "_work"
    work.mkdir(exist_ok=True)
    pasta_xls = work / DIR_EXPORTAR
    pasta_xls.mkdir(exist_ok=True)
    with zipfile.ZipFile(zip_stage1, "r") as zf:
        extrair_zipfile_para_pasta(zf, pasta_xls)
    _log_etapa(1, total_etapas, "Preparar workspace stage2 e extrair stage1", "FIM")

    pasta_fotos_pdf = work / DIR_IMAGENS_PDF
    _log_etapa(2, total_etapas, "Importar imagens PDF (quando enviadas)")
    if imagens_pdf_pasta_preparada is not None and imagens_pdf_pasta_preparada.is_dir():
        pasta_fotos_pdf.mkdir(parents=True, exist_ok=True)
        _purge_dir_contents(pasta_fotos_pdf)
        _limpar_cache_indices_foto()
        n_zip_in = 0
        for item in sorted(imagens_pdf_pasta_preparada.iterdir(), key=lambda x: x.name):
            if item.is_file():
                shutil.copy2(item, pasta_fotos_pdf / item.name)
                n_zip_in += 1
        if n_zip_in == 0:
            logger.warning("Imagens PDF pré-extraídas (pasta): 0 ficheiros copiados.")
    elif imagens_pdf_arquivos:
        pasta_fotos_pdf.mkdir(parents=True, exist_ok=True)
        _purge_dir_contents(pasta_fotos_pdf)
        _limpar_cache_indices_foto()
        n_zip_in = 0
        for nome, data in (imagens_pdf_arquivos or {}).items():
            final = Path((nome or "").replace("\\", "/")).name
            if not final:
                continue
            out = pasta_fotos_pdf / final
            out.write_bytes(data)
            n_zip_in += 1
        if n_zip_in == 0:
            logger.warning("Imagens PDF pré-extraídas: 0 ficheiros recebidos.")
    elif imagens_pdf_zip_bytes:
        pasta_fotos_pdf.mkdir(parents=True, exist_ok=True)
        _purge_dir_contents(pasta_fotos_pdf)
        _limpar_cache_indices_foto()
        try:
            n_zip_in = _nc_extrair_zip_para_pasta_seguro(imagens_pdf_zip_bytes, pasta_fotos_pdf)
            if n_zip_in == 0:
                logger.warning(
                    "ZIP de imagens PDF: 0 ficheiros extraídos (bytes=%d). Verifique o formato do ZIP.",
                    len(imagens_pdf_zip_bytes),
                )
        except zipfile.BadZipFile as e:
            logger.warning("ZIP de imagens PDF corrompido ou inválido: %s", e)
    _log_etapa(2, total_etapas, "Importar imagens PDF (quando enviadas)", "FIM")

    pasta_fotos_dir = pasta_fotos_pdf if pasta_fotos_pdf.is_dir() else None
    pasta_fotos_nc_email = pasta_fotos_dir

    ws.final.mkdir(parents=True, exist_ok=True)
    _log_etapa(3, total_etapas, "Gerar e-mails (.eml)")
    pasta_emails = ws.final / "emails"
    pasta_emails.mkdir(parents=True, exist_ok=True)
    qtd_eml = 0
    try:
        res_email = mod_criar_email.executar(
            pasta_xls=pasta_xls,
            pasta_fotos_pdf=pasta_fotos_dir,
            pasta_fotos_nc=pasta_fotos_nc_email,
            usar_outlook=False,
            pasta_saida_eml=pasta_emails,
        )
        qtd_eml = len((res_email or {}).get("eml") or [])
        if qtd_eml == 0:
            logger.warning("Módulo NC Email: nenhum .eml gerado no pipeline.")
    except Exception as e_email:
        logger.exception("Módulo NC Email (após M01) falhou: %s", e_email)
    _log_etapa(3, total_etapas, f"Gerar e-mails (.eml) — total: {qtd_eml}", "FIM")

    lote_mod = (lote or "").strip() or None

    pasta_kria = (work / DIR_KARTADO_RELATORIO_FOTOS).resolve()
    pasta_resp = (work / DIR_RESPOSTAS_PENDENTES).resolve()
    pasta_kria.mkdir(parents=True, exist_ok=True)
    pasta_resp.mkdir(parents=True, exist_ok=True)
    _purge_dir_contents(pasta_kria)
    _purge_dir_contents(pasta_resp)

    p_modelo_kria = work / "modelo_kria.xlsx"
    p_modelo_resp = work / "modelo_resp.xlsx"
    modelos_m02_ok = False
    try:
        p_modelo_kria.write_bytes(_carregar_modelo_kria(lote_mod))
        p_modelo_resp.write_bytes(_carregar_modelo_resp(lote_mod))
        modelos_m02_ok = True
    except HTTPException as he:
        logger.warning(
            "Pipeline: modelos M02 indisponíveis (%s). Continua sem M02.",
            getattr(he, "detail", he),
        )
    except Exception as e_mod:
        logger.warning("Pipeline: modelos M02 indisponíveis (%s). Continua sem M02.", e_mod)

    pasta_fotos_nc_m02 = pasta_fotos_dir

    _log_etapa(4, total_etapas, "Executar M02 (Kria + Resposta)")
    if modelos_m02_ok and p_modelo_kria.is_file() and p_modelo_resp.is_file():
        try:
            res_m02 = mod_modelo.executar(
                pasta_xls=pasta_xls.resolve(),
                modelo_kria=p_modelo_kria,
                pasta_saida_kria=pasta_kria,
                modelo_resposta=p_modelo_resp,
                pasta_saida_resp=pasta_resp,
                pasta_fotos_nc=pasta_fotos_nc_m02,
                pasta_fotos_pdf=pasta_fotos_dir,
            )
            if (res_m02 or {}).get("erros"):
                logger.warning("M02: %s ficheiro(s) com erro.", len(res_m02["erros"]))
        except Exception as e_m02:
            logger.warning("M02 (gerar modelo foto): %s", e_m02)
    qtd_kria = len(list(pasta_kria.glob("*.xlsx"))) if pasta_kria.is_dir() else 0
    qtd_resp = len(list(pasta_resp.glob("*.xlsx"))) if pasta_resp.is_dir() else 0
    _log_etapa(4, total_etapas, f"Executar M02 (Kria + Resposta) — Kria: {qtd_kria}, Respostas: {qtd_resp}", "FIM")

    # Kartado (M01): só com Excel **layout Kartado** (templates por atividade).
    # Pacotes KTD de M02 não entram na entrega final deste fluxo.
    tem_base_m01 = any(
        f.is_file()
        and f.suffix.lower() == ".xlsx"
        and not f.name.startswith("~")
        and not f.name.startswith("_")
        for f in pasta_xls.rglob("*.xlsx")
    )
    m01_eh_kartado = tem_base_m01 and _nc_exportar_contem_excel_layout_kartado(pasta_xls)
    _log_etapa(5, total_etapas, "Fundir Kartado (um Excel consolidado)")
    if m01_eh_kartado:
        _nc_gravar_pacotes_kartado_de_m01(pasta_xls, pasta_fotos_dir)
    n_kart_xlsx = sum(
        1
        for f in pasta_xls.rglob("*.xlsx")
        if f.is_file()
        and not f.name.startswith("~")
        and not f.name.startswith("_")
        and _nc_workbook_primeira_linha_eh_layout_kartado(f)
    )
    _log_etapa(5, total_etapas, f"Fundir Kartado — .xlsx layout Kartado na pasta Exportar: {n_kart_xlsx}", "FIM")

    ws.final.mkdir(parents=True, exist_ok=True)
    entrega = ws.final / "_entrega"
    if entrega.is_dir():
        shutil.rmtree(entrega, ignore_errors=True)
    entrega.mkdir(parents=True)

    # Pastas legíveis no ZIP final (caminhos curtos nos ficheiros internos continuam limitados por MAX_PATH).
    p02 = entrega / "Kartado"
    p04 = entrega / "acumulado"
    p05 = entrega / "calendario"
    p06 = entrega / "emails"

    if pasta_kria.is_dir() and any(pasta_kria.glob("*.xlsx")):
        pk = entrega / DIR_KARTADO_RELATORIO_FOTOS
        pk.mkdir(parents=True, exist_ok=True)
        _nc_copiar_xlsx_de_pasta(pasta_kria, pk)
    if pasta_resp.is_dir() and any(pasta_resp.glob("*.xlsx")):
        pr = entrega / DIR_PENDENTES_ENTREGA
        pr.mkdir(parents=True, exist_ok=True)
        _nc_copiar_xlsx_de_pasta(pasta_resp, pr)

    # Pedido do fluxo atual: não incluir pasta Pacotes_KTD na entrega final.
    _nc_copiar_kartado_para_entrega(pasta_xls, p02)

    p_mae_kartado_exportar = pasta_xls / EXPORTAR_KARTADO_MAE_SUBDIR
    if p_mae_kartado_exportar.is_dir() and any(
        f.is_file() and not f.name.startswith("~")
        for f in p_mae_kartado_exportar.rglob("*.xlsx")
    ):
        p_entrega_mae_kartado = entrega / EXPORTAR_KARTADO_MAE_SUBDIR
        p_entrega_mae_kartado.mkdir(parents=True, exist_ok=True)
        _nc_copiar_xlsx_de_pasta(p_mae_kartado_exportar, p_entrega_mae_kartado)

    p04.mkdir(parents=True, exist_ok=True)
    _log_etapa(6, total_etapas, "Gerar acumulado (M04) e calendário (M06)")
    acum_path = p04 / "Acumulado.xlsx"
    if not _nc_gerar_acumulado_xlsx(ws.input, acum_path, pasta_fallback=pasta_xls):
        (p04 / "README.txt").write_text(
            "Não foi possível gerar o Acumulado.xlsx.\n"
            "Verifique: (1) ficheiro Acumulado.xlsx ou _Planilha Modelo Kcor-Kria.xlsx em nc_artesp/assets/templates/ "
            "(ou variável ARTESP_M04_TEMPLATE_ACUMULADO_KCOR_KRIA no servidor);\n"
            "(2) EAF com código na coluna C a partir da linha 5, ou saída M01 Kartado com código na linha 2;\n"
            "(3) logs do serviço com mensagens «Acumulado Kcor-Kria».\n",
            encoding="utf-8",
        )

    p05.mkdir(parents=True, exist_ok=True)
    if acum_path.is_file():
        try:
            mod_calendario.executar(
                acum_path,
                usar_outlook=False,
                pasta_saida_ics=p05,
                executar_mod08=False,
            )
        except Exception as e_m06:
            logger.exception("M06 (exportar calendário .ics): %s", e_m06)
            (p05 / "README.txt").write_text(
                "Falha ao gerar o .ics a partir do Acumulado.xlsx.\n"
                f"Erro: {type(e_m06).__name__}: {e_m06}\n"
                "Verifique se o Acumulado é um .xlsx válido (não ponteiro Git LFS) e se há linhas com tipo NC (col E).\n",
                encoding="utf-8",
            )
    else:
        (p05 / "README.txt").write_text(
            "Sem Acumulado.xlsx — calendário não gerado.",
            encoding="utf-8",
        )
    _log_etapa(6, total_etapas, "Gerar acumulado (M04) e calendário (M06)", "FIM")

    p06.mkdir(parents=True, exist_ok=True)
    n_eml_copiados = 0
    if pasta_emails.is_dir():
        for eml in pasta_emails.rglob("*"):
            if eml.is_file():
                rel = eml.relative_to(pasta_emails)
                out_eml = p06 / rel
                out_eml.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(eml, out_eml)
                if eml.suffix.lower() == ".eml":
                    n_eml_copiados += 1
    if n_eml_copiados == 0:
        (p06 / "README.txt").write_text(
            "Nenhum ficheiro .eml foi gerado nesta execução (verifique planilhas em Exportar/ "
            "e o módulo de e-mail nos logs). A pasta emails/ é mantida no ZIP para referência.\n",
            encoding="utf-8",
        )

    # Gera o ZIP de backup de imagens extraídas e embute no ZIP final (não fica separado).
    zip_imagens_backup_name: Optional[str] = None
    zip_imagens_backup_path = ws.stage2 / "_work" / f"nc_{ws.job_id}_imagens_extraidas_backup.zip"
    n_img_backup = _nc_zip_imagens_extraidas_backup(
        work,
        pasta_fotos_pdf,
        zip_imagens_backup_path,
        zip_bytes_pipeline_pdf=imagens_pdf_zip_bytes,
    )
    if n_img_backup > 0 and zip_imagens_backup_path.is_file():
        zip_imagens_backup_name = zip_imagens_backup_path.name
        logger.info(
            "ZIP backup imagens extraídas embutido no principal: %s (%d ficheiro(s))",
            zip_imagens_backup_name,
            n_img_backup,
        )

    zip_final_name = f"Kartado_NC_{ws.job_id}_artesp.zip"
    _log_etapa(7, total_etapas, "Empacotar entrega final (.zip)")
    zip_final_path = ws.final / zip_final_name
    _used_final: set[str] = set()
    with zipfile.ZipFile(zip_final_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in entrega.rglob("*"):
            if f.is_file():
                arc = f.relative_to(entrega).as_posix()
                arc = _nc_arcnome_zip_para_extracao_windows(arc, usados=_used_final)
                zf.write(f, arc)
        if zip_imagens_backup_name and zip_imagens_backup_path.is_file():
            arc_backup = _nc_arcnome_zip_para_extracao_windows(
                f"backup/{zip_imagens_backup_name}",
                usados=_used_final,
            )
            zf.write(zip_imagens_backup_path, arc_backup)

    shutil.rmtree(entrega, ignore_errors=True)
    _log_etapa(7, total_etapas, f"Empacotar entrega final (.zip) — arquivo: {zip_final_name}", "FIM")
    if pasta_emails.is_dir():
        shutil.rmtree(pasta_emails, ignore_errors=True)

    _purge_dir_contents(ws.stage1)
    _purge_dir_contents(ws.stage2)

    _update_job_json(
        ws,
        status="finished",
        log_summary={"errors": 0, "warnings": 0},
        retain_hours=2.0,
    )
    prefix = f"/outputs/nc/{ws.job_id}"
    links = [f"{prefix}/final/{zip_final_name}"]
    out: Dict[str, Any] = {
        "job_id": ws.job_id,
        "download_links": links,
        "download_zip": zip_final_name,
        "kartado_merge_avisos": _nc_kartado_merge_avisos_payload(),
    }
    return out


@router.post(
    "/completo",
    summary="EAF + PDF(s) → ZIP final numa só chamada (sem ZIP intermédio)",
)
async def nc_completo(
    request: Request,
    arquivo: Annotated[Optional[UploadFile], File(description="Legado: um único EAF")] = None,
    arquivos: Annotated[
        Optional[List[UploadFile]],
        File(description="Um ou mais EAF — mesmo campo repetido no multipart"),
    ] = None,
    pdfs: Optional[List[UploadFile]] = File(
        None,
        description="Um ou mais PDFs de constatação (opcional). Mesma extração que /nc/extrair-pdf; "
        "com um único EAF e m01_kartado=true: também alimenta «Observações» (AA) do Kartado.",
    ),
    lote: str = Form(
        "",
        description="Lote ARTESP (13, 21 ou 26) para Kartado; 50 = Artemig (sem Kartado). Obrigatório se enviar PDFs.",
    ),
    dpi: Optional[int] = Form(
        None,
        description="DPI PyMuPDF na extração; padrão do projeto se omitido.",
    ),
    nomear_por_indice_fiscalizacao: bool = Form(
        False,
        description="Se True, nomes das fotos por índice (como em /nc/extrair-pdf).",
    ),
    m01_kartado: bool = Form(
        False,
        description="True: M01 Kartado (lotes 13/21/26). False (padrão): Art_011 — omissão preserva fluxo legado.",
    ),
    m01_consolidado: bool = Form(
        True,
        description="True (padrão com Art_011): um .xlsx por EAF consolidado. False: vários Excels por grupo.",
    ),
):
    """
    Pipeline **start + stage2** num único pedido: M01; imagens dos PDFs no servidor quando enviadas;
    em seguida e-mail → **M02** (Kria + Resposta) → **M04** (acumulado a partir da EAF em input/) → **M06** (`.ics`).
    **M03** inserir Kria não corre neste pipeline; **M05** (coluna Y) continua manual.
    ZIP final: `Kartado_relatorio_fotos/`, `pendentes/`,
    `Kartado/` (apenas se M01 gerou layout Kartado), `acumulado/`, `calendario/`, `emails/`, `backup/`.
    """
    _check_auth(request)
    mod_sep = _importar_modulo("separar_nc")
    m01_kartado_ativo = _m01_kartado_ativo_para_lote(m01_kartado, lote)
    ws = create_nc_workspace()
    try:
        eafs = _lista_uploads_eaf(arquivo, arquivos)
        if not eafs:
            raise HTTPException(400, detail="Envie pelo menos um ficheiro EAF (.xlsx ou .xls).")
        ws.input.mkdir(parents=True, exist_ok=True)
        arqs_all: list[Path] = []
        for i, eaf in enumerate(eafs):
            data_eaf = await eaf.read()
            if len(data_eaf) > MAX_BYTES:
                raise HTTPException(413, detail=f"EAF {eaf.filename!r} excede {MAX_MB} MB.")
            nome_safe = _safe_input_filename(eaf.filename or "eaf.xlsx")
            if i > 0:
                stem, suff = Path(nome_safe).stem, Path(nome_safe).suffix
                nome_safe = _safe_input_filename(f"{stem}_{i}{suff}")
            input_path = ws.input / nome_safe
            input_path.write_bytes(data_eaf)
            _gravar_pdf_observacao_ao_lado_do_eaf_kartado(
                input_path, pdfs, m01_kartado_ativo=m01_kartado_ativo, num_eafs=len(eafs)
            )
            arqs = mod_sep.executar(
                input_path,
                pasta_destino=ws.stage1,
                sobrescrever=True,
                copia_planilha_mae=not m01_kartado_ativo,
                unico_arquivo_organizado=(None if m01_consolidado else False),
            )
            arqs_all.extend(arqs or [])
        if not arqs_all:
            raise HTTPException(500, detail="M01 não gerou arquivos.")

        arqs_all = _nc_m01_kartado_consolidar_multiplos_excels(
            mod_sep, arqs_all, ws.stage1, m01_kartado=m01_kartado_ativo
        )

        zip_path = ws.stage1 / "nc_separados.zip"
        _used_c: set[str] = set()
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            _nc_zip_stage1_nc_separados(zf, ws.stage1, arqs_all, _used_c)

        img_tmp_root: Optional[Path] = None
        try:
            img_tmp_root, img_out = await _nc_pdfs_uploads_para_arquivos_imagens(
                pdfs or [],
                lote,
                dpi,
                nomear_por_indice_fiscalizacao,
            )

            _touch_job_access(ws)
            _update_job_json(ws, status="running")
            try:
                return _nc_executar_pipeline_stage2_interno(
                    ws,
                    lote=(lote or "").strip() or None,
                    imagens_pdf_pasta_preparada=img_out,
                )
            except HTTPException:
                raise
            except Exception as e:
                logger.error("nc/completo (pipeline): %s", traceback.format_exc())
                try:
                    _update_job_json(
                        ws,
                        status="failed",
                        log_summary={"errors": 1, "warnings": 0, "message": str(e)[:200]},
                        retain_hours=2.0,
                    )
                except Exception:
                    pass
                raise HTTPException(500, str(e))
        finally:
            if img_tmp_root is not None and img_tmp_root.is_dir():
                shutil.rmtree(img_tmp_root, ignore_errors=True)
    except HTTPException:
        raise
    except Exception as e:
        logger.error("nc/completo: %s", traceback.format_exc())
        raise HTTPException(500, str(e))


@router.post("/stage2", summary="Etapa 2 — Processamento (M01+EML) com upload opcional de imagens PDF")
async def nc_stage2(
    request: Request,
    job_id: str = Form(..., description="ID do job (stage1 já executado)"),
    imagens_pdf_zip: Optional[UploadFile] = File(None, description="ZIP opcional com imagens PDF (N).jpg (saída do Extrair PDF). Se enviado, os e-mails embutirão as fotos."),
    lote: Optional[str] = Form(None, description="Lote 50 = ARTEMIG (templates em nc_artemig/assets/Template)"),
):
    """
    Recebe job_id + parâmetros e, opcionalmente, um ZIP com imagens extraídas do PDF.
    Lê stage1/, executa o **mesmo pipeline** que ``/nc/completo`` (e-mail → M02 → acumulado → .ics; M03 desativado).
    Se imagens_pdf_zip for enviado, é extraído em Imagens Provisórias - PDF antes do fluxo, para M02 e .eml usarem as fotos.
    Imagens PDF continuam opcionais e são mantidas para backup de extração embutido no ZIP final.
    Resposta: apenas ZIP final (com `backup/` quando houver imagens extraídas).
    Para **uma só chamada** com EAF + PDFs (sem ZIP intermédio), use POST /nc/separar (padrão) ou POST /nc/completo.
    """
    _check_auth(request)
    ws = resolve_nc_workspace(job_id)
    if not ws.job_dir.is_dir():
        raise HTTPException(404, detail="Workspace não encontrado. Execute /nc/start primeiro.")
    _touch_job_access(ws)

    zip_stage1 = ws.stage1 / "nc_separados.zip"
    if not zip_stage1.is_file():
        raise HTTPException(400, detail="stage1/nc_separados.zip não encontrado. Execute /nc/start primeiro.")

    _update_job_json(ws, status="running")
    img_zip_bytes: Optional[bytes] = None
    if imagens_pdf_zip and imagens_pdf_zip.filename:
        img_zip_bytes = await imagens_pdf_zip.read()
    try:
        return _nc_executar_pipeline_stage2_interno(
            ws,
            lote=lote,
            imagens_pdf_zip_bytes=img_zip_bytes,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error("nc/stage2: %s", traceback.format_exc())
        try:
            _update_job_json(
                ws,
                status="failed",
                log_summary={"errors": 1, "warnings": 0, "message": str(e)[:200]},
                retain_hours=2.0,
            )
        except Exception:
            pass
        raise HTTPException(500, str(e))


@router.post("/admin/cleanup-jobs", summary="Admin — Remove jobs expirados do disco")
async def nc_admin_cleanup():
    """
    Remove manualmente todos os jobs expirados em OUTPUT_PATH/nc/.
    Útil para liberar disco quando o servidor está cheio.
    A limpeza também ocorre automaticamente a cada novo workspace criado.
    """
    removed = await asyncio.to_thread(_cleanup_expired_jobs)
    base = _nc_output_path() / "nc"
    total_jobs = sum(1 for p in base.iterdir() if p.is_dir()) if base.is_dir() else 0
    return {
        "removidos": removed,
        "jobs_restantes": total_jobs,
        "output_path": str(base),
    }


@router.get("/", summary="Status do módulo NC Artesp")
async def nc_info():
    return {
        "modulo": "nc_artesp",
        "pipeline": "M01 → Email → M02 → M04 (acum. EAF) → M06 (.ics) em /nc/completo e /nc/stage2; M03 inserir Kria desativado no automático; M05/M08 manuais",
        "nc_proj_disponivel": _nc_proj_disponivel(),
        "nc_proj_path": str(_NC_PROJ),
        "nc_output_path": str(_nc_output_path()),
        "endpoints": [
            "POST /nc/completo               → EAF + PDF(s)? → ZIP (respostas_kartado_fotos/, Kartado/, acumulado/, calendario/, emails/, backup/)",
            "POST /nc/start                  → Etapa 1: upload 1 arquivo → input/ + stage1/ (job_id)",
            "POST /nc/stage2                 → job_id + ZIP imagens? → ZIP final (com backup/ se aplicável)",
            "GET  /outputs/nc/{job_id}/{subpath} → Download arquivo do job",
            "POST /nc/job                    → Criar workspace vazio (job_id)",
            "GET  /nc/job/{job_id}           → Info do workspace",
            "POST /nc/extrair-pdf             → PDF NC → ZIP nc(N).jpg + PDF(N).jpg",
            "POST /nc/analisar-pdf           → PDF NC → PDF análise (gaps KM, emergenciais, por tipo)",
            "POST /nc/separar                → M01 + pipeline (padrão) ou só ZIP XLS (entrega_completa=false)",
            "POST /nc/criar-email             → ZIP XLS + opcional imagens PDF → ZIP .eml",
            "POST /nc/gerar-modelo-foto      → M02: XLS ZIP → ZIP Kria + Resposta",
            "POST /nc/inserir-conservacao    → M03: Kria → ZIP Kcor-Kria Conservação",
            "POST /nc/inserir-meio-ambiente  → M07: Kria MA → ZIP Kcor-Kria MA",
            "POST /nc/juntar                 → M04: Kcor → XLSX Acumulado",
            "POST /nc/inserir-numero         → M05 manual: preencher coluna Y no acumulado",
            "POST /nc/exportar-calendario    → M06: Acumulado → .ics (iCalendar)",
            "POST /nc/organizar-imagens      → M08: Acumulado + ZIP imagens → ZIP classificado por tipo",
            "POST /nc/admin/cleanup-jobs     → Remove jobs expirados do disco (manutenção)",
        ],
    }
