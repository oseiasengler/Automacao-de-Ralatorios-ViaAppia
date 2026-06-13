"""Acumulado EAF→Kcor: coluna W (Arquivos) com pdf completo e coluna sequência detetada."""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

_REPO = Path(__file__).resolve().parents[2]
_NC_ARTESP = _REPO / "nc_artesp"
if str(_NC_ARTESP) not in sys.path:
    sys.path.insert(0, str(_NC_ARTESP))

from modulos.juntar_arquivos import (  # noqa: E402
    _chave_dedupe_registro_kcor,
    _eaf_linha_para_registro_kcor,
)
from modulos import separar_nc as sn  # noqa: E402


def _linha_minima_eaf(ws, row: int, seq_col: int, seq_val, codigo: str = "NC.99.1") -> None:
    ws.cell(row=row, column=sn.COL_CODIGO).value = codigo
    ws.cell(row=row, column=sn.COL_DATA_CON).value = datetime(2026, 5, 8)
    ws.cell(row=row, column=sn.COL_RODOVIA).value = "SP-075"
    ws.cell(row=row, column=sn.COL_KM_I_FULL).value = "100+000"
    ws.cell(row=row, column=sn.COL_SENTIDO).value = "Norte"
    ws.cell(row=row, column=sn.COL_TIPO_NC).value = "Pavimento"
    ws.cell(row=row, column=seq_col).value = seq_val


def test_acumulado_w_usa_coluna_seq_foto_detetada_quando_nao_e_v():
    wb = Workbook()
    ws = wb.active
    seq_col = 24
    ws.cell(row=2, column=seq_col).value = "Sequência foto"
    _linha_minima_eaf(ws, 5, seq_col=seq_col, seq_val=42)
    det = sn._detectar_col_seq_foto(ws, fallback=sn.COL_SEQ_FOTO)
    assert det == seq_col
    linha = _eaf_linha_para_registro_kcor(
        ws,
        5,
        col_data_reparo=sn.COL_DATA_NC,
        col_data_envio=19,
        col_tipo_nc=sn.COL_TIPO_NC,
        col_seq_foto=det,
    )
    arquivos = linha[22]
    assert arquivos
    assert ";pdf (42).jpg" in arquivos


def test_acumulado_w_fallback_pdf_usa_codigo_se_seq_vazia():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=2, column=22).value = "Seq foto"
    _linha_minima_eaf(ws, 5, seq_col=22, seq_val=None, codigo="HE.13.0999")
    det = sn._detectar_col_seq_foto(ws, fallback=22)
    linha = _eaf_linha_para_registro_kcor(
        ws,
        5,
        col_data_reparo=sn.COL_DATA_NC,
        col_data_envio=19,
        col_tipo_nc=sn.COL_TIPO_NC,
        col_seq_foto=det,
    )
    arquivos = linha[22]
    assert ";pdf (HE.13.0999).jpg" in arquivos


def test_acumulado_w_seq_zero_usa_codigo_no_pdf():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=2, column=22).value = "Seq foto"
    _linha_minima_eaf(ws, 5, seq_col=22, seq_val=0, codigo="HE.13.0999")
    det = 22
    linha = _eaf_linha_para_registro_kcor(
        ws,
        5,
        col_data_reparo=sn.COL_DATA_NC,
        col_data_envio=19,
        col_tipo_nc=sn.COL_TIPO_NC,
        col_seq_foto=det,
    )
    arquivos = linha[22]
    assert ";pdf (HE.13.0999).jpg" in arquivos


def test_acumulado_w_pdf_usa_codigo_se_seq_e_nome_proprio():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=2, column=22).value = "Seq foto"
    _linha_minima_eaf(
        ws,
        5,
        seq_col=22,
        seq_val="Rogerio Aparecido de Aguiar",
        codigo="HE.13.0999",
    )
    det = 22
    linha = _eaf_linha_para_registro_kcor(
        ws,
        5,
        col_data_reparo=sn.COL_DATA_NC,
        col_data_envio=19,
        col_tipo_nc=sn.COL_TIPO_NC,
        col_seq_foto=det,
    )
    assert ";pdf (HE.13.0999).jpg" in linha[22]
    assert "Rogerio" not in linha[22]


def test_detectar_col_seq_foto_ignora_cabecalho_responsavel():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=21, value="Responsável")
    ws.cell(row=1, column=22, value="Seq. foto")
    assert sn._detectar_col_seq_foto(ws, fallback=sn.COL_SEQ_FOTO) == 22


def test_chave_dedupe_usa_codigo_fiscal_quando_presente():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=2, column=22).value = "Seq foto"
    _linha_minima_eaf(ws, 5, seq_col=22, seq_val=1, codigo="HE.13.0001")
    _linha_minima_eaf(ws, 6, seq_col=22, seq_val=1, codigo="HE.13.0002")
    det = 22
    a = _eaf_linha_para_registro_kcor(
        ws,
        5,
        col_data_reparo=sn.COL_DATA_NC,
        col_data_envio=19,
        col_tipo_nc=sn.COL_TIPO_NC,
        col_seq_foto=det,
    )
    b = _eaf_linha_para_registro_kcor(
        ws,
        6,
        col_data_reparo=sn.COL_DATA_NC,
        col_data_envio=19,
        col_tipo_nc=sn.COL_TIPO_NC,
        col_seq_foto=det,
    )
    a26 = list(a) + ["HE.13.0001"]
    b26 = list(b) + ["HE.13.0002"]
    assert _chave_dedupe_registro_kcor(a26) != _chave_dedupe_registro_kcor(b26)


def test_chave_dedupe_distinta_para_duas_ncs():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=2, column=22).value = "Seq foto"
    _linha_minima_eaf(ws, 5, seq_col=22, seq_val=1, codigo="HE.13.0001")
    _linha_minima_eaf(ws, 6, seq_col=22, seq_val=2, codigo="HE.13.0002")
    det = 22
    a = _eaf_linha_para_registro_kcor(
        ws,
        5,
        col_data_reparo=sn.COL_DATA_NC,
        col_data_envio=19,
        col_tipo_nc=sn.COL_TIPO_NC,
        col_seq_foto=det,
    )
    b = _eaf_linha_para_registro_kcor(
        ws,
        6,
        col_data_reparo=sn.COL_DATA_NC,
        col_data_envio=19,
        col_tipo_nc=sn.COL_TIPO_NC,
        col_seq_foto=det,
    )
    assert _chave_dedupe_registro_kcor(a) != _chave_dedupe_registro_kcor(b)


def test_chave_dedupe_igual_para_mesma_linha_lida_duas_vezes():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=2, column=22).value = "Seq foto"
    _linha_minima_eaf(ws, 5, seq_col=22, seq_val=7, codigo="HE.13.0999")
    det = 22
    a = _eaf_linha_para_registro_kcor(
        ws,
        5,
        col_data_reparo=sn.COL_DATA_NC,
        col_data_envio=19,
        col_tipo_nc=sn.COL_TIPO_NC,
        col_seq_foto=det,
    )
    b = _eaf_linha_para_registro_kcor(
        ws,
        5,
        col_data_reparo=sn.COL_DATA_NC,
        col_data_envio=19,
        col_tipo_nc=sn.COL_TIPO_NC,
        col_seq_foto=det,
    )
    assert _chave_dedupe_registro_kcor(a) == _chave_dedupe_registro_kcor(b)
