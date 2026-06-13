"""nc_artemig: normalização de texto PDF e datas Artemig."""

from types import SimpleNamespace


def test_excel_complemento_nao_mescla_campos_texto_kcor_lote50():
    from nc_artesp.modulos.analisar_pdf_nc import _excel_complemento_pode_mesclar_campo

    assert _excel_complemento_pode_mesclar_campo("50", "tipo_atividade") is False
    assert _excel_complemento_pode_mesclar_campo("50", "atividade") is False
    assert _excel_complemento_pode_mesclar_campo("50", "prazo_str") is False
    assert _excel_complemento_pode_mesclar_campo("50", "km_ini_str") is False
    assert _excel_complemento_pode_mesclar_campo("50", "km_fim_str") is False
    assert _excel_complemento_pode_mesclar_campo("50", "horario_fiscalizacao") is True
    assert _excel_complemento_pode_mesclar_campo("13", "tipo_atividade") is True


def test_relatorio_xlsx_lote50_col_w_igual_exportar_kcor_arquivos():
    import io

    from openpyxl import load_workbook

    from nc_artemig.exportar_kcor_planilha import (
        _escapar_inicio_formula_excel,
        _excel_valor_texto_ou_none,
        _montar_v_w_kcor,
    )
    from nc_artesp.modulos.analisar_pdf_nc import NcItem, gerar_relatorio_xlsx

    nc = NcItem(
        codigo="CE2607782",
        data_con="01/01/2026",
        rodovia="MG 050",
        lote="50",
        artemig_pdf_stem="NOT_2026_001",
        artemig_kcor_nomes_arquivos=["nc (2607782).jpg"],
    )
    _, w_arq = _montar_v_w_kcor(nc)
    esperado = _excel_valor_texto_ou_none(
        _escapar_inicio_formula_excel(w_arq) if w_arq else ""
    )
    raw = gerar_relatorio_xlsx([nc], "50", "")
    wb = load_workbook(io.BytesIO(raw))
    ws = wb.active
    assert ws.cell(row=5, column=23).value == esperado


def test_detectar_colunas_saida_q_observacoes_fiscalizacao_nao_e_obs_kcor():
    """Col. Q «Observações da fiscalização» (ex-Atividade): dados em atividade; obs Kcor continua na U."""
    from openpyxl import Workbook

    from nc_artesp.modulos.analisar_pdf_nc import _detectar_colunas_saida_template

    wb = Workbook()
    ws = wb.active
    ws.cell(row=4, column=17, value="Observações da fiscalização")
    ws.cell(row=4, column=21, value="Observações")
    m = _detectar_colunas_saida_template(ws, cabecalho_fim=4)
    assert m["atividade"] == 17
    assert m.get("obs_linha_kcor") == 21


def test_normalizar_texto_extraido_pdf_nbsp_zw():
    from nc_artemig.texto_pdf import normalizar_texto_extraido_pdf

    raw = "MG\u00a0050\u200bSH02"
    out = normalizar_texto_extraido_pdf(raw)
    assert "\u00a0" not in out
    assert "\u200b" not in out
    assert "MG" in out and "050" in out and "SH02" in out


def test_data_artemig_dd_mm_yyyy_com_ruido():
    from nc_artesp.modulos.analisar_pdf_nc import _data_artemig_dd_mm_yyyy

    assert _data_artemig_dd_mm_yyyy("15/03/26 ") == "15/03/2026"
    assert _data_artemig_dd_mm_yyyy("Data \u00a0 01/12/2025 fim") == "01/12/2025"


def test_local_coluna_j_faixa_dominio():
    from nc_artemig.exportar_kcor_planilha import _local_coluna_j

    nc = SimpleNamespace(atividade="Trecho FAIXA DE DOMINIO", tipo_atividade="", grupo_atividade="")
    assert _local_coluna_j(nc) == "Faixa de Domínio"
    nc2 = SimpleNamespace(atividade="FX. marginal", tipo_atividade="", grupo_atividade="")
    assert _local_coluna_j(nc2) == "Faixa de Domínio"
    nc3 = SimpleNamespace(atividade="Pista rolamento", tipo_atividade="", grupo_atividade="")
    assert _local_coluna_j(nc3) == "Faixa de Rolamento"


def test_patologia_para_kcor_macro_inexistencia_defensa():
    from nc_artemig.exportar_kcor_planilha import _patologia_para_kcor

    # "Inexistência de elementos refletivos" → Barreira rígida (parametrização confirmada)
    k, cl = _patologia_para_kcor("Inexistência de elementos refletivos", "", "")
    assert "Barreira" in k
    assert cl == "Eng. QID"

    # "Inexistência de defensa metálica" → Defensa metálica
    k2, _ = _patologia_para_kcor("Inexistência de defensa metálica", "", "")
    assert "Defensa" in k2


def test_patologia_para_kcor_macro_guarda_corpo():
    from nc_artemig.exportar_kcor_planilha import _patologia_para_kcor

    k, _ = _patologia_para_kcor("Guarda corpo metálico danificado", "", "")
    assert "Barreira" in k


def test_patologia_para_kcor_macro_placas_advertencia():
    from nc_artemig.exportar_kcor_planilha import _patologia_para_kcor

    k, _ = _patologia_para_kcor("Vandalismo placas de advertência na via", "", "")
    assert "Placas - Regulam" in k


def test_patologia_para_kcor_sem_mapeamento_fica_em_branco():
    """Patologia sem regra no macro VBA → Col. E vazia (aguarda parametrização no Kcor)."""
    from nc_artemig.exportar_kcor_planilha import _patologia_para_kcor

    # "Prédio e Pátio" não bate em nenhuma condição do macro
    k, cl = _patologia_para_kcor("Prédio e Pátio", "", "")
    assert k == ""
    assert cl == "Eng. QID"

    k2, _ = _patologia_para_kcor("Elemento de concreto danificado", "", "")
    assert k2 == ""


def test_patologia_para_kcor_buracos_macro_sempre_emergencial():
    """Macro VBA: só ``Left(patologia,7)='Buracos'`` → Emergencial (sem Reparo técnico no mapa)."""
    from types import SimpleNamespace

    from nc_artemig.exportar_kcor_planilha import _patologia_para_kcor

    pat = "Buracos e/ou panelas na pista"
    ind = "Buracos / Panelas"
    atv = "Panela na pista"
    nc_em = SimpleNamespace(
        emergencial=True,
        prazo_dias=1,
        tipo_atividade="Panelas - Emergencial",
    )
    nc_rt = SimpleNamespace(
        emergencial=False,
        prazo_dias=15,
        tipo_atividade="Panelas - Reparo Técnico",
    )
    ke, _ = _patologia_para_kcor(pat, ind, atv, nc_em)
    kt, _ = _patologia_para_kcor(pat, ind, atv, nc_rt)
    assert "Emergencial" in ke
    assert "Emergencial" in kt
    assert "Reparo" not in ke and "Reparo" not in kt


def test_export_kcor_panela_duas_linhas_col_e_emergencial_e_reparo_tecnico():
    """Evento emergencial → «Buracos e panelas - Emergencial»; reparo técnico → «Buracos e panelas - Reparo técnico»."""
    import io
    from types import SimpleNamespace

    from openpyxl import load_workbook

    from nc_artemig.config import COL_KCOR_KRIA
    from nc_artemig.exportar_kcor_planilha import gerar_exportar_kcor_xlsx_bytes

    def _nc(cod: str, emerg: bool, pd: int, tipo_act: str, act: str):
        return SimpleNamespace(
            lote="50",
            codigo=cod,
            data_con="10/04/2026",
            prazo_dias=pd,
            emergencial=emerg,
            tipo_panela=True,
            km_ini=100.0,
            km_fim=100.0,
            km_ini_str="100,000",
            km_fim_str="100,000",
            rodovia="MG-050",
            sentido="Crescente",
            atividade=act,
            tipo_atividade=tipo_act,
            grupo_atividade="Pavimento",
            observacao="",
            num_consol="2609991",
            sh_artemig="SH02",
            patologia_artemig="Buracos e/ou panelas na pista",
            indicador_artemig="Buracos / Panelas",
            artemig_pdf_stem="X",
            artemig_kcor_nomes_arquivos=[],
            artemig_kcor_paginas_jpg=[],
            nome_fiscal="",
            prazo_str="11/04/2026" if emerg else "25/04/2026",
        )

    ncs = [
        _nc(
            "202609991001",
            True,
            1,
            "Panelas - Emergencial",
            "Panela na pista — Panelas - Emergencial",
        ),
        _nc(
            "202609991002",
            False,
            15,
            "Panelas - Reparo Técnico",
            "Panela na pista — Panelas - Reparo Técnico",
        ),
    ]

    b, meta = gerar_exportar_kcor_xlsx_bytes(ncs)
    assert meta.get("ok") and b
    wb = load_workbook(io.BytesIO(b))
    ws = wb["Dados"] if "Dados" in wb.sheetnames else wb.active
    ck = COL_KCOR_KRIA
    t2 = str(ws.cell(2, ck["Tipo"]).value or "")
    t3 = str(ws.cell(3, ck["Tipo"]).value or "")
    assert "Emergencial" in t2, f"Linha emergencial deve ter «Emergencial»; obtido: {t2!r}"
    assert "Reparo" in t3, f"Linha reparo técnico deve ter «Reparo»; obtido: {t3!r}"
    assert "Reparo" not in t2, f"Linha emergencial não deve ter «Reparo»; obtido: {t2!r}"
    assert ws.cell(2, ck["Prazo"]).value == 1
    assert ws.cell(3, ck["Prazo"]).value == 15
    # Data de Suspensão deve ficar vazia
    assert ws.cell(2, ck["Data_Suspensao"]).value in (None, ""), "Data_Suspensao deve estar vazia"
    assert ws.cell(3, ck["Data_Suspensao"]).value in (None, ""), "Data_Suspensao deve estar vazia"


def test_km_normalizado_nas01_nao_divide_km_decimal_artemig():
    """653,4 km não pode virar 0,6534 (regra antiga «>500 → /1000»)."""
    from nc_artemig.exportar_kcor_planilha import _km_normalizado_nas01

    assert _km_normalizado_nas01(653.4) == 653.4
    assert _km_normalizado_nas01(653.0) == 653.0
    assert _km_normalizado_nas01(653400) == 653.4
    assert _km_normalizado_nas01(12000) == 12.0


def test_export_kcor_col_p_dt_fim_prog_apos_contagem_emergencial():
    import io
    from types import SimpleNamespace

    from openpyxl import load_workbook

    from nc_artemig.config import COL_KCOR_KRIA
    from nc_artemig.exportar_kcor_planilha import gerar_exportar_kcor_xlsx_bytes

    nc = SimpleNamespace(
        lote="50",
        codigo="202603080000",
        data_con="10/04/2026",
        prazo_dias=24,
        emergencial=True,
        tipo_panela=False,
        km_ini=100.0,
        km_fim=100.0,
        km_ini_str="100,000",
        km_fim_str="100,000",
        rodovia="MG-050",
        sentido="Crescente",
        atividade="Buracos e/ou panelas",
        tipo_atividade="Panelas",
        grupo_atividade="Pavimento",
        observacao="",
        num_consol="2607782",
        sh_artemig="SH02",
        patologia_artemig="Buracos e/ou panelas na pista",
        indicador_artemig="Buracos / Panelas",
        artemig_pdf_stem="X",
        artemig_kcor_nomes_arquivos=[],
        artemig_kcor_paginas_jpg=[],
        nome_fiscal="",
        prazo_str="11/04/2026",
    )
    b, meta = gerar_exportar_kcor_xlsx_bytes([nc])
    assert meta.get("ok") and b
    wb = load_workbook(io.BytesIO(b))
    ws = wb["Dados"] if "Dados" in wb.sheetnames else wb.active
    ck = COL_KCOR_KRIA
    r = 2
    ini = str(ws.cell(r, ck["Dt_Inicio_Prog"]).value or "").strip()
    fim = str(ws.cell(r, ck["Dt_Fim_Prog"]).value or "").strip()
    assert ini.startswith("10/04/2026")
    assert fim.startswith("11/04/2026")
    assert fim.split()[0] != ini.split()[0]
    prazo_cell = ws.cell(r, ck["Prazo"])
    assert prazo_cell.alignment is not None
    assert prazo_cell.alignment.horizontal == "center"


def test_stem_subpasta_fotos_col_v_usa_stem_pdf_nao_pavimento_fixo():
    from types import SimpleNamespace

    from nc_artemig.exportar_kcor_planilha import _stem_subpasta_fotos

    nc = SimpleNamespace(
        codigo="202506768",
        num_consol="2516934",
        artemig_pdf_stem="NOT-25-06768_DRENAGEM_CE2516934",
        grupo_atividade="",
        tipo_atividade="",
        indicador_artemig="",
        patologia_artemig="",
        atividade="",
    )
    s = _stem_subpasta_fotos(nc)
    assert "DRENAGEM" in s
    assert "PAVIMENTO" not in s


def test_montar_col_w_lista_nomes_multiplos_jpg():
    """Col W inclui todas as fotos: base + sufixos (N) gerados por _nome_unico."""
    from types import SimpleNamespace

    from nc_artemig.exportar_kcor_planilha import _montar_v_w_kcor

    # Formato real gerado por _nome_unico: nc (COD) (N).jpg
    nc = SimpleNamespace(
        codigo="202506999",
        num_consol="",
        emergencial=False,
        prazo_dias=None,
        artemig_pdf_stem="Relato1",
        artemig_kcor_nomes_arquivos=[
            "nc (202506999).jpg",
            "nc (202506999) (1).jpg",
            "nc (202506999) (2).jpg",
        ],
        artemig_kcor_paginas_jpg=[],
        atividade="",
        tipo_atividade="",
        grupo_atividade="",
    )
    _v, w = _montar_v_w_kcor(nc)
    assert "nc (202506999).jpg" in w
    assert "nc (202506999) (1).jpg" in w
    assert "nc (202506999) (2).jpg" in w
    assert w.count(";") >= 3


def test_montar_col_w_lista_nomes_formato_legado_underscore():
    """Col W também aceita o formato legado nc (COD)_N.jpg."""
    from types import SimpleNamespace

    from nc_artemig.exportar_kcor_planilha import _montar_v_w_kcor

    nc = SimpleNamespace(
        codigo="202506999",
        num_consol="",
        emergencial=False,
        prazo_dias=None,
        artemig_pdf_stem="Relato1",
        artemig_kcor_nomes_arquivos=[
            "nc (202506999).jpg",
            "nc (202506999)_1.jpg",
            "nc (202506999)_2.jpg",
        ],
        artemig_kcor_paginas_jpg=[],
        atividade="",
        tipo_atividade="",
        grupo_atividade="",
    )
    _v, w = _montar_v_w_kcor(nc)
    assert "nc (202506999).jpg" in w
    # após _lista_arquivos_coluna_w_sanear, _ → espaço
    assert "nc (202506999) 1.jpg" in w
    assert "nc (202506999) 2.jpg" in w
    assert w.count(";") >= 3


def test_prazo_artemig_em_ate_24_horas_com_parenteses():
    from nc_artesp.modulos.analisar_pdf_nc import _prazo_artemig

    texto = (
        "Prazo para Atendimento à Notificação: Remendo Emergencial: "
        "em até 24 (vinte e quatro) horas, a partir da data do recebimento desta notificação."
    )
    prazo_str, prazo_dias, emerg = _prazo_artemig(texto, "10/04/2026")
    assert prazo_dias == 1
    assert emerg is True
    assert prazo_str == "11/04/2026"


def test_prazo_artemig_em_ate_ndias_com_parenteses_extenso():
    from nc_artesp.modulos.analisar_pdf_nc import _prazo_artemig

    texto = (
        "Prazo para Atendimento à Notificação: "
        "em até 30 (trinta) dias corridos, a partir da data do recebimento desta notificação."
    )
    prazo_str, prazo_dias, emerg = _prazo_artemig(texto, "01/04/2026")
    assert prazo_dias == 30
    assert emerg is False
    assert prazo_str == "01/05/2026"


def test_prazo_artemig_prazo_maximo_de_ndias_na_janela():
    from nc_artesp.modulos.analisar_pdf_nc import _prazo_artemig

    texto = (
        "Prazo para Atendimento à Notificação: "
        "prazo máximo de 60 (sessenta) dias corridos."
    )
    prazo_str, prazo_dias, emerg = _prazo_artemig(texto, "01/01/2026")
    assert prazo_dias == 60
    assert emerg is False
    assert prazo_str == "02/03/2026"


def test_prazo_artemig_em_at_ndias_sem_letra_apos_at():
    """PDFs/encodings em que «até» perde o «e»/«é» («Em at 5 dias»)."""
    from nc_artesp.modulos.analisar_pdf_nc import _prazo_artemig

    texto = (
        "Prazo para Atendimento à Notificação: "
        "Em at 5 (cinco) dias, a partir da data do recebimento desta notificação."
    )
    prazo_str, prazo_dias, emerg = _prazo_artemig(texto, "01/04/2026")
    assert prazo_dias == 5
    assert emerg is False
    assert prazo_str == "06/04/2026"


def test_prazo_artemig_buracos_so_primeiro_prazo_emergencial_ignora_tecnico():
    from nc_artesp.modulos.analisar_pdf_nc import _prazo_artemig

    texto = (
        "Prazo para Atendimento à Notificação: "
        "Remendo Emergencial: em até 24 (vinte e quatro) horas, a partir da data.\n"
        "Remendo Técnico: no prazo máximo de 15 (quinze) dias, a partir da data."
    )
    prazo_str, prazo_dias, emerg = _prazo_artemig(texto, "15/09/2025")
    assert prazo_dias == 1
    assert emerg is True
    assert prazo_str == "16/09/2025"


def test_prazo_artemig_solo_remendo_tecnico_15_dias():
    from nc_artesp.modulos.analisar_pdf_nc import _prazo_artemig

    texto = (
        "Prazo para Atendimento à Notificação: "
        "Remendo Técnico: no prazo máximo de 15 (quinze) dias, a partir da data."
    )
    prazo_str, prazo_dias, emerg = _prazo_artemig(texto, "15/09/2025")
    assert prazo_dias == 15
    assert emerg is False
    assert prazo_str == "30/09/2025"


def test_indicador_patologia_drenagem_subterranea_patologia_completa():
    from nc_artesp.modulos.analisar_pdf_nc import _indicador_patologia_de_resto_artemig

    resto = "Drenagem Subterrânea Drenagem subterrânea obstruída"
    ind, pat = _indicador_patologia_de_resto_artemig(resto)
    assert "Subterr" in ind
    assert "obstru" in pat.lower()


def test_indicador_patologia_parametros_gerais_patologia_completa():
    from nc_artesp.modulos.analisar_pdf_nc import _indicador_patologia_de_resto_artemig

    resto = "Parâmetros Gerais\nInexistência de tachas e tachões"
    ind, pat = _indicador_patologia_de_resto_artemig(resto)
    assert "Parâmetros" in ind and "Gerais" in ind
    assert "Inexistência" in pat
    assert "tachas" in pat.lower()


def test_indicador_patologia_gerais_parametros_ordem_colunas_pdf():
    """PDF linearizado: «Gerais» + «Parâmetros» (ordem de colunas) não vira só «Gerais (Parâmetros)»."""
    from nc_artesp.modulos.analisar_pdf_nc import _indicador_patologia_de_resto_artemig

    resto = "Gerais Parâmetros Inexistência de sinalização vertical"
    ind, pat = _indicador_patologia_de_resto_artemig(resto)
    assert "Parâmetros" in ind and "Gerais" in ind
    assert "Inexistência" in pat
    assert "sinalização vertical" in pat.lower()
    assert "Gerais (Parâmetros)" not in pat


def test_indicador_patologia_resto_multilinha_antes_de_e_ou():
    from nc_artesp.modulos.analisar_pdf_nc import _indicador_patologia_de_resto_artemig

    resto = (
        "Buracos / Panelas e \n"
        "Deformação permanente\n"
        "Buracos e/ou panelas na pista de \n"
        "rolamento"
    )
    ind, pat = _indicador_patologia_de_resto_artemig(resto)
    assert "Deformação permanente" in ind
    assert "Panelas e" in ind
    assert "e/ou" in pat
    assert "rolamento" in pat


def test_indicador_patologia_resto_buracos_e_ou_colado():
    from nc_artesp.modulos.analisar_pdf_nc import _indicador_patologia_de_resto_artemig

    r = (
        "Buracos / Panelas e Deformação permanente "
        "Buracos e/ou panelas na pista de rolamento."
    )
    ind, pat = _indicador_patologia_de_resto_artemig(r)
    assert "Deformação permanente" in ind
    assert "e/ou" in pat
    assert "pista de rolamento" in pat


def test_col_u_enriquece_patologia_curta_com_tipo_planilha():
    from nc_artemig.exportar_kcor_planilha import _texto_observacoes_nas01

    nc = SimpleNamespace(
        patologia_artemig="Panelas e",
        tipo_atividade="Panelas e buracos emergencial na pista",
        grupo_atividade="Buracos / Panelas e Deformação permanente",
        indicador_artemig="Buracos",
        atividade="",
        observacao="",
        codigo="202603080",
        num_consol="2607782",
        sh_artemig="SH02",
    )
    u = _texto_observacoes_nas01(nc)
    assert "202603080" in u and "SH02" in u and "2607782" in u
    assert "Buracos" in u and "Panelas e buracos emergencial na pista" in u
    assert " · " in u
    assert "causandoaspecto" not in u.lower()


def test_col_u_patologia_panelas_e_completa_so_com_tipo_pdf_sem_atividade():
    """NC real: patologia PDF curta + tipo (mesmo PDF) longo; atividade pode vir vazia."""
    from types import SimpleNamespace

    from nc_artemig.exportar_kcor_planilha import _texto_observacoes_nas01

    nc = SimpleNamespace(
        patologia_artemig="Panelas e",
        tipo_atividade="Panelas e buracos emergencial na pista",
        grupo_atividade="Buracos / Panelas e Deformação permanente",
        indicador_artemig="Buracos",
        atividade="",
        observacao="",
        codigo="202501364",
        num_consol="2516928",
        sh_artemig="SH18",
    )
    u = _texto_observacoes_nas01(nc)
    assert "202501364" in u and "SH18" in u and "2516928" in u
    assert "Buracos" in u and "Panelas e buracos emergencial na pista" in u
    assert " · " in u


def test_col_u_patologia_completa_desde_atividade_quando_pdf_truncado():
    from types import SimpleNamespace

    from nc_artemig.exportar_kcor_planilha import _texto_observacoes_nas01

    nc = SimpleNamespace(
        patologia_artemig="Panelas e",
        tipo_atividade="",
        grupo_atividade="",
        indicador_artemig="Buracos",
        atividade="Buracos e/ou panelas na pista de rolamento",
        observacao="",
        codigo="202501364",
        num_consol="2516928",
        sh_artemig="SH18",
    )
    u = _texto_observacoes_nas01(nc)
    assert "202501364" in u and "SH18" in u and "2516928" in u
    assert "Buracos e/ou panelas na pista de rolamento" in u
    assert " · " in u


def test_col_u_nao_duplica_indicador_patologia_iguais_gerais():
    from types import SimpleNamespace

    from nc_artemig.exportar_kcor_planilha import _texto_observacoes_nas01

    g = "Gerais (Parâmetros)"
    nc = SimpleNamespace(
        patologia_artemig=g,
        tipo_atividade="",
        grupo_atividade=g,
        indicador_artemig=g,
        atividade="",
        observacao="",
        codigo="202501363",
        num_consol="2516927",
        sh_artemig="SH05",
    )
    u = _texto_observacoes_nas01(nc)
    assert "202501363" in u and "SH05" in u and "2516927" in u
    assert "Gerais (Parâmetros)" in u
    assert u.count("Gerais (Parâmetros)") == 1
    assert " · " in u


def test_col_u_patologia_panelas_e_completa_com_tipo_curto_mais_3_chars():
    """Antes exigia len(tipo) > len(pat)+14; «Panelas e buracos» não substituía «Panelas e»."""
    from types import SimpleNamespace

    from nc_artemig.exportar_kcor_planilha import _texto_observacoes_nas01

    nc = SimpleNamespace(
        patologia_artemig="Panelas e",
        tipo_atividade="Panelas e buracos",
        grupo_atividade="Buracos",
        indicador_artemig="Buracos",
        atividade="",
        observacao="",
        codigo="202501365",
        num_consol="2516929",
        sh_artemig="SH20",
    )
    u = _texto_observacoes_nas01(nc)
    assert "202501365" in u and "SH20" in u and "2516929" in u
    assert "Buracos" in u and "Panelas e buracos" in u
    assert " · " in u


def test_col_u_observacoes_usa_patologia_tipo_excel_nao_grupo_indicador():
    """Col. U: rótulos PDF — Indicador (grupo Excel) antes de Patologia (tipo), sem confundir os campos."""
    from nc_artemig.exportar_kcor_planilha import _texto_observacoes_nas01

    nc = SimpleNamespace(
        patologia_artemig="",
        tipo_atividade="Panelas e buracos emergencial",
        grupo_atividade="Pavimento",
        indicador_artemig="",
        atividade="Limpeza da área",
        observacao="",
        codigo="202506123456",
        num_consol="987654",
        sh_artemig="SH02",
    )
    u = _texto_observacoes_nas01(nc)
    assert "202506123456" in u
    assert "SH02" in u
    assert "987654" in u
    assert "Pavimento" in u
    assert "Panelas e buracos emergencial" in u
    assert "Limpeza da área" in u
    assert " · " in u


def test_col_u_vegetacao_parametros_gerais_texto_legivel_com_separadores():
    from nc_artemig.exportar_kcor_planilha import _texto_observacoes_nas01

    nc = SimpleNamespace(
        patologia_artemig="Vegetação fora padrão",
        tipo_atividade="Vegetação fora padrão",
        grupo_atividade="Parâmetros Gerais",
        indicador_artemig="Parâmetros Gerais",
        atividade="",
        observacao="",
        codigo="202603720",
        num_consol="2609065",
        sh_artemig="SH01",
    )
    u = _texto_observacoes_nas01(nc)
    assert "Vegetação fora padrão (Parâmetros Gerais)" in u
    assert "Prazo para Atendimento à Notificação:" in u
    assert "Em até 5 (cinco) dias" in u
    assert "Foi verificado" in u and "monitoramento" in u
    assert "causando aspecto visual" in u
    assert "propagação de incêndios" in u
    assert "condições de segurança" in u
    assert "causandoaspecto" not in u.lower()
    assert "condiçõesdesegurança" not in u.lower()
    assert "SH01" in u
    assert "Notificação: 202603720" in u
    assert "Nº Consol: 2609065" in u
    assert u.count("Prazo para Atendimento à Notificação:") == 1
    assert "\n" not in u
    assert "Descrição" not in u


def test_extrair_descricao_artemig_remove_descrição_duplicada():
    from nc_artesp.modulos.analisar_pdf_nc import _extrair_descricao_atividade_artemig

    texto = (
        "Bloco\nDescrição:\nDescrição: Foi verificado, em rotina diária de monitoramento, "
        "a existência de vandalismo na passagem inferior de veículos e pedestres, "
        "ocasionando desconforto e aspecto visual desagradável aos usuários da rodovia. "
        "EAF: CONSOL"
    )
    out = _extrair_descricao_atividade_artemig(texto)
    assert out.startswith("Foi verificado")
    assert "Descrição" not in out
    assert "vandalismo" in out.lower()


def test_extrair_descricao_QID_conso_antes_do_paragrafo_NOT25_estilo():
    """NOT-25-01363…: primeiro «Descrição:» seguido de Nº CONSOL/tabela; o parágrafo vem no segundo rótulo."""
    from nc_artesp.modulos.analisar_pdf_nc import _extrair_descricao_atividade_artemig

    texto = (
        "NOTIFICAÇÃO\nDescrição:\nNº da CONSOL:\n2516927\nREGISTRO\nMG-050 SH05\n"
        "Em até 5 (cinco) dias.\n"
        "Descrição: Foi verificado, em rotina diária de monitoramento, vandalismo na via.\n"
        "LOCALIZAÇÃO\n202501363\n"
    )
    out = _extrair_descricao_atividade_artemig(texto)
    assert "Foi verificado" in out
    assert "vandalismo" in out.lower()
    assert "2516927" not in out
    assert "MG-050" not in out


def test_col_t_obs_gestor_remove_marcador_escape_excel():
    from nc_artemig.exportar_kcor_planilha import _bloco_obs_gestor_nas01

    nc = SimpleNamespace(
        sh_artemig="SH01_x000D_",
        codigo="202603720_x000D_",
        num_consol="2609065_x000D_",
    )
    t = _bloco_obs_gestor_nas01(nc)
    assert "_x000D_" not in t
    assert "Trecho Homogênio: SH01" in t
    assert "Notificação: 202603720" in t
    assert "Consol: 2609065" in t


def test_nascentes_panela_gera_eventos_emergencial_e_reparo_tecnico():
    from nc_artesp.modulos.analisar_pdf_nc import (
        NcItem,
        _expandir_eventos_panela_nascentes,
    )

    nc = NcItem(
        codigo="900001",
        data_con="10/04/2026",
        atividade="Panela na pista",
        tipo_atividade="Panelas",
        prazo_str="11/04/2026",
        prazo_dias=1,
        emergencial=True,
        tipo_panela=True,
    )
    out = _expandir_eventos_panela_nascentes([nc], "50")
    assert len(out) == 2
    prazos = sorted([x.prazo_dias for x in out if x.prazo_dias is not None])
    assert prazos == [1, 15]
    tipos = " | ".join((x.tipo_atividade or "") for x in out)
    assert "Emergencial" in tipos
    assert "Reparo Técnico" in tipos
    assert any((x.prazo_str or "") == "11/04/2026" for x in out)
    assert any((x.prazo_str or "") == "25/04/2026" for x in out)
    assert any(" — Panelas - Emergencial" in (x.atividade or "") for x in out)
    assert any(" — Panelas - Reparo Técnico" in (x.atividade or "") for x in out)


def test_outros_lotes_nao_expandem_eventos_panela():
    from nc_artesp.modulos.analisar_pdf_nc import (
        NcItem,
        _expandir_eventos_panela_nascentes,
    )

    nc = NcItem(
        codigo="900001",
        data_con="10/04/2026",
        atividade="Panela na pista",
        tipo_atividade="Panelas",
        prazo_str="11/04/2026",
        prazo_dias=1,
        emergencial=True,
        tipo_panela=True,
    )
    out = _expandir_eventos_panela_nascentes([nc], "13")
    assert len(out) == 1
    assert out[0].prazo_dias == 1


def test_parse_artemig_texto_km_inicial_final_rotulos_pdf():
    from nc_artesp.modulos.analisar_pdf_nc import _parse_artemig_texto

    texto = (
        "NOTIFICAÇÃO CONSOL MG-050\n"
        "202509999 01/01/26 10:00 Parâmetros Gerais Inexistência de tachas\n"
        "LOCALIZAÇÃO\n"
        "202509999\n"
        "Rodovia: MG-050 SH16 Km Inicial: 543+500 Km Final: 544+000 Sentido: CRESCENTE\n"
        "Nº da CONSOL: 2516932\n"
    )
    nc = _parse_artemig_texto(texto)
    assert nc is not None
    assert abs(nc.km_ini - 543.5) < 1e-9
    assert abs(nc.km_fim - 544.0) < 1e-9
    assert "543" in (nc.km_ini_str or "") and "500" in (nc.km_ini_str or "")


def test_parse_artemig_texto_km_valores_antes_rotulos_tabela_pdf():
    """PyMuPDF: «653,400» nas linhas antes de «Km Inicial» / «Km Final» (BR-265)."""
    from nc_artesp.modulos.analisar_pdf_nc import _parse_artemig_texto

    texto = (
        "NOTIFICAÇÃO CONSOL\n"
        "Nº da CONSOL:\n2516932\n"
        "BR-265\nSH20\n653,400\n653,400\nCRESCENTE\nPISTA\n"
        "Rodovia\nSH\nKm Inicial\nKm Final\nSentido\nLocal\n"
        "LOCALIZAÇÃO\n202501368\n"
    )
    nc = _parse_artemig_texto(texto)
    assert nc is not None
    assert abs(nc.km_ini - 653.4) < 1e-9
    assert abs(nc.km_fim - 653.4) < 1e-9


def test_parse_artemig_texto_tabela_pipes_antes_localizacao_sinalizacao():
    """Layout com Indicador|Patologia|…|Local| antes do bloco LOCALIZAÇÃO + código na linha seguinte."""
    from nc_artesp.modulos.analisar_pdf_nc import _parse_artemig_texto

    texto = (
        "NOTIFICAÇÃO CONSOL\n"
        "Indicador|Patologia|202506787|15/09/25|10:26|"
        "Sinalização Horizontal|Placas de advertência danificadas||Local|Em até 5 dias|\n"
        "LOCALIZAÇÃO\n"
        "202506787\n"
        "MG-050 SH02 100,500 CRESCENTE PISTA\n"
        "Nº da CONSOL: 2516953\n"
    )
    nc = _parse_artemig_texto(texto)
    assert nc is not None
    assert nc.codigo == "202506787"
    assert "Sinaliza" in (nc.indicador_artemig or "")
    assert "Horizontal" in (nc.indicador_artemig or "")
    assert "Placas" in (nc.patologia_artemig or "")
    assert "Gerais (Parâmetros)" not in (nc.patologia_artemig or "")


def test_arcnome_artemig_lote50_zip_e_coluna_w_alinhados():
    """ZIP lote 50: PDF mantém ``{stem}.pdf``; nc (*.jpg) recebe prefixo ``{stem} `` (single ou multi)."""
    from render_api.nc_router import _nc_arcnome_artemig_lote50

    stem_raw = "NOT-25-01369_PAVIMENTO_CE2516933"
    stem_s = "NOT-25-01369 PAVIMENTO CE2516933"

    assert _nc_arcnome_artemig_lote50(stem_raw, f"{stem_raw}.pdf") == f"{stem_s}.pdf"
    assert (
        _nc_arcnome_artemig_lote50(stem_raw, "nc (202501369).jpg")
        == f"{stem_s} nc (202501369).jpg"
    )
    assert (
        _nc_arcnome_artemig_lote50(stem_raw, "nc (202501369)_1.jpg")
        == f"{stem_s} nc (202501369) 1.jpg"
    )
    assert (
        _nc_arcnome_artemig_lote50(stem_raw, f"{stem_s} nc (202501369).jpg")
        == f"{stem_s} nc (202501369).jpg"
    )
    assert _nc_arcnome_artemig_lote50(stem_raw, f"{stem_raw}_{stem_raw}.pdf") == f"{stem_s}.pdf"
    assert (
        _nc_arcnome_artemig_lote50(stem_raw, f"{stem_raw}_{stem_raw}_{stem_raw}.pdf") == f"{stem_s}.pdf"
    )


def test_montar_col_w_kcor_prefixa_nc_jpg_com_stem_pdf():
    """Col. W espelha o nome do ficheiro no ZIP: ``{stem} nc (cod).jpg`` (stem alinhado ao PDF, sem «_»)."""
    from types import SimpleNamespace

    from nc_artemig.exportar_kcor_planilha import _montar_v_w_kcor

    nc = SimpleNamespace(
        codigo="202501367",
        num_consol="2516931",
        emergencial=False,
        prazo_dias=None,
        artemig_pdf_stem="NOT-25-01367_PAVIMENTO_CE2516931",
        artemig_kcor_nomes_arquivos=[
            "nc (202501367).jpg",
            "nc (202501367)_1.jpg",
        ],
        artemig_kcor_paginas_jpg=[],
        atividade="",
        tipo_atividade="",
        grupo_atividade="",
    )
    _v, w = _montar_v_w_kcor(nc)
    assert "NOT-25-01367 PAVIMENTO CE2516931.pdf" in w
    assert "NOT-25-01367 PAVIMENTO CE2516931 nc (202501367).jpg" in w
    assert "NOT-25-01367 PAVIMENTO CE2516931 nc (202501367) 1.jpg" in w
