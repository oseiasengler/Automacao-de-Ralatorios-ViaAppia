"""Layout planilha-mãe pós-M01 Kartado (colunas R / O-P / A)."""

from __future__ import annotations

import shutil
import sys
import zipfile
from pathlib import Path

import pytest

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Border, Side

_REPO = Path(__file__).resolve().parents[2]
_NC_ARTESP = _REPO / "nc_artesp"
if str(_NC_ARTESP) not in sys.path:
    sys.path.insert(0, str(_NC_ARTESP))

from modulos.separar_nc import (  # noqa: E402
    aplicar_layout_exportacao_planilha_mae_pos_coleta_kartado,
    gravar_planilha_mae_kartado_pasta_exportar,
    _ler_cabecalhos_planilha_hierarquico,
    _limitar_nome_ficheiro_exportar_windows,
)


def test_aplicar_layout_remove_col_r_insere_o_p_e_duas_em_a():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=2, column=1, value="A1")
    ws.cell(row=2, column=15, value="O_DATA")
    ws.cell(row=2, column=16, value="P_DATA")
    ws.cell(row=2, column=17, value="Q_DATA")
    ws.cell(row=2, column=18, value="R_GONE")
    ws.cell(row=2, column=19, value="S_AFTER")

    aplicar_layout_exportacao_planilha_mae_pos_coleta_kartado(ws)

    assert ws.cell(row=2, column=1).value is None
    assert ws.cell(row=2, column=2).value is None
    assert ws.cell(row=2, column=3).value == "A1"
    assert ws.cell(row=2, column=17).value == "O_DATA"
    assert ws.cell(row=2, column=18).value is None
    assert ws.cell(row=2, column=19).value == "P_DATA"
    assert ws.cell(row=2, column=20).value == "Q_DATA"
    assert ws.cell(row=2, column=21).value == "S_AFTER"


def test_gravar_planilha_mae_kartado_grava_sobre_template_exportar(tmp_path):
    mae = tmp_path / "mae_eaf.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.cell(row=5, column=3, value="NC001")
    wb.save(mae)
    nome = "20260101 - CONSTATAÇÕES NC LOTE 13 (SP 300 - TESTE) - Prazo - 15-01-2026.xlsx"
    out = gravar_planilha_mae_kartado_pasta_exportar(mae, tmp_path, nome)
    assert len(out) == 1 and out[0].is_file()
    from openpyxl import load_workbook as lw

    w = lw(out[0])
    t = w.active
    assert t.max_row >= 5
    assert t.cell(row=4, column=8).value == " Km Inicial"
    assert t.cell(row=5, column=3).value == "NC001"
    w.close()


def _snapshot_header(ws, linha_fim: int = 4, max_col: int | None = None) -> dict:
    data: dict = {
        "cells": {},
        "merges": sorted(str(rng) for rng in ws.merged_cells.ranges if rng.max_row <= linha_fim),
        "heights": {},
    }
    max_col = max_col if max_col is not None else ws.max_column
    for r in range(1, linha_fim + 1):
        dim = ws.row_dimensions.get(r)
        data["heights"][r] = dim.height if dim is not None else None
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            data["cells"][(r, c)] = (
                cell.value,
                cell.number_format,
                cell.font.name,
                cell.font.bold,
                cell.font.italic,
                cell.font.color.type if cell.font.color else None,
                cell.font.color.rgb if cell.font.color else None,
                cell.fill.fill_type,
                cell.fill.fgColor.type if cell.fill.fgColor else None,
                cell.fill.fgColor.rgb if cell.fill.fgColor else None,
                cell.alignment.horizontal,
                cell.alignment.vertical,
            )
    return data


def test_gravar_planilha_mae_preserva_cabecalho_template(monkeypatch, tmp_path):
    template = tmp_path / "Exportar.xlsx"
    wb_t = Workbook()
    ws_t = wb_t.active
    ws_t["A1"] = "TITULO"
    ws_t["A1"].font = Font(bold=True, name="Calibri")
    ws_t.merge_cells("A1:D1")
    ws_t["C3"] = "HEADER_MERGE"
    ws_t["C3"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    ws_t.merge_cells("C3:C4")
    ws_t.row_dimensions[2].height = 24
    ws_t["H4"] = " Km Inicial"
    wb_t.save(template)
    wb_t.close()

    mae = tmp_path / "mae.xlsx"
    wb_m = Workbook()
    ws_m = wb_m.active
    ws_m.cell(row=5, column=3, value="NC-100")
    ws_m.cell(row=6, column=4, value="NC-200")
    wb_m.save(mae)
    wb_m.close()

    monkeypatch.setattr("modulos.separar_nc.TEMPLATE_EXPORTAR_ROTINA", str(template))
    out = gravar_planilha_mae_kartado_pasta_exportar(mae, tmp_path, "saida.xlsx")
    assert len(out) == 1 and out[0].is_file()

    wb_tpl = load_workbook(template)
    wb_out = load_workbook(out[0])
    try:
        wt = wb_tpl.active
        mc = wt.max_column
        assert _snapshot_header(wb_out.active, max_col=mc) == _snapshot_header(wt, max_col=mc)
    finally:
        wb_tpl.close()
        wb_out.close()


def test_gravar_planilha_mae_escreve_apenas_corpo(monkeypatch, tmp_path):
    template = tmp_path / "Exportar.xlsx"
    wb_t = Workbook()
    ws_t = wb_t.active
    ws_t["A1"] = "HEAD"
    ws_t["B4"] = "Sub"
    ws_t.cell(row=5, column=3, value="LIXO")
    ws_t.cell(row=5, column=3).alignment = ws_t.cell(row=5, column=3).alignment.copy(horizontal="center")
    ws_t.cell(row=5, column=3).border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    ws_t.cell(row=6, column=4, value="LIXO2")
    wb_t.save(template)
    wb_t.close()

    mae = tmp_path / "mae.xlsx"
    wb_m = Workbook()
    ws_m = wb_m.active
    ws_m.cell(row=5, column=3, value="NC001")
    ws_m.cell(row=5, column=4, value="NC002")
    ws_m.cell(row=6, column=3, value="NC001")
    ws_m.cell(row=6, column=4, value="NC002")
    wb_m.save(mae)
    wb_m.close()

    monkeypatch.setattr("modulos.separar_nc.TEMPLATE_EXPORTAR_ROTINA", str(template))
    out = gravar_planilha_mae_kartado_pasta_exportar(mae, tmp_path, "saida_corpo.xlsx")
    assert len(out) == 1 and out[0].is_file()

    wb_out = load_workbook(out[0])
    try:
        ws_out = wb_out.active
        assert ws_out.cell(row=1, column=1).value == "HEAD"
        assert ws_out.cell(row=4, column=2).value == "Sub"
        # O template deve manter o mapeamento original de colunas (sem deslocar cabeçalho/corpo).
        assert ws_out.cell(row=5, column=3).value == "NC001"
        assert ws_out.cell(row=6, column=4).value == "NC002"
        assert ws_out.cell(row=5, column=18).value is None
        assert ws_out.cell(row=6, column=18).value is None
        assert ws_out.cell(row=6, column=3).alignment.horizontal == "center"
        assert ws_out.cell(row=6, column=3).border.left.style == "thin"
        assert ws_out.cell(row=6, column=3).border.right.style == "thin"
        assert ws_out.cell(row=6, column=3).border.top.style == "thin"
        assert ws_out.cell(row=6, column=3).border.bottom.style == "thin"
        assert ws_out.cell(row=5, column=22).value == 1
        assert ws_out.cell(row=6, column=22).value == 2
        # Harmonização das colunas finais S/T/U.
        assert ws_out.cell(row=5, column=20).value in ("LIXO", "LIXO2", None, "HEAD")
    finally:
        wb_out.close()


def test_cabecalho_hierarquico_duas_colunas_m_usam_km_a_esquerda_na_sub():
    """Exportar.xlsx: «Trecho» em H3 e I3–K3 vazios faziam «trecho|m» duplicado e K sumia do mapa."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=3, column=8, value="Trecho")
    ws.cell(row=4, column=8, value="Km Inicial")
    ws.cell(row=4, column=9, value="m")
    ws.cell(row=4, column=10, value="Km Final")
    ws.cell(row=4, column=11, value="m")
    hdr = _ler_cabecalhos_planilha_hierarquico(ws, 3, 4)
    assert hdr.get("km inicial | m") == 9
    assert hdr.get("km final | m") == 11


def test_gravar_mapeia_coluna_metros_km_inicial_alias_na_mae(monkeypatch, tmp_path):
    """Template «km inicial | m»; mãe só «metros inicial» — coluna K (metros) deve copiar."""
    template = tmp_path / "Exportar.xlsx"
    wb_t = Workbook()
    ws_t = wb_t.active
    ws_t.cell(row=3, column=3, value="Código de fiscalização")
    ws_t.cell(row=3, column=10, value="Km Inicial")
    ws_t.cell(row=4, column=10, value="km")
    ws_t.cell(row=4, column=11, value="m")
    ws_t.cell(row=5, column=11, value="_tpl_placeholder")
    wb_t.save(template)
    wb_t.close()

    mae = tmp_path / "mae.xlsx"
    wb_m = Workbook()
    ws_m = wb_m.active
    ws_m.cell(row=3, column=3, value="Código de fiscalização")
    ws_m.cell(row=3, column=11, value="metros inicial")
    ws_m.cell(row=5, column=3, value="NC001")
    ws_m.cell(row=5, column=11, value=777)
    wb_m.save(mae)
    wb_m.close()

    monkeypatch.setattr("modulos.separar_nc.TEMPLATE_EXPORTAR_ROTINA", str(template))
    out = gravar_planilha_mae_kartado_pasta_exportar(mae, tmp_path, "saida_metros_kmi.xlsx")
    assert len(out) == 1 and out[0].is_file()
    w = load_workbook(out[0])
    try:
        assert w.active.cell(row=5, column=11).value == 777
    finally:
        w.close()


def test_limitar_nome_exportar_trunca_nomes_longos_windows(tmp_path):
    pasta = tmp_path / ("sub_" + "p" * 80)
    pasta.mkdir(parents=True)
    longe = "A" * 300 + " - Prazo - 15-01-2026.xlsx"
    curto = _limitar_nome_ficheiro_exportar_windows(pasta, longe)
    assert curto.endswith(".xlsx")
    limite = min(180, max(40, 240 - len(str(pasta)) - 1))
    assert len(curto) <= limite + 2
    assert "Prazo" in curto


def test_exportar_rotina_separa_ficheiros_quando_prazo_diferente(monkeypatch, tmp_path):
    template = tmp_path / "Exportar.xlsx"
    wb_t = Workbook()
    ws_t = wb_t.active
    ws_t.cell(row=3, column=3, value="Código de fiscalização")
    ws_t.cell(row=3, column=6, value="Rodovia")
    ws_t.cell(row=3, column=17, value="Tipo NC")
    ws_t.cell(row=3, column=20, value="Data Reparo")
    ws_t.cell(row=5, column=3, value="_")
    wb_t.save(template)
    wb_t.close()

    mae = tmp_path / "mae.xlsx"
    wb_m = Workbook()
    ws_m = wb_m.active
    ws_m.cell(row=3, column=3, value="Código de fiscalização")
    ws_m.cell(row=3, column=6, value="Rodovia")
    ws_m.cell(row=3, column=17, value="Tipo NC")
    ws_m.cell(row=3, column=20, value="Data Reparo")
    ws_m.cell(row=5, column=3, value="NC001")
    ws_m.cell(row=5, column=6, value="SP 300")
    ws_m.cell(row=5, column=17, value="Servico Igual")
    ws_m.cell(row=5, column=20, value="2026-01-15")
    ws_m.cell(row=6, column=3, value="NC002")
    ws_m.cell(row=6, column=6, value="SP 300")
    ws_m.cell(row=6, column=17, value="Servico Igual")
    ws_m.cell(row=6, column=20, value="2026-02-15")
    wb_m.save(mae)
    wb_m.close()

    monkeypatch.setattr("modulos.separar_nc.TEMPLATE_EXPORTAR_ROTINA", str(template))
    out = gravar_planilha_mae_kartado_pasta_exportar(mae, tmp_path, "")
    assert len(out) == 2
    codigos_por_ficheiro = []
    for p in out:
        w = load_workbook(p)
        try:
            ws = w.active
            linha_dados = 4
            codigos_por_ficheiro.append(ws.cell(row=linha_dados, column=3).value)
            assert ws.cell(row=linha_dados, column=22).value == 1
        finally:
            w.close()
    assert set(codigos_por_ficheiro) == {"NC001", "NC002"}


def test_exportar_rotina_um_ficheiro_mesmo_tipo_rodovia_prazo(monkeypatch, tmp_path):
    """Mesmo tipo + rodovia + prazo + responsável → um único relatório (códigos/km diferentes no mesmo ficheiro)."""
    template = tmp_path / "Exportar.xlsx"
    wb_t = Workbook()
    ws_t = wb_t.active
    ws_t.cell(row=3, column=3, value="Código de fiscalização")
    ws_t.cell(row=3, column=6, value="Rodovia")
    ws_t.cell(row=3, column=8, value="Km")
    ws_t.cell(row=3, column=17, value="Tipo NC")
    ws_t.cell(row=3, column=20, value="Data Reparo")
    ws_t.cell(row=3, column=21, value="Responsável Técnico")
    ws_t.cell(row=5, column=3, value="_")
    wb_t.save(template)
    wb_t.close()

    mae = tmp_path / "mae.xlsx"
    wb_m = Workbook()
    ws_m = wb_m.active
    ws_m.cell(row=3, column=3, value="Código de fiscalização")
    ws_m.cell(row=3, column=6, value="Rodovia")
    ws_m.cell(row=3, column=8, value="Km")
    ws_m.cell(row=3, column=17, value="Tipo NC")
    ws_m.cell(row=3, column=20, value="Data Reparo")
    ws_m.cell(row=3, column=21, value="Responsável Técnico")
    ws_m.cell(row=5, column=3, value="NC001")
    ws_m.cell(row=5, column=6, value="SP 075")
    ws_m.cell(row=5, column=8, value="10+100")
    ws_m.cell(row=5, column=17, value="Pavimento")
    ws_m.cell(row=5, column=20, value="2026-05-08")
    ws_m.cell(row=5, column=21, value="Maria Fiscal")
    ws_m.cell(row=6, column=3, value="NC002")
    ws_m.cell(row=6, column=6, value="SP 075")
    ws_m.cell(row=6, column=8, value="20+200")
    ws_m.cell(row=6, column=17, value="Pavimento")
    ws_m.cell(row=6, column=20, value="2026-05-08")
    ws_m.cell(row=6, column=21, value="Maria Fiscal")
    wb_m.save(mae)
    wb_m.close()

    monkeypatch.setattr("modulos.separar_nc.TEMPLATE_EXPORTAR_ROTINA", str(template))
    out = gravar_planilha_mae_kartado_pasta_exportar(mae, tmp_path, "")
    assert len(out) == 1
    w = load_workbook(out[0])
    try:
        ws = w.active
        linha_dados = 4
        assert ws.cell(row=linha_dados, column=3).value == "NC001"
        assert ws.cell(row=linha_dados + 1, column=3).value == "NC002"
        assert ws.cell(row=linha_dados, column=22).value == 1
        assert ws.cell(row=linha_dados + 1, column=22).value == 2
        assert ws.cell(row=linha_dados, column=8).value == "10+100"
        assert ws.cell(row=linha_dados + 1, column=8).value == "20+200"
    finally:
        w.close()


def test_exportar_rotina_dois_ficheiros_responsavel_tecnico_diferente(monkeypatch, tmp_path):
    template = tmp_path / "Exportar.xlsx"
    wb_t = Workbook()
    ws_t = wb_t.active
    ws_t.cell(row=3, column=3, value="Código de fiscalização")
    ws_t.cell(row=3, column=6, value="Rodovia")
    ws_t.cell(row=3, column=17, value="Tipo NC")
    ws_t.cell(row=3, column=20, value="Data Reparo")
    ws_t.cell(row=3, column=21, value="Responsável Técnico")
    ws_t.cell(row=5, column=3, value="_")
    wb_t.save(template)
    wb_t.close()

    mae = tmp_path / "mae.xlsx"
    wb_m = Workbook()
    ws_m = wb_m.active
    ws_m.cell(row=3, column=3, value="Código de fiscalização")
    ws_m.cell(row=3, column=6, value="Rodovia")
    ws_m.cell(row=3, column=17, value="Tipo NC")
    ws_m.cell(row=3, column=20, value="Data Reparo")
    ws_m.cell(row=3, column=21, value="Responsável Técnico")
    ws_m.cell(row=5, column=3, value="NC-A")
    ws_m.cell(row=5, column=6, value="SP 075")
    ws_m.cell(row=5, column=17, value="Pavimento")
    ws_m.cell(row=5, column=20, value="2026-05-08")
    ws_m.cell(row=5, column=21, value="Fiscal Um")
    ws_m.cell(row=6, column=3, value="NC-B")
    ws_m.cell(row=6, column=6, value="SP 075")
    ws_m.cell(row=6, column=17, value="Pavimento")
    ws_m.cell(row=6, column=20, value="2026-05-08")
    ws_m.cell(row=6, column=21, value="Fiscal Dois")
    wb_m.save(mae)
    wb_m.close()

    monkeypatch.setattr("modulos.separar_nc.TEMPLATE_EXPORTAR_ROTINA", str(template))
    out = gravar_planilha_mae_kartado_pasta_exportar(mae, tmp_path, "")
    assert len(out) == 2
    codigos = set()
    for p in out:
        w = load_workbook(p)
        try:
            codigos.add(w.active.cell(row=4, column=3).value)
        finally:
            w.close()
    assert codigos == {"NC-A", "NC-B"}


def test_layout_com_merge_horizontal_abrangendo_coluna_r_executa_sem_excecao():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=2, column=15, value="O")
    ws.cell(row=2, column=16, value="P")
    ws.cell(row=2, column=17, value="Q_HEAD")
    ws.merge_cells("Q2:S2")
    aplicar_layout_exportacao_planilha_mae_pos_coleta_kartado(ws)
    assert ws.cell(row=2, column=17).value == "O"


def test_gravar_planilha_mae_repo_exportar_preserva_ooxml_printer_e_rels(monkeypatch, tmp_path):
    repo_tpl = _REPO / "nc_artesp" / "assets" / "templates" / "Exportar.xlsx"
    if not repo_tpl.is_file():
        pytest.skip("Template Exportar.xlsx ausente no clone")
    template = tmp_path / "Exportar.xlsx"
    shutil.copyfile(repo_tpl, template)

    mae = tmp_path / "mae.xlsx"
    wb_m = Workbook()
    wb_m.active.cell(row=5, column=3, value="NC001")
    wb_m.save(mae)
    wb_m.close()

    monkeypatch.setattr("modulos.separar_nc.TEMPLATE_EXPORTAR_ROTINA", str(template))
    out = gravar_planilha_mae_kartado_pasta_exportar(mae, tmp_path, "saida_ooxml.xlsx")
    assert len(out) == 1

    with zipfile.ZipFile(template) as tpl_z, zipfile.ZipFile(out[0]) as out_z:
        tpl_names = set(tpl_z.namelist())
        out_names = set(out_z.namelist())
        if "xl/worksheets/_rels/sheet1.xml.rels" in tpl_names:
            assert "xl/worksheets/_rels/sheet1.xml.rels" in out_names
        if "xl/printerSettings/printerSettings1.bin" in tpl_names:
            assert "xl/printerSettings/printerSettings1.bin" in out_names
        tpl_sheet = tpl_z.read("xl/worksheets/sheet1.xml").decode("utf-8", errors="replace")
        out_sheet = out_z.read("xl/worksheets/sheet1.xml").decode("utf-8", errors="replace")
        if "r:id=" in tpl_sheet and "pageSetup" in tpl_sheet:
            assert "r:id=" in out_sheet
