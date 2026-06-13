"""Reintegração OOXML após openpyxl: formas do template + desenho da foto no mesmo drawing."""

from __future__ import annotations

import re
import shutil
import sys
import tempfile
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook
from PIL import Image

_REPO = Path(__file__).resolve().parents[2]
_NC_ARTESP = _REPO / "nc_artesp"
if str(_NC_ARTESP) not in sys.path:
    sys.path.insert(0, str(_NC_ARTESP))

from modulos.gerar_modelo_foto import (  # noqa: E402
    _definir_valor_preservando_estilo,
    _inserir_imagem,
)
from utils.helpers import preservar_ooxml_planilha_pos_openpyxl  # noqa: E402


def test_preservar_modelo_resposta_mantem_forma_data_execucao_com_foto_pdf():
    modelo = _REPO / "nc_artesp" / "assets" / "templates" / "Modelo.xlsx"
    if not modelo.is_file():
        pytest.skip("Modelo.xlsx ausente")
    td = Path(tempfile.mkdtemp())
    img = td / "t.jpg"
    Image.new("RGB", (40, 30), color=(200, 50, 50)).save(img, "jpeg")
    out = td / "out.xlsx"
    shutil.copyfile(modelo, out)
    wb = load_workbook(out)
    try:
        ws = wb.active
        _definir_valor_preservando_estilo(ws, 2, 2, "cabeçalho teste")
        _inserir_imagem(ws, "B2", img, 960, 401)
        wb.save(out)
    finally:
        wb.close()
    preservar_ooxml_planilha_pos_openpyxl(modelo, out)
    z = zipfile.ZipFile(out)
    try:
        xml = z.read("xl/drawings/drawing1.xml").decode("utf-8", errors="replace")
        assert "pic" in xml
        assert "sp" in xml or ":sp" in xml
        assert "Execu" in xml or "execu" in xml.lower()
        rels = z.read("xl/worksheets/_rels/sheet1.xml.rels").decode("utf-8", errors="replace")
        assert "printerSettings" in rels
        assert "drawing" in rels.lower()
        assert "ns0:" not in rels
        ct = z.read("[Content_Types].xml").decode("utf-8", errors="replace")
        assert 'Extension="bin"' in ct
        assert "printerSettings" in ct
        sheet = z.read("xl/worksheets/sheet1.xml").decode("utf-8", errors="replace")
        m = re.search(r"<pageSetup[^/]*/>", sheet)
        assert m and 'r:id="rId2"' in m.group(0)
        assert 'r:id="rId1"' not in m.group(0)
    finally:
        z.close()
    load_workbook(out).close()
