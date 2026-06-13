"""Resposta Pendentes: replicação de shapes/layout do 1.º bloco para blocos seguintes."""

from __future__ import annotations

import shutil
import sys
import tempfile
import zipfile
import xml.etree.ElementTree as ET
from datetime import date
from pathlib import Path

import pytest
from PIL import Image

_REPO = Path(__file__).resolve().parents[2]
_NC_ARTESP = _REPO / "nc_artesp"
if str(_NC_ARTESP) not in sys.path:
    sys.path.insert(0, str(_NC_ARTESP))

from modulos.gerar_modelo_foto import _gerar_resposta  # noqa: E402


def _nc_base(codigo: str) -> dict:
    return {
        "codigo": codigo,
        "rod_tag": "SP 280",
        "rod_codigo": "SP 280",
        "km_i": "143+800",
        "sentido": "Norte",
        "tipo_nc": "Pavimento",
        "data_con": date(2026, 5, 8),
        "data_reparo": date(2026, 5, 9),
        "foto_id": 1,
        "num_foto": 1,
    }


def test_resposta_multiplos_blocos_replicam_shape_template():
    modelo = _REPO / "nc_artesp" / "assets" / "templates" / "Modelo.xlsx"
    if not modelo.is_file():
        pytest.skip("Modelo.xlsx ausente")
    td = Path(tempfile.mkdtemp())
    try:
        # PDF (1).jpg para as duas NCs; o objetivo aqui é validar shape sem foto.
        Image.new("RGB", (120, 80), color=(180, 20, 20)).save(td / "PDF (1).jpg", "jpeg")
        ncs = [_nc_base("923501"), _nc_base("923502")]
        out = _gerar_resposta(ncs, modelo, td, td)
        assert out is not None and out.is_file()
        from openpyxl import load_workbook as lw

        wbr = lw(out, data_only=True)
        try:
            wsr = wbr.active
            v9 = wsr.cell(9, 22).value
            w9 = wsr.cell(9, 23).value
            assert v9 and ("Imagens" in str(v9) or "CONSERVA" in str(v9).upper())
            assert w9 and "pdf" in str(w9).lower() and ".jpg" in str(w9).lower()
        finally:
            wbr.close()
        z = zipfile.ZipFile(out)
        try:
            xml = z.read("xl/drawings/drawing1.xml")
            root = ET.fromstring(xml)
            ns = {"xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"}
            non_pic = [
                a
                for a in root
                if a.tag.endswith("Anchor") and a.find("xdr:pic", ns) is None
            ]
            assert len(non_pic) >= 2
            rows = []
            for a in non_pic:
                frm = a.find("xdr:from/xdr:row", ns)
                if frm is not None and (frm.text or "").strip():
                    rows.append(int(frm.text))
            # Shape-base no modelo fica na faixa ~25; réplica do 2.º bloco deve aparecer em +28.
            assert any(r >= 25 for r in rows)
            assert any(r >= 53 for r in rows)
        finally:
            z.close()
    finally:
        shutil.rmtree(td, ignore_errors=True)
