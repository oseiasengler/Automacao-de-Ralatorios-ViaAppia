"""Deteção da coluna «Data Reparo» / sinónimos (Kartado) na planilha-mãe."""

from __future__ import annotations

import logging
import sys
from pathlib import Path

import openpyxl

_REPO = Path(__file__).resolve().parents[2]
_NC_ARTESP = _REPO / "nc_artesp"
if str(_NC_ARTESP) not in sys.path:
    sys.path.insert(0, str(_NC_ARTESP))

from modulos.separar_nc import (  # noqa: E402
    COL_SCAN_MAX_CABECALHO_DATAS,
    _detectar_col_data_reparo,
    _score_header_data_reparo,
)


def test_score_data_reparo_sinonimos():
    assert _score_header_data_reparo("data reparo") == 100
    assert _score_header_data_reparo("data do reparo") == 99
    assert _score_header_data_reparo("prazo") == 70
    assert _score_header_data_reparo("dtfim_prog") == 80


def test_detecta_data_reparo_col_t_sem_warning(caplog):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=20, value="Data Reparo")
    with caplog.at_level(logging.WARNING):
        col = _detectar_col_data_reparo(ws, fallback=20)
    assert col == 20
    assert "nao encontrada" not in caplog.text.lower()


def test_detecta_prazo_alem_coluna_30_sem_warning(caplog):
    wb = openpyxl.Workbook()
    ws = wb.active
    c = min(45, COL_SCAN_MAX_CABECALHO_DATAS)
    ws.cell(row=2, column=c, value="Prazo")
    with caplog.at_level(logging.WARNING):
        col = _detectar_col_data_reparo(ws, fallback=20)
    assert col == c
    assert "nao encontrada" not in caplog.text.lower()


def test_sem_cabecalho_emite_warning_e_fallback(caplog):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=10, column=3, value=12345)
    with caplog.at_level(logging.WARNING):
        col = _detectar_col_data_reparo(ws, fallback=20)
    assert col == 20
    assert "nao encontrada" in caplog.text.lower()
