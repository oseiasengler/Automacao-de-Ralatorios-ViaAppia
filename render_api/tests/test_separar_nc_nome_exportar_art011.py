"""Nome ficheiro Exportar: texto da coluna Atividade (detetada) + abreviatura Art_011."""

from __future__ import annotations

import sys
from pathlib import Path

_REPO = Path(__file__).resolve().parents[2]
_NC_ARTESP = _REPO / "nc_artesp"
if str(_NC_ARTESP) not in sys.path:
    sys.path.insert(0, str(_NC_ARTESP))

from modulos.separar_nc import (  # noqa: E402
    COL_TIPO_NC,
    _detectar_col_tipo_nc,
    _nome_arquivo,
    _nome_arquivo_consolidado_eaf,
    _tipo_nc_texto_para_nome_exportar_art011,
)


def test_nome_arquivo_usa_abrev_art011_buraco_panela():
    nome = _nome_arquivo("SP 300", "Buraco ou panela", "01/01/2026", "15/01/2026")
    assert "PANELA" in nome
    assert "Pav - Panela" not in nome


def test_tipo_para_nome_prefere_coluna_detetada_q_quando_eh_tipo_nc():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    # Coluna detetada como Tipo NC/Atividade = Q (17): texto Art_011; N (14) = macro «Tipo de Atividade» (não entra no nome).
    r = 5
    ws.cell(row=r, column=14, value="Tipo de Atividade genérico")
    ws.cell(row=r, column=COL_TIPO_NC, value="Buraco ou panela")
    ws.cell(row=r, column=3, value="26-0001")
    t = _tipo_nc_texto_para_nome_exportar_art011(ws, r, COL_TIPO_NC, 3)
    assert t == "Buraco ou panela"


def test_tipo_para_nome_prefere_atividade_n_quando_q_e_grupo():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=14, value="Atividade")
    ws.cell(row=1, column=17, value="Grupo de atividade")
    r = 5
    ws.cell(row=r, column=14, value="Pintura")
    ws.cell(row=r, column=17, value="Remoção de lixo e entulho")
    ws.cell(row=r, column=3, value="26-0001")
    col_ativ = _detectar_col_tipo_nc(ws)
    assert col_ativ == 14
    t = _tipo_nc_texto_para_nome_exportar_art011(ws, r, col_ativ, 3)
    assert t == "Pintura"


def test_nome_consolidado_inclui_serv_art011_e_sufixo():
    linhas_info = [
        ("15/01/2026", "SP 300", "Buraco ou panela", "01/01/2026"),
    ]
    nome = _nome_arquivo_consolidado_eaf(
        linhas_info, [5], 5, arquivo_mae=Path("EAF_Lote26.xlsx")
    )
    assert "PANELA" in nome
    assert "Consolidado" in nome


def test_tipo_para_nome_ignora_data_na_coluna_q_usa_fallback():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    r = 5
    ws.cell(row=r, column=COL_TIPO_NC, value="06/06/2026")
    ws.cell(row=r, column=14, value="Buraco ou panela")
    ws.cell(row=r, column=3, value="26-0001")
    t = _tipo_nc_texto_para_nome_exportar_art011(ws, r, 14, 3)
    assert t == "Buraco ou panela"


def test_tipo_para_nome_fallback_se_q_vazia():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    r = 5
    ws.cell(row=r, column=COL_TIPO_NC, value=None)
    ws.cell(row=r, column=14, value="Só coluna detetada")
    ws.cell(row=r, column=3, value="X")
    t = _tipo_nc_texto_para_nome_exportar_art011(ws, r, 14, 3)
    assert t == "Só coluna detetada"
